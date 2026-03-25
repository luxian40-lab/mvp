from django.contrib import admin
from django.contrib import messages
from django.utils.html import format_html
from django.shortcuts import render, redirect
from django.urls import path
from django.utils import timezone
import openpyxl
from django.http import HttpResponse, JsonResponse
from .models import (
    Estudiante, Plantilla, Campana, EnvioLog, Linea, WhatsappLog,
    Curso, Modulo, ProgresoEstudiante,
    AliadoEmpleabilidad, LogroEstudiante,
    PreguntaAbierta, RespuestaAbierta,
)
from .services import ejecutar_campana_servicio

# =================================================
# 1. ACCIÓN: EXPORTAR A EXCEL (Estilo Andrés)
# =================================================
@admin.action(description='📥 Descargar Reporte Excel')
def exportar_logs_excel(modeladmin, request, queryset):
    """
    Genera un Excel idéntico a la tabla que ve Andrés.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_envios.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reporte de Envíos'

    # Encabezados (Iguales a la imagen)
    columns = ['ID', 'Receptor (Estudiante)', 'Teléfono', 'Estado', 'Fecha Envío', 'Campaña', 'Plantilla', 'Respuesta API']
    worksheet.append(columns)

    # Datos
    for log in queryset:
        # Quitamos la zona horaria para que Excel no moleste
        fecha_sin_tz = log.fecha_envio.replace(tzinfo=None) if log.fecha_envio else ""
        
        row = [
            log.id,
            log.estudiante.nombre,
            log.estudiante.telefono,
            log.estado,
            fecha_sin_tz,
            log.campana.nombre,
            log.campana.plantilla.nombre_interno,
            log.respuesta_api
        ]
        worksheet.append(row)

    workbook.save(response)
    return response

# =================================================
# 2. ACCIÓN: IMPORTAR ESTUDIANTES MASIVAMENTE
# =================================================
@admin.action(description='📤 Importar Estudiantes desde Excel')
def importar_estudiantes_desde_logs(modeladmin, request, queryset):
    """
    Importa estudiantes desde un archivo Excel en la vista de logs.
    """
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_estudiantes')
        
        if not archivo:
            modeladmin.message_user(request, '❌ Por favor selecciona un archivo Excel', level=messages.ERROR)
            return redirect(request.META.get('HTTP_REFERER', '/admin/core/enviolog/'))
        
        try:
            # Validar extensión
            if not archivo.name.endswith(('.xlsx', '.xls')):
                modeladmin.message_user(request, '❌ El archivo debe ser .xlsx o .xls', level=messages.ERROR)
                return redirect(request.META.get('HTTP_REFERER', '/admin/core/enviolog/'))
            
            # Cargar Excel
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            creados = 0
            actualizados = 0
            errores = []
            
            # Procesar filas (saltar encabezado en fila 1)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    nombre = row[0]
                    telefono = row[1]
                    
                    if not nombre or not telefono:
                        continue
                    
                    telefono_str = str(telefono).strip()
                    nombre_str = str(nombre).strip()
                    
                    estudiante, created = Estudiante.objects.update_or_create(
                        telefono=telefono_str,
                        defaults={'nombre': nombre_str, 'activo': True}
                    )
                    
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
            
            # Mensaje de éxito
            total = creados + actualizados
            msg = f"✅ Importación completada: {creados} nuevos, {actualizados} actualizados, {total} total"
            
            if errores:
                msg += f"\n⚠️ {len(errores)} errores encontrados"
            
            modeladmin.message_user(request, msg, level=messages.SUCCESS)
            
        except Exception as e:
            modeladmin.message_user(request, f'❌ Error al procesar: {str(e)}', level=messages.ERROR)
        
        return redirect(request.META.get('HTTP_REFERER', '/admin/core/enviolog/'))
    
    # Mostrar formulario para subir archivo
    return render(request, 'admin/importar_en_logs.html', {
        'action': 'importar_estudiantes_desde_logs',
        'site_header': 'Importar Estudiantes'
    })

# =================================================
# 2. CONFIGURACIÓN DE TABLAS
# =================================================

@admin.register(Estudiante)
class EstudianteAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'telefono', 'activo', 'fecha_registro')
    search_fields = ('nombre', 'telefono')
    list_per_page = 20
    actions = ['exportar_estudiantes_excel']
    
    def get_urls(self):
        """Agregar URL personalizada para importar masivamente"""
        urls = super().get_urls()
        custom_urls = [
            path('importar-masivamente/', self.admin_site.admin_view(self.importar_masivamente), name='core_estudiante_importar'),
        ]
        return custom_urls + urls
    
    def importar_masivamente(self, request):
        """Vista personalizada para importar estudiantes masivamente"""
        if request.method == 'POST':
            archivo = request.FILES.get('archivo_excel')
            
            if not archivo:
                messages.error(request, '❌ Por favor selecciona un archivo Excel')
                return render(request, 'admin/importar_estudiantes_masivamente.html', {
                    'site_header': self.admin_site.site_header,
                    'site_title': self.admin_site.site_title,
                })
            
            try:
                # Validar extensión
                if not archivo.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, '❌ El archivo debe ser .xlsx o .xls')
                    return render(request, 'admin/importar_estudiantes_masivamente.html', {
                        'site_header': self.admin_site.site_header,
                        'site_title': self.admin_site.site_title,
                    })
                
                # Cargar Excel
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                
                creados = 0
                actualizados = 0
                errores = []
                
                # Procesar filas (saltar encabezado en fila 1)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        nombre = row[0]
                        telefono = row[1]
                        
                        if not nombre or not telefono:
                            continue
                        
                        telefono_str = str(telefono).strip()
                        nombre_str = str(nombre).strip()
                        
                        estudiante, created = Estudiante.objects.update_or_create(
                            telefono=telefono_str,
                            defaults={'nombre': nombre_str, 'activo': True}
                        )
                        
                        if created:
                            creados += 1
                        else:
                            actualizados += 1
                    
                    except Exception as e:
                        errores.append(f"Fila {row_idx}: {str(e)}")
                
                # Mensaje de éxito
                total = creados + actualizados
                mensaje = f"✅ Importación completada: {creados} nuevos, {actualizados} actualizados, {total} total procesados"
                
                if errores:
                    mensaje += f"\n⚠️ {len(errores)} errores encontrados"
                
                messages.success(request, mensaje)
                
                return render(request, 'admin/importar_estudiantes_masivamente.html', {
                    'site_header': self.admin_site.site_header,
                    'site_title': self.admin_site.site_title,
                    'creados': creados,
                    'actualizados': actualizados,
                    'total': total,
                    'errores': errores,
                })
                
            except Exception as e:
                messages.error(request, f'❌ Error al procesar: {str(e)}')
                return render(request, 'admin/importar_estudiantes_masivamente.html', {
                    'site_header': self.admin_site.site_header,
                    'site_title': self.admin_site.site_title,
                })
        
        # GET: Mostrar formulario
        return render(request, 'admin/importar_estudiantes_masivamente.html', {
            'site_header': self.admin_site.site_header,
            'site_title': self.admin_site.site_title,
        })
    
    def exportar_estudiantes_excel(self, request, queryset):
        """Exportar estudiantes seleccionados a Excel."""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="estudiantes_export.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estudiantes'
        
        # Encabezados
        headers = ['Nombre', 'Teléfono', 'Activo', 'Fecha Registro']
        ws.append(headers)
        
        # Datos
        for est in queryset:
            ws.append([
                est.nombre,
                est.telefono,
                'Sí' if est.activo else 'No',
                est.fecha_registro.strftime('%Y-%m-%d %H:%M') if est.fecha_registro else ''
            ])
        
        wb.save(response)
        return response
    
    exportar_estudiantes_excel.short_description = '📥 Descargar Estudiantes (Excel)'
    
    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        """Personalizar el formulario para agregar un botón de importar masivamente"""
        from django.urls import reverse
        extra_context = extra_context or {}
        # Usar reverse para obtener la URL personalizada
        extra_context['importar_url'] = reverse('admin:core_estudiante_importar')
        extra_context['show_importar_btn'] = True
        return super().changeform_view(
            request, object_id=object_id, form_url=form_url, extra_context=extra_context
        )

@admin.register(Plantilla)
class PlantillaAdmin(admin.ModelAdmin):
    list_display = ('nombre_interno', 'vista_previa', 'tiene_imagen', 'preview_imagen')
    search_fields = ('nombre_interno',)
    list_filter = ('tiene_imagen',)
    
    fieldsets = (
        ('Información Básica', {
            'fields': ('nombre_interno', 'cuerpo_mensaje')
        }),
        ('Configuración de Imagen', {
            'fields': ('tiene_imagen', 'url_imagen'),
            'description': 'Configura una imagen para enviar junto con el mensaje de WhatsApp'
        }),
    )
    
    def vista_previa(self, obj):
        return obj.cuerpo_mensaje[:50] + "..."
    vista_previa.short_description = "Cuerpo del Mensaje"
    
    def preview_imagen(self, obj):
        if obj.tiene_imagen and obj.url_imagen:
            return format_html(
                '<a href="{}" target="_blank">🔗 Ver imagen</a>',
                obj.url_imagen
            )
        return "-"
    preview_imagen.short_description = "URL Imagen"

@admin.register(Campana)
class CampanaAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'estado_visual', 'conteo_destinatarios', 'fecha_creacion')
    list_filter = ('ejecutada', 'fecha_creacion')
    search_fields = ('nombre',)
    filter_horizontal = ('destinatarios',)
    
    # Configuramos para que el formulario de crear sea limpio
    fieldsets = (
        ('📝 Datos Básicos', {
            'fields': ('nombre', 'plantilla', 'canal_envio', 'linea_origen')
        }),
        ('📂 Importar (Opcional)', {
            'fields': ('archivo_excel',),
            'description': 'Sube un Excel con columnas A (Nombre) y B (Teléfono).'
        }),
        ('👥 Audiencia', {
            'fields': ('destinatarios',)
        }),
    )

    # Botón para enviar desde la lista
    actions = ['enviar_campana_accion']

    @admin.action(description='🚀 Ejecutar Campaña (Enviar Mensajes)')
    def enviar_campana_accion(self, request, queryset):
        for campana in queryset:
            if campana.ejecutada:
                self.message_user(request, f"⚠️ '{campana.nombre}' ya fue enviada antes.", level=messages.WARNING)
                continue
            
            res = ejecutar_campana_servicio(campana)
            self.message_user(request, f"✅ '{campana.nombre}': {res['exitosos']} enviados, {res['fallidos']} errores.", level=messages.SUCCESS)

    def estado_visual(self, obj):
        if obj.ejecutada:
            return format_html('<span style="color: green;">✅ Ejecutada</span>')
        return format_html('<span style="color: orange;">⏳ Pendiente</span>')
    estado_visual.short_description = "Estado"
    
    def conteo_destinatarios(self, obj):
        return obj.destinatarios.count()
    conteo_destinatarios.short_description = "# Destinatarios"


# ESTA ES LA TABLA QUE SE PARECE A LA IMAGEN
@admin.register(EnvioLog)
class EnvioLogAdmin(admin.ModelAdmin):
    # Columnas exactas de la imagen
    list_display = ('id', 'estudiante_nombre', 'estado_color', 'fecha_envio', 'nombre_plantilla', 'detalle_mensaje')
    
    # Filtros laterales (Simulan los tabs y fechas)
    list_filter = ('estado', 'fecha_envio', 'campana__nombre')
    
    # Buscador general
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'campana__nombre')
    
    # Botones de exportar e importar
    actions = [exportar_logs_excel, importar_estudiantes_desde_logs]
    
    # Solo lectura (historial no se debe editar)
    readonly_fields = ('campana', 'estudiante', 'estado', 'respuesta_api', 'fecha_envio')

    def estudiante_nombre(self, obj):
        return f"{obj.estudiante.nombre} ({obj.estudiante.telefono})"
    estudiante_nombre.short_description = "Receptor"

    def nombre_plantilla(self, obj):
        return obj.campana.plantilla.nombre_interno
    nombre_plantilla.short_description = "Plantilla"

    def detalle_mensaje(self, obj):
        # Mostramos el inicio del mensaje
        msg = obj.campana.plantilla.cuerpo_mensaje.replace("{nombre}", obj.estudiante.nombre)
        return msg[:40] + "..."
    detalle_mensaje.short_description = "Detalle"

    def estado_color(self, obj):
        if obj.estado == 'ENVIADO':
            return format_html('<b style="color:green;">ENVIADO</b>')
        elif obj.estado == 'FALLIDO':
            return format_html('<b style="color:red;">FALLIDO</b>')
        return obj.estado
    estado_color.short_description = "Estado"


# TABLA DE LOGS DE WHATSAPP
@admin.register(WhatsappLog)
class WhatsappLogAdmin(admin.ModelAdmin):
    # Columnas principales
    list_display = ('id', 'telefono_formateado', 'tipo_mensaje', 'estado_color', 'fecha', 'mensaje_preview', 'mensaje_id')
    
    # Filtros laterales
    list_filter = ('estado', 'fecha')
    
    # Búsqueda
    search_fields = ('telefono', 'mensaje', 'mensaje_id')
    
    # Solo lectura (son logs, no se deben editar)
    readonly_fields = ('telefono', 'mensaje', 'mensaje_id', 'estado', 'fecha')
    
    # Organización del formulario
    fieldsets = (
        ('📱 Información de Contacto', {
            'fields': ('telefono',)
        }),
        ('💬 Mensaje', {
            'fields': ('mensaje',)
        }),
        ('📊 Estado y Metadata', {
            'fields': ('estado', 'mensaje_id', 'fecha')
        }),
    )
    
    def telefono_formateado(self, obj):
        return format_html('<strong>{}</strong>', obj.telefono)
    telefono_formateado.short_description = "Teléfono"
    
    def tipo_mensaje(self, obj):
        if obj.estado == 'INCOMING':
            return format_html('<span style="color: #007bff;">📥 Entrante</span>')
        else:
            return format_html('<span style="color: #28a745;">📤 Saliente</span>')
    tipo_mensaje.short_description = "Tipo"
    
    def estado_color(self, obj):
        if obj.estado == 'SENT':
            return format_html('<span style="background: #28a745; color: white; padding: 3px 8px; border-radius: 3px;">✅ ENVIADO</span>')
        elif obj.estado == 'INCOMING':
            return format_html('<span style="background: #007bff; color: white; padding: 3px 8px; border-radius: 3px;">📥 RECIBIDO</span>')
        elif obj.estado == 'PENDING':
            return format_html('<span style="background: #ffc107; color: black; padding: 3px 8px; border-radius: 3px;">⏳ PENDIENTE</span>')
        elif obj.estado == 'ERROR':
            return format_html('<span style="background: #dc3545; color: white; padding: 3px 8px; border-radius: 3px;">❌ ERROR</span>')
        else:
            return format_html('<span style="background: #6c757d; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>', obj.estado)
    estado_color.short_description = "Estado"
    
    def mensaje_preview(self, obj):
        if not obj.mensaje:
            return "—"
        preview = obj.mensaje[:50]
        if len(obj.mensaje) > 50:
            preview += "..."
        return preview
    mensaje_preview.short_description = "Mensaje"
    
    def get_ordering(self, request):
        # Ordenar por fecha descendente por defecto
        return ['-fecha']


# ============================================================
# MÓDULO: ENTREGAS DIFERIDAS (DRIP CONTENT)
# ============================================================

class ModuloInline(admin.TabularInline):
    model = Modulo
    extra = 1
    fields = ('orden', 'nombre', 'contenido')
    ordering = ('orden',)


class PreguntaAbiertaInline(admin.TabularInline):
    model = PreguntaAbierta
    extra = 1
    fields = ('orden', 'pregunta', 'activa')
    ordering = ('orden',)


@admin.register(Curso)
class CursoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'dias_espera_entre_modulos', 'activo', 'total_modulos')
    list_filter = ('activo',)
    search_fields = ('nombre',)
    inlines = [ModuloInline]

    def total_modulos(self, obj):
        return obj.modulos.count()
    total_modulos.short_description = '# Módulos'


@admin.register(Modulo)
class ModuloAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'curso', 'orden')
    list_filter = ('curso',)
    search_fields = ('nombre', 'curso__nombre')
    ordering = ('curso', 'orden')
    inlines = [PreguntaAbiertaInline]


@admin.register(ProgresoEstudiante)
class ProgresoEstudianteAdmin(admin.ModelAdmin):
    list_display = ('estudiante', 'curso', 'modulo_actual', 'estado', 'fecha_ultimo_avance', 'puede_avanzar_display')
    list_filter = ('estado', 'curso')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'curso__nombre')
    readonly_fields = ('fecha_inicio',)

    def puede_avanzar_display(self, obj):
        puede, fecha = obj.puede_avanzar()
        if puede:
            return format_html('<span style="color:green;">✅ Sí</span>')
        fecha_str = fecha.strftime('%d/%m/%Y') if fecha else '?'
        return format_html('<span style="color:orange;">🔒 No ({})</span>', fecha_str)
    puede_avanzar_display.short_description = '¿Puede avanzar?'


# ============================================================
# MÓDULO: GAMIFICACIÓN GEOLOCALIZADA
# ============================================================

@admin.register(AliadoEmpleabilidad)
class AliadoEmpleabilidadAdmin(admin.ModelAdmin):
    list_display = ('nombre_empresa', 'latitud', 'longitud', 'vacantes_activas', 'codigo_secreto')
    list_filter = ('vacantes_activas',)
    search_fields = ('nombre_empresa',)


@admin.register(LogroEstudiante)
class LogroEstudianteAdmin(admin.ModelAdmin):
    list_display = ('estudiante', 'tipo', 'descripcion', 'aliado', 'fecha')
    list_filter = ('tipo', 'fecha')
    search_fields = ('estudiante__nombre', 'descripcion')
    readonly_fields = ('fecha',)


# ============================================================
# MÓDULO: PREGUNTAS ABIERTAS (EVALUACIÓN POR FACILITADORA)
# ============================================================

class RespuestaAbiertaInline(admin.TabularInline):
    model = RespuestaAbierta
    extra = 0
    fields = ('estudiante', 'respuesta', 'calificacion', 'comentario_facilitador', 'fecha_respuesta')
    readonly_fields = ('estudiante', 'respuesta', 'fecha_respuesta')
    can_delete = False


@admin.register(PreguntaAbierta)
class PreguntaAbiertaAdmin(admin.ModelAdmin):
    list_display = ('modulo', 'orden', 'pregunta_preview', 'activa', 'total_respuestas')
    list_filter = ('activa', 'modulo__curso')
    search_fields = ('pregunta', 'modulo__nombre')
    inlines = [RespuestaAbiertaInline]

    def pregunta_preview(self, obj):
        return obj.pregunta[:60] + ('...' if len(obj.pregunta) > 60 else '')
    pregunta_preview.short_description = 'Pregunta'

    def total_respuestas(self, obj):
        return obj.respuestas.count()
    total_respuestas.short_description = '# Respuestas'


@admin.register(RespuestaAbierta)
class RespuestaAbiertaAdmin(admin.ModelAdmin):
    list_display = (
        'estudiante', 'pregunta_preview', 'respuesta_preview',
        'calificacion', 'fecha_respuesta', 'calificado_display'
    )
    list_filter = ('pregunta__modulo__curso', 'fecha_respuesta')
    search_fields = ('estudiante__nombre', 'respuesta', 'pregunta__pregunta')
    readonly_fields = ('estudiante', 'pregunta', 'respuesta', 'fecha_respuesta', 'fecha_calificacion', 'calificado_por')

    fieldsets = (
        ('📋 Respuesta del Estudiante', {
            'fields': ('estudiante', 'pregunta', 'respuesta', 'fecha_respuesta')
        }),
        ('✏️ Calificación de la Facilitadora', {
            'fields': ('calificacion', 'comentario_facilitador', 'calificado_por', 'fecha_calificacion')
        }),
    )

    def pregunta_preview(self, obj):
        return obj.pregunta.pregunta[:50] + ('...' if len(obj.pregunta.pregunta) > 50 else '')
    pregunta_preview.short_description = 'Pregunta'

    def respuesta_preview(self, obj):
        return obj.respuesta[:60] + ('...' if len(obj.respuesta) > 60 else '')
    respuesta_preview.short_description = 'Respuesta'

    def calificado_display(self, obj):
        if obj.calificacion is not None:
            return format_html('<span style="color:green;">✅ {}</span>', obj.calificacion)
        return format_html('<span style="color:orange;">⏳ Pendiente</span>')
    calificado_display.short_description = 'Estado Calificación'

    def save_model(self, request, obj, form, change):
        """Al guardar una calificación, registra quién calificó y cuándo."""
        if obj.calificacion is not None and 'calificacion' in form.changed_data:
            obj.calificado_por = request.user
            obj.fecha_calificacion = timezone.now()
        super().save_model(request, obj, form, change)

