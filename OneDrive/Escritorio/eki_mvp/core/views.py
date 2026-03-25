from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    Campana, EnvioLog, Estudiante, WhatsappLog,
    ProgresoEstudiante, AliadoEmpleabilidad, LogroEstudiante,
    PreguntaAbierta, RespuestaAbierta,
)
from .utils import enviar_whatsapp, haversine_distance
from .intent_detector import detect_intent
from .response_templates import get_response_for_intent

@staff_member_required
def dashboard_view(request):
    # 1. Calcular Métricas
    total_campanas = Campana.objects.count()
    
    # Contamos logs
    total_envios = EnvioLog.objects.count()
    exitosos = EnvioLog.objects.filter(estado='ENVIADO').count()
    fallidos = EnvioLog.objects.filter(estado='FALLIDO').count()
    pendientes = EnvioLog.objects.filter(estado='PENDIENTE').count()
    
    estudiantes_activos = Estudiante.objects.filter(activo=True).count()
    
    # WhatsApp logs (últimos 10)
    whatsapp_logs = WhatsappLog.objects.all().order_by('-fecha')[:10]
    whatsapp_total = WhatsappLog.objects.count()
    whatsapp_enviados = WhatsappLog.objects.filter(estado='SENT').count()
    whatsapp_entrantes = WhatsappLog.objects.filter(estado='INCOMING').count()

    context = {
        'total_campanas': total_campanas,
        'total_envios': total_envios,
        'exitosos': exitosos,
        'fallidos': fallidos,
        'pendientes': pendientes,
        'estudiantes_activos': estudiantes_activos,
        'whatsapp_logs': whatsapp_logs,
        'whatsapp_total': whatsapp_total,
        'whatsapp_enviados': whatsapp_enviados,
        'whatsapp_entrantes': whatsapp_entrantes,
    }
    
    # Renderizamos la plantilla que vamos a crear en el paso 2
    return render(request, 'admin/dashboard_metrics.html', context)


# ---------- Vista de importación de estudiantes ----------
@staff_member_required
def importar_estudiantes(request):
    """Vista para importar estudiantes desde un archivo Excel."""
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        
        if not archivo:
            return JsonResponse({
                'exito': False,
                'mensaje': 'Por favor selecciona un archivo Excel'
            })
        
        try:
            # Verificar que sea Excel
            if not archivo.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'exito': False,
                    'mensaje': 'El archivo debe ser .xlsx o .xls'
                })
            
            # Cargar el libro de trabajo
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            estudiantes_creados = 0
            estudiantes_actualizados = 0
            errores = []
            
            # Iteramos desde la fila 2 (saltar encabezados)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Esperamos: Nombre en columna A, Teléfono en columna B
                    nombre = row[0]
                    telefono = row[1]
                    
                    # Validar que no estén vacíos
                    if not nombre or not telefono:
                        continue
                    
                    # Limpiar teléfono
                    telefono_str = str(telefono).strip()
                    nombre_str = str(nombre).strip()
                    
                    # Crear o actualizar estudiante
                    estudiante, creado = Estudiante.objects.update_or_create(
                        telefono=telefono_str,
                        defaults={
                            'nombre': nombre_str,
                            'activo': True
                        }
                    )
                    
                    if creado:
                        estudiantes_creados += 1
                    else:
                        estudiantes_actualizados += 1
                    
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
                    if len(errores) >= 10:  # Limitar errores mostrados
                        break
            
            return JsonResponse({
                'exito': True,
                'creados': estudiantes_creados,
                'actualizados': estudiantes_actualizados,
                'total': estudiantes_creados + estudiantes_actualizados,
                'errores': errores
            })
        
        except Exception as e:
            return JsonResponse({
                'exito': False,
                'mensaje': f'Error al procesar el archivo: {str(e)}'
            })
    
    # GET: Mostrar formulario
    return render(request, 'admin/importar_estudiantes.html')


# ---------- Vista de descarga de reportes ----------
@staff_member_required
def descargar_reportes(request):
    """Vista para descargar reportes en Excel filtrando por fechas."""
    context = {}
    
    if request.method == 'POST':
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        tipo_reporte = request.POST.get('tipo_reporte', 'todos')  # todos, envios, whatsapp
        
        try:
            # Parsear fechas
            inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d') if fecha_inicio else None
            fin = datetime.strptime(fecha_fin, '%Y-%m-%d') if fecha_fin else None
            
            # Ajustar fin de día
            if fin:
                fin = fin.replace(hour=23, minute=59, second=59)
            
            # Crear workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Eliminar hoja por defecto
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # ========== ENVÍOS ==========
            if tipo_reporte in ['todos', 'envios']:
                ws_envios = wb.create_sheet('Envíos')
                
                # Filtrar por fecha
                queryset = EnvioLog.objects.all()
                if inicio:
                    queryset = queryset.filter(fecha_envio__gte=inicio)
                if fin:
                    queryset = queryset.filter(fecha_envio__lte=fin)
                queryset = queryset.order_by('-fecha_envio')
                
                # Encabezados
                headers = ['ID', 'Estudiante', 'Teléfono', 'Campaña', 'Plantilla', 'Estado', 'Fecha', 'Respuesta API']
                ws_envios.append(headers)
                
                # Aplicar estilos a encabezados
                for cell in ws_envios[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Datos
                for log in queryset:
                    fecha_str = log.fecha_envio.strftime('%Y-%m-%d %H:%M:%S') if log.fecha_envio else ''
                    row = [
                        log.id,
                        log.estudiante.nombre,
                        log.estudiante.telefono,
                        log.campana.nombre,
                        log.campana.plantilla.nombre_interno,
                        log.estado,
                        fecha_str,
                        log.respuesta_api or ''
                    ]
                    ws_envios.append(row)
                
                # Ajustar ancho de columnas
                ws_envios.column_dimensions['A'].width = 8
                ws_envios.column_dimensions['B'].width = 20
                ws_envios.column_dimensions['C'].width = 15
                ws_envios.column_dimensions['D'].width = 20
                ws_envios.column_dimensions['E'].width = 20
                ws_envios.column_dimensions['F'].width = 12
                ws_envios.column_dimensions['G'].width = 20
                ws_envios.column_dimensions['H'].width = 30
            
            # ========== WHATSAPP ==========
            if tipo_reporte in ['todos', 'whatsapp']:
                ws_whatsapp = wb.create_sheet('WhatsApp')
                
                # Filtrar por fecha
                queryset = WhatsappLog.objects.all()
                if inicio:
                    queryset = queryset.filter(fecha__gte=inicio)
                if fin:
                    queryset = queryset.filter(fecha__lte=fin)
                queryset = queryset.order_by('-fecha')
                
                # Encabezados
                headers = ['ID', 'Teléfono', 'Tipo', 'Estado', 'Mensaje', 'Fecha', 'ID Mensaje']
                ws_whatsapp.append(headers)
                
                # Aplicar estilos a encabezados
                for cell in ws_whatsapp[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Datos
                for log in queryset:
                    fecha_str = log.fecha.strftime('%Y-%m-%d %H:%M:%S') if log.fecha else ''
                    tipo = '📥 Entrante' if log.estado == 'INCOMING' else '📤 Saliente'
                    row = [
                        log.id,
                        log.telefono,
                        tipo,
                        log.estado,
                        log.mensaje or '',
                        fecha_str,
                        log.mensaje_id or ''
                    ]
                    ws_whatsapp.append(row)
                
                # Ajustar ancho de columnas
                ws_whatsapp.column_dimensions['A'].width = 8
                ws_whatsapp.column_dimensions['B'].width = 15
                ws_whatsapp.column_dimensions['C'].width = 15
                ws_whatsapp.column_dimensions['D'].width = 12
                ws_whatsapp.column_dimensions['E'].width = 50
                ws_whatsapp.column_dimensions['F'].width = 20
                ws_whatsapp.column_dimensions['G'].width = 25
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            response['Content-Disposition'] = f'attachment; filename="Reporte_Eki_{fecha_str}.xlsx"'
            wb.save(response)
            return response
        
        except Exception as e:
            context['error'] = f"Error al generar reporte: {str(e)}"
    
    # GET: mostrar formulario
    # Calcular primer día del mes actual y último día
    hoy = datetime.now()
    primer_dia_mes = hoy.replace(day=1)
    if hoy.month == 12:
        ultimo_dia_mes = primer_dia_mes.replace(year=hoy.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        ultimo_dia_mes = primer_dia_mes.replace(month=hoy.month + 1, day=1) - timedelta(days=1)
    
    context['fecha_inicio_default'] = primer_dia_mes.strftime('%Y-%m-%d')
    context['fecha_fin_default'] = ultimo_dia_mes.strftime('%Y-%m-%d')
    
    return render(request, 'admin/descargar_reportes.html', context)


# ---------- Vista de importar estudiantes ----------
@staff_member_required
def importar_estudiantes(request):
    """Vista para importar estudiantes desde Excel."""
    context = {}
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        
        if not archivo:
            context['error'] = "Por favor selecciona un archivo Excel"
            return render(request, 'admin/importar_estudiantes.html', context)
        
        try:
            # Cargar el archivo Excel
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            estudiantes_creados = 0
            estudiantes_actualizados = 0
            errores = []
            
            # Leer filas (columna A = nombre, columna B = teléfono)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    nombre = row[0]
                    telefono = row[1]
                    
                    if not nombre or not telefono:
                        continue
                    
                    # Normalizar teléfono
                    telefono_str = str(telefono).strip()
                    
                    # Crear o actualizar estudiante
                    estudiante, creado = Estudiante.objects.update_or_create(
                        telefono=telefono_str,
                        defaults={'nombre': str(nombre).strip(), 'activo': True}
                    )
                    
                    if creado:
                        estudiantes_creados += 1
                    else:
                        estudiantes_actualizados += 1
                        
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
            
            context['exito'] = True
            context['creados'] = estudiantes_creados
            context['actualizados'] = estudiantes_actualizados
            context['total'] = estudiantes_creados + estudiantes_actualizados
            
            if errores:
                context['advertencias'] = errores[:10]  # Mostrar primeras 10
            
        except Exception as e:
            context['error'] = f"Error al procesar el archivo: {str(e)}"
    
    return render(request, 'admin/importar_estudiantes.html', context)


# ---------- Webhook para WhatsApp Cloud API ----------
@csrf_exempt
def whatsapp_webhook(request):
    """GET: Verificación del token (hub.verify_token).
       POST: Procesa mensajes entrantes, detecta intención y responde dinámicamente.
       Soporta: Drip Content, Geolocalización, Preguntas Abiertas.
    """
    if request.method == 'GET':
        verify_token = request.GET.get('hub.verify_token')
        challenge = request.GET.get('hub.challenge')
        expected = getattr(settings, 'WHATSAPP_VERIFY_TOKEN', None)
        if verify_token and expected and verify_token == expected:
            return HttpResponse(challenge)
        return HttpResponse('Forbidden', status=403)

    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except Exception:
            return JsonResponse({'ok': False, 'error': 'invalid_json'}, status=400)

        entries = payload.get('entry', [])

        for entry in entries:
            changes = entry.get('changes', [])
            for change in changes:
                value = change.get('value', {})

                # Mensajes entrantes
                messages = value.get('messages', [])
                for m in messages:
                    phone = m.get('from')
                    msg_id = m.get('id')
                    msg_type = m.get('type', 'text')

                    # ── Extraer texto o coordenadas según tipo de mensaje ──
                    text = ''
                    latitude = None
                    longitude = None

                    if msg_type == 'text' and isinstance(m.get('text'), dict):
                        text = m['text'].get('body', '')
                    elif msg_type == 'location' and isinstance(m.get('location'), dict):
                        latitude = m['location'].get('latitude')
                        longitude = m['location'].get('longitude')

                    # 1. Guardar registro entrante
                    WhatsappLog.objects.create(
                        telefono=phone,
                        mensaje=text if text else f'[{msg_type}]',
                        mensaje_id=msg_id,
                        estado='INCOMING'
                    )

                    # 2. Obtener datos del estudiante
                    try:
                        estudiante = Estudiante.objects.get(telefono=phone)
                        nombre_usuario = estudiante.nombre
                    except Estudiante.DoesNotExist:
                        estudiante = None
                        nombre_usuario = 'Estudiante'

                    # 3. Procesar mensaje de ubicación (Modo Pokémon GO)
                    if msg_type == 'location' and latitude is not None and longitude is not None:
                        texto_respuesta = _procesar_ubicacion(
                            estudiante, nombre_usuario, latitude, longitude
                        )
                        _enviar_y_loguear(phone, texto_respuesta)
                        continue

                    # 4. Detectar intención del texto
                    intent = detect_intent(text)

                    # 5. Lógica según intent
                    texto_respuesta = _procesar_intent(
                        intent, text, estudiante, nombre_usuario
                    )

                    # 6. Enviar respuesta
                    _enviar_y_loguear(phone, texto_respuesta)

                # Estados (delivery receipts)
                statuses = value.get('statuses', [])
                for s in statuses:
                    msg_id = s.get('id')
                    status = s.get('status')
                    if msg_id:
                        WhatsappLog.objects.filter(mensaje_id=msg_id).update(estado=status)

        return JsonResponse({'ok': True})


# ── Helpers del webhook ──────────────────────────────────────────────────────

def _enviar_y_loguear(phone: str, texto: str):
    """Envía un mensaje WhatsApp y registra el resultado."""
    resultado = enviar_whatsapp(phone, texto)
    if resultado.get('success'):
        WhatsappLog.objects.create(
            telefono=phone,
            mensaje=texto,
            mensaje_id=resultado.get('mensaje_id'),
            estado='SENT'
        )


def _procesar_ubicacion(estudiante, nombre_usuario: str, lat: float, lon: float) -> str:
    """
    Procesa un mensaje de ubicación enviado por WhatsApp.
    Busca el aliado de empleabilidad más cercano (con vacantes activas)
    y retorna el mensaje apropiado según la distancia.
    """
    aliados = AliadoEmpleabilidad.objects.filter(vacantes_activas=True)
    if not aliados.exists():
        return get_response_for_intent('ubicacion_lejos', nombre_usuario, sector='el área central')

    # Calcular distancia a cada aliado
    aliado_cercano = None
    distancia_min = float('inf')
    for aliado in aliados:
        dist = haversine_distance(lat, lon, aliado.latitud, aliado.longitud)
        if dist < distancia_min:
            distancia_min = dist
            aliado_cercano = aliado

    if distancia_min <= 100:
        return get_response_for_intent(
            'ubicacion_cerca', nombre_usuario,
            metros=int(distancia_min),
            empresa=aliado_cercano.nombre_empresa
        )
    else:
        return get_response_for_intent('ubicacion_lejos', nombre_usuario, sector='el área central')


def _procesar_intent(intent: str, text: str, estudiante, nombre_usuario: str) -> str:
    """
    Genera la respuesta según el intent detectado.
    Maneja: drip content, preguntas abiertas, códigos secretos y flujo estándar.
    """
    # ── Continuar Lección (Drip Content) ──
    if intent == 'continuar_leccion' and estudiante:
        return _procesar_continuar_leccion(estudiante, nombre_usuario)

    # ── Código secreto (Gamificación) ──
    if estudiante and text.strip():
        respuesta_codigo = _verificar_codigo_secreto(estudiante, nombre_usuario, text.strip())
        if respuesta_codigo:
            return respuesta_codigo

    # ── Pregunta Abierta pendiente ──
    if estudiante:
        respuesta_pregunta = _procesar_pregunta_abierta(estudiante, nombre_usuario, text)
        if respuesta_pregunta:
            return respuesta_pregunta

    # ── Flujo estándar ──
    if estudiante:
        total_envios = EnvioLog.objects.filter(estudiante=estudiante).count()
        exitosos = EnvioLog.objects.filter(estudiante=estudiante, estado='ENVIADO').count()
        progreso_porcentaje = int((exitosos / total_envios * 100) if total_envios > 0 else 0)

        siguiente_tarea_obj = EnvioLog.objects.filter(
            estudiante=estudiante, estado='PENDIENTE'
        ).order_by('fecha_envio').first()
        siguiente_tarea = (
            siguiente_tarea_obj.campana.nombre if siguiente_tarea_obj else "No hay tareas pendientes"
        )

        datos_respuesta = {
            'progreso': f'{progreso_porcentaje}%',
            'modulo_actual': 'Introducción a la Plataforma',
            'siguiente_tarea': siguiente_tarea,
            'fecha_vence': 'hoy',
        }
    else:
        datos_respuesta = {}

    return get_response_for_intent(intent, nombre_usuario, **datos_respuesta)


def _procesar_continuar_leccion(estudiante, nombre_usuario: str) -> str:
    """
    Evalúa el Drip Content: si el estudiante puede avanzar, envía el siguiente módulo
    y actualiza fecha_ultimo_avance; si no, retorna mensaje de bloqueo.
    """
    progreso = ProgresoEstudiante.objects.filter(
        estudiante=estudiante, estado='en_curso'
    ).select_related('curso', 'modulo_actual').first()

    if not progreso:
        return get_response_for_intent('desconocido', nombre_usuario)

    puede, fecha_desbloqueo = progreso.puede_avanzar()

    if not puede:
        fecha_str = fecha_desbloqueo.strftime('%d/%m/%Y') if fecha_desbloqueo else '?'
        return get_response_for_intent(
            'continuar_leccion_bloqueado', nombre_usuario, fecha_desbloqueo=fecha_str
        )

    # Avanzar al siguiente módulo
    modulos = list(progreso.curso.modulos.order_by('orden'))
    modulo_actual = progreso.modulo_actual
    idx_actual = next((i for i, m in enumerate(modulos) if m == modulo_actual), -1)
    idx_siguiente = idx_actual + 1

    if idx_siguiente < len(modulos):
        modulo_siguiente = modulos[idx_siguiente]
        progreso.modulo_actual = modulo_siguiente
        progreso.fecha_ultimo_avance = timezone.now()
        progreso.save(update_fields=['modulo_actual', 'fecha_ultimo_avance'])

        # Verificar si el nuevo módulo es el último y tiene preguntas abiertas
        es_ultimo_modulo = (idx_siguiente == len(modulos) - 1)
        if es_ultimo_modulo:
            pregunta = modulo_siguiente.preguntas_abiertas.filter(activa=True).order_by('orden').first()
            if pregunta:
                return get_response_for_intent(
                    'pregunta_abierta', nombre_usuario, pregunta=pregunta.pregunta
                )

        return get_response_for_intent(
            'continuar_leccion_libre', nombre_usuario, modulo_siguiente=modulo_siguiente.nombre
        )
    else:
        # Curso completado
        progreso.estado = 'completado'
        progreso.fecha_ultimo_avance = timezone.now()
        progreso.save(update_fields=['estado', 'fecha_ultimo_avance'])

        LogroEstudiante.objects.get_or_create(
            estudiante=estudiante,
            tipo='curso',
            defaults={'descripcion': f'Completó el curso: {progreso.curso.nombre}'}
        )

        return get_response_for_intent(
            'continuar_leccion_completado', nombre_usuario, curso=progreso.curso.nombre
        )


def _verificar_codigo_secreto(estudiante, nombre_usuario: str, text: str):
    """
    Verifica si el texto enviado corresponde a un código secreto de algún aliado.
    Retorna el mensaje de éxito/error si coincide, o None si no es un código.
    """
    try:
        aliado = AliadoEmpleabilidad.objects.get(codigo_secreto__iexact=text, vacantes_activas=True)
    except AliadoEmpleabilidad.DoesNotExist:
        return None

    # Registrar logro
    LogroEstudiante.objects.get_or_create(
        estudiante=estudiante,
        tipo='empleabilidad',
        aliado=aliado,
        defaults={'descripcion': f'Conexión laboral con {aliado.nombre_empresa}'}
    )

    return get_response_for_intent('codigo_correcto', nombre_usuario, empresa=aliado.nombre_empresa)


def _procesar_pregunta_abierta(estudiante, nombre_usuario: str, text: str):
    """
    Si hay una pregunta abierta pendiente de respuesta para el estudiante,
    registra la respuesta y retorna el mensaje de confirmación.
    Retorna None si no hay preguntas abiertas pendientes.
    """
    progreso = ProgresoEstudiante.objects.filter(
        estudiante=estudiante, estado__in=['en_curso', 'completado']
    ).select_related('modulo_actual').first()

    if not progreso or not progreso.modulo_actual:
        return None

    preguntas = progreso.modulo_actual.preguntas_abiertas.filter(activa=True).order_by('orden')
    for pregunta in preguntas:
        ya_respondio = RespuestaAbierta.objects.filter(
            pregunta=pregunta, estudiante=estudiante
        ).exists()
        if not ya_respondio:
            RespuestaAbierta.objects.create(
                pregunta=pregunta,
                estudiante=estudiante,
                respuesta=text
            )
            return get_response_for_intent('respuesta_registrada', nombre_usuario)

    return None
