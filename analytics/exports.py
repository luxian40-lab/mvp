"""
Exportación Excel del dashboard unificado — todas las pestañas y gráficos como datos tabulares.
"""

from __future__ import annotations

from datetime import datetime

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='3b5bdb', end_color='3b5bdb', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _style_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _write_kv_sheet(ws, title: str, pairs: list[tuple[str, object]]):
    ws.title = title[:31]
    _style_header(ws, ['Indicador', 'Valor'])
    for idx, (k, v) in enumerate(pairs, 2):
        ws.cell(row=idx, column=1, value=k)
        ws.cell(row=idx, column=2, value=v)


def _write_chart_series(ws, title: str, labels: list, values: list, label_col='Etiqueta', value_col='Valor'):
    ws.title = title[:31]
    _style_header(ws, [label_col, value_col])
    for idx, (lbl, val) in enumerate(zip(labels, values), 2):
        ws.cell(row=idx, column=1, value=lbl)
        ws.cell(row=idx, column=2, value=val)


def _sheet_estudiantes(wb, estudiantes_detalle):
    ws = wb.create_sheet('Estudiantes B2B')
    headers = [
        'Nombre', 'Cédula', 'Teléfono', 'Organización', 'Municipio', 'Grupo(s)', 'Curso',
        'Módulo actual', 'Módulos', 'Estado avance', 'Avance %', 'Puntos',
    ]
    _style_header(ws, headers)
    for row_idx, est in enumerate(estudiantes_detalle, 2):
        ws.cell(row=row_idx, column=1, value=est['nombre'])
        ws.cell(row=row_idx, column=2, value=est['cedula'])
        ws.cell(row=row_idx, column=3, value=est.get('telefono', '-'))
        ws.cell(row=row_idx, column=4, value=est['organizacion'])
        ws.cell(row=row_idx, column=5, value=est['municipio'])
        ws.cell(row=row_idx, column=6, value=est.get('grupos', '-'))
        ws.cell(row=row_idx, column=7, value=est['curso'])
        ws.cell(row=row_idx, column=8, value=est.get('modulo_actual', '-'))
        ws.cell(row=row_idx, column=9, value=est.get('modulos_completados', '-'))
        ws.cell(row=row_idx, column=10, value=est.get('estado_avance', '-'))
        ws.cell(row=row_idx, column=11, value=est['avance'])
        ws.cell(row=row_idx, column=12, value=est['puntos'])


def _sheet_executive(wb, context, resumen_payload):
    kpis = resumen_payload.get('kpis', {})
    ws_meta = wb.active
    ws_meta.title = 'Executive KPIs'
    meta_pairs = [
        ('Pestaña exportada', 'executive'),
        ('Generado', resumen_payload.get('generated_at', '')),
        ('Organización ID', context.get('cliente_filtro') or 'Todas'),
        ('Curso ID', context.get('curso_filtro') or 'Todos'),
        ('Desde', context.get('fecha_inicio') or '—'),
        ('Hasta', context.get('fecha_fin') or '—'),
        ('Estudiantes activos', kpis.get('total_estudiantes')),
        ('Tasa completación %', kpis.get('tasa_completacion')),
        ('Certificados', kpis.get('total_certificados')),
        ('Cursos activos', kpis.get('total_cursos')),
        ('Organizaciones', kpis.get('total_clientes')),
        ('Leads B2B', kpis.get('total_prospectos')),
        ('Módulos completados', kpis.get('total_modulos_completados')),
        ('Puntos promedio gamificación', kpis.get('puntos_promedio')),
    ]
    _write_kv_sheet(ws_meta, 'Executive KPIs', meta_pairs)

    ws_curso = wb.create_sheet('Progreso por curso')
    _style_header(ws_curso, ['Curso', 'Estudiantes', 'Completados', '%'])
    for row_idx, curso in enumerate(context.get('progreso_por_curso', []), 2):
        total = curso.total_estudiantes or 0
        comp = curso.completados or 0
        pct = round(comp / total * 100, 1) if total else 0
        ws_curso.cell(row=row_idx, column=1, value=curso.nombre)
        ws_curso.cell(row=row_idx, column=2, value=total)
        ws_curso.cell(row=row_idx, column=3, value=comp)
        ws_curso.cell(row=row_idx, column=4, value=pct)

    ws_rank = wb.create_sheet('Ranking gamificación')
    _style_header(ws_rank, ['#', 'Nombre', 'Cédula', 'Puntos', 'Organización'])
    for row_idx, perfil in enumerate(context.get('ranking_gamificacion_completo', [])[:500], 2):
        est = perfil.estudiante
        ws_rank.cell(row=row_idx, column=1, value=row_idx - 1)
        ws_rank.cell(row=row_idx, column=2, value=est.nombre if est else '—')
        ws_rank.cell(row=row_idx, column=3, value=est.cedula if est else '—')
        ws_rank.cell(row=row_idx, column=4, value=perfil.puntos_totales)
        ws_rank.cell(
            row=row_idx, column=5,
            value=est.cliente.nombre if est and est.cliente_id else '—',
        )

    chart_msg = resumen_payload.get('chart_mensajes', {})
    ws_msg = wb.create_sheet('Gráfico mensajes 7d')
    _write_chart_series(
        ws_msg, 'Gráfico mensajes 7d',
        chart_msg.get('labels', []),
        chart_msg.get('values', []),
        'Día', 'Mensajes',
    )

    chart_ubi = resumen_payload.get('chart_ubicaciones', {})
    ws_ubi = wb.create_sheet('Gráfico municipios')
    _write_chart_series(
        ws_ubi, 'Gráfico municipios',
        chart_ubi.get('labels', []),
        chart_ubi.get('values', []),
        'Municipio', 'Estudiantes',
    )

    chart_tipos = resumen_payload.get('chart_tipos', {})
    ws_tipos = wb.create_sheet('Gráfico tipos WA')
    _write_chart_series(
        ws_tipos, 'Gráfico tipos WA',
        chart_tipos.get('labels', []),
        chart_tipos.get('values', []),
        'Tipo mensaje', 'Cantidad',
    )


def _sheet_ai_ops(wb, context, resumen_payload):
    ws = wb.active
    kpis = resumen_payload.get('kpis', {})
    pairs = [
        ('Pestaña exportada', 'ai_ops'),
        ('Mensajes WhatsApp total', kpis.get('total_mensajes_whatsapp')),
        ('Enviados', kpis.get('mensajes_enviados')),
        ('Recibidos', kpis.get('mensajes_recibidos')),
        ('Entregados (DELIVERED)', kpis.get('wa_entregados')),
        ('Leídos (READ)', kpis.get('wa_leidos')),
        ('En tránsito', kpis.get('wa_en_transito')),
        ('Bot comercial enviados', kpis.get('wa_bot_comercial_sent')),
        ('Bot comercial leídos', kpis.get('wa_bot_comercial_read')),
        ('Audios', kpis.get('total_audios')),
        ('Activaciones agentes IA', kpis.get('total_agentes_ia')),
    ]
    _write_kv_sheet(ws, 'AI Ops resumen', pairs)

    chart_msg = resumen_payload.get('chart_mensajes', {})
    ws2 = wb.create_sheet('Actividad 7 días')
    _write_chart_series(ws2, 'Actividad 7 días', chart_msg.get('labels', []), chart_msg.get('values', []))

    chart_tipos = resumen_payload.get('chart_tipos', {})
    ws3 = wb.create_sheet('Tipos mensaje')
    _write_chart_series(ws3, 'Tipos mensaje', chart_tipos.get('labels', []), chart_tipos.get('values', []))

    ws_ev = wb.create_sheet('Eventos IA recientes')
    _style_header(ws_ev, ['Fecha', 'Tipo', 'Trace', 'Agente', 'Canal', 'Regla', 'Input', 'Output'])
    for row_idx, ev in enumerate(context.get('eventos_ia_recientes', [])[:200], 2):
        ws_ev.cell(row=row_idx, column=1, value=ev.created_at.strftime('%Y-%m-%d %H:%M') if ev.created_at else '')
        ws_ev.cell(row=row_idx, column=2, value=ev.get_tipo_display() if hasattr(ev, 'get_tipo_display') else ev.tipo)
        ws_ev.cell(row=row_idx, column=3, value=str(ev.trace_id))
        ws_ev.cell(row=row_idx, column=4, value=ev.agente or '')
        ws_ev.cell(row=row_idx, column=5, value=ev.canal or '')
        ws_ev.cell(row=row_idx, column=6, value=ev.regla_aplicada or '')
        ws_ev.cell(row=row_idx, column=7, value=(ev.input_preview or '')[:300])
        ws_ev.cell(row=row_idx, column=8, value=(ev.output_preview or '')[:300])


def _sheet_commercial(wb, context):
    from core.domains.analytics.metricas import calcular_metricas_nati

    ws = wb.active
    cid = context.get('cliente_filtro')
    desde = context.get('fecha_inicio') or None
    hasta = context.get('fecha_fin') or None
    payload = calcular_metricas_nati(cliente_id=cid, desde=desde, hasta=hasta)

    pairs = [('Pestaña exportada', 'commercial')]
    resumen = payload.get('resumen') or {}
    for k, v in resumen.items():
        pairs.append((k.replace('_', ' ').title(), v))
    pcts = payload.get('porcentajes') or {}
    for k, v in pcts.items():
        pairs.append((f'Pct {k}', v))
    _write_kv_sheet(ws, 'Nati KPIs', pairs)

    ws2 = wb.create_sheet('Nati por día')
    serie = (payload.get('series') or {}).get('temporal') or []
    _style_header(ws2, ['Fecha', 'Enviados', 'Leídos', 'Recibidos'])
    for row_idx, row in enumerate(serie, 2):
        ws2.cell(row=row_idx, column=1, value=row.get('fecha', ''))
        ws2.cell(row=row_idx, column=2, value=row.get('enviados', 0))
        ws2.cell(row=row_idx, column=3, value=row.get('leidos', 0))
        ws2.cell(row=row_idx, column=4, value=row.get('recibidos', 0))

    semaforos = payload.get('semaforos') or {}
    labels = payload.get('semaforo_labels') or {}
    if semaforos:
        ws3 = wb.create_sheet('Semáforos Nati')
        _style_header(ws3, ['Métrica', 'Estado', 'Etiqueta'])
        metas = payload.get('metas') or {}
        for row_idx, (nombre, estado) in enumerate(semaforos.items(), 2):
            ws3.cell(row=row_idx, column=1, value=nombre)
            ws3.cell(row=row_idx, column=2, value=estado)
            ws3.cell(row=row_idx, column=3, value=labels.get(nombre, ''))
            if nombre in metas:
                ws3.cell(row=row_idx, column=4, value=f"Meta: {metas[nombre]}")


def _sheet_learning_empresa(wb, context):
    from core.domains.analytics.metricas import calcular_metricas_empresa

    cid = context.get('cliente_filtro')
    curso_id = context.get('curso_filtro')
    payload = calcular_metricas_empresa(
        cliente_id=cid,
        curso_id=curso_id,
        desde=context.get('fecha_inicio') or None,
        hasta=context.get('fecha_fin') or None,
    )

    ws = wb.active
    pairs = [('Pestaña', 'learning / métricas empresa')]
    for k, v in (payload.get('resumen') or {}).items():
        pairs.append((k.replace('_', ' ').title(), v))
    for k, v in (payload.get('porcentajes') or {}).items():
        pairs.append((f'Pct {k}', v))
    _write_kv_sheet(ws, 'Métricas empresa', pairs)

    distrib = (payload.get('series') or {}).get('distribucion_avance') or []
    if distrib:
        ws2 = wb.create_sheet('Distribución avance')
        _style_header(ws2, ['Estado', 'Cantidad'])
        for row_idx, row in enumerate(distrib, 2):
            ws2.cell(row=row_idx, column=1, value=row.get('label', row.get('estado', '')))
            ws2.cell(row=row_idx, column=2, value=row.get('total', row.get('cantidad', 0)))

    serie = (payload.get('series') or {}).get('temporal') or []
    ws3 = wb.create_sheet('Actividad diaria')
    _style_header(ws3, ['Fecha', 'Mensajes enviados', 'Avances'])
    for row_idx, row in enumerate(serie, 2):
        ws3.cell(row=row_idx, column=1, value=row.get('fecha', ''))
        ws3.cell(row=row_idx, column=2, value=row.get('mensajes_enviados', 0))
        ws3.cell(row=row_idx, column=3, value=row.get('avances', 0))

    ws4 = wb.create_sheet('Por organización')
    _style_header(ws4, ['Organización', 'Cursos', 'Estudiantes', 'Uso audio', 'Uso IA', 'Completados'])
    for row_idx, c in enumerate(context.get('clientes_detalle', []), 2):
        ws4.cell(row=row_idx, column=1, value=c['nombre'])
        ws4.cell(row=row_idx, column=2, value=c['cursos'])
        ws4.cell(row=row_idx, column=3, value=c['estudiantes'])
        ws4.cell(row=row_idx, column=4, value=c['uso_audio'])
        ws4.cell(row=row_idx, column=5, value=c['uso_ia'])
        ws4.cell(row=row_idx, column=6, value=c['cursos_completados'])


def export_dashboard_excel(
    *,
    context: dict,
    resumen_payload: dict,
    tab: str = 'executive',
    learning_section: str = 'reportes',
) -> HttpResponse:
    """Genera workbook multi-hoja según pestaña activa del dashboard."""
    wb = openpyxl.Workbook()

    if tab == 'executive':
        _sheet_executive(wb, context, resumen_payload)
    elif tab == 'ai_ops':
        _sheet_ai_ops(wb, context, resumen_payload)
    elif tab == 'commercial':
        _sheet_commercial(wb, context)
    elif tab == 'learning':
        if learning_section == 'metricas_empresa':
            _sheet_learning_empresa(wb, context)
        else:
            wb.active.title = 'Resumen'
            _write_kv_sheet(wb.active, 'Resumen', [('Pestaña', 'learning / reportes B2B')])
            _sheet_estudiantes(wb, context.get('estudiantes_detalle', []))
    else:
        _sheet_executive(wb, context, resumen_payload)
        _sheet_estudiantes(wb, context.get('estudiantes_detalle', []))

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'eki_dashboard_{tab}_{fecha_str}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
