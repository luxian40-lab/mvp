"""Asistencia y calificaciones del aula docente (portal profesor)."""

from __future__ import annotations

from datetime import date
from decimal import Decimal, InvalidOperation

from django.db.models import Count, F, Sum
from django.utils.dateparse import parse_date

from core.gamificacion import EvaluacionNotaGamificacion, PerfilGamificacion
from core.gamificacion_modo import (
    gamificacion_activa,
    curso_usa_calificacion,
    curso_usa_puntos,
    registrar_nota_gamificacion,
    resumen_calificaciones_estudiante,
)
from core.models import Curso, Estudiante, ProgresoEstudiante

from .models import AsistenciaAula, EntregaTarea, TareaCurso

PESO_ASISTENCIA = Decimal('1')
PASO_FALTA = Decimal('0.2')
NOTA_ASISTENCIA_MAX = Decimal('5')
NOTA_ASISTENCIA_MIN = Decimal('1')
DETALLE_ASISTENCIA_CURSO = 'Asistencia del curso'


def estudiantes_inscritos_curso(curso: Curso):
    """Estudiantes activos con progreso en el curso."""
    return (
        Estudiante.objects.filter(
            progresos__curso=curso,
            activo=True,
        )
        .distinct()
        .order_by('nombre')
    )


def nota_asistencia_desde_faltas(faltas: int) -> Decimal:
    """
    Nota cerrada 1–5: parte en 5 y resta 0,2 por falta (piso 1).
    0 faltas→5 · 1→4,8 · 5→4,0 · 20+→1.
    """
    n = int(faltas or 0)
    if n < 0:
        n = 0
    raw = NOTA_ASISTENCIA_MAX - (Decimal(n) * PASO_FALTA)
    if raw < NOTA_ASISTENCIA_MIN:
        raw = NOTA_ASISTENCIA_MIN
    return raw.quantize(Decimal('0.1'))


def _detalle_asistencia(fecha: date) -> str:
    return f'Asistencia {fecha.isoformat()}'


def resumen_asistencia_estudiante(curso: Curso, estudiante_id: int) -> dict:
    agg = AsistenciaAula.objects.filter(curso=curso, estudiante_id=estudiante_id).aggregate(
        total=Count('id'),
        presentes=Count('id', filter=F('presente')),
    )
    total = int(agg['total'] or 0)
    presentes = int(agg['presentes'] or 0)
    faltas = max(0, total - presentes)
    return {
        'total': total,
        'presentes': presentes,
        'faltas': faltas,
        'nota': nota_asistencia_desde_faltas(faltas) if total else None,
    }


def recalcular_nota_asistencia_estudiante(curso: Curso, estudiante: Estudiante) -> Decimal | None:
    """
    Una sola nota de asistencia por estudiante/curso (modo calificación).
    Borra notas por-día legacy y escribe el agregado.
    """
    cliente = curso.cliente
    if not cliente or not curso_usa_calificacion(cliente, curso):
        return None

    EvaluacionNotaGamificacion.objects.filter(
        estudiante=estudiante,
        curso=curso,
        tipo='asistencia',
    ).delete()

    asist = resumen_asistencia_estudiante(curso, estudiante.pk)
    total = int(asist['total'] or 0)
    if total <= 0:
        return None

    nota = asist['nota']
    registrar_nota_gamificacion(
        estudiante,
        nota,
        'asistencia',
        curso=curso,
        detalle=DETALLE_ASISTENCIA_CURSO,
        peso=PESO_ASISTENCIA,
    )
    return nota


def recalcular_notas_asistencia_curso(curso: Curso) -> int:
    """Recalcula asistencia agregada para todos los inscritos. Devuelve # con nota."""
    n = 0
    for est in estudiantes_inscritos_curso(curso):
        if recalcular_nota_asistencia_estudiante(curso, est) is not None:
            n += 1
    return n


def _sync_asistencia_gamificacion(asistencia: AsistenciaAula) -> None:
    """Tras marcar un día: puntos (+1 si presente) o recalcular nota agregada 1–5."""
    cliente = asistencia.curso.cliente
    if not cliente:
        return
    if curso_usa_calificacion(cliente, asistencia.curso):
        recalcular_nota_asistencia_estudiante(asistencia.curso, asistencia.estudiante)
        return
    if curso_usa_puntos(cliente, asistencia.curso) and asistencia.presente:
        perfil, _ = PerfilGamificacion.objects.get_or_create(estudiante=asistencia.estudiante)
        motivo = _detalle_asistencia(asistencia.fecha)
        if not perfil.transacciones.filter(razon=motivo).exists():
            perfil.ajustar_puntos_manual(1, motivo)


def sincronizar_nota_tarea_entrega(entrega: EntregaTarea) -> None:
    """Refleja la nota de tarea en el ranking por calificaciones (peso 1)."""
    if entrega.nota is None:
        EvaluacionNotaGamificacion.objects.filter(
            estudiante=entrega.estudiante,
            curso=entrega.tarea.curso,
            tipo='tarea_aula',
            detalle=f'Tarea: {entrega.tarea.titulo}'[:200],
        ).delete()
        return

    cliente = entrega.tarea.curso.cliente
    if not cliente or not curso_usa_calificacion(cliente, entrega.tarea.curso):
        return

    detalle = f'Tarea: {entrega.tarea.titulo}'[:200]
    prev = EvaluacionNotaGamificacion.objects.filter(
        estudiante=entrega.estudiante,
        curso=entrega.tarea.curso,
        tipo='tarea_aula',
        detalle=detalle,
    ).first()
    if prev:
        prev.nota = Decimal(str(entrega.nota))
        prev.peso = Decimal('1')
        prev.save(update_fields=['nota', 'peso'])
    else:
        registrar_nota_gamificacion(
            entrega.estudiante,
            entrega.nota,
            'tarea_aula',
            curso=entrega.tarea.curso,
            detalle=detalle,
            peso=Decimal('1'),
        )


def guardar_asistencia_sesion(request, curso: Curso, fecha: date, presente_ids: set[int]) -> int:
    """Registra asistencia de todos los inscritos para una fecha. Devuelve # presentes."""
    pu = getattr(request, 'portal_usuario', None)
    user = pu.user if pu else None
    count = 0
    for est in estudiantes_inscritos_curso(curso):
        presente = est.pk in presente_ids
        if presente:
            count += 1
        asistencia, _ = AsistenciaAula.objects.update_or_create(
            curso=curso,
            estudiante=est,
            fecha=fecha,
            defaults={'presente': presente, 'registrado_por': user},
        )
        _sync_asistencia_gamificacion(asistencia)
    return count


def borrar_asistencia_sesion(curso: Curso, fecha: date) -> int:
    """
    Elimina el registro de asistencia de ese día y recalcula la nota agregada.
    Útil cuando el profesor cargó mal la fecha o los presentes.
    """
    from core.gamificacion import TransaccionPuntos

    detalle = _detalle_asistencia(fecha)
    afectados = list(
        AsistenciaAula.objects.filter(curso=curso, fecha=fecha).values_list(
            'estudiante_id', flat=True
        )
    )
    n = len(afectados)
    AsistenciaAula.objects.filter(curso=curso, fecha=fecha).delete()

    cliente = curso.cliente
    if cliente and curso_usa_calificacion(cliente, curso):
        # Legacy por-día + agregado: se reescribe en el recálculo.
        for est in Estudiante.objects.filter(pk__in=afectados):
            recalcular_nota_asistencia_estudiante(curso, est)
        # Si nadie quedó con sesiones, limpia huérfanos de esa fecha legacy.
        EvaluacionNotaGamificacion.objects.filter(
            curso=curso,
            tipo='asistencia',
            detalle=detalle,
        ).delete()
    elif cliente and curso_usa_puntos(cliente, curso):
        razon = f'Manual: {detalle}'
        for tx in TransaccionPuntos.objects.filter(razon=razon).select_related('perfil'):
            perfil = tx.perfil
            perfil.puntos_totales = max(0, perfil.puntos_totales - abs(tx.puntos))
            perfil.calcular_nivel()
            perfil.save()
            tx.delete()
    return n


def filas_asistencia_curso(curso: Curso, fecha: date) -> list[dict]:
    """Lista inscritos con checkbox de asistencia para una fecha."""
    registros = {
        a.estudiante_id: a
        for a in AsistenciaAula.objects.filter(curso=curso, fecha=fecha)
    }
    filas = []
    for est in estudiantes_inscritos_curso(curso):
        reg = registros.get(est.pk)
        filas.append({
            'estudiante': est,
            'presente': reg.presente if reg else False,
            'registrado': reg is not None,
        })
    return filas


def filas_calificacion_curso(curso: Curso) -> list[dict]:
    """Resumen por estudiante: asistencia, tareas, evaluaciones y promedio."""
    filas = []
    for est in estudiantes_inscritos_curso(curso):
        asist = resumen_asistencia_estudiante(curso, est.pk)
        entregas = (
            EntregaTarea.objects.filter(estudiante=est, tarea__curso=curso, nota__isnull=False)
            .select_related('tarea')
            .order_by('tarea__titulo')
        )
        evals = (
            EvaluacionNotaGamificacion.objects.filter(estudiante=est, curso=curso)
            .order_by('-fecha')
        )
        resumen = resumen_calificaciones_estudiante(est, curso_id=curso.pk)
        promedio = float(resumen['promedio']) if resumen['promedio'] is not None else None
        filas.append({
            'estudiante': est,
            'asistencia': asist,
            'entregas': list(entregas),
            'evaluaciones': list(evals),
            'promedio': promedio,
            'num_evaluaciones': resumen['cantidad'],
            'suma_pesos': float(resumen['suma_pesos']) if resumen['suma_pesos'] else 0,
        })
    return filas


def parse_fecha_asistencia(raw: str) -> date | None:
    return parse_date((raw or '').strip())


def fechas_asistencia_marcadas(curso: Curso, fecha: date | None = None) -> list[date]:
    """Días con asistencia registrada. Si `fecha` se indica, solo ese día (si existe)."""
    qs = AsistenciaAula.objects.filter(curso=curso)
    if fecha:
        qs = qs.filter(fecha=fecha)
    return list(qs.values_list('fecha', flat=True).distinct().order_by('fecha'))


def slug_archivo_asistencia(nombre: str) -> str:
    import re
    from unicodedata import normalize

    s = normalize('NFKD', nombre or '').encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[^A-Za-z0-9._-]+', '_', s).strip('._') or 'curso'
    return s[:60]


def generar_excel_asistencia_curso(curso: Curso, fechas: list[date] | None = None) -> bytes:
    """
    Genera un .xlsx con la asistencia de los días marcados.
    Una fila por estudiante y una columna por fecha (Presente/Ausente).
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if fechas is None:
        fechas = fechas_asistencia_marcadas(curso)
    if not fechas:
        raise ValueError('No hay días con asistencia marcada para descargar.')

    inscritos = list(estudiantes_inscritos_curso(curso))
    registros = {
        (a.estudiante_id, a.fecha): a.presente
        for a in AsistenciaAula.objects.filter(curso=curso, fecha__in=fechas)
    }

    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencia'

    headers = ['Estudiante', 'Documento'] + [f.strftime('%d/%m/%Y') for f in fechas] + [
        'Presentes', 'Total días', '% asistencia',
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='9A6CAC', end_color='9A6CAC', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    presente_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    ausente_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for est in inscritos:
        presentes = 0
        estados = []
        for fecha in fechas:
            presente = registros.get((est.pk, fecha), False)
            if presente:
                presentes += 1
            estados.append('Presente' if presente else 'Ausente')
        total = len(fechas)
        pct = round((presentes / total) * 100, 1) if total else 0
        ws.append([est.nombre or '', est.cedula or '', *estados, presentes, total, pct])

        row_idx = ws.max_row
        for col_idx, estado in enumerate(estados, start=3):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = presente_fill if estado == 'Presente' else ausente_fill

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    for i in range(3, 3 + len(fechas)):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 12
    for i in range(3 + len(fechas), 3 + len(fechas) + 3):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 14

    # Hoja detalle largo (útil para filtrar por fecha)
    ws2 = wb.create_sheet('Detalle por día')
    ws2.append(['Fecha', 'Estudiante', 'Documento', 'Estado'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for fecha in fechas:
        for est in inscritos:
            presente = registros.get((est.pk, fecha), False)
            ws2.append([
                fecha.strftime('%d/%m/%Y'),
                est.nombre or '',
                est.cedula or '',
                'Presente' if presente else 'Ausente',
            ])
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def matriz_tareas_calificacion_curso(curso: Curso) -> tuple[list[TareaCurso], list[dict]]:
    """Estudiantes × tareas activas con entrega (si existe)."""
    tareas = list(
        TareaCurso.objects.filter(curso=curso, activa=True).order_by('titulo', 'id')
    )
    entregas = EntregaTarea.objects.filter(tarea__curso=curso).select_related('tarea', 'estudiante')
    entrega_map = {(e.estudiante_id, e.tarea_id): e for e in entregas}
    filas = []
    for est in estudiantes_inscritos_curso(curso):
        celdas = []
        for t in tareas:
            celdas.append({
                'tarea': t,
                'entrega': entrega_map.get((est.pk, t.pk)),
            })
        filas.append({'estudiante': est, 'celdas': celdas})
    return tareas, filas


def calificar_matriz_tareas_curso(request, curso: Curso) -> tuple[int, str | None]:
    """Guarda notas desde matriz (nota_<estudiante_id>_<tarea_id>)."""
    from aprende.tarea_service import _aplicar_calificacion_entrega, _parse_nota

    guardadas = 0
    pu = getattr(request, 'portal_usuario', None)
    calificador = pu.user if pu else None
    entregas = EntregaTarea.objects.filter(tarea__curso=curso).select_related('tarea', 'estudiante')
    for entrega in entregas:
        raw = request.POST.get(f'nota_{entrega.estudiante_id}_{entrega.tarea_id}', '')
        nota, err = _parse_nota(raw)
        if err:
            return guardadas, err
        if nota is None:
            continue
        _aplicar_calificacion_entrega(
            entrega,
            nota,
            request.POST.get(
                f'comentario_{entrega.estudiante_id}_{entrega.tarea_id}', '',
            ).strip(),
            calificador,
        )
        guardadas += 1
    if guardadas == 0:
        return 0, 'Indica al menos una nota para guardar.'
    return guardadas, None


def generar_excel_calificaciones_curso(curso: Curso) -> bytes:
    """Excel: filas estudiantes, columnas tareas + promedio."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    tareas, filas = matriz_tareas_calificacion_curso(curso)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Calificaciones'

    headers = ['Estudiante', 'Documento'] + [t.titulo[:40] for t in tareas] + ['Promedio']
    ws.append(headers)
    header_fill = PatternFill(start_color='9A6CAC', end_color='9A6CAC', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for fila in filas:
        est = fila['estudiante']
        notas: list = []
        for celda in fila['celdas']:
            e = celda['entrega']
            notas.append(e.nota if e and e.nota is not None else '')
        resumen = resumen_calificaciones_estudiante(est, curso_id=curso.pk)
        prom = resumen['promedio']
        prom_txt = float(prom) if prom is not None else ''
        ws.append([est.nombre or '', est.cedula or '', *notas, prom_txt])

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    for i in range(3, 3 + len(tareas) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def actualizar_nota_evaluacion(eval_id: int, curso: Curso, nota_raw: str) -> EvaluacionNotaGamificacion:
    try:
        nota = Decimal(str(nota_raw).replace(',', '.'))
    except (InvalidOperation, ValueError):
        raise ValueError('Nota inválida. Use un número entre 1 y 5.')
    if not Decimal('1') <= nota <= Decimal('5'):
        raise ValueError('La nota debe estar entre 1 y 5.')

    ev = EvaluacionNotaGamificacion.objects.get(pk=eval_id, curso=curso)
    ev.nota = nota
    ev.save(update_fields=['nota'])
    return ev


def registrar_nota_manual_curso(
    curso: Curso,
    estudiante_id: int,
    nota_raw: str,
    detalle: str = '',
    peso_raw: str = '',
) -> dict:
    try:
        est = Estudiante.objects.get(pk=estudiante_id, cliente=curso.cliente, activo=True)
    except Estudiante.DoesNotExist:
        raise ValueError('Estudiante no válido.')

    try:
        nota = Decimal(str(nota_raw).replace(',', '.'))
    except (InvalidOperation, ValueError):
        raise ValueError('Nota inválida.')
    peso = Decimal(str(peso_raw).replace(',', '.')) if peso_raw else Decimal('1')

    registrar_nota_gamificacion(
        est,
        nota,
        'manual',
        curso=curso,
        detalle=(detalle or 'Evaluación docente')[:200],
        peso=peso,
    )
    resumen = resumen_calificaciones_estudiante(est, curso_id=curso.pk)
    return {
        'estudiante': est,
        'promedio': float(resumen['promedio']) if resumen['promedio'] is not None else None,
    }


def contexto_modo_calificacion(org, curso=None) -> dict:
    return {
        'gamif_activa': gamificacion_activa(org) or (
            bool(curso) and getattr(curso, 'es_modo_clases', lambda: False)()
        ),
        'usa_calificacion': curso_usa_calificacion(org, curso),
        'usa_puntos': curso_usa_puntos(org, curso),
    }
