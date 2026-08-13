"""Tests básicos del aula web /aprende/."""

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings

from aprende.models import EntregaTarea, TareaCurso

from core.models import Cliente, Curso, Estudiante, Modulo, ProgresoEstudiante
from portal.models import PortalUsuario


@override_settings(
    SECURE_SSL_REDIRECT=False,
    ALLOWED_HOSTS=['testserver', 'localhost', '127.0.0.1', 'aprende.eki.technology', 'aula.eki.technology'],
    STORAGES={
        'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    },
)
class AprendeWebTests(TestCase):
    def setUp(self):
        from django.core.cache import cache
        cache.clear()
        self.http = Client()
        self.cliente = Cliente.objects.create(
            nombre='Org Aprende',
            contacto_principal='A',
            email='ap@test.com',
            telefono='573009999001',
            activo=True,
        )
        self.curso = Curso.objects.create(nombre='Curso Web', cliente=self.cliente, activo=True)
        self.modulo = Modulo.objects.create(
            curso=self.curso,
            numero=1,
            titulo='Intro',
            descripcion='d',
            contenido='Hola desde la web',
        )
        self.est = Estudiante.objects.create(
            cedula='web1',
            nombre='Est Web',
            telefono='573009999002',
            cliente=self.cliente,
            activo=True,
        )
        ProgresoEstudiante.objects.create(estudiante=self.est, curso=self.curso, modulo_actual=self.modulo)
        self.user = User.objects.create_user('prof_ap', 'p@t.com', 'pass')
        PortalUsuario.objects.create(user=self.user, organizacion=self.cliente, rol='profesor')

    def _login_estudiante(self, telefono='573009999002', cedula='web1', password='clave123'):
        """Simula *aula* → OTP → crear clave (si falta) → sesión."""
        from aprende.acceso_whatsapp import emitir_acceso_desde_whatsapp
        from aprende.credencial_service import tiene_clave

        msg = emitir_acceso_desde_whatsapp(self.est)
        import re
        m = re.search(r'\*(\d{6})\*', msg)
        self.assertIsNotNone(m, msg=msg)
        self.assertIn('/aprende/estudiante/login/', msg)
        self.assertNotIn('/aprende/handoff/', msg)
        self.assertIn('contraseña', msg.lower())
        codigo = m.group(1)
        from aprende.models import CodigoAccesoAprende
        self.assertTrue(CodigoAccesoAprende.objects.filter(codigo=codigo, estudiante=self.est).exists())
        r2 = self.http.post('/aprende/estudiante/login/', {'codigo': codigo, 'accion': 'codigo'})
        self.assertEqual(r2.status_code, 302)
        if not tiene_clave(self.est):
            self.assertIn('/aprende/estudiante/clave/', r2['Location'])
            r3 = self.http.post('/aprende/estudiante/clave/', {
                'password': password,
                'password2': password,
            })
            self.assertEqual(r3.status_code, 302)
            return r3
        return r2

    def test_inicio_carga(self):
        r = self.http.get('/aprende/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'aprende')
        self.assertContains(r, '*aula*')
        self.assertContains(r, 'Cómo entro')
        self.assertNotContains(r, 'Aula virtual')

    def test_login_estudiante_celebra_whatsapp(self):
        r = self.http.get('/aprende/estudiante/login/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Código')
        self.assertContains(r, 'name="codigo"')
        self.assertContains(r, '*aula*')
        self.assertContains(r, 'Documento + clave')
        self.assertContains(r, 'Olvidé')
        self.assertContains(r, 'og:image')
        self.assertContains(r, 'og-aprende-v3')

    def test_sin_clave_otp_pide_crear_password(self):
        from aprende.acceso_whatsapp import emitir_acceso_desde_whatsapp
        import re
        msg = emitir_acceso_desde_whatsapp(self.est)
        codigo = re.search(r'\*(\d{6})\*', msg).group(1)
        r = self.http.post('/aprende/estudiante/login/', {'codigo': codigo, 'accion': 'codigo'})
        self.assertEqual(r.status_code, 302)
        self.assertIn('/aprende/estudiante/clave/', r['Location'])
        r2 = self.http.get('/aprende/estudiante/clave/')
        self.assertEqual(r2.status_code, 200)
        self.assertContains(r2, 'Crea tu contraseña')
        r3 = self.http.post('/aprende/estudiante/clave/', {
            'password': 'secreta1',
            'password2': 'secreta1',
        })
        self.assertEqual(r3.status_code, 302)
        from aprende.credencial_service import tiene_clave
        self.assertTrue(tiene_clave(self.est))
        # Re-login con documento + clave
        self.http.get('/aprende/estudiante/logout/')
        r4 = self.http.post('/aprende/estudiante/login/', {
            'accion': 'clave',
            'documento': 'web1',
            'password': 'secreta1',
        })
        self.assertEqual(r4.status_code, 302)
        r5 = self.http.get('/aprende/estudiante/')
        self.assertEqual(r5.status_code, 200)

    def test_olvide_clave_via_otp_reset(self):
        from aprende.credencial_service import guardar_clave
        guardar_clave(self.est, 'vieja123')
        from aprende.acceso_whatsapp import emitir_acceso_desde_whatsapp
        import re
        codigo = re.search(r'\*(\d{6})\*', emitir_acceso_desde_whatsapp(self.est)).group(1)
        r = self.http.post('/aprende/estudiante/login/', {
            'codigo': codigo,
            'accion': 'codigo',
            'tab': 'olvide',
        })
        self.assertEqual(r.status_code, 302)
        self.assertIn('/clave/', r['Location'])
        r2 = self.http.post('/aprende/estudiante/clave/', {
            'password': 'nueva456',
            'password2': 'nueva456',
        })
        self.assertEqual(r2.status_code, 302)
        self.http.get('/aprende/estudiante/logout/')
        bad = self.http.post('/aprende/estudiante/login/', {
            'accion': 'clave', 'documento': 'web1', 'password': 'vieja123',
        })
        self.assertEqual(bad.status_code, 200)
        ok = self.http.post('/aprende/estudiante/login/', {
            'accion': 'clave', 'documento': 'web1', 'password': 'nueva456',
        })
        self.assertEqual(ok.status_code, 302)

    def test_handoff_crawler_whatsapp_no_consume_token(self):
        from studio.aprende_bridge import crear_token_handoff
        token = crear_token_handoff(estudiante_id=self.est.pk)
        r = self.http.get(
            f'/aprende/handoff/?t={token}',
            HTTP_USER_AGENT='WhatsApp/2.0',
        )
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'og:image')
        self.assertContains(r, 'eki aprende')
        # Token sigue válido: humano puede usarlo
        r2 = self.http.get(f'/aprende/handoff/?t={token}')
        self.assertEqual(r2.status_code, 302)
        self.assertIn('/aprende/estudiante/', r2['Location'])

    def test_estudiante_login_y_ve_modulo(self):
        self._login_estudiante()
        r3 = self.http.get(f'/aprende/estudiante/modulo/{self.modulo.id}/')
        self.assertEqual(r3.status_code, 200)
        self.assertContains(r3, 'Hola desde la web')

    def test_estudiante_home_continuar_y_puente_wa(self):
        self._login_estudiante()
        r = self.http.get('/aprende/estudiante/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Continuar')
        self.assertContains(r, 'Seguir módulo')
        self.assertContains(r, 'WhatsApp + Aprende')
        self.assertContains(r, '*aula*')
        self.assertContains(r, 'is-active')
        self.assertContains(r, self.curso.nombre)

    def test_quiz_web_modulo_practica(self):
        from core.models import PreguntaModulo
        PreguntaModulo.objects.create(
            modulo=self.modulo,
            pregunta='¿Color de eki?',
            opcion_a='Morado',
            opcion_b='Rojo',
            respuesta_correcta='A',
            explicacion='Marca eki',
            activa=True,
        )
        self._login_estudiante()
        r = self.http.get(f'/aprende/estudiante/modulo/{self.modulo.id}/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Práctica del módulo')
        self.assertContains(r, '¿Color de eki?')
        r2 = self.http.post(f'/aprende/estudiante/modulo/{self.modulo.id}/', {
            'accion': 'quiz_modulo',
            f'q_{PreguntaModulo.objects.get(modulo=self.modulo).id}': 'A',
        })
        self.assertEqual(r2.status_code, 200)
        self.assertContains(r2, 'Aprobado')
        from aprende.models import IntentoQuizModulo
        it = IntentoQuizModulo.objects.get(estudiante=self.est, modulo=self.modulo)
        self.assertTrue(it.aprobado)
        self.assertEqual(it.correctas, 1)

    def test_media_h5p_clasifica(self):
        from aprende.media_aula import clasificar_media_url, media_desde_url
        self.assertEqual(clasificar_media_url('https://h5p.org/h5p/embed/123'), 'h5p')
        m = media_desde_url('Interactivo', 'https://example.org/h5p/embed/9')
        self.assertEqual(m.tipo, 'h5p')

    def test_estudiante_login_codigo_whatsapp(self):
        self._login_estudiante()
        r = self.http.get('/aprende/estudiante/')
        self.assertEqual(r.status_code, 200)

    def test_estudiante_login_codigo_invalido_no_entra(self):
        r = self.http.post('/aprende/estudiante/login/', {'codigo': '000000'})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'inv')
        r2 = self.http.get('/aprende/estudiante/')
        self.assertEqual(r2.status_code, 302)

    @override_settings(APRENDE_OTP_MAX_ATTEMPTS=3, APRENDE_OTP_LOCKOUT_SECONDS=600)
    def test_otp_lockout_tras_fallos(self):
        from django.core.cache import cache

        cache.clear()
        for _ in range(3):
            r = self.http.post('/aprende/estudiante/login/', {'codigo': '000000'})
            self.assertEqual(r.status_code, 200)
        r_lock = self.http.post('/aprende/estudiante/login/', {'codigo': '000000'})
        self.assertEqual(r_lock.status_code, 200)
        self.assertContains(r_lock, 'Demasiados intentos')

    def test_pwa_manifest_y_sw(self):
        r = self.http.get('/aprende/manifest.webmanifest')
        self.assertEqual(r.status_code, 200)
        self.assertIn('application/manifest+json', r['Content-Type'])
        data = r.json()
        self.assertEqual(data.get('name'), 'eki aprende')
        self.assertEqual(data.get('start_url'), '/aprende/')
        self.assertTrue(data.get('icons'))
        r_sw = self.http.get('/aprende/sw.js')
        self.assertEqual(r_sw.status_code, 200)
        self.assertIn('skipWaiting', r_sw.content.decode('utf-8'))

    def test_ayuda_y_theme_color_en_base(self):
        r = self.http.get('/aprende/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'theme-color')
        self.assertContains(r, 'eki-aula-help')
        self.assertContains(r, 'Guía del aula')
        self.assertContains(r, 'manifest.webmanifest')

    def test_auth_contrato_otp_limpia_docente(self):
        """Login estudiante WhatsApp limpia portal_usuario_id en el mismo host."""
        from portal.middleware import PORTAL_SESSION_KEY

        self.http.post('/aprende/profesor/login/', {'username': 'prof_ap', 'password': 'pass'})
        self.assertTrue(self.http.session.get(PORTAL_SESSION_KEY))
        self._login_estudiante()
        self.assertTrue(self.http.session.get('aprende_estudiante_id'))
        self.assertEqual(self.http.session.get('aprende_auth_via'), 'whatsapp')
        self.assertFalse(self.http.session.get(PORTAL_SESSION_KEY))

    def test_auth_contrato_docente_limpia_estudiante(self):
        from portal.middleware import PORTAL_SESSION_KEY

        self._login_estudiante()
        self.assertTrue(self.http.session.get('aprende_estudiante_id'))
        self.http.post('/aprende/profesor/login/', {'username': 'prof_ap', 'password': 'pass'})
        self.assertTrue(self.http.session.get(PORTAL_SESSION_KEY))
        self.assertFalse(self.http.session.get('aprende_estudiante_id'))
        self.assertFalse(self.http.session.get('aprende_auth_via'))

    def test_handoff_whatsapp_via_rechazado(self):
        """Tokens con via=whatsapp no abren sesión (handoff solo Studio)."""
        from django.core import signing
        from studio.aprende_bridge import HANDOFF_SALT

        token = signing.dumps(
            {'eid': int(self.est.pk), 'next': '/aprende/estudiante/', 'via': 'whatsapp'},
            salt=HANDOFF_SALT,
            compress=True,
        )
        r = self.http.get(f'/aprende/handoff/?t={token}')
        self.assertEqual(r.status_code, 302)
        self.assertIn('/aprende/estudiante/login/', r.url)
        self.assertFalse(self.http.session.get('aprende_estudiante_id'))

    def test_mensaje_aula_emite_acceso(self):
        from aprende.acceso_whatsapp import mensaje_pide_acceso_aula, emitir_acceso_desde_whatsapp
        self.assertTrue(mensaje_pide_acceso_aula('aula'))
        self.assertTrue(mensaje_pide_acceso_aula('Entrar al aula'))
        self.assertFalse(mensaje_pide_acceso_aula('listo'))
        txt = emitir_acceso_desde_whatsapp(self.est)
        self.assertIn('eki Aprende', txt)
        self.assertIn('listo', txt.lower())
        self.assertIn('/aprende/estudiante/login/', txt)
        self.assertNotIn('/aprende/handoff/', txt)
        self.assertRegex(txt, r'\*\d{6}\*')

    def test_profesor_ve_cursos(self):
        self.http.post('/aprende/profesor/login/', {'username': 'prof_ap', 'password': 'pass'})
        r = self.http.get('/aprende/profesor/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Curso Web')

    def test_profesor_crea_leccion_con_bloques(self):
        from core.models import SeccionModulo

        self.http.post('/aprende/profesor/login/', {'username': 'prof_ap', 'password': 'pass'})
        r_form = self.http.get(f'/aprende/profesor/curso/{self.curso.id}/modulo/nuevo/')
        self.assertEqual(r_form.status_code, 200)
        self.assertContains(r_form, 'Bloques del recorrido')
        r = self.http.post(f'/aprende/profesor/curso/{self.curso.id}/modulo/nuevo/', {
            'titulo': 'Lección con bloques',
            'descripcion': 'd',
            'contenido': '',
            'bloques_rapidos': 'Intro\nCampo\nCierre',
        })
        self.assertEqual(r.status_code, 302)
        mod = Modulo.objects.get(curso=self.curso, titulo='Lección con bloques')
        titulos = list(
            SeccionModulo.objects.filter(modulo=mod).order_by('orden').values_list('titulo', flat=True)
        )
        self.assertEqual(titulos, ['Intro', 'Campo', 'Cierre'])
        r_edit = self.http.get(f'/aprende/profesor/modulo/{mod.pk}/')
        self.assertEqual(r_edit.status_code, 200)
        self.assertContains(r_edit, 'Bloques actuales')
        self.assertContains(r_edit, 'Intro')

    def test_subdominio_aprende_redirige_raiz(self):
        r = self.http.get('/', HTTP_HOST='aprende.eki.technology')
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.url, '/aprende/')

    def test_subdominio_aula_redirige_raiz(self):
        r = self.http.get('/', HTTP_HOST='aula.eki.technology')
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.url, '/aprende/')

    def test_profesor_crea_tarea_y_estudiante_entrega(self):
        self.http.post('/aprende/profesor/login/', {'username': 'prof_ap', 'password': 'pass'})
        r = self.http.post(f'/aprende/profesor/curso/{self.curso.id}/tarea/nueva/', {
            'titulo': 'Informe final',
            'instrucciones': 'Sube PDF',
        })
        self.assertEqual(r.status_code, 302)
        tarea = TareaCurso.objects.get(titulo='Informe final')
        self._login_estudiante(telefono="3009999002")
        archivo = SimpleUploadedFile('tarea.pdf', b'%PDF-1.4 test', content_type='application/pdf')
        r2 = self.http.post(f'/aprende/estudiante/tarea/{tarea.id}/', {
            'archivo': archivo,
            'comentario': 'Listo profe',
        })
        self.assertEqual(r2.status_code, 302)
        self.assertTrue(EntregaTarea.objects.filter(tarea=tarea, estudiante=self.est).exists())

    def test_profesor_califica_entrega(self):
        tarea = TareaCurso.objects.create(curso=self.curso, titulo='T1', instrucciones='x')
        archivo = SimpleUploadedFile('t.pdf', b'%PDF', content_type='application/pdf')
        entrega = EntregaTarea.objects.create(
            tarea=tarea, estudiante=self.est, archivo=archivo, nombre_archivo='t.pdf',
        )
        self.http.post('/aprende/profesor/login/', {'username': 'prof_ap', 'password': 'pass'})
        r = self.http.post(f'/aprende/profesor/tarea/{tarea.id}/entregas/', {
            'entrega_id': entrega.id,
            'nota': '4',
            'comentario_profesor': 'Buen trabajo',
        })
        self.assertEqual(r.status_code, 302)
        entrega.refresh_from_db()
        self.assertEqual(entrega.nota, 4)
        self.assertEqual(entrega.comentario_profesor, 'Buen trabajo')

    def test_aula_solo_modulos_habilitados_por_admin(self):
        from datetime import timedelta

        from django.utils import timezone

        from core.models import HabilitacionModuloEstudiante

        self.cliente.drip_modulos_solo_estudiantes_listados = True
        self.cliente.save(update_fields=['drip_modulos_solo_estudiantes_listados'])
        m2 = Modulo.objects.create(
            curso=self.curso, numero=2, titulo='Modulo 2', descripcion='', contenido='Secreto',
        )
        HabilitacionModuloEstudiante.objects.create(
            estudiante=self.est,
            curso=self.curso,
            modulo=self.modulo,
            habilitado_desde=timezone.now() - timedelta(hours=1),
        )
        self._login_estudiante(telefono="3009999002")
        r = self.http.get(f'/aprende/estudiante/curso/{self.curso.id}/')
        self.assertContains(r, 'Intro')
        self.assertNotContains(r, 'Modulo 2')
        r2 = self.http.get(f'/aprende/estudiante/modulo/{m2.id}/')
        self.assertEqual(r2.status_code, 302)

    def test_aula_sin_lista_solo_hasta_modulo_actual(self):
        m2 = Modulo.objects.create(
            curso=self.curso, numero=2, titulo='Futuro', descripcion='', contenido='No aún',
        )
        self._login_estudiante(telefono="3009999002")
        r = self.http.get(f'/aprende/estudiante/curso/{self.curso.id}/')
        self.assertContains(r, 'Intro')
        self.assertNotContains(r, 'Futuro')
        r2 = self.http.get(f'/aprende/estudiante/modulo/{m2.id}/')
        self.assertEqual(r2.status_code, 302)

    def test_biblioteca_muestra_archivos_modulo_liberado(self):
        from core.models_extras import ArchivoModulo

        ArchivoModulo.objects.create(
            modulo=self.modulo,
            tipo='pdf',
            titulo='Guía WA',
            activo=True,
        )
        ArchivoModulo.objects.create(
            modulo=self.modulo,
            tipo='imagen',
            titulo='Foto campo',
            url_externa='https://eki-produccion.s3.us-east-2.amazonaws.com/media/test.jpg',
            activo=True,
        )
        self._login_estudiante(telefono="3009999002")
        r = self.http.get('/aprende/estudiante/biblioteca/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Biblioteca de recursos')
        self.assertContains(r, 'Guía WA')
        self.assertContains(r, 'Módulo 1')
        self.assertContains(r, self.curso.nombre)
        self.assertContains(r, 'eki-bib-grid')
        self.assertContains(r, 'eki-bib-thumb')
        self.assertContains(r, 'aria-label="Ver Foto campo"')
        self.assertContains(r, 'width="35" height="35"')
        self.assertContains(r, 'eki-bib-lightbox')
        self.assertContains(r, 'title="Guía WA"')

    def test_biblioteca_incluye_media_microcontenido(self):
        from core.models import PasoModulo, SeccionModulo

        sec = SeccionModulo.objects.create(
            modulo=self.modulo, orden=1, titulo='Fundamentos',
        )
        PasoModulo.objects.create(
            modulo=self.modulo,
            seccion=sec,
            orden=1,
            titulo='Video intro micro',
            tipo=PasoModulo.TIPO_CONTENIDO,
            contenido='Texto del paso.',
            media_url='https://www.youtube.com/watch?v=dQw4w9WgXcQ',
        )
        PasoModulo.objects.create(
            modulo=self.modulo,
            seccion=sec,
            orden=2,
            tipo=PasoModulo.TIPO_CONTENIDO,
            contenido='Solo texto, sin media.',
        )
        self._login_estudiante(telefono="3009999002")
        r = self.http.get('/aprende/estudiante/biblioteca/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Video intro micro')
        self.assertContains(r, 'aria-label="Ver Video intro micro"')
        self.assertContains(r, 'img.youtube.com/vi/dQw4w9WgXcQ/default.jpg')

    def test_modulo_muestra_secciones_y_media_solo_consulta(self):
        from core.models import PasoModulo, SeccionModulo

        sec = SeccionModulo.objects.create(
            modulo=self.modulo, orden=1, titulo='Fundamentos',
        )
        PasoModulo.objects.create(
            modulo=self.modulo,
            seccion=sec,
            orden=1,
            tipo=PasoModulo.TIPO_CONTENIDO,
            contenido='Contenido del micro paso.',
            media_url='https://www.youtube.com/watch?v=dQw4w9WgXcQ',
        )
        self._login_estudiante(telefono="3009999002")
        r = self.http.get(f'/aprende/estudiante/modulo/{self.modulo.id}/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Fundamentos')
        self.assertContains(r, 'Contenido del micro paso')
        self.assertContains(r, 'youtube-nocookie.com/embed/')
        self.assertContains(r, 'consulta en línea')
        self.assertNotContains(r, 'Descargar documento')

    def test_aula_oculta_modulo_con_drip_tras_completar_anterior(self):
        from datetime import timedelta

        from django.utils import timezone

        from core.models import ModuloCompletado

        m2 = Modulo.objects.create(
            curso=self.curso, numero=2, titulo='Bloqueado drip', descripcion='', contenido='Espera',
        )
        m2.habilitado_desde = timezone.now() + timedelta(days=5)
        m2.save(update_fields=['habilitado_desde'])
        progreso = ProgresoEstudiante.objects.get(estudiante=self.est, curso=self.curso)
        ModuloCompletado.objects.create(progreso=progreso, modulo=self.modulo)
        progreso.modulo_actual = m2
        progreso.fecha_ultimo_avance = timezone.now()
        progreso.save(update_fields=['modulo_actual', 'fecha_ultimo_avance'])

        self._login_estudiante(telefono="3009999002")
        r = self.http.get(f'/aprende/estudiante/curso/{self.curso.id}/')
        self.assertContains(r, 'Intro')
        self.assertNotContains(r, 'Bloqueado drip')
        r2 = self.http.get(f'/aprende/estudiante/modulo/{m2.id}/')
        self.assertEqual(r2.status_code, 302)

    def test_estudiante_perfil_y_puntos(self):
        from core.gamificacion import PerfilGamificacion

        perfil, _ = PerfilGamificacion.objects.get_or_create(estudiante=self.est)
        perfil.puntos_totales = 120
        perfil.nivel = 3
        perfil.save(update_fields=['puntos_totales', 'nivel'])
        self._login_estudiante(telefono="3009999002")
        r = self.http.get('/aprende/estudiante/perfil/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, '120')
        self.assertContains(r, 'Mi perfil')

        r2 = self.http.post('/aprende/estudiante/perfil/', {
            'accion': 'perfil',
            'nombre': 'Est Web Actualizado',
            'municipio': 'Medellín',
            'departamento': 'Antioquia',
            'genero': 'M',
            'edad': '28',
        })
        self.assertEqual(r2.status_code, 302)
        self.est.refresh_from_db()
        self.assertEqual(self.est.nombre, 'Est Web Actualizado')
        self.assertEqual(self.est.municipio, 'Medellín')

    def test_estudiante_pagina_tareas_sin_subida_en_modulo(self):
        TareaCurso.objects.create(curso=self.curso, titulo='Entrega 1', instrucciones='Sube PDF')
        self._login_estudiante(telefono="3009999002")
        r = self.http.get('/aprende/estudiante/tareas/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Entrega 1')
        self.assertContains(r, 'Pendiente')
        r2 = self.http.get(f'/aprende/estudiante/modulo/{self.modulo.id}/')
        self.assertEqual(r2.status_code, 200)
        self.assertNotContains(r2, 'Subir documento')
        self.assertContains(r2, '*listo*')

    def test_ranking_grupo_en_perfil(self):
        from core.gamificacion import PerfilGamificacion
        from core.models_extras import GrupoEstudiantes

        est2 = Estudiante.objects.create(
            cedula='web_rank',
            nombre='Compañero',
            telefono='573009999099',
            cliente=self.cliente,
            activo=True,
        )
        grupo = GrupoEstudiantes.objects.create(nombre='Grupo Norte', cliente=self.cliente, activo=True)
        grupo.estudiantes.add(self.est, est2)
        grupo.cursos.add(self.curso)
        p1, _ = PerfilGamificacion.objects.get_or_create(estudiante=self.est)
        p1.puntos_totales = 50
        p1.save(update_fields=['puntos_totales'])
        p2, _ = PerfilGamificacion.objects.get_or_create(estudiante=est2)
        p2.puntos_totales = 200
        p2.save(update_fields=['puntos_totales'])

        self._login_estudiante(telefono="3009999002")
        r = self.http.get('/aprende/estudiante/perfil/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Grupo Norte')
        self.assertContains(r, 'Compañero')
        self.assertContains(r, 'Tú')

    def test_curso_pestanas_modulos_y_tareas(self):
        TareaCurso.objects.create(curso=self.curso, titulo='Tarea curso', instrucciones='x')
        self._login_estudiante(telefono="3009999002")
        r = self.http.get(f'/aprende/estudiante/curso/{self.curso.id}/')
        self.assertContains(r, 'Módulos')
        self.assertContains(r, 'Tareas')
        self.assertContains(r, 'Ranking')
        r2 = self.http.get(f'/aprende/estudiante/curso/{self.curso.id}/tareas/')
        self.assertEqual(r2.status_code, 200)
        self.assertContains(r2, 'Tarea curso')

    def test_ranking_curso_por_grupo(self):
        from core.gamificacion import PerfilGamificacion
        from core.models_extras import GrupoEstudiantes

        est2 = Estudiante.objects.create(
            cedula='web_rank2',
            nombre='Líder grupo',
            telefono='573009999088',
            cliente=self.cliente,
            activo=True,
        )
        grupo = GrupoEstudiantes.objects.create(nombre='Grupo Curso', cliente=self.cliente, activo=True)
        grupo.estudiantes.add(self.est, est2)
        grupo.cursos.add(self.curso)

        p1, _ = PerfilGamificacion.objects.get_or_create(estudiante=self.est)
        p1.puntos_totales = 30
        p1.save(update_fields=['puntos_totales'])
        p2, _ = PerfilGamificacion.objects.get_or_create(estudiante=est2)
        p2.puntos_totales = 120
        p2.save(update_fields=['puntos_totales'])

        self._login_estudiante(telefono="3009999002")
        r = self.http.get(f'/aprende/estudiante/curso/{self.curso.id}/ranking/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Ranking del curso')
        self.assertContains(r, 'Grupo Curso')
        self.assertContains(r, 'Líder Grupo')
        self.assertContains(r, 'eki-leaderboard-score__num">120')
        self.assertContains(r, 'Tu puesto')
        self.assertContains(r, 'Te faltan')


@override_settings(SECURE_SSL_REDIRECT=False)
class AprendeProfesorAuthTests(TestCase):
    """Portal B2B y aula docente comparten sesión; is_staff no debe bloquear.
    Superuser sí se bloquea (no mezclar con admin).
    """

    def setUp(self):
        self.http = Client()
        self.cliente = Cliente.objects.create(
            nombre='Org Docente',
            contacto_principal='A',
            email='doc@test.com',
            telefono='573008888001',
            activo=True,
            fecha_fin_suscripcion='2099-12-31',
        )
        self.curso = Curso.objects.create(nombre='Curso Doc', cliente=self.cliente, activo=True)
        self.user_staff = User.objects.create_user('coord_admin', 'c@t.com', 'pass1234')
        self.user_staff.is_staff = True
        self.user_staff.save(update_fields=['is_staff'])
        PortalUsuario.objects.create(
            user=self.user_staff, organizacion=self.cliente, rol='admin',
        )
        self.viewer = User.objects.create_user('solo_ver', 'v@t.com', 'pass1234')
        PortalUsuario.objects.create(
            user=self.viewer, organizacion=self.cliente, rol='viewer',
        )

    def test_staff_admin_entra_aprende_profesor(self):
        r = self.http.post('/aprende/profesor/login/', {
            'username': 'coord_admin',
            'password': 'pass1234',
        })
        self.assertIn(r.status_code, (301, 302))
        if r.status_code == 301:
            r = self.http.get(r['Location'].replace('https://testserver', '').replace('http://testserver', ''))
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.url, '/aprende/profesor/')
        r2 = self.http.get('/aprende/profesor/')
        self.assertEqual(r2.status_code, 200)
        self.assertContains(r2, 'Curso Doc')

    def test_sesion_portal_abre_aula_sin_relogin(self):
        self.http.post('/portal/login/', {'username': 'coord_admin', 'password': 'pass1234'})
        r = self.http.get('/aprende/profesor/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Curso Doc')

    def test_viewer_no_entra_aula_docente(self):
        r = self.http.post('/aprende/profesor/login/', {
            'username': 'solo_ver',
            'password': 'pass1234',
        })
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Administrador o Profesor')

    def test_superuser_no_entra_profesor_login(self):
        su = User.objects.create_superuser('su_luisa', 'su@t.com', 'pass1234')
        PortalUsuario.objects.create(user=su, organizacion=self.cliente, rol='admin')
        r = self.http.post('/aprende/profesor/login/', {
            'username': 'su_luisa',
            'password': 'pass1234',
        })
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'superadmin')


@override_settings(
    SECURE_SSL_REDIRECT=False,
    STORAGES={
        'default': {'BACKEND': 'django.core.files.storage.FileSystemStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    },
)
class AprendeProfesorGestionTests(TestCase):
    """Tareas editar/borrar, asistencia y calificaciones en aula docente."""

    def setUp(self):
        self.http = Client()
        self.cliente = Cliente.objects.create(
            nombre='Org Gestión',
            contacto_principal='A',
            email='gest@test.com',
            telefono='573007777001',
            activo=True,
            modo_gamificacion='calificacion',
        )
        self.curso = Curso.objects.create(nombre='Curso Gest', cliente=self.cliente, activo=True)
        self.modulo = Modulo.objects.create(
            curso=self.curso, numero=1, titulo='M1', descripcion='', contenido='x',
        )
        self.est = Estudiante.objects.create(
            cedula='gest1', nombre='Est Gest', telefono='573007777002',
            cliente=self.cliente, activo=True,
        )
        ProgresoEstudiante.objects.create(estudiante=self.est, curso=self.curso, modulo_actual=self.modulo)
        self.user = User.objects.create_user('prof_gest', 'pg@t.com', 'pass')
        PortalUsuario.objects.create(user=self.user, organizacion=self.cliente, rol='profesor')
        self.http.post('/aprende/profesor/login/', {'username': 'prof_gest', 'password': 'pass'})

    def test_editar_y_desactivar_tarea(self):
        tarea = TareaCurso.objects.create(curso=self.curso, titulo='Original', instrucciones='a')
        archivo = SimpleUploadedFile('x.pdf', b'%PDF', content_type='application/pdf')
        EntregaTarea.objects.create(
            tarea=tarea, estudiante=self.est,
            archivo=archivo, nombre_archivo='x.pdf',
        )
        r = self.http.post(f'/aprende/profesor/tarea/{tarea.id}/editar/', {
            'titulo': 'Actualizada',
            'instrucciones': 'b',
            'activa': '',
        })
        self.assertEqual(r.status_code, 302)
        tarea.refresh_from_db()
        self.assertEqual(tarea.titulo, 'Actualizada')
        self.assertFalse(tarea.activa)

        r2 = self.http.post(f'/aprende/profesor/tarea/{tarea.id}/eliminar/')
        self.assertEqual(r2.status_code, 302)
        self.assertTrue(TareaCurso.objects.filter(pk=tarea.pk).exists())
        tarea.refresh_from_db()
        self.assertFalse(tarea.activa)

    def test_eliminar_tarea_sin_entregas(self):
        tarea = TareaCurso.objects.create(curso=self.curso, titulo='Borrar', instrucciones='x')
        r = self.http.post(f'/aprende/profesor/tarea/{tarea.id}/eliminar/')
        self.assertEqual(r.status_code, 302)
        self.assertFalse(TareaCurso.objects.filter(pk=tarea.pk).exists())

    def test_asistencia_y_calificacion(self):
        from core.gamificacion import EvaluacionNotaGamificacion

        r = self.http.post(f'/aprende/profesor/curso/{self.curso.id}/asistencia/', {
            'fecha': '2026-07-06',
            'presente': [str(self.est.pk)],
        })
        self.assertEqual(r.status_code, 302)
        self.assertTrue(
            EvaluacionNotaGamificacion.objects.filter(
                estudiante=self.est, curso=self.curso, tipo='asistencia',
            ).exists()
        )

        r2 = self.http.get(f'/aprende/profesor/curso/{self.curso.id}/calificaciones/')
        self.assertEqual(r2.status_code, 200)
        self.assertContains(r2, 'Est Gest')

        r3 = self.http.post(f'/aprende/profesor/curso/{self.curso.id}/calificaciones/', {
            'accion': 'nota_manual',
            'estudiante_id': str(self.est.pk),
            'nota': '4',
            'detalle': 'Participación',
        })
        self.assertEqual(r3.status_code, 302)
        self.assertTrue(
            EvaluacionNotaGamificacion.objects.filter(estudiante=self.est, tipo='manual').exists()
        )

    def test_profesor_ve_ranking_del_curso(self):
        r = self.http.get(f'/aprende/profesor/curso/{self.curso.id}/ranking/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Ranking de estudiantes')
        self.assertContains(r, f'/aprende/profesor/curso/{self.curso.id}/ranking/')
        # CSS debe ir en <style> (extra_css), no como texto suelto antes del page-head
        html = r.content.decode('utf-8')
        head, _, body = html.partition('<body')
        self.assertIn('.eki-lb', head)
        self.assertNotIn('.eki-lb {', body)
        self.assertContains(r, 'eki-lb-board')

    def test_borrar_asistencia_del_dia(self):
        from aprende.models import AsistenciaAula
        from core.gamificacion import EvaluacionNotaGamificacion

        self.http.post(f'/aprende/profesor/curso/{self.curso.id}/asistencia/', {
            'fecha': '2026-07-06',
            'accion': 'guardar',
            'presente': [str(self.est.pk)],
        })
        self.assertTrue(
            AsistenciaAula.objects.filter(curso=self.curso, fecha='2026-07-06').exists()
        )

        r = self.http.post(f'/aprende/profesor/curso/{self.curso.id}/asistencia/', {
            'fecha': '2026-07-06',
            'accion': 'borrar',
        })
        self.assertEqual(r.status_code, 302)
        self.assertFalse(
            AsistenciaAula.objects.filter(curso=self.curso, fecha='2026-07-06').exists()
        )
        self.assertFalse(
            EvaluacionNotaGamificacion.objects.filter(
                estudiante=self.est, curso=self.curso, tipo='asistencia',
                detalle='Asistencia 2026-07-06',
            ).exists()
        )

    def test_descargar_asistencia_excel_dias_marcados(self):
        from io import BytesIO

        from openpyxl import load_workbook

        self.http.post(f'/aprende/profesor/curso/{self.curso.id}/asistencia/', {
            'fecha': '2026-07-06',
            'accion': 'guardar',
            'presente': [str(self.est.pk)],
        })
        self.http.post(f'/aprende/profesor/curso/{self.curso.id}/asistencia/', {
            'fecha': '2026-07-07',
            'accion': 'guardar',
            'presente': [],
        })

        pagina = self.http.get(f'/aprende/profesor/curso/{self.curso.id}/asistencia/')
        self.assertEqual(pagina.status_code, 200)
        self.assertContains(pagina, 'Descargar Excel')
        self.assertContains(pagina, 'asistencia/excel/')

        r = self.http.get(f'/aprende/profesor/curso/{self.curso.id}/asistencia/excel/')
        self.assertEqual(r.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            r['Content-Type'],
        )
        self.assertIn('attachment', r['Content-Disposition'])
        self.assertIn('.xlsx', r['Content-Disposition'])

        wb = load_workbook(BytesIO(r.content))
        self.assertIn('Asistencia', wb.sheetnames)
        self.assertIn('Detalle por día', wb.sheetnames)
        ws = wb['Asistencia']
        headers = [c.value for c in ws[1]]
        self.assertIn('Estudiante', headers)
        self.assertIn('06/07/2026', headers)
        self.assertIn('07/07/2026', headers)
        # Fila del estudiante: presente el 06, ausente el 07
        fila = [c.value for c in ws[2]]
        self.assertEqual(fila[0], 'Est Gest')
        self.assertEqual(fila[2], 'Presente')
        self.assertEqual(fila[3], 'Ausente')

        r_dia = self.http.get(
            f'/aprende/profesor/curso/{self.curso.id}/asistencia/excel/?fecha=2026-07-06'
        )
        self.assertEqual(r_dia.status_code, 200)
        wb_dia = load_workbook(BytesIO(r_dia.content))
        headers_dia = [c.value for c in wb_dia['Asistencia'][1]]
        self.assertIn('06/07/2026', headers_dia)
        self.assertNotIn('07/07/2026', headers_dia)

        r_vacio = self.http.get(
            f'/aprende/profesor/curso/{self.curso.id}/asistencia/excel/?fecha=2026-01-01'
        )
        self.assertEqual(r_vacio.status_code, 302)
