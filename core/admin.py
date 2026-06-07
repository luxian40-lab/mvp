"""
Admin completo: Estudiantes, Plantillas, Campañas, EnvioLog, Sistema Educativo
CON función de envío directo desde Plantillas Y gestión de cursos/módulos/exámenes
"""
from django.contrib import admin
from django import forms
from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.admin import helpers
from django.urls import path, reverse
from django.db import models, transaction  # Q(), atomic/on_commit
from django.db.models import Count  # ✅ Para agregaciones
from django.conf import settings  # ✅ Para acceder a settings
import openpyxl
from django.http import HttpResponse
from .models import (
    Estudiante, WhatsappLog, Plantilla, Campana, EnvioLog, Linea,
    Curso, ConfiguracionDripCliente, HabilitacionModuloDripCliente, HabilitacionModuloEstudiante,
    Modulo, SeccionModulo, PasoModulo, ProgresoEstudiante, ModuloCompletado,
    Examen, PreguntaExamen, ResultadoExamen, Cliente,
    PerfilGamificacion, Badge, BadgeEstudiante, TransaccionPuntos,
    SolicitudSoporte, PreguntaModulo,  # 🆘 NUEVO + 📝 PREGUNTA MODULO
    Certificado, PlantillaCertificado, # 📜 CERTIFICADOS
    CampanaUnica, RespuestaCampanaUnica,
    ProspectoB2B, CampanaB2B,  # 🤝 LEADS B2B
    AliadoEmpleabilidad, MisionEmpleabilidad, PreguntaAbiertaFinalCurso, RespuestaAbiertaFinal,
    DocumentoRAG,  # 📚 RAG Multi-Tenant
    DocumentoRAGComercial,  # 🛒 RAG Comercial
    ProductoComercial,  # 💰 Precios Nat (Postgres)
    ProductoCatalogo,  # 📦 Catálogo recomendación Nat
    MetaMetricaEmpresa, MetaMetricaNati,
    EventoIA,
    ContextoAgroSession,
    ConversacionRAGCandidata,
)
from .admin_campana_actualizado import CampanaUnicaAdmin, RespuestaCampanaUnicaAdmin
from .models_extras import (
    GrupoEstudiantes, EnvioProgramado, PQRS, ArchivoModulo,
    MensajePush, EnvioMensajePush,
    GrupoWhatsApp, InvitacionGrupo  # 📦 NUEVOS MODELOS
)
from .models_audit import AuditLog  # 🔐 Auditoría
from .recompensas import Recompensa, CanjeRecompensa
from .utils import enviar_whatsapp_twilio, enviar_whatsapp  # Asegurar que tenemos la función
from .widgets import ColorPickerWidget  # 🎨 Color picker para certificados
from portal.models import PortalUsuario
from django.utils import timezone  # Para timestamps
import logging
from datetime import datetime
import json
import os

logger = logging.getLogger(__name__)


def guardar_upload_admin_media(uploaded_file, *, carpeta='admin_media', prefix='media'):
    """Guarda un upload del admin en el storage activo y devuelve su URL pública."""
    from django.core.files.storage import default_storage
    from django.utils.text import get_valid_filename

    now = timezone.now()
    filename = get_valid_filename(uploaded_file.name)
    path = f'{carpeta}/{now:%Y/%m}/{prefix}_{now:%Y%m%d%H%M%S}_{filename}'
    saved_path = default_storage.save(path, uploaded_file)
    return default_storage.url(saved_path)


# ================================================
# 🔍 FILTRO PERSONALIZADO PARA CURSOS
# ================================================

class CursosEstudianteFilter(admin.SimpleListFilter):
    """Filtro para ver estudiantes por curso inscrito"""
    title = 'Curso inscrito'
    parameter_name = 'curso'

    def lookups(self, request, model_admin):
        """Obtiene todos los cursos disponibles"""
        cursos = Curso.objects.all().order_by('nombre')
        return [(curso.id, curso.nombre) for curso in cursos] + [('sin_curso', '❌ Sin curso')]

    def queryset(self, request, queryset):
        """Filtra los estudiantes según el curso seleccionado"""
        if self.value() == 'sin_curso':
            # Estudiantes sin ningún curso
            return queryset.exclude(progresos__isnull=False)
        elif self.value():
            # Estudiantes inscritos en el curso específico
            return queryset.filter(progresos__curso_id=self.value()).distinct()
        return queryset


class GruposEstudianteFilter(admin.SimpleListFilter):
    """Filtro para ver estudiantes por grupo"""
    title = '👥 Grupo asignado'
    parameter_name = 'grupo'

    def lookups(self, request, model_admin):
        """Obtiene todos los grupos disponibles"""
        grupos = GrupoEstudiantes.objects.all().order_by('nombre')
        return [(grupo.id, f"{grupo.emoji} {grupo.nombre}") for grupo in grupos] + [('sin_grupo', '❌ Sin grupo')]

    def queryset(self, request, queryset):
        """Filtra los estudiantes según el grupo seleccionado"""
        if self.value() == 'sin_grupo':
            # Estudiantes sin ningún grupo
            return queryset.filter(grupos__isnull=True)
        elif self.value():
            # Estudiantes en el grupo específico
            return queryset.filter(grupos__id=self.value()).distinct()
        return queryset


# ========== CLIENTE (NUEVO) ==========
class ConfiguracionDripClienteInline(admin.TabularInline):
    """Override de días entre módulos por curso (misma fila = un curso por cliente)."""
    model = ConfiguracionDripCliente
    extra = 0
    fields = ('curso', 'dias_espera_entre_modulos', 'activo')
    autocomplete_fields = ('curso',)
    verbose_name = 'Drip curso'
    verbose_name_plural = '⏱️ Ritmo drip por curso (override)'


class HabilitacionModuloEstudianteInline(admin.TabularInline):
    model = HabilitacionModuloEstudiante
    extra = 1
    fields = ('curso', 'modulo', 'habilitado_desde', 'activo', 'notas')
    autocomplete_fields = ('curso', 'modulo')
    verbose_name = 'Módulo individual'
    verbose_name_plural = (
        'Módulos habilitados solo para este estudiante '
        '(requiere «solo por lista» activo en el cliente)'
    )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'curso' and request.resolver_match:
            obj_id = request.resolver_match.kwargs.get('object_id')
            if obj_id:
                try:
                    est = Estudiante.objects.only('cliente_id').get(pk=obj_id)
                    kwargs['queryset'] = Curso.objects.filter(cliente_id=est.cliente_id, activo=True)
                except Estudiante.DoesNotExist:
                    pass
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(HabilitacionModuloEstudiante)
class HabilitacionModuloEstudianteAdmin(admin.ModelAdmin):
    """Listado masivo de qué estudiante puede ver qué módulo."""
    list_display = ('estudiante', 'curso', 'modulo', 'habilitado_desde', 'activo', 'notas')
    list_filter = ('activo', 'curso__cliente', 'curso')
    search_fields = (
        'estudiante__nombre', 'estudiante__cedula', 'estudiante__telefono',
        'curso__nombre', 'modulo__nombre',
    )
    autocomplete_fields = ('estudiante', 'curso', 'modulo')
    list_select_related = ('estudiante', 'curso', 'modulo', 'curso__cliente')
    ordering = ('-id',)


class HabilitacionModuloDripClienteInline(admin.TabularInline):
    """Fecha/hora en que un módulo del curso se habilita para este cliente (sustituye la fecha global del módulo)."""
    model = HabilitacionModuloDripCliente
    extra = 0
    fields = ('curso', 'modulo', 'habilitado_desde', 'activo')
    autocomplete_fields = ('curso', 'modulo')
    verbose_name = 'Calendario módulo'
    verbose_name_plural = '📅 Habilitación de módulos por calendario'


class ProductoCatalogoInline(admin.TabularInline):
    model = ProductoCatalogo
    extra = 0
    fields = (
        'nombre', 'categoria', 'cultivos_objetivo',
        'precio_cop', 'unidad', 'url_producto', 'activo',
    )
    show_change_link = True


class PortalUsuarioInline(admin.TabularInline):
    model = PortalUsuario
    extra = 0
    fields = ('user', 'rol', 'portal_user_link')
    readonly_fields = ('portal_user_link',)
    autocomplete_fields = ('user',)
    verbose_name = 'Usuario del portal'
    verbose_name_plural = 'Usuarios del portal'

    def portal_user_link(self, obj):
        if not obj or not obj.user_id:
            return '-'
        url = reverse('admin:auth_user_change', args=[obj.user_id])
        return format_html('<a href="{}">Editar usuario / contraseña</a>', url)
    portal_user_link.short_description = 'Acceso'


class CrearUsuarioPortalForm(forms.Form):
    username = forms.CharField(label='Usuario', max_length=150)
    first_name = forms.CharField(label='Nombre', max_length=150, required=False)
    last_name = forms.CharField(label='Apellido', max_length=150, required=False)
    email = forms.EmailField(label='Email', required=False)
    password1 = forms.CharField(label='Contraseña', widget=forms.PasswordInput)
    password2 = forms.CharField(label='Confirmar contraseña', widget=forms.PasswordInput)
    rol = forms.ChoiceField(label='Rol', choices=PortalUsuario.ROL_CHOICES, initial='viewer')
    is_active = forms.BooleanField(label='Usuario activo', required=False, initial=True)

    def clean_username(self):
        username = self.cleaned_data['username'].strip()
        User = get_user_model()
        if User.objects.filter(username__iexact=username).exists():
            raise forms.ValidationError('Ya existe un usuario con ese username.')
        return username

    def clean(self):
        cleaned = super().clean()
        password1 = cleaned.get('password1')
        password2 = cleaned.get('password2')
        if password1 and password2 and password1 != password2:
            raise forms.ValidationError('Las contraseñas no coinciden.')
        return cleaned

    def save(self, cliente):
        User = get_user_model()
        user = User(
            username=self.cleaned_data['username'],
            first_name=self.cleaned_data.get('first_name', ''),
            last_name=self.cleaned_data.get('last_name', ''),
            email=self.cleaned_data.get('email', ''),
            is_staff=False,
            is_superuser=False,
            is_active=self.cleaned_data.get('is_active', True),
        )
        user.set_password(self.cleaned_data['password1'])
        user.save()
        PortalUsuario.objects.create(
            user=user,
            organizacion=cliente,
            rol=self.cleaned_data['rol'],
        )
        return user


@admin.register(Cliente)
class ClienteAdmin(admin.ModelAdmin):
    """Gestión de clientes/organizaciones"""
    change_form_template = 'admin/core/cliente/change_form.html'

    def get_form(self, request, obj=None, **kwargs):
        from portal.forms import ClientePortalAdminForm
        kwargs['form'] = ClientePortalAdminForm
        return super().get_form(request, obj, **kwargs)

    inlines = [
        PortalUsuarioInline,
        ConfiguracionDripClienteInline,
        HabilitacionModuloDripClienteInline,
        ProductoCatalogoInline,
    ]
    list_display = ('nombre', 'contacto_principal', 'email', 'numero_meta_badge', 'estudiantes_activos', 'cursos_asignados', 'activo', 'fecha_registro')
    list_filter = ('activo', 'enviar_certificados_email', 'fecha_registro')
    search_fields = ('nombre', 'nit', 'contacto_principal', 'email')
    list_per_page = 50
    ordering = ('-fecha_registro',)
    readonly_fields = ('portal_usuarios_acciones', 'cobertura_y_drip_acciones', 'usar_gamificacion')
    
    fieldsets = (
        ('Datos del cliente', {
            'fields': (
                'nombre',
                'nit',
                'contacto_principal',
                'email',
                'telefono',
                'activo',
                'notas_internas',
            ),
        }),
        ('Portal B2B', {
            'fields': (
                'tipo_proyecto',
                'portal_modulos',
                'fecha_inicio_suscripcion',
                'fecha_fin_suscripcion',
                'logo_url',
                'portal_subtitulo',
                'portal_usuarios_acciones',
                'whatsapp_numero',
                'twilio_account_sid',
                'twilio_auth_token',
                'twilio_whatsapp_from',
            ),
            'description': (
                'Acceso web, branding y credenciales Twilio del cliente. '
                'Logo y subtítulo también en <code>/portal/perfil/</code> (rol admin).'
            ),
        }),
        ('WhatsApp, legal y grupo', {
            'fields': (
                'numero_whatsapp_autorizado',
                'enlace_habeas_data',
                'content_sid_habeas_data_twilio',
                'modo_avance_modulo',
                'content_sid_boton_listo',
                'enlace_grupo_whatsapp',
            ),
            'description': (
                'Número autorizado en Meta Business; política de datos (Habeas) y plantilla Twilio propia; '
                'avance por texto o botón «Listo» al cerrar cada módulo; enlace de invitación al grupo.'
            ),
        }),
        ('Certificados, drip y gamificación', {
            'fields': (
                'enviar_certificados_email',
                'exigir_nota_minima_certificado',
                'nota_minima_certificado',
                'drip_modulos_solo_estudiantes_listados',
                'cobertura_y_drip_acciones',
                'modo_gamificacion',
                'usar_gamificacion',
            ),
            'description': (
                'Certificados por email y nota mínima; drip por estudiante (ver tablas al final); '
                'modo puntos o calificación 1–5.'
            ),
        }),
        ('Ventanas por fechas', {
            'fields': (
                'habilitar_pregunta_abierta_final',
                'fecha_inicio_pregunta_abierta_final',
                'fecha_fin_pregunta_abierta_final',
                'habilitar_gamificacion_proximidad',
                'fecha_inicio_gamificacion_proximidad',
                'fecha_fin_gamificacion_proximidad',
            ),
            'description': (
                'Cuándo se activa la pregunta abierta final y el radar de empleabilidad por proximidad.'
            ),
        }),
        ('Empleabilidad, IA y bot comercial', {
            'fields': (
                'empleabilidad_exploracion_activa',
                'empleabilidad_radio_metros',
                'empleabilidad_cooldown_horas',
                'empleabilidad_max_misiones_dia',
                'empleabilidad_puntos_validacion',
                'nombre_agente_tutor',
                'nombre_agente_asistente',
                'nombre_bot',
                'system_prompt_extra',
            ),
            'description': (
                'Exploración de empleabilidad, nombres de agentes educativos y bot Nat/comercial. '
                'Catálogo e inlines al final del formulario.'
            ),
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:cliente_id>/crear-usuario-portal/',
                self.admin_site.admin_view(self.crear_usuario_portal_view),
                name='core_cliente_crear_usuario_portal',
            ),
        ]
        return custom_urls + urls

    def portal_usuarios_acciones(self, obj):
        if not obj or not obj.pk:
            return 'Guarda el cliente para crear usuarios del portal.'

        crear_url = reverse('admin:core_cliente_crear_usuario_portal', args=[obj.pk])
        usuarios_url = (
            reverse('admin:portal_portalusuario_changelist')
            + f'?organizacion__id__exact={obj.pk}'
        )
        portal_url = '/portal/login/'
        total = obj.usuarios_portal.count()

        return format_html(
            '<div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;">'
            '<a class="button" href="{}">➕ Crear usuario portal</a>'
            '<a class="button" href="{}">Ver usuarios ({})</a>'
            '<a class="button" href="{}" target="_blank" rel="noopener">Abrir portal</a>'
            '</div>'
            '<p style="margin:8px 0 0;color:#666;">'
            'Crea aquí usuarios no-staff para esta organización. '
            'La contraseña queda en Django Auth; el rol y la organización quedan en PortalUsuario.'
            '</p>',
            crear_url,
            usuarios_url,
            total,
            portal_url,
        )
    portal_usuarios_acciones.short_description = 'Accesos portal'

    def cobertura_y_drip_acciones(self, obj):
        if not obj or not obj.pk:
            return 'Guarde el cliente para ver enlaces.'
        mapa_url = reverse('admin_cobertura_mapa') + f'?cliente={obj.pk}'
        drip_url = reverse('admin_drip_estudiantes') + f'?cliente={obj.pk}'
        return format_html(
            '<div style="display:flex;gap:10px;flex-wrap:wrap;">'
            '<a class="button" href="{}">Mapa cobertura</a>'
            '<a class="button" href="{}">Módulos por estudiante (tabla)</a>'
            '</div>',
            mapa_url,
            drip_url,
        )

    cobertura_y_drip_acciones.short_description = 'Mapa y listas'

    def crear_usuario_portal_view(self, request, cliente_id):
        cliente = self.get_object(request, str(cliente_id))
        if not cliente:
            self.message_user(request, 'Cliente no encontrado.', level=messages.ERROR)
            return redirect('admin:core_cliente_changelist')
        if not self.has_change_permission(request, cliente):
            raise PermissionDenied

        initial = {
            'email': cliente.email,
            'first_name': cliente.contacto_principal,
            'is_active': True,
        }
        if request.method == 'POST':
            form = CrearUsuarioPortalForm(request.POST)
            if form.is_valid():
                with transaction.atomic():
                    user = form.save(cliente)
                self.message_user(
                    request,
                    f'Usuario portal "{user.username}" creado para {cliente.nombre}.',
                    level=messages.SUCCESS,
                )
                return redirect('admin:core_cliente_change', cliente.pk)
        else:
            form = CrearUsuarioPortalForm(initial=initial)

        context = {
            **self.admin_site.each_context(request),
            'title': f'Crear usuario portal para {cliente.nombre}',
            'opts': self.model._meta,
            'original': cliente,
            'cliente': cliente,
            'form': form,
            'change_url': reverse('admin:core_cliente_change', args=[cliente.pk]),
        }
        return render(request, 'admin/core/cliente/crear_usuario_portal.html', context)
    
    def estudiantes_activos(self, obj):
        count = obj.total_estudiantes()
        if count > 0:
            return format_html('<span style="background:#4caf50;color:white;padding:4px 12px;border-radius:12px;font-weight:bold;">{}</span>', count)
        return format_html('<span style="color:#999;">0</span>')
    estudiantes_activos.short_description = "👥 Estudiantes"
    
    def numero_meta_badge(self, obj):
        if obj.numero_whatsapp_autorizado:
            return format_html(
                '<span style="background:#25d366;color:white;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">✅ Meta OK</span><br>'
                '<span style="font-size:10px;color:#666;">{}</span>',
                obj.numero_whatsapp_autorizado
            )
        return format_html('<span style="background:#ff9800;color:white;padding:4px 10px;border-radius:12px;font-size:11px;">⚠️ Sin Meta</span>')
    numero_meta_badge.short_description = "📱 WhatsApp Meta"

    def cursos_asignados(self, obj):
        count = obj.total_cursos()
        if count > 0:
            return format_html('<span style="background:#2196f3;color:white;padding:4px 12px;border-radius:12px;font-weight:bold;">{}</span>', count)
        return format_html('<span style="color:#999;">0</span>')
    cursos_asignados.short_description = "📚 Cursos"

    actions = ['copiar_todos_cursos_a_analytics_pruebas']

    @admin.action(description='📋 Copiar todos los cursos → Analytics (Pruebas)')
    def copiar_todos_cursos_a_analytics_pruebas(self, request, queryset):
        from core.copiar_cursos import (
            ClienteAnalyticsNoEncontrado,
            copiar_cursos_a_pruebas,
            obtener_cliente_analytics_origen,
        )

        if queryset.count() != 1:
            self.message_user(request, 'Selecciona un solo cliente.', level='error')
            return
        cliente = queryset.first()
        try:
            origen = obtener_cliente_analytics_origen()
        except ClienteOrigenNoEncontrado as e:
            self.message_user(request, str(e), level='error')
            return
        if cliente.pk != origen.pk:
            self.message_user(
                request,
                f'Selecciona el cliente «{origen.nombre}» (origen Alitic).',
                level='error',
            )
            return
        result = copiar_cursos_a_pruebas()
        self.message_user(
            request,
            f'✅ {result.total_copiados} curso(s) copiados a {result.destino.nombre}. '
            f'Omitidos (ya existían): {len(result.omitidos)}.',
        )


# ========== PLANTILLA ==========
@admin.register(Plantilla)
class PlantillaDashboardAdmin(admin.ModelAdmin):
    """Gestión de plantillas de mensajes"""
    list_display = ('nombre_con_emoji', 'categoria_display', 'estado_template', 'activa', 'veces_usada', 'fecha_modificacion')
    list_filter = ('activa', 'categoria', 'aprobada_twilio', 'fecha_modificacion')
    search_fields = ('nombre_interno', 'cuerpo_mensaje', 'twilio_template_sid')
    list_per_page = 50
    actions = ['crear_template_en_twilio', 'sincronizar_templates_twilio']

    fieldsets = (
        ('📝 Información Básica', {
            'fields': ('nombre_interno', 'categoria', 'activa'),
            'description': 'Configura el nombre y categoría de la plantilla'
        }),
        ('🎨 Personalización Visual', {
            'fields': ('emoji',),
            'description': mark_safe('''<div style="background:#f5f5f5;padding:15px;border-radius:8px;margin-top:10px;">
                <strong>💡 El emoji se autocompletará según la categoría seleccionada arriba</strong><br><br>
                <strong>Categorías disponibles:</strong><br>
                🌾 Cultivos • 🐄 Ganadería • 🌱 General Agrícola • 📚 Educación • 💼 Gestión<br><br>
                <em>Puedes cambiar el emoji manualmente si lo deseas</em>
            </div>''')
        }),
        ('📄 Contenido del Mensaje', {
            'fields': ('cuerpo_mensaje',),
            'description': 'Escribe el mensaje. Usa {nombre} para personalizar con el nombre del estudiante.'
        }),
        ('� Twilio (Opcional - Para Casos Avanzados)', {
            'fields': ('twilio_template_sid', 'twilio_template_nombre', 'aprobada_twilio'),
            'description': mark_safe('''<div style="background:#fff3e0;padding:15px;border-radius:8px;border-left:4px solid #ff9800;">
                <strong>⚠️ IMPORTANTE: Para campañas con Content Templates</strong><br><br>
                <strong>Flujo correcto:</strong>
                <ol style="margin:10px 0;">
                    <li>🔵 Ve a <a href="https://console.twilio.com/us1/develop/sms/content-editor" target="_blank" style="color:#2196F3;">Twilio Content Editor</a></li>
                    <li>📝 Crea tu plantilla de WhatsApp y obtén el <strong>Content SID</strong> (ej: HX1234...)</li>
                    <li>⚙️ Configura el SID arriba y marca como "Aprobada en Twilio"</li>
                    <li>✅ Las campañas usarán este template automáticamente</li>
                </ol>
                <em>💡 Si no configuras esto, las campañas usarán envío directo (sin template).</em>
            </div>'''),
            'classes': ('collapse',)
        }),
    )

    def nombre_con_emoji(self, obj):
        return str(obj)
    nombre_con_emoji.short_description = "Plantilla"

    def categoria_display(self, obj):
        return obj.get_categoria_display()
    categoria_display.short_description = "Categoría"
    
    def estado_template(self, obj):
        if obj.twilio_template_sid and obj.aprobada_twilio:
            return format_html('<span style="background:#4caf50;color:white;padding:4px 8px;border-radius:12px;font-size:11px;">✅ TWILIO</span>')
        elif obj.twilio_template_sid and not obj.aprobada_twilio:
            return format_html('<span style="background:#ff9800;color:white;padding:4px 8px;border-radius:12px;font-size:11px;">⏳ PENDIENTE</span>')
        else:
            return format_html('<span style="background:#2196f3;color:white;padding:4px 8px;border-radius:12px;font-size:11px;">📱 DIRECTO</span>')
    estado_template.short_description = "Estado"
    
    def total_plantillas(self, obj):
        count = obj.plantillas.count()
        if count > 0:
            return format_html('<span style="background:#4caf50;color:white;padding:4px 12px;border-radius:12px;font-weight:bold;">{}</span>', count)
        return format_html('<span style="color:#999;">0</span>')
    total_plantillas.short_description = "📄 Plantillas"
    
    def total_campanas(self, obj):
        count = obj.campanas.count()
        if count > 0:
            return format_html('<span style="background:#2196f3;color:white;padding:4px 12px;border-radius:12px;font-weight:bold;">{}</span>', count)
        return format_html('<span style="color:#999;">0</span>')
    total_campanas.short_description = "📢 Campañas"
    
    @admin.action(description='🔧 Crear Template en Twilio')
    def crear_template_en_twilio(self, request, queryset):
        """Crea automáticamente un Content Template en Twilio para las plantillas seleccionadas"""
        from .enviar_plantillas import crear_template_twilio
        
        creados = 0
        errores = []
        
        for plantilla in queryset:
            # Verificar que no tenga ya un template
            if plantilla.twilio_template_sid:
                errores.append(f"{plantilla.nombre_interno}: Ya tiene un template configurado")
                continue
            
            # Crear template en Twilio
            # Convertir {nombre} a {{1}} para Twilio
            contenido_twilio = plantilla.cuerpo_mensaje.replace('{nombre}', '{{1}}')
            
            resultado = crear_template_twilio(
                nombre=f"eki_{plantilla.nombre_interno.lower().replace(' ', '_')}",
                contenido=contenido_twilio,
                variables=['nombre']
            )
            
            if resultado['success']:
                # Guardar el Content SID en la plantilla
                plantilla.twilio_template_sid = resultado['content_sid']
                plantilla.twilio_template_nombre = f"eki_{plantilla.nombre_interno.lower().replace(' ', '_')}"
                plantilla.save()
                creados += 1
                logger.info(f"✅ Template creado para {plantilla.nombre_interno}: {resultado['content_sid']}")
            else:
                errores.append(f"{plantilla.nombre_interno}: {resultado['response']}")
        
        # Mensajes al usuario
        if creados > 0:
            self.message_user(
                request,
                f"✅ {creados} template(s) creado(s) en Twilio. "
                f"IMPORTANTE: Debes ir a Twilio Console para aprobar los templates antes de usarlos.",
                level=messages.SUCCESS
            )
        
        if errores:
            self.message_user(
                request,
                f"⚠️ Errores: {', '.join(errores)}",
                level=messages.WARNING
            )
    
    @admin.action(description='🔄 Sincronizar con Twilio')
    def sincronizar_templates_twilio(self, request, queryset):
        """Obtiene la lista de templates de Twilio y actualiza el estado de aprobación"""
        from .enviar_plantillas import listar_templates_twilio
        
        # Obtener templates de Twilio
        templates_twilio = listar_templates_twilio()
        
        if not templates_twilio:
            self.message_user(
                request,
                "⚠️ No se pudieron obtener los templates de Twilio. Verifica las credenciales.",
                level=messages.WARNING
            )
            return
        
        # Crear diccionario de SIDs para búsqueda rápida
        sids_twilio = {t['sid']: t for t in templates_twilio}
        
        actualizados = 0
        for plantilla in queryset:
            if plantilla.twilio_template_sid and plantilla.twilio_template_sid in sids_twilio:
                # Template existe en Twilio - marcarlo como aprobado
                if not plantilla.aprobada_twilio:
                    plantilla.aprobada_twilio = True
                    plantilla.save()
                    actualizados += 1
                    logger.info(f"✅ Template aprobado: {plantilla.nombre_interno}")
        
        self.message_user(
            request,
            f"🔄 Sincronización completa. {actualizados} template(s) marcado(s) como aprobado(s). "
            f"Total templates en Twilio: {len(templates_twilio)}",
            level=messages.SUCCESS
        )


@admin.register(Estudiante)
class EstudianteAdmin(admin.ModelAdmin):
    """Gestión de estudiantes/campesinos"""
    change_form_template = 'admin/core/estudiante/change_form.html'
    inlines = [HabilitacionModuloEstudianteInline]
    list_display = ('cedula_formateada', 'nombre', 'telefono_formateado', 'municipio', 'departamento', 'genero', 'edad', 'cliente_nombre', 'grupos_display', 'cursos_inscritos', 'activo', 'fecha_registro')
    list_filter = ('activo', 'cliente', 'genero', 'departamento', 'fecha_registro', CursosEstudianteFilter, GruposEstudianteFilter)
    search_fields = ('nombre', 'cedula', 'telefono', 'cliente__nombre')
    list_per_page = 50
    ordering = ('-fecha_registro',)
    actions = [
        'enviar_mensaje_masivo',
        'enviar_anuncio_grupal',
        'invitar_a_grupo_whatsapp',
        'exportar_estudiantes_por_curso',
        'exportar_plantilla_importacion',
        'asignar_a_grupo_accion',
        'asignar_cliente_masivo',
        'eliminar_estudiantes_seguro',
    ]
    
    # ✨ AGREGAR BOTÓN DE IMPORTAR EN LA PARTE SUPERIOR
    change_list_template = 'admin/estudiante_changelist.html'
    
    fieldsets = (
        ('Identificación y contacto', {
            'fields': ('tipo_documento', 'cedula', 'nombre', 'telefono', 'cliente', 'activo'),
            'description': 'Documento único; teléfono WhatsApp y organización del campesino.',
        }),
        ('Ubicación y perfil', {
            'fields': ('municipio', 'departamento', 'ubicacion_detalle', 'genero', 'edad', 'rango_edad'),
        }),
        ('Módulos drip (lista)', {
            'fields': (),
            'description': (
                'Si el cliente tiene activo «Módulos solo por lista», agregue en la tabla inferior '
                'cada combinación curso + módulo que este estudiante puede abrir.'
            ),
        }),
        ('Cursos inscritos', {
            'fields': ('mostrar_cursos_inscritos',),
        }),
    )
    readonly_fields = ('mostrar_cursos_inscritos',)
    
    def cedula_formateada(self, obj):
        """Muestra cédula con formato y tipo"""
        return format_html(
            '<strong style="color:#2196F3;">🪪 {} {}</strong>',
            obj.tipo_documento,
            obj.cedula
        )
    cedula_formateada.short_description = "Documento"
    
    def mostrar_cursos_inscritos(self, obj):
        """Muestra los cursos en los que está inscrito el estudiante"""
        progresos = ProgresoEstudiante.objects.filter(estudiante=obj).select_related('curso')
        if not progresos:
            return format_html('<p style="color:#999;">No está inscrito en ningún curso</p>')
        
        html = '<div style="margin-top:10px;">'
        for progreso in progresos:
            porcentaje = progreso.porcentaje_avance()
            color = '#4CAF50' if porcentaje == 100 else '#FF9800' if porcentaje > 0 else '#999'
            html += f'''
                <div style="margin-bottom:15px;padding:10px;background:#f5f5f5;border-radius:5px;">
                    <strong style="color:{color};">📚 {progreso.curso.nombre}</strong><br>
                    <small>Progreso: {porcentaje}%</small>
                    <div style="background:#e0e0e0;height:8px;border-radius:4px;margin-top:5px;">
                        <div style="background:{color};width:{porcentaje}%;height:100%;border-radius:4px;"></div>
                    </div>
                </div>
            '''
        html += '</div>'
        return format_html(html)
    mostrar_cursos_inscritos.short_description = "Cursos"
    
    def cursos_inscritos(self, obj):
        """Muestra cantidad de cursos inscritos"""
        count = ProgresoEstudiante.objects.filter(estudiante=obj).count()
        if count == 0:
            return format_html('<span style="color:#999;">Sin cursos</span>')
        return format_html(
            '<span style="background:#e3f2fd;padding:4px 10px;border-radius:12px;font-size:11px;">{} curso{}</span>',
            count, 's' if count != 1 else ''
        )
    cursos_inscritos.short_description = "📚 Cursos"
    
    def telefono_formateado(self, obj):
        """Muestra teléfono con formato WhatsApp"""
        return f"+{obj.telefono}"
    telefono_formateado.short_description = "📱 WhatsApp"
    
    def cliente_nombre(self, obj):
        """Muestra el cliente al que pertenece"""
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;">Sin cliente</span>')
    cliente_nombre.short_description = "🏢 Cliente"
    
    def grupos_display(self, obj):
        """Muestra los grupos a los que pertenece el estudiante"""
        grupos = obj.grupos.all()
        if not grupos:
            return format_html('<span style="color:#999;">Sin grupos</span>')
        
        badges = []
        for grupo in grupos[:3]:  # Mostrar máximo 3
            badges.append(f'<span style="background:#e8f5e9;color:#2e7d32;padding:3px 8px;border-radius:8px;font-size:10px;margin-right:3px;">{grupo.emoji} {grupo.nombre}</span>')
        
        html = ''.join(badges)
        if grupos.count() > 3:
            html += f' <span style="color:#999;font-size:10px;">+{grupos.count() - 3} más</span>'
        
        return format_html(html)
    grupos_display.short_description = "👥 Grupos"
    
    # Ver conversaciones removido por solicitud del usuario
    
    def get_urls(self):
        """Agregar URL personalizada para importar estudiantes"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('importar/', self.admin_site.admin_view(self.importar_estudiantes_view), name='importar_estudiantes'),
            path('exportar-plantilla/', self.admin_site.admin_view(self.exportar_plantilla_importacion), name='exportar_plantilla_importacion'),
        ]
        return custom_urls + urls
    
    def importar_estudiantes_view(self, request):
        """Vista para importar estudiantes desde Excel.
        Campos obligatorios: Cédula | Nombre | Teléfono.
        Campos opcionales: Municipio | Departamento | Género | Edad | Curso | Cliente.
        """
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from django.db import IntegrityError
        import re
        
        if request.method == 'POST':
            archivo = request.FILES.get('archivo_excel')
            
            if not archivo:
                messages.error(request, "⚠️ Debes seleccionar un archivo Excel")
                return redirect('admin:core_estudiante_changelist')
            
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                
                creados = 0
                actualizados = 0
                inscritos = 0
                errores = []
                
                def _normalizar_celda(val):
                    if val is None:
                        return ''
                    if isinstance(val, float):
                        if val == int(val):
                            return str(int(val))
                        return str(val)
                    if isinstance(val, int):
                        return str(val)
                    return str(val).strip()
                
                def _limpiar_texto(val):
                    if not val:
                        return ''
                    return re.sub(r'\s+', ' ', val.strip().lower())
                
                def _normalizar_telefono(raw):
                    tel = re.sub(r'\D', '', raw)
                    if tel.startswith('57') and len(tel) == 12:
                        return tel
                    if len(tel) == 10 and tel.startswith('3'):
                        return '57' + tel
                    if len(tel) == 7 or len(tel) == 10:
                        return '57' + tel
                    return tel
                
                GENEROS_VALIDOS = {'m': 'M', 'f': 'F', 'o': 'O', 'masculino': 'M', 'femenino': 'F',
                                   'otro': 'O', 'hombre': 'M', 'mujer': 'F', 'nr': 'NR', 'no reporta': 'NR'}
                
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row[:3]):
                        continue
                    
                    try:
                        cedula = _normalizar_celda(row[0]) if len(row) > 0 else ''
                        nombre = _normalizar_celda(row[1]) if len(row) > 1 else ''
                        telefono_raw = _normalizar_celda(row[2]) if len(row) > 2 else ''
                        municipio = _limpiar_texto(_normalizar_celda(row[3])) if len(row) > 3 else ''
                        departamento = _limpiar_texto(_normalizar_celda(row[4])) if len(row) > 4 else ''
                        genero_raw = _limpiar_texto(_normalizar_celda(row[5])) if len(row) > 5 else ''
                        edad_raw = _normalizar_celda(row[6]) if len(row) > 6 else ''
                        curso_nombre = _normalizar_celda(row[7]) if len(row) > 7 else ''
                        cliente_nombre = _normalizar_celda(row[8]) if len(row) > 8 else ''
                    except IndexError:
                        errores.append(f"Fila {idx}: Columnas insuficientes")
                        continue
                    
                    # Validar obligatorios mínimos
                    campos_faltantes = []
                    if not cedula: campos_faltantes.append('Cédula')
                    if not nombre: campos_faltantes.append('Nombre')
                    if not telefono_raw: campos_faltantes.append('Teléfono')
                    
                    if campos_faltantes:
                        errores.append(f"Fila {idx}: Faltan: {', '.join(campos_faltantes)}")
                        continue
                    
                    telefono = _normalizar_telefono(telefono_raw)
                    if not telefono or len(telefono) < 10:
                        errores.append(f"Fila {idx}: Teléfono inválido '{telefono_raw}'")
                        continue
                    
                    genero = GENEROS_VALIDOS.get(genero_raw, '') if genero_raw else ''
                    if not genero:
                        genero = 'NR'
                    
                    # Validar edad
                    edad = None
                    if edad_raw:
                        try:
                            edad = int(re.sub(r'\D', '', edad_raw))
                            if edad < 1 or edad > 120:
                                errores.append(f"Fila {idx}: Edad '{edad_raw}' fuera de rango (1-120)")
                                continue
                        except (ValueError, TypeError):
                            errores.append(f"Fila {idx}: Edad '{edad_raw}' no es un número válido")
                            continue
                    
                    try:
                        cliente = None
                        if cliente_nombre:
                            try:
                                cliente = Cliente.objects.get(nombre__iexact=cliente_nombre.strip())
                            except Cliente.DoesNotExist:
                                errores.append(f"Fila {idx}: Cliente '{cliente_nombre}' no encontrado")
                        
                        try:
                            defaults = {
                                'nombre': nombre.strip().title(),
                                'telefono': telefono,
                                'municipio': municipio,
                                'departamento': departamento,
                                'genero': genero,
                                'edad': edad,
                                'tipo_documento': 'CC',
                                'estado_onboarding': 'completado',
                                'estado_chat': 'ACTIVO',
                                'acepto_terminos': True,
                                'activo': True,
                            }
                            if cliente:
                                defaults['cliente'] = cliente
                            
                            estudiante, created = Estudiante.objects.update_or_create(
                                cedula=cedula,
                                defaults=defaults
                            )
                            if created:
                                creados += 1
                            else:
                                actualizados += 1
                        except IntegrityError as e:
                            if 'telefono' in str(e).lower():
                                errores.append(f"Fila {idx}: Teléfono '{telefono}' ya existe para otro estudiante")
                            else:
                                errores.append(f"Fila {idx}: Error de integridad - {str(e)}")
                            continue
                        
                        if curso_nombre:
                            try:
                                curso = Curso.objects.get(nombre__iexact=curso_nombre.strip())
                                progreso, creado_prog = ProgresoEstudiante.objects.get_or_create(
                                    estudiante=estudiante,
                                    curso=curso,
                                    defaults={'progreso': 0, 'completado': False}
                                )
                                if creado_prog:
                                    inscritos += 1
                            except Curso.DoesNotExist:
                                errores.append(f"Fila {idx}: Curso '{curso_nombre}' no encontrado")
                    
                    except Exception as e:
                        errores.append(f"Fila {idx}: {str(e)}")
                
                if creados > 0:
                    messages.success(request, f"✅ {creados} estudiante(s) creado(s)")
                if actualizados > 0:
                    messages.info(request, f"ℹ️ {actualizados} estudiante(s) actualizado(s)")
                if inscritos > 0:
                    messages.success(request, f"🎓 {inscritos} inscripción(es) en cursos")
                if errores:
                    messages.warning(request, f"⚠️ {len(errores)} error(es)")
                    for error in errores[:5]:
                        messages.warning(request, error)
                
                return redirect('admin:core_estudiante_changelist')
            
            except Exception as e:
                messages.error(request, f"❌ Error procesando archivo: {str(e)}")
                return redirect('admin:core_estudiante_changelist')
        
        return render(request, 'admin/importar_estudiantes.html', {
            'title': 'Importar Estudiantes desde Excel',
            'site_header': 'Importar Estudiantes',
        })
    
    def total_mensajes(self, obj):
        """Cuenta cuántos mensajes ha enviado el estudiante"""
        count = WhatsappLog.objects.filter(telefono=obj.telefono).count()
        return format_html(
            '<span style="background:#e3f2fd;padding:4px 8px;border-radius:4px;">{} mensajes</span>',
            count
        )
    total_mensajes.short_description = "💬 Total"
    
    # ========================================
    # 📢 MEGÁFONO: ENVÍO MASIVO DE MENSAJES
    # ========================================
    
    @admin.action(description='📢 Megáfono: Enviar mensaje a seleccionados')
    def enviar_mensaje_masivo(self, request, queryset):
        """
        Permite enviar un mensaje personalizado a múltiples estudiantes seleccionados.
        Útil para avisos de mantenimiento, anuncios importantes, etc.
        
        IMPORTANTE: Valida la ventana de 24 horas de WhatsApp para evitar bloqueos.
        """
        from datetime import timedelta
        
        if 'aplicar' in request.POST:
            # El usuario confirmó el envío
            mensaje = request.POST.get('mensaje')
            if not mensaje:
                self.message_user(request, "⚠️ Debes escribir un mensaje", level=messages.ERROR)
                return
            
            # Verificar que no esté vacío
            if not mensaje.strip():
                self.message_user(request, "⚠️ El mensaje no puede estar vacío", level=messages.ERROR)
                return
            
            enviados = 0
            errores = 0
            errores_detalle = []
            
            for estudiante in queryset:
                try:
                    # Personalizar mensaje con nombre del estudiante
                    mensaje_personalizado = mensaje.replace('{nombre}', estudiante.nombre)
                    mensaje_personalizado = mensaje_personalizado.replace('{cedula}', estudiante.cedula)
                    
                    # Enviar mensaje
                    resultado = enviar_whatsapp(estudiante.telefono, mensaje_personalizado)
                    
                    if resultado:
                        enviados += 1
                        logger.info(f"📢 Megáfono: Mensaje enviado a {estudiante.nombre}")
                        
                        # Registrar en WhatsappLog
                        WhatsappLog.objects.create(
                            telefono=estudiante.telefono,
                            mensaje=mensaje_personalizado,
                            tipo='SENT',
                            estado='ENVIADO',
                            estudiante=estudiante,
                            mensaje_id=f'megafono_{timezone.now().timestamp()}'
                        )
                    else:
                        errores += 1
                        errores_detalle.append(f"{estudiante.nombre} ({estudiante.telefono})")
                        logger.error(f"❌ Megáfono: Error enviando a {estudiante.nombre}")
                        
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"{estudiante.nombre}: {str(e)}")
                    logger.error(f"❌ Megáfono: Error con {estudiante.nombre}: {e}")
            
            # Mostrar resultados
            if enviados > 0:
                self.message_user(
                    request,
                    f"✅ Mensaje enviado exitosamente a {enviados} estudiante(s)",
                    level=messages.SUCCESS
                )
            
            if errores > 0:
                mensaje_error = f"⚠️ Hubo {errores} error(es) al enviar"
                if errores_detalle:
                    mensaje_error += f": {', '.join(errores_detalle[:5])}"
                    if len(errores_detalle) > 5:
                        mensaje_error += f" y {len(errores_detalle) - 5} más..."
                self.message_user(request, mensaje_error, level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # Validar ventana de 24 horas ANTES de mostrar formulario
        hace_24h = timezone.now() - timedelta(hours=24)
        activos = []
        inactivos = []
        sin_interaccion = []
        
        for estudiante in queryset:
            # Obtener último mensaje RECIBIDO del estudiante (no enviado por nosotros)
            ultimo_msg = WhatsappLog.objects.filter(
                estudiante=estudiante,
                tipo_mensaje='recibido'  # Solo mensajes que EL estudiante nos envió
            ).order_by('-timestamp').first()
            
            if ultimo_msg:
                if ultimo_msg.timestamp >= hace_24h:
                    activos.append({
                        'estudiante': estudiante,
                        'ultima_interaccion': ultimo_msg.timestamp,
                        'hace': timezone.now() - ultimo_msg.timestamp
                    })
                else:
                    inactivos.append({
                        'estudiante': estudiante,
                        'ultima_interaccion': ultimo_msg.timestamp,
                        'hace': timezone.now() - ultimo_msg.timestamp
                    })
            else:
                # Nunca ha escrito
                sin_interaccion.append(estudiante)
        
        # Mostrar formulario de confirmación con advertencias
        return render(request, 'admin/enviar_mensaje_masivo.html', {
            'estudiantes': queryset,
            'total_estudiantes': queryset.count(),
            'activos': activos,
            'inactivos': inactivos,
            'sin_interaccion': sin_interaccion,
            'total_activos': len(activos),
            'total_inactivos': len(inactivos),
            'total_sin_interaccion': len(sin_interaccion),
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
            'title': 'Enviar Mensaje Masivo (Megáfono)',
        })
    
    # ========================================
    # 📢 ANUNCIO GRUPAL (Difusión Simulada)
    # ========================================
    
    @admin.action(description='📣 Enviar anuncio grupal (difusión)')
    def enviar_anuncio_grupal(self, request, queryset):
        """
        Envía el mismo mensaje a múltiples estudiantes usando template aprobado de Twilio.
        Ideal para: PQRS, anuncios de clases, notificaciones importantes.
        """
        from .utils import enviar_whatsapp_twilio_content_template
        
        if 'aplicar' in request.POST:
            mensaje = request.POST.get('mensaje', '').strip()
            
            # Debug: Ver qué datos llegan
            logger.info(f"📋 POST data: {dict(request.POST)}")
            
            if not mensaje:
                self.message_user(request, "⚠️ Debes escribir un mensaje", level=messages.ERROR)
                # Regresar al formulario con los estudiantes seleccionados
                return render(request, 'admin/enviar_anuncio_grupal.html', {
                    'estudiantes': queryset,
                    'total_estudiantes': queryset.count(),
                    'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
                    'title': 'Enviar Anuncio Grupal (Difusión)',
                    'mensaje': mensaje,  # Preservar mensaje ingresado
                })
            
            # Reconstruir queryset desde los IDs en POST
            estudiante_ids = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
            if estudiante_ids:
                queryset = Estudiante.objects.filter(id__in=estudiante_ids)
            
            if not queryset.exists():
                self.message_user(request, "⚠️ No se encontraron estudiantes seleccionados", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')
            
            # Verificar que el template esté configurado
            content_sid = getattr(settings, 'TWILIO_TEMPLATE_ANUNCIO_GRUPAL', None)
            logger.info(f"🔧 Content SID: {content_sid}")
            
            if not content_sid:
                self.message_user(
                    request, 
                    "⚠️ Template de Twilio no configurado. Configura TWILIO_TEMPLATE_ANUNCIO_GRUPAL en .env",
                    level=messages.ERROR
                )
                # Regresar al formulario con el mensaje preservado
                return render(request, 'admin/enviar_anuncio_grupal.html', {
                    'estudiantes': queryset,
                    'total_estudiantes': queryset.count(),
                    'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
                    'title': 'Enviar Anuncio Grupal (Difusión)',
                    'mensaje': mensaje,
                })
            
            enviados = 0
            errores = 0
            errores_detalle = []
            
            for estudiante in queryset:
                try:
                    # Variables del template:
                    # {{1}} = Nombre del estudiante
                    # {{2}} = Contenido del mensaje
                    # {{3}} = Enlace del grupo
                    variables = {
                        '1': estudiante.nombre,
                        '2': mensaje
                    }
                    
                    # Enviar con template de Twilio
                    resultado = enviar_whatsapp_twilio_content_template(
                        telefono=estudiante.telefono,
                        content_sid=content_sid,
                        variables=variables
                    )
                    
                    if resultado.get('success'):
                        enviados += 1
                        logger.info(f"📣 Anuncio grupal enviado a {estudiante.nombre}")
                    else:
                        errores += 1
                        errores_detalle.append(estudiante.nombre)
                        logger.error(f"❌ Error enviando anuncio a {estudiante.nombre}: {resultado.get('response')}")
                        
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"{estudiante.nombre}: {str(e)}")
                    logger.error(f"❌ Error con {estudiante.nombre}: {e}")
            
            # Mostrar resultados
            if enviados > 0:
                self.message_user(
                    request,
                    f"✅ Anuncio enviado exitosamente a {enviados} estudiante(s)",
                    level=messages.SUCCESS
                )
            
            if errores > 0:
                mensaje_error = f"⚠️ Hubo {errores} error(es)"
                if errores_detalle:
                    mensaje_error += f": {', '.join(errores_detalle[:3])}"
                    if len(errores_detalle) > 3:
                        mensaje_error += f" y {len(errores_detalle) - 3} más"
                self.message_user(request, mensaje_error, level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # Mostrar formulario
        return render(request, 'admin/enviar_anuncio_grupal.html', {
            'estudiantes': queryset,
            'total_estudiantes': queryset.count(),
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
            'title': 'Enviar Anuncio Grupal (Difusión)',
        })
    
    # ========================================
    # 👥 INVITACIÓN A GRUPO DE WHATSAPP
    # ========================================
    
    @admin.action(description='🔗 Invitar a grupo de WhatsApp')
    def invitar_a_grupo_whatsapp(self, request, queryset):
        """
        Envía invitación con enlace al grupo de WhatsApp.
        Puede usar el enlace del curso o del cliente, o uno personalizado.
        """
        if 'aplicar' in request.POST:
            enlace_tipo = request.POST.get('enlace_tipo')  # 'curso', 'cliente', 'personalizado'
            curso_id = request.POST.get('curso_id')
            enlace_personalizado = request.POST.get('enlace_personalizado')
            mensaje_intro = request.POST.get('mensaje_intro', '')
            
            # Determinar el enlace a usar
            enlace_grupo = None
            
            if enlace_tipo == 'curso' and curso_id:
                try:
                    curso = Curso.objects.get(id=curso_id)
                    enlace_grupo = curso.enlace_grupo_whatsapp
                    if not enlace_grupo:
                        self.message_user(
                            request,
                            f"⚠️ El curso '{curso.nombre}' no tiene un enlace de grupo configurado",
                            level=messages.ERROR
                        )
                        return
                except Curso.DoesNotExist:
                    self.message_user(request, "⚠️ Curso no encontrado", level=messages.ERROR)
                    return
            
            elif enlace_tipo == 'cliente':
                # Usar el enlace del cliente del primer estudiante
                primer_estudiante = queryset.first()
                if primer_estudiante and primer_estudiante.cliente:
                    enlace_grupo = primer_estudiante.cliente.enlace_grupo_whatsapp
                    if not enlace_grupo:
                        self.message_user(
                            request,
                            f"⚠️ El cliente '{primer_estudiante.cliente.nombre}' no tiene un enlace de grupo configurado",
                            level=messages.ERROR
                        )
                        return
                else:
                    self.message_user(
                        request,
                        "⚠️ Los estudiantes no tienen cliente asignado o el cliente no tiene enlace de grupo",
                        level=messages.ERROR
                    )
                    return
            
            elif enlace_tipo == 'personalizado':
                enlace_grupo = enlace_personalizado
                if not enlace_grupo or not enlace_grupo.startswith('https://chat.whatsapp.com/'):
                    self.message_user(
                        request,
                        "⚠️ Enlace personalizado inválido. Debe ser: https://chat.whatsapp.com/xxxxx",
                        level=messages.ERROR
                    )
                    return
            
            if not enlace_grupo:
                self.message_user(request, "⚠️ No se pudo determinar el enlace del grupo", level=messages.ERROR)
                return
            
            # Construir mensaje
            mensaje_base = mensaje_intro if mensaje_intro else "¡Hola! Te invitamos a unirte a nuestro grupo de WhatsApp:"
            
            enviados = 0
            errores = 0
            errores_detalle = []
            
            for estudiante in queryset:
                try:
                    from .utils import enviar_whatsapp_twilio_content_template
                    
                    # Verificar que el template esté configurado
                    content_sid = settings.TWILIO_TEMPLATE_INVITACION_GRUPO
                    if not content_sid:
                        self.message_user(
                            request,
                            "⚠️ Template de invitación no configurado. Configura TWILIO_TEMPLATE_INVITACION_GRUPO en .env",
                            level=messages.ERROR
                        )
                        return
                    
                    # Variables del template:
                    # {{1}} = Nombre del estudiante
                    # {{2}} = Mensaje de introducción (opcional)
                    # {{3}} = Enlace del grupo
                    mensaje_intro_final = mensaje_intro if mensaje_intro else "¡Hola! Te invitamos a unirte a nuestro grupo de WhatsApp."
                    
                    variables = {
                        '1': estudiante.nombre,
                        '2': mensaje_intro_final,
                        '3': enlace_grupo
                    }
                    
                    # Enviar con template de Twilio
                    resultado = enviar_whatsapp_twilio_content_template(
                        telefono=estudiante.telefono,
                        content_sid=content_sid,
                        variables=variables
                    )
                    
                    if resultado.get('success'):
                        enviados += 1
                        logger.info(f"🔗 Invitación a grupo enviada a {estudiante.nombre}")
                    else:
                        errores += 1
                        errores_detalle.append(estudiante.nombre)
                        logger.error(f"❌ Error enviando invitación a {estudiante.nombre}: {resultado.get('response')}")
                        
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"{estudiante.nombre}: {str(e)}")
                    logger.error(f"❌ Error con {estudiante.nombre}: {e}")
            
            # Mostrar resultados
            if enviados > 0:
                self.message_user(
                    request,
                    f"✅ Invitación enviada exitosamente a {enviados} estudiante(s)",
                    level=messages.SUCCESS
                )
            
            if errores > 0:
                mensaje_error = f"⚠️ {errores} error(es)"
                if errores_detalle:
                    mensaje_error += f": {', '.join(errores_detalle[:3])}"
                self.message_user(request, mensaje_error, level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # GET - Mostrar formulario
        # Obtener cursos y clientes con enlaces configurados
        cursos_con_enlace = Curso.objects.exclude(enlace_grupo_whatsapp='').values('id', 'nombre', 'emoji')
        
        # Verificar si los estudiantes tienen cliente
        clientes_ids = queryset.values_list('cliente_id', flat=True).distinct()
        clientes_con_enlace = Cliente.objects.filter(
            id__in=clientes_ids
        ).exclude(enlace_grupo_whatsapp='').values('id', 'nombre')
        
        return render(request, 'admin/invitar_grupo_whatsapp.html', {
            'estudiantes': queryset,
            'total_estudiantes': queryset.count(),
            'cursos_con_enlace': cursos_con_enlace,
            'clientes_con_enlace': clientes_con_enlace,
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
            'title': 'Invitar a Grupo de WhatsApp',
        })
    
    # ✅ Acción para enviar mensaje manual de prueba
    def enviar_mensaje_manual(self, request, queryset):
        """Permite enviar un mensaje de prueba a los estudiantes seleccionados"""
        if 'aplicar' in request.POST:
            # El usuario confirmó el envío
            mensaje = request.POST.get('mensaje')
            if not mensaje:
                self.message_user(request, "⚠️ Debes escribir un mensaje", level=messages.ERROR)
                return
            
            enviados = 0
            errores = 0
            for estudiante in queryset:
                try:
                    telefono = estudiante.telefono
                    if not telefono.startswith('whatsapp:'):
                        telefono = f'whatsapp:{telefono}'
                    
                    # Enviar con Twilio
                    resultado = enviar_whatsapp_twilio(
                        telefono=telefono,
                        texto=mensaje,
                        mensaje_id_referencia=None
                    )
                    
                    if resultado.get('success'):
                        # Registrar en log
                        WhatsappLog.objects.create(
                            telefono=estudiante.telefono,
                            mensaje=mensaje,
                            mensaje_id=resultado.get('mensaje_id'),
                            tipo='SENT',
                            estado='sent'
                        )
                        enviados += 1
                        logger.info(f"✅ Mensaje manual enviado a {estudiante.nombre} ({estudiante.telefono})")
                    else:
                        errores += 1
                        logger.error(f"❌ Error al enviar a {estudiante.telefono}: {resultado.get('error')}")
                
                except Exception as e:
                    errores += 1
                    logger.error(f"❌ Excepción al enviar a {estudiante.telefono}: {str(e)}")
            
            if enviados > 0:
                self.message_user(request, f"✅ Mensaje enviado exitosamente a {enviados} estudiante(s)", level=messages.SUCCESS)
            if errores > 0:
                self.message_user(request, f"⚠️ Hubo {errores} error(es) al enviar", level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # Mostrar formulario de confirmación
        return render(request, 'admin/enviar_mensaje_manual.html', {
            'estudiantes': queryset,
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        })
    
    enviar_mensaje_manual.short_description = "📤 Enviar mensaje de prueba"
    
    @admin.action(description='📊 Exportar estudiantes con cursos a Excel')
    def exportar_estudiantes_por_curso(self, request, queryset):
        """Exporta estudiantes con información detallada de cursos"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        from core.export_estudiantes import limpiar_telefono
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Estudiantes por Curso"
        
        # Encabezados
        headers = [
            'Nombre', 'Apellido', 'Documento', 'Teléfono', 'Email',
            'Municipio', 'Departamento', 'Ciudad',
            'Organización', 'Grupo', 'Curso', 'Estado', 'Progreso (%)',
            'Fecha registro', 'Activo',
            'Cursos Inscritos', 'Cursos Completados', 'Progreso Promedio',
            'Total Mensajes',
        ]
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for estudiante in queryset.select_related('cliente').prefetch_related('grupos'):
            # Calcular estadísticas de cursos
            progresos = ProgresoEstudiante.objects.filter(estudiante=estudiante).select_related('curso')
            total_cursos = progresos.count()
            cursos_completados = progresos.filter(completado=True).count()
            
            # Calcular promedio de progreso
            if total_cursos > 0:
                progreso_promedio = sum([p.porcentaje_avance() for p in progresos]) / total_cursos
            else:
                progreso_promedio = 0

            primer_progreso = progresos.first()
            curso_nombre = primer_progreso.curso.nombre if primer_progreso and primer_progreso.curso_id else ''
            if total_cursos > 1:
                curso_nombre = f"{curso_nombre} (+{total_cursos - 1})" if curso_nombre else f"{total_cursos} cursos"
            estado_prog = 'Sin inscripción'
            pct_prog = 0
            if primer_progreso:
                pct_prog = primer_progreso.porcentaje_avance()
                estado_prog = 'Completado' if primer_progreso.completado else (
                    'En curso' if pct_prog > 0 else 'Sin avance'
                )
            
            # Total mensajes
            total_mensajes = WhatsappLog.objects.filter(telefono=estudiante.telefono).count()
            grupos_txt = ', '.join(g.nombre for g in estudiante.grupos.all()[:5])
            municipio = (estudiante.municipio or '').strip()
            departamento = (getattr(estudiante, 'departamento', '') or '').strip()
            ciudad = municipio
            
            ws.append([
                estudiante.nombre,
                '',
                estudiante.cedula,
                limpiar_telefono(estudiante.telefono),
                '',
                municipio,
                departamento,
                ciudad,
                estudiante.cliente.nombre if estudiante.cliente else 'Sin cliente',
                grupos_txt,
                curso_nombre,
                estado_prog,
                f"{pct_prog:.1f}%",
                estudiante.fecha_registro.strftime('%Y-%m-%d %H:%M'),
                'Sí' if estudiante.activo else 'No',
                total_cursos,
                cursos_completados,
                f"{progreso_promedio:.1f}%",
                total_mensajes,
            ])
        
        # Crear segunda hoja con detalle de cursos
        ws2 = wb.create_sheet("Detalle por Curso")
        headers2 = [
            'Documento', 'Estudiante', 'Teléfono', 'Municipio', 'Departamento',
            'Curso', 'Progreso', 'Estado', 'Fecha Inicio',
        ]
        ws2.append(headers2)
        
        # Estilo
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Detalle de cada curso
        for estudiante in queryset:
            progresos = ProgresoEstudiante.objects.filter(estudiante=estudiante).select_related('curso')
            for progreso in progresos:
                ws2.append([
                    estudiante.cedula,
                    estudiante.nombre,
                    limpiar_telefono(estudiante.telefono),
                    estudiante.municipio or '',
                    getattr(estudiante, 'departamento', '') or '',
                    progreso.curso.nombre,
                    f"{progreso.porcentaje_avance()}%",
                    "Completado" if progreso.completado else "En progreso",
                    progreso.fecha_inicio.strftime('%Y-%m-%d'),
                ])
        
        # Ajustar ancho de columnas
        for ws in [wb.active, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=estudiantes_cursos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        self.message_user(request, f"✅ Exportados {queryset.count()} estudiantes con detalle de cursos", level=messages.SUCCESS)
        return response
    
    def exportar_plantilla_importacion(self, request):
        """Genera plantilla Excel mejorada con curso y cliente"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.comments import Comment
        # from openpyxl.data_validation import DataValidation  # Eliminado, no se usa
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Estudiantes"
        
        # Encabezados mejorados (9 columnas, 7 obligatorias)
        headers = ['Cédula', 'Nombre Completo', 'Teléfono', 'Municipio', 'Departamento', 'Género', 'Edad', 'Curso', 'Cliente']
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Comentarios con instrucciones
        ws['A1'].comment = Comment("📝 Cédula sin puntos ni espacios\nEjemplo: 1234567890", "eki")
        ws['B1'].comment = Comment("👤 Nombre completo del estudiante\nEjemplo: Juan Pérez García", "eki")
        ws['C1'].comment = Comment("📱 WhatsApp con código de país\nEjemplo: 573001234567 o 3001234567", "eki")
        ws['D1'].comment = Comment("🏙️ Municipio del estudiante (obligatorio)\nEjemplo: Manizales", "eki")
        ws['E1'].comment = Comment("🗺️ Departamento del estudiante (obligatorio)\nEjemplo: Caldas", "eki")
        ws['F1'].comment = Comment("👫 Género del estudiante (obligatorio)\nValores: masculino, femenino, otro, no reporta", "eki")
        ws['G1'].comment = Comment("🎂 Edad del estudiante (obligatorio)\nEjemplo: 35", "eki")
        ws['H1'].comment = Comment("📚 Nombre del curso (opcional)\nEjemplo: Curso de Café\nDeja vacío si no aplica", "eki")
        ws['I1'].comment = Comment("🏢 Nombre del cliente (opcional)\nEjemplo: FNC\nDeja vacío si no aplica", "eki")
        
        # Obtener cursos y clientes para validación
        cursos = Curso.objects.filter(activo=True).order_by('nombre')
        clientes = Cliente.objects.filter(activo=True).order_by('nombre')
        
        # Agregar ejemplos con datos reales
        if cursos.exists() and clientes.exists():
            curso_ejemplo = cursos.first().nombre
            cliente_ejemplo = clientes.first().nombre
            ws.append(['1234567890', 'Juan Pérez García', '573001234567', 'Manizales', 'Caldas', 'masculino', 35, curso_ejemplo, cliente_ejemplo])
            ws.append(['9876543210', 'María López Rodríguez', '3109876543', 'Bogotá', 'Cundinamarca', 'femenino', 28, curso_ejemplo, cliente_ejemplo])
            ws.append(['5555555555', 'Pedro Gómez Hernández', '3201234567', 'Medellín', 'Antioquia', 'otro', 52, '', ''])
        else:
            ws.append(['1234567890', 'Juan Pérez García', '573001234567', 'Manizales', 'Caldas', 'masculino', 35, 'Curso de Café', 'FNC'])
            ws.append(['9876543210', 'María López Rodríguez', '3109876543', 'Bogotá', 'Cundinamarca', 'femenino', 28, 'Curso de Aguacate', 'Fedecacao'])
            ws.append(['5555555555', 'Pedro Gómez Hernández', '3201234567', 'Medellín', 'Antioquia', 'otro', 52, '', ''])
        
        # Fila vacía para empezar
        ws.append(['', '', '', '', '', '', '', '', ''])
        
        # Estilo para ejemplos
        example_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        for row in [2, 3, 4]:
            for cell in ws[row]:
                cell.fill = example_fill
                cell.font = Font(italic=True, color="666666", size=10)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 25
        
        # Crear hoja con listas de valores disponibles
        ws_ref = wb.create_sheet("Valores Disponibles")
        ws_ref.append(["CURSOS DISPONIBLES", "CLIENTES DISPONIBLES"])
        ws_ref['A1'].font = Font(bold=True, size=12, color="2196F3")
        ws_ref['B1'].font = Font(bold=True, size=12, color="2196F3")
        
        max_rows = max(cursos.count(), clientes.count())
        for i in range(max_rows):
            curso_nombre = cursos[i].nombre if i < cursos.count() else ""
            cliente_nombre = clientes[i].nombre if i < clientes.count() else ""
            ws_ref.append([curso_nombre, cliente_nombre])
        
        ws_ref.column_dimensions['A'].width = 35
        ws_ref.column_dimensions['B'].width = 35
        
        # Agregar hoja de instrucciones
        ws_inst = wb.create_sheet("Instrucciones")
        instrucciones = [
            ["📋 GUÍA RÁPIDA - IMPORTAR ESTUDIANTES A eki"],
            [""],
            ["✅ CAMPOS OBLIGATORIOS (7):"],
            ["   • Cédula: Sin puntos ni espacios (Ej: 1234567890)"],
            ["   • Nombre: Nombre completo del estudiante"],
            ["   • Teléfono: Con código 57 o sin él (Ej: 573001234567 o 3001234567)"],
            ["   • Municipio: Ciudad o municipio (Ej: Manizales)"],
            ["   • Departamento: Departamento (Ej: Caldas)"],
            ["   • Género: masculino, femenino, otro o no reporta"],
            ["   • Edad: Edad en años (Ej: 35)"],
            [""],
            ["📝 CAMPOS OPCIONALES (2):"],
            ["   • Curso: Nombre del curso a inscribir (ver hoja 'Valores Disponibles')"],
            ["   • Cliente: Organización del estudiante (ver hoja 'Valores Disponibles')"],
            [""],
            ["💡 TIPS:"],
            ["   1. Si dejas vacío 'Curso', el estudiante NO se inscribirá automáticamente"],
            ["   2. Si dejas vacío 'Cliente', el estudiante quedará sin organización"],
            ["   3. Los nombres de Curso y Cliente deben coincidir EXACTAMENTE con los disponibles"],
            ["   4. Copia y pega desde la hoja 'Valores Disponibles' para evitar errores"],
            ["   5. Los 7 primeros campos son OBLIGATORIOS — filas incompletas serán rechazadas"],
            ["   6. Los textos se normalizan automáticamente a minúsculas"],
            [""],
            ["📊 PROCESO:"],
            ["   1. Completa la hoja 'Plantilla Estudiantes' con tus datos"],
            ["   2. Guarda el archivo Excel"],
            ["   3. Ve a eki Admin → Estudiantes → Botón 'Importar desde Excel'"],
            ["   4. Sube el archivo y confirma"],
            ["   5. ¡Listo! El sistema procesará automáticamente"],
            [""],
            ["✨ EJEMPLOS:"],
            [""],
            ["   Cédula      | Nombre              | Teléfono      | Municipio   | Departamento  | Género    | Edad | Curso             | Cliente"],
            ["   1234567890  | Juan Pérez García   | 573001234567  | Manizales   | Caldas        | masculino | 35   | Curso de Café     | FNC"],
            ["   9876543210  | María López         | 3109876543    | Bogotá      | Cundinamarca  | femenino  | 28   | Curso de Aguacate | Fedecacao"],
            ["   5555555555  | Pedro Gómez         | 3201234567    | Medellín    | Antioquia     | otro      | 52   |                   |"],
            [""],
            ["⚠️ ERRORES COMUNES:"],
            ["   • Cédula duplicada: Cada cédula debe ser única"],
            ["   • Teléfono duplicado: Cada teléfono debe ser único"],
            ["   • Curso inexistente: Verifica en 'Valores Disponibles'"],
            ["   • Campos obligatorios vacíos: Los 7 primeros campos son requeridos"],
            ["   • Filas vacías: No dejes filas vacías entre estudiantes"],
            [""],
            ["📞 Soporte: contacto@eki.com | WhatsApp: +57 300 123 4567"],
        ]
        
        for row_data in instrucciones:
            ws_inst.append(row_data)
        
        # Estilo del título
        ws_inst['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws_inst['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar ancho
        ws_inst.column_dimensions['A'].width = 80
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=PLANTILLA_IMPORTACION_ESTUDIANTES_eki.xlsx'
        wb.save(response)
        
        # Solo mostrar mensaje si es una acción de admin
        self.message_user(request, "Plantilla de importación descargada. Comparte este archivo con tus clientes.", level=messages.SUCCESS)
        
        return response
    
    def delete_model(self, request, obj):
        """Eliminar un estudiante individual con manejo seguro de relaciones CASCADE"""
        try:
            nombre = obj.nombre or obj.cedula
            # Eliminar relaciones en orden para evitar timeouts
            from .models import (
                WhatsappLog, EnvioLog, ProgresoEstudiante, ModuloCompletado,
                ResultadoExamen, RespuestaEjercicio, InteraccionLog,
                SolicitudSoporte, Certificado
            )
            from .gamificacion import PerfilGamificacion, BadgeEstudiante, TransaccionPuntos
            from .models_extras import EnvioProgramado, PQRS, InvitacionGrupo, GrupoEstudiantes
            from .recompensas import CanjeRecompensa
            
            # Limpiar relaciones vinculadas al teléfono
            if obj.telefono:
                WhatsappLog.objects.filter(telefono=obj.telefono).delete()
            
            # Limpiar WhatsappLog por FK (SET_NULL)
            WhatsappLog.objects.filter(estudiante=obj).update(estudiante=None)
            
            # Limpiar relaciones por FK directas
            EnvioLog.objects.filter(estudiante=obj).delete()
            InteraccionLog.objects.filter(estudiante=obj).delete()
            SolicitudSoporte.objects.filter(estudiante=obj).delete()
            RespuestaEjercicio.objects.filter(estudiante=obj).delete()
            ResultadoExamen.objects.filter(estudiante=obj).delete()
            
            # Modelos extras
            EnvioProgramado.objects.filter(estudiante=obj).delete()
            PQRS.objects.filter(estudiante=obj).delete()
            InvitacionGrupo.objects.filter(estudiante=obj).delete()
            CanjeRecompensa.objects.filter(estudiante=obj).delete()
            
            # M2M relations
            for grupo in GrupoEstudiantes.objects.filter(estudiantes=obj):
                grupo.estudiantes.remove(obj)
            
            # Progresos y sus dependencias
            progresos = ProgresoEstudiante.objects.filter(estudiante=obj)
            for prog in progresos:
                ModuloCompletado.objects.filter(progreso=prog).delete()
            progresos.delete()
            
            Certificado.objects.filter(estudiante=obj).delete()
            TransaccionPuntos.objects.filter(perfil__estudiante=obj).delete()
            BadgeEstudiante.objects.filter(estudiante=obj).delete()
            PerfilGamificacion.objects.filter(estudiante=obj).delete()
            
            obj.delete()
            self.message_user(request, f"✅ Estudiante '{nombre}' eliminado correctamente", level=messages.SUCCESS)
        except Exception as e:
            self.message_user(request, f"❌ Error eliminando estudiante: {str(e)}", level=messages.ERROR)

    def delete_queryset(self, request, queryset):
        """Eliminar múltiples estudiantes con manejo seguro"""
        total = queryset.count()
        errores = 0
        for estudiante in queryset:
            try:
                self.delete_model(request, estudiante)
            except Exception:
                errores += 1
        if errores:
            self.message_user(request, f"⚠️ {total - errores} eliminados, {errores} con error", level=messages.WARNING)

    def eliminar_estudiantes_seguro(self, request, queryset):
        """Eliminar estudiantes seleccionados con limpieza de datos relacionados"""
        total = queryset.count()
        errores = 0
        for estudiante in queryset:
            try:
                self.delete_model(request, estudiante)
            except Exception as e:
                errores += 1
        if errores == 0:
            self.message_user(request, f"✅ {total} estudiante(s) eliminado(s) correctamente", level=messages.SUCCESS)
        else:
            self.message_user(request, f"⚠️ {total - errores} eliminados, {errores} con error", level=messages.WARNING)
    eliminar_estudiantes_seguro.short_description = "🗑️ Eliminar estudiantes (seguro)"

    @admin.action(description='🏢 Escoger cliente (asignación masiva)')
    def asignar_cliente_masivo(self, request, queryset):
        """Asigna el mismo Cliente (organización) a todos los estudiantes seleccionados."""
        if 'aplicar' in request.POST:
            cliente_id = (request.POST.get('cliente_id') or '').strip()
            if not cliente_id:
                self.message_user(request, "Seleccione un cliente.", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')

            cliente = Cliente.objects.filter(id=cliente_id, activo=True).first()
            if not cliente:
                self.message_user(request, "Cliente no válido o inactivo.", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')

            ids = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
            if not ids:
                self.message_user(request, "No se recibieron estudiantes seleccionados.", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')

            qs = Estudiante.objects.filter(id__in=ids)
            n = qs.update(cliente=cliente)
            self.message_user(
                request,
                f"Se asignó el cliente «{cliente.nombre}» a {n} estudiante(s).",
                level=messages.SUCCESS,
            )
            return redirect('admin:core_estudiante_changelist')

        clientes = Cliente.objects.filter(activo=True).order_by('nombre')
        return render(
            request,
            'admin/asignar_cliente_estudiantes.html',
            {
                'estudiantes': queryset,
                'total_estudiantes': queryset.count(),
                'clientes': clientes,
                'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
                'title': 'Escoger cliente — asignación masiva',
            },
        )

    def asignar_a_grupo_accion(self, request, queryset):
        """Asigna múltiples estudiantes a un grupo (existente o nuevo)"""
        if request.POST.get('aplicar') or 'apply' in request.POST:
            # Verificar si es grupo existente o nuevo
            grupo_tipo = request.POST.get('grupo_tipo', 'existente')
            
            if grupo_tipo == 'nuevo':
                # Crear nuevo grupo
                nuevo_nombre = request.POST.get('nuevo_nombre', '').strip()
                nuevo_emoji = request.POST.get('nuevo_emoji', '👥')
                nuevo_descripcion = request.POST.get('nuevo_descripcion', '').strip()
                
                if not nuevo_nombre:
                    self.message_user(request, "⚠️ Debes ingresar un nombre para el grupo", level=messages.WARNING)
                    return
                
                # Obtener el cliente del primer estudiante o None
                primer_estudiante = queryset.first()
                cliente = primer_estudiante.cliente if primer_estudiante else None
                
                # Crear el grupo
                grupo = GrupoEstudiantes.objects.create(
                    nombre=nuevo_nombre,
                    emoji=nuevo_emoji,
                    descripcion=nuevo_descripcion,
                    cliente=cliente,
                    creado_por=request.user,
                    activo=True
                )
                
                # Asignar estudiantes al nuevo grupo
                contador = 0
                for estudiante in queryset:
                    grupo.estudiantes.add(estudiante)
                    contador += 1
                
                self.message_user(
                    request,
                    f"✅ Grupo '{grupo.emoji} {grupo.nombre}' creado y {contador} estudiante(s) asignado(s)",
                    level=messages.SUCCESS
                )
                return redirect('admin:core_estudiante_changelist')
            
            else:
                # Usar grupo existente
                grupo_id = request.POST.get('grupo')
                if not grupo_id:
                    self.message_user(request, "⚠️ Debes seleccionar un grupo", level=messages.WARNING)
                    return
                
                try:
                    grupo = GrupoEstudiantes.objects.get(id=grupo_id)
                    contador = 0
                    for estudiante in queryset:
                        grupo.estudiantes.add(estudiante)
                        contador += 1
                    
                    self.message_user(
                        request,
                        f"✅ {contador} estudiante(s) asignado(s) al grupo '{grupo.nombre}'",
                        level=messages.SUCCESS
                    )
                    return redirect('admin:core_estudiante_changelist')
                except GrupoEstudiantes.DoesNotExist:
                    self.message_user(request, "❌ El grupo seleccionado no existe", level=messages.ERROR)
                    return redirect('admin:core_estudiante_changelist')
        
        # Mostrar formulario de selección de grupo
        grupos = GrupoEstudiantes.objects.all().order_by('nombre')
        context = {
            'title': f'Asignar {queryset.count()} estudiantes a un grupo',
            'queryset': queryset,
            'grupos': grupos,
            'opts': self.model._meta,
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, 'admin/asignar_grupo.html', context)
    asignar_a_grupo_accion.short_description = "👥 Asignar estudiantes a un grupo"


# @admin.register(Plantilla)  # Registro duplicado - movido arriba
class PlantillaAdminDuplicado(admin.ModelAdmin):
    """Gestión de plantillas de mensajes con Twilio Content Templates"""
    list_display = ('nombre_interno', 'categoria_emoji', 'temas_badge', 'twilio_status_badge', 'vista_previa', 'activa', 'veces_usada', 'fecha_modificacion')
    list_filter = ('categoria', 'temas', 'activa', 'aprobada_twilio', 'fecha_creacion')
    search_fields = ('nombre_interno', 'cuerpo_mensaje', 'twilio_template_sid', 'twilio_template_nombre')
    actions = ['enviar_plantilla_directa', 'duplicar_plantilla', 'activar_plantillas', 'desactivar_plantillas']
    readonly_fields = ('veces_usada', 'fecha_creacion', 'fecha_modificacion', 'preview_personalizado')
    filter_horizontal = ('temas',)
    
    fieldsets = (
        ('Información de la Plantilla', {
            'fields': ('nombre_interno', 'categoria', 'temas', 'activa'),
            'description': 'Dale un nombre descriptivo a tu plantilla, categoría y temas relacionados'
        }),
        ('Contenido del Mensaje', {
            'fields': ('cuerpo_mensaje',),
            'description': mark_safe('<strong>Variables disponibles:</strong> {nombre} {telefono} {curso}<br>'
                          '<strong>Ejemplo:</strong> "Hola {nombre}, te damos la bienvenida al curso {curso}"')
        }),
        ('Configuración Twilio Content Templates', {
            'fields': ('twilio_template_sid', 'twilio_template_nombre', 'aprobada_twilio'),
            'description': mark_safe('<div style="background:#e3f2fd;padding:15px;border-radius:4px;margin-bottom:15px;">'
                          '<strong>INSTRUCCIONES PARA USAR TWILIO:</strong><br><br>'
                          '1. Ve a <a href="https://console.twilio.com/us1/develop/sms/content-editor" target="_blank">Twilio Content Editor</a><br>'
                          '2. Crea una nueva plantilla con este mensaje<br>'
                          '3. Espera la aprobación de Twilio (1-2 días hábiles)<br>'
                          '4. Copia el <strong>Content SID</strong> (ej: HX1234...)<br>'
                          '5. Pégalo en el campo "Twilio Content SID" arriba<br>'
                          '6. Marca "Aprobada en Twilio" cuando esté lista<br><br>'
                          '<strong>Una vez aprobada, podrás usarla en campañas</strong></div>')
        }),
        ('Estadísticas y Vista Previa', {
            'fields': ('preview_personalizado', 'veces_usada', 'fecha_creacion', 'fecha_modificacion'),
            'classes': ('collapse',),
        }),
    )
    
    def temas_badge(self, obj):
        """Muestra temas asociados con badges"""
        temas = obj.temas.all()
        if not temas:
            return format_html('<span style="color:#999;">Sin temas</span>')
        
        badges = []
        for tema in temas:
            badges.append(f'<span style="background:#e3f2fd;color:#1976d2;padding:3px 10px;border-radius:12px;margin:2px;display:inline-block;">{tema}</span>')
        return format_html(''.join(badges))
    temas_badge.short_description = "Temas"
    
    def twilio_status_badge(self, obj):
        """Muestra estado de aprobación en Twilio"""
        if not obj.twilio_template_sid:
            return format_html('<span style="background:#f5f5f5;color:#666;padding:4px 10px;border-radius:12px;font-size:11px;">📝 Sin SID</span>')
        
        if obj.aprobada_twilio:
            return format_html(
                '<span style="background:#e8f5e9;color:#2e7d32;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:500;">✅ Aprobada</span>'
            )
        else:
            return format_html(
                '<span style="background:#fff3e0;color:#f57c00;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:500;">⏳ Pendiente</span>'
            )
    twilio_status_badge.short_description = "🔵 Estado Twilio"
    
    def categoria_emoji(self, obj):
        """Muestra categoría con emoji"""
        return obj.get_categoria_display()
    categoria_emoji.short_description = "📂 Categoría"
    
    def vista_previa(self, obj):
        """Muestra preview del mensaje"""
        preview = obj.vista_previa()
        return format_html('<span style="color:#666;font-style:italic;">{}</span>', preview)
    vista_previa.short_description = "📄 Vista Previa"
    
    def preview_personalizado(self, obj):
        """Muestra cómo se vería el mensaje personalizado"""
        ejemplo = obj.cuerpo_mensaje.replace('{nombre}', 'Juan Pérez')
        ejemplo = ejemplo.replace('{telefono}', '+573001234567')
        ejemplo = ejemplo.replace('{curso}', 'Cultivo de Aguacate Hass')
        return format_html(
            '<div style="background:#f5f5f5;padding:15px;border-left:4px solid #4CAF50;border-radius:4px;">'
            '<strong>📱 Vista Previa Personalizada:</strong><br><br>{}</div>',
            ejemplo
        )
    preview_personalizado.short_description = "Vista Previa con Variables"
    
    @admin.action(description='📄 Duplicar plantilla(s) seleccionada(s)')
    def duplicar_plantilla(self, request, queryset):
        """Duplica plantillas seleccionadas"""
        duplicadas = 0
        for plantilla in queryset:
            plantilla.pk = None
            plantilla.nombre_interno = f"{plantilla.nombre_interno} (Copia)"
            plantilla.veces_usada = 0
            plantilla.save()
            duplicadas += 1
        self.message_user(request, f"✅ {duplicadas} plantilla(s) duplicada(s)", level=messages.SUCCESS)
    
    @admin.action(description='✅ Activar plantilla(s) seleccionada(s)')
    def activar_plantillas(self, request, queryset):
        """Activa plantillas seleccionadas"""
        actualizadas = queryset.update(activa=True)
        self.message_user(request, f"✅ {actualizadas} plantilla(s) activada(s)", level=messages.SUCCESS)
    
    @admin.action(description='❌ Desactivar plantilla(s) seleccionada(s)')
    def desactivar_plantillas(self, request, queryset):
        """Desactiva plantillas seleccionadas"""
        actualizadas = queryset.update(activa=False)
        self.message_user(request, f"⚠️ {actualizadas} plantilla(s) desactivada(s)", level=messages.WARNING)
    
    @admin.action(description='📱 Enviar a Meta para aprobación')
    def enviar_a_meta_accion(self, request, queryset):
        """Envía plantillas a Meta WhatsApp Business para aprobación"""
        from .meta_templates import enviar_plantilla_a_meta
        
        enviadas = 0
        errores = 0
        
        for plantilla in queryset:
            try:
                # Verificar si ya fue enviada
                if plantilla.enviada_a_meta and plantilla.meta_template_status == 'APPROVED':
                    self.message_user(
                        request,
                        f"⚠️ '{plantilla.nombre_interno}' ya está aprobada en Meta",
                        level=messages.WARNING
                    )
                    continue
                
                # Enviar a Meta
                resultado = enviar_plantilla_a_meta(
                    nombre_plantilla=plantilla.nombre_interno,
                    contenido=plantilla.cuerpo_mensaje,
                    categoria=plantilla.categoria if plantilla.categoria in ['MARKETING', 'UTILITY'] else 'MARKETING',
                    idioma='es'
                )
                
                if resultado['success']:
                    # Actualizar plantilla con información de Meta
                    plantilla.enviada_a_meta = True
                    plantilla.meta_template_id = resultado['template_id']
                    plantilla.meta_template_status = resultado['status']
                    plantilla.meta_template_name = resultado['nombre_meta']
                    plantilla.save()
                    
                    enviadas += 1
                    self.message_user(
                        request,
                        f"✅ '{plantilla.nombre_interno}' enviada a Meta. ID: {resultado['template_id']}",
                        level=messages.SUCCESS
                    )
                else:
                    errores += 1
                    self.message_user(
                        request,
                        f"❌ Error con '{plantilla.nombre_interno}': {resultado['message']}",
                        level=messages.ERROR
                    )
            
            except Exception as e:
                errores += 1
                logger.error(f"Error enviando plantilla a Meta: {str(e)}")
                self.message_user(
                    request,
                    f"❌ Excepción con '{plantilla.nombre_interno}': {str(e)}",
                    level=messages.ERROR
                )
        
        # Resumen final
        if enviadas > 0:
            self.message_user(
                request,
                f"🎉 {enviadas} plantilla(s) enviada(s) a Meta para revisión",
                level=messages.SUCCESS
            )
        if errores > 0:
            self.message_user(
                request,
                f"⚠️ {errores} plantilla(s) con errores. Verifica las credenciales de Meta.",
                level=messages.WARNING
            )
    
    @admin.action(description='📤 Enviar plantilla a estudiantes')
    def enviar_plantilla_directa(self, request, queryset):
        """Permite enviar una plantilla directamente a estudiantes seleccionados"""
        
        if queryset.count() > 1:
            self.message_user(request, "⚠️ Solo puedes enviar una plantilla a la vez", level=messages.WARNING)
            return
        
        plantilla = queryset.first()
        
        # Si es POST con confirmación
        if 'aplicar' in request.POST:
            # Obtener estudiantes seleccionados
            estudiantes_ids = request.POST.getlist('estudiantes_seleccionados')
            if not estudiantes_ids:
                self.message_user(request, "⚠️ Debes seleccionar al menos un estudiante", level=messages.ERROR)
            else:
                enviados = 0
                errores = 0
                
                for est_id in estudiantes_ids:
                    try:
                        estudiante = Estudiante.objects.get(id=est_id)
                        
                        # Personalizar mensaje con nombre
                        mensaje = plantilla.cuerpo_mensaje.replace('{nombre}', estudiante.nombre)
                        mensaje = mensaje.replace('{estudiante}', estudiante.nombre)
                        
                        telefono = estudiante.telefono
                        if not telefono.startswith('whatsapp:'):
                            telefono = f'whatsapp:{telefono}'
                        
                        resultado = enviar_whatsapp_twilio(
                            telefono=telefono,
                            texto=mensaje,
                            mensaje_id_referencia=None
                        )
                        
                        if resultado.get('success'):
                            enviados += 1
                            # Registrar en WhatsappLog
                            WhatsappLog.objects.create(
                                telefono=estudiante.telefono,
                                mensaje=mensaje,
                                mensaje_id=resultado.get('mensaje_id'),
                                tipo='SENT',
                                estado='SENT'
                            )
                            logger.info(f"✅ Plantilla '{plantilla.nombre_interno}' enviada a {estudiante.nombre}")
                        else:
                            errores += 1
                            logger.error(f"❌ Error al enviar plantilla a {estudiante.telefono}")
                    
                    except Exception as e:
                        errores += 1
                        logger.error(f"❌ Excepción al enviar: {str(e)}")
                
                if enviados > 0:
                    self.message_user(request, f"✅ Plantilla enviada a {enviados} estudiante(s)", level=messages.SUCCESS)
                if errores > 0:
                    self.message_user(request, f"⚠️ Hubo {errores} error(es)", level=messages.WARNING)
                
                # Redirigir a la lista de plantillas después de enviar
                from django.urls import reverse
                return redirect(reverse('admin:core_plantilla_changelist'))
        
        # Mostrar formulario de selección de estudiantes
        estudiantes = Estudiante.objects.filter(activo=True).order_by('nombre')
        
        return render(request, 'admin/enviar_plantilla_directa.html', {
            'plantilla': plantilla,
            'estudiantes': estudiantes,
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        })


class EnvioProgramadoForm(forms.ModelForm):
    media_file_upload = forms.FileField(
        label='Subir archivo desde PC',
        required=False,
        help_text='Opcional. Guarda el archivo en storage/S3 y completa URL del Archivo.',
    )

    class Meta:
        model = EnvioProgramado
        exclude = (
            'fecha_envio_real',
            'total_destinatarios',
            'total_enviados',
            'total_fallidos',
            'error',
            'fecha_creacion',
            'creado_por',
        )

    def save(self, commit=True):
        instance = super().save(commit=False)
        uploaded_file = self.cleaned_data.get('media_file_upload')
        if uploaded_file:
            instance.media_url = guardar_upload_admin_media(
                uploaded_file,
                carpeta='envios_programados',
                prefix='media',
            )
            instance.incluir_media = True
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class EnvioProgramadoInline(admin.TabularInline):
    """Envíos programados dentro de una campaña"""
    model = EnvioProgramado
    form = EnvioProgramadoForm
    extra = 0
    fields = (
        'nombre', 'tipo', 'fecha_programada', 'estado', 'mensaje',
        'incluir_media', 'media_url', 'media_file_upload', 'fecha_envio_real',
    )
    readonly_fields = ('fecha_envio_real',)
    show_change_link = True


@admin.register(Campana)
class CampanaAdmin(admin.ModelAdmin):
    """Gestión de campañas masivas"""
    list_display = ('nombre', 'cliente_nombre', 'tipo_audiencia_display', 'categoria_badge', 'plantilla_estado', 'estado_visual', 'conteo_destinatarios', 'programada_display', 'fecha_creacion')
    list_filter = ('ejecutada', 'cliente', 'categoria', 'tipo_audiencia', 'fecha_creacion', 'plantilla__aprobada_twilio')
    search_fields = ('nombre', 'cliente__nombre')
    filter_horizontal = ('destinatarios',)
    actions = ['enviar_campana_accion']
    inlines = [EnvioProgramadoInline]
    
    fieldsets = (
        ('📝 Información Básica', {
            'fields': ('nombre', 'cliente'),
            'description': 'Configura el nombre y cliente de la campaña'
        }),
        ('📨 Template de Twilio (Content SID)', {
            'fields': ('template_twilio_id',),
            'description': mark_safe('''<div style="background:#e8f5e9;padding:15px;border-radius:8px;border-left:4px solid #4CAF50;">
                <strong>✅ Método recomendado:</strong> Pega el Content SID de Twilio directamente (ej: HX1234...).<br>
                Crea tu template en <a href="https://console.twilio.com/us1/develop/sms/content-editor" target="_blank" style="color:#2196F3;">Twilio Content Editor</a>.
            </div>''')
        }),
        ('🚀 Campaña de Inicio de Curso', {
            'fields': ('es_campana_curso', 'curso_destino'),
            'description': mark_safe('''<div style="background:#fff3e0;padding:15px;border-radius:8px;border-left:4px solid #ff9800;">
                <strong>📚 Si esta es una campaña para iniciar un curso:</strong><br>
                1. Marca "Es campaña de inicio de curso"<br>
                2. Selecciona el curso destino<br>
                3. Al enviar, los estudiantes entrarán al flujo: Habeas Data → Verificación → Curso
            </div>''')
        }),
        ('📄 Plantilla Django (Alternativa)', {
            'fields': ('plantilla',),
            'classes': ('collapse',),
            'description': 'Solo si NO usas Content SID directo. Requiere plantilla creada en eki.'
        }),
        ('👥 Audiencia', {
            'fields': ('tipo_audiencia', 'grupo', 'grupo_estudiantes_preview', 'destinatarios'),
            'description': mark_safe('✨ <strong>Individual:</strong> Selecciona estudiantes específicos | <strong>Grupo:</strong> Envía a todo un grupo')
        }),
        ('⏰ Programación (Opcional)', {
            'fields': ('fecha_programada',),
            'description': '📅 Si seleccionas una fecha, el envío se realizará automáticamente en ese momento',
            'classes': ('collapse',)
        }),
        ('📂 Importar desde Excel (Opcional)', {
            'fields': ('archivo_excel',),
            'description': 'Sube un Excel con columnas A (Nombre) y B (Teléfono)',
            'classes': ('collapse',)
        }),
        ('📊 Estadísticas', {
            'fields': ('total_enviados', 'respuestas_si', 'respuestas_no'),
            'classes': ('collapse',),
            'description': 'Estadísticas de envío y respuestas'
        }),
    )
    
    def cliente_nombre(self, obj):
        """Muestra el cliente de la campaña"""
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;font-style:italic;">Sin cliente</span>')
    cliente_nombre.short_description = "🏢 Cliente"

    def get_readonly_fields(self, request, obj=None):
        return tuple(super().get_readonly_fields(request, obj)) + ('grupo_estudiantes_preview',)

    def grupo_estudiantes_preview(self, obj):
        grupos = GrupoEstudiantes.objects.prefetch_related('estudiantes').order_by('nombre')
        map_grupos = {}
        for grupo in grupos:
            estudiantes = [
                {
                    "nombre": (e.nombre or '').strip() or f"ID {e.id}",
                    "telefono": e.telefono or "",
                }
                for e in grupo.estudiantes.all().order_by('nombre')[:300]
            ]
            map_grupos[str(grupo.id)] = {
                "nombre": f"{grupo.emoji or '👥'} {grupo.nombre}",
                "total": grupo.estudiantes.count(),
                "estudiantes": estudiantes,
            }

        data_json = json.dumps(map_grupos, ensure_ascii=False)
        return mark_safe(f"""
            <div id="grupo-preview-box" style="border:1px solid #e0e0e0;background:#fafafa;border-radius:8px;padding:10px;max-width:860px;">
                <strong>👀 Estudiantes del grupo seleccionado</strong>
                <div id="grupo-preview-content" style="margin-top:8px;color:#666;">Seleccione un grupo para ver sus destinatarios.</div>
            </div>
            <script>
            (function() {{
                const data = {data_json};
                const selectGrupo = document.getElementById("id_grupo");
                const selectTipo = document.getElementById("id_tipo_audiencia");
                const content = document.getElementById("grupo-preview-content");
                if (!selectGrupo || !content) return;
                function renderGrupo() {{
                    const gid = (selectGrupo.value || "").toString();
                    const tipo = (selectTipo && selectTipo.value) ? selectTipo.value : "";
                    if (tipo && tipo !== "grupo") {{
                        content.innerHTML = "<span style='color:#888;'>La campaña está en modo individual.</span>";
                        return;
                    }}
                    if (!gid || !data[gid]) {{
                        content.innerHTML = "<span style='color:#888;'>Seleccione un grupo para ver sus destinatarios.</span>";
                        return;
                    }}
                    const g = data[gid];
                    const rows = (g.estudiantes || []).slice(0, 80).map(e => `<li>${{e.nombre}} <span style="color:#999;">(${{e.telefono || "sin teléfono"}})</span></li>`).join("");
                    const extra = g.total > 80 ? `<div style="margin-top:6px;color:#999;">... y ${{g.total - 80}} estudiante(s) más.</div>` : "";
                    content.innerHTML = `
                        <div><strong>${{g.nombre}}</strong> — <strong>${{g.total}}</strong> estudiante(s)</div>
                        <ul style="margin:8px 0 0 18px;max-height:220px;overflow:auto;">${{rows || "<li>Sin estudiantes</li>"}}</ul>
                        ${{extra}}
                    `;
                }}
                selectGrupo.addEventListener("change", renderGrupo);
                if (selectTipo) selectTipo.addEventListener("change", renderGrupo);
                renderGrupo();
            }})();
            </script>
        """)
    grupo_estudiantes_preview.short_description = "Vista previa de destinatarios del grupo"
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Filtrar plantillas según la categoría seleccionada en la campaña"""
        if db_field.name == "plantilla":
            # Intentar obtener la categoría desde POST (cuando está guardando)
            categoria = request.POST.get('categoria') or request.GET.get('categoria')
            
            if categoria and categoria != 'todas':
                kwargs["queryset"] = Plantilla.objects.filter(categoria=categoria, activa=True)
            else:
                # Si no hay categoría o es 'todas', mostrar todas las plantillas activas
                kwargs["queryset"] = Plantilla.objects.filter(activa=True)
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def categoria_badge(self, obj):
        """Muestra la categoría con badge"""
        if obj.categoria and obj.categoria != 'todas':
            return format_html(
                '<span style="background:#e3f2fd;color:#1976d2;padding:6px 12px;border-radius:12px;font-weight:bold;">{}</span>',
                obj.get_categoria_display()
            )
        return format_html('<span style="color:#999;">Todas las categorías</span>')
    categoria_badge.short_description = "🏷️ Categoría"
    
    def tipo_audiencia_display(self, obj):
        """Muestra si es por grupo o individual"""
        if hasattr(obj, 'tipo_audiencia') and obj.tipo_audiencia == 'grupo' and obj.grupo:
            return format_html(
                '<span style="background:#fff3e0;color:#e65100;padding:4px 10px;border-radius:12px;font-size:11px;">👥 Grupo: {}</span>',
                obj.grupo.nombre
            )
        return format_html('<span style="color:#666;">👤 Individual</span>')
    tipo_audiencia_display.short_description = "Tipo"
    
    @admin.action(description='🚀 Ejecutar Campaña (Enviar Mensajes)')
    def enviar_campana_accion(self, request, queryset):
        """Encola el envío masivo en segundo plano (evita 504 con +90 destinatarios)."""
        from .services import encolar_ejecutar_campana
        
        for campana in queryset:
            # Permitir re-envío de campañas ya ejecutadas
            if campana.ejecutada:
                self.message_user(
                    request, 
                    f"ℹ️ '{campana.nombre}' ya fue enviada antes. Re-enviando...", 
                    level=messages.INFO
                )
            
            # VALIDACIÓN: necesita template_twilio_id O plantilla aprobada
            if campana.template_twilio_id:
                # Envío directo con Content SID
                pass  # Válido
            elif campana.plantilla and campana.plantilla.content_sid and campana.plantilla.aprobada_twilio:
                # Envío con plantilla Django aprobada
                pass  # Válido
            else:
                self.message_user(
                    request,
                    f"🚨 '{campana.nombre}': Necesita Content SID de Twilio o una plantilla aprobada. "
                    f"Configura un Content Template en Twilio primero.",
                    level=messages.ERROR
                )
                continue
            
            # Validar que tenga destinatarios (o grupo con al menos un estudiante activo)
            destinatarios_count = campana.destinatarios.filter(activo=True).count()
            if hasattr(campana, 'tipo_audiencia') and campana.tipo_audiencia == 'grupo':
                if not campana.grupo:
                    self.message_user(
                        request,
                        f"⚠️ '{campana.nombre}' no tiene un grupo seleccionado.",
                        level=messages.WARNING
                    )
                    continue
                destinatarios_count = campana.grupo.estudiantes.filter(activo=True).count()

            if destinatarios_count == 0:
                self.message_user(
                    request,
                    f"⚠️ '{campana.nombre}' no tiene destinatarios: no se envía a nadie.",
                    level=messages.WARNING
                )
                continue
            
            try:
                modo = encolar_ejecutar_campana(campana.id)
                detalle_modo = 'Celery' if modo == 'celery' else 'proceso en background'
                self.message_user(
                    request,
                    f"⏳ '{campana.nombre}' encolada ({detalle_modo}): "
                    f"se enviará a {destinatarios_count} destinatarios en segundo plano. "
                    f"Actualiza el listado en unos minutos para ver el resultado.",
                    level=messages.SUCCESS,
                )
            except Exception as e:
                self.message_user(
                    request,
                    f"❌ Error encolando '{campana.nombre}': {str(e)}",
                    level=messages.ERROR
                )
    
    enviar_campana_accion.short_description = "📤 Ejecutar campañas seleccionadas (envío real por WhatsApp)"
    
    def estado_visual(self, obj):
        if obj.ejecutada:
            return format_html('<span style="color: green;">✅ Ejecutada</span>')
        return format_html('<span style="color: orange;">⏳ Pendiente</span>')
    estado_visual.short_description = "Estado"
    
    def conteo_destinatarios(self, obj):
        """Muestra cantidad de destinatarios según tipo de audiencia"""
        if hasattr(obj, 'tipo_audiencia') and obj.tipo_audiencia == 'grupo' and obj.grupo:
            count = obj.grupo.estudiantes.count()
            return format_html(
                '<span style="background:#fff3e0;padding:3px 8px;border-radius:8px;font-size:11px;">👥 {} estudiantes</span>',
                count
            )
        count = obj.destinatarios.count()
        return format_html(
            '<span style="background:#e3f2fd;padding:3px 8px;border-radius:8px;font-size:11px;">👤 {} individual{}</span>',
            count, 'es' if count != 1 else ''
        )
    conteo_destinatarios.short_description = "Destinatarios"
    
    def programada_display(self, obj):
        """Muestra si está programada"""
        if obj.fecha_programada:
            return format_html(
                '<span style="background:#e8f5e9;color:#2e7d32;padding:3px 8px;border-radius:8px;font-size:11px;">📅 {}</span>',
                obj.fecha_programada.strftime('%d/%m %H:%M')
            )
        return format_html('<span style="color:#999;">-</span>')
    programada_display.short_description = "Programada"

    def plantilla_estado(self, obj):
        """Muestra el estado del Content Template de la plantilla"""
        if obj.plantilla:
            if obj.plantilla.twilio_template_sid and obj.plantilla.aprobada_twilio:
                return format_html('<span style="background:#4caf50;color:white;padding:2px 6px;border-radius:8px;font-size:10px;">✅ TWILIO</span>')
            elif obj.plantilla.twilio_template_sid and not obj.plantilla.aprobada_twilio:
                return format_html('<span style="background:#ff9800;color:white;padding:2px 6px;border-radius:8px;font-size:10px;">⏳ PENDIENTE</span>')
            else:
                return format_html('<span style="background:#2196f3;color:white;padding:2px 6px;border-radius:8px;font-size:10px;">📱 DIRECTO</span>')
        return format_html('<span style="color:#999;font-size:10px;">Sin plantilla</span>')
    plantilla_estado.short_description = "Template"


@admin.register(MensajePush)
class MensajePushAdmin(admin.ModelAdmin):
    """Recordatorios WhatsApp que no reinician el curso (responder *listo*)."""
    list_display = ('nombre', 'tipo', 'cliente', 'curso', 'activo', 'fecha_creacion')
    list_filter = ('tipo', 'activo', 'cliente')
    search_fields = ('nombre', 'cuerpo_fallback', 'twilio_content_sid')
    actions = ['enviar_push_seleccionados']
    fieldsets = (
        (None, {'fields': ('nombre', 'tipo', 'activo')}),
        ('Audiencia', {'fields': ('cliente', 'curso')}),
        ('Contenido Twilio / texto', {
            'fields': ('twilio_content_sid', 'plantilla', 'cuerpo_fallback', 'incluir_boton_continuar'),
            'description': 'Use Content SID (HX…) o plantilla Django. Variables: {nombre}, {curso}.',
        }),
    )

    @admin.action(description='📲 Enviar push a audiencia del mensaje')
    def enviar_push_seleccionados(self, request, queryset):
        from core.mensajes_push import enviar_mensaje_push_masivo
        for mp in queryset.filter(activo=True):
            r = enviar_mensaje_push_masivo(mp)
            self.message_user(
                request,
                f'{mp.nombre}: {r["enviados"]} enviados, {r["errores"]} errores.',
            )


@admin.register(EnvioMensajePush)
class EnvioMensajePushAdmin(admin.ModelAdmin):
    list_display = ('mensaje_push', 'estudiante', 'telefono', 'exito', 'fecha')
    list_filter = ('exito', 'mensaje_push')
    readonly_fields = ('mensaje_push', 'estudiante', 'telefono', 'exito', 'detalle', 'fecha')


@admin.register(EnvioLog)
class EnvioLogAdmin(admin.ModelAdmin):
    """Historial de envíos de campañas"""
    list_display = ('campana_nombre', 'estudiante_info', 'estado_badge', 'fecha_envio')
    list_filter = ('estado', 'campana', 'fecha_envio')
    search_fields = ('estudiante__nombre', 'estudiante__cedula', 'estudiante__telefono', 'campana__nombre')
    readonly_fields = ('campana', 'estudiante', 'estado', 'respuesta_api', 'fecha_envio')
    date_hierarchy = 'fecha_envio'
    list_per_page = 50
    
    fieldsets = (
        ('📤 Información del Envío', {
            'fields': ('campana', 'estudiante', 'fecha_envio')
        }),
        ('📊 Estado', {
            'fields': ('estado', 'respuesta_api')
        }),
    )
    
    def campana_nombre(self, obj):
        return format_html(
            '<strong>{}</strong><br><small style="color:#666;">{}</small>',
            obj.campana.nombre,
            obj.campana.plantilla.nombre_interno if obj.campana.plantilla else ''
        )
    campana_nombre.short_description = "Campaña"
    
    def estudiante_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small style="color:#666;">📱 +{}</small>',
            obj.estudiante.nombre,
            obj.estudiante.telefono
        )
    estudiante_info.short_description = "Estudiante"
    
    def estado_badge(self, obj):
        colores = {
            'ENVIADO': ('#e8f5e9', '#2e7d32'),
            'FALLIDO': ('#ffebee', '#c62828'),
            'PENDIENTE': ('#fff3e0', '#e65100'),
        }
        bg, color = colores.get(obj.estado, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;font-weight:bold;">{}</span>',
            bg, color, obj.estado
        )
    estado_badge.short_description = "Estado"


@admin.register(WhatsappLog)
class WhatsappLogAdmin(admin.ModelAdmin):
    """Registro de todas las conversaciones del chatbot"""
    list_display = ('fecha', 'estudiante_nombre', 'telefono_corto', 'tipo_badge', 'mensaje_preview', 'estado_badge', 'estado_visual', 'actividad_badge')
    list_filter = ('tipo', 'estado', 'fecha', 'estudiante__activo', 'agente_usado')
    search_fields = ('telefono', 'mensaje', 'mensaje_id', 'estudiante__nombre')  # ✅ Búsqueda por nombre
    date_hierarchy = 'fecha'
    list_per_page = 100
    ordering = ('-fecha',)
    readonly_fields = ('fecha', 'mensaje_id', 'estudiante')
    actions = ['exportar_conversaciones_excel', 'exportar_conversaciones_csv', 'marcar_como_procesado']
    autocomplete_fields = ['estudiante']  # ✅ Autocompletar estudiante
    
    fieldsets = (
        ('Información del Mensaje', {
            'fields': ('telefono', 'estudiante', 'tipo', 'mensaje', 'estado')
        }),
        ('Metadatos', {
            'fields': ('mensaje_id', 'fecha'),
            'classes': ('collapse',)
        }),
    )
    
    def estudiante_nombre(self, obj):
        """Muestra nombre del estudiante si está asignado"""
        if obj.estudiante:
            return obj.estudiante.nombre
        return format_html('<span style="color:#999;font-style:italic;">Sin asignar</span>')
    estudiante_nombre.short_description = "👤 Estudiante"
    estudiante_nombre.admin_order_field = 'estudiante__nombre'
    
    def telefono_corto(self, obj):
        """Muestra solo los últimos 4 dígitos"""
        return f"...{obj.telefono[-4:]}"
    telefono_corto.short_description = "📱"
    
    def tipo_badge(self, obj):
        """Badge visual para tipo de mensaje"""
        if obj.tipo == 'INCOMING':
            return format_html(
                '<span style="background:#4caf50;color:white;padding:3px 8px;border-radius:12px;font-size:11px;">⬇️ RECIBIDO</span>'
            )
        return format_html(
            '<span style="background:#2196f3;color:white;padding:3px 8px;border-radius:12px;font-size:11px;">⬆️ ENVIADO</span>'
        )
    tipo_badge.short_description = "Tipo"
    
    def mensaje_preview(self, obj):
        """Muestra preview del mensaje"""
        texto = obj.mensaje[:60] + "..." if len(obj.mensaje) > 60 else obj.mensaje
        return texto
    mensaje_preview.short_description = "💬 Mensaje"
    
    def estado_badge(self, obj):
        """Badge visual para estado"""
        colores = {
            'RECIBIDO': '#4caf50',
            'SENT': '#2196f3',
            'PENDING': '#ff9800',
            'ERROR': '#f44336'
        }
        color = colores.get(obj.estado, '#999')
        return format_html(
            '<span style="background:{};color:white;padding:3px 8px;border-radius:12px;font-size:11px;">{}</span>',
            color, obj.estado
        )
    estado_badge.short_description = "Estado"
    
    def estado_visual(self, obj):
        """Indicador visual de éxito/error"""
        if obj.estado == 'SENT' or obj.estado == 'RECIBIDO':
            return format_html('<span style="font-size:18px;color:#4caf50;">✅</span>')
        elif obj.estado == 'ERROR':
            return format_html('<span style="font-size:18px;color:#f44336;">❌</span>')
        else:
            return format_html('<span style="font-size:18px;color:#ff9800;">⏳</span>')
    estado_visual.short_description = "✓"
    
    def actividad_badge(self, obj):
        """Badge de actividad del mensaje"""
        from django.utils import timezone
        from datetime import timedelta
        
        now = timezone.now()
        delta = now - obj.fecha
        
        if delta.total_seconds() < 86400:  # 24 horas
            return format_html('<span style="color:#f44336;font-weight:bold;">🔴 Nueva</span>')
        elif delta.days < 7:
            return format_html('<span style="color:#4caf50;font-weight:bold;">🟢 Activa</span>')
        elif delta.days < 30:
            return format_html('<span style="color:#ff9800;">🟡 Reciente</span>')
        else:
            return format_html('<span style="color:#999;">⚪ Antigua</span>')
    actividad_badge.short_description = "⏰ Actividad"
    
    def get_queryset(self, request):
        """Ordena por fecha descendente por defecto"""
        qs = super().get_queryset(request)
        return qs.order_by('-fecha')
    
    @admin.action(description='📊 Exportar conversaciones a Excel')
    def exportar_conversaciones_excel(self, request, queryset):
        """Exporta conversaciones seleccionadas a archivo Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conversaciones"
        
        # Encabezados
        headers = ['Fecha', 'Teléfono', 'Tipo', 'Mensaje', 'Estado', 'ID Mensaje']
        ws.append(headers)
        
        # Estilo
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for log in queryset.order_by('fecha'):
            ws.append([
                log.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                log.telefono,
                log.tipo,
                log.mensaje[:500],  # Limitar tamaño
                log.estado,
                log.mensaje_id or 'N/A'
            ])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 35
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'conversaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @admin.action(description='📄 Exportar conversaciones a CSV')
    def exportar_conversaciones_csv(self, request, queryset):
        """Exporta conversaciones seleccionadas a archivo CSV"""
        import csv
        from datetime import datetime
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'conversaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Fecha', 'Teléfono', 'Tipo', 'Mensaje', 'Estado', 'ID Mensaje'])
        
        for log in queryset.order_by('fecha'):
            writer.writerow([
                log.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                log.telefono,
                log.tipo,
                log.mensaje[:500],
                log.estado,
                log.mensaje_id or 'N/A'
            ])
        
        return response
    
    @admin.action(description='✅ Marcar como procesado')
    def marcar_como_procesado(self, request, queryset):
        """Marca mensajes como procesados"""
        actualizado = queryset.filter(estado__in=['RECIBIDO', 'PENDING']).update(estado='SENT')
        self.message_user(request, f'✅ {actualizado} mensaje(s) marcado(s) como procesado', messages.SUCCESS)



# Personalizar el admin site
admin.site.site_header = "eki - Chatbot Agro 🌱"


# ==========================================
# SISTEMA EDUCATIVO - ADMINISTRACIÓN
# ==========================================

class ModuloInline(admin.TabularInline):
    """Módulos dentro del curso"""
    model = Modulo
    extra = 1
    fields = ('numero', 'titulo', 'descripcion', 'duracion_dias')
    ordering = ['numero']


class DocumentoRAGInline(admin.StackedInline):
    """Documentos RAG para la base de conocimiento IA del curso"""
    model = DocumentoRAG
    extra = 0
    verbose_name = '📄 Documento RAG'
    verbose_name_plural = 'DOCUMENTOS RAG — Base de Conocimiento para Agentes IA'
    readonly_fields = ('estado_badge', 'chunks_indexados', 'fecha_subida', 'fecha_indexado')
    fields = ('nombre', 'archivo', 'tipo', 'descripcion', 'estado_badge', 'chunks_indexados', 'fecha_subida', 'fecha_indexado')

    def estado_badge(self, obj):
        if not obj.pk:
            return '-'
        colors = {'pendiente': '#ffc107', 'indexado': '#28a745', 'error': '#dc3545'}
        color = colors.get(obj.estado, '#6c757d')
        return format_html(
            '<span style="background:{};color:white;padding:3px 10px;border-radius:12px;font-size:11px;">{}</span>',
            color, obj.get_estado_display()
        )
    estado_badge.short_description = 'Estado RAG'


class PreguntaAbiertaFinalInline(admin.TabularInline):
    """Preguntas abiertas finales dentro del curso (máximo 3)."""
    model = PreguntaAbiertaFinalCurso
    extra = 1
    max_num = 3
    fields = ('orden', 'pregunta', 'activa')
    ordering = ('orden', 'id')
    verbose_name = '📝 Pregunta Abierta Final'
    verbose_name_plural = 'PREGUNTAS ABIERTAS FINALES (MAX 3)'


@admin.register(Curso)
class CursoAdmin(admin.ModelAdmin):
    """Administración de cursos"""
    change_form_template = 'admin/core/curso/change_form.html'
    list_display = ('nombre', 'cliente_nombre', 'total_modulos_display', 'docs_rag_count', 'duracion_semanas', 'ver_modulos_link', 'activo', 'tiene_formulario_gei', 'usar_agentes_ia', 'orden')
    list_filter = ('activo', 'cliente', 'usar_gamificacion', 'usar_agentes_ia', 'habilitar_pregunta_abierta_final', 'tiene_formulario_gei')
    search_fields = ('nombre', 'descripcion', 'cliente__nombre')
    list_editable = ('orden',)
    inlines = [ModuloInline, DocumentoRAGInline, PreguntaAbiertaFinalInline]
    actions = [
        'ver_todos_modulos', 'indexar_documentos_rag', 'indexar_contenido_modulos',
        'activar_cursos', 'desactivar_cursos', 'copiar_a_otro_cliente', 'copiar_a_analytics_pruebas',
    ]
    # change_list_template = 'admin/curso_changelist.html'  # Eliminado para usar el template estándar de Django
    
    fieldsets = (
        ('Datos del curso', {
            'fields': ('nombre', 'descripcion', 'cliente', 'duracion_semanas', 'activo', 'orden'),
        }),
        ('Ritmo drip y acceso', {
            'fields': (
                'dias_espera_entre_modulos',
                'usar_gamificacion',
                'habilitar_pregunta_abierta_final',
            ),
            'description': mark_safe(
                '<p><strong>0</strong> = avance inmediato con <em>listo</em>. '
                '<strong>&gt; 0</strong> = días de espera entre módulos.</p>'
                '<p>Override por cliente: tabla <em>Ritmo drip por curso</em> en la ficha Cliente.</p>'
            ),
        }),
        ('IA y retos', {
            'fields': (
                'usar_agentes_ia',
                'nombre_agente_tutor',
                'nombre_agente_asistente',
                'preguntas_ejemplo_ia',
            ),
            'description': mark_safe(
                '<p>Retos Darío/Claudia, nombres de agentes (override) y preguntas ejemplo para el tutor.</p>'
                '<p>Si los nombres van vacíos, se usan los del Cliente o los por defecto.</p>'
            ),
        }),
        ('GEI y WhatsApp', {
            'fields': ('tiene_formulario_gei', 'enlace_grupo_whatsapp'),
            'description': mark_safe(
                '<p>Recolección GEI al completar el módulo disparador (Formulario → Tipos de formulario). '
                '<a href="/admin/gei/panel/" target="_blank">Panel GEI</a></p>'
            ),
        }),
    )
    
    def cliente_nombre(self, obj):
        """Muestra si es curso específico de un cliente"""
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;font-style:italic;">General (eki)</span>')
    cliente_nombre.short_description = "🏢 Cliente"
    
    def total_modulos_display(self, obj):
        count = obj.modulos.count()
        return format_html(
            '<span style="background:#e3f2fd;padding:4px 8px;border-radius:4px;">{} módulos</span>',
            count
        )
    total_modulos_display.short_description = "Módulos"
    
    def ver_modulos_link(self, obj):
        """Link para ver todos los módulos del curso"""
        url = f"/admin/core/modulo/?curso__id__exact={obj.id}"
        count = obj.modulos.count()
        
        # Contar archivos multimedia totales
        from django.db.models import Count
        total_archivos = ArchivoModulo.objects.filter(modulo__curso=obj, activo=True).count()
        
        html = f'<a href="{url}" style="color:#2196F3;">📋 {count} módulo(s)</a>'
        if total_archivos > 0:
            html += f' <span style="color:#999;font-size:11px;">• {total_archivos} archivos</span>'
        return format_html(html)
    ver_modulos_link.short_description = "Gestión"

    def docs_rag_count(self, obj):
        """Muestra cantidad de documentos RAG indexados"""
        total = obj.documentos_rag.count()
        indexados = obj.documentos_rag.filter(estado='indexado').count()
        if total == 0:
            return format_html('<span style="color:#999;font-size:11px;">Sin docs</span>')
        color = '#28a745' if indexados == total else '#ffc107'
        return format_html(
            '<span style="background:{};color:white;padding:3px 8px;border-radius:4px;font-size:11px;">'
            '📚 {}/{}</span>',
            color, indexados, total
        )
    docs_rag_count.short_description = "RAG"

    @admin.action(description='🤖 Indexar documentos RAG de cursos seleccionados')
    def indexar_documentos_rag(self, request, queryset):
        """Indexa todos los documentos RAG pendientes de los cursos seleccionados."""
        total_indexados = 0
        errores = 0
        for curso in queryset:
            for doc in curso.documentos_rag.filter(estado__in=['pendiente', 'error']):
                n = doc.indexar()
                if n > 0:
                    total_indexados += 1
                else:
                    errores += 1
        msg = f"✅ {total_indexados} documentos indexados correctamente."
        if errores:
            msg += f" ⚠️ {errores} con errores."
        self.message_user(request, msg)

    @admin.action(description='📝 Indexar contenido de módulos en RAG')
    def indexar_contenido_modulos(self, request, queryset):
        """Indexa el contenido educativo de los módulos en la BD vectorial."""
        from core.rag_manager import rag_manager
        total = 0
        for curso in queryset:
            n = rag_manager.indexar_modulos_curso(curso.id)
            total += n
        self.message_user(request, f"✅ {total} chunks indexados desde contenido de módulos.")

    def save_formset(self, request, form, formset, change):
        """Al guardar DocumentoRAG inline, auto-indexar."""
        instances = formset.save(commit=False)
        for instance in instances:
            if isinstance(instance, DocumentoRAG):
                if not instance.subido_por_id:
                    instance.subido_por = request.user
                instance.save()
                # Auto-indexar documentos nuevos (Celery tras commit; evita 504 en admin)
                if instance.estado == 'pendiente' and instance.archivo:
                    if getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False):
                        instance.indexar()
                    else:
                        _enqueue_indexar_rag_task('DocumentoRAG', instance.pk)
            else:
                instance.save()
        formset.save_m2m()
        # Manejar eliminaciones
        for obj in formset.deleted_objects:
            if isinstance(obj, DocumentoRAG):
                from core.rag_manager import rag_manager
                rag_manager.eliminar_documento(obj.cliente_id, obj.curso_id, obj.nombre)
            obj.delete()

    @admin.action(description='📋 Ver módulos de cursos seleccionados')
    def ver_todos_modulos(self, request, queryset):
        """Redirige a la vista de módulos filtrando por los cursos seleccionados"""
        from django.shortcuts import redirect
        curso_ids = ','.join(str(c.id) for c in queryset)
        return redirect(f'/admin/core/modulo/?curso__id__in={curso_ids}')

    @admin.action(description='✅ Activar cursos seleccionados')
    def activar_cursos(self, request, queryset):
        count = queryset.update(activo=True)
        self.message_user(request, f"✅ {count} curso(s) activado(s)")

    @admin.action(description='❌ Desactivar cursos seleccionados')
    def desactivar_cursos(self, request, queryset):
        count = queryset.update(activo=False)
        self.message_user(request, f"❌ {count} curso(s) desactivado(s)")

    @admin.action(description='📋 Copiar a otro cliente…')
    def copiar_a_otro_cliente(self, request, queryset):
        from django.shortcuts import redirect

        ids = ','.join(str(c.pk) for c in queryset)
        return redirect(f'/admin/copiar-curso/?cursos={ids}')

    @admin.action(description='📋 Copiar a Analytics (Pruebas) — legacy')
    def copiar_a_analytics_pruebas(self, request, queryset):
        from core.copiar_cursos import (
            CLIENTE_ORIGEN_NOMBRE,
            ClienteOrigenNoEncontrado,
            copiar_cursos_a_pruebas,
            obtener_cliente_analytics_origen,
        )

        try:
            origen = obtener_cliente_analytics_origen()
        except ClienteOrigenNoEncontrado as e:
            self.message_user(request, str(e), level='error')
            return
        fuera = queryset.exclude(cliente=origen)
        if fuera.exists():
            self.message_user(
                request,
                f'Solo se copian cursos del cliente «{CLIENTE_ORIGEN_NOMBRE}» '
                f'(id={origen.pk}). {fuera.count()} curso(s) de otro cliente ignorados.',
                level='warning',
            )
        ids = list(queryset.filter(cliente=origen).values_list('pk', flat=True))
        if not ids:
            self.message_user(request, 'No hay cursos de Analytics seleccionados.', level='error')
            return
        result = copiar_cursos_a_pruebas(curso_ids=ids)
        self.message_user(
            request,
            f'✅ {result.total_copiados} curso(s) copiados a {result.destino.nombre}.',
        )


class PreguntaModuloInline(admin.StackedInline):
    """Preguntas de validación del módulo (Mini examen)"""
    model = PreguntaModulo
    extra = 0
    can_delete = True
    show_change_link = True
    verbose_name = 'Pregunta de Mini Examen'
    verbose_name_plural = 'Mini examen (después de los pasos)'

    fieldsets = (
        ('Pregunta', {
            'fields': ('pregunta',),
            'description': (
                'Se envía <strong>cuando el estudiante ya recorrió todos los microcontenidos activos</strong> '
                '(no durante el flujo de *listo* intermedio). '
                'Si la dejás <strong>activa</strong>, debe contestar bien para cerrar el módulo y seguir '
                '(igual que el comportamiento sin pasos internos). '
                '<br>Esto es <em>independiente</em> de las evaluaciones por letras en cada microcontenido (que frenan el avance dentro del drip).'
            ),
        }),
        ('Opciones de Respuesta', {
            'fields': ('opcion_a', 'opcion_b', 'opcion_c', 'opcion_d', 'respuesta_correcta')
        }),
        ('Estado', {
            'fields': ('activa',)
        }),
    )


class SeccionModuloInline(admin.TabularInline):
    """Agrupa pasos en el admin; el título de la sección no se envía por WhatsApp."""
    model = SeccionModulo
    extra = 0
    can_delete = True
    ordering = ('orden', 'id')
    show_change_link = True
    verbose_name = 'Bloque'
    # Texto corto: Jazzmin usa esto en pestañas + ancla #slug-tab; títulos largos rompen el cambio de pestaña.
    verbose_name_plural = 'Secciones'
    fields = ('orden', 'activa', 'titulo', 'resumen_pasos')
    readonly_fields = ('resumen_pasos',)

    @admin.display(description='Pasos activos')
    def resumen_pasos(self, obj):
        if not obj or not obj.pk:
            return format_html(
                '<span style="color:#94a3b8;font-size:12px;">—</span>'
            )
        n = obj.pasos.filter(activo=True).count()
        return format_html(
            '<span style="font-weight:600;color:#0f766e;">{}</span>'
            '<span style="color:#64748b;font-size:11px;"> en esta sección</span>',
            n,
        )


class PasoModuloForm(forms.ModelForm):
    media_file_upload = forms.FileField(
        label='Subir archivo desde PC',
        required=False,
        help_text=(
            'Opcional. Si subes un archivo aquí, se guarda en el storage configurado '
            '(S3 en producción) y se copia su URL pública a “Media url”.'
        ),
    )

    class Meta:
        model = PasoModulo
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        inst = self.instance
        if inst and inst.pk and getattr(inst, 'tipo', None) == PasoModulo.TIPO_EVAL_OPC:
            data = inst.opciones_json
            if isinstance(data, dict):
                for letter in ('A', 'B', 'C', 'D'):
                    fname = f'eval_opcion_{letter.lower()}'
                    cur = (getattr(inst, fname, None) or '').strip()
                    if cur:
                        continue
                    val = data.get(letter)
                    if val:
                        self.initial[fname] = str(val)

    def save(self, commit=True):
        instance = super().save(commit=False)
        uploaded_file = self.cleaned_data.get('media_file_upload')
        if uploaded_file:
            modulo_id = instance.modulo_id or 'sin_modulo'
            instance.media_url = guardar_upload_admin_media(
                uploaded_file,
                carpeta='modulos/pasos',
                prefix=f'modulo_{modulo_id}',
            )

        if commit:
            instance.save()
            self.save_m2m()
        return instance


class PasoModuloInline(admin.StackedInline):
    """Microcontenidos WhatsApp dentro del módulo (orden + listo)."""
    model = PasoModulo
    form = PasoModuloForm
    extra = 0
    can_delete = True
    ordering = ('orden', 'id')
    show_change_link = True
    verbose_name = 'Paso'
    verbose_name_plural = 'Microcontenidos'
    fieldsets = (
        (None, {
            'fields': ('orden', 'seccion', 'activo', 'requiere_listo_para_avanzar', 'tipo'),
        }),
        ('Texto y multimedia (WhatsApp)', {
            'fields': ('titulo', 'contenido', 'media_url', 'media_file_upload'),
            'description': (
                'El título es referencia interna. Lo que recibe el estudiante va en «Contenido». '
                'Puedes pegar una URL o subir archivo desde tu PC. Si subes archivo, se guarda en S3 '
                'en producción y se completa Media URL automáticamente.'
            ),
        }),
        ('Evaluación tipo opciones', {
            'fields': (
                'eval_opcion_a', 'eval_opcion_b', 'eval_opcion_c', 'eval_opcion_d',
                'respuesta_correcta',
            ),
        }),
        ('Retroalimentación', {
            'fields': ('feedback_correcto', 'feedback_incorrecto'),
            'classes': ('collapse',),
        }),
        ('Avanzado / legado', {
            'fields': ('opciones_json',),
            'classes': ('collapse',),
        }),
    )
    formfield_overrides = {
        models.TextField: {
            'widget': forms.Textarea(attrs={'rows': 3, 'cols': 50, 'style': 'min-width:280px;'}),
        },
    }

    def get_formset(self, request, obj=None, **kwargs):
        self._modulo_inline_parent = obj
        return super().get_formset(request, obj, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'seccion':
            mod = getattr(self, '_modulo_inline_parent', None)
            if mod is not None and mod.pk:
                kwargs['queryset'] = SeccionModulo.objects.filter(modulo_id=mod.pk).order_by(
                    'orden', 'id'
                )
            # Módulo nuevo sin PK: no filtramos; el admin no muestra este inline (ver ModuloAdmin).
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class ArchivoModuloInline(admin.StackedInline):
    """Archivos multimedia del módulo (imágenes, videos, infografías, PDFs)"""
    model = ArchivoModulo
    extra = 1
    can_delete = True
    show_change_link = True
    verbose_name = '📎 Multimedia'
    verbose_name_plural = 'Multimedia'
    readonly_fields = ('preview_multimedia',)
    
    fieldsets = (
        (None, {
            'fields': ('tipo', 'titulo', 'descripcion'),
            'description': (
                '👉 Paso 1: Selecciona el TIPO de archivo (video, imagen, infografía, pdf, audio).<br>'
                '💡 <b>TODOS los tipos se envían como adjunto por WhatsApp</b> — videos, imágenes, '
                'infografías, PDFs y audios se entregan automáticamente al estudiante.<br>'
                '📎 Puedes agregar MÚLTIPLES archivos por módulo — todos se enviarán en orden.'
            )
        }),
        ('📤 Subir Archivo o URL', {
            'fields': ('archivo', 'preview_multimedia', 'url_externa'),
            'description': '''
            👉 Paso 2: Elige UNA opción:
            • SUBIR ARCHIVO: Sube desde tu PC (se guardará en S3 automáticamente)
            • URL EXTERNA: Pega link de YouTube, Vimeo, Google Drive, etc.
            ⚠️ IMPORTANTE: Verifica que la URL sea pública y accesible. URLs privadas no se podrán enviar.
            '''
        }),
        ('Configuración', {
            'fields': ('disponible_offline', 'orden', 'activo'),
            'classes': ('collapse',)
        }),
    )
    
    def preview_multimedia(self, obj):
        """Vista previa del archivo multimedia"""
        if not obj.archivo:
            if obj.url_externa:
                return format_html(
                    '<div style="background:#e0f2fe;padding:12px;border-radius:6px;border-left:4px solid #0284c7;">'
                    '<strong>🔗 URL Externa:</strong><br>'
                    '<a href="{}" target="_blank" style="color:#0284c7;word-break:break-all;">{}</a>'
                    '</div>',
                    obj.url_externa, obj.url_externa
                )
            return format_html('<span style="color:#999;font-style:italic;">⚠️ Sin archivo subido</span>')
        
        url = obj.archivo.url
        
        if obj.tipo == 'imagen':
            return format_html(
                '<div style="text-align:center;background:#f9fafb;padding:16px;border-radius:8px;border:2px solid #e5e7eb;">'
                '<img src="{}" style="max-width:100%;max-height:400px;border-radius:6px;box-shadow:0 4px 6px rgba(0,0,0,0.1);" />'
                '<p style="margin-top:12px;color:#6b7280;font-size:12px;">📸 Imagen cargada correctamente</p>'
                '</div>',
                url
            )
        elif obj.tipo == 'video':
            return format_html(
                '<div style="background:#f9fafb;padding:16px;border-radius:8px;border:2px solid #e5e7eb;">'
                '<video controls style="max-width:100%;border-radius:6px;box-shadow:0 4px 6px rgba(0,0,0,0.1);">'
                '<source src="{}" type="video/mp4">'
                'Tu navegador no soporta video HTML5.'
                '</video>'
                '<p style="margin-top:12px;color:#6b7280;font-size:12px;">🎥 Video cargado - URL: <code style="background:#e5e7eb;padding:2px 6px;border-radius:4px;font-size:11px;">{}</code></p>'
                '</div>',
                url, url
            )
        elif obj.tipo == 'pdf':
            return format_html(
                '<div style="background:#fef2f2;padding:14px;border-radius:6px;border-left:4px solid #dc2626;">'
                '<a href="{}" target="_blank" style="background:#dc2626;color:white;padding:10px 20px;text-decoration:none;border-radius:6px;display:inline-block;font-weight:bold;">'
                '📄 Abrir PDF en Nueva Pestaña'
                '</a>'
                '<p style="margin-top:10px;color:#991b1b;font-size:12px;">Archivo: {}</p>'
                '</div>',
                url, obj.archivo.name
            )
        elif obj.tipo == 'audio':
            return format_html(
                '<div style="background:#f9fafb;padding:16px;border-radius:8px;border:2px solid #e5e7eb;">'
                '<audio controls style="width:100%;">'
                '<source src="{}" type="audio/mpeg">'
                'Tu navegador no soporta audio HTML5.'
                '</audio>'
                '<p style="margin-top:12px;color:#6b7280;font-size:12px;">🎵 Audio cargado</p>'
                '</div>',
                url
            )
        else:
            return format_html(
                '<div style="background:#f0fdf4;padding:14px;border-radius:6px;border-left:4px solid #16a34a;">'
                '<a href="{}" target="_blank" style="color:#16a34a;font-weight:bold;">📎 Ver Archivo</a>'
                '<p style="margin-top:8px;color:#166534;font-size:12px;">{}</p>'
                '</div>',
                url, obj.archivo.name
            )
    preview_multimedia.short_description = "Vista Previa"


@admin.register(Modulo)
class ModuloAdmin(admin.ModelAdmin):
    """Administración de módulos"""
    class Media:
        css = {
            'all': ('admin/css/modulo_whatsapp_bloques.css',),
        }

    list_display = (
        'numero_titulo',
        'curso',
        'duracion_dias',
        'modo_entrega_badge',
        'pasos_activos_count',
        'examen_badge',
        'archivos_link',
        'tiene_pregunta',
        'contenido_preview',
        'ver_curso_link',
    )
    list_filter = ('curso', 'examen_obligatorio', 'modo_entrega')
    search_fields = ('titulo', 'descripcion', 'contenido')
    list_per_page = 50
    ordering = ['curso', 'numero']
    readonly_fields = ('guia_microcontenidos_whatsapp',)
    inlines = [SeccionModuloInline, PasoModuloInline, ArchivoModuloInline, PreguntaModuloInline]
    actions = ['enviar_archivos_multimedia', 'ver_archivos_multimedia', 'renumerar_modulos']

    def get_inline_instances(self, request, obj=None):
        """Módulo nuevo: solo secciones hasta la 1.ª guardada (luego se elige sección por módulo)."""
        instances = []
        for inline_class in self.inlines:
            if inline_class is PasoModuloInline and obj is None:
                continue
            instances.append(inline_class(self.model, self.admin_site))
        return instances

    def ver_curso_link(self, obj):
        """Link directo al curso padre"""
        url = reverse('admin:core_curso_change', args=[obj.curso.id])
        return format_html('<a href="{}" style="color:#2196F3;">📚 Ver Curso</a>', url)
    ver_curso_link.short_description = "Curso"

    @admin.display(description='')
    def guia_microcontenidos_whatsapp(self, obj):
        """Panel de ayuda visual (no persiste en BD)."""
        return format_html(
            '<div style="background:linear-gradient(180deg,#ecfdf5 0%,#f8fafc 100%);'
            'border:1px solid #6ee7b7;border-radius:10px;padding:16px 18px;margin:0 0 4px 0;'
            'max-width:920px;">'
            '<p style="margin:0 0 10px 0;font-size:14px;color:#065f46;font-weight:600;">'
            'Flujo recomendado (modo <em>Pasos</em> o <em>Automático</em> con pasos activos)'
            '</p>'
            '<table style="width:100%;border-collapse:collapse;font-size:13px;color:#334155;">'
            '<tr style="vertical-align:top;">'
            '<td style="width:34%;padding:8px 10px;background:#fff;border-radius:8px;border:1px solid #e2e8f0;">'
            '<strong style="color:#0f766e;">① Secciones</strong> (cuadro de abajo)<br>'
            '<span style="font-size:12px;line-height:1.45;">Cada fila es un <b>bloque</b>. '
            'El <b>orden</b> define el recorrido. El <b>título</b> es solo para el equipo (no se envía por WhatsApp); '
            'el texto al estudiante va en cada microcontenido.</span>'
            '</td>'
            '<td style="width:34%;padding:8px 10px;background:#fff;border-radius:8px;border:1px solid #e2e8f0;">'
            '<strong style="color:#0369a1;">② Microcontenidos</strong><br>'
            '<span style="font-size:12px;line-height:1.45;">Cada paso debe enlazarse a una '
            '<b>sección</b> del paso anterior. Podés poner muchos pasos dentro del mismo bloque. '
            'El estudiante avanza con <b>*listo*</b> según el ajuste del siguiente bloque.</span>'
            '</td>'
            '<td style="width:32%;padding:8px 10px;background:#fff;border-radius:8px;border:1px solid #e2e8f0;">'
            '<strong style="color:#a16207;">③ Campo «Secciones por *listo*»</strong><br>'
            '<span style="font-size:12px;line-height:1.45;">Número entero <b>1–5</b>: cuántos '
            'bloques (filas de secciones consecutivas) se mandan en cada <b>*listo*</b>. Con <b>1</b> '
            'equivale a un bloque por *listo*; con <b>2</b> manda dos bloques seguidos, etc.</span>'
            '</td>'
            '</tr></table>'
            '<p style="margin:12px 0 0 0;font-size:12px;color:#64748b;line-height:1.55;border-top:1px solid #e2e8f0;padding-top:10px;">'
            '<strong style="color:#334155;">Preguntas y bloqueo de avance</strong><br>'
            '• <strong>Dentro del drip</strong>: en «Evaluación (opciones)», la pregunta va en «Contenido» y las alternativas en A–D + letra correcta; '
            'el estudiante <strong>no sigue</strong> hasta responder bien (o cumplir reto/entrega según el tipo).<br>'
            '• <strong>Mini examen</strong> (pestaña más abajo): va <em>al final</em>, cuando ya terminó todos los pasos; si está activa, exige la respuesta correcta para '
            'dar por cerrado el módulo (independiente del JSON de cada paso).<br>'
            '• <strong>Módulo nuevo</strong>: guardá una vez con las filas de Secciones; al reabrir el registro cargás Microcontenidos y el desplegable «sección» solo lista bloques de este módulo.'
            '</p>'
            '<p style="margin:10px 0 0 0;font-size:12px;color:#64748b;line-height:1.5;">'
            'Modo <b>Legacy</b>: ignora estos pasos y envía el texto del módulo completo. '
            'Mini examen y multimedia siguen en sus pestañas.'
            '</p>'
            '</div>'
        )
    
    fieldsets = (
        (
            '📱 Guía: envío por WhatsApp (*listo*)',
            {
                'fields': ('guia_microcontenidos_whatsapp',),
                'classes': ('wide',),
                'description': (
                    'Referencia visual; no se guarda. Orden práctico: definí <strong>Secciones</strong>, '
                    '<strong>Guardá</strong>, luego pestaña <strong>Microcontenidos</strong> '
                    '(asigná cada paso a un bloque). Primera vez sin guardar solo verás Secciones.'
                ),
            },
        ),
        ('📖 Información del Módulo', {
            'fields': ('curso', 'numero', 'titulo', 'descripcion'),
            'description': (
                '<strong>Número del módulo:</strong> entero <strong>≥ 0</strong> (0 = bienvenida u onboarding; '
                'luego 1, 2, 3…). El curso ordena por este número. '
                'Sin decimales en este campo.'
            ),
        }),
        ('📝 Contenido Educativo', {
            'fields': ('contenido',),
            'description': 'Texto del módulo completo (en modo <b>Legacy</b> es lo principal que ve el estudiante).',
        }),
        ('✅ Examen Obligatorio', {
            'fields': ('examen_obligatorio', 'puntaje_minimo_aprobacion'),
            'description': 'Si activas "Examen Obligatorio", el estudiante NO podrá avanzar al siguiente módulo hasta aprobar',
            'classes': ('collapse',)
        }),
        ('📦 Entrega al estudiante y checkpoint IA', {
            'fields': ('modo_entrega', 'secciones_por_listo', 'facilitador_checkpoint'),
            'description': (
                '<div style="font-size:13px;line-height:1.5;color:#334155;">'
                '<strong>Modo de entrega:</strong> Pasos / Automático / Legacy (este último ignora la tabla de microcontenidos). '
                '<strong>Secciones por *listo*:</strong> cuántos bloques consecutivos (según el orden de secciones) salen en cada «listo». '
                '<strong>Checkpoint facilitadora:</strong> solo si el curso usa agentes IA; aplica al <em>cerrar</em> este módulo.'
                '</div>'
            ),
        }),
        ('⏱️ Duración, calendario y multimedia', {
            'fields': ('duracion_dias', 'habilitado_desde'),
            'description': (
                'Multimedia del módulo: usá la tabla <strong>ARCHIVOS MULTIMEDIA</strong> más abajo '
                '(videos, imágenes, PDFs, etc.). '
                '<strong>Disponible desde</strong>: opcional; bloquea el envío de este módulo hasta esa fecha para todos los estudiantes, '
                'salvo que en el <em>Cliente</em> exista una habilitación distinta para el mismo módulo.'
            ),
        }),
    )
    
    def modo_entrega_badge(self, obj):
        from .models import Modulo
        labels = dict(Modulo.MODOS_ENTREGA)
        t = labels.get(obj.modo_entrega, obj.modo_entrega)
        return format_html('<span style="font-size:11px;">{}</span>', t)
    modo_entrega_badge.short_description = 'Entrega'

    def pasos_activos_count(self, obj):
        count = obj.pasos.filter(activo=True, seccion__activa=True).count()
        if obj.modo_entrega == 'pasos' and count == 0:
            return format_html(
                '<span style="color:#c62828;font-weight:bold;">'
                '⚠️ Modo PASOS pero sin pasos activos (count: {})'
                '</span>',
                count,
            )
        return count
    pasos_activos_count.short_description = 'Pasos activos'

    def numero_titulo(self, obj):
        return f"Módulo {obj.numero}: {obj.titulo}"
    numero_titulo.short_description = "Módulo"
    
    def examen_badge(self, obj):
        if obj.examen_obligatorio:
            return format_html(
                '<span style="background:#ffebee;color:#c62828;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">🔒 OBLIGATORIO ({0}%)</span>',
                obj.puntaje_minimo_aprobacion
            )
        return format_html('<span style="color:#999;">Sin examen</span>')
    examen_badge.short_description = "Examen"
    
    def archivos_link(self, obj):
        count = obj.archivos_multimedia.filter(activo=True).count()
        if count > 0:
            url = f"/admin/core/archivomodulo/?modulo__id__exact={obj.id}"
            return format_html(
                '<a href="{}" style="background:#e3f2fd;color:#1976d2;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;text-decoration:none;">📁 {} archivo(s)</a>',
                url, count
            )
        return format_html('<span style="color:#999;">Sin archivos</span>')
    archivos_link.short_description = "Multimedia"
    
    def tiene_pregunta(self, obj):
        count = obj.preguntas.filter(activa=True).count()
        if count > 0:
            return format_html('<span style="color:green;">✅ {} pregunta(s)</span>', count)
        return format_html('<span style="color:red;">❌ Sin pregunta</span>')
    tiene_pregunta.short_description = "Mini Examen"
    
    def contenido_preview(self, obj):
        preview = obj.contenido[:60] + "..." if len(obj.contenido) > 60 else obj.contenido
        return format_html('<span style="color:#666;font-style:italic;">{}</span>', preview)
    contenido_preview.short_description = "Vista Previa"
    
    def enviar_archivos_multimedia(self, request, queryset):
        """Envía los archivos multimedia de los módulos seleccionados a estudiantes inscritos"""
        from .utils import enviar_whatsapp_twilio
        import json
        
        enviados = 0
        errores = 0
        
        for modulo in queryset:
            archivos = modulo.archivos_multimedia.filter(activo=True)
            if not archivos.exists():
                continue
            
            # Obtener estudiantes inscritos en el curso
            estudiantes = Estudiante.objects.filter(
                progreso__curso=modulo.curso,
                activo=True
            ).distinct()
            
            for estudiante in estudiantes:
                try:
                    # Mensaje con lista de archivos
                    mensaje = f"📚 *{modulo.titulo}*\n\n"
                    mensaje += f"Tienes {archivos.count()} archivo(s) multimedia disponible(s):\n\n"
                    
                    for i, archivo in enumerate(archivos, 1):
                        icono = {
                            'video': '🎥',
                            'imagen': '🖼️',
                            'infografia': '📊',
                            'pdf': '📄',
                            'audio': '🎵'
                        }.get(archivo.tipo, '📁')
                        
                        mensaje += f"{icono} *{i}. {archivo.titulo}*\n"
                        if archivo.descripcion:
                            mensaje += f"   {archivo.descripcion}\n"
                        
                        if archivo.disponible_offline:
                            url_descarga = f"{request.build_absolute_uri('/media/descargar-archivo/')}{archivo.id}/"
                            mensaje += f"   🔗 Descarga: {url_descarga}\n"
                        
                        if archivo.url_externa:
                            mensaje += f"   🌐 Ver online: {archivo.url_externa}\n"
                        
                        mensaje += "\n"
                    
                    # Enviar por WhatsApp
                    enviar_whatsapp_twilio(estudiante.telefono, mensaje)
                    enviados += 1
                    
                except Exception as e:
                    print(f"Error enviando archivos a {estudiante.telefono}: {e}")
                    errores += 1
        
        self.message_user(
            request,
            f'✅ {enviados} mensaje(s) enviado(s) con archivos multimedia. ❌ {errores} error(es).'
        )
    enviar_archivos_multimedia.short_description = "📤 Enviar archivos multimedia a estudiantes"
    
    @admin.action(description='� Renumerar módulos (1, 2, 3...)')
    def renumerar_modulos(self, request, queryset):
        """Renumera los módulos seleccionados empezando desde 1"""
        # Agrupar por curso
        modulos_por_curso = {}
        for modulo in queryset.order_by('curso', 'numero', 'id'):
            if modulo.curso not in modulos_por_curso:
                modulos_por_curso[modulo.curso] = []
            modulos_por_curso[modulo.curso].append(modulo)
        
        total_renumerados = 0
        for curso, modulos in modulos_por_curso.items():
            for idx, modulo in enumerate(modulos, start=1):
                if modulo.numero != idx:
                    modulo.numero = idx
                    modulo.save()
                    total_renumerados += 1
        
        self.message_user(
            request,
            f"✅ {total_renumerados} módulos renumerados correctamente",
            level='success'
        )
    renumerar_modulos.short_description = "🔢 Renumerar módulos (1, 2, 3...)"
    
    @admin.action(description='�📁 Ver archivos multimedia de módulos')
    def ver_archivos_multimedia(self, request, queryset):
        """Redirige a ver los archivos multimedia de los módulos seleccionados"""
        from django.shortcuts import redirect
        modulo_ids = ','.join(str(m.id) for m in queryset)
        return redirect(f'/admin/core/archivomodulo/?modulo__id__in={modulo_ids}')


class PreguntaExamenInline(admin.TabularInline):
    """Preguntas dentro del examen"""
    model = PreguntaExamen
    extra = 1
    fields = ('numero', 'pregunta', 'respuesta_correcta', 'puntos')
    ordering = ['numero']


class ExamenAdmin(admin.ModelAdmin):
    """Administración de exámenes"""
    list_display = ('curso_nombre', 'total_preguntas_display', 'puntaje_minimo')
    list_filter = ('curso',)
    search_fields = ('curso__nombre', 'instrucciones')
    inlines = [PreguntaExamenInline]
    
    fieldsets = (
        ('📝 Configuración del Examen', {
            'fields': ('curso', 'instrucciones', 'puntaje_minimo')
        }),
    )
    
    def curso_nombre(self, obj):
        return f"{obj.curso.emoji} {obj.curso.nombre}"
    curso_nombre.short_description = "Curso"
    
    def total_preguntas_display(self, obj):
        count = obj.preguntas.count()
        return format_html(
            '<span style="background:#fff3cd;padding:4px 8px;border-radius:4px;">{} preguntas</span>',
            count
        )
    total_preguntas_display.short_description = "Preguntas"


class PreguntaExamenAdmin(admin.ModelAdmin):
    """Administración de preguntas de examen"""
    list_display = ('numero_pregunta', 'examen', 'puntos', 'pregunta_preview')
    list_filter = ('examen__curso',)
    search_fields = ('pregunta', 'respuesta_correcta')
    ordering = ['examen', 'numero']
    
    fieldsets = (
        ('❓ Pregunta', {
            'fields': ('examen', 'numero', 'pregunta')
        }),
        ('✅ Respuesta', {
            'fields': ('respuesta_correcta', 'puntos'),
            'description': 'Palabras clave separadas por comas (la IA evaluará si están presentes)'
        }),
    )
    
    def numero_pregunta(self, obj):
        return f"Pregunta {obj.numero}"
    numero_pregunta.short_description = "N°"
    
    def pregunta_preview(self, obj):
        preview = obj.pregunta[:80] + "..." if len(obj.pregunta) > 80 else obj.pregunta
        return preview
    pregunta_preview.short_description = "Pregunta"


@admin.register(ProgresoEstudiante)
class ProgresoEstudianteAdmin(admin.ModelAdmin):
    """Seguimiento del progreso de estudiantes"""
    list_display = ('estudiante', 'curso', 'barra_progreso', 'modulo_actual', 'completado_badge', 'certificado_status', 'fecha_ultimo_avance', 'fecha_inicio')
    list_filter = ('completado', 'curso', 'fecha_inicio')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'curso__nombre')
    readonly_fields = ('fecha_inicio', 'porcentaje_avance', 'info_certificado')
    list_per_page = 50
    ordering = ('-fecha_inicio',)
    actions = ['exportar_progreso_excel', 'exportar_progreso_csv', 'generar_certificados_pendientes']  # ✅ Nuevas acciones
    
    fieldsets = (
        ('👤 Estudiante y Curso', {
            'fields': ('estudiante', 'curso')
        }),
        ('📊 Progreso', {
            'fields': ('modulo_actual', 'completado', 'porcentaje_avance')
        }),
        ('📅 Fechas', {
            'fields': ('fecha_inicio', 'fecha_ultimo_avance', 'fecha_completado')
        }),
    )
    
    def barra_progreso(self, obj):
        """Muestra una barra de progreso visual"""
        porcentaje = obj.porcentaje_avance()
        
        # Colores según progreso
        if porcentaje >= 80:
            color = '#4caf50'
        elif porcentaje >= 50:
            color = '#ff9800'
        else:
            color = '#f44336'
        
        return format_html(
            '<div style="width:100px;height:20px;background:#f0f0f0;border-radius:10px;overflow:hidden;border:1px solid #ddd;">'
            '<div style="width:{}%;height:100%;background:{};transition:width 0.3s;display:flex;align-items:center;justify-content:center;color:white;font-size:11px;font-weight:bold;">'
            '{}%'
            '</div></div>',
            porcentaje, color, porcentaje
        )
    barra_progreso.short_description = "Progreso"
    
    def certificado_status(self, obj):
        """Muestra si hay certificado generado"""
        if obj.completado:
            certificado = Certificado.objects.filter(estudiante=obj.estudiante, curso=obj.curso, emitido=True).first()
            if certificado:
                return format_html(
                    '<span style="color:#4caf50;font-weight:bold;">🏆 Emitido</span>'
                )
            else:
                return format_html(
                    '<span style="color:#ff9800;font-weight:bold;">⏳ Pendiente</span>'
                )
        return format_html('<span style="color:#999;">-</span>')
    certificado_status.short_description = "Certificado"
    
    def info_certificado(self, obj):
        """Información del certificado en los detalles"""
        if obj.completado:
            certificado = Certificado.objects.filter(estudiante=obj.estudiante, curso=obj.curso).first()
            if certificado:
                return format_html(
                    '<div style="background:#f5f5f5;padding:10px;border-radius:4px;border-left:4px solid #4caf50;">'
                    '<strong>✅ Certificado Generado</strong><br>'
                    'Código: <code>{}</code><br>'
                    'Calificación: <strong>{}</strong>%<br>'
                    'Emitido: {}'
                    '</div>',
                    certificado.codigo_verificacion,
                    int(certificado.calificacion_final),
                    certificado.fecha_emision.strftime('%d/%m/%Y') if certificado.fecha_emision else 'N/A'
                )
            else:
                return format_html(
                    '<div style="background:#fff3cd;padding:10px;border-radius:4px;border-left:4px solid #ff9800;">'
                    '<strong>⏳ Certificado Pendiente</strong><br>'
                    'El curso está completo pero el certificado aún no se ha generado'
                    '</div>'
                )
        return format_html(
            '<div style="background:#f5f5f5;padding:10px;border-radius:4px;border-left:4px solid #999;">'
            '<span style="color:#999;">El curso aún no está completo</span>'
            '</div>'
        )
    info_certificado.short_description = "Estado del Certificado"
    
    def porcentaje_badge(self, obj):
        porcentaje = obj.porcentaje_avance()
        if porcentaje >= 80:
            color = '#4caf50'
        elif porcentaje >= 50:
            color = '#ff9800'
        else:
            color = '#f44336'
        return format_html(
            '<span style="background:{};color:white;padding:4px 12px;border-radius:12px;font-weight:bold;"{}%</span>',
            color, porcentaje
        )
    porcentaje_badge.short_description = "Avance"
    
    def completado_badge(self, obj):
        if obj.completado:
            return format_html('<span style="color:green;">✅ Completo</span>')
        return format_html('<span style="color:orange;">⏳ En progreso</span>')
    completado_badge.short_description = "Estado"
    
    @admin.action(description='🏆 Generar certificados pendientes')
    def generar_certificados_pendientes(self, request, queryset):
        """Genera certificados para cursos completados que no los tienen"""
        from .certificado_service import generar_y_guardar_certificado, crear_certificado_automatico, enviar_certificado_whatsapp
        
        generados = 0
        errores = 0
        
        for progreso in queryset.filter(completado=True):
            certificado = Certificado.objects.filter(estudiante=progreso.estudiante, curso=progreso.curso).first()
            if not certificado:
                try:
                    certificado = crear_certificado_automatico(progreso.estudiante, progreso.curso)
                    if certificado and not certificado.emitido:
                        generar_y_guardar_certificado(certificado)
                    if certificado and certificado.emitido:
                        enviar_certificado_whatsapp(certificado)
                    generados += 1
                except Exception as e:
                    logger.error(f"Error al generar certificado para {progreso.estudiante.nombre}: {str(e)}")
                    errores += 1
        
        if generados > 0:
            self.message_user(request, f'✅ {generados} certificado(s) generado(s) y enviado(s)', messages.SUCCESS)
        if errores > 0:
            self.message_user(request, f'❌ {errores} error(es) al generar', messages.ERROR)
    
    @admin.action(description='📊 Exportar progreso a Excel')
    def exportar_progreso_excel(self, request, queryset):
        """Exporta el progreso de estudiantes a Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Progreso Estudiantes"
        
        # Encabezados
        headers = ['Estudiante', 'Teléfono', 'Curso', 'Módulo Actual', 'Avance %', 'Completado', 'Fecha Inicio', 'Fecha Completado']
        ws.append(headers)
        
        # Estilo
        header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for progreso in queryset:
            ws.append([
                progreso.estudiante.nombre,
                f"+{progreso.estudiante.telefono}",
                progreso.curso.nombre,
                progreso.modulo_actual or 'No iniciado',
                progreso.porcentaje_avance(),
                "Sí" if progreso.completado else "No",
                progreso.fecha_inicio.strftime('%Y-%m-%d %H:%M'),
                progreso.fecha_completado.strftime('%Y-%m-%d %H:%M') if progreso.fecha_completado else 'N/A'
            ])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'progreso_estudiantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @admin.action(description='📄 Exportar progreso a CSV')
    def exportar_progreso_csv(self, request, queryset):
        """Exporta el progreso de estudiantes a CSV"""
        import csv
        from datetime import datetime
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'progreso_estudiantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Estudiante', 'Teléfono', 'Curso', 'Módulo Actual', 'Avance %', 'Completado', 'Fecha Inicio', 'Fecha Completado'])
        
        for progreso in queryset:
            writer.writerow([
                progreso.estudiante.nombre,
                f"+{progreso.estudiante.telefono}",
                progreso.curso.nombre,
                progreso.modulo_actual or 'No iniciado',
                progreso.porcentaje_avance(),
                "Sí" if progreso.completado else "No",
                progreso.fecha_inicio.strftime('%Y-%m-%d %H:%M'),
                progreso.fecha_completado.strftime('%Y-%m-%d %H:%M') if progreso.fecha_completado else 'N/A'
            ])
        
        return response


@admin.register(ModuloCompletado)
class ModuloCompletadoAdmin(admin.ModelAdmin):
    """Registro de módulos completados"""
    list_display = ('estudiante_nombre', 'modulo_info', 'fecha_completado')
    list_filter = ('fecha_completado', 'modulo__curso')
    search_fields = ('progreso__estudiante__nombre', 'modulo__titulo')
    readonly_fields = ('fecha_completado',)
    ordering = ('-fecha_completado',)
    
    def estudiante_nombre(self, obj):
        return obj.progreso.estudiante.nombre
    estudiante_nombre.short_description = "Estudiante"
    
    def modulo_info(self, obj):
        return f"{obj.modulo.curso.emoji} {obj.modulo.titulo}"
    modulo_info.short_description = "Módulo"


class ResultadoExamenAdmin(admin.ModelAdmin):
    """Resultados de exámenes"""
    list_display = ('estudiante', 'examen_info', 'puntaje_badge', 'aprobado_badge', 'fecha_realizado')
    list_filter = ('aprobado', 'examen__curso', 'fecha_realizado')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'examen__curso__nombre')
    readonly_fields = ('fecha_realizado', 'respuestas', 'feedback')
    ordering = ('-fecha_realizado',)
    
    fieldsets = (
        ('👤 Estudiante y Examen', {
            'fields': ('estudiante', 'examen')
        }),
        ('📊 Resultado', {
            'fields': ('puntaje', 'aprobado')
        }),
        ('📝 Respuestas y Retroalimentación', {
            'fields': ('respuestas', 'feedback'),
            'classes': ('collapse',)
        }),
        ('📅 Fecha', {
            'fields': ('fecha_realizado',)
        }),
    )
    
    def examen_info(self, obj):
        return f"{obj.examen.curso.emoji} {obj.examen.curso.nombre}"
    examen_info.short_description = "Examen"
    
    def puntaje_badge(self, obj):
        if obj.puntaje >= 80:
            color = '#4caf50'
        elif obj.puntaje >= 70:
            color = '#ff9800'
        else:
            color = '#f44336'
        return format_html(
            '<span style="background:{};color:white;padding:6px 12px;border-radius:12px;font-weight:bold;font-size:14px;">{}/100</span>',
            color, obj.puntaje
        )
    puntaje_badge.short_description = "Puntaje"
    
    def aprobado_badge(self, obj):
        if obj.aprobado:
            return format_html('<span style="background:#4caf50;color:white;padding:4px 12px;border-radius:12px;">✅ APROBADO</span>')
        return format_html('<span style="background:#f44336;color:white;padding:4px 12px;border-radius:12px;">❌ REPROBADO</span>')
    aprobado_badge.short_description = "Estado"


# Personalizar el admin site
admin.site.site_header = "eki - Chatbot Agro 🌱"
admin.site.site_title = "Administración eki"
admin.site.index_title = "Panel de Control - Chatbot Educativo"


# ==========================================
# PERSONALIZACIÓN DEL INDEX DEL ADMIN
# ==========================================
from django.contrib.admin import AdminSite
from django.urls import reverse
from django.utils.html import format_html

# Sobrescribir el template del index para agregar enlaces personalizados
def index_view(self, request, extra_context=None):
    """Vista personalizada del index del admin con enlaces a conversaciones"""
    extra_context = extra_context or {}
    
    # Agregar enlace a conversaciones en el contexto
    extra_context['conversaciones_url'] = reverse('conversaciones')
    extra_context['dashboard_url'] = reverse('dashboard_unificado')
    extra_context['dashboard_control_url'] = reverse('dashboard_unificado')  # Para compatibilidad con template
    extra_context['dashboard_analytics_url'] = reverse('dashboard_analytics')
    
    return AdminSite.index(self, request, extra_context)

# Aplicar la vista personalizada
admin.site.index = index_view.__get__(admin.site, AdminSite)


# ========================================
# 🎮 GAMIFICACIÓN
# ========================================
# 🏆 ADMIN UNIFICADO DE GAMIFICACIÓN (TODO EN UNO)
# ========================================

class PerfilGamificacionAdmin(admin.ModelAdmin):
    """
    🏆 GESTIÓN UNIFICADA DE GAMIFICACIÓN
    Desde aquí puedes gestionar:
    - Perfiles de Gamificación
    - Ver Badges (insignias)
    - Ver Recompensas
    - Ver Canjes de Recompensas
    """
    list_display = ('estudiante_info', 'nivel_display', 'puntos_totales', 'racha_display', 'badges_link', 'recompensas_link', 'posicion_ranking')
    list_filter = ('nivel', 'racha_dias_actual')
    search_fields = ('estudiante__nombre', 'estudiante__telefono')
    readonly_fields = ('puntos_totales', 'nivel', 'experiencia_nivel_actual', 'fecha_creacion', 'fecha_actualizacion', 'posicion_ranking')
    list_per_page = 50
    ordering = ['-puntos_totales']
    actions = ['ver_badges', 'ver_recompensas', 'resetear_racha']
    
    fieldsets = (
        ('👤 Estudiante', {
            'fields': ('estudiante',)
        }),
        ('🎯 Nivel y Puntos', {
            'fields': ('nivel', 'puntos_totales', 'experiencia_nivel_actual')
        }),
        ('🔥 Rachas', {
            'fields': ('racha_dias_actual', 'racha_dias_maxima', 'ultima_actividad')
        }),
        ('📊 Estadísticas', {
            'fields': ('modulos_completados', 'examenes_aprobados', 'preguntas_respondidas', 'audios_enviados')
        }),
        ('🏆 Ranking', {
            'fields': ('posicion_ranking',)
        }),
    )
    
    def estudiante_info(self, obj):
        return f"{obj.estudiante.nombre}"
    estudiante_info.short_description = "Estudiante"
    
    def nivel_display(self, obj):
        colores = {
            1: '#9e9e9e', 2: '#795548', 3: '#4caf50', 4: '#03a9f4',
            5: '#3f51b5', 6: '#9c27b0', 7: '#e91e63', 8: '#ff5722',
            9: '#ff9800', 10: '#ffc107'
        }
        color = colores.get(obj.nivel, '#000')
        porcentaje = obj.porcentaje_nivel()
        return format_html(
            '<div style="background:{};color:white;padding:8px 16px;border-radius:20px;font-weight:bold;text-align:center;">'
            'Nivel {} <br><small>{}% progreso</small></div>',
            color, obj.nivel, porcentaje
        )
    nivel_display.short_description = "Nivel"
    
    def racha_display(self, obj):
        if obj.racha_dias_actual >= 7:
            color = '#ff5722'
            emoji = '🔥🔥'
        elif obj.racha_dias_actual >= 3:
            color = '#ff9800'
            emoji = '🔥'
        else:
            color = '#9e9e9e'
            emoji = '📅'
        
        return format_html(
            '<span style="background:{};color:white;padding:6px 12px;border-radius:12px;font-weight:bold;">'
            '{} {} días</span>',
            color, emoji, obj.racha_dias_actual
        )
    racha_display.short_description = "Racha Actual"
    
    def badges_link(self, obj):
        """Link directo para ver badges del estudiante"""
        count = obj.get_badges().count()
        
        return format_html(
            '<a href="/admin/core/badge/" style="background:#ffc107;color:#000;padding:6px 12px;border-radius:12px;text-decoration:none;font-size:11px;font-weight:600;">🏅 {} Badges</a>',
            count
        )
    badges_link.short_description = "Insignias"
    
    def recompensas_link(self, obj):
        """Link directo para ver canjes de recompensas"""
        count = CanjeRecompensa.objects.filter(estudiante=obj.estudiante).count()
        
        return format_html(
            '<a href="/admin/core/canjerecompensa/?estudiante__id__exact={}" style="background:#9c27b0;color:white;padding:6px 12px;border-radius:12px;text-decoration:none;font-size:11px;font-weight:600;">🎁 {} Canjes</a>',
            obj.estudiante.id,
            count
        )
    recompensas_link.short_description = "Recompensas"
    
    def ver_badges(self, request, queryset):
        """Ver todos los badges disponibles"""
        from django.shortcuts import redirect
        return redirect('/admin/core/badge/')
    ver_badges.short_description = "🏅 Ver catálogo de badges"
    
    def ver_recompensas(self, request, queryset):
        """Ver todas las recompensas disponibles"""
        from django.shortcuts import redirect
        return redirect('/admin/core/recompensa/')
    ver_recompensas.short_description = "🎁 Ver catálogo de recompensas"
    
    def resetear_racha(self, request, queryset):
        """Resetear racha de estudiantes seleccionados"""
        queryset.update(racha_dias_actual=0)
        self.message_user(request, f"✅ Racha reseteada para {queryset.count()} estudiante(s)")
    resetear_racha.short_description = "🔄 Resetear racha"
    
    def badges_count(self, obj):
        count = obj.get_badges().count()
        if count > 0:
            return format_html(
                '<span style="background:#ffc107;color:#000;padding:6px 12px;border-radius:12px;font-weight:bold;">'
                '🏆 {} badges</span>',
                count
            )
        return format_html('<span style="color:#999;">0</span>')
    badges_count.short_description = "Badges"


class BadgeAdmin(admin.ModelAdmin):
    """Administración de badges/insignias (Ver también desde Perfil Gamificación)"""
    list_display = ('icono_nombre', 'tipo', 'descripcion_corta', 'criterios', 'puntos_bonus', 'total_obtenidos_display', 'activo')
    list_filter = ('tipo', 'activo', 'es_secreto')
    search_fields = ('nombre', 'descripcion')
    list_editable = ('activo',)
    list_per_page = 50
    ordering = ['orden', 'tipo', 'nombre']
    
    fieldsets = (
        ('🏆 Información del Badge', {
            'fields': ('nombre', 'descripcion', 'icono', 'tipo')
        }),
        ('✅ Criterios de Obtención', {
            'fields': ('nivel_requerido', 'valor_requerido', 'curso_requerido', 'puntos_bonus')
        }),
        ('⚙️ Configuración', {
            'fields': ('es_secreto', 'activo', 'orden')
        }),
    )
    
    actions = ['duplicar_badge', 'activar_badges', 'desactivar_badges']
    
    def icono_nombre(self, obj):
        return f"{obj.icono} {obj.nombre}"
    icono_nombre.short_description = "Badge"
    
    def descripcion_corta(self, obj):
        if len(obj.descripcion) > 60:
            return obj.descripcion[:60] + '...'
        return obj.descripcion
    descripcion_corta.short_description = "Descripción"
    
    def criterios(self, obj):
        """Muestra los criterios para obtener el badge"""
        criterios = []
        if obj.nivel_requerido:
            criterios.append(f"Nivel {obj.nivel_requerido}")
        if obj.valor_requerido:
            if obj.tipo == 'RACHA':
                criterios.append(f"{obj.valor_requerido} días de racha")
            elif obj.tipo == 'CURSO':
                criterios.append(f"{obj.valor_requerido} cursos completados")
            else:
                criterios.append(f"Valor: {obj.valor_requerido}")
        if obj.curso_requerido:
            criterios.append(f"Curso: {obj.curso_requerido.nombre}")
        
        if not criterios:
            return format_html('<span style="color:#999;font-style:italic;">Sin criterios</span>')
        
        return format_html('<span style="color:#666;">{}</span>', ' | '.join(criterios))
    criterios.short_description = "Criterios"
    
    def total_obtenidos_display(self, obj):
        count = obj.total_obtenidos()
        if count > 0:
            return format_html(
                '<span style="background:#4caf50;color:white;padding:4px 12px;border-radius:12px;font-weight:bold;">{} estudiantes</span>',
                count
            )
        return format_html('<span style="color:#999;">Nadie aún</span>')
    total_obtenidos_display.short_description = "Obtenido por"
    
    def duplicar_badge(self, request, queryset):
        """Duplica badges seleccionados"""
        count = 0
        for badge in queryset:
            badge.pk = None
            badge.nombre = f"{badge.nombre} (Copia)"
            badge.save()
            count += 1
        self.message_user(request, f"{count} badge(s) duplicado(s)")
    duplicar_badge.short_description = "📋 Duplicar badges"
    
    def activar_badges(self, request, queryset):
        count = queryset.update(activo=True)
        self.message_user(request, f"{count} badge(s) activado(s)")
    activar_badges.short_description = "✅ Activar badges"
    
    def desactivar_badges(self, request, queryset):
        count = queryset.update(activo=False)
        self.message_user(request, f"{count} badge(s) desactivado(s)")
    desactivar_badges.short_description = "❌ Desactivar badges"


class BadgeEstudianteAdmin(admin.ModelAdmin):
    """Administración de badges obtenidos por estudiantes"""
    list_display = ('estudiante', 'badge_display', 'fecha_obtenido')
    list_filter = ('badge__tipo', 'fecha_obtenido')
    search_fields = ('estudiante__nombre', 'badge__nombre')
    readonly_fields = ('fecha_obtenido',)
    list_per_page = 50

    def badge_display(self, obj):
        return f"{obj.badge.icono} {obj.badge.nombre}"
    badge_display.short_description = "Badge"


class AliadoEmpleabilidadAdmin(admin.ModelAdmin):
    list_display = (
        'nombre_empresa',
        'cliente',
        'vacantes_activas',
        'cupos_disponibles',
        'prioridad',
        'vigencia_desde',
        'vigencia_hasta',
        'latitud',
        'longitud',
        'codigo_secreto',
    )
    list_filter = ('vacantes_activas', 'cliente', 'prioridad', 'vigencia_desde', 'vigencia_hasta')
    search_fields = ('nombre_empresa', 'codigo_secreto')
    list_editable = ('vacantes_activas', 'cupos_disponibles', 'prioridad')


class MisionEmpleabilidadAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'estudiante',
        'aliado',
        'cliente',
        'estado',
        'estado_flujo',
        'puntaje_prioridad',
        'distancia_metros',
        'codigo_validado',
        'puntos_otorgados',
        'fecha_descubierta',
    )
    list_filter = ('estado', 'estado_flujo', 'cliente', 'codigo_validado', 'fecha_descubierta')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'aliado__nombre_empresa')
    readonly_fields = (
        'fecha_descubierta',
        'fecha_reclamada',
        'fecha_completada',
        'fecha_interes',
        'fecha_postulacion',
        'fecha_entrevista',
        'fecha_vinculacion',
    )
    list_per_page = 100


class PreguntaAbiertaFinalCursoAdmin(admin.ModelAdmin):
    list_display = ('curso', 'orden', 'activa', 'fecha_creacion')
    list_filter = ('activa', 'curso')
    search_fields = ('curso__nombre', 'pregunta')
    ordering = ('curso', 'orden', 'id')


class RespuestaAbiertaFinalAdmin(admin.ModelAdmin):
    list_display = ('estudiante', 'curso', 'estado', 'calificacion', 'fecha_respuesta', 'fecha_calificacion')
    list_filter = ('estado', 'curso', 'fecha_respuesta')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'respuesta_texto')
    readonly_fields = ('estudiante', 'curso', 'pregunta', 'progreso', 'respuesta_texto', 'fecha_respuesta')

    fieldsets = (
        ('Respuesta del estudiante', {
            'fields': ('estudiante', 'curso', 'pregunta', 'progreso', 'respuesta_texto', 'fecha_respuesta')
        }),
        ('Calificación facilitadora', {
            'fields': ('estado', 'calificacion', 'retroalimentacion', 'calificada_por', 'fecha_calificacion')
        }),
    )

    def save_model(self, request, obj, form, change):
        if obj.calificacion is not None:
            obj.estado = 'calificada'
            obj.calificada_por = request.user
            if not obj.fecha_calificacion:
                obj.fecha_calificacion = timezone.now()
        super().save_model(request, obj, form, change)

    ordering = ['-fecha_respuesta']


try:
    admin.site.unregister(RespuestaAbiertaFinal)
except admin.sites.NotRegistered:
    pass


class TransaccionPuntosAdmin(admin.ModelAdmin):
    """Historial de transacciones de puntos"""
    list_display = ('estudiante_nombre', 'puntos_display', 'tipo', 'razon', 'fecha')
    list_filter = ('tipo', 'fecha')
    search_fields = ('perfil__estudiante__nombre', 'razon')
    readonly_fields = ('fecha',)
    list_per_page = 100
    ordering = ['-fecha']
    
    def estudiante_nombre(self, obj):
        return obj.perfil.estudiante.nombre
    estudiante_nombre.short_description = "Estudiante"
    
    def puntos_display(self, obj):
        if obj.tipo in ['GANANCIA', 'BONUS']:
            color = '#4caf50'
            signo = '+'
        else:
            color = '#f44336'
            signo = '-'
        
        return format_html(
            '<span style="background:{};color:white;padding:4px 12px;border-radius:8px;font-weight:bold;">{}{}</span>',
            color, signo, obj.puntos
        )
    puntos_display.short_description = "Puntos"


# ========== RECOMPENSAS ==========
class RecompensaAdmin(admin.ModelAdmin):
    """Gestión de recompensas canjeables (Ver también desde Perfil Gamificación)"""
    list_display = ('icono_nombre', 'puntos_requeridos', 'tipo', 'estado', 'cantidad_info', 'nivel_minimo', 'destacado', 'canjes_totales')
    list_filter = ('tipo', 'estado', 'destacado', 'activo')
    search_fields = ('nombre', 'descripcion')
    list_editable = ('destacado',)
    
    fieldsets = (
        ('Información Básica', {
            'fields': ('nombre', 'descripcion', 'icono', 'imagen_url')
        }),
        ('Configuración', {
            'fields': ('tipo', 'puntos_requeridos', 'estado', 'cantidad_disponible', 'nivel_minimo')
        }),
        ('Disponibilidad Temporal', {
            'fields': ('fecha_inicio', 'fecha_fin'),
            'classes': ('collapse',)
        }),
        ('Entrega', {
            'fields': ('instrucciones_entrega', 'enlace_descarga'),
            'classes': ('collapse',)
        }),
        ('Visualización', {
            'fields': ('orden', 'destacado', 'activo')
        }),
    )
    
    def icono_nombre(self, obj):
        destacado = '⭐' if obj.destacado else ''
        return format_html(
            '<span style="font-size:18px;">{} {}</span> {}',
            obj.icono, obj.nombre, destacado
        )
    icono_nombre.short_description = "Recompensa"
    
    def cantidad_info(self, obj):
        restante = obj.cantidad_restante()
        if restante is None:
            return format_html('<span style="color:#4caf50;">∞ Ilimitado</span>')
        
        color = '#4caf50' if restante > 10 else '#ff9800' if restante > 0 else '#f44336'
        return format_html(
            '<span style="color:{};">{} / {}</span>',
            color, restante, obj.cantidad_disponible
        )
    cantidad_info.short_description = "Disponible"
    
    def canjes_totales(self, obj):
        return format_html(
            '<span style="background:#2196f3;color:white;padding:4px 8px;border-radius:4px;">{} canjes</span>',
            obj.cantidad_canjeada
        )
    canjes_totales.short_description = "Canjeado"
    
    actions = ['duplicar_recompensa', 'marcar_destacado', 'marcar_agotado']
    
    def duplicar_recompensa(self, request, queryset):
        for recompensa in queryset:
            recompensa.pk = None
            recompensa.nombre = f"{recompensa.nombre} (Copia)"
            recompensa.cantidad_canjeada = 0
            recompensa.save()
        self.message_user(request, f"{queryset.count()} recompensa(s) duplicada(s)")
    duplicar_recompensa.short_description = "Duplicar recompensas seleccionadas"
    
    def marcar_destacado(self, request, queryset):
        queryset.update(destacado=True)
        self.message_user(request, f"{queryset.count()} recompensa(s) marcada(s) como destacadas")
    marcar_destacado.short_description = "Marcar como destacado"
    
    def marcar_agotado(self, request, queryset):
        queryset.update(estado='AGOTADO')
        self.message_user(request, f"{queryset.count()} recompensa(s) marcada(s) como agotadas")
    marcar_agotado.short_description = "Marcar como agotado"


class CanjeRecompensaAdmin(admin.ModelAdmin):
    """Gestión de canjes de recompensas (Ver también desde Perfil Gamificación)"""
    list_display = ('estudiante_nombre', 'recompensa_info', 'puntos_gastados', 'estado_display', 'fecha_canje', 'fecha_entrega', 'atendido_por')
    list_filter = ('estado', 'fecha_canje', 'recompensa__tipo')
    search_fields = ('estudiante__nombre', 'recompensa__nombre')
    readonly_fields = ('estudiante', 'recompensa', 'puntos_gastados', 'fecha_canje')
    
    fieldsets = (
        ('Información del Canje', {
            'fields': ('estudiante', 'recompensa', 'puntos_gastados', 'fecha_canje', 'estado')
        }),
        ('Entrega', {
            'fields': ('fecha_entrega', 'nota_entrega', 'atendido_por')
        }),
    )
    
    def estudiante_nombre(self, obj):
        return obj.estudiante.nombre
    estudiante_nombre.short_description = "Estudiante"
    
    def recompensa_info(self, obj):
        return format_html(
            '{} <b>{}</b>',
            obj.recompensa.icono, obj.recompensa.nombre
        )
    recompensa_info.short_description = "Recompensa"
    
    def estado_display(self, obj):
        colores = {
            'PENDIENTE': '#ff9800',
            'PROCESANDO': '#2196f3',
            'ENTREGADO': '#4caf50',
            'CANCELADO': '#f44336'
        }
        return format_html(
            '<span style="background:{};color:white;padding:4px 12px;border-radius:12px;">{}</span>',
            colores.get(obj.estado, '#999'), obj.get_estado_display()
        )
    estado_display.short_description = "Estado"
    
    actions = ['marcar_entregado', 'marcar_procesando']
    
    def marcar_entregado(self, request, queryset):
        count = 0
        for canje in queryset:
            canje.marcar_entregado(nota="Marcado como entregado desde admin")
            count += 1
        self.message_user(request, f"{count} canje(s) marcado(s) como entregados")
    marcar_entregado.short_description = "Marcar como entregado"
    
    def marcar_procesando(self, request, queryset):
        queryset.update(estado='PROCESANDO')
        self.message_user(request, f"{queryset.count()} canje(s) en procesamiento")
    marcar_procesando.short_description = "Marcar como procesando"


# ========================================
# 🆘 ADMIN UNIFICADO DE SOPORTE (TODO EN UNO)
# ========================================

@admin.register(SolicitudSoporte)
class SolicitudSoporteAdmin(admin.ModelAdmin):
    """
    🆘 GESTIÓN UNIFICADA DE SOPORTE Y PQRS
    Todo en un solo lugar: soporte técnico + peticiones, quejas, reclamos, sugerencias
    """
    change_list_template = 'admin/core/solicitudsoporte/change_list.html'
    change_form_template = 'admin/core/solicitudsoporte/change_form.html'
    list_display = ('tipo_badge', 'estudiante_info', 'asunto_o_keyword', 'categoria_badge', 'estado_badge', 'prioridad_badge', 'fecha_solicitud', 'tiempo_espera', 'atendido_por_info')
    list_filter = ('tipo_solicitud', 'categoria', 'resuelto_por_agente', 'estado', 'prioridad', 'keyword_usada', 'fecha_solicitud')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'mensaje_original', 'asunto', 'respuesta')
    readonly_fields = (
        'fecha_solicitud', 'categoria', 'resuelto_por_agente', 'notas_internas',
        'fecha_respuesta', 'respondido_por',
    )
    list_per_page = 50
    ordering = ('-fecha_solicitud',)
    actions = ['marcar_en_atencion', 'marcar_resuelta', 'marcar_prioridad_alta']
    
    fieldsets = (
        ('📋 Tipo y Clasificación', {
            'fields': ('tipo_solicitud', 'asunto', 'curso_relacionado', 'prioridad')
        }),
        ('📞 Información del Estudiante', {
            'fields': ('estudiante', 'keyword_usada', 'fecha_solicitud')
        }),
        ('💬 Mensaje Original', {
            'fields': ('mensaje_original',)
        }),
        ('🎯 Gestión', {
            'fields': ('estado', 'atendido_por', 'categoria', 'resuelto_por_agente'),
            'description': 'Categoría y resuelto_por_agente los completa el agente PQRS automático.',
        }),
        ('Respuesta', {
            'fields': ('respuesta', 'respuesta_portal', 'fecha_atencion', 'fecha_resolucion', 'fecha_respuesta', 'respondido_por'),
            'description': 'Registro interno. Para notificar al estudiante, use el formulario al final de la página.',
        }),
        ('📋 Notas Internas', {
            'fields': ('notas_internas',),
            'classes': ('collapse',),
            'description': 'Estas notas son privadas y no se muestran al estudiante'
        }),
        ('⭐ Calificación del Estudiante', {
            'fields': ('calificacion', 'comentario_calificacion'),
            'classes': ('collapse',),
        }),
    )
    
    def tipo_badge(self, obj):
        colores = {
            'soporte': ('#f44336', '#ffffff'),
            'peticion': ('#e3f2fd', '#1976d2'),
            'queja': ('#fff3e0', '#e65100'),
            'reclamo': ('#ffebee', '#c62828'),
            'sugerencia': ('#f3e5f5', '#7b1fa2'),
            'felicitacion': ('#e8f5e9', '#2e7d32'),
        }
        iconos = {
            'soporte': '🆘', 'peticion': '📋', 'queja': '😤',
            'reclamo': '📢', 'sugerencia': '💡', 'felicitacion': '🌟',
        }
        bg, color = colores.get(obj.tipo_solicitud, ('#f5f5f5', '#666'))
        icono = iconos.get(obj.tipo_solicitud, '')
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">'
            '{} {}'
            '</span>',
            bg, color, icono, obj.get_tipo_solicitud_display()
        )
    tipo_badge.short_description = "Tipo"
    
    def asunto_o_keyword(self, obj):
        """Muestra asunto si es PQRS o keyword si es soporte"""
        if obj.asunto:
            return obj.asunto[:50]
        if obj.keyword_usada:
            return format_html(
                '<span style="background:#e3f2fd;color:#1565c0;padding:4px 10px;border-radius:12px;font-size:11px;">'
                '🔑 {}'
                '</span>',
                obj.keyword_usada.upper()
            )
        return format_html('<span style="color:#999;">-</span>')
    asunto_o_keyword.short_description = "Asunto / Keyword"
    
    def estudiante_info(self, obj):
        return format_html(
            '<strong>{}</strong><br>'
            '<small style="color:#666;">📱 +{}</small>',
            obj.estudiante.nombre,
            obj.estudiante.telefono
        )
    estudiante_info.short_description = "Estudiante"
    
    def categoria_badge(self, obj):
        if not obj.categoria:
            return format_html('<span style="color:#bbb;">—</span>')
        colores = {
            'acceso': ('#fff8e1', '#f9a825'),
            'contenido': ('#e3f2fd', '#1565c0'),
            'tecnico': ('#fbe9e7', '#bf360c'),
            'otro': ('#f5f5f5', '#666'),
        }
        bg, color = colores.get(obj.categoria, ('#f5f5f5', '#666'))
        marca = '🤖✓' if obj.resuelto_por_agente else '🆘'
        return format_html(
            '<span title="{}" style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">{} {}</span>',
            'Resuelto por agente IA' if obj.resuelto_por_agente else 'Escalado',
            bg, color, marca, obj.get_categoria_display(),
        )
    categoria_badge.short_description = 'Categoría PQRS'

    def estado_badge(self, obj):
        colores = {
            'pendiente': '#ff9800',
            'en_atencion': '#2196f3',
            'resuelta': '#4caf50',
            'cerrada': '#999'
        }
        iconos = {
            'pendiente': '⏳',
            'en_atencion': '👀',
            'resuelta': '✅',
            'cerrada': '🔒'
        }
        return format_html(
            '<span style="background:{};color:white;padding:6px 12px;border-radius:12px;font-weight:600;">'
            '{} {}'
            '</span>',
            colores.get(obj.estado, '#999'),
            iconos.get(obj.estado, ''),
            obj.get_estado_display()
        )
    estado_badge.short_description = "Estado"
    
    def prioridad_badge(self, obj):
        colores = {
            'baja': '#4caf50',
            'media': '#ff9800',
            'alta': '#f44336',
            'critica': '#d32f2f'
        }
        return format_html(
            '<span style="background:{};color:white;padding:6px 12px;border-radius:12px;font-weight:600;">'
            '{}'
            '</span>',
            colores.get(obj.prioridad, '#999'),
            obj.get_prioridad_display()
        )
    prioridad_badge.short_description = "Prioridad"
    
    def tiempo_espera(self, obj):
        """Muestra cuánto tiempo lleva esperando si no ha sido resuelta"""
        if obj.estado in ['resuelta', 'cerrada']:
            return format_html('<span style="color:#4caf50;">✅ Resuelta</span>')
        
        from django.utils import timezone
        tiempo = timezone.now() - obj.fecha_solicitud
        horas = int(tiempo.total_seconds() / 3600)
        
        if horas < 1:
            minutos = int(tiempo.total_seconds() / 60)
            color = '#4caf50'
            texto = f'{minutos} min'
        elif horas < 24:
            color = '#ff9800' if horas > 4 else '#4caf50'
            texto = f'{horas} horas'
        else:
            dias = int(horas / 24)
            color = '#f44336'
            texto = f'{dias} días'
        
        return format_html(
            '<span style="color:{};">⏰ {}</span>',
            color, texto
        )
    tiempo_espera.short_description = "⏰ Tiempo"
    
    def atendido_por_info(self, obj):
        if obj.atendido_por:
            return format_html(
                '<span style="background:#e8f5e9;color:#2e7d32;padding:4px 10px;border-radius:12px;font-size:11px;">'
                '👤 {}'
                '</span>',
                obj.atendido_por
            )
        return format_html('<span style="color:#999;">Sin asignar</span>')
    atendido_por_info.short_description = "Atendido por"
    
    @admin.action(description='👀 Marcar como "En Atención"')
    def marcar_en_atencion(self, request, queryset):
        from django.utils import timezone
        queryset.update(
            estado='en_atencion',
            fecha_atencion=timezone.now()
        )
        self.message_user(request, f"✅ {queryset.count()} solicitud(es) marcada(s) como 'En Atención'")
    
    @admin.action(description='✅ Marcar como "Resuelta"')
    def marcar_resuelta(self, request, queryset):
        from django.utils import timezone
        queryset.update(
            estado='resuelta',
            fecha_resolucion=timezone.now()
        )
        self.message_user(request, f"✅ {queryset.count()} solicitud(es) marcada(s) como 'Resuelta'")
    
    @admin.action(description='🚨 Marcar como "Prioridad Alta"')
    def marcar_prioridad_alta(self, request, queryset):
        queryset.update(prioridad='alta')
        self.message_user(request, f"🚨 {queryset.count()} solicitud(es) marcada(s) como 'Prioridad Alta'")

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        err = request.session.pop('pqrs_envio_error', None)
        if err:
            extra_context['envio_error'] = err
        if request.session.pop('pqrs_envio_ok', False):
            extra_context['envio_ok'] = True

        if request.method == 'POST' and '_enviar_whatsapp' in request.POST and object_id:
            from core.pqrs_respuesta import aplicar_respuesta_pqrs

            obj = self.get_object(request, object_id)
            texto = (
                request.POST.get('respuesta_whatsapp_admin', '').strip()
                or request.POST.get('respuesta', '').strip()
            )
            ok, error = aplicar_respuesta_pqrs(obj, texto, request.user)
            if ok:
                self.message_user(
                    request,
                    f'Respuesta enviada por WhatsApp a {obj.estudiante.nombre}.',
                    messages.SUCCESS,
                )
                request.session['pqrs_envio_ok'] = True
                return redirect(reverse('admin:core_solicitudsoporte_change', args=[obj.pk]))
            request.session['pqrs_envio_error'] = error
            return redirect(reverse('admin:core_solicitudsoporte_change', args=[obj.pk]))

        return super().changeform_view(request, object_id, form_url, extra_context)

    def changelist_view(self, request, extra_context=None):
        """Vista unificada de Soporte y PQRS con estadísticas"""
        extra_context = extra_context or {}
        
        # Estadísticas unificadas
        total = SolicitudSoporte.objects.count()
        pendientes = SolicitudSoporte.objects.filter(estado='pendiente').count()
        en_atencion = SolicitudSoporte.objects.filter(estado='en_atencion').count()
        resueltas = SolicitudSoporte.objects.filter(estado='resuelta').count()
        
        # Por tipo
        soporte_count = SolicitudSoporte.objects.filter(tipo_solicitud='soporte').count()
        pqrs_count = SolicitudSoporte.objects.exclude(tipo_solicitud='soporte').count()
        
        extra_context.update({
            'pqrs_panel_html': format_html('''
                <div style="background:linear-gradient(135deg,#f5f5f5,#e8eaf6);padding:20px;border-radius:12px;margin-bottom:20px;border:1px solid #c5cae9;">
                    <h3 style="margin:0 0 15px 0;color:#283593;">🆘 Panel Unificado: Soporte y PQRS</h3>
                    <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:15px;">
                        <div style="background:#fff;padding:12px 20px;border-radius:8px;border-left:4px solid #ff9800;flex:1;min-width:120px;">
                            <div style="font-size:24px;font-weight:bold;color:#ff9800;">{}</div>
                            <div style="color:#666;font-size:12px;">Pendientes</div>
                        </div>
                        <div style="background:#fff;padding:12px 20px;border-radius:8px;border-left:4px solid #2196f3;flex:1;min-width:120px;">
                            <div style="font-size:24px;font-weight:bold;color:#2196f3;">{}</div>
                            <div style="color:#666;font-size:12px;">En Atención</div>
                        </div>
                        <div style="background:#fff;padding:12px 20px;border-radius:8px;border-left:4px solid #4caf50;flex:1;min-width:120px;">
                            <div style="font-size:24px;font-weight:bold;color:#4caf50;">{}</div>
                            <div style="color:#666;font-size:12px;">Resueltas</div>
                        </div>
                        <div style="background:#fff;padding:12px 20px;border-radius:8px;border-left:4px solid #f44336;flex:1;min-width:120px;">
                            <div style="font-size:24px;font-weight:bold;color:#f44336;">{}</div>
                            <div style="color:#666;font-size:12px;">Soporte Técnico</div>
                        </div>
                        <div style="background:#fff;padding:12px 20px;border-radius:8px;border-left:4px solid #7b1fa2;flex:1;min-width:120px;">
                            <div style="font-size:24px;font-weight:bold;color:#7b1fa2;">{}</div>
                            <div style="color:#666;font-size:12px;">PQRS</div>
                        </div>
                    </div>
                    <p style="color:#666;font-size:12px;margin:0;">Usa los filtros de la derecha para filtrar por tipo (Soporte, Petición, Queja, etc.)</p>
                </div>
            ''', pendientes, en_atencion, resueltas, soporte_count, pqrs_count),
        })
        
        return super().changelist_view(request, extra_context=extra_context)


# ========================================
# 📜 CERTIFICADOS DIGITALES
# ========================================

class PlantillaCertificadoAdmin(admin.ModelAdmin):
    """Plantillas de Certificados — Simplificado: sube imagen a S3"""
    list_display = ('nombre', 'curso_info', 'cliente_info', 'tipo_plantilla', 'por_defecto', 'activa')
    list_filter = ('activa', 'por_defecto', 'cliente', 'curso')
    search_fields = ('nombre', 'descripcion', 'cliente__nombre', 'curso__nombre')
    list_per_page = 50
    
    fieldsets = (
        ('📝 Información Básica', {
            'fields': ('nombre', 'descripcion', 'curso', 'cliente', 'activa', 'por_defecto'),
            'description': mark_safe('''<div style="background:#e3f2fd;padding:12px;border-radius:8px;border-left:4px solid #2196F3;margin:10px 0;">
                <strong>📌 IMPORTANTE:</strong> Selecciona el <strong>Curso</strong> para que el certificado se genere automáticamente al completar ese curso.<br>
                Si no seleccionas curso, se usará solo si está marcada como "Por defecto".
            </div>''')
        }),
        ('🖼️ Imagen del Certificado (S3)', {
            'fields': ('formato_certificado', 'archivo_plantilla_imagen', 'url_plantilla_imagen'),
            'description': mark_safe('''<div style="background:#e8f5e9;padding:15px;border-radius:8px;border-left:4px solid #4CAF50;margin:10px 0;">
                <strong>✅ COMO PREPARAR TU PLANTILLA DE CERTIFICADO</strong><br><br>
                1. Diseña tu certificado en Canva, Word, Photoshop, etc.<br>
                2. Coloca <strong>3 marcadores de color</strong> donde quieras la informacion:<br>
                &nbsp;&nbsp;&nbsp;⬜ <strong>GRIS</strong> (128,128,128) = Donde ira el <strong>NOMBRE</strong> del estudiante<br>
                &nbsp;&nbsp;&nbsp;🟥 <strong>ROJO</strong> (255,0,0) = Donde ira la <strong>CEDULA</strong><br>
                &nbsp;&nbsp;&nbsp;🟦 <strong>AZUL</strong> (0,0,255) = Donde ira el <strong>CODIGO QR</strong><br>
                3. Exporta como <strong>PNG o JPG</strong><br>
                4. Sube la imagen aqui o pega la URL de S3<br><br>
                <strong>💡 Cada curso puede tener su propia plantilla.</strong> Al actualizar la plantilla, los certificados existentes se regeneran automaticamente.<br>
                <strong>⚠️ Al pegar URL:</strong> Asegurate de pegar solo UNA vez la URL completa (https://...).
            </div>''')
        }),
        ('📄 PDF Personalizado (Avanzado)', {
            'fields': ('archivo_plantilla_pdf', 'variable_nombre', 'variable_curso', 'variable_fecha'),
            'classes': ('collapse',),
            'description': mark_safe('''<div style="background:#f5f5f5;padding:10px;border-radius:8px;margin:10px 0;">
                <em>Opcional: sube un PDF con variables {nombre}, {curso}, {fecha}</em>
            </div>''')
        }),
        ('🎨 Diseño eki (Avanzado)', {
            'fields': ('imagen_fondo', 'logo_institucion', 'color_primario', 'color_secundario', 'texto_superior', 'texto_certificado'),
            'classes': ('collapse',),
            'description': 'Solo si no subes imagen ni PDF'
        }),
    )
    
    def get_form(self, request, obj=None, **kwargs):
        """Personaliza el formulario para usar color picker"""
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['color_primario'].widget = ColorPickerWidget()
        form.base_fields['color_secundario'].widget = ColorPickerWidget()
        return form
    
    def curso_info(self, obj):
        """Muestra el curso asignado"""
        if obj.curso:
            return format_html(
                '<span style="background:#c8e6c9;padding:4px 10px;border-radius:12px;font-size:11px;">📚 {}</span>',
                obj.curso.nombre
            )
        return format_html('<span style="color:#999;font-style:italic;">Sin curso (general)</span>')
    curso_info.short_description = "Curso"
    
    def cliente_info(self, obj):
        """Muestra el cliente asociado"""
        if obj.cliente:
            return format_html(
                '<span style="background:#e3f2fd;padding:4px 10px;border-radius:12px;font-size:11px;">🏢 {}</span>',
                obj.cliente.nombre
            )
        return format_html('<span style="color:#999;font-style:italic;">General (eki)</span>')
    cliente_info.short_description = "Cliente"
    
    def tipo_plantilla(self, obj):
        """Muestra si usa PDF personalizado, imagen o diseño eki"""
        if obj.archivo_plantilla_imagen or obj.url_plantilla_imagen:
            label = '🖼️ Imagen' + (' (URL)' if obj.url_plantilla_imagen and not obj.archivo_plantilla_imagen else '')
            return format_html(
                '<span style="background:#f59f00;color:white;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">{}</span>',
                label
            )
        if obj.archivo_plantilla_pdf:
            return format_html(
                '<span style="background:#4CAF50;color:white;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">📄 PDF Personalizado</span>'
            )
        return format_html(
            '<span style="background:#2196F3;color:white;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">🎨 Diseño eki</span>'
        )
    tipo_plantilla.short_description = "Tipo"
    
    def vista_colores(self, obj):
        """Muestra preview de los colores (solo si no usa PDF)"""
        if obj.archivo_plantilla_pdf:
            return format_html('<span style="color:#999;">-</span>')
        
        return format_html(
            '<div style="display:flex;gap:8px;align-items:center;">'
            '<div style="background:{};width:40px;height:40px;border-radius:6px;border:2px solid #ddd;box-shadow:0 2px 4px rgba(0,0,0,0.1);"></div>'
            '<div style="background:{};width:40px;height:40px;border-radius:6px;border:2px solid #ddd;box-shadow:0 2px 4px rgba(0,0,0,0.1);"></div>'
            '</div>',
            obj.color_primario,
            obj.color_secundario
        )
    vista_colores.short_description = "🎨 Colores"
    
    def total_certificados(self, obj):
        """Cuenta certificados generados con esta plantilla"""
        # Por ahora retorna placeholder, puedes implementar el conteo real si agregas FK en Certificado
        return format_html('<span style="color:#999;">-</span>')
    total_certificados.short_description = "Certificados"
    
    def save_model(self, request, obj, form, change):
        """Al guardar plantilla: auto-regenerar certificados existentes que usen esta plantilla"""
        super().save_model(request, obj, form, change)
        if change:  # Solo al editar (no al crear)
            # Buscar certificados que usen esta plantilla y regenerarlos
            from .models_certificados import Certificado
            from .certificado_service import generar_y_guardar_certificado
            import logging
            logger = logging.getLogger(__name__)
            
            # Certificados del curso de esta plantilla
            cert_qs = Certificado.objects.filter(emitido=True)
            if obj.curso:
                cert_qs = cert_qs.filter(curso=obj.curso)
            elif obj.cliente:
                cert_qs = cert_qs.filter(estudiante__cliente=obj.cliente)
            else:
                # Plantilla por defecto — regenerar solo los que NO tienen plantilla especifica
                if obj.por_defecto:
                    from .models_certificados import PlantillaCertificado
                    cursos_con_plantilla = PlantillaCertificado.objects.filter(
                        activa=True, curso__isnull=False
                    ).exclude(pk=obj.pk).values_list('curso_id', flat=True)
                    cert_qs = cert_qs.exclude(curso_id__in=cursos_con_plantilla)
                else:
                    cert_qs = cert_qs.none()
            
            count = cert_qs.count()
            if count > 0 and count <= 50:  # Limite de seguridad
                regenerados = 0
                for cert in cert_qs:
                    try:
                        if generar_y_guardar_certificado(cert, plantilla=obj, force=True):
                            regenerados += 1
                    except Exception as e:
                        logger.error(f"Error regenerando cert {cert.codigo_verificacion}: {e}")
                if regenerados > 0:
                    self.message_user(
                        request,
                        f"🔄 {regenerados} certificado(s) regenerado(s) automaticamente con la nueva plantilla",
                        messages.SUCCESS
                    )
            elif count > 50:
                self.message_user(
                    request,
                    f"⚠️ Hay {count} certificados que usan esta plantilla. Usa la accion 'Regenerar' en Certificados para actualizarlos.",
                    messages.WARNING
                )

    @admin.action(description='👁️ Previsualizar con nombre de prueba')
    def previsualizar_certificado_accion(self, request, queryset):
        """Genera un certificado de prueba para ver cómo se verá"""
        if queryset.count() > 1:
            self.message_user(request, "⚠️ Selecciona solo una plantilla para previsualizar", level=messages.WARNING)
            return
        
        plantilla = queryset.first()
        
        if not plantilla.archivo_plantilla_pdf:
            self.message_user(
                request,
                "⚠️ Esta plantilla no tiene PDF personalizado. Previsualización solo disponible para PDFs personalizados.",
                level=messages.WARNING
            )
            return
        
        # TODO: Implementar generación de certificado de prueba
        # Por ahora solo muestra el PDF
        self.message_user(
            request,
            f"✅ Plantilla: {plantilla.nombre} - Archivo: {plantilla.archivo_plantilla_pdf.name}",
            level=messages.SUCCESS
        )


class CertificadoAdmin(admin.ModelAdmin):
    """
    📜 GESTIÓN UNIFICADA DE CERTIFICADOS
    Desde aquí puedes gestionar certificados y ver plantillas
    """
    list_display = (
        'codigo_verificacion',
        'estudiante_info',
        'curso_info',
        'calificacion_badge',
        'mencion_badge',
        'estado_emision',
        'enviado_whatsapp_badge',
        'plantilla_link',
        'fecha_completado'
    )
    list_filter = ('emitido', 'enviado_whatsapp', 'fecha_completado', 'curso')
    search_fields = ('codigo_verificacion', 'estudiante__nombre', 'estudiante__telefono', 'curso__nombre')
    readonly_fields = (
        'codigo_verificacion',
        'creado_en',
        'actualizado_en',
        'vista_previa_certificado',
        'url_verificacion',
        'duracion_dias'
    )
    list_per_page = 50
    ordering = ('-fecha_emision', '-creado_en')
    
    fieldsets = (
        ('📋 Información', {
            'fields': ('codigo_verificacion', 'estudiante', 'curso', 'url_verificacion')
        }),
        ('🎓 Datos Académicos', {
            'fields': ('calificacion_final', 'fecha_inicio', 'fecha_completado', 'duracion_dias')
        }),
        ('📄 Certificado', {
            'fields': ('emitido', 'fecha_emision', 'archivo_pdf', 'archivo_imagen', 'vista_previa_certificado'),
            'description': '📄 PDF o 🖼️ Imagen del certificado generado'
        }),
        ('📲 Envío', {
            'fields': ('enviado_whatsapp', 'fecha_envio')
        }),
        ('🕐 Timestamps', {
            'fields': ('creado_en', 'actualizado_en'),
            'classes': ('collapse',)
        }),
    )
    
    actions = [
        'generar_certificados', 
        'enviar_por_whatsapp', 
        'regenerar_certificados', 
        'descargar_todos_pdf',
        'descargar_por_cliente',
        'descargar_por_curso',
        'descargar_por_grupo',
        'generar_pdf_consolidado',
        'enviar_email_a_clientes'
    ]
    
    def estudiante_info(self, obj):
        nombre = obj.estudiante.nombre
        telefono = obj.estudiante.telefono
        return format_html(
            '<div style="line-height:1.4;">'
            '<strong>{}</strong><br>'
            '<span style="color:#666;font-size:11px;">📱 {}</span>'
            '</div>',
            nombre, telefono
        )
    estudiante_info.short_description = "👤 Estudiante"
    
    def curso_info(self, obj):
        return format_html(
            '<span style="background:#e3f2fd;color:#1976d2;padding:4px 10px;border-radius:12px;font-size:11px;">'
            '📚 {}'
            '</span>',
            obj.curso.nombre
        )
    curso_info.short_description = "Curso"
    
    def calificacion_badge(self, obj):
        calificacion = float(obj.calificacion_final)
        if calificacion >= 90:
            color = '#4caf50'
            emoji = '🌟'
        elif calificacion >= 80:
            color = '#2196f3'
            emoji = '⭐'
        elif calificacion >= 70:
            color = '#ff9800'
            emoji = '📊'
        else:
            color = '#f44336'
            emoji = '📉'
        
        return format_html(
            '<span style="background:{};color:white;padding:4px 12px;border-radius:12px;font-weight:bold;font-size:12px;">'
            '{} {}%'
            '</span>',
            color, emoji, calificacion
        )
    calificacion_badge.short_description = "📊 Calificación"
    
    def mencion_badge(self, obj):
        mencion = obj.obtener_mencion()
        if mencion:
            return format_html(
                '<span style="background:#ffd700;color:#804000;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:bold;">'
                '🏆 {}'
                '</span>',
                mencion
            )
        return format_html('<span style="color:#999;">-</span>')
    mencion_badge.short_description = "🏆 Mención"
    
    def estado_emision(self, obj):
        if obj.emitido and obj.fecha_emision:
            fecha = obj.fecha_emision.strftime('%d/%m/%Y %H:%M')
            return format_html(
                '<span style="background:#4caf50;color:white;padding:4px 10px;border-radius:12px;font-size:11px;">'
                '✅ Emitido<br>'
                '<span style="font-size:10px;">{}</span>'
                '</span>',
                fecha
            )
        return format_html(
            '<span style="background:#ff9800;color:white;padding:4px 10px;border-radius:12px;font-size:11px;">'
            '⏳ Pendiente'
            '</span>'
        )
    estado_emision.short_description = "Estado"
    
    def enviado_whatsapp_badge(self, obj):
        if obj.enviado_whatsapp and obj.fecha_envio:
            return format_html(
                '<span style="background:#25d366;color:white;padding:4px 10px;border-radius:12px;font-size:11px;">'
                '✅ Enviado'
                '</span>'
            )
        return format_html(
            '<span style="background:#e0e0e0;color:#666;padding:4px 10px;border-radius:12px;font-size:11px;">'
            '📤 No enviado'
            '</span>'
        )
    enviado_whatsapp_badge.short_description = "📲 WhatsApp"
    
    def plantilla_link(self, obj):
        """Link directo para ver plantillas de certificados"""
        count = PlantillaCertificado.objects.filter(cliente=obj.estudiante.cliente).count()
        if count == 0:
            count = PlantillaCertificado.objects.filter(cliente__isnull=True).count()
        
        return format_html(
            '<a href="/admin/core/plantillacertificado/" style="background:#673ab7;color:white;padding:6px 12px;border-radius:12px;text-decoration:none;font-size:11px;font-weight:600;">🎨 {} Plantillas</a>',
            count
        )
    plantilla_link.short_description = "Plantillas"
    
    def url_verificacion(self, obj):
        if obj.codigo_verificacion:
            url = obj.obtener_url_verificacion()
            return format_html(
                '<a href="{}" target="_blank" style="color:#2196f3;">'
                '🔗 Ver certificado público'
                '</a><br>'
                '<code style="font-size:10px;background:#f5f5f5;padding:4px;display:block;margin-top:4px;">{}</code>',
                url, url
            )
        return '-'
    url_verificacion.short_description = "🔗 URL Verificación"
    
    def duracion_dias(self, obj):
        dias = obj.duracion_curso()
        if dias > 0:
            return format_html(
                '<span style="color:#666;">📅 {} día{}</span>',
                dias,
                's' if dias != 1 else ''
            )
        return '-'
    duracion_dias.short_description = "Duración"
    
    def vista_previa_certificado(self, obj):
        if obj.archivo_pdf:
            return format_html(
                '<a href="{}" target="_blank" style="color:#f44336;">'
                '📄 Descargar PDF'
                '</a>',
                obj.archivo_pdf.url
            )
        return format_html('<span style="color:#999;">No generado aún</span>')
    vista_previa_certificado.short_description = "Vista Previa"
    
    @admin.action(description='📄 Generar certificados PDF')
    def generar_certificados(self, request, queryset):
        """Genera los PDFs de los certificados seleccionados"""
        from .certificado_service import generar_y_guardar_certificado
        from django.utils import timezone
        
        generados = 0
        errores = 0
        
        for certificado in queryset:
            try:
                success = generar_y_guardar_certificado(certificado)
                if success:
                    certificado.emitido = True
                    certificado.fecha_emision = timezone.now()
                    certificado.save()
                    generados += 1
                else:
                    errores += 1
            except Exception as e:
                logger.error(f"Error generando certificado {certificado.codigo_verificacion}: {e}")
                errores += 1
        
        if generados > 0:
            self.message_user(
                request,
                f"✅ {generados} certificado(s) generado(s) exitosamente",
                messages.SUCCESS
            )
        if errores > 0:
            self.message_user(
                request,
                f"❌ {errores} error(es) al generar certificados",
                messages.ERROR
            )
    
    @admin.action(description='📲 Enviar certificados por WhatsApp')
    def enviar_por_whatsapp(self, request, queryset):
        """Envía los certificados por WhatsApp"""
        from .certificado_service import enviar_certificado_whatsapp
        
        # Filtrar emitidos con PDF o imagen
        queryset = queryset.filter(emitido=True).filter(
            models.Q(archivo_pdf__isnull=False) | models.Q(archivo_imagen__isnull=False)
        ).exclude(archivo_pdf='', archivo_imagen='')
        
        if not queryset.exists():
            self.message_user(
                request,
                "⚠️ Primero debes generar los certificados (PDF o imagen)",
                messages.WARNING
            )
            return
        
        enviados = 0
        errores = 0
        
        for certificado in queryset:
            try:
                success = enviar_certificado_whatsapp(certificado)
                if success:
                    enviados += 1
                else:
                    errores += 1
            except Exception as e:
                logger.error(f"Error enviando certificado {certificado.codigo_verificacion}: {e}")
                errores += 1
        
        if enviados > 0:
            self.message_user(
                request,
                f"✅ {enviados} certificado(s) enviado(s) por WhatsApp",
                messages.SUCCESS
            )
        if errores > 0:
            self.message_user(
                request,
                f"❌ {errores} error(es) al enviar certificados",
                messages.ERROR
            )
    
    @admin.action(description='🔄 Regenerar certificados PDF')
    def regenerar_certificados(self, request, queryset):
        """Regenera los PDFs (útil si cambió el diseño)"""
        from .certificado_service import generar_y_guardar_certificado
        
        regenerados = 0
        errores = 0
        
        for certificado in queryset:
            try:
                success = generar_y_guardar_certificado(certificado, force=True)
                if success:
                    regenerados += 1
                else:
                    errores += 1
            except Exception as e:
                logger.error(f"Error regenerando certificado {certificado.codigo_verificacion}: {e}")
                errores += 1
        
        if regenerados > 0:
            self.message_user(
                request,
                f"✅ {regenerados} certificado(s) regenerado(s)",
                messages.SUCCESS
            )
        if errores > 0:
            self.message_user(
                request,
                f"❌ {errores} error(es) al regenerar",
                messages.ERROR
            )
    
    @admin.action(description='📥 Descargar todos los certificados en ZIP')
    def descargar_todos_pdf(self, request, queryset):
        """Descarga todos los certificados seleccionados en un archivo ZIP"""
        import zipfile
        from django.http import HttpResponse
        import io
        from django.utils import timezone
        
        # Filtrar solo certificados con PDF
        queryset = queryset.filter(emitido=True, archivo_pdf__isnull=False)
        
        if not queryset.exists():
            self.message_user(
                request,
                "⚠️ No hay certificados con PDF generado",
                messages.WARNING
            )
            return
        
        # Crear archivo ZIP en memoria
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for certificado in queryset:
                try:
                    # Leer el archivo PDF
                    pdf_file = certificado.archivo_pdf.open('rb')
                    pdf_content = pdf_file.read()
                    pdf_file.close()
                    
                    # Nombre del archivo en el ZIP
                    filename = f"{certificado.estudiante.nombre.replace(' ', '_')}_{certificado.curso.nombre.replace(' ', '_')}_{certificado.codigo_verificacion}.pdf"
                    
                    # Agregar al ZIP
                    zip_file.writestr(filename, pdf_content)
                    
                except Exception as e:
                    logger.error(f"Error agregando certificado {certificado.codigo_verificacion} al ZIP: {e}")
        
        # Preparar respuesta HTTP
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="certificados_{timezone.now().strftime("%Y%m%d_%H%M%S")}.zip"'
        
        return response
    
    @admin.action(description='� Descargar por Cliente (ZIP)')
    def descargar_por_cliente(self, request, queryset):
        """Descarga certificados agrupados por cliente en ZIPs separados"""
        import zipfile
        from django.http import HttpResponse
        import io
        from django.utils import timezone
        
        # Filtrar certificados con PDF
        queryset = queryset.filter(emitido=True, archivo_pdf__isnull=False)
        
        if not queryset.exists():
            self.message_user(request, "⚠️ No hay certificados con PDF", messages.WARNING)
            return
        
        # Agrupar por cliente
        certificados_por_cliente = {}
        for cert in queryset:
            cliente = cert.estudiante.cliente
            cliente_nombre = cliente.nombre if cliente else "Sin_Cliente"
            if cliente_nombre not in certificados_por_cliente:
                certificados_por_cliente[cliente_nombre] = []
            certificados_por_cliente[cliente_nombre].append(cert)
        
        # Si es un solo cliente, descargar ZIP directo
        if len(certificados_por_cliente) == 1:
            cliente_nombre = list(certificados_por_cliente.keys())[0]
            certs = certificados_por_cliente[cliente_nombre]
            
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for cert in certs:
                    try:
                        pdf_content = cert.archivo_pdf.read()
                        filename = f"{cert.estudiante.nombre.replace(' ', '_')}_{cert.curso.nombre.replace(' ', '_')}.pdf"
                        zip_file.writestr(filename, pdf_content)
                    except Exception as e:
                        logger.error(f"Error: {e}")
            
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="Certificados_{cliente_nombre}_{timezone.now().strftime("%Y%m%d")}.zip"'
            return response
        
        # Múltiples clientes: crear ZIP maestro con subdirectorios
        master_zip = io.BytesIO()
        with zipfile.ZipFile(master_zip, 'w', zipfile.ZIP_DEFLATED) as zip_master:
            for cliente_nombre, certs in certificados_por_cliente.items():
                for cert in certs:
                    try:
                        pdf_content = cert.archivo_pdf.read()
                        filename = f"{cliente_nombre}/{cert.estudiante.nombre.replace(' ', '_')}_{cert.curso.nombre.replace(' ', '_')}.pdf"
                        zip_master.writestr(filename, pdf_content)
                    except Exception as e:
                        logger.error(f"Error: {e}")
        
        master_zip.seek(0)
        response = HttpResponse(master_zip.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="Certificados_Por_Cliente_{timezone.now().strftime("%Y%m%d")}.zip"'
        
        self.message_user(request, f"✅ Descargados {queryset.count()} certificados de {len(certificados_por_cliente)} cliente(s)", messages.SUCCESS)
        return response
    
    @admin.action(description='📚 Descargar por Curso (ZIP)')
    def descargar_por_curso(self, request, queryset):
        """Descarga certificados agrupados por curso"""
        import zipfile
        from django.http import HttpResponse
        import io
        from django.utils import timezone
        
        queryset = queryset.filter(emitido=True, archivo_pdf__isnull=False)
        
        if not queryset.exists():
            self.message_user(request, "⚠️ No hay certificados con PDF", messages.WARNING)
            return
        
        # Agrupar por curso
        certificados_por_curso = {}
        for cert in queryset:
            curso_nombre = cert.curso.nombre
            if curso_nombre not in certificados_por_curso:
                certificados_por_curso[curso_nombre] = []
            certificados_por_curso[curso_nombre].append(cert)
        
        # Si es un solo curso
        if len(certificados_por_curso) == 1:
            curso_nombre = list(certificados_por_curso.keys())[0]
            certs = certificados_por_curso[curso_nombre]
            
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for cert in certs:
                    try:
                        pdf_content = cert.archivo_pdf.read()
                        filename = f"{cert.estudiante.nombre.replace(' ', '_')}_{cert.codigo_verificacion}.pdf"
                        zip_file.writestr(filename, pdf_content)
                    except Exception as e:
                        logger.error(f"Error: {e}")
            
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="Certificados_{curso_nombre.replace(" ", "_")}_{timezone.now().strftime("%Y%m%d")}.zip"'
            return response
        
        # Múltiples cursos
        master_zip = io.BytesIO()
        with zipfile.ZipFile(master_zip, 'w', zipfile.ZIP_DEFLATED) as zip_master:
            for curso_nombre, certs in certificados_por_curso.items():
                for cert in certs:
                    try:
                        pdf_content = cert.archivo_pdf.read()
                        filename = f"{curso_nombre.replace(' ', '_')}/{cert.estudiante.nombre.replace(' ', '_')}.pdf"
                        zip_master.writestr(filename, pdf_content)
                    except Exception as e:
                        logger.error(f"Error: {e}")
        
        master_zip.seek(0)
        response = HttpResponse(master_zip.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="Certificados_Por_Curso_{timezone.now().strftime("%Y%m%d")}.zip"'
        
        self.message_user(request, f"✅ Descargados {queryset.count()} certificados de {len(certificados_por_curso)} curso(s)", messages.SUCCESS)
        return response
    
    @admin.action(description='👥 Descargar por Grupo (ZIP)')
    def descargar_por_grupo(self, request, queryset):
        """Descarga certificados agrupados por grupo de estudiantes"""
        import zipfile
        from django.http import HttpResponse
        import io
        from django.utils import timezone
        from .models_extras import GrupoEstudiantes
        
        queryset = queryset.filter(emitido=True, archivo_pdf__isnull=False)
        
        if not queryset.exists():
            self.message_user(request, "⚠️ No hay certificados con PDF", messages.WARNING)
            return
        
        # Agrupar por grupo
        certificados_por_grupo = {}
        sin_grupo = []
        
        for cert in queryset:
            estudiante = cert.estudiante
            # Buscar grupos del estudiante
            grupos = GrupoEstudiantes.objects.filter(estudiantes=estudiante)
            
            if grupos.exists():
                for grupo in grupos:
                    grupo_nombre = grupo.nombre
                    if grupo_nombre not in certificados_por_grupo:
                        certificados_por_grupo[grupo_nombre] = []
                    certificados_por_grupo[grupo_nombre].append(cert)
            else:
                sin_grupo.append(cert)
        
        if not certificados_por_grupo:
            self.message_user(request, "⚠️ Los estudiantes seleccionados no pertenecen a ningún grupo", messages.WARNING)
            return
        
        # Crear ZIP
        master_zip = io.BytesIO()
        with zipfile.ZipFile(master_zip, 'w', zipfile.ZIP_DEFLATED) as zip_master:
            # Certificados por grupo
            for grupo_nombre, certs in certificados_por_grupo.items():
                for cert in certs:
                    try:
                        pdf_content = cert.archivo_pdf.read()
                        filename = f"{grupo_nombre.replace(' ', '_')}/{cert.estudiante.nombre.replace(' ', '_')}_{cert.curso.nombre.replace(' ', '_')}.pdf"
                        zip_master.writestr(filename, pdf_content)
                    except Exception as e:
                        logger.error(f"Error: {e}")
            
            # Certificados sin grupo
            if sin_grupo:
                for cert in sin_grupo:
                    try:
                        pdf_content = cert.archivo_pdf.read()
                        filename = f"Sin_Grupo/{cert.estudiante.nombre.replace(' ', '_')}_{cert.curso.nombre.replace(' ', '_')}.pdf"
                        zip_master.writestr(filename, pdf_content)
                    except Exception as e:
                        logger.error(f"Error: {e}")
        
        master_zip.seek(0)
        response = HttpResponse(master_zip.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="Certificados_Por_Grupo_{timezone.now().strftime("%Y%m%d")}.zip"'
        
        self.message_user(request, f"✅ Descargados {queryset.count()} certificados de {len(certificados_por_grupo)} grupo(s)", messages.SUCCESS)
        return response
    
    @admin.action(description='📄 Generar PDF Consolidado (Un archivo)')
    def generar_pdf_consolidado(self, request, queryset):
        """Genera un único PDF con todos los certificados seleccionados"""
        from PyPDF2 import PdfMerger
        from django.http import HttpResponse
        import io
        from django.utils import timezone
        
        queryset = queryset.filter(emitido=True, archivo_pdf__isnull=False)
        
        if not queryset.exists():
            self.message_user(request, "⚠️ No hay certificados con PDF", messages.WARNING)
            return
        
        # Crear merger
        merger = PdfMerger()
        
        # Agregar cada certificado
        for cert in queryset:
            try:
                pdf_file = cert.archivo_pdf.open('rb')
                merger.append(pdf_file)
                pdf_file.close()
            except Exception as e:
                logger.error(f"Error agregando certificado {cert.codigo_verificacion}: {e}")
        
        # Guardar PDF consolidado
        output = io.BytesIO()
        merger.write(output)
        merger.close()
        output.seek(0)
        
        # Preparar respuesta
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Certificados_Consolidados_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        self.message_user(request, f"✅ Generado PDF consolidado con {queryset.count()} certificado(s)", messages.SUCCESS)
        return response
    
    @admin.action(description='�📧 Enviar certificados por email a clientes')
    def enviar_email_a_clientes(self, request, queryset):
        """Envía los certificados por email al cliente (no al estudiante) usando API"""
        from .email_service import enviar_certificados_a_cliente
        
        # Filtrar solo certificados con PDF
        queryset = queryset.filter(emitido=True, archivo_pdf__isnull=False)
        
        if not queryset.exists():
            self.message_user(
                request,
                "⚠️ No hay certificados con PDF generado",
                messages.WARNING
            )
            return
        
        # Agrupar certificados por cliente
        certificados_por_cliente = {}
        sin_cliente = []
        
        for certificado in queryset:
            estudiante = certificado.estudiante
            cliente = estudiante.cliente
            
            if cliente and cliente.enviar_certificados_email:
                if cliente.id not in certificados_por_cliente:
                    certificados_por_cliente[cliente.id] = {
                        'cliente': cliente,
                        'certificados': []
                    }
                certificados_por_cliente[cliente.id]['certificados'].append(certificado)
            else:
                sin_cliente.append(certificado)
        
        enviados = 0
        errores = 0
        
        # Enviar emails agrupados por cliente usando el servicio
        for cliente_id, data in certificados_por_cliente.items():
            cliente = data['cliente']
            certificados = data['certificados']
            
            try:
                success = enviar_certificados_a_cliente(cliente, certificados)
                if success:
                    enviados += len(certificados)
                else:
                    errores += len(certificados)
                    
            except Exception as e:
                logger.error(f"❌ Error enviando certificados a {cliente.email}: {e}")
                errores += len(certificados)
        
        # Mensaje de resultado
        if enviados > 0:
            self.message_user(
                request,
                f"✅ {enviados} certificado(s) enviado(s) por email a {len(certificados_por_cliente)} cliente(s)",
                messages.SUCCESS
            )
        
        if sin_cliente:
            self.message_user(
                request,
                f"⚠️ {len(sin_cliente)} certificado(s) sin cliente o con envío deshabilitado",
                messages.WARNING
            )
        
        if errores > 0:
            self.message_user(
                request,
                f"❌ {errores} error(es) al enviar emails",
                messages.ERROR
            )


# ========================================
# 🔐 AUDITORÍA - AUDIT LOG
# ========================================

class AuditLogAdmin(admin.ModelAdmin):
    """Admin para registro de auditoría de certificados"""
    
    list_display = ('get_resumen_corto', 'accion_badge', 'get_exitoso_display', 'fecha_accion', 'usuario')
    list_filter = ('accion', 'exitoso', 'fecha_accion')
    search_fields = ('certificado_codigo', 'estudiante_nombre', 'curso_nombre', 'ip_address')
    readonly_fields = ('fecha_accion', 'ip_address')
    date_hierarchy = 'fecha_accion'
    ordering = ('-fecha_accion',)
    
    fieldsets = (
        ('📋 Información General', {
            'fields': ('accion', 'exitoso', 'usuario', 'fecha_accion')
        }),
        ('📜 Certificado & Estudiante', {
            'fields': ('certificado_codigo', 'estudiante_nombre', 'curso_nombre')
        }),
        ('📝 Detalles', {
            'fields': ('descripcion', 'mensaje_error'),
            'classes': ('collapse',)
        }),
        ('🔍 Auditoría Técnica', {
            'fields': ('ip_address',),
            'classes': ('collapse',)
        }),
    )
    
    def get_resumen_corto(self, obj):
        """Muestra resumen en lista"""
        return f"{obj.get_accion_display()}: {obj.estudiante_nombre or 'Sistema'}"
    get_resumen_corto.short_description = 'Acción'
    
    def accion_badge(self, obj):
        """Muestra acción con badge de color"""
        from django.utils.html import format_html
        
        color_map = {
            'GENERAR': '#28a745',  # Verde
            'ENVIAR': '#17a2b8',   # Azul
            'DESCARGAR': '#007bff', # Azul claro
            'VERIFICAR': '#6c757d',  # Gris
            'MODIFICAR': '#ffc107',  # Amarillo
            'ELIMINAR': '#dc3545',   # Rojo
            'REGENERAR': '#fd7e14',  # Naranja
            'AUTO_GENERAR': '#20c997', # Verde agua
            'ERROR': '#dc3545',      # Rojo
        }
        
        color = color_map.get(obj.accion, '#6c757d')
        return format_html(
            '<span style="background:{};color:white;padding:3px 8px;border-radius:3px;font-weight:bold;">{}</span>',
            color,
            obj.get_accion_display()
        )
    accion_badge.short_description = 'Acción'
    
    def get_exitoso_display(self, obj):
        """Muestra estado de éxito con icono"""
        from django.utils.html import format_html
        return format_html(
            '<span style="font-size: 18px;">{}</span>',
            '✅' if obj.exitoso else '❌'
        )
    get_exitoso_display.short_description = 'Estado'


# ========================================
# 🎯 CARGAR PERSONALIZACIÓN DEL DASHBOARD
# ========================================

# Importar y ejecutar la personalización del admin
try:
    from .admin_dashboard import setup_custom_admin_dashboard
    setup_custom_admin_dashboard()
except Exception as e:
    logger.warning(f"No se pudo cargar la personalización del dashboard: {str(e)}")


# ========================================
# 👥 ADMIN UNIFICADO DE GRUPOS (TODO EN UNO)
# ========================================

@admin.register(GrupoEstudiantes)
class GrupoEstudiantesAdmin(admin.ModelAdmin):
    """
    📦 GESTIÓN UNIFICADA DE GRUPOS
    Los miembros se gestionan en «Gestionar miembros» (lista/Excel), no con el selector doble.
    """
    list_display = (
        'nombre_completo', 'cliente_nombre', 'cantidad_estudiantes',
        'gestionar_miembros_link', 'cursos_asociados_display',
        'whatsapp_grupos_link', 'invitaciones_link', 'fecha_creacion',
    )
    list_filter = ('cliente', 'activo', 'fecha_creacion')
    search_fields = ('nombre', 'descripcion', 'cliente__nombre')
    filter_horizontal = ('cursos',)
    exclude = ('estudiantes',)
    readonly_fields = ('panel_gestion_miembros',)
    autocomplete_fields = ('cliente',)
    actions = ['crear_grupo_whatsapp', 'enviar_invitaciones']

    fieldsets = (
        ('📋 Información del Grupo', {
            'fields': ('nombre', 'emoji', 'descripcion', 'cliente', 'activo'),
        }),
        ('👥 Miembros del grupo', {
            'fields': ('panel_gestion_miembros',),
            'description': (
                'Use «Gestionar miembros» para pegar cédulas/teléfonos o subir Excel. '
                'Para muchos estudiantes: lista Estudiantes → marcar → acción «Asignar a un grupo».'
            ),
        }),
        ('📚 Cursos asociados (opcional)', {
            'fields': ('cursos',),
            'classes': ('collapse',),
        }),
    )

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom = [
            path(
                '<path:object_id>/gestionar-miembros/',
                self.admin_site.admin_view(self.gestionar_miembros_view),
                name='core_grupoestudiantes_gestionar_miembros',
            ),
        ]
        return custom + urls

    def gestionar_miembros_view(self, request, object_id):
        from django.contrib import messages
        from django.shortcuts import get_object_or_404, render

        from core.grupos_miembros import (
            agregar_miembros_por_identificadores,
            leer_identificadores_desde_excel,
            parsear_lineas_identificadores,
            quitar_miembros_por_identificadores,
        )

        grupo = get_object_or_404(GrupoEstudiantes, pk=object_id)

        if request.method == 'POST':
            accion = request.POST.get('accion')
            modo = request.POST.get('modo_busqueda', 'auto')
            ids = parsear_lineas_identificadores(request.POST.get('lista_identificadores', ''))

            archivo = request.FILES.get('archivo_excel')
            if archivo and accion == 'agregar':
                try:
                    ids_excel = leer_identificadores_desde_excel(archivo, columna='cedula')
                    if not ids_excel:
                        ids_excel = leer_identificadores_desde_excel(archivo, columna='telefono')
                    ids = list(dict.fromkeys(ids + ids_excel))
                except Exception as exc:
                    messages.error(request, f'No se pudo leer el Excel: {exc}')

            if not ids:
                messages.warning(request, 'No ingresó cédulas, teléfonos ni archivo válido.')
            elif accion == 'agregar':
                r = agregar_miembros_por_identificadores(grupo, ids, modo=modo)
                messages.success(
                    request,
                    f'Agregados: {r["agregados"]}. Ya estaban: {r["ya_estaban"]}. '
                    f'No encontrados: {len(r["no_encontrados"])}.',
                )
                if r['no_encontrados'][:5]:
                    messages.warning(
                        request,
                        'Ejemplos no encontrados: ' + ', '.join(r['no_encontrados'][:5]),
                    )
            elif accion == 'quitar':
                r = quitar_miembros_por_identificadores(grupo, ids, modo=modo)
                messages.success(
                    request,
                    f'Quitados: {r["quitados"]}. No estaban en el grupo: {len(r["no_en_grupo"])}. '
                    f'No encontrados: {len(r["no_encontrados"])}.',
                )

        total = grupo.estudiantes.count()
        miembros = list(
            grupo.estudiantes.order_by('nombre').values('cedula', 'nombre', 'telefono')[:50]
        )
        ctx = {
            'title': f'Gestionar miembros — {grupo.emoji} {grupo.nombre}',
            'grupo': grupo,
            'total_miembros': total,
            'miembros_muestra': miembros,
            'opts': self.model._meta,
        }
        return render(request, 'admin/grupo_gestionar_miembros.html', ctx)

    def panel_gestion_miembros(self, obj):
        if not obj or not obj.pk:
            return format_html(
                '<p style="color:#666;">Guarde el grupo primero; luego podrá agregar miembros.</p>'
            )
        n = obj.estudiantes.count()
        url = reverse('admin:core_grupoestudiantes_gestionar_miembros', args=[obj.pk])
        lista_url = (
            f'{reverse("admin:core_estudiante_changelist")}?grupos__id__exact={obj.pk}'
        )
        return format_html(
            '<p><strong>{}</strong> estudiante(s) en este grupo.</p>'
            '<p><a href="{}" class="button" style="background:#1976d2;color:#fff!important;'
            'padding:8px 16px;border-radius:4px;text-decoration:none;margin-right:8px;">'
            'Gestionar miembros (lista / Excel)</a>'
            '<a href="{}" style="margin-left:4px;">Ver en lista de estudiantes</a></p>'
            '<p style="color:#666;font-size:12px;margin-top:12px;">'
            'También: Admin → Estudiantes → filtrar → marcar → acción «Asignar estudiantes a un grupo».</p>',
            n, url, lista_url,
        )
    panel_gestion_miembros.short_description = 'Miembros'

    def gestionar_miembros_link(self, obj):
        if not obj.pk:
            return '-'
        url = reverse('admin:core_grupoestudiantes_gestionar_miembros', args=[obj.pk])
        return format_html(
            '<a href="{}" style="background:#1976d2;color:#fff;padding:5px 10px;'
            'border-radius:6px;text-decoration:none;font-size:11px;">Gestionar</a>',
            url,
        )
    gestionar_miembros_link.short_description = 'Miembros'

    def nombre_completo(self, obj):
        """Muestra el nombre con emoji"""
        return format_html(
            '<span style="font-size:14px;">{} <strong>{}</strong></span>',
            obj.emoji, obj.nombre
        )
    nombre_completo.short_description = "Grupo"
    
    def cliente_nombre(self, obj):
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;">Sin cliente</span>')
    cliente_nombre.short_description = "🏢 Cliente"
    
    def cantidad_estudiantes(self, obj):
        count = obj.estudiantes.count()
        return format_html(
            '<span style="background:#e3f2fd;padding:4px 10px;border-radius:12px;font-size:11px;">{} estudiante{}</span>',
            count, 's' if count != 1 else ''
        )
    cantidad_estudiantes.short_description = "👥 Estudiantes"
    
    def cursos_asociados_display(self, obj):
        cursos = obj.cursos.all()
        if not cursos:
            return format_html('<span style="color:#999;">Sin cursos</span>')
        
        badges = []
        for curso in cursos[:2]:
            badges.append(f'<span style="background:#fff3e0;color:#e65100;padding:3px 8px;border-radius:8px;font-size:10px;margin-right:3px;">{curso.nombre}</span>')
        
        html = ''.join(badges)
        if cursos.count() > 2:
            html += f' <span style="color:#999;font-size:10px;">+{cursos.count() - 2} más</span>'
        
        return format_html(html)
    cursos_asociados_display.short_description = "📚 Cursos"
    
    def whatsapp_grupos_link(self, obj):
        """Link directo para gestionar grupos de WhatsApp"""
        cursos_ids = obj.cursos.values_list('id', flat=True)
        count = GrupoWhatsApp.objects.filter(curso_id__in=cursos_ids).count()
        
        return format_html(
            '<a href="/admin/core/grupowhatsapp/?curso__id__in={}" style="background:#25d366;color:white;padding:6px 12px;border-radius:12px;text-decoration:none;font-size:11px;font-weight:600;">💬 {} WhatsApp</a>',
            ','.join(map(str, cursos_ids)) if cursos_ids else '0',
            count
        )
    whatsapp_grupos_link.short_description = "Grupos WhatsApp"
    
    def invitaciones_link(self, obj):
        """Link directo para gestionar invitaciones"""
        estudiantes_ids = obj.estudiantes.values_list('id', flat=True)
        count = InvitacionGrupo.objects.filter(estudiante_id__in=estudiantes_ids).count()
        
        return format_html(
            '<a href="/admin/core/invitaciongrupo/?estudiante__id__in={}" style="background:#2196f3;color:white;padding:6px 12px;border-radius:12px;text-decoration:none;font-size:11px;font-weight:600;">✉️ {} Invitaciones</a>',
            ','.join(map(str, estudiantes_ids)) if estudiantes_ids else '0',
            count
        )
    invitaciones_link.short_description = "Invitaciones"
    
    def crear_grupo_whatsapp(self, request, queryset):
        """Acción rápida: Crear grupo de WhatsApp"""
        from django.shortcuts import redirect
        if queryset.count() == 1:
            grupo = queryset.first()
            # Redirigir a crear grupo WhatsApp con el curso preseleccionado
            return redirect(f'/admin/core/grupowhatsapp/add/')
        else:
            self.message_user(request, "Selecciona solo un grupo para crear el grupo de WhatsApp", level='warning')
    crear_grupo_whatsapp.short_description = "➕ Crear grupo de WhatsApp para este grupo"
    
    def enviar_invitaciones(self, request, queryset):
        """Acción rápida: Enviar invitaciones"""
        from django.shortcuts import redirect
        return redirect('/admin/core/invitaciongrupo/add/')
    enviar_invitaciones.short_description = "✉️ Enviar invitaciones a estudiantes"
    enviar_invitaciones.short_description = "✉️ Enviar invitaciones a estudiantes"


@admin.register(EnvioProgramado)
class EnvioProgramadoAdmin(admin.ModelAdmin):
    """Gestión de envíos programados"""
    form = EnvioProgramadoForm
    list_display = ('nombre', 'tipo', 'fecha_programada', 'estado_badge', 'fecha_envio_real')
    list_filter = ('estado', 'tipo', 'fecha_programada')
    search_fields = ('nombre', 'mensaje')
    readonly_fields = ('fecha_envio_real', 'total_destinatarios', 'total_enviados', 'total_fallidos')
    
    fieldsets = (
        ('📋 Información Básica', {
            'fields': ('nombre', 'tipo', 'campana', 'grupo', 'estudiante', 'mensaje')
        }),
        ('📎 Multimedia (Opcional)', {
            'fields': ('incluir_media', 'media_url', 'media_file_upload'),
            'classes': ('collapse',)
        }),
        ('📅 Programación', {
            'fields': ('fecha_programada', 'estado')
        }),
        ('📊 Resultados', {
            'fields': ('fecha_envio_real', 'total_destinatarios', 'total_enviados', 'total_fallidos'),
            'classes': ('collapse',)
        }),
    )
    
    def estado_badge(self, obj):
        colores = {
            'pendiente': ('#fff3e0', '#e65100'),
            'enviando': ('#e3f2fd', '#1976d2'),
            'enviado': ('#e8f5e9', '#2e7d32'),
            'fallido': ('#ffebee', '#c62828'),
            'cancelado': ('#f5f5f5', '#666'),
        }
        bg, color = colores.get(obj.estado, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;">{}</span>',
            bg, color, obj.get_estado_display()
        )
    estado_badge.short_description = "Estado"


class PQRSAdmin(admin.ModelAdmin):
    """Gestión auxiliar de PQRS (Ver principalmente desde Solicitudes de Soporte)"""
    change_list_template = 'admin/core/pqrs/change_list.html'
    list_display = ('tipo_badge', 'estudiante', 'asunto', 'prioridad_badge', 'estado_badge', 'soporte_link', 'fecha_creacion')
    list_filter = ('tipo', 'estado', 'prioridad', 'fecha_creacion')
    search_fields = ('asunto', 'descripcion', 'estudiante__nombre')
    readonly_fields = ('fecha_creacion', 'fecha_respuesta', 'fecha_cierre')
    actions = ['crear_solicitud_soporte']
    
    fieldsets = (
        ('📝 Información', {
            'fields': ('tipo', 'estudiante', 'asunto', 'descripcion', 'curso_relacionado', 'prioridad')
        }),
        ('📊 Estado y Respuesta', {
            'fields': ('estado', 'respuesta', 'notas_internas', 'atendido_por')
        }),
        ('📅 Fechas', {
            'fields': ('fecha_creacion', 'fecha_respuesta', 'fecha_cierre'),
            'classes': ('collapse',)
        }),
        ('⭐ Calificación', {
            'fields': ('calificacion', 'comentario_calificacion'),
            'classes': ('collapse',)
        }),
    )
    
    def tipo_badge(self, obj):
        colores = {
            'peticion': ('#e3f2fd', '#1976d2'),
            'queja': ('#fff3e0', '#e65100'),
            'reclamo': ('#ffebee', '#c62828'),
            'sugerencia': ('#f3e5f5', '#7b1fa2'),
            'felicitacion': ('#e8f5e9', '#2e7d32'),
        }
        bg, color = colores.get(obj.tipo, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;">{}</span>',
            bg, color, obj.get_tipo_display()
        )
    tipo_badge.short_description = "Tipo"
    
    def prioridad_badge(self, obj):
        colores = {
            'baja': ('#f5f5f5', '#666'),
            'media': ('#fff3e0', '#e65100'),
            'alta': ('#ffebee', '#c62828'),
            'urgente': ('#b71c1c', '#ffffff'),
        }
        bg, color = colores.get(obj.prioridad, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;">{}</span>',
            bg, color, obj.get_prioridad_display()
        )
    prioridad_badge.short_description = "Prioridad"
    
    def estado_badge(self, obj):
        colores = {
            'pendiente': ('#fff3e0', '#e65100'),
            'en_proceso': ('#e3f2fd', '#1976d2'),
            'resuelto': ('#e8f5e9', '#2e7d32'),
            'cerrado': ('#f5f5f5', '#666'),
            'rechazado': ('#ffebee', '#c62828'),
        }
        bg, color = colores.get(obj.estado, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;">{}</span>',
            bg, color, obj.get_estado_display()
        )
    estado_badge.short_description = "Estado"
    
    def soporte_link(self, obj):
        """Link para ver solicitudes de soporte del estudiante"""
        count = SolicitudSoporte.objects.filter(estudiante=obj.estudiante).count()
        
        return format_html(
            '<a href="/admin/core/solicitudsoporte/?estudiante__id__exact={}" style="background:#f44336;color:white;padding:6px 12px;border-radius:12px;text-decoration:none;font-size:11px;font-weight:600;">🆘 {} Soporte</a>',
            obj.estudiante.id,
            count
        )
    soporte_link.short_description = "Solicitudes"
    
    def crear_solicitud_soporte(self, request, queryset):
        """Crear solicitud de soporte desde PQRS urgente"""
        from django.shortcuts import redirect
        if queryset.count() == 1:
            pqrs = queryset.first()
            # Redirigir para crear solicitud de soporte
            return redirect(f'/admin/core/solicitudsoporte/add/?estudiante={pqrs.estudiante.id}')
        else:
            self.message_user(request, "Selecciona solo una PQRS para crear solicitud", level='warning')
    crear_solicitud_soporte.short_description = "🆘 Crear solicitud de soporte urgente"

    def changelist_view(self, request, extra_context=None):
        """Agrega link de vuelta al panel unificado de Soporte y PQRS"""
        extra_context = extra_context or {}
        extra_context['title'] = '📮 PQRS — Parte del panel Soporte y PQRS'
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(ArchivoModulo)
class ArchivoModuloAdmin(admin.ModelAdmin):
    """Gestión de archivos multimedia de módulos"""
    list_display = ('modulo', 'tipo_badge', 'titulo', 'disponible_offline', 'descargar_btn', 'activo')
    list_filter = ('tipo', 'disponible_offline', 'activo', 'modulo__curso')
    search_fields = ('titulo', 'descripcion', 'modulo__titulo')
    readonly_fields = ('preview_archivo', 'fecha_creacion')
    
    fieldsets = (
        ('📁 Información del Archivo', {
            'fields': ('modulo', 'tipo', 'titulo', 'descripcion', 'orden')
        }),
        ('📎 Archivo', {
            'fields': ('archivo', 'preview_archivo', 'url_externa')
        }),
        ('⚙️ Configuración', {
            'fields': ('disponible_offline', 'activo', 'fecha_creacion')
        }),
    )
    
    actions = ['enviar_video_whatsapp_action', 'marcar_disponible_offline', 'marcar_no_disponible_offline']

    @admin.action(description='📤 Enviar video por WhatsApp')
    def enviar_video_whatsapp_action(self, request, queryset):
        from .whatsapp_service import enviar_video_whatsapp
        from django.contrib import messages
        enviados = 0
        errores = 0
        for archivo in queryset:
            if archivo.tipo == 'video' and archivo.url_externa:
                # Solicitar número destino
                numero = request.POST.get('numero_destino')
                if not numero:
                    self.message_user(request, 'Debes ingresar el número destino en formato internacional (+57...).', level='error')
                    errores += 1
                    continue
                try:
                    enviar_video_whatsapp(numero, archivo.url_externa)
                    enviados += 1
                except Exception as e:
                    self.message_user(request, f'Error enviando video: {e}', level='error')
                    errores += 1
            else:
                self.message_user(request, 'Solo se pueden enviar archivos de tipo video con URL externa.', level='error')
                errores += 1
        if enviados:
            self.message_user(request, f'{enviados} video(s) enviados correctamente por WhatsApp.')
        if errores:
            self.message_user(request, f'{errores} error(es) al intentar enviar videos.', level='error')
    
    def tipo_badge(self, obj):
        colores = {
            'video': ('#f3e5f5', '#7b1fa2'),
            'imagen': ('#e3f2fd', '#1976d2'),
            'infografia': ('#e8f5e9', '#2e7d32'),
            'pdf': ('#ffebee', '#c62828'),
            'audio': ('#fff3e0', '#e65100'),
        }
        bg, color = colores.get(obj.tipo, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;">{}</span>',
            bg, color, obj.get_tipo_display()
        )
    tipo_badge.short_description = "Tipo"
    
    def tamanio_display(self, obj):
        if obj.tamano_bytes:
            mb = obj.tamano_bytes / (1024 * 1024)
            return f"{mb:.2f} MB"
        return "N/A"
    tamanio_display.short_description = "Tamaño"
    
    def preview_archivo(self, obj):
        url = obj.get_url_para_envio()
        if not url:
            return format_html('<span style="color:#999;">Sin archivo</span>')
        if obj.tipo == 'imagen':
            return format_html(
                '<img src="{}" style="max-width:400px;max-height:300px;border-radius:8px;border:2px solid #ddd;" />',
                url
            )
        elif obj.tipo == 'video':
            return format_html(
                '<video controls style="max-width:500px;border-radius:8px;"><source src="{}" type="video/mp4"></video>',
                url
            )
        elif obj.tipo == 'pdf':
            return format_html(
                '<a href="{}" target="_blank" class="button" style="background:#dc2626;color:white;padding:10px 20px;text-decoration:none;border-radius:6px;">📄 Abrir PDF</a>',
                url
            )
        else:
            return format_html(
                '<a href="{}" target="_blank">Ver archivo</a>',
                url
            )
    preview_archivo.short_description = "Vista Previa"
    
    def descargar_btn(self, obj):
        if obj.archivo:
            return format_html(
                '<a href="{}" download class="button" style="background:#2563eb;color:white;padding:6px 14px;text-decoration:none;border-radius:6px;font-size:12px;">⬇️ Descargar</a>',
                obj.archivo.url
            )
        return '-'
    descargar_btn.short_description = "Descarga"
    
    def marcar_disponible_offline(self, request, queryset):
        updated = queryset.update(disponible_offline=True)
        self.message_user(request, f'{updated} archivo(s) marcado(s) como disponibles offline')
    marcar_disponible_offline.short_description = "✅ Marcar disponible offline"
    
    def marcar_no_disponible_offline(self, request, queryset):
        updated = queryset.update(disponible_offline=False)
        self.message_user(request, f'{updated} archivo(s) marcado(s) como NO disponibles offline')
    marcar_no_disponible_offline.short_description = "❌ Desactivar descarga offline"


# ⚠️ MODELOS AUXILIARES - NO APARECEN EN EL MENÚ PRINCIPAL
# Estos modelos se gestionan desde "Grupos de Estudiantes"
# Pero los registramos para que estén disponibles via acciones/links

@admin.register(GrupoWhatsApp)
class GrupoWhatsAppAdmin(admin.ModelAdmin):
    """Gestión auxiliar de grupos de WhatsApp (acceso desde Grupos de Estudiantes)"""
    list_display = ('nombre', 'curso', 'link_invitacion_corto', 'ocupacion_badge', 'activo', 'fecha_creacion')
    list_filter = ('activo', 'curso', 'fecha_creacion')
    search_fields = ('nombre', 'descripcion')
    
    def link_invitacion_corto(self, obj):
        if obj.link_invitacion:
            return format_html(
                '<a href="{}" target="_blank" style="color:#25d366;">🔗 Abrir</a>',
                obj.link_invitacion
            )
        return '-'
    link_invitacion_corto.short_description = "Link"
    
    def ocupacion_badge(self, obj):
        porcentaje = obj.porcentaje_ocupacion()
        color = '#4caf50' if porcentaje < 70 else '#ff9800' if porcentaje < 90 else '#f44336'
        return format_html(
            '<span style="background:{};color:white;padding:4px 10px;border-radius:12px;font-size:11px;">{}/{} ({}%)</span>',
            color, obj.miembros_actuales, obj.capacidad_maxima, porcentaje
        )
    ocupacion_badge.short_description = "Ocupación"
    
    class Meta:
        verbose_name = 'Grupo de WhatsApp'
        verbose_name_plural = 'Grupos de WhatsApp (Ver desde Grupos)'

@admin.register(InvitacionGrupo)
class InvitacionGrupoAdmin(admin.ModelAdmin):
    """Gestión auxiliar de invitaciones (acceso desde Grupos de Estudiantes)"""
    list_display = ('estudiante', 'grupo', 'estado_badge', 'fecha_envio', 'fecha_respuesta')
    list_filter = ('estado', 'fecha_envio', 'grupo')
    search_fields = ('estudiante__nombre', 'grupo__nombre')
    readonly_fields = ('fecha_envio', 'fecha_respuesta', 'fecha_creacion')
    
    def estado_badge(self, obj):
        colores = {
            'pendiente': ('#fff3e0', '#e65100'),
            'enviada': ('#e3f2fd', '#1976d2'),
            'aceptada': ('#e8f5e9', '#2e7d32'),
            'rechazada': ('#ffebee', '#c62828'),
            'expirada': ('#f5f5f5', '#666'),
        }
        return format_html(
            '<span style="background:{};color:white;padding:6px 12px;border-radius:12px;font-weight:600;">{}</span>',
            colores.get(obj.estado, '#999'),
            obj.get_estado_display()
        )
    estado_badge.short_description = "Estado"
    
    class Meta:
        verbose_name = 'Invitación a Grupo'
        verbose_name_plural = 'Invitaciones (Ver desde Grupos)'
    
    def estado_badge(self, obj):
        colores = {
            'pendiente': ('#fff3e0', '#e65100'),
            'enviada': ('#e3f2fd', '#1976d2'),
            'aceptada': ('#e8f5e9', '#2e7d32'),
            'rechazada': ('#ffebee', '#c62828'),
            'expirada': ('#f5f5f5', '#666'),
        }
        bg, color = colores.get(obj.estado, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;">{}</span>',
            bg, color, obj.get_estado_display()
        )
    estado_badge.short_description = "Estado"


# ================================================
# 📊 DASHBOARD DE PLANTILLAS Y TEMPLATES
# ================================================

class PlantillaDashboardAdmin(admin.ModelAdmin):
    """Dashboard para ver el estado de plantillas y Content Templates"""
    list_display = ('nombre_con_emoji', 'categoria_display', 'estado_template', 'activa', 'veces_usada', 'fecha_modificacion')
    list_filter = ('activa', 'categoria', 'aprobada_twilio', 'fecha_modificacion')
    search_fields = ('nombre_interno', 'cuerpo_mensaje', 'twilio_template_sid')
    list_per_page = 50

    fieldsets = (
        ('📝 Información Básica', {
            'fields': ('nombre_interno', 'categoria', 'activa'),
            'description': 'Configura el nombre y categoría de la plantilla'
        }),
        ('🎨 Personalización Visual', {
            'fields': ('emoji',),
            'description': mark_safe('''<div style="background:#f5f5f5;padding:15px;border-radius:8px;margin-top:10px;">
                <strong>💡 El emoji se autocompletará según la categoría seleccionada arriba</strong><br><br>
                <strong>Categorías disponibles:</strong><br>
                🌾 Cultivos • 🐄 Ganadería • 🌱 General Agrícola • 📚 Educación • 💼 Gestión<br><br>
                <em>Puedes cambiar el emoji manualmente si lo deseas</em>
            </div>''')
        }),
        ('📄 Contenido del Mensaje', {
            'fields': ('cuerpo_mensaje',),
            'description': 'Escribe el mensaje. Usa {nombre} para personalizar con el nombre del estudiante.'
        }),
        ('🔗 Twilio Content Template (Opcional)', {
            'fields': ('twilio_template_sid', 'twilio_template_nombre', 'aprobada_twilio'),
            'description': mark_safe('''<div style="background:#fff3e0;padding:15px;border-radius:8px;border-left:4px solid #ff9800;">
                <strong>⚠️ IMPORTANTE: Para campañas con Content Templates</strong><br><br>
                <strong>Flujo correcto:</strong>
                <ol style="margin:10px 0;">
                    <li>🔵 Ve a <a href="https://console.twilio.com/us1/develop/sms/content-editor" target="_blank" style="color:#2196F3;">Twilio Content Editor</a></li>
                    <li>📝 Crea tu plantilla de WhatsApp y obtén el <strong>Content SID</strong> (ej: HX1234...)</li>
                    <li>⚙️ Configura el SID arriba y marca como "Aprobada en Twilio"</li>
                    <li>✅ Las campañas usarán este template automáticamente</li>
                </ol>
                <em>💡 Si no configuras esto, las campañas usarán envío directo (sin template).</em>
            </div>'''),
            'classes': ('collapse',)
        }),
    )

    def nombre_con_emoji(self, obj):
        return str(obj)
    nombre_con_emoji.short_description = "Plantilla"

    def categoria_display(self, obj):
        return obj.get_categoria_display()
    categoria_display.short_description = "Categoría"
    
    def estado_template(self, obj):
        if obj.twilio_template_sid and obj.aprobada_twilio:
            return format_html('<span style="background:#4caf50;color:white;padding:4px 8px;border-radius:12px;font-size:11px;">✅ TWILIO</span>')
        elif obj.twilio_template_sid and not obj.aprobada_twilio:
            return format_html('<span style="background:#ff9800;color:white;padding:4px 8px;border-radius:12px;font-size:11px;">⏳ PENDIENTE</span>')
        else:
            return format_html('<span style="background:#2196f3;color:white;padding:4px 8px;border-radius:12px;font-size:11px;">📱 DIRECTO</span>')
    estado_template.short_description = "Estado"
    
    def total_plantillas(self, obj):
        count = obj.plantillas.count()
        if count > 0:
            return format_html('<span style="background:#4caf50;color:white;padding:4px 12px;border-radius:12px;font-weight:bold;">{}</span>',
                count
            )
        return format_html('<span style="color:#999;">0</span>')
    total_plantillas.short_description = "📄 Plantillas"
    
    def total_campanas(self, obj):
        count = obj.campanas.count()
        if count > 0:
            return format_html('<span style="background:#2196f3;color:white;padding:4px 12px;border-radius:12px;font-weight:bold;">{}</span>',
                count
            )
        return format_html('<span style="color:#999;">0</span>')
    total_campanas.short_description = "📢 Campañas"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('plantilla-dashboard/', self.admin_site.admin_view(self.plantilla_dashboard_view), name='plantilla_dashboard'),
        ]
        return custom_urls + urls

    def plantilla_dashboard_view(self, request):
        """Vista del dashboard de plantillas"""
        plantillas = Plantilla.objects.all().order_by('categoria', 'nombre_interno')

        # Agrupar por categoría
        categorias = {}
        for plantilla in plantillas:
            cat = plantilla.get_categoria_display()
            if cat not in categorias:
                categorias[cat] = []
            categorias[cat].append(plantilla)

        # Estadísticas
        stats = {
            'total': plantillas.count(),
            'twilio_configuradas': plantillas.filter(twilio_template_sid__isnull=False, aprobada_twilio=True).count(),
            'twilio_pendientes': plantillas.filter(twilio_template_sid__isnull=False, aprobada_twilio=False).count(),
            'directo': plantillas.filter(twilio_template_sid__isnull=True).count(),
        }

        context = {
            'categorias': categorias,
            'stats': stats,
            'title': '📊 Dashboard de Plantillas y Content Templates',
        }

        return render(request, 'admin/plantilla_dashboard.html', context)

    def changelist_view(self, request, extra_context=None):
        """Agregar enlace al dashboard en la lista de plantillas"""
        extra_context = extra_context or {}
        extra_context['dashboard_url'] = 'plantilla-dashboard/'
        return super().changelist_view(request, extra_context)


# ========================================
# 🤝 ADMIN DE PROSPECTOS B2B (LEADS)
# ========================================

class ProspectoB2BAdmin(admin.ModelAdmin):
    """
    🤝 GESTIÓN DE LEADS B2B
    Prospectos capturados desde WhatsApp (Phase 0 del webhook).
    """
    list_display = ('telefono', 'nombre_contacto', 'empresa', 'email', 'estado_badge', 'origen', 'fecha_captura')
    list_filter = ('estado', 'origen', 'fecha_captura')
    search_fields = ('telefono', 'email', 'empresa', 'nombre_contacto')
    list_per_page = 50
    ordering = ('-fecha_captura',)
    readonly_fields = ('fecha_captura',)
    actions = ['enviar_campana_b2b']
    change_list_template = 'admin/prospectob2b_changelist.html'

    fieldsets = (
        ('📱 Contacto', {
            'fields': ('telefono', 'nombre_contacto', 'email', 'empresa')
        }),
        ('📊 Estado', {
            'fields': ('estado', 'origen', 'fecha_captura', 'notas')
        }),
    )

    def estado_badge(self, obj):
        colores = {
            'nuevo': '#2196f3',
            'contactado': '#ff9800',
            'convertido': '#4caf50',
            'descartado': '#999',
        }
        color = colores.get(obj.estado, '#999')
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 10px;border-radius:12px;font-size:11px;">{}</span>',
            color, obj.get_estado_display()
        )
    estado_badge.short_description = "Estado"

    @admin.action(description='📤 Enviar campaña B2B a seleccionados')
    def enviar_campana_b2b(self, request, queryset):
        """Envía una campaña B2B elegida por el admin a los prospectos seleccionados."""
        from .models import CampanaB2B
        from .utils import enviar_whatsapp_twilio
        from .whatsapp_service import enviar_template_twilio

        campanas = CampanaB2B.objects.all().order_by('-fecha_creacion')
        if not campanas.exists():
            self.message_user(request, "❌ No hay campañas B2B creadas. Crea una primero.", level='error')
            return

        # Paso 2: Si el admin ya eligió la campaña, enviar
        if 'campana_id' in request.POST:
            try:
                campana = CampanaB2B.objects.get(id=request.POST['campana_id'])
            except CampanaB2B.DoesNotExist:
                self.message_user(request, "❌ Campaña no encontrada.", level='error')
                return

            enviados = 0
            errores = 0

            for prospecto in queryset:
                try:
                    if campana.twilio_template_sid:
                        variables = {}
                        if prospecto.nombre_contacto:
                            variables['1'] = prospecto.nombre_contacto
                        resultado = enviar_template_twilio(
                            prospecto.telefono,
                            campana.twilio_template_sid,
                            variables=variables if variables else None
                        )
                        if resultado.get('success'):
                            enviados += 1
                        else:
                            errores += 1

                        if campana.url_media and resultado.get('success'):
                            import time
                            time.sleep(1)
                            enviar_whatsapp_twilio(
                                prospecto.telefono,
                                "📄 Adjunto:",
                                media_url=campana.url_media
                            )
                    else:
                        texto = campana.mensaje.replace('{nombre}', prospecto.nombre_contacto or 'Estimado/a')
                        media = campana.url_media or None
                        resultado = enviar_whatsapp_twilio(
                            prospecto.telefono,
                            texto,
                            media_url=media
                        )
                        if resultado.get('success'):
                            enviados += 1
                        else:
                            errores += 1

                    prospecto.estado = 'contactado'
                    prospecto.fecha_ultimo_contacto = timezone.now()
                    prospecto.save()

                except Exception as e:
                    errores += 1
                    import logging
                    logging.getLogger(__name__).error(f"Error enviando a {prospecto.telefono}: {e}")

            campana.total_enviados += enviados
            campana.total_errores += errores
            campana.estado = 'enviada'
            campana.fecha_envio = timezone.now()
            campana.save()

            self.message_user(
                request,
                f"📤 Campaña '{campana.nombre}' enviada: {enviados} exitosos, {errores} errores"
            )
            return

        # Paso 1: Mostrar formulario para elegir campaña
        from django.template.response import TemplateResponse
        return TemplateResponse(request, 'admin/elegir_campana_b2b.html', {
            'title': 'Elegir campaña B2B',
            'campanas': campanas,
            'prospectos': queryset,
            'prospectos_ids': ','.join(str(p.pk) for p in queryset),
            'action': 'enviar_campana_b2b',
            'opts': self.model._meta,
        })


# ========================================
# � ADMIN DE CAMPAÑAS B2B
# ========================================

class CampanaB2BAdminForm(forms.ModelForm):
    media_file_upload = forms.FileField(
        label='Subir archivo desde PC',
        required=False,
        help_text='Opcional. Guarda el archivo en storage/S3 y completa URL del PDF/Media.',
    )

    class Meta:
        model = CampanaB2B
        fields = '__all__'

    def save(self, commit=True):
        instance = super().save(commit=False)
        uploaded_file = self.cleaned_data.get('media_file_upload')
        if uploaded_file:
            instance.url_media = guardar_upload_admin_media(
                uploaded_file,
                carpeta='campanas_b2b',
                prefix='media',
            )
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class CampanaB2BAdmin(admin.ModelAdmin):
    """
    📤 CAMPAÑAS B2B — Envío de plantillas/PDF a prospectos sin registrarlos.
    """
    form = CampanaB2BAdminForm
    list_display = ('nombre', 'estado_badge', 'tiene_template', 'tiene_media', 'total_enviados', 'total_errores', 'fecha_creacion')
    list_filter = ('estado',)
    search_fields = ('nombre',)
    readonly_fields = ('total_enviados', 'total_errores', 'fecha_envio')
    ordering = ('-fecha_creacion',)

    fieldsets = (
        ('📤 Campaña', {
            'fields': ('nombre', 'estado')
        }),
        ('💬 Contenido del Mensaje', {
            'fields': ('mensaje', 'twilio_template_sid', 'url_media', 'media_file_upload'),
            'description': '📝 Si especificas un Content SID de Twilio, se usará ese template. Si no, se envía el mensaje de texto. Puedes pegar URL o subir archivo desde PC; el PDF/media se adjunta en ambos casos.'
        }),
        ('📊 Resultados', {
            'fields': ('total_enviados', 'total_errores', 'fecha_envio'),
            'classes': ('collapse',)
        }),
    )

    def estado_badge(self, obj):
        color = '#ffc107' if obj.estado == 'borrador' else '#28a745'
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 10px;border-radius:12px;font-size:11px;">{}</span>',
            color, obj.get_estado_display()
        )
    estado_badge.short_description = "Estado"

    def tiene_template(self, obj):
        if obj.twilio_template_sid:
            return format_html('<span style="color:green;">✅ {}</span>', obj.twilio_template_sid[:15] + '...')
        return format_html('<span style="color:#999;">—</span>')
    tiene_template.short_description = "Template"

    def tiene_media(self, obj):
        if obj.url_media:
            return format_html('<span style="color:green;">📎 Sí</span>')
        return format_html('<span style="color:#999;">—</span>')
    tiene_media.short_description = "Media"


# ========================================
# 📚 ADMIN DE DOCUMENTOS RAG
# ========================================

def _nombre_documento_desde_nombre_archivo(filename: str) -> str:
    """Identificador a partir del archivo subido (sin extensión), máx. 200 caracteres."""
    base = os.path.basename((filename or '').strip())
    stem, _ext = os.path.splitext(base)
    stem = (stem or base or 'documento').strip()
    return stem[:200]


def _nombre_rag_comercial_unico(cliente, canal: str, nombre_base: str, exclude_pk=None):
    """Evita violar unique_together (cliente, canal, nombre)."""
    nombre = nombre_base[:200]
    q = DocumentoRAGComercial.objects.filter(canal=canal, nombre=nombre)
    q = q.filter(cliente__isnull=True) if cliente is None else q.filter(cliente=cliente)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    if not q.exists():
        return nombre
    for i in range(2, 200):
        suf = f'_{i}'
        candidato = (nombre_base[: 200 - len(suf)] + suf)[:200]
        q2 = DocumentoRAGComercial.objects.filter(canal=canal, nombre=candidato)
        q2 = q2.filter(cliente__isnull=True) if cliente is None else q2.filter(cliente=cliente)
        if exclude_pk:
            q2 = q2.exclude(pk=exclude_pk)
        if not q2.exists():
            return candidato
    return (nombre_base[:190] + '_dup')[:200]


def _nombre_rag_curso_unico(curso, nombre_base: str, exclude_pk=None):
    nombre = nombre_base[:200]
    q = DocumentoRAG.objects.filter(curso=curso, nombre=nombre)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    if not q.exists():
        return nombre
    for i in range(2, 200):
        suf = f'_{i}'
        candidato = (nombre_base[: 200 - len(suf)] + suf)[:200]
        q2 = DocumentoRAG.objects.filter(curso=curso, nombre=candidato)
        if exclude_pk:
            q2 = q2.exclude(pk=exclude_pk)
        if not q2.exists():
            return candidato
    return (nombre_base[:190] + '_dup')[:200]


class DocumentoRAGComercialAdminForm(forms.ModelForm):
    archivo_segundo = forms.FileField(
        required=False,
        label='Segundo archivo (opcional)',
        help_text='Mismo cliente, canal y tipo. Se crea otro documento y se indexa en segundo plano.',
    )

    class Meta:
        model = DocumentoRAGComercial
        fields = '__all__'

    def clean(self):
        data = super().clean()
        archivo = data.get('archivo')
        nombre = (data.get('nombre') or '').strip()
        if archivo and not nombre:
            data['nombre'] = _nombre_documento_desde_nombre_archivo(getattr(archivo, 'name', '') or '')
        return data


class DocumentoRAGAdminForm(forms.ModelForm):
    archivo_segundo = forms.FileField(
        required=False,
        label='Segundo archivo (opcional)',
        help_text='Mismo curso y tipo. Se crea otro documento y se indexa en segundo plano.',
    )

    class Meta:
        model = DocumentoRAG
        fields = '__all__'

    def clean(self):
        data = super().clean()
        archivo = data.get('archivo')
        nombre = (data.get('nombre') or '').strip()
        if archivo and not nombre:
            data['nombre'] = _nombre_documento_desde_nombre_archivo(getattr(archivo, 'name', '') or '')
        return data


def _enqueue_indexar_rag_task(model_class_name: str, pk: int):
    """Registra indexación en Celery después del commit (archivo/DB visibles para el worker)."""

    def _run():
        try:
            from core.tasks import indexar_documento_rag_por_id

            indexar_documento_rag_por_id.delay('core', model_class_name, pk)
        except Exception:
            logger.exception('[RAG] Fallo encolando %s id=%s', model_class_name, pk)

    transaction.on_commit(_run)


def _encolar_o_indexar_rag_doc(request, modeladmin, model_class_name: str, obj, show_index_message=True):
    """
    La indexación RAG puede superar el timeout de nginx/ALB; en producción
    se delega a Celery o a un hilo en background si no hay worker Redis.
    """
    pk = obj.pk

    def _run():
        try:
            if getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False):
                import threading

                def _bg():
                    try:
                        doc = DocumentoRAGComercial.objects.filter(pk=pk).first()
                        if doc and doc.archivo:
                            doc.indexar()
                    except Exception:
                        logger.exception('[RAG] Indexación background falló %s id=%s', model_class_name, pk)

                threading.Thread(target=_bg, daemon=True, name=f'rag-idx-{pk}').start()
            else:
                from core.tasks import indexar_documento_rag_por_id

                indexar_documento_rag_por_id.delay('core', model_class_name, pk)
        except Exception:
            logger.exception('[RAG] Fallo encolando %s id=%s', model_class_name, pk)

    transaction.on_commit(_run)
    if show_index_message:
        modeladmin.message_user(
            request,
            'La indexación RAG se ejecuta en segundo plano; actualiza el listado en unos minutos para ver el estado.',
            messages.INFO,
        )
    return None


def _encolar_zip_rag_comercial(
    storage_path: str,
    cliente_id,
    canal: str,
    tipo: str,
    descripcion: str,
    user_id,
) -> None:
    """ZIP → worker Celery; si no hay broker, hilo en background (evita 504 en EB)."""
    args = (storage_path, cliente_id, canal, tipo, descripcion, user_id)

    def _run():
        import threading

        from core.tasks import procesar_zip_rag_comercial

        def _bg():
            try:
                procesar_zip_rag_comercial.apply(args=args)
            except Exception:
                logger.exception('[RAG comercial] Procesamiento ZIP en background falló')

        try:
            if getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False):
                threading.Thread(target=_bg, daemon=True, name='rag-zip').start()
            else:
                procesar_zip_rag_comercial.delay(*args)
        except Exception:
            logger.warning('[RAG comercial] Celery no disponible; ZIP en hilo background')
            threading.Thread(target=_bg, daemon=True, name='rag-zip').start()

    transaction.on_commit(_run)


# Subida masiva solo RAG comercial (varios archivos → un DocumentoRAGComercial por archivo).
RAG_COMERCIAL_SUBIDA_MASIVA_MIN = 2
# Subida masiva solo RAG comercial (varios archivos → un DocumentoRAGComercial por archivo).
RAG_COMERCIAL_SUBIDA_MASIVA_MAX = 50
RAG_COMERCIAL_ZIP_MAX_BYTES = 150 * 1024 * 1024  # 150 MB
RAG_COMERCIAL_ZIP_MAX_FILES = 100
_RAG_COMERCIAL_EXT_OK = {'.pdf', '.docx', '.txt', '.xlsx', '.xlsm', '.xls'}


def _extension_archivo_comercial_ok(nombre: str) -> bool:
    ext = os.path.splitext((nombre or '').lower())[1]
    return ext in _RAG_COMERCIAL_EXT_OK


class MetaMetricaEmpresaAdmin(admin.ModelAdmin):
    list_display = (
        "cliente", "curso", "meta_finalizacion_porcentaje",
        "meta_inicio_porcentaje", "meta_max_no_iniciados_porcentaje", "activa",
    )
    list_filter = ("activa", "cliente")
    search_fields = ("cliente__nombre", "curso__nombre")
    autocomplete_fields = ("cliente", "curso")


class MetaMetricaNatiAdmin(admin.ModelAdmin):
    list_display = ("cliente", "meta_lectura_porcentaje", "meta_respuesta_porcentaje", "activa")
    list_filter = ("activa",)
    search_fields = ("cliente__nombre",)
    autocomplete_fields = ("cliente",)


class DocumentoRAGAdmin(admin.ModelAdmin):
    """
    📚 GESTIÓN DE DOCUMENTOS RAG — Base de Conocimiento para Agentes IA
    Multi-Tenant: cada documento está aislado por Cliente + Curso.
    """
    form = DocumentoRAGAdminForm
    list_display = ('nombre', 'curso_link', 'cliente_display', 'tipo_badge', 'estado_rag_badge', 'chunks_indexados', 'fecha_subida')
    list_filter = ('estado', 'tipo', 'curso__cliente', 'curso')
    search_fields = ('nombre', 'descripcion', 'curso__nombre', 'curso__cliente__nombre')
    list_per_page = 50
    ordering = ('-fecha_subida',)
    readonly_fields = ('estado', 'chunks_indexados', 'fecha_subida', 'fecha_indexado', 'subido_por')
    actions = ['indexar_seleccionados', 'reindexar_seleccionados', 'eliminar_del_rag']

    fieldsets = (
        ('📄 Documento', {
            'fields': ('curso', 'nombre', 'archivo', 'archivo_segundo', 'tipo', 'descripcion'),
            'description': 'Si dejás "Nombre" vacío al subir archivo, se usa el nombre del archivo (sin extensión). Podés adjuntar un segundo archivo en la misma carga.',
        }),
        ('🤖 Estado RAG', {
            'fields': ('estado', 'chunks_indexados', 'fecha_subida', 'fecha_indexado', 'subido_por'),
            'description': 'Estado de indexación en la base de datos vectorial. Los documentos indexados son usados por los agentes IA.'
        }),
    )

    def curso_link(self, obj):
        url = reverse('admin:core_curso_change', args=[obj.curso_id])
        return format_html('<a href="{}">{}</a>', url, obj.curso.nombre)
    curso_link.short_description = "Curso"

    def cliente_display(self, obj):
        if obj.curso.cliente:
            return obj.curso.cliente.nombre
        return format_html('<span style="color:#999;">General (eki)</span>')
    cliente_display.short_description = "🏢 Cliente"

    def tipo_badge(self, obj):
        colores = {'contenido': '#2196F3', 'manual': '#9C27B0', 'faq': '#FF9800', 'guia': '#4CAF50', 'normativa': '#607D8B'}
        color = colores.get(obj.tipo, '#999')
        return format_html(
            '<span style="background:{};color:white;padding:3px 8px;border-radius:12px;font-size:11px;">{}</span>',
            color, obj.get_tipo_display()
        )
    tipo_badge.short_description = "Tipo"

    def estado_rag_badge(self, obj):
        colores = {'pendiente': '#ffc107', 'indexado': '#28a745', 'error': '#dc3545'}
        color = colores.get(obj.estado, '#6c757d')
        label = obj.get_estado_display()
        return format_html(
            '<span style="background:{};color:white;padding:3px 10px;border-radius:12px;font-size:11px;">{}</span>',
            color, label
        )
    estado_rag_badge.short_description = "Estado RAG"

    def save_model(self, request, obj, form, change):
        arch2 = form.cleaned_data.get('archivo_segundo') if form and hasattr(form, 'cleaned_data') else None
        if obj.archivo and (not obj.nombre or not str(obj.nombre).strip()):
            obj.nombre = _nombre_documento_desde_nombre_archivo(obj.archivo.name)
        if obj.curso_id and obj.nombre:
            obj.nombre = _nombre_rag_curso_unico(obj.curso, obj.nombre, exclude_pk=obj.pk if change else None)
        if not obj.subido_por_id:
            obj.subido_por = request.user
        super().save_model(request, obj, form, change)
        if obj.estado == 'pendiente' and obj.archivo:
            _encolar_o_indexar_rag_doc(request, self, 'DocumentoRAG', obj)
        if arch2:
            n2 = _nombre_documento_desde_nombre_archivo(arch2.name)
            n2 = _nombre_rag_curso_unico(obj.curso, n2)
            doc2 = DocumentoRAG(
                curso=obj.curso,
                nombre=n2,
                archivo=arch2,
                tipo=obj.tipo,
                descripcion=obj.descripcion or '',
                subido_por=request.user,
                estado='pendiente',
            )
            doc2.save()
            _encolar_o_indexar_rag_doc(request, self, 'DocumentoRAG', doc2)
            self.message_user(
                request,
                f'Se creó también el documento «{n2}» en el mismo curso (indexación en segundo plano).',
                messages.SUCCESS,
            )

    @admin.action(description='🤖 Indexar documentos seleccionados en RAG')
    def indexar_seleccionados(self, request, queryset):
        ok, err, queued = 0, 0, 0
        eager = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
        for doc in queryset.filter(estado='pendiente'):
            if not eager:
                _enqueue_indexar_rag_task('DocumentoRAG', doc.pk)
                queued += 1
                continue
            n = doc.indexar()
            if n > 0:
                ok += 1
            else:
                err += 1
        msg = []
        if queued:
            msg.append(f'{queued} en cola (segundo plano)')
        if ok or err:
            msg.append(f'{ok} indexados en línea' + (f', {err} con error' if err else ''))
        self.message_user(request, '✅ ' + '. '.join(msg) if msg else 'Nada que indexar.')

    @admin.action(description='🔄 Re-indexar documentos seleccionados')
    def reindexar_seleccionados(self, request, queryset):
        ok, queued = 0, 0
        eager = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
        for doc in queryset:
            doc.estado = 'pendiente'
            doc.save(update_fields=['estado'])
            if not eager:
                _enqueue_indexar_rag_task('DocumentoRAG', doc.pk)
                queued += 1
                continue
            n = doc.indexar()
            if n > 0:
                ok += 1
        parts = [f'{ok} re-indexados en línea'] if ok else []
        if queued:
            parts.append(f'{queued} en cola')
        self.message_user(request, '✅ ' + ', '.join(parts) if parts else 'Listo.')

    @admin.action(description='🗑️ Eliminar del RAG (sin borrar archivo)')
    def eliminar_del_rag(self, request, queryset):
        from core.rag_manager import rag_manager
        for doc in queryset:
            rag_manager.eliminar_documento(doc.cliente_id, doc.curso_id, doc.nombre)
            doc.estado = 'pendiente'
            doc.chunks_indexados = 0
            doc.save(update_fields=['estado', 'chunks_indexados'])
        self.message_user(request, f"✅ {queryset.count()} documentos eliminados del índice RAG.")

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['title'] = '📚 Documentos RAG — Base de Conocimiento para Agentes IA'
        return super().changelist_view(request, extra_context)


@admin.register(ProductoCatalogo)
class ProductoCatalogoAdmin(admin.ModelAdmin):
    list_display = (
        'nombre', 'cliente', 'categoria',
        'cultivos_objetivo', 'precio_cop', 'unidad', 'activo',
    )
    list_filter = ('cliente', 'categoria', 'activo')
    search_fields = ('nombre', 'descripcion', 'problema_que_resuelve', 'ingrediente_activo')
    list_editable = ('activo',)
    readonly_fields = ('fecha_actualizacion',)
    fieldsets = (
        ('Identificación', {
            'fields': ('cliente', 'nombre', 'categoria', 'cultivos_objetivo', 'activo'),
        }),
        ('Información técnica', {
            'fields': ('descripcion', 'problema_que_resuelve', 'ingrediente_activo', 'dosis'),
        }),
        ('Información comercial', {
            'fields': ('precio_cop', 'unidad', 'url_producto'),
        }),
        ('Auditoría', {
            'fields': ('fecha_actualizacion',),
            'classes': ('collapse',),
        }),
    )
    actions = ['desactivar_productos', 'activar_productos']

    @admin.action(description='Desactivar productos seleccionados')
    def desactivar_productos(self, request, queryset):
        count = queryset.update(activo=False)
        self.message_user(request, f'{count} productos desactivados.')

    @admin.action(description='Activar productos seleccionados')
    def activar_productos(self, request, queryset):
        count = queryset.update(activo=True)
        self.message_user(request, f'{count} productos activados.')


@admin.register(ProductoComercial)
class ProductoComercialAdmin(admin.ModelAdmin):
    """💰 Catálogo de precios estructurado para Nat (Postgres)."""

    change_list_template = 'admin/productocomercial_changelist.html'
    list_display = (
        'sku', 'nombre', 'presentacion', 'precio', 'moneda',
        'cliente', 'categoria', 'activo', 'fecha_actualizacion',
    )
    list_filter = ('activo', 'cliente', 'categoria', 'moneda')
    search_fields = ('sku', 'nombre', 'categoria', 'notas')
    list_editable = ('precio', 'activo')
    ordering = ('-fecha_actualizacion',)
    list_per_page = 50

    def get_urls(self):
        info = self.model._meta.app_label, self.model._meta.model_name
        custom = [
            path(
                'importar-precios/',
                self.admin_site.admin_view(self.importar_precios_view),
                name='%s_%s_importar_precios' % info,
            ),
        ]
        return custom + super().get_urls()

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        opts = self.model._meta
        extra_context['title'] = '💰 Precios comerciales — Nat (Postgres)'
        extra_context['importar_precios_url'] = reverse(
            'admin:%s_%s_importar_precios' % (opts.app_label, opts.model_name)
        )
        return super().changelist_view(request, extra_context)

    def importar_precios_view(self, request):
        import os
        import tempfile

        from django.core.exceptions import PermissionDenied

        from core.precios_import import importar_precios_desde_archivo

        if not self.has_add_permission(request):
            raise PermissionDenied

        opts = self.model._meta
        changelist_url = reverse('admin:%s_%s_changelist' % (opts.app_label, opts.model_name))
        clientes = Cliente.objects.filter(activo=True).order_by('nombre')
        context = {
            **self.admin_site.each_context(request),
            'title': 'Importar precios comerciales (Excel → Postgres)',
            'opts': opts,
            'changelist_url': changelist_url,
            'clientes': clientes,
            'preview': None,
            'cliente_sel': '',
            'vigencia_desde': '',
            'vigencia_hasta': '',
            'desactivar_ausentes': False,
        }

        if request.method != 'POST':
            return render(request, 'admin/importar_precios_comercial.html', context)

        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Seleccioná un archivo Excel o JSON.')
            return render(request, 'admin/importar_precios_comercial.html', context)

        cliente_id_raw = (request.POST.get('cliente') or '').strip()
        cliente_id = int(cliente_id_raw) if cliente_id_raw else None
        context['cliente_sel'] = cliente_id_raw
        context['vigencia_desde'] = request.POST.get('vigencia_desde') or ''
        context['vigencia_hasta'] = request.POST.get('vigencia_hasta') or ''
        context['desactivar_ausentes'] = request.POST.get('desactivar_ausentes') == '1'

        ext = os.path.splitext(archivo.name or '')[1].lower()
        if ext not in {'.xlsx', '.xlsm', '.xls', '.json'}:
            messages.error(request, 'Formato no soportado. Use .xlsx, .xlsm, .xls o .json')
            return render(request, 'admin/importar_precios_comercial.html', context)

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                for chunk in archivo.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            dry_run = (request.POST.get('accion') or '').strip().lower() == 'vista previa'
            result = importar_precios_desde_archivo(
                tmp_path,
                cliente_id=cliente_id,
                desactivar_ausentes=context['desactivar_ausentes'],
                dry_run=dry_run,
                vigencia_desde=context['vigencia_desde'] or None,
                vigencia_hasta=context['vigencia_hasta'] or None,
            )

            if result.errores:
                messages.error(
                    request,
                    'Errores de validación:\n' + '\n'.join(result.errores[:12]),
                )
                context['preview'] = result
                return render(request, 'admin/importar_precios_comercial.html', context)

            if dry_run:
                context['preview'] = result
                messages.info(
                    request,
                    f'Vista previa: {result.total_validos} producto(s) válido(s) para «{result.cliente_nombre}».',
                )
                return render(request, 'admin/importar_precios_comercial.html', context)

            messages.success(
                request,
                f'Importación lista: {result.creados} creados, {result.actualizados} actualizados'
                + (f', {result.desactivados} desactivados' if result.desactivados else '')
                + f' — cliente «{result.cliente_nombre}».',
            )
            return redirect(changelist_url)
        except (ValueError, FileNotFoundError) as e:
            messages.error(request, str(e))
            return render(request, 'admin/importar_precios_comercial.html', context)
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass


class DocumentoRAGComercialAdmin(admin.ModelAdmin):
    """🛒 Documentos del RAG comercial (aislado del educativo)."""

    form = DocumentoRAGComercialAdminForm
    change_list_template = 'admin/documentoragcomercial_changelist.html'
    list_display = (
        'nombre',
        'cliente_display',
        'canal',
        'tipo_badge',
        'estado_rag_badge',
        'chunks_indexados',
        'fecha_subida',
    )
    list_filter = ('estado', 'tipo', 'canal', 'cliente')
    search_fields = ('nombre', 'descripcion', 'cliente__nombre')
    list_per_page = 50
    ordering = ('-fecha_subida',)
    readonly_fields = ('estado', 'chunks_indexados', 'fecha_subida', 'fecha_indexado', 'subido_por')
    actions = [
        'indexar_seleccionados',
        'reindexar_seleccionados',
        'eliminar_del_rag',
        'importar_precios_a_catalogo',
    ]

    def get_urls(self):
        info = self.model._meta.app_label, self.model._meta.model_name
        custom = [
            path(
                'subida-masiva/',
                self.admin_site.admin_view(self.subida_masiva_view),
                name='%s_%s_subida_masiva' % info,
            ),
        ]
        return custom + super().get_urls()

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        opts = self.model._meta
        extra_context['subida_masiva_max'] = RAG_COMERCIAL_SUBIDA_MASIVA_MAX
        extra_context['subida_masiva_url'] = reverse(
            'admin:%s_%s_subida_masiva' % (opts.app_label, opts.model_name)
        )
        return super().changelist_view(request, extra_context)

    fieldsets = (
        ('📄 Documento Comercial', {
            'fields': ('cliente', 'canal', 'nombre', 'archivo', 'archivo_segundo', 'tipo', 'descripcion'),
            'description': 'Si dejás "Nombre" vacío al subir archivo, se usa el nombre del archivo (sin extensión). Podés adjuntar un segundo archivo en la misma carga.',
        }),
        ('🤖 Estado RAG Comercial', {
            'fields': ('estado', 'chunks_indexados', 'fecha_subida', 'fecha_indexado', 'subido_por'),
            'description': 'Estos documentos alimentan el bot comercial y NO el bot educativo de cursos.'
        }),
    )

    def cliente_display(self, obj):
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;">General</span>')
    cliente_display.short_description = '🏢 Cliente'

    def tipo_badge(self, obj):
        colores = {
            'producto': '#0ea5e9',
            'precio': '#ef4444',
            'informe_tecnico': '#475569',
            'faq': '#f59e0b',
            'politica': '#64748b',
            'promo': '#10b981',
            'general': '#3b82f6',
        }
        color = colores.get(obj.tipo, '#999')
        return format_html(
            '<span style="background:{};color:white;padding:3px 8px;border-radius:12px;font-size:11px;">{}</span>',
            color,
            obj.get_tipo_display(),
        )
    tipo_badge.short_description = 'Tipo'

    def estado_rag_badge(self, obj):
        colores = {'pendiente': '#ffc107', 'indexado': '#28a745', 'error': '#dc3545'}
        color = colores.get(obj.estado, '#6c757d')
        return format_html(
            '<span style="background:{};color:white;padding:3px 10px;border-radius:12px;font-size:11px;">{}</span>',
            color,
            obj.get_estado_display(),
        )
    estado_rag_badge.short_description = 'Estado RAG'

    def subida_masiva_view(self, request):
        """Varios archivos → un documento comercial por archivo; el nombre sale del nombre del archivo."""
        from django.core.exceptions import PermissionDenied

        if not self.has_add_permission(request):
            raise PermissionDenied

        opts = self.model._meta
        changelist_url = reverse('admin:%s_%s_changelist' % (opts.app_label, opts.model_name))

        clientes = Cliente.objects.filter(activo=True).order_by('nombre')

        context = {
            **self.admin_site.each_context(request),
            'title': 'Subida masiva — RAG comercial',
            'opts': opts,
            'changelist_url': changelist_url,
            'clientes': clientes,
            'canal_choices': DocumentoRAGComercial.CANAL_CHOICES,
            'tipo_choices': DocumentoRAGComercial.TIPO_CHOICES,
            'min_files': RAG_COMERCIAL_SUBIDA_MASIVA_MIN,
            'max_files': RAG_COMERCIAL_SUBIDA_MASIVA_MAX,
            'ext_help': ', '.join(sorted(_RAG_COMERCIAL_EXT_OK)),
        }

        if request.method != 'POST':
            return render(request, 'admin/subida_masiva_rag_comercial.html', context)

        canal = (request.POST.get('canal') or 'bot_comercial').strip()
        tipo = (request.POST.get('tipo') or 'general').strip()
        descripcion = (request.POST.get('descripcion') or '').strip()
        cliente_id = request.POST.get('cliente') or ''
        cliente = None
        if cliente_id:
            try:
                cliente = Cliente.objects.get(pk=int(cliente_id))
            except (ValueError, TypeError, Cliente.DoesNotExist):
                messages.error(request, 'Cliente inválido.')
                return render(request, 'admin/subida_masiva_rag_comercial.html', context)

        valid_canal = {c for c, _ in DocumentoRAGComercial.CANAL_CHOICES}
        if canal not in valid_canal:
            messages.error(request, 'Canal inválido.')
            return render(request, 'admin/subida_masiva_rag_comercial.html', context)

        valid_tipo = {t for t, _ in DocumentoRAGComercial.TIPO_CHOICES}
        if tipo not in valid_tipo:
            messages.error(request, 'Tipo de documento inválido.')
            return render(request, 'admin/subida_masiva_rag_comercial.html', context)

        archivo_zip = request.FILES.get('archivo_zip')
        archivos = request.FILES.getlist('archivos')

        if archivo_zip:
            import uuid

            from django.core.files.storage import default_storage

            if not (archivo_zip.name or '').lower().endswith('.zip'):
                messages.error(request, 'El archivo debe ser un .zip')
                return render(request, 'admin/subida_masiva_rag_comercial.html', context)
            if archivo_zip.size > RAG_COMERCIAL_ZIP_MAX_BYTES:
                messages.error(
                    request,
                    f'ZIP demasiado grande (máx {RAG_COMERCIAL_ZIP_MAX_BYTES // (1024 * 1024)} MB).',
                )
                return render(request, 'admin/subida_masiva_rag_comercial.html', context)
            storage_path = default_storage.save(
                f'rag_zip_uploads/{uuid.uuid4().hex}.zip',
                archivo_zip,
            )
            _encolar_zip_rag_comercial(
                storage_path,
                cliente.pk if cliente else None,
                canal,
                tipo,
                descripcion,
                request.user.pk,
            )
            messages.success(
                request,
                f'ZIP recibido. Se procesará en segundo plano (hasta {RAG_COMERCIAL_ZIP_MAX_FILES} archivos válidos). '
                'Refrescá el listado en unos minutos.',
            )
            return redirect(changelist_url)

        n = len(archivos)
        if n < RAG_COMERCIAL_SUBIDA_MASIVA_MIN:
            messages.error(
                request,
                f'Seleccioná al menos {RAG_COMERCIAL_SUBIDA_MASIVA_MIN} archivos, o subí un ZIP.',
            )
            return render(request, 'admin/subida_masiva_rag_comercial.html', context)
        if n > RAG_COMERCIAL_SUBIDA_MASIVA_MAX:
            messages.error(
                request,
                f'Máximo {RAG_COMERCIAL_SUBIDA_MASIVA_MAX} archivos por carga. Dividí en varias tandas.',
            )
            return render(request, 'admin/subida_masiva_rag_comercial.html', context)

        rechazados = []
        for f in archivos:
            if not _extension_archivo_comercial_ok(getattr(f, 'name', '') or ''):
                rechazados.append(getattr(f, 'name', '?'))
        if rechazados:
            messages.error(
                request,
                'Extensión no permitida en: ' + ', '.join(rechazados[:12])
                + (f' (+{len(rechazados) - 12} más)' if len(rechazados) > 12 else '')
                + f'. Permitidos: {context["ext_help"]}',
            )
            return render(request, 'admin/subida_masiva_rag_comercial.html', context)

        creados = []
        try:
            with transaction.atomic():
                for f in archivos:
                    nombre = _nombre_documento_desde_nombre_archivo(f.name)
                    nombre = _nombre_rag_comercial_unico(cliente, canal, nombre)
                    doc = DocumentoRAGComercial(
                        cliente=cliente,
                        canal=canal,
                        nombre=nombre,
                        archivo=f,
                        tipo=tipo,
                        descripcion=descripcion,
                        subido_por=request.user,
                        estado='pendiente',
                    )
                    doc.save()
                    creados.append(doc)
        except Exception:
            logger.exception('[RAG comercial] Subida masiva falló')
            messages.error(request, 'No se pudieron guardar los documentos. Revisá los logs o probá con menos archivos.')
            return render(request, 'admin/subida_masiva_rag_comercial.html', context)

        for doc in creados:
            if doc.estado == 'pendiente' and doc.archivo:
                _encolar_o_indexar_rag_doc(
                    request, self, 'DocumentoRAGComercial', doc, show_index_message=False
                )

        messages.success(
            request,
            f'Se crearon {len(creados)} documentos comerciales. El título de cada uno es el nombre del archivo '
            '(ajustado si ya existía). La indexación corre en segundo plano; refrescá el listado en unos minutos.',
        )
        return redirect(changelist_url)

    def save_model(self, request, obj, form, change):
        arch2 = form.cleaned_data.get('archivo_segundo') if form and hasattr(form, 'cleaned_data') else None
        if obj.archivo and (not obj.nombre or not str(obj.nombre).strip()):
            obj.nombre = _nombre_documento_desde_nombre_archivo(obj.archivo.name)
        obj.nombre = _nombre_rag_comercial_unico(obj.cliente, obj.canal, obj.nombre, exclude_pk=obj.pk if change else None)
        if not obj.subido_por_id:
            obj.subido_por = request.user
        super().save_model(request, obj, form, change)
        if obj.estado == 'pendiente' and obj.archivo:
            _encolar_o_indexar_rag_doc(request, self, 'DocumentoRAGComercial', obj)
        if arch2:
            n2 = _nombre_documento_desde_nombre_archivo(arch2.name)
            n2 = _nombre_rag_comercial_unico(obj.cliente, obj.canal, n2)
            doc2 = DocumentoRAGComercial(
                cliente=obj.cliente,
                canal=obj.canal,
                nombre=n2,
                archivo=arch2,
                tipo=obj.tipo,
                descripcion=obj.descripcion or '',
                subido_por=request.user,
                estado='pendiente',
            )
            doc2.save()
            _encolar_o_indexar_rag_doc(request, self, 'DocumentoRAGComercial', doc2)
            self.message_user(
                request,
                f'Se creó también el documento comercial «{n2}» (indexación en segundo plano).',
                messages.SUCCESS,
            )

    @admin.action(description='🤖 Indexar documentos comerciales seleccionados')
    def indexar_seleccionados(self, request, queryset):
        ok, err, queued = 0, 0, 0
        eager = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
        for doc in queryset.filter(estado='pendiente'):
            if not eager:
                _enqueue_indexar_rag_task('DocumentoRAGComercial', doc.pk)
                queued += 1
                continue
            n = doc.indexar()
            if n > 0:
                ok += 1
            else:
                err += 1
        msg = []
        if queued:
            msg.append(f'{queued} en cola (segundo plano)')
        if ok or err:
            msg.append(f'{ok} indexados en línea' + (f', {err} con error' if err else ''))
        self.message_user(request, '✅ ' + '. '.join(msg) if msg else 'Nada que indexar.')

    @admin.action(description='🔄 Re-indexar documentos comerciales seleccionados')
    def reindexar_seleccionados(self, request, queryset):
        ok, queued = 0, 0
        eager = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
        for doc in queryset:
            doc.estado = 'pendiente'
            doc.save(update_fields=['estado'])
            if not eager:
                _enqueue_indexar_rag_task('DocumentoRAGComercial', doc.pk)
                queued += 1
                continue
            n = doc.indexar()
            if n > 0:
                ok += 1
        parts = [f'{ok} re-indexados en línea'] if ok else []
        if queued:
            parts.append(f'{queued} en cola')
        self.message_user(request, '✅ ' + ', '.join(parts) if parts else 'Listo.')

    @admin.action(description='📥 Importar precios a catálogo (Excel → Postgres)')
    def importar_precios_a_catalogo(self, request, queryset):
        from core.precios_import import importar_precios_desde_documento_rag

        ok = err = 0
        for doc in queryset:
            try:
                result = importar_precios_desde_documento_rag(doc)
                if result.errores:
                    err += 1
                    self.message_user(
                        request,
                        f'«{doc.nombre}»: ' + '; '.join(result.errores[:3]),
                        level=messages.ERROR,
                    )
                    continue
                ok += 1
                self.message_user(
                    request,
                    f'«{doc.nombre}»: {result.creados} creados, {result.actualizados} actualizados '
                    f'(cliente «{result.cliente_nombre}»).',
                    level=messages.SUCCESS,
                )
            except (ValueError, FileNotFoundError) as e:
                err += 1
                self.message_user(request, f'«{doc.nombre}»: {e}', level=messages.ERROR)
        if ok and not err:
            self.message_user(request, f'✅ {ok} lista(s) importada(s) al catálogo de precios.', level=messages.SUCCESS)

    @admin.action(description='🗑️ Eliminar del RAG comercial (sin borrar archivo)')
    def eliminar_del_rag(self, request, queryset):
        from core.rag_comercial_manager import rag_comercial_manager

        for doc in queryset:
            rag_comercial_manager.eliminar_documento(doc.cliente_scope_id, doc.canal, doc.nombre)
            doc.estado = 'pendiente'
            doc.chunks_indexados = 0
            doc.save(update_fields=['estado', 'chunks_indexados'])
        self.message_user(request, f"✅ {queryset.count()} documentos eliminados del índice comercial.")


from .models import ConfiguracionGlobal as _ConfiguracionGlobal


@admin.register(_ConfiguracionGlobal)
class ConfiguracionGlobalAdmin(admin.ModelAdmin):
    """Singleton: solo permite editar la fila id=1, no agregar ni borrar."""
    list_display = ('__str__', 'content_sid_habeas_data_global', 'fecha_actualizacion')
    readonly_fields = ('fecha_actualizacion',)
    fieldsets = (
        ('🛡️ Habeas Data — Plantilla Twilio general', {
            'fields': ('content_sid_habeas_data_global',),
            'description': (
                'Content SID (HX...) de la plantilla Twilio aprobada que se usa por defecto '
                'para enviar el Habeas Data cuando un cliente no tiene la suya propia. '
                'Cada Cliente puede sobrescribir este valor desde su ficha en '
                '"Habeas Data → Plantilla Twilio (Content SID del cliente)".'
            ),
        }),
        ('Auditoría', {
            'fields': ('fecha_actualizacion',),
        }),
    )

    def has_add_permission(self, request):
        return not _ConfiguracionGlobal.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        from django.shortcuts import redirect
        from django.urls import reverse
        obj = _ConfiguracionGlobal.get_solo()
        return redirect(reverse('admin:core_configuracionglobal_change', args=[obj.pk]))


class EventoIAAdmin(admin.ModelAdmin):
    """Solo lectura — audit trail IA (Parte 2A/2B)."""

    list_display = (
        'created_at', 'tipo', 'trace_id', 'estudiante', 'regla_aplicada',
        'agente', 'es_reto', 'canal',
    )
    list_filter = ('tipo', 'canal', 'regla_aplicada', 'created_at')
    search_fields = ('trace_id', 'input_preview', 'output_preview', 'agente')
    readonly_fields = [
        f.name for f in EventoIA._meta.fields
    ]
    ordering = ('-created_at',)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


class ContextoAgroSessionAdmin(admin.ModelAdmin):
    list_display = ('sesion', 'cultivo', 'etapa', 'region', 'problema', 'updated_at')
    list_filter = ('cultivo', 'region')
    search_fields = ('sesion__telefono', 'cultivo', 'problema', 'region', 'municipio')
    readonly_fields = ('updated_at',)


class ConversacionRAGCandidataAdmin(admin.ModelAdmin):
    list_display = (
        'fecha_creacion', 'estado', 'telefono', 'cliente', 'pregunta_corta', 'revisado_por',
    )
    list_filter = ('estado', 'fecha_creacion', 'cliente')
    search_fields = ('telefono', 'pregunta', 'respuesta_nati')
    readonly_fields = ('fecha_creacion', 'fecha_revision', 'trace_id')
    actions = ['marcar_aprobada']

    @admin.display(description='Pregunta')
    def pregunta_corta(self, obj):
        return (obj.pregunta or '')[:80]

    @admin.action(description='Marcar aprobada (sin publicar)')
    def marcar_aprobada(self, request, queryset):
        from core.knowledge_studio import revisar_candidata

        for c in queryset.filter(estado=ConversacionRAGCandidata.ESTADO_PENDIENTE):
            revisar_candidata(c, usuario=request.user, accion='aprobar')

