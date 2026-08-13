from core.admin._common import *  # noqa: F401,F403
from core.admin.estudiantes import EnvioProgramadoInline

@admin.register(Campana)
class CampanaAdmin(admin.ModelAdmin):
    """Gestión de campañas masivas"""
    change_form_template = 'admin/core/campana/change_form.html'
    list_display = ('nombre', 'cliente_nombre', 'tipo_audiencia_display', 'categoria_badge', 'plantilla_estado', 'estado_visual', 'conteo_destinatarios', 'programada_display', 'fecha_creacion')
    list_filter = ('ejecutada', 'cliente', 'categoria', 'tipo_audiencia', 'fecha_creacion', 'plantilla__aprobada_twilio')
    search_fields = ('nombre', 'cliente__nombre')
    filter_horizontal = ('destinatarios',)
    actions = ['enviar_campana_accion']
    inlines = [EnvioProgramadoInline]
    
    fieldsets = (
        ('Datos', {
            'classes': ['tab'],
            'fields': ('nombre', 'cliente'),
            'description': 'Nombre interno y organización de este lanzamiento.',
        }),
        ('Mensaje inicial', {
            'classes': ['tab'],
            'fields': ('template_twilio_id',),
            'description': (
                'Mensaje de bienvenida / inicio que recibe el estudiante por WhatsApp. '
                'Internamente usa Content SID (HX…). '
                'Consola Twilio: https://console.twilio.com/us1/develop/sms/content-editor'
            ),
        }),
        ('Inicio de curso / aviso Aprende', {
            'classes': ['tab'],
            'fields': ('es_campana_curso', 'curso_destino'),
            'description': (
                '<strong>Curso Clases / Aprende (informativo):</strong> deje «inicio de curso» en NO, '
                'elija el curso de clases (ej. Cenipalma — Clases Aprende) y ponga el Content SID del aviso. '
                'Inscribe en Aprende sin Habeas/*listo*. '
                '<br><strong>Curso WhatsApp clásico:</strong> marque inicio = SÍ + curso destino '
                '(Habeas → verificación → avance con *listo*).'
            ),
        }),
        ('Plantilla', {
            'classes': ['tab'],
            'fields': ('plantilla',),
            'description': 'Diseño del mensaje en eki (alternativa al Content SID directo). Requiere plantilla aprobada.',
        }),
        ('Audiencia', {
            'classes': ['tab'],
            'fields': ('tipo_audiencia', 'grupo', 'grupo_estudiantes_preview', 'destinatarios'),
            'description': 'Participantes: individual = elegidos. Grupo = todos los del grupo.',
        }),
        ('Programar', {
            'classes': ['tab'],
            'fields': ('fecha_programada',),
            'description': 'Opcional: envío automático en esa fecha/hora.',
        }),
        ('Excel', {
            'classes': ['tab'],
            'fields': ('archivo_excel',),
            'description': 'Opcional: columnas A (Nombre) y B (Teléfono).',
        }),
        ('Resultados', {
            'classes': ['tab'],
            'fields': ('total_enviados', 'respuestas_si', 'respuestas_no'),
            'description': 'Contadores de envío y respuestas.',
        }),
    )

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        participantes = '—'
        extra_context['eki_camp_audiencia'] = None
        extra_context['eki_camp_envios'] = None
        extra_context['eki_camp_export_fallidos_url'] = None
        if object_id:
            obj = self.get_object(request, object_id)
            if obj:
                if getattr(obj, 'tipo_audiencia', None) == 'grupo' and obj.grupo_id:
                    participantes = f'{obj.grupo.estudiantes.filter(activo=True).count()} estudiantes'
                else:
                    n = obj.destinatarios.filter(activo=True).count()
                    participantes = f'{n} estudiante(s)' if n else 'Sin destinatarios'
                try:
                    from core.campana_resultados import (
                        revisar_audiencia_campana,
                        resumen_envios_campana,
                    )
                    from django.urls import reverse

                    extra_context['eki_camp_audiencia'] = revisar_audiencia_campana(obj)
                    extra_context['eki_camp_envios'] = resumen_envios_campana(obj)
                    extra_context['eki_camp_export_fallidos_url'] = reverse(
                        'admin:core_campana_export_fallidos',
                        args=[obj.pk],
                    )
                except Exception:
                    pass
        extra_context['eki_camp_participantes'] = participantes
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)

    def get_urls(self):
        from django.urls import path

        urls = super().get_urls()
        custom = [
            path(
                '<path:object_id>/exportar-fallidos/',
                self.admin_site.admin_view(self.export_fallidos_view),
                name='core_campana_export_fallidos',
            ),
        ]
        return custom + urls

    def export_fallidos_view(self, request, object_id):
        from django.http import HttpResponse
        from django.shortcuts import get_object_or_404

        from core.campana_resultados import csv_fallidos_campana
        from core.models import Campana

        campana = get_object_or_404(Campana, pk=object_id)
        csv_data = csv_fallidos_campana(campana)
        resp = HttpResponse(csv_data, content_type='text/csv; charset=utf-8')
        safe = ''.join(c if c.isalnum() or c in '-_' else '_' for c in (campana.nombre or 'campana'))[:40]
        resp['Content-Disposition'] = f'attachment; filename="fallidos_{safe}.csv"'
        return resp
    
    def cliente_nombre(self, obj):
        """Muestra el cliente de la campaña"""
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;font-style:italic;">Sin cliente</span>')
    cliente_nombre.short_description = "Cliente"

    def get_readonly_fields(self, request, obj=None):
        return tuple(super().get_readonly_fields(request, obj)) + ('grupo_estudiantes_preview',)

    def grupo_estudiantes_preview(self, obj):
        grupos = GrupoEstudiantes.objects.prefetch_related('estudiantes').order_by('nombre')
        map_grupos = {}
        for grupo in grupos:
            estudiantes = [
                {
                    "nombre": (e.nombre or '').strip() or f"ID {e.id}",
                    "telefono": e.telefono or "",
                }
                for e in grupo.estudiantes.all().order_by('nombre')[:300]
            ]
            map_grupos[str(grupo.id)] = {
                "nombre": f"{grupo.emoji or '👥'} {grupo.nombre}",
                "total": grupo.estudiantes.count(),
                "estudiantes": estudiantes,
            }

        data_json = json.dumps(map_grupos, ensure_ascii=False)
        return mark_safe(f"""
            <div id="grupo-preview-box" style="border:1px solid #e0e0e0;background:#fafafa;border-radius:8px;padding:10px;max-width:860px;">
                <strong>👀 Estudiantes del grupo seleccionado</strong>
                <div id="grupo-preview-content" style="margin-top:8px;color:#666;">Seleccione un grupo para ver sus destinatarios.</div>
            </div>
            <script>
            (function() {{
                const data = {data_json};
                const selectGrupo = document.getElementById("id_grupo");
                const selectTipo = document.getElementById("id_tipo_audiencia");
                const content = document.getElementById("grupo-preview-content");
                if (!selectGrupo || !content) return;
                function renderGrupo() {{
                    const gid = (selectGrupo.value || "").toString();
                    const tipo = (selectTipo && selectTipo.value) ? selectTipo.value : "";
                    if (tipo && tipo !== "grupo") {{
                        content.innerHTML = "<span style='color:#888;'>La campaña está en modo individual.</span>";
                        return;
                    }}
                    if (!gid || !data[gid]) {{
                        content.innerHTML = "<span style='color:#888;'>Seleccione un grupo para ver sus destinatarios.</span>";
                        return;
                    }}
                    const g = data[gid];
                    const rows = (g.estudiantes || []).slice(0, 80).map(e => `<li>${{e.nombre}} <span style="color:#999;">(${{e.telefono || "sin teléfono"}})</span></li>`).join("");
                    const extra = g.total > 80 ? `<div style="margin-top:6px;color:#999;">... y ${{g.total - 80}} estudiante(s) más.</div>` : "";
                    content.innerHTML = `
                        <div><strong>${{g.nombre}}</strong> — <strong>${{g.total}}</strong> estudiante(s)</div>
                        <ul style="margin:8px 0 0 18px;max-height:220px;overflow:auto;">${{rows || "<li>Sin estudiantes</li>"}}</ul>
                        ${{extra}}
                    `;
                }}
                selectGrupo.addEventListener("change", renderGrupo);
                if (selectTipo) selectTipo.addEventListener("change", renderGrupo);
                renderGrupo();
            }})();
            </script>
        """)
    grupo_estudiantes_preview.short_description = "Vista previa de destinatarios del grupo"
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Filtrar plantillas según la categoría seleccionada en la campaña"""
        if db_field.name == "plantilla":
            # Intentar obtener la categoría desde POST (cuando está guardando)
            categoria = request.POST.get('categoria') or request.GET.get('categoria')
            
            if categoria and categoria != 'todas':
                kwargs["queryset"] = Plantilla.objects.filter(categoria=categoria, activa=True)
            else:
                # Si no hay categoría o es 'todas', mostrar todas las plantillas activas
                kwargs["queryset"] = Plantilla.objects.filter(activa=True)
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def categoria_badge(self, obj):
        """Muestra la categoría con badge"""
        if obj.categoria and obj.categoria != 'todas':
            return format_html(
                '<span style="background:#e3f2fd;color:#1976d2;padding:6px 12px;border-radius:12px;font-weight:bold;">{}</span>',
                obj.get_categoria_display()
            )
        return format_html('<span style="color:#999;">Todas las categorías</span>')
    categoria_badge.short_description = "🏷️ Categoría"
    
    def tipo_audiencia_display(self, obj):
        """Muestra si es por grupo o individual"""
        if hasattr(obj, 'tipo_audiencia') and obj.tipo_audiencia == 'grupo' and obj.grupo:
            return format_html(
                '<span style="background:#fff3e0;color:#e65100;padding:4px 10px;border-radius:12px;font-size:11px;">👥 Grupo: {}</span>',
                obj.grupo.nombre
            )
        return format_html('<span style="color:#666;">👤 Individual</span>')
    tipo_audiencia_display.short_description = "Tipo"
    
    @admin.action(description='🚀 Ejecutar Campaña (Enviar Mensajes)')
    def enviar_campana_accion(self, request, queryset):
        """Encola el envío masivo en segundo plano (evita 504 con +90 destinatarios)."""
        from core.services import encolar_ejecutar_campana
        
        for campana in queryset:
            # Permitir re-envío de campañas ya ejecutadas
            if campana.ejecutada:
                self.message_user(
                    request, 
                    f"ℹ️ '{campana.nombre}' ya fue enviada antes. Re-enviando...", 
                    level=messages.INFO
                )
            
            # VALIDACIÓN: necesita template_twilio_id O plantilla aprobada
            if campana.template_twilio_id:
                # Envío directo con Content SID
                pass  # Válido
            elif campana.plantilla and campana.plantilla.content_sid and campana.plantilla.aprobada_twilio:
                # Envío con plantilla Django aprobada
                pass  # Válido
            else:
                self.message_user(
                    request,
                    f"🚨 '{campana.nombre}': Necesita Content SID de Twilio o una plantilla aprobada. "
                    f"Configura un Content Template en Twilio primero.",
                    level=messages.ERROR
                )
                continue
            
            # Validar que tenga destinatarios (o grupo con al menos un estudiante activo)
            destinatarios_count = campana.destinatarios.filter(activo=True).count()
            if hasattr(campana, 'tipo_audiencia') and campana.tipo_audiencia == 'grupo':
                if not campana.grupo:
                    self.message_user(
                        request,
                        f"⚠️ '{campana.nombre}' no tiene un grupo seleccionado.",
                        level=messages.WARNING
                    )
                    continue
                destinatarios_count = campana.grupo.estudiantes.filter(activo=True).count()

            if destinatarios_count == 0:
                self.message_user(
                    request,
                    f"⚠️ '{campana.nombre}' no tiene destinatarios: no se envía a nadie.",
                    level=messages.WARNING
                )
                continue

            try:
                from core.campana_resultados import revisar_audiencia_campana

                aud = revisar_audiencia_campana(campana)
                if aud['n_error']:
                    self.message_user(
                        request,
                        f"⚠️ '{campana.nombre}': {aud['n_error']} teléfono(s) dudoso(s) "
                        f"(faltan dígitos o código de país). Abra la campaña para ver el detalle; "
                        f"el envío continúa con el resto.",
                        level=messages.WARNING,
                    )
            except Exception:
                pass

            try:
                modo = encolar_ejecutar_campana(campana.id)
                detalle_modo = 'Celery' if modo == 'celery' else 'proceso en background'
                self.message_user(
                    request,
                    f"⏳ '{campana.nombre}' encolada ({detalle_modo}): "
                    f"se enviará a {destinatarios_count} destinatarios en segundo plano. "
                    f"Actualiza el listado en unos minutos para ver el resultado.",
                    level=messages.SUCCESS,
                )
            except Exception as e:
                self.message_user(
                    request,
                    f"❌ Error encolando '{campana.nombre}': {str(e)}",
                    level=messages.ERROR
                )
    
    enviar_campana_accion.short_description = "📤 Ejecutar campañas seleccionadas (envío real por WhatsApp)"
    
    def estado_visual(self, obj):
        if obj.ejecutada:
            return format_html('<span style="color: green;">✅ Ejecutada</span>')
        return format_html('<span style="color: orange;">⏳ Pendiente</span>')
    estado_visual.short_description = "Estado"
    
    def conteo_destinatarios(self, obj):
        """Muestra cantidad de destinatarios según tipo de audiencia"""
        if hasattr(obj, 'tipo_audiencia') and obj.tipo_audiencia == 'grupo' and obj.grupo:
            count = obj.grupo.estudiantes.count()
            return format_html(
                '<span style="background:#fff3e0;padding:3px 8px;border-radius:8px;font-size:11px;">👥 {} estudiantes</span>',
                count
            )
        count = obj.destinatarios.count()
        return format_html(
            '<span style="background:#e3f2fd;padding:3px 8px;border-radius:8px;font-size:11px;">👤 {} individual{}</span>',
            count, 'es' if count != 1 else ''
        )
    conteo_destinatarios.short_description = "Destinatarios"
    
    def programada_display(self, obj):
        """Muestra si está programada"""
        if obj.fecha_programada:
            return format_html(
                '<span style="background:#e8f5e9;color:#2e7d32;padding:3px 8px;border-radius:8px;font-size:11px;">📅 {}</span>',
                obj.fecha_programada.strftime('%d/%m %H:%M')
            )
        return format_html('<span style="color:#999;">-</span>')
    programada_display.short_description = "Programada"

    def plantilla_estado(self, obj):
        """Muestra el estado del Content Template de la plantilla"""
        if obj.plantilla:
            if obj.plantilla.twilio_template_sid and obj.plantilla.aprobada_twilio:
                return format_html('<span style="background:#4caf50;color:white;padding:2px 6px;border-radius:8px;font-size:10px;">✅ TWILIO</span>')
            elif obj.plantilla.twilio_template_sid and not obj.plantilla.aprobada_twilio:
                return format_html('<span style="background:#ff9800;color:white;padding:2px 6px;border-radius:8px;font-size:10px;">⏳ PENDIENTE</span>')
            else:
                return format_html('<span style="background:#2196f3;color:white;padding:2px 6px;border-radius:8px;font-size:10px;">📱 DIRECTO</span>')
        return format_html('<span style="color:#999;font-size:10px;">Sin plantilla</span>')
    plantilla_estado.short_description = "Template"


@admin.register(MensajePush)
class MensajePushAdmin(admin.ModelAdmin):
    """Recordatorios WhatsApp que no reinician el curso (responder *listo*)."""
    list_display = ('nombre', 'tipo', 'cliente', 'curso', 'activo', 'fecha_creacion')
    list_filter = ('tipo', 'activo', 'cliente')
    search_fields = ('nombre', 'cuerpo_fallback', 'twilio_content_sid')
    actions = ['enviar_push_seleccionados']
    fieldsets = (
        (None, {'fields': ('nombre', 'tipo', 'activo')}),
        ('Audiencia', {'fields': ('cliente', 'curso')}),
        ('Contenido Twilio / texto', {
            'fields': ('twilio_content_sid', 'plantilla', 'cuerpo_fallback', 'incluir_boton_continuar'),
            'description': (
                'Recordatorio únicamente: no cambia módulo ni progreso. '
                'Texto libre (cuerpo fallback) basta sin Twilio; o use Content SID (HX…). '
                'Variables: {nombre}, {curso}.'
            ),
        }),
    )

    @admin.action(description='📲 Enviar push a audiencia del mensaje')
    def enviar_push_seleccionados(self, request, queryset):
        from core.mensajes_push import enviar_mensaje_push_masivo
        for mp in queryset.filter(activo=True):
            r = enviar_mensaje_push_masivo(mp)
            self.message_user(
                request,
                f'{mp.nombre}: {r["enviados"]} enviados, {r["errores"]} errores.',
            )


@admin.register(EnvioMensajePush)
class EnvioMensajePushAdmin(admin.ModelAdmin):
    list_display = ('mensaje_push', 'estudiante', 'telefono', 'exito', 'fecha')
    list_filter = ('exito', 'mensaje_push')
    readonly_fields = ('mensaje_push', 'estudiante', 'telefono', 'exito', 'detalle', 'fecha')


@admin.register(EnlaceFormularioExterno)
class EnlaceFormularioExternoAdmin(admin.ModelAdmin):
    """Google Form → habilitar módulo al enviar respuesta."""
    list_display = ('nombre', 'cliente', 'curso', 'modulo_resumen', 'campo_identificador', 'activo', 'creado_en')
    list_filter = ('activo', 'cliente', 'campo_identificador')
    search_fields = ('nombre', 'token', 'cliente__nombre', 'curso__nombre')
    readonly_fields = ('token', 'webhook_url', 'instrucciones_google', 'creado_en')
    autocomplete_fields = ('cliente', 'curso', 'modulo')
    fieldsets = (
        (None, {'fields': ('nombre', 'activo', 'cliente', 'curso', 'modulo', 'campo_identificador', 'notas')}),
        ('Webhook (copiar a Google Apps Script)', {
            'fields': ('token', 'webhook_url', 'instrucciones_google'),
        }),
        ('Auditoría', {'fields': ('creado_en',), 'classes': ('collapse',)}),
    )

    def modulo_resumen(self, obj):
        if obj.modulo_id:
            return f'M{obj.modulo.numero}'
        return 'Último módulo'
    modulo_resumen.short_description = 'Módulo'

    def webhook_url(self, obj):
        if not obj or not obj.token:
            return 'Guarde el registro para generar el token.'
        from django.urls import reverse
        path = reverse('api_form_externo_webhook', args=[obj.token])
        base = 'https://eki-prod-final.eba-32krwxas.us-east-2.elasticbeanstalk.com'
        return format_html('<code style="word-break:break-all;">{}{}</code>', base, path)
    webhook_url.short_description = 'URL POST'

    def instrucciones_google(self, obj):
        if not obj or not obj.pk:
            return 'Guarde primero para ver el script de ejemplo.'
        campo = obj.campo_identificador
        campos_txt = {
            'cedula': 'solo documento/ID (cédula, DUI, CURP, DNI…)',
            'telefono': 'teléfono WhatsApp del curso',
            'cedula_y_telefono': 'documento + teléfono WhatsApp (recomendado: ambos deben coincidir)',
            'cedula_y_nombre': 'documento + nombre completo (el nombre debe parecerse al registro)',
        }.get(campo, campo)
        payload_ej = {
            'cedula': '{"cedula": cedula}',
            'telefono': '{"telefono": telefono}',
            'cedula_y_telefono': '{"cedula": cedula, "telefono": telefono}',
            'cedula_y_nombre': '{"cedula": cedula, "nombre": nombre}',
        }.get(campo, '{"cedula": cedula}')
        return format_html(
            '<div style="font-size:12px;line-height:1.55;max-width:720px;">'
            '<p><strong>1.</strong> En Google Form → ⋮ → Editor de secuencias de comandos.</p>'
            '<p><strong>2.</strong> Activadores → Al enviar el formulario → pegar función que llame a la URL POST.</p>'
            '<p><strong>3.</strong> El formulario debe pedir: <strong>{}</strong>.</p>'
            '<p><strong>4.</strong> Si falla la validación, el estudiante no queda habilitado; '
            'revise en «Registros formulario externo» o habilite manual en Acceso módulos.</p>'
            '<p><strong>5.</strong> Active «Módulos solo por lista» en el cliente para que el acceso surta efecto.</p>'
            '<pre style="background:#f1f5f9;padding:10px;border-radius:8px;overflow:auto;">'
            'function onSubmit(e) {{\n'
            '  var cedula = "", telefono = "", nombre = "";\n'
            '  e.response.getItemResponses().forEach(function(r) {{\n'
            '    var t = r.getItem().getTitle().toLowerCase();\n'
            '    if (t.indexOf("cedula") >= 0 || t.indexOf("documento") >= 0 || t.indexOf("identificacion") >= 0 || t.indexOf("dui") >= 0 || t.indexOf("curp") >= 0 || t.indexOf("dni") >= 0)\n'
            '      cedula = String(r.getResponse()).replace(/[\\s.\\-]/g, "").toUpperCase();\n'
            '  if (t.indexOf("whatsapp") >= 0 || t.indexOf("telefono") >= 0 || t.indexOf("celular") >= 0)\n'
            '      telefono = String(r.getResponse()).replace(/\\D/g, "");\n'
            '    if (t.indexOf("nombre") >= 0)\n'
            '      nombre = r.getResponse();\n'
            '  }});\n'
            '  UrlFetchApp.fetch("{url}", {{\n'
            '    method: "post",\n'
            '    contentType: "application/json",\n'
            '    payload: JSON.stringify({payload}),\n'
            '    muteHttpExceptions: true\n'
            '  }});\n'
            '}}'
            '</pre></div>',
            campos_txt,
            url=f'https://eki-prod-final.eba-32krwxas.us-east-2.elasticbeanstalk.com/api/integracion/form-externo/{obj.token}/',
            payload=payload_ej,
        )
    instrucciones_google.short_description = 'Cómo conectar Google Form'


@admin.register(RegistroFormularioExterno)
class RegistroFormularioExternoAdmin(admin.ModelAdmin):
    list_display = ('fecha', 'enlace', 'estudiante', 'identificador_recibido', 'exito', 'detalle')
    list_filter = ('exito', 'enlace')
    readonly_fields = ('enlace', 'estudiante', 'identificador_recibido', 'exito', 'detalle', 'payload', 'fecha')
    ordering = ('-fecha',)


@admin.register(EnvioLog)
class EnvioLogAdmin(admin.ModelAdmin):
    """Historial de envíos de campañas"""
    list_display = ('campana_nombre', 'estudiante_info', 'estado_badge', 'fecha_envio')
    list_filter = ('estado', 'campana', 'fecha_envio')
    search_fields = ('estudiante__nombre', 'estudiante__cedula', 'estudiante__telefono', 'campana__nombre')
    readonly_fields = ('campana', 'estudiante', 'estado', 'respuesta_api', 'fecha_envio')
    date_hierarchy = 'fecha_envio'
    list_per_page = 50
    
    fieldsets = (
        ('📤 Información del Envío', {
            'fields': ('campana', 'estudiante', 'fecha_envio')
        }),
        ('📊 Estado', {
            'fields': ('estado', 'respuesta_api')
        }),
    )
    
    def campana_nombre(self, obj):
        return format_html(
            '<strong>{}</strong><br><small style="color:#666;">{}</small>',
            obj.campana.nombre,
            obj.campana.plantilla.nombre_interno if obj.campana.plantilla else ''
        )
    campana_nombre.short_description = "Campaña"
    
    def estudiante_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small style="color:#666;">📱 +{}</small>',
            obj.estudiante.nombre,
            obj.estudiante.telefono
        )
    estudiante_info.short_description = "Estudiante"
    
    def estado_badge(self, obj):
        colores = {
            'ENVIADO': ('#e8f5e9', '#2e7d32'),
            'FALLIDO': ('#ffebee', '#c62828'),
            'PENDIENTE': ('#fff3e0', '#e65100'),
        }
        bg, color = colores.get(obj.estado, ('#f5f5f5', '#666'))
        return format_html(
            '<span style="background:{};color:{};padding:4px 10px;border-radius:12px;font-size:11px;font-weight:bold;">{}</span>',
            bg, color, obj.estado
        )
    estado_badge.short_description = "Estado"


@admin.register(WhatsappLog)
class WhatsappLogAdmin(admin.ModelAdmin):
    """Registro de todas las conversaciones del chatbot"""
    list_display = ('fecha', 'estudiante_nombre', 'telefono_corto', 'tipo_badge', 'mensaje_preview', 'estado_badge', 'estado_visual', 'actividad_badge')
    list_filter = ('tipo', 'estado', 'fecha', 'estudiante__activo', 'agente_usado')
    search_fields = ('telefono', 'mensaje', 'mensaje_id', 'estudiante__nombre')  # ✅ Búsqueda por nombre
    date_hierarchy = 'fecha'
    list_per_page = 100
    ordering = ('-fecha',)
    readonly_fields = ('fecha', 'mensaje_id', 'estudiante')
    actions = ['exportar_conversaciones_excel', 'exportar_conversaciones_csv', 'marcar_como_procesado']
    autocomplete_fields = ['estudiante']  # ✅ Autocompletar estudiante
    
    fieldsets = (
        ('Información del Mensaje', {
            'fields': ('telefono', 'estudiante', 'tipo', 'mensaje', 'estado')
        }),
        ('Metadatos', {
            'fields': ('mensaje_id', 'fecha'),
            'classes': ('collapse',)
        }),
    )
    
    def estudiante_nombre(self, obj):
        """Muestra nombre del estudiante si está asignado"""
        if obj.estudiante:
            return obj.estudiante.nombre
        return format_html('<span style="color:#999;font-style:italic;">Sin asignar</span>')
    estudiante_nombre.short_description = "👤 Estudiante"
    estudiante_nombre.admin_order_field = 'estudiante__nombre'
    
    def telefono_corto(self, obj):
        """Muestra solo los últimos 4 dígitos"""
        return f"...{obj.telefono[-4:]}"
    telefono_corto.short_description = "📱"
    
    def tipo_badge(self, obj):
        """Badge visual para tipo de mensaje"""
        if obj.tipo == 'INCOMING':
            return format_html(
                '<span style="background:#4caf50;color:white;padding:3px 8px;border-radius:12px;font-size:11px;">⬇️ RECIBIDO</span>'
            )
        return format_html(
            '<span style="background:#2196f3;color:white;padding:3px 8px;border-radius:12px;font-size:11px;">⬆️ ENVIADO</span>'
        )
    tipo_badge.short_description = "Tipo"
    
    def mensaje_preview(self, obj):
        """Muestra preview del mensaje"""
        texto = obj.mensaje[:60] + "..." if len(obj.mensaje) > 60 else obj.mensaje
        return texto
    mensaje_preview.short_description = "💬 Mensaje"
    
    def estado_badge(self, obj):
        """Badge visual para estado"""
        colores = {
            'RECIBIDO': '#4caf50',
            'SENT': '#2196f3',
            'PENDING': '#ff9800',
            'ERROR': '#f44336'
        }
        color = colores.get(obj.estado, '#999')
        return format_html(
            '<span style="background:{};color:white;padding:3px 8px;border-radius:12px;font-size:11px;">{}</span>',
            color, obj.estado
        )
    estado_badge.short_description = "Estado"
    
    def estado_visual(self, obj):
        """Indicador visual de éxito/error"""
        if obj.estado == 'SENT' or obj.estado == 'RECIBIDO':
            return format_html('<span style="font-size:18px;color:#4caf50;">✅</span>')
        elif obj.estado == 'ERROR':
            return format_html('<span style="font-size:18px;color:#f44336;">❌</span>')
        else:
            return format_html('<span style="font-size:18px;color:#ff9800;">⏳</span>')
    estado_visual.short_description = "✓"
    
    def actividad_badge(self, obj):
        """Badge de actividad del mensaje"""
        from django.utils import timezone
        from datetime import timedelta
        
        now = timezone.now()
        delta = now - obj.fecha
        
        if delta.total_seconds() < 86400:  # 24 horas
            return format_html('<span style="color:#f44336;font-weight:bold;">🔴 Nueva</span>')
        elif delta.days < 7:
            return format_html('<span style="color:#4caf50;font-weight:bold;">🟢 Activa</span>')
        elif delta.days < 30:
            return format_html('<span style="color:#ff9800;">🟡 Reciente</span>')
        else:
            return format_html('<span style="color:#999;">⚪ Antigua</span>')
    actividad_badge.short_description = "⏰ Actividad"
    
    def get_queryset(self, request):
        """Ordena por fecha descendente por defecto"""
        qs = super().get_queryset(request)
        return qs.order_by('-fecha')
    
    @admin.action(description='📊 Exportar conversaciones a Excel')
    def exportar_conversaciones_excel(self, request, queryset):
        """Exporta conversaciones seleccionadas a archivo Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conversaciones"
        
        # Encabezados
        headers = ['Fecha', 'Teléfono', 'Tipo', 'Mensaje', 'Estado', 'ID Mensaje']
        ws.append(headers)
        
        # Estilo
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for log in queryset.order_by('fecha'):
            ws.append([
                log.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                log.telefono,
                log.tipo,
                log.mensaje[:500],  # Limitar tamaño
                log.estado,
                log.mensaje_id or 'N/A'
            ])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 35
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'conversaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @admin.action(description='📄 Exportar conversaciones a CSV')
    def exportar_conversaciones_csv(self, request, queryset):
        """Exporta conversaciones seleccionadas a archivo CSV"""
        import csv
        from datetime import datetime
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'conversaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Fecha', 'Teléfono', 'Tipo', 'Mensaje', 'Estado', 'ID Mensaje'])
        
        for log in queryset.order_by('fecha'):
            writer.writerow([
                log.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                log.telefono,
                log.tipo,
                log.mensaje[:500],
                log.estado,
                log.mensaje_id or 'N/A'
            ])
        
        return response
    
    @admin.action(description='✅ Marcar como procesado')
    def marcar_como_procesado(self, request, queryset):
        """Marca mensajes como procesados"""
        actualizado = queryset.filter(estado__in=['RECIBIDO', 'PENDING']).update(estado='SENT')
        self.message_user(request, f'✅ {actualizado} mensaje(s) marcado(s) como procesado', messages.SUCCESS)



# Personalizar el admin site
admin.site.site_header = "eki - Chatbot Agro 🌱"


