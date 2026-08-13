from core.admin._common import *  # noqa: F401,F403
from core.orden_bloques import (
    intercambiar_orden,
    preparar_ordenes_temporales,
    renumerar_orden_1_based,
)
from unfold.decorators import action
from unfold.widgets import UnfoldAdminFileFieldWidget

# ==========================================
# SISTEMA EDUCATIVO - ADMINISTRACIÓN
# ==========================================

# Alta de módulo: 1 sección + N microcontenidos vacíos (inactivos hasta completar).
MODULO_ALTA_PASOS_PLANTILLA = 3


def _botones_mover_bloque(obj, kind: str):
    """Enlaces ↑↓ nativos (Unfold): solo si el registro ya tiene PK."""
    if not obj or not obj.pk or not getattr(obj, 'modulo_id', None):
        return format_html(
            '<span style="color:#94a3b8;font-size:12px;">Guarde para reordenar</span>'
        )
    up = reverse(
        'admin:core_modulo_mover_bloque',
        args=[obj.modulo_id, kind, obj.pk, 'up'],
    )
    down = reverse(
        'admin:core_modulo_mover_bloque',
        args=[obj.modulo_id, kind, obj.pk, 'down'],
    )
    return format_html(
        '<span class="eki-orden-btns" style="display:inline-flex;gap:4px;align-items:center;">'
        '<a class="button" href="{}" title="Subir" style="padding:2px 8px;min-width:2rem;text-align:center;">↑</a>'
        '<a class="button" href="{}" title="Bajar" style="padding:2px 8px;min-width:2rem;text-align:center;">↓</a>'
        '<span style="color:#6b6575;font-size:12px;margin-left:4px;">#{}</span>'
        '</span>',
        up,
        down,
        obj.orden,
    )


def sembrar_plantilla_modulo(modulo, *, n_pasos: int = MODULO_ALTA_PASOS_PLANTILLA) -> dict:
    """
    Tras el alta: garantiza 1 sección y N microcontenidos vacíos (activo=False).
    No pisa secciones/pasos que el operador ya haya cargado en el mismo POST.
    Cursos modo clases: 1 solo paso, sin exigir *listo*.
    """
    created = {'seccion': False, 'pasos': 0}
    if not modulo or not modulo.pk:
        return created

    curso = getattr(modulo, 'curso', None)
    modo_clases = bool(curso and getattr(curso, 'es_modo_clases', lambda: False)())
    if modo_clases:
        n_pasos = 1

    seccion = modulo.secciones.order_by('orden', 'id').first()
    if seccion is None:
        if modo_clases:
            titulo_sec = (modulo.titulo or '').strip() or 'Clase'
        else:
            titulo_sec = 'Bloque 1'
        seccion = SeccionModulo.objects.create(
            modulo=modulo,
            orden=1,
            titulo=titulo_sec[:200],
            activa=True,
        )
        created['seccion'] = True

    if modulo.pasos.exists():
        return created

    for i in range(1, max(1, int(n_pasos)) + 1):
        PasoModulo.objects.create(
            modulo=modulo,
            seccion=seccion,
            orden=i,
            titulo=('Bienvenida' if modo_clases and i == 1 else f'Microcontenido {i}'),
            contenido='',
            activo=False,
            requiere_listo_para_avanzar=not modo_clases,
        )
        created['pasos'] += 1
    return created


def asegurar_seccion_y_primer_paso(modulo) -> tuple:
    """Garantiza 1 sección + 1er paso (no borra pasos extra)."""
    seccion = modulo.secciones.order_by('orden', 'id').first()
    if seccion is None:
        titulo_sec = (modulo.titulo or '').strip() or f'Clase {modulo.numero}'
        seccion = SeccionModulo.objects.create(
            modulo=modulo,
            orden=1,
            titulo=titulo_sec[:200],
            activa=True,
        )
    elif not (seccion.titulo or '').strip():
        seccion.titulo = ((modulo.titulo or '').strip() or f'Clase {modulo.numero}')[:200]
        seccion.save(update_fields=['titulo'])

    paso = modulo.pasos.order_by('orden', 'id').first()
    curso = getattr(modulo, 'curso', None)
    modo_clases = bool(curso and getattr(curso, 'es_modo_clases', lambda: False)())
    if paso is None:
        paso = PasoModulo.objects.create(
            modulo=modulo,
            seccion=seccion,
            orden=1,
            titulo='Bienvenida',
            contenido='',
            activo=False,
            requiere_listo_para_avanzar=not modo_clases,
        )
    elif paso.seccion_id != seccion.id:
        paso.seccion = seccion
        paso.save(update_fields=['seccion'])
    return seccion, paso


def aplicar_clase_simple_desde_form(form, formsets=None) -> PasoModulo | None:
    """
    Sincroniza pestaña Clase → 1er PasoModulo (texto / media / activo).
    Solo campos Clase realmente editados. Si Materiales guardó el 1er paso
    en el mismo POST, no pisar su contenido (bug Agrosavia / varios micros).
    """
    modulo = form.instance
    if not modulo or not modulo.pk:
        return None
    if not _form_toco_clase_simple(form):
        return None
    cd = getattr(form, 'cleaned_data', None) or {}
    changed = getattr(form, 'changed_data', None)
    strict = changed is not None
    seccion, paso = asegurar_seccion_y_primer_paso(modulo)

    inline_guardo_contenido = _formsets_guardaron_contenido_paso(formsets, paso)

    update_fields = []
    texto = (cd.get('clase_texto') or '').strip()
    tocar_texto = (
        (strict and 'clase_texto' in changed)
        or (not strict and bool(texto))
    )
    if tocar_texto and not inline_guardo_contenido:
        if texto != (paso.contenido or ''):
            paso.contenido = texto
            update_fields.append('contenido')

    titulo_paso = (cd.get('titulo') or modulo.titulo or '').strip() or 'Bienvenida'
    if not (paso.titulo or '').strip() or paso.titulo.startswith('Microcontenido'):
        paso.titulo = titulo_paso[:200]
        update_fields.append('titulo')

    pending = getattr(form, '_clase_pending_media_url', None)
    url_manual = (cd.get('clase_url') or '').strip()
    tocar_url = bool(pending) or (
        (strict and 'clase_url' in changed and url_manual)
        or (not strict and url_manual)
    )
    new_url = (pending or url_manual or '').strip()
    if tocar_url and new_url and new_url != (paso.media_url or ''):
        paso.media_url = new_url
        update_fields.append('media_url')

    tocar_activo = (
        (strict and 'clase_activo' in changed)
        or (not strict and 'clase_activo' in cd)
    )
    if tocar_activo:
        activo = bool(cd.get('clase_activo'))
        if activo != bool(paso.activo):
            paso.activo = activo
            update_fields.append('activo')

    if seccion.titulo != ((modulo.titulo or '').strip() or seccion.titulo):
        seccion.titulo = ((modulo.titulo or '').strip() or seccion.titulo)[:200]
        seccion.save(update_fields=['titulo'])

    if update_fields:
        paso.save(update_fields=list(dict.fromkeys(update_fields)))
    return paso if update_fields else None


def _formsets_guardaron_contenido_paso(formsets, paso) -> bool:
    """True si el inline Materiales envió contenido para el mismo paso en este POST."""
    if not formsets or not paso:
        return False
    for fs in formsets:
        if getattr(fs, 'model', None) is not PasoModulo:
            continue
        for f in fs.forms:
            cleaned = getattr(f, 'cleaned_data', None) or {}
            if cleaned.get('DELETE'):
                continue
            inst = getattr(f, 'instance', None)
            if inst is None:
                continue
            same_pk = getattr(inst, 'pk', None) and inst.pk == paso.pk
            if not same_pk:
                continue
            changed = getattr(f, 'changed_data', None)
            if changed is not None and 'contenido' in changed:
                return True
            if (cleaned.get('contenido') or '').strip() and f.has_changed():
                return True
    return False


def _form_toco_clase_simple(form) -> bool:
    """True si el POST/form tocó Clase (evita overwrite Materiales en Guardar genérico)."""
    if getattr(form, '_clase_pending_media_url', None):
        return True
    cd = getattr(form, 'cleaned_data', None) or {}
    if cd.get('clase_archivo') or (cd.get('clase_url') or '').strip():
        return True
    changed = getattr(form, 'changed_data', None)
    if changed is not None:
        return any(
            name in changed
            for name in ('clase_texto', 'clase_archivo', 'clase_url', 'clase_activo')
        )
    if (cd.get('clase_texto') or '').strip():
        return True
    if cd.get('clase_activo') is True:
        return True
    return False




class ModuloInline(admin.TabularInline):
    """Módulos dentro del curso"""
    model = Modulo
    extra = 1
    fields = ('numero', 'titulo', 'descripcion', 'duracion_dias')
    ordering = ['numero']


class DocumentoRAGInline(admin.StackedInline):
    """Documentos RAG para la base de conocimiento IA del curso"""
    model = DocumentoRAG
    extra = 0
    verbose_name = '📄 Documento RAG'
    verbose_name_plural = 'DOCUMENTOS RAG — Base de Conocimiento para Agentes IA'
    readonly_fields = ('estado_badge', 'chunks_indexados', 'fecha_subida', 'fecha_indexado')
    fields = ('nombre', 'archivo', 'tipo', 'descripcion', 'estado_badge', 'chunks_indexados', 'fecha_subida', 'fecha_indexado')

    def estado_badge(self, obj):
        if not obj.pk:
            return '-'
        colors = {'pendiente': '#ffc107', 'indexado': '#28a745', 'error': '#dc3545'}
        color = colors.get(obj.estado, '#6c757d')
        return format_html(
            '<span style="background:{};color:white;padding:3px 10px;border-radius:12px;font-size:11px;">{}</span>',
            color, obj.get_estado_display()
        )
    estado_badge.short_description = 'Estado RAG'


class PreguntaAbiertaFinalInline(admin.TabularInline):
    """Preguntas abiertas finales dentro del curso (máximo 3)."""
    model = PreguntaAbiertaFinalCurso
    extra = 1
    max_num = 3
    fields = ('orden', 'pregunta', 'activa')
    ordering = ('orden', 'id')
    verbose_name = '📝 Pregunta Abierta Final'
    verbose_name_plural = 'PREGUNTAS ABIERTAS FINALES (MAX 3)'


@admin.register(Curso)
class CursoAdmin(admin.ModelAdmin):
    """Administración de cursos"""
    change_form_template = 'admin/core/curso/change_form.html'
    # Listado limpio P1: ≤6 columnas (RAG/GEI/IA/orden → ficha o acciones).
    list_display = (
        'nombre',
        'cliente_nombre',
        'experiencia_display',
        'total_modulos_display',
        'activo',
        'visible_en_studio',
    )
    list_filter = (
        'activo', 'modo_aula', 'visible_en_studio', 'visible_en_aula', 'cliente',
        'usar_gamificacion', 'usar_agentes_ia', 'tiene_formulario_gei',
    )
    search_fields = ('nombre', 'descripcion', 'cliente__nombre')
    list_editable = ()
    inlines = [ModuloInline, DocumentoRAGInline, PreguntaAbiertaFinalInline]
    actions = [
        'ver_todos_modulos', 'añadir_clase_rapida',
        'indexar_documentos_rag', 'indexar_contenido_modulos',
        'activar_cursos', 'desactivar_cursos', 'copiar_a_otro_cliente', 'copiar_a_analytics_pruebas',
    ]
    # change_list_template = 'admin/curso_changelist.html'  # Eliminado para usar el template estándar de Django
    
    fieldsets = (
        ('Datos del curso', {
            'fields': (
                'nombre', 'descripcion', 'cliente', 'duracion_semanas',
                'activo', 'visible_en_studio', 'visible_en_aula', 'modo_aula', 'orden',
            ),
            'description': mark_safe(
                '<p><strong>Cómo elegir el proceso del curso</strong> '
                '(campo <em>Experiencia en aula</em> — también se ve en el listado)</p>'
                '<ul style="margin:0.4rem 0 0;padding-left:1.2rem;">'
                '<li><strong>Módulos (WhatsApp + avance)</strong> — flujo clásico con *listo*, '
                'puntos/ranking si la org los tiene activos.</li>'
                '<li><strong>Clases / biblioteca</strong> — informativo en Aprende: '
                'Biblioteca = «mis clases», <em>sin gamificación ni *listo* por WA</em> '
                '(se apagan solos al guardar). Use un nombre claro para la lista Excel '
                '(ej. «Cenipalma - Clases Aprende», guión simple).</li>'
                '</ul>'
            ),
        }),
        ('Ritmo drip y acceso', {
            'classes': ('collapse',),
            'fields': (
                'dias_espera_entre_modulos',
                'usar_gamificacion',
                'habilitar_pregunta_abierta_final',
            ),
            'description': mark_safe(
                '<p><strong>0</strong> = avance inmediato con <em>listo</em>. '
                '<strong>&gt; 0</strong> = días de espera entre módulos.</p>'
                '<p>Override por cliente: tabla <em>Ritmo drip por curso</em> en la ficha Cliente.</p>'
            ),
        }),
        ('IA y retos', {
            'classes': ('collapse',),
            'fields': (
                'usar_agentes_ia',
                'nombre_agente_tutor',
                'nombre_agente_asistente',
                'preguntas_ejemplo_ia',
            ),
            'description': mark_safe(
                '<p>Retos Darío/Claudia, nombres de agentes (override) y preguntas ejemplo para el tutor.</p>'
                '<p>Si los nombres van vacíos, se usan los del Cliente o los por defecto.</p>'
            ),
        }),
        ('GEI y WhatsApp', {
            'classes': ('collapse',),
            'fields': ('tiene_formulario_gei', 'enlace_grupo_whatsapp'),
            'description': mark_safe(
                '<p>Recolección GEI al completar el módulo disparador (Formulario → Tipos de formulario). '
                '<a href="/admin/gei/panel/" target="_blank">Panel GEI</a></p>'
            ),
        }),
    )
    
    def cliente_nombre(self, obj):
        """Muestra si es curso específico de un cliente"""
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;font-style:italic;">General (eki)</span>')
    cliente_nombre.short_description = "Organización"

    @admin.display(description='Experiencia', ordering='modo_aula')
    def experiencia_display(self, obj):
        if obj.es_modo_clases():
            return format_html(
                '<span style="background:#efe8f5;color:#5F3A6E;padding:3px 8px;'
                'border-radius:6px;font-size:11px;font-weight:700;">Clases · Aprende</span>'
            )
        return format_html(
            '<span style="background:#e8f0fe;color:#1a56a0;padding:3px 8px;'
            'border-radius:6px;font-size:11px;font-weight:700;">Módulos · WhatsApp</span>'
        )
    
    def total_modulos_display(self, obj):
        count = obj.modulos.count()
        return format_html(
            '<span style="font-variant-numeric:tabular-nums;">{}</span>',
            count
        )
    total_modulos_display.short_description = "Módulos"
    
    def ver_modulos_link(self, obj):
        """Link para ver todos los módulos del curso"""
        url = f"/admin/core/modulo/?curso__id__exact={obj.id}"
        count = obj.modulos.count()
        
        # Contar archivos multimedia totales
        from django.db.models import Count
        total_archivos = ArchivoModulo.objects.filter(modulo__curso=obj, activo=True).count()
        
        html = f'<a href="{url}" style="color:#2196F3;">📋 {count} módulo(s)</a>'
        if total_archivos > 0:
            html += f' <span style="color:#999;font-size:11px;">• {total_archivos} archivos</span>'
        return format_html(html)
    ver_modulos_link.short_description = "Gestión"

    def docs_rag_count(self, obj):
        """Muestra cantidad de documentos RAG indexados"""
        total = obj.documentos_rag.count()
        indexados = obj.documentos_rag.filter(estado='indexado').count()
        if total == 0:
            return format_html('<span style="color:#999;font-size:11px;">Sin docs</span>')
        color = '#28a745' if indexados == total else '#ffc107'
        return format_html(
            '<span style="background:{};color:white;padding:3px 8px;border-radius:4px;font-size:11px;">'
            '📚 {}/{}</span>',
            color, indexados, total
        )
    docs_rag_count.short_description = "RAG"

    @admin.action(description='🤖 Indexar documentos RAG de cursos seleccionados')
    def indexar_documentos_rag(self, request, queryset):
        """Encola indexación RAG (nunca bloquea el request HTTP)."""
        from core.admin.commercial import _encolar_o_indexar_rag_doc

        queued = 0
        for curso in queryset:
            for doc in curso.documentos_rag.filter(estado__in=['pendiente', 'error']):
                if not doc.archivo:
                    continue
                doc.estado = 'pendiente'
                doc.save(update_fields=['estado'])
                _encolar_o_indexar_rag_doc(
                    request, self, 'DocumentoRAG', doc, show_index_message=False,
                )
                queued += 1
        self.message_user(
            request,
            f'✅ {queued} documentos en cola (segundo plano).'
            if queued else 'Nada pendiente que indexar.',
        )

    @admin.action(description='📝 Indexar contenido de módulos en RAG')
    def indexar_contenido_modulos(self, request, queryset):
        """Indexa el contenido educativo de los módulos en la BD vectorial."""
        from core.rag_manager import rag_manager
        total = 0
        for curso in queryset:
            n = rag_manager.indexar_modulos_curso(curso.id)
            total += n
        self.message_user(request, f"✅ {total} chunks indexados desde contenido de módulos.")

    def save_formset(self, request, form, formset, change):
        """Al guardar DocumentoRAG inline, encolar indexación (no bloquear request)."""
        from core.admin.commercial import _encolar_o_indexar_rag_doc

        instances = formset.save(commit=False)
        for instance in instances:
            if isinstance(instance, DocumentoRAG):
                if not instance.subido_por_id:
                    instance.subido_por = request.user
                instance.save()
                if instance.estado == 'pendiente' and instance.archivo:
                    _encolar_o_indexar_rag_doc(
                        request, self, 'DocumentoRAG', instance, show_index_message=False,
                    )
            else:
                instance.save()
        formset.save_m2m()
        # Manejar eliminaciones
        for obj in formset.deleted_objects:
            if isinstance(obj, DocumentoRAG):
                from core.rag_manager import rag_manager
                rag_manager.eliminar_documento(obj.cliente_id, obj.curso_id, obj.nombre)
            obj.delete()

    @admin.action(description='📋 Ver módulos de cursos seleccionados')
    def ver_todos_modulos(self, request, queryset):
        """Redirige a la vista de módulos filtrando por los cursos seleccionados"""
        from django.shortcuts import redirect
        curso_ids = ','.join(str(c.id) for c in queryset)
        return redirect(f'/admin/core/modulo/?curso__id__in={curso_ids}')

    @admin.action(description='➕ Añadir clase rápida (abre editor Clase)')
    def añadir_clase_rapida(self, request, queryset):
        """Crea módulo + sección + 1 paso vacío y abre pestaña Clase."""
        if queryset.count() != 1:
            self.message_user(
                request,
                'Seleccione un solo curso para añadir la clase.',
                level=messages.ERROR,
            )
            return
        curso = queryset.first()
        last = curso.modulos.order_by('-numero').first()
        numero = (last.numero + 1) if last else 1
        modo_clases = curso.es_modo_clases()
        with transaction.atomic():
            modulo = Modulo.objects.create(
                curso=curso,
                numero=numero,
                titulo=f'Clase {numero}',
                descripcion='',
                contenido='',
                duracion_dias=7,
                modo_entrega=(
                    Modulo.MODO_ENTREGA_LEGACY if modo_clases else Modulo.MODO_ENTREGA_PASOS
                ),
            )
            sec = SeccionModulo.objects.create(
                modulo=modulo,
                orden=1,
                titulo=f'Clase {numero}',
                activa=True,
            )
            PasoModulo.objects.create(
                modulo=modulo,
                seccion=sec,
                orden=1,
                titulo='Bienvenida',
                contenido='',
                activo=False,
                requiere_listo_para_avanzar=not modo_clases,
            )
        self.message_user(
            request,
            f'Clase {numero} creada. Suba el archivo en la pestaña Clase y active.',
            level=messages.SUCCESS,
        )
        return redirect('admin:core_modulo_change', modulo.pk)

    @admin.action(description='✅ Activar cursos seleccionados')
    def activar_cursos(self, request, queryset):
        count = queryset.update(activo=True)
        self.message_user(request, f"✅ {count} curso(s) activado(s)")

    @admin.action(description='❌ Desactivar cursos seleccionados')
    def desactivar_cursos(self, request, queryset):
        count = queryset.update(activo=False)
        self.message_user(request, f"❌ {count} curso(s) desactivado(s)")

    @admin.action(description='📋 Copiar a otro cliente…')
    def copiar_a_otro_cliente(self, request, queryset):
        from django.shortcuts import redirect

        ids = ','.join(str(c.pk) for c in queryset)
        return redirect(f'/admin/copiar-curso/?cursos={ids}')

    @admin.action(description='📋 Copiar a Analytics (Pruebas) — legacy')
    def copiar_a_analytics_pruebas(self, request, queryset):
        from core.copiar_cursos import (
            CLIENTE_ORIGEN_NOMBRE,
            ClienteOrigenNoEncontrado,
            copiar_cursos_a_pruebas,
            obtener_cliente_analytics_origen,
        )

        try:
            origen = obtener_cliente_analytics_origen()
        except ClienteOrigenNoEncontrado as e:
            self.message_user(request, str(e), level='error')
            return
        fuera = queryset.exclude(cliente=origen)
        if fuera.exists():
            self.message_user(
                request,
                f'Solo se copian cursos del cliente «{CLIENTE_ORIGEN_NOMBRE}» '
                f'(id={origen.pk}). {fuera.count()} curso(s) de otro cliente ignorados.',
                level='warning',
            )
        ids = list(queryset.filter(cliente=origen).values_list('pk', flat=True))
        if not ids:
            self.message_user(request, 'No hay cursos de Analytics seleccionados.', level='error')
            return
        result = copiar_cursos_a_pruebas(curso_ids=ids)
        self.message_user(
            request,
            f'✅ {result.total_copiados} curso(s) copiados a {result.destino.nombre}.',
        )


class PreguntaModuloInline(admin.StackedInline):
    """Preguntas de validación del módulo (Mini examen)"""
    model = PreguntaModulo
    extra = 0
    tab = True
    can_delete = True
    show_change_link = True
    verbose_name = 'Pregunta'
    verbose_name_plural = 'Examen'

    fieldsets = (
        ('Pregunta', {
            'fields': ('pregunta',),
            'description': 'Opcional. Solo si el módulo tiene mini examen al final.',
        }),
        ('Opciones de respuesta', {
            'fields': ('opcion_a', 'opcion_b', 'opcion_c', 'opcion_d', 'respuesta_correcta')
        }),
        ('Estado', {
            'fields': ('activa',)
        }),
    )


class SeccionModuloForm(forms.ModelForm):
    """
    Orden oculto por Unfold: filas nuevas llegan sin orden → «campo obligatorio»
    invisible + UniqueConstraint → no guarda el bloque. Igual patrón que Pasos.
    """

    class Meta:
        model = SeccionModulo
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if 'orden' in self.fields:
            self.fields['orden'].required = False
            self.fields['orden'].widget = forms.HiddenInput()
        if 'titulo' in self.fields:
            self.fields['titulo'].help_text = 'Nombre interno del bloque (no se envía por WhatsApp).'

    def _get_validation_exclusions(self):
        exclude = super()._get_validation_exclusions()
        exclude.add('orden')
        return exclude

    def clean_orden(self):
        try:
            idx = int(str(self.prefix or '0').rsplit('-', 1)[-1])
        except (TypeError, ValueError):
            idx = 0
        raw = None
        if self.data is not None:
            raw = self.data.get(self.add_prefix('orden'))
        if self.instance and self.instance.pk:
            if raw not in (None, '') and str(raw).strip().lstrip('-').isdigit():
                return int(str(raw).strip())
            if self.instance.orden:
                return self.instance.orden
            return idx + 1
        return 900000 + idx

    def validate_unique(self):
        exclude = self._get_validation_exclusions()
        try:
            self.instance.validate_unique(exclude=exclude)
        except ValidationError as e:
            self._update_errors(e)

    def has_changed(self):
        """Fila «Agregar Bloque» vacía no debe invalidar el guardado."""
        if getattr(self, 'empty_permitted', False) and not (self.instance and self.instance.pk):
            if self.data and (self.data.get(self.add_prefix('titulo')) or '').strip():
                return True
            if self.data and self.data.get(self.add_prefix('DELETE')):
                return True
            return False
        return super().has_changed()

    def clean(self):
        cleaned = super().clean()
        if not cleaned.get('orden'):
            cleaned['orden'] = self.clean_orden() if 'orden' in self.fields else 1
        return cleaned


class SeccionModuloInlineFormSet(BaseInlineFormSet):
    def add_fields(self, form, index):
        super().add_fields(form, index)
        if 'orden' in form.fields:
            form.fields['orden'].required = False
            form.fields['orden'].widget = forms.HiddenInput()
            if index is not None and not form.initial.get('orden') and not getattr(form.instance, 'orden', None):
                form.fields['orden'].initial = 900000 + index

    def validate_unique(self):
        return

    def clean(self):
        super().clean()
        seen = set()
        for i, form in enumerate(self.forms):
            if not hasattr(form, 'cleaned_data') or form.cleaned_data is None:
                continue
            if form.cleaned_data.get('DELETE'):
                continue
            o = form.cleaned_data.get('orden')
            if o is None or o in seen:
                o = 900000 + i
                form.cleaned_data['orden'] = o
            seen.add(o)
        summaries = []
        for i, form in enumerate(self.forms, start=1):
            if not form.errors:
                continue
            if getattr(form, 'empty_permitted', False) and not form.has_changed():
                continue
            titulo = ''
            if getattr(form, 'cleaned_data', None):
                titulo = (form.cleaned_data.get('titulo') or '').strip()
            if not titulo and form.data:
                titulo = (form.data.get(form.add_prefix('titulo')) or '').strip()
            label = f'«{titulo}»' if titulo else f'#{i}'
            bits = []
            for field, errs in form.errors.items():
                field_label = field
                if field in form.fields:
                    field_label = str(form.fields[field].label or field)
                for e in errs:
                    bits.append(f'{field_label}: {e}')
            if bits:
                summaries.append(f'Bloque {label}: ' + ' · '.join(bits))
        if summaries:
            raise ValidationError(summaries)


class SeccionModuloInline(admin.TabularInline):
    """Estructura: bloques que agrupan microcontenidos (título no se envía por WhatsApp)."""
    model = SeccionModulo
    form = SeccionModuloForm
    formset = SeccionModuloInlineFormSet
    extra = 0
    tab = True
    can_delete = True
    ordering = ('orden', 'id')
    ordering_field = 'orden'
    hide_ordering_field = True
    show_change_link = True
    verbose_name = 'Bloque'
    verbose_name_plural = 'Estructura'
    fields = ('mover_orden', 'orden', 'activa', 'titulo', 'resumen_pasos')
    readonly_fields = ('mover_orden', 'resumen_pasos')

    @admin.display(description='Orden')
    def mover_orden(self, obj):
        return _botones_mover_bloque(obj, 'seccion')

    @admin.display(description='Materiales')
    def resumen_pasos(self, obj):
        if not obj or not obj.pk:
            return format_html(
                '<span style="color:#94a3b8;font-size:12px;">Guarde para ver</span>'
            )
        n = obj.pasos.filter(activo=True).count()
        return format_html(
            '<span style="font-weight:600;color:#0f766e;">{}</span>'
            '<span style="color:#64748b;font-size:11px;"> activo(s)</span>',
            n,
        )


class PasoModuloForm(forms.ModelForm):
    media_file_upload = forms.FileField(
        label='Subir desde el PC',
        required=False,
        help_text=(
            'Elija PDF, imagen, audio o video. Al Guardar se sube a S3. '
            'Si el guardado falla, vuelva a elegir el archivo.'
        ),
        widget=UnfoldAdminFileFieldWidget(attrs={
            'class': 'eki-paso-file-input',
            'accept': 'video/*,image/*,application/pdf,audio/*',
        }),
    )

    class Meta:
        model = PasoModulo
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if 'orden' in self.fields:
            # Unfold hide_ordering_field + collapse dejan orden vacío en pasos NUEVOS
            # → "Este campo es obligatorio" y el archivo NUNCA se guarda en S3.
            self.fields['orden'].required = False
            self.fields['orden'].widget = forms.HiddenInput()
            if not self.is_bound and not (self.initial.get('orden') or getattr(self.instance, 'orden', None)):
                try:
                    idx = int(str(self.prefix or '0').rsplit('-', 1)[-1])
                    self.fields['orden'].initial = idx + 1
                except (TypeError, ValueError):
                    self.fields['orden'].initial = 1
        if 'seccion' in self.fields:
            self.fields['seccion'].help_text = (
                'Obligatorio en cada paso. Si la lista está vacía: Estructura → crear bloque → Guardar.'
            )
            self.fields['seccion'].error_messages = {
                **getattr(self.fields['seccion'], 'error_messages', {}),
                'required': 'Elija una sección (bloque). Cree una en Estructura si aún no hay.',
            }
        if 'contenido' in self.fields:
            self.fields['contenido'].help_text = (
                'Texto que ve el estudiante (opcional si solo sube archivo).'
            )
        if 'media_url' in self.fields:
            self.fields['media_url'].label = 'URL en S3 (se completa al Guardar)'
            self.fields['media_url'].required = False
            self.fields['media_url'].help_text = (
                'Tras Guardar con «Subir desde el PC» debe aparecer https://eki-produccion… '
                'No escriba el nombre del archivo aquí.'
            )
            self.fields['media_url'].error_messages = {
                **getattr(self.fields['media_url'], 'error_messages', {}),
                'invalid': (
                    'URL inválida. Deje vacío y use «Subir desde el PC», '
                    'o pegue https://… completo.'
                ),
            }
        inst = self.instance
        if inst and inst.pk and getattr(inst, 'tipo', None) == PasoModulo.TIPO_EVAL_OPC:
            data = inst.opciones_json
            if isinstance(data, dict):
                for letter in ('A', 'B', 'C', 'D'):
                    fname = f'eval_opcion_{letter.lower()}'
                    cur = (getattr(inst, fname, None) or '').strip()
                    if cur:
                        continue
                    val = data.get(letter)
                    if val:
                        self.initial[fname] = str(val)

    def _get_validation_exclusions(self):
        """
        Excluir orden: UniqueConstraint (modulo, orden) corre en full_clean →
        validate_constraints (Django 5.2) ANTES de save_formset renumerar.
        Sin esto: «dato duplicado para orden» al agregar micros o al drag.
        """
        exclude = super()._get_validation_exclusions()
        exclude.add('orden')
        return exclude

    def clean_orden(self):
        """
        Orden: Unfold/drag a menudo manda vacíos o duplicados en filas nuevas.
        - Existente: respeta POST (drag) o el valor en BD.
        - Nuevo: orden temporal único; ModuloAdmin.save_formset renumerar_orden_1_based.
        """
        try:
            idx = int(str(self.prefix or '0').rsplit('-', 1)[-1])
        except (TypeError, ValueError):
            idx = 0

        raw = None
        if self.data is not None:
            raw = self.data.get(self.add_prefix('orden'))

        if self.instance and self.instance.pk:
            if raw not in (None, '') and str(raw).strip().lstrip('-').isdigit():
                return int(str(raw).strip())
            if self.instance.orden:
                return self.instance.orden
            return idx + 1

        # Alta: no usar 1..n (choca UniqueConstraint con pasos ya guardados).
        return 900000 + idx

    def validate_unique(self):
        """orden ya va en exclusions; renumerar_orden_1_based al guardar."""
        exclude = self._get_validation_exclusions()
        try:
            self.instance.validate_unique(exclude=exclude)
        except ValidationError as e:
            self._update_errors(e)

    def has_changed(self):
        """Filas extra vacías no deben fallar solo por tipo=contenido por defecto."""
        if getattr(self, 'empty_permitted', False) and not (self.instance and self.instance.pk):
            if self._empty_extra_row_has_user_input():
                return True
            return False
        return super().has_changed()

    def _empty_extra_row_has_user_input(self) -> bool:
        if self.files:
            upload_name = self.add_prefix('media_file_upload')
            if self.files.get(upload_name):
                return True
        data = self.data
        if not data:
            return False
        if data.get(self.add_prefix('DELETE')):
            return True
        for name in ('seccion', 'titulo', 'contenido', 'media_url'):
            raw = data.get(self.add_prefix(name))
            if raw is None:
                continue
            if str(raw).strip():
                return True
        return False

    def clean_media_url(self):
        url = (self.cleaned_data.get('media_url') or '').strip()
        if not url:
            return ''
        if '://' not in url:
            raise ValidationError(
                'No use el nombre del archivo aquí. Deje vacío y use «Subir desde el PC», '
                'o pegue una URL https:// completa.'
            )
        return url

    def clean(self):
        cleaned = super().clean()
        self._pending_media_url = None
        self._pending_media_wa_apto = None
        if not cleaned.get('orden'):
            cleaned['orden'] = self.clean_orden() if 'orden' in self.fields else 1
        uploaded_file = cleaned.get('media_file_upload')
        if not uploaded_file:
            return cleaned
        # Si aún hay errores de otros campos, no subir a S3 (evita huérfanos).
        if self.errors:
            return cleaned
        modulo_id = getattr(self.instance, 'modulo_id', None) or 'sin_modulo'
        try:
            from core.admin._common import guardar_upload_admin_media_resultado

            resultado = guardar_upload_admin_media_resultado(
                uploaded_file,
                carpeta='modulos/pasos',
                prefix=f'modulo_{modulo_id}',
            )
            url = resultado.get('url')
            self._pending_media_wa_apto = resultado.get('media_wa_apto')
        except ValidationError as exc:
            msgs = getattr(exc, 'messages', None) or [str(exc)]
            self.add_error('media_file_upload', msgs[0] if len(msgs) == 1 else msgs)
            return cleaned
        if not (url or '').strip():
            self.add_error(
                'media_file_upload',
                'El archivo se subió pero no se obtuvo URL pública. '
                'Reintente o pegue la URL en «URL en S3».',
            )
            return cleaned
        self._pending_media_url = url
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        pending = getattr(self, '_pending_media_url', None)
        if pending:
            instance.media_url = pending
            apto = getattr(self, '_pending_media_wa_apto', None)
            if apto is not None:
                instance.media_wa_apto = bool(apto)
        if not instance.orden:
            instance.orden = 1
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class ModuloAdminForm(forms.ModelForm):
    """Form módulo + pestaña Clase simple (escribe el 1er PasoModulo)."""

    clase_texto = forms.CharField(
        label='Texto de la clase',
        required=False,
        widget=forms.Textarea(attrs={'rows': 4, 'style': 'min-width:280px;width:100%;'}),
        help_text='Lo que ve el estudiante (opcional si solo sube archivo/video).',
    )
    clase_archivo = forms.FileField(
        label='Subir material desde el PC',
        required=False,
        help_text=(
            'PDF, imagen, audio o video. Al Guardar se sube a S3. '
            'Si el guardado falla, vuelva a elegir el archivo.'
        ),
        widget=UnfoldAdminFileFieldWidget(attrs={
            'class': 'eki-paso-file-input eki-clase-file-input',
            'accept': 'video/*,image/*,application/pdf,audio/*',
        }),
    )
    clase_url = forms.CharField(
        label='O pegar URL https (opcional)',
        required=False,
        help_text='Solo si ya tiene un enlace público. No escriba video.mp4 a secas.',
        widget=forms.URLInput(attrs={'placeholder': 'https://…', 'style': 'width:100%;max-width:40rem;'}),
    )
    clase_activo = forms.BooleanField(
        label='Clase visible / activa',
        required=False,
        help_text='Márquelo cuando el material esté listo para estudiantes.',
    )
    clase_media_actual = forms.CharField(
        label='Material guardado',
        required=False,
        disabled=True,
        help_text='Se completa solo tras Guardar con archivo o URL.',
    )

    class Meta:
        model = Modulo
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        from core.module_steps import cuenta_microcontenidos_modulo

        # Form valida contenido con Clase/Estructura; no repetir en Modulo.clean.
        self.instance._eki_skip_contenido_model_clean = True

        self.fields['contenido'].required = False
        if 'modo_entrega' in self.fields:
            self.fields['modo_entrega'].help_text = (
                'Recomendado: «Por pasos con listo». Legacy envía todo el módulo de una vez '
                'e ignora microcontenidos. Automático hereda según haya pasos o no.'
            )
        n_micro = cuenta_microcontenidos_modulo(self.instance)
        if not self.instance.pk:
            self.fields['contenido'].help_text = (
                'Opcional. Prefiera la pestaña Clase (texto + archivo). '
                'Este campo es solo para modo Legacy.'
            )
        elif n_micro > 0:
            self.fields['contenido'].help_text = (
                f'Opcional: este módulo ya tiene {n_micro} microcontenido(s). '
                'Use pestaña Clase o Avanzado · Microcontenidos.'
            )
        else:
            self.fields['contenido'].help_text = (
                'Obligatorio solo si aún no hay microcontenidos. Use la pestaña Clase.'
            )

        paso = None
        if self.instance and self.instance.pk:
            paso = self.instance.pasos.order_by('orden', 'id').first()
        if paso:
            self.fields['clase_texto'].initial = paso.contenido or ''
            self.fields['clase_activo'].initial = bool(paso.activo)
            self.fields['clase_url'].initial = ''
            media = (paso.media_url or '').strip()
            self.fields['clase_media_actual'].initial = media or '(sin archivo aún)'
        else:
            self.fields['clase_activo'].initial = False
            self.fields['clase_media_actual'].initial = '(sin archivo aún)'

        curso = None
        if self.instance and getattr(self.instance, 'curso_id', None):
            curso = getattr(self.instance, 'curso', None)
        prefer_simple = True
        if curso is not None:
            prefer_simple = curso.es_modo_clases() or (
                self.instance.pk and self.instance.pasos.count() <= 1
            )
        elif not (self.instance and self.instance.pk):
            prefer_simple = True
        self.prefer_clase_simple = prefer_simple

    def clean_clase_url(self):
        url = (self.cleaned_data.get('clase_url') or '').strip()
        if not url:
            return ''
        if '://' not in url:
            raise ValidationError(
                'URL inválida. Deje vacío y use «Subir material», o pegue https://… completo.'
            )
        return url

    def clean(self):
        cleaned = super().clean()
        self._clase_pending_media_url = None
        uploaded = cleaned.get('clase_archivo')
        if not uploaded:
            return cleaned
        # No subir a S3 si ya hay errores de otros campos (evita huérfanos).
        if self.errors:
            return cleaned
        modulo_id = getattr(self.instance, 'pk', None) or 'nuevo'
        try:
            url = guardar_upload_admin_media(
                uploaded,
                carpeta='modulos/pasos',
                prefix=f'modulo_{modulo_id}',
            )
        except ValidationError as exc:
            msgs = getattr(exc, 'messages', None) or [str(exc)]
            self.add_error('clase_archivo', msgs[0] if len(msgs) == 1 else msgs)
            return cleaned
        if not (url or '').strip():
            self.add_error(
                'clase_archivo',
                'No se obtuvo URL pública. Reintente o pegue https:// en «O pegar URL».',
            )
            return cleaned
        self._clase_pending_media_url = url
        return cleaned

    def clean_contenido(self):
        from core.module_steps import validar_contenido_modulo

        contenido = self.cleaned_data.get('contenido', '')
        if not self.instance.pk:
            return contenido
        tiene_clase = bool(
            (self.cleaned_data.get('clase_texto') or '').strip()
            or self.cleaned_data.get('clase_archivo')
            or (self.cleaned_data.get('clase_url') or '').strip()
            or getattr(self, '_clase_pending_media_url', None)
        )
        validar_contenido_modulo(
            contenido,
            self.instance,
            tiene_clase_simple=tiene_clase,
            data=self.data,
        )
        return contenido


class PasoModuloInlineFormSet(BaseInlineFormSet):
    def add_fields(self, form, index):
        super().add_fields(form, index)
        if 'orden' in form.fields:
            form.fields['orden'].required = False
            form.fields['orden'].widget = forms.HiddenInput()
            if index is not None and not form.initial.get('orden') and not getattr(form.instance, 'orden', None):
                # Temporal; clean_orden / save_formset normalizan.
                form.fields['orden'].initial = 900000 + index

    def validate_unique(self):
        """Evita «dato duplicado para orden» entre filas; save_formset renumerar_orden_1_based."""
        return

    def clean(self):
        super().clean()
        from core.module_steps import validar_contenido_modulo

        # Asegura ordenes distintos en cleaned_data (por si Unfold mandó duplicados).
        seen = set()
        for i, form in enumerate(self.forms):
            if not hasattr(form, 'cleaned_data') or form.cleaned_data is None:
                continue
            if form.cleaned_data.get('DELETE'):
                continue
            o = form.cleaned_data.get('orden')
            if o is None or o in seen:
                o = 900000 + i
                form.cleaned_data['orden'] = o
            seen.add(o)

        # Errores por fila: sección vacía ya marca el campo; reforzamos mensaje de grupo.
        for form in self.forms:
            if not hasattr(form, 'cleaned_data') or form.cleaned_data is None:
                continue
            if form.cleaned_data.get('DELETE'):
                continue
            if form.errors.get('seccion'):
                continue
            if not form.cleaned_data.get('seccion') and (
                form.cleaned_data.get('titulo')
                or form.cleaned_data.get('contenido')
                or form.cleaned_data.get('media_file_upload')
                or form.cleaned_data.get('media_url')
            ):
                form.add_error(
                    'seccion',
                    'Elija una sección. Si no hay opciones: pestaña Estructura → crear bloque → Guardar.',
                )

        # Anti-intercalado (Module Builder WA): simular orden final de filas no borradas.
        from types import SimpleNamespace

        from core.module_structure import (
            detectar_secciones_intercaladas,
            mensaje_error_intercalado,
        )

        simulados = []
        for form in self.forms:
            if not hasattr(form, 'cleaned_data') or form.cleaned_data is None:
                continue
            if form.cleaned_data.get('DELETE'):
                continue
            sec = form.cleaned_data.get('seccion')
            if not sec:
                continue
            simulados.append(
                SimpleNamespace(
                    seccion_id=getattr(sec, 'pk', sec),
                    orden=form.cleaned_data.get('orden') or 0,
                    pk=getattr(form.instance, 'pk', None),
                )
            )
        simulados.sort(key=lambda x: (x.orden, x.pk or 0))
        hall = detectar_secciones_intercaladas(simulados)
        if hall:
            raise ValidationError(mensaje_error_intercalado(hall))

        summaries = []
        for i, form in enumerate(self.forms, start=1):
            if not form.errors:
                continue
            if getattr(form, 'empty_permitted', False) and not form.has_changed():
                continue
            titulo = ''
            if getattr(form, 'cleaned_data', None):
                titulo = (form.cleaned_data.get('titulo') or '').strip()
            if not titulo and form.data:
                titulo = (form.data.get(form.add_prefix('titulo')) or '').strip()
            label = f'«{titulo}»' if titulo else f'#{i}'
            bits = []
            for field, errs in form.errors.items():
                field_label = field
                if field in form.fields:
                    field_label = str(form.fields[field].label or field)
                for e in errs:
                    bits.append(f'{field_label}: {e}')
            if bits:
                summaries.append(f'Microcontenido {label}: ' + ' · '.join(bits))

        contenido = (self.data.get('contenido') or '').strip()
        tiene_clase = bool(
            (self.data.get('clase_texto') or '').strip()
            or (self.data.get('clase_url') or '').strip()
            or (self.files and self.files.get('clase_archivo'))
        )
        try:
            validar_contenido_modulo(
                contenido,
                self.instance,
                pasos_formset=self,
                tiene_clase_simple=tiene_clase,
                data=self.data,
            )
        except ValidationError as exc:
            summaries.extend(list(exc.messages))

        if summaries:
            raise ValidationError(summaries)


class PasoModuloInline(admin.StackedInline):
    """Microcontenidos WhatsApp dentro del módulo (orden + listo)."""
    model = PasoModulo
    form = PasoModuloForm
    formset = PasoModuloInlineFormSet
    extra = 0
    tab = True
    can_delete = True
    ordering = ('orden', 'id')
    ordering_field = 'orden'
    hide_ordering_field = True
    show_change_link = True
    verbose_name = 'Material'
    verbose_name_plural = 'Materiales'
    fieldsets = (
        ('Material', {
            'fields': (
                'mover_orden', 'orden', 'seccion', 'titulo', 'contenido',
                'media_file_upload', 'media_url', 'media_wa_apto', 'activo',
            ),
            'description': (
                '1) Elija bloque  2) Suba archivo  3) Active  4) Guarde. '
                'Video: se comprime a WA-safe; si no queda apto (&gt;16MB), el Guardar falla. '
                '«Apto WhatsApp» se completa solo en uploads nuevos.'
            ),
        }),
        ('WhatsApp (*listo* / tipo)', {
            'fields': ('tipo', 'requiere_listo_para_avanzar'),
            'classes': ('collapse',),
        }),
        ('Evaluación A–D', {
            'fields': (
                'eval_opcion_a', 'eval_opcion_b', 'eval_opcion_c', 'eval_opcion_d',
                'respuesta_correcta',
            ),
            'classes': ('collapse',),
        }),
        ('Retroalimentación', {
            'fields': ('feedback_correcto', 'feedback_incorrecto'),
            'classes': ('collapse',),
        }),
        ('JSON legado', {
            'fields': ('opciones_json',),
            'classes': ('collapse',),
        }),
    )
    readonly_fields = ('mover_orden', 'media_wa_apto')
    formfield_overrides = {
        models.TextField: {
            'widget': forms.Textarea(attrs={'rows': 3, 'cols': 50, 'style': 'min-width:280px;'}),
        },
    }

    @admin.display(description='Mover')
    def mover_orden(self, obj):
        return _botones_mover_bloque(obj, 'paso')

    def get_formset(self, request, obj=None, **kwargs):
        self._modulo_inline_parent = obj
        return super().get_formset(request, obj, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'seccion':
            mod = getattr(self, '_modulo_inline_parent', None)
            if mod is not None and mod.pk:
                kwargs['queryset'] = SeccionModulo.objects.filter(modulo_id=mod.pk).order_by(
                    'orden', 'id'
                )
            # Módulo nuevo sin PK: no filtramos; el admin no muestra este inline (ver ModuloAdmin).
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class ArchivoModuloInline(admin.StackedInline):
    """Multimedia a nivel módulo (Legacy / complemento). Preferir media en cada microcontenido."""
    model = ArchivoModulo
    extra = 0
    tab = True
    can_delete = True
    ordering = ('orden', 'id')
    ordering_field = 'orden'
    hide_ordering_field = True
    show_change_link = True
    verbose_name = 'Archivo'
    verbose_name_plural = 'Media (legacy)'

    fieldsets = (
        (None, {
            'fields': ('mover_orden', 'tipo', 'titulo', 'descripcion'),
            'description': 'Opcional. Preferir subir en Clase o en Materiales.',
        }),
        ('Archivo o URL', {
            'fields': ('archivo', 'preview_multimedia', 'url_externa'),
        }),
        ('Configuración', {
            'fields': ('disponible_offline', 'orden', 'activo'),
            'classes': ('collapse',),
        }),
    )
    readonly_fields = ('preview_multimedia', 'mover_orden')

    @admin.display(description='Mover')
    def mover_orden(self, obj):
        return _botones_mover_bloque(obj, 'archivo')

    def preview_multimedia(self, obj):
        """Vista previa del archivo multimedia"""
        if not obj.archivo:
            if obj.url_externa:
                return format_html(
                    '<div style="background:#e0f2fe;padding:12px;border-radius:6px;border-left:4px solid #0284c7;">'
                    '<strong>🔗 URL Externa:</strong><br>'
                    '<a href="{}" target="_blank" style="color:#0284c7;word-break:break-all;">{}</a>'
                    '</div>',
                    obj.url_externa, obj.url_externa
                )
            return format_html('<span style="color:#999;font-style:italic;">⚠️ Sin archivo subido</span>')

        url = obj.archivo.url

        if obj.tipo == 'imagen':
            return format_html(
                '<div style="text-align:center;background:#f9fafb;padding:16px;border-radius:8px;border:2px solid #e5e7eb;">'
                '<img src="{}" style="max-width:100%;max-height:400px;border-radius:6px;box-shadow:0 4px 6px rgba(0,0,0,0.1);" />'
                '<p style="margin-top:12px;color:#6b7280;font-size:12px;">📸 Imagen cargada correctamente</p>'
                '</div>',
                url
            )
        if obj.tipo == 'video':
            return format_html(
                '<div style="background:#f9fafb;padding:16px;border-radius:8px;border:2px solid #e5e7eb;">'
                '<video controls style="max-width:100%;border-radius:6px;box-shadow:0 4px 6px rgba(0,0,0,0.1);">'
                '<source src="{}" type="video/mp4">'
                'Tu navegador no soporta video HTML5.'
                '</video>'
                '<p style="margin-top:12px;color:#6b7280;font-size:12px;">🎥 Video cargado - URL: '
                '<code style="background:#e5e7eb;padding:2px 6px;border-radius:4px;font-size:11px;">{}</code></p>'
                '</div>',
                url, url
            )
        if obj.tipo == 'pdf':
            return format_html(
                '<div style="background:#fef2f2;padding:14px;border-radius:6px;border-left:4px solid #dc2626;">'
                '<a href="{}" target="_blank" style="background:#dc2626;color:white;padding:10px 20px;'
                'text-decoration:none;border-radius:6px;display:inline-block;font-weight:bold;">'
                '📄 Abrir PDF en Nueva Pestaña'
                '</a>'
                '<p style="margin-top:10px;color:#991b1b;font-size:12px;">Archivo: {}</p>'
                '</div>',
                url, obj.archivo.name
            )
        if obj.tipo == 'audio':
            return format_html(
                '<div style="background:#f9fafb;padding:16px;border-radius:8px;border:2px solid #e5e7eb;">'
                '<audio controls style="width:100%;">'
                '<source src="{}" type="audio/mpeg">'
                'Tu navegador no soporta audio HTML5.'
                '</audio>'
                '<p style="margin-top:12px;color:#6b7280;font-size:12px;">🎵 Audio cargado</p>'
                '</div>',
                url
            )
        return format_html(
            '<div style="background:#f0fdf4;padding:14px;border-radius:6px;border-left:4px solid #16a34a;">'
            '<a href="{}" target="_blank" style="color:#16a34a;font-weight:bold;">📎 Ver Archivo</a>'
            '<p style="margin-top:8px;color:#166534;font-size:12px;">{}</p>'
            '</div>',
            url, obj.archivo.name
        )
    preview_multimedia.short_description = "Vista Previa"


@admin.register(Modulo)
class ModuloAdmin(admin.ModelAdmin):
    """Administración de módulos"""
    form = ModuloAdminForm
    class Media:
        css = {
            'all': ('admin/css/modulo_whatsapp_bloques.css',),
        }
        js = ('admin/js/eki_modulo_jump.js',)

    # Listado ops: curso + número; preview largo → ficha del módulo.
    list_display = (
        'numero_titulo',
        'curso',
        'module_builder_link',
        'modo_entrega_badge',
        'pasos_activos_count',
        'examen_badge',
        'archivos_link',
        'ver_curso_link',
    )
    list_filter = ('curso', 'modo_entrega', 'examen_obligatorio')
    search_fields = ('titulo', 'descripcion', 'curso__nombre')
    list_select_related = ('curso', 'curso__cliente')
    list_per_page = 50
    ordering = ['curso__nombre', 'numero']
    readonly_fields = ('guia_microcontenidos_whatsapp',)
    inlines = [SeccionModuloInline, PasoModuloInline, ArchivoModuloInline, PreguntaModuloInline]
    actions = ['enviar_archivos_multimedia', 'ver_archivos_multimedia', 'renumerar_modulos']
    actions_detail = ['abrir_module_builder']

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                '<int:modulo_id>/mover/<str:kind>/<int:obj_id>/<str:direction>/',
                self.admin_site.admin_view(self.mover_bloque_view),
                name='core_modulo_mover_bloque',
            ),
            path(
                '<int:modulo_id>/mover/<str:kind>/<int:obj_id>/<str:direction>',
                self.admin_site.admin_view(self.mover_bloque_view),
            ),
        ]
        return custom + urls

    @action(
        description='Module Builder',
        url_path='abrir-module-builder',
        icon='view_timeline',
    )
    def abrir_module_builder(self, request, object_id):
        """Botón arriba del change form → canvas Module Builder."""
        from core.module_builder import module_builder_habilitado_para_curso

        modulo = Modulo.objects.select_related('curso').filter(pk=object_id).first()
        curso = modulo.curso if modulo else None
        if not module_builder_habilitado_para_curso(curso, request):
            messages.error(
                request,
                'Module Builder desactivado para este curso. Active EKI_MODULE_BUILDER_BETA=1, '
                'use ?builder=1 como superusuario, o añada el curso a EKI_MODULE_BUILDER_CURSOS.',
            )
            return redirect(f'/admin/core/modulo/{object_id}/change/')
        return redirect('admin_module_builder', modulo_id=int(object_id))

    def has_abrir_module_builder_permission(self, request, object_id=None):
        from core.module_builder import (
            module_builder_habilitado,
            module_builder_habilitado_para_curso,
        )

        if not request.user.is_staff:
            return False
        if object_id:
            modulo = Modulo.objects.select_related('curso').filter(pk=object_id).first()
            curso = modulo.curso if modulo else None
            return module_builder_habilitado_para_curso(curso, request)
        return module_builder_habilitado(request)

    def module_builder_link(self, obj):
        from core.module_builder import module_builder_habilitado_para_curso

        curso = getattr(obj, 'curso', None) if obj else None
        if not obj or not obj.pk or not module_builder_habilitado_para_curso(curso):
            return '—'
        return format_html(
            '<a href="/admin/module-builder/{}/" style="font-weight:700;color:#7A4E8E;">Builder</a>',
            obj.pk,
        )
    module_builder_link.short_description = 'Builder'

    def mover_bloque_view(self, request, modulo_id, kind, obj_id, direction):
        """↑↓ de secciones / pasos / multimedia sin salir del change del módulo."""
        from core.models import ArchivoModulo

        if not self.has_change_permission(request):
            raise PermissionDenied

        kind_map = {
            'seccion': SeccionModulo,
            'paso': PasoModulo,
            'archivo': ArchivoModulo,
        }
        model = kind_map.get(kind)
        if model is None or direction not in ('up', 'down'):
            messages.error(request, 'Movimiento no válido.')
            return redirect('admin:core_modulo_change', modulo_id)

        try:
            modulo = Modulo.objects.get(pk=modulo_id)
            obj = model.objects.get(pk=obj_id, modulo_id=modulo.pk)
        except (Modulo.DoesNotExist, model.DoesNotExist):
            messages.error(request, 'Bloque no encontrado.')
            return redirect('admin:core_modulo_changelist')

        if intercambiar_orden(obj, direction):
            messages.success(request, 'Orden actualizado.')
        else:
            messages.info(request, 'Ya está al extremo; no hay cambio.')
        return redirect('admin:core_modulo_change', modulo_id)

    def get_inline_instances(self, request, obj=None):
        """Módulo nuevo: Microcontenidos aparecen tras el 1.er guardado (necesitan PK + sección)."""
        instances = []
        for inline_class in self.inlines:
            if inline_class is PasoModuloInline and obj is None:
                continue
            instances.append(inline_class(self.model, self.admin_site))
        return instances

    def get_changeform_initial_data(self, request):
        data = super().get_changeform_initial_data(request)
        data.setdefault('modo_entrega', Modulo.MODO_ENTREGA_PASOS)
        return data

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        if not change:
            created = sembrar_plantilla_modulo(form.instance)
            if created['seccion'] or created['pasos']:
                bits = []
                if created['seccion']:
                    bits.append('1 bloque')
                if created['pasos']:
                    bits.append(f'{created["pasos"]} micro(s)')
                self.message_user(
                    request,
                    'Plantilla: ' + ' + '.join(bits) + '. Complete la pestaña Clase (archivo + activar).',
                    level=messages.INFO,
                )
        paso = aplicar_clase_simple_desde_form(form, formsets=formsets)
        if paso and (
            getattr(form, '_clase_pending_media_url', None)
            or (form.cleaned_data.get('clase_url') or '').strip()
        ):
            self.message_user(
                request,
                f'Material de clase listo. URL: {paso.media_url[:120]}'
                + ('…' if len(paso.media_url or '') > 120 else ''),
                level=messages.SUCCESS,
            )

    def response_add(self, request, obj, post_url_continue=None):
        """Tras crear: Module Builder si está habilitado; si no, ficha Clase."""
        if '_addanother' in request.POST:
            return super().response_add(request, obj, post_url_continue)
        from core.module_builder import module_builder_habilitado_para_curso

        if module_builder_habilitado_para_curso(getattr(obj, 'curso', None), request):
            self.message_user(
                request,
                'Módulo creado. Arme secciones y micros en el Module Builder.',
                level=messages.SUCCESS,
            )
            return redirect('admin_module_builder', modulo_id=obj.pk)
        self.message_user(
            request,
            'Módulo creado. En pestaña Clase: texto + subir archivo → Activar → Guardar.',
            level=messages.SUCCESS,
        )
        return redirect('admin:core_modulo_change', obj.pk)

    def response_change(self, request, obj):
        """Tras guardar: ir al Module Builder (camino fácil A+B)."""
        if '_addanother' in request.POST:
            return super().response_change(request, obj)
        if '_continue' in request.POST:
            return super().response_change(request, obj)
        from core.module_builder import module_builder_habilitado_para_curso

        if module_builder_habilitado_para_curso(getattr(obj, 'curso', None), request):
            self.message_user(
                request,
                'Guardado. Puede seguir en el Module Builder.',
                level=messages.SUCCESS,
            )
            return redirect('admin_module_builder', modulo_id=obj.pk)
        return super().response_change(request, obj)

    def ver_curso_link(self, obj):
        """Link directo al curso padre"""
        url = reverse('admin:core_curso_change', args=[obj.curso.id])
        return format_html('<a href="{}" style="color:#2196F3;">Ver curso</a>', url)
    ver_curso_link.short_description = "Curso"

    @admin.display(description='')
    def guia_microcontenidos_whatsapp(self, obj):
        """Guía Clase simple + prefs para JS (tab por defecto) + link Module Builder."""
        from core.module_builder import module_builder_habilitado_para_curso

        prefer = '1'
        modo_clases = '0'
        builder_html = ''
        if obj and obj.pk and obj.curso_id:
            modo_clases = '1' if obj.curso.es_modo_clases() else '0'
            prefer = '1' if (obj.curso.es_modo_clases() or obj.pasos.count() <= 1) else '0'
            if module_builder_habilitado_para_curso(obj.curso, None):
                builder_html = (
                    f'<p class="eki-modulo-guia__line">'
                    f'<a href="/admin/module-builder/{obj.pk}/" style="font-weight:700;color:#7A4E8E;">'
                    f'→ Abrir Module Builder (secciones + micros)</a>'
                    f' · camino recomendado. Edita <b>Materiales / pasos WA</b> (no borra drip ni '
                    f'«Disponible desde»). El admin clásico sigue abajo si lo necesita.</p>'
                )
        return format_html(
            '<div class="eki-modulo-guia" id="eki-modulo-prefs" '
            'data-default-tab="clase" data-prefer-simple="{}" data-modo-clases="{}">'
            '{}'
            '<p class="eki-modulo-guia__line">'
            '<b>Subir:</b> elija archivo → Active → <b>Guardar</b>. '
            'Varios materiales: pestañas Estructura + Materiales (el drag se mantiene).'
            '</p>'
            '</div>',
            prefer,
            modo_clases,
            mark_safe(builder_html),
        )

    def save_formset(self, request, form, formset, change):
        """Drag Unfold escribe orden 0..n; UniqueConstraint (modulo, orden) exige temp + renúmero 1..n."""
        models_con_unique = (SeccionModulo, PasoModulo)
        if formset.model in models_con_unique and form.instance.pk:
            with transaction.atomic():
                preparar_ordenes_temporales(formset.model, form.instance.pk)
                super().save_formset(request, form, formset, change)
                renumerar_orden_1_based(formset.model, form.instance.pk)
            return
        super().save_formset(request, form, formset, change)

    fieldsets = (
        (
            'Clase',
            {
                'classes': ['tab'],
                'fields': (
                    'guia_microcontenidos_whatsapp',
                    'curso',
                    'numero',
                    'titulo',
                    'clase_texto',
                    'clase_archivo',
                    'clase_url',
                    'clase_media_actual',
                    'clase_activo',
                ),
                'description': (
                    'Flujo recomendado para Aprende / clases. '
                    'No pegue el nombre del archivo; use Subir material.'
                ),
            },
        ),
        (
            'Más opciones',
            {
                'classes': ['tab'],
                'fields': (
                    'descripcion',
                    'modo_entrega',
                    'secciones_por_listo',
                    'facilitador_checkpoint',
                    'contenido',
                    'examen_obligatorio',
                    'puntaje_minimo_aprobacion',
                    'duracion_dias',
                    'habilitado_desde',
                ),
                'description': (
                    'Entrega WhatsApp, texto legacy (solo si no usa Clase/Materiales), '
                    'examen y calendario.'
                ),
            },
        ),
    )
    
    def modo_entrega_badge(self, obj):
        from core.models import Modulo
        labels = dict(Modulo.MODOS_ENTREGA)
        t = labels.get(obj.modo_entrega, obj.modo_entrega)
        return format_html('<span style="font-size:11px;">{}</span>', t)
    modo_entrega_badge.short_description = 'Entrega'

    def pasos_activos_count(self, obj):
        count = obj.pasos.filter(activo=True, seccion__activa=True).count()
        if obj.modo_entrega == 'pasos' and count == 0:
            return format_html(
                '<span style="color:#c62828;font-weight:bold;">'
                '⚠️ Modo PASOS pero sin pasos activos (count: {})'
                '</span>',
                count,
            )
        return count
    pasos_activos_count.short_description = 'Pasos activos'

    def numero_titulo(self, obj):
        return f"Módulo {obj.numero}: {obj.titulo}"
    numero_titulo.short_description = "Módulo"
    
    def examen_badge(self, obj):
        if obj.examen_obligatorio:
            return format_html(
                '<span style="background:#ffebee;color:#c62828;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;">🔒 OBLIGATORIO ({0}%)</span>',
                obj.puntaje_minimo_aprobacion
            )
        return format_html('<span style="color:#999;">Sin examen</span>')
    examen_badge.short_description = "Examen"
    
    def archivos_link(self, obj):
        count = obj.archivos_multimedia.filter(activo=True).count()
        if count > 0:
            url = f"/admin/core/archivomodulo/?modulo__id__exact={obj.id}"
            return format_html(
                '<a href="{}" style="background:#e3f2fd;color:#1976d2;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;text-decoration:none;">📁 {} archivo(s)</a>',
                url, count
            )
        return format_html('<span style="color:#999;">Sin archivos</span>')
    archivos_link.short_description = "Multimedia"
    
    def tiene_pregunta(self, obj):
        count = obj.preguntas.filter(activa=True).count()
        if count > 0:
            return format_html('<span style="color:green;">✅ {} pregunta(s)</span>', count)
        return format_html('<span style="color:red;">❌ Sin pregunta</span>')
    tiene_pregunta.short_description = "Mini Examen"
    
    def contenido_preview(self, obj):
        preview = obj.contenido[:60] + "..." if len(obj.contenido) > 60 else obj.contenido
        return format_html('<span style="color:#666;font-style:italic;">{}</span>', preview)
    contenido_preview.short_description = "Vista Previa"
    
    def enviar_archivos_multimedia(self, request, queryset):
        """Envía los archivos multimedia de los módulos seleccionados a estudiantes inscritos"""
        from core.utils import enviar_whatsapp_twilio
        import json
        
        enviados = 0
        errores = 0
        
        for modulo in queryset:
            archivos = modulo.archivos_multimedia.filter(activo=True)
            if not archivos.exists():
                continue
            
            # Obtener estudiantes inscritos en el curso
            estudiantes = Estudiante.objects.filter(
                progreso__curso=modulo.curso,
                activo=True
            ).distinct()
            
            for estudiante in estudiantes:
                try:
                    # Mensaje con lista de archivos
                    mensaje = f"📚 *{modulo.titulo}*\n\n"
                    mensaje += f"Tienes {archivos.count()} archivo(s) multimedia disponible(s):\n\n"
                    
                    for i, archivo in enumerate(archivos, 1):
                        icono = {
                            'video': '🎥',
                            'imagen': '🖼️',
                            'infografia': '📊',
                            'pdf': '📄',
                            'audio': '🎵'
                        }.get(archivo.tipo, '📁')
                        
                        mensaje += f"{icono} *{i}. {archivo.titulo}*\n"
                        if archivo.descripcion:
                            mensaje += f"   {archivo.descripcion}\n"
                        
                        if archivo.disponible_offline:
                            url_descarga = f"{request.build_absolute_uri('/media/descargar-archivo/')}{archivo.id}/"
                            mensaje += f"   🔗 Descarga: {url_descarga}\n"
                        
                        if archivo.url_externa:
                            mensaje += f"   🌐 Ver online: {archivo.url_externa}\n"
                        
                        mensaje += "\n"
                    
                    # Enviar por WhatsApp
                    enviar_whatsapp_twilio(estudiante.telefono, mensaje)
                    enviados += 1
                    
                except Exception as e:
                    print(f"Error enviando archivos a {estudiante.telefono}: {e}")
                    errores += 1
        
        self.message_user(
            request,
            f'✅ {enviados} mensaje(s) enviado(s) con archivos multimedia. ❌ {errores} error(es).'
        )
    enviar_archivos_multimedia.short_description = "📤 Enviar archivos multimedia a estudiantes"
    
    @admin.action(description='� Renumerar módulos (1, 2, 3...)')
    def renumerar_modulos(self, request, queryset):
        """Renumera los módulos seleccionados empezando desde 1"""
        # Agrupar por curso
        modulos_por_curso = {}
        for modulo in queryset.order_by('curso', 'numero', 'id'):
            if modulo.curso not in modulos_por_curso:
                modulos_por_curso[modulo.curso] = []
            modulos_por_curso[modulo.curso].append(modulo)
        
        total_renumerados = 0
        for curso, modulos in modulos_por_curso.items():
            for idx, modulo in enumerate(modulos, start=1):
                if modulo.numero != idx:
                    modulo.numero = idx
                    modulo.save()
                    total_renumerados += 1
        
        self.message_user(
            request,
            f"✅ {total_renumerados} módulos renumerados correctamente",
            level='success'
        )
    renumerar_modulos.short_description = "🔢 Renumerar módulos (1, 2, 3...)"
    
    @admin.action(description='�📁 Ver archivos multimedia de módulos')
    def ver_archivos_multimedia(self, request, queryset):
        """Redirige a ver los archivos multimedia de los módulos seleccionados"""
        from django.shortcuts import redirect
        modulo_ids = ','.join(str(m.id) for m in queryset)
        return redirect(f'/admin/core/archivomodulo/?modulo__id__in={modulo_ids}')


class PreguntaExamenInline(admin.TabularInline):
    """Preguntas dentro del examen"""
    model = PreguntaExamen
    extra = 1
    fields = ('numero', 'pregunta', 'respuesta_correcta', 'puntos')
    ordering = ['numero']


class ExamenAdmin(admin.ModelAdmin):
    """Administración de exámenes"""
    list_display = ('curso_nombre', 'total_preguntas_display', 'puntaje_minimo')
    list_filter = ('curso',)
    search_fields = ('curso__nombre', 'instrucciones')
    inlines = [PreguntaExamenInline]
    
    fieldsets = (
        ('📝 Configuración del Examen', {
            'fields': ('curso', 'instrucciones', 'puntaje_minimo')
        }),
    )
    
    def curso_nombre(self, obj):
        return f"{obj.curso.emoji} {obj.curso.nombre}"
    curso_nombre.short_description = "Curso"
    
    def total_preguntas_display(self, obj):
        count = obj.preguntas.count()
        return format_html(
            '<span style="background:#fff3cd;padding:4px 8px;border-radius:4px;">{} preguntas</span>',
            count
        )
    total_preguntas_display.short_description = "Preguntas"


class PreguntaExamenAdmin(admin.ModelAdmin):
    """Administración de preguntas de examen"""
    list_display = ('numero_pregunta', 'examen', 'puntos', 'pregunta_preview')
    list_filter = ('examen__curso',)
    search_fields = ('pregunta', 'respuesta_correcta')
    ordering = ['examen', 'numero']
    
    fieldsets = (
        ('❓ Pregunta', {
            'fields': ('examen', 'numero', 'pregunta')
        }),
        ('✅ Respuesta', {
            'fields': ('respuesta_correcta', 'puntos'),
            'description': 'Palabras clave separadas por comas (la IA evaluará si están presentes)'
        }),
    )
    
    def numero_pregunta(self, obj):
        return f"Pregunta {obj.numero}"
    numero_pregunta.short_description = "N°"
    
    def pregunta_preview(self, obj):
        preview = obj.pregunta[:80] + "..." if len(obj.pregunta) > 80 else obj.pregunta
        return preview
    pregunta_preview.short_description = "Pregunta"


@admin.register(ProgresoEstudiante)
class ProgresoEstudianteAdmin(admin.ModelAdmin):
    """Seguimiento del progreso de estudiantes"""
    list_display = ('estudiante', 'curso', 'barra_progreso', 'modulo_actual', 'completado_badge', 'certificado_status', 'fecha_ultimo_avance', 'fecha_inicio')
    list_filter = ('completado', 'curso', 'fecha_inicio')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'curso__nombre')
    readonly_fields = ('fecha_inicio', 'porcentaje_avance', 'info_certificado')
    list_per_page = 50
    ordering = ('-fecha_inicio',)
    actions = ['exportar_progreso_excel', 'exportar_progreso_csv', 'generar_certificados_pendientes']  # ✅ Nuevas acciones
    
    fieldsets = (
        ('👤 Estudiante y Curso', {
            'fields': ('estudiante', 'curso')
        }),
        ('📊 Progreso', {
            'fields': ('modulo_actual', 'completado', 'porcentaje_avance')
        }),
        ('📅 Fechas', {
            'fields': ('fecha_inicio', 'fecha_ultimo_avance', 'fecha_completado')
        }),
    )
    
    def barra_progreso(self, obj):
        """Muestra una barra de progreso visual"""
        porcentaje = obj.porcentaje_avance()
        
        # Colores según progreso
        if porcentaje >= 80:
            color = '#4caf50'
        elif porcentaje >= 50:
            color = '#ff9800'
        else:
            color = '#f44336'
        
        return format_html(
            '<div style="width:100px;height:20px;background:#f0f0f0;border-radius:10px;overflow:hidden;border:1px solid #ddd;">'
            '<div style="width:{}%;height:100%;background:{};transition:width 0.3s;display:flex;align-items:center;justify-content:center;color:white;font-size:11px;font-weight:bold;">'
            '{}%'
            '</div></div>',
            porcentaje, color, porcentaje
        )
    barra_progreso.short_description = "Progreso"
    
    def certificado_status(self, obj):
        """Muestra si hay certificado generado"""
        if obj.completado:
            certificado = Certificado.objects.filter(estudiante=obj.estudiante, curso=obj.curso, emitido=True).first()
            if certificado:
                return format_html(
                    '<span style="color:#4caf50;font-weight:bold;">🏆 Emitido</span>'
                )
            else:
                return format_html(
                    '<span style="color:#ff9800;font-weight:bold;">⏳ Pendiente</span>'
                )
        return format_html('<span style="color:#999;">-</span>')
    certificado_status.short_description = "Certificado"
    
    def info_certificado(self, obj):
        """Información del certificado en los detalles"""
        if obj.completado:
            certificado = Certificado.objects.filter(estudiante=obj.estudiante, curso=obj.curso).first()
            if certificado:
                return format_html(
                    '<div style="background:#f5f5f5;padding:10px;border-radius:4px;border-left:4px solid #4caf50;">'
                    '<strong>✅ Certificado Generado</strong><br>'
                    'Código: <code>{}</code><br>'
                    'Calificación: <strong>{}</strong>%<br>'
                    'Emitido: {}'
                    '</div>',
                    certificado.codigo_verificacion,
                    int(certificado.calificacion_final),
                    certificado.fecha_emision.strftime('%d/%m/%Y') if certificado.fecha_emision else 'N/A'
                )
            else:
                return format_html(
                    '<div style="background:#fff3cd;padding:10px;border-radius:4px;border-left:4px solid #ff9800;">'
                    '<strong>⏳ Certificado Pendiente</strong><br>'
                    'El curso está completo pero el certificado aún no se ha generado'
                    '</div>'
                )
        return format_html(
            '<div style="background:#f5f5f5;padding:10px;border-radius:4px;border-left:4px solid #999;">'
            '<span style="color:#999;">El curso aún no está completo</span>'
            '</div>'
        )
    info_certificado.short_description = "Estado del Certificado"
    
    def porcentaje_badge(self, obj):
        porcentaje = obj.porcentaje_avance()
        if porcentaje >= 80:
            color = '#4caf50'
        elif porcentaje >= 50:
            color = '#ff9800'
        else:
            color = '#f44336'
        return format_html(
            '<span style="background:{};color:white;padding:4px 12px;border-radius:12px;font-weight:bold;"{}%</span>',
            color, porcentaje
        )
    porcentaje_badge.short_description = "Avance"
    
    def completado_badge(self, obj):
        if obj.completado:
            return format_html('<span style="color:green;">✅ Completo</span>')
        return format_html('<span style="color:orange;">⏳ En progreso</span>')
    completado_badge.short_description = "Estado"
    
    @admin.action(description='🏆 Generar certificados pendientes')
    def generar_certificados_pendientes(self, request, queryset):
        """Genera certificados para cursos completados que no los tienen"""
        from core.certificado_service import generar_y_guardar_certificado, crear_certificado_automatico, enviar_certificado_whatsapp
        
        generados = 0
        errores = 0
        
        for progreso in queryset.filter(completado=True):
            certificado = Certificado.objects.filter(estudiante=progreso.estudiante, curso=progreso.curso).first()
            if not certificado:
                try:
                    certificado = crear_certificado_automatico(progreso.estudiante, progreso.curso)
                    if certificado and not certificado.emitido:
                        generar_y_guardar_certificado(certificado)
                    if certificado and certificado.emitido:
                        enviar_certificado_whatsapp(certificado)
                    generados += 1
                except Exception as e:
                    logger.error(f"Error al generar certificado para {progreso.estudiante.nombre}: {str(e)}")
                    errores += 1
        
        if generados > 0:
            self.message_user(request, f'✅ {generados} certificado(s) generado(s) y enviado(s)', messages.SUCCESS)
        if errores > 0:
            self.message_user(request, f'❌ {errores} error(es) al generar', messages.ERROR)
    
    @admin.action(description='📊 Exportar progreso a Excel')
    def exportar_progreso_excel(self, request, queryset):
        """Exporta el progreso de estudiantes a Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Progreso Estudiantes"
        
        # Encabezados
        headers = ['Estudiante', 'Teléfono', 'Curso', 'Módulo Actual', 'Avance %', 'Completado', 'Fecha Inicio', 'Fecha Completado']
        ws.append(headers)
        
        # Estilo
        header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for progreso in queryset:
            ws.append([
                progreso.estudiante.nombre,
                f"+{progreso.estudiante.telefono}",
                progreso.curso.nombre,
                progreso.modulo_actual or 'No iniciado',
                progreso.porcentaje_avance(),
                "Sí" if progreso.completado else "No",
                progreso.fecha_inicio.strftime('%Y-%m-%d %H:%M'),
                progreso.fecha_completado.strftime('%Y-%m-%d %H:%M') if progreso.fecha_completado else 'N/A'
            ])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'progreso_estudiantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @admin.action(description='📄 Exportar progreso a CSV')
    def exportar_progreso_csv(self, request, queryset):
        """Exporta el progreso de estudiantes a CSV"""
        import csv
        from datetime import datetime
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'progreso_estudiantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Estudiante', 'Teléfono', 'Curso', 'Módulo Actual', 'Avance %', 'Completado', 'Fecha Inicio', 'Fecha Completado'])
        
        for progreso in queryset:
            writer.writerow([
                progreso.estudiante.nombre,
                f"+{progreso.estudiante.telefono}",
                progreso.curso.nombre,
                progreso.modulo_actual or 'No iniciado',
                progreso.porcentaje_avance(),
                "Sí" if progreso.completado else "No",
                progreso.fecha_inicio.strftime('%Y-%m-%d %H:%M'),
                progreso.fecha_completado.strftime('%Y-%m-%d %H:%M') if progreso.fecha_completado else 'N/A'
            ])
        
        return response


@admin.register(ModuloCompletado)
class ModuloCompletadoAdmin(admin.ModelAdmin):
    """Registro de módulos completados"""
    list_display = ('estudiante_nombre', 'modulo_info', 'fecha_completado')
    list_filter = ('fecha_completado', 'modulo__curso')
    search_fields = ('progreso__estudiante__nombre', 'modulo__titulo')
    readonly_fields = ('fecha_completado',)
    ordering = ('-fecha_completado',)
    
    def estudiante_nombre(self, obj):
        return obj.progreso.estudiante.nombre
    estudiante_nombre.short_description = "Estudiante"
    
    def modulo_info(self, obj):
        return f"{obj.modulo.curso.emoji} {obj.modulo.titulo}"
    modulo_info.short_description = "Módulo"


class ResultadoExamenAdmin(admin.ModelAdmin):
    """Resultados de exámenes"""
    list_display = ('estudiante', 'examen_info', 'puntaje_badge', 'aprobado_badge', 'fecha_realizado')
    list_filter = ('aprobado', 'examen__curso', 'fecha_realizado')
    search_fields = ('estudiante__nombre', 'estudiante__telefono', 'examen__curso__nombre')
    readonly_fields = ('fecha_realizado', 'respuestas', 'feedback')
    ordering = ('-fecha_realizado',)
    
    fieldsets = (
        ('👤 Estudiante y Examen', {
            'fields': ('estudiante', 'examen')
        }),
        ('📊 Resultado', {
            'fields': ('puntaje', 'aprobado')
        }),
        ('📝 Respuestas y Retroalimentación', {
            'fields': ('respuestas', 'feedback'),
            'classes': ('collapse',)
        }),
        ('📅 Fecha', {
            'fields': ('fecha_realizado',)
        }),
    )
    
    def examen_info(self, obj):
        return f"{obj.examen.curso.emoji} {obj.examen.curso.nombre}"
    examen_info.short_description = "Examen"
    
    def puntaje_badge(self, obj):
        if obj.puntaje >= 80:
            color = '#4caf50'
        elif obj.puntaje >= 70:
            color = '#ff9800'
        else:
            color = '#f44336'
        return format_html(
            '<span style="background:{};color:white;padding:6px 12px;border-radius:12px;font-weight:bold;font-size:14px;">{}/100</span>',
            color, obj.puntaje
        )
    puntaje_badge.short_description = "Puntaje"
    
    def aprobado_badge(self, obj):
        if obj.aprobado:
            return format_html('<span style="background:#4caf50;color:white;padding:4px 12px;border-radius:12px;">✅ APROBADO</span>')
        return format_html('<span style="background:#f44336;color:white;padding:4px 12px;border-radius:12px;">❌ REPROBADO</span>')
    aprobado_badge.short_description = "Estado"


# Personalizar el admin site
admin.site.site_header = "eki - Chatbot Agro 🌱"
admin.site.site_title = "Administración eki"
admin.site.index_title = "Panel de Control - Chatbot Educativo"


