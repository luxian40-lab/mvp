from core.admin._common import *  # noqa: F401,F403
from core.admin.clientes import HabilitacionModuloEstudianteInline
from django import forms
from django.db.models import Q


class CursoInscribirChoiceField(forms.ModelMultipleChoiceField):
    def label_from_instance(self, obj):
        org = obj.cliente.nombre if getattr(obj, 'cliente_id', None) else 'General'
        modo = 'Clases · Aprende' if obj.es_modo_clases() else 'Módulos · WhatsApp'
        return f'{obj.nombre}  —  {org} ({modo})'


class EstudianteAdminForm(forms.ModelForm):
    """Alta/edición: permite elegir curso(s) a inscribir sin pasar por Excel."""

    cursos_a_inscribir = CursoInscribirChoiceField(
        queryset=Curso.objects.none(),
        required=False,
        label='Inscribir en curso(s)',
        help_text=(
            'Opcional. Al guardar se crea el progreso. '
            'Si eligió Cliente, se listan sobre todo los cursos de esa org.'
        ),
        widget=forms.CheckboxSelectMultiple,
    )

    class Meta:
        model = Estudiante
        fields = (
            'tipo_documento', 'cedula', 'nombre', 'telefono', 'cliente', 'activo',
            'municipio', 'departamento', 'ubicacion_detalle', 'genero', 'edad', 'rango_edad',
        )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        qs = (
            Curso.objects.filter(activo=True)
            .select_related('cliente')
            .order_by('cliente__nombre', 'nombre')
        )
        cliente_id = None
        if self.data.get('cliente'):
            try:
                cliente_id = int(self.data.get('cliente'))
            except (TypeError, ValueError):
                cliente_id = None
        elif self.instance and self.instance.pk and self.instance.cliente_id:
            cliente_id = self.instance.cliente_id
        if cliente_id:
            qs = qs.filter(Q(cliente_id=cliente_id) | Q(cliente__isnull=True))
        if self.instance and self.instance.pk:
            ya = ProgresoEstudiante.objects.filter(
                estudiante=self.instance,
            ).values_list('curso_id', flat=True)
            qs = qs.exclude(pk__in=ya)
            self.fields['cursos_a_inscribir'].help_text = (
                'Marque cursos nuevos. Los ya inscritos aparecen en la lista de abajo.'
            )
        self.fields['cursos_a_inscribir'].queryset = qs


@admin.register(Estudiante)
class EstudianteAdmin(admin.ModelAdmin):
    """Gestión de estudiantes/campesinos"""
    form = EstudianteAdminForm
    change_form_template = 'admin/core/estudiante/change_form.html'
    inlines = [HabilitacionModuloEstudianteInline]
    # Listado limpio P1: 5 columnas operativas (cursos → ficha / filtros).
    list_display = (
        'nombre',
        'cedula_formateada',
        'telefono_formateado',
        'cliente_nombre',
        'conversacion_link',
        'activo',
    )
    list_filter = (
        'cliente',
        'activo',
        EstudianteSinProgresoFilter,
        CursosEstudianteFilter,
        GruposEstudianteFilter,
        'departamento',
    )
    search_fields = ('nombre', 'cedula', 'telefono', 'cliente__nombre')
    list_select_related = ('cliente',)
    list_per_page = 50
    ordering = ('cliente__nombre', 'nombre')
    # Primero ops de datos; WhatsApp al final del menú Acciones.
    actions = [
        'asignar_a_grupo_accion',
        'asignar_cliente_masivo',
        'inscribir_curso_masivo',
        'exportar_estudiantes_por_curso',
        'exportar_seleccion_ligera',
        'enviar_mensaje_masivo',
        'enviar_anuncio_grupal',
        'invitar_a_grupo_whatsapp',
        'eliminar_estudiantes_seguro',
    ]
    
    # ✨ AGREGAR BOTÓN DE IMPORTAR EN LA PARTE SUPERIOR
    change_list_template = 'admin/estudiante_changelist.html'
    
    fieldsets = (
        ('Identificación y contacto', {
            'fields': ('tipo_documento', 'cedula', 'nombre', 'telefono', 'cliente', 'activo'),
            'description': 'Documento único; teléfono WhatsApp y organización.',
        }),
        ('Inscripción a curso', {
            'fields': ('cursos_a_inscribir', 'mostrar_cursos_inscritos'),
            'description': (
                'Elija el curso aquí al crear o editar (ej. Cenipalma — Clases Aprende). '
                'Para muchos alumnos use Importar Excel (columna Curso).'
            ),
        }),
        ('Ubicación y perfil', {
            'classes': ('collapse',),
            'fields': ('municipio', 'departamento', 'ubicacion_detalle', 'genero', 'edad', 'rango_edad'),
        }),
    )
    readonly_fields = ('mostrar_cursos_inscritos',)

    def get_fieldsets(self, request, obj=None):
        if obj is None:
            return (
                ('Identificación y contacto', {
                    'fields': ('tipo_documento', 'cedula', 'nombre', 'telefono', 'cliente', 'activo'),
                    'description': 'Documento único; teléfono WhatsApp y organización.',
                }),
                ('Inscripción a curso', {
                    'fields': ('cursos_a_inscribir',),
                    'description': (
                        'Marque el curso al que entra este estudiante. '
                        'Si eligió Cliente arriba, verá los cursos de esa org.'
                    ),
                }),
                ('Ubicación y perfil', {
                    'classes': ('collapse',),
                    'fields': (
                        'municipio', 'departamento', 'ubicacion_detalle',
                        'genero', 'edad', 'rango_edad',
                    ),
                }),
            )
        return self.fieldsets

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        cursos = form.cleaned_data.get('cursos_a_inscribir')
        if not cursos:
            return
        from core.inscripcion_curso import inscribir_estudiante_en_curso

        nuevos = 0
        for curso in cursos:
            _prog, creado = inscribir_estudiante_en_curso(obj, curso)
            if creado:
                nuevos += 1
        if nuevos:
            self.message_user(
                request,
                f'Inscrito en {nuevos} curso(s).',
                level=messages.SUCCESS,
            )
    
    def cedula_formateada(self, obj):
        """Muestra cédula con tipo (sin emoji en listado)."""
        return format_html(
            '<span style="font-variant-numeric:tabular-nums;">{} {}</span>',
            obj.tipo_documento,
            obj.cedula
        )
    cedula_formateada.short_description = "Documento"
    
    def mostrar_cursos_inscritos(self, obj):
        """Muestra los cursos en los que está inscrito el estudiante"""
        progresos = ProgresoEstudiante.objects.filter(estudiante=obj).select_related('curso')
        if not progresos:
            return format_html('<p style="color:#999;">No está inscrito en ningún curso</p>')
        
        html = '<div style="margin-top:10px;">'
        for progreso in progresos:
            porcentaje = progreso.porcentaje_avance()
            color = '#4CAF50' if porcentaje == 100 else '#FF9800' if porcentaje > 0 else '#999'
            html += f'''
                <div style="margin-bottom:15px;padding:10px;background:#f5f5f5;border-radius:5px;">
                    <strong style="color:{color};">📚 {progreso.curso.nombre}</strong><br>
                    <small>Progreso: {porcentaje}%</small>
                    <div style="background:#e0e0e0;height:8px;border-radius:4px;margin-top:5px;">
                        <div style="background:{color};width:{porcentaje}%;height:100%;border-radius:4px;"></div>
                    </div>
                </div>
            '''
        html += '</div>'
        return format_html(html)
    mostrar_cursos_inscritos.short_description = "Cursos"
    
    def cursos_inscritos(self, obj):
        """Muestra cantidad de cursos inscritos"""
        count = ProgresoEstudiante.objects.filter(estudiante=obj).count()
        if count == 0:
            return format_html('<span style="color:#999;">Sin cursos</span>')
        return format_html(
            '<span style="background:#e3f2fd;padding:4px 10px;border-radius:12px;font-size:11px;">{} curso{}</span>',
            count, 's' if count != 1 else ''
        )
    cursos_inscritos.short_description = "📚 Cursos"
    
    def telefono_formateado(self, obj):
        """Muestra teléfono con formato WhatsApp"""
        return f"+{obj.telefono}"
    telefono_formateado.short_description = "WhatsApp"
    
    def cliente_nombre(self, obj):
        """Muestra el cliente al que pertenece"""
        if obj.cliente:
            return obj.cliente.nombre
        return format_html('<span style="color:#999;">Sin cliente</span>')
    cliente_nombre.short_description = "Organización"
    cliente_nombre.admin_order_field = 'cliente__nombre'

    def conversacion_link(self, obj):
        if not obj or not obj.pk:
            return '—'
        url = reverse('conversaciones') + f'?estudiante={obj.pk}'
        return format_html(
            '<a href="{}" style="font-weight:700;color:#7A4E8E;">Chat</a>',
            url,
        )
    conversacion_link.short_description = 'WA'

    def grupos_display(self, obj):
        """Muestra los grupos a los que pertenece el estudiante"""
        grupos = obj.grupos.all()
        if not grupos:
            return format_html('<span style="color:#999;">Sin grupos</span>')
        
        badges = []
        for grupo in grupos[:3]:  # Mostrar máximo 3
            badges.append(f'<span style="background:#e8f5e9;color:#2e7d32;padding:3px 8px;border-radius:8px;font-size:10px;margin-right:3px;">{grupo.emoji} {grupo.nombre}</span>')
        
        html = ''.join(badges)
        if grupos.count() > 3:
            html += f' <span style="color:#999;font-size:10px;">+{grupos.count() - 3} más</span>'
        
        return format_html(html)
    grupos_display.short_description = "👥 Grupos"
    
    # Ver conversaciones removido por solicitud del usuario
    
    def get_urls(self):
        """Agregar URL personalizada para importar estudiantes"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('importar/', self.admin_site.admin_view(self.importar_estudiantes_view), name='core_estudiante_importar'),
            path('exportar-plantilla/', self.admin_site.admin_view(self.exportar_plantilla_importacion), name='core_estudiante_exportar_plantilla'),
        ]
        return custom_urls + urls

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        obj = self.get_object(request, object_id)
        if obj:
            from django.urls import reverse

            from core.models import ProgresoEstudiante, WhatsappLog
            from portal.branding import contexto_identidad_org

            extra_context.update(contexto_identidad_org(getattr(obj, 'cliente', None)))
            extra_context['eki_est_inicial'] = (obj.nombre or '?').strip()[:1].upper()
            progresos = (
                ProgresoEstudiante.objects.filter(estudiante=obj)
                .select_related('curso', 'modulo_actual')
                .order_by('-fecha_ultimo_avance', '-fecha_inicio')[:6]
            )
            extra_context['eki_est_progresos'] = [
                {
                    'curso': p.curso.nombre if p.curso_id else '—',
                    'pct': int(round(p.porcentaje_avance() or 0)),
                    'modulo': (
                        f'M{p.modulo_actual.numero}'
                        if p.modulo_actual_id and p.modulo_actual
                        else '—'
                    ),
                    'completado': p.completado,
                }
                for p in progresos
            ]
            ultimo = (
                WhatsappLog.objects.filter(telefono=obj.telefono)
                .order_by('-fecha')
                .first()
            )
            if ultimo is None and obj.pk:
                ultimo = (
                    WhatsappLog.objects.filter(estudiante_id=obj.pk)
                    .order_by('-fecha')
                    .first()
                )
            extra_context['eki_est_ultimo_wa'] = ultimo
            extra_context['eki_est_conv_url'] = (
                reverse('conversaciones')
                + f'?estudiante={obj.pk}'
            )
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def importar_estudiantes_view(self, request):
        """Vista para importar estudiantes desde Excel.
        Campos obligatorios: Cédula | Nombre | Teléfono.
        Campos opcionales: Municipio | Departamento | Género | Edad | Curso | Cliente.
        """
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from django.db import IntegrityError
        import re
        
        if request.method == 'POST':
            archivo = request.FILES.get('archivo_excel')
            
            if not archivo:
                messages.error(request, "⚠️ Debes seleccionar un archivo Excel")
                return redirect('admin:core_estudiante_changelist')
            
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                
                creados = 0
                actualizados = 0
                inscritos = 0
                errores = []
                avisos = []
                
                from core.documento_identidad import (
                    normalizar_numero_documento,
                    normalizar_tipo_documento,
                )
                from core.excel_celdas import celda_excel_a_texto
                from core.import_estudiantes_excel import (
                    extraer_fila_estudiante,
                    mapear_columnas_estudiante,
                )
                from core.utils_telefono import normalizar_telefono, validar_telefono_whatsapp

                def _normalizar_celda(val):
                    return celda_excel_a_texto(val)
                
                def _limpiar_texto(val):
                    if not val:
                        return ''
                    return re.sub(r'\s+', ' ', val.strip().lower())

                # Detectar encabezados (por nombre de columna, no solo posición)
                primera = [_normalizar_celda(c).lower() for c in (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []) or [])]
                colmap = mapear_columnas_estudiante(primera)
                tiene_tipo_col = bool(primera) and (
                    'tipo' in (primera[0] or '')
                    or (primera[0] or '') in ('tipo documento', 'tipo_documento', 'tipodocumento')
                )
                data_start = 2 if (primera and (
                    'cedula' in (primera[0] if not tiene_tipo_col else (primera[1] if len(primera) > 1 else ''))
                    or 'documento' in ' '.join(primera[:4])
                    or 'nombre' in ' '.join(primera[:5])
                    or 'tel' in ' '.join(primera[:5])
                    or colmap is not None
                )) else 2
                
                GENEROS_VALIDOS = {'m': 'M', 'f': 'F', 'o': 'O', 'masculino': 'M', 'femenino': 'F',
                                   'otro': 'O', 'hombre': 'M', 'mujer': 'F', 'nr': 'NR', 'no reporta': 'NR'}
                
                for idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row[:3]):
                        continue
                    
                    try:
                        campos = extraer_fila_estudiante(row, colmap, tiene_tipo_col)
                        tipo_raw = campos['tipo_raw']
                        cedula = campos['cedula']
                        nombre = campos['nombre']
                        telefono_raw = campos['telefono_raw']
                        municipio = _limpiar_texto(campos['municipio'])
                        departamento = _limpiar_texto(campos['departamento'])
                        genero_raw = _limpiar_texto(campos['genero_raw'])
                        edad_raw = campos['edad_raw']
                        curso_nombre = campos['curso_nombre']
                        cliente_nombre = campos['cliente_nombre']
                    except IndexError:
                        errores.append(f"Fila {idx}: Columnas insuficientes")
                        continue
                    
                    tipo_documento = normalizar_tipo_documento(tipo_raw or 'CC')
                    cedula = normalizar_numero_documento(cedula)
                    
                    # Validar obligatorios mínimos
                    campos_faltantes = []
                    if not cedula: campos_faltantes.append('Documento')
                    if not nombre: campos_faltantes.append('Nombre')
                    if not telefono_raw: campos_faltantes.append('Teléfono')
                    
                    if campos_faltantes:
                        errores.append(f"Fila {idx}: Faltan: {', '.join(campos_faltantes)}")
                        continue
                    
                    telefono = campos['telefono_normalizado'] or normalizar_telefono(telefono_raw)
                    check = validar_telefono_whatsapp(telefono_raw)
                    if not check['ok']:
                        errores.append(f"Fila {idx}: {check['mensaje']}")
                        continue
                    telefono = check['telefono'] or telefono
                    if check['severity'] == 'warn':
                        avisos.append(f"Fila {idx}: {check['mensaje']}")
                    
                    genero = GENEROS_VALIDOS.get(genero_raw, '') if genero_raw else ''
                    if not genero:
                        genero = 'NR'
                    
                    # Validar edad
                    edad = None
                    if edad_raw:
                        try:
                            edad = int(re.sub(r'\D', '', edad_raw))
                            if edad < 1 or edad > 120:
                                errores.append(f"Fila {idx}: Edad '{edad_raw}' fuera de rango (1-120)")
                                continue
                        except (ValueError, TypeError):
                            errores.append(f"Fila {idx}: Edad '{edad_raw}' no es un número válido")
                            continue
                    
                    try:
                        cliente = None
                        if cliente_nombre:
                            try:
                                cliente = Cliente.objects.get(nombre__iexact=cliente_nombre.strip())
                            except Cliente.DoesNotExist:
                                errores.append(f"Fila {idx}: Cliente '{cliente_nombre}' no encontrado")
                        
                        try:
                            defaults = {
                                'nombre': nombre.strip().title(),
                                'telefono': telefono,
                                'municipio': municipio,
                                'departamento': departamento,
                                'genero': genero,
                                'edad': edad,
                                'tipo_documento': tipo_documento,
                                'estado_onboarding': 'completado',
                                'estado_chat': 'ACTIVO',
                                'acepto_terminos': True,
                                'activo': True,
                            }
                            if cliente:
                                defaults['cliente'] = cliente
                            
                            estudiante, created = Estudiante.objects.update_or_create(
                                cedula=cedula,
                                defaults=defaults
                            )
                            if created:
                                creados += 1
                            else:
                                actualizados += 1
                        except IntegrityError as e:
                            if 'telefono' in str(e).lower():
                                errores.append(
                                    f"Fila {idx}: El teléfono '{telefono}' ya está en otro estudiante. "
                                    f"Use otro número o actualice ese registro (misma cédula)."
                                )
                            else:
                                errores.append(f"Fila {idx}: Error de integridad - {str(e)}")
                            continue
                        
                        if curso_nombre:
                            try:
                                from core.inscripcion_curso import (
                                    inscribir_estudiante_en_curso,
                                    resolver_curso_por_nombre,
                                )

                                curso = resolver_curso_por_nombre(
                                    curso_nombre,
                                    cliente_nombre=cliente_nombre or None,
                                )
                                if curso is None:
                                    errores.append(
                                        f"Fila {idx}: Curso '{curso_nombre}' no encontrado"
                                    )
                                else:
                                    _progreso, creado_prog = inscribir_estudiante_en_curso(
                                        estudiante, curso,
                                    )
                                    if creado_prog:
                                        inscritos += 1
                            except Exception as exc_insc:
                                errores.append(f"Fila {idx}: Inscripción curso — {exc_insc}")
                    
                    except Exception as e:
                        errores.append(f"Fila {idx}: {str(e)}")
                
                if creados > 0:
                    messages.success(request, f"✅ {creados} estudiante(s) creado(s)")
                if actualizados > 0:
                    messages.info(request, f"ℹ️ {actualizados} estudiante(s) actualizado(s)")
                if inscritos > 0:
                    messages.success(request, f"🎓 {inscritos} inscripción(es) en cursos")
                if errores:
                    messages.warning(request, f"⚠️ {len(errores)} error(es)")
                    for error in errores[:5]:
                        messages.warning(request, error)
                if avisos:
                    messages.info(request, f"ℹ️ {len(avisos)} aviso(s) de teléfono (se guardaron igual)")
                    for aviso in avisos[:5]:
                        messages.info(request, aviso)
                
                return redirect('admin:core_estudiante_changelist')
            
            except Exception as e:
                messages.error(request, f"❌ Error procesando archivo: {str(e)}")
                return redirect('admin:core_estudiante_changelist')
        
        return render(request, 'admin/importar_estudiantes.html', {
            'title': 'Importar Estudiantes desde Excel',
            'site_header': 'Importar Estudiantes',
        })
    
    def total_mensajes(self, obj):
        """Cuenta cuántos mensajes ha enviado el estudiante"""
        count = WhatsappLog.objects.filter(telefono=obj.telefono).count()
        return format_html(
            '<span style="background:#e3f2fd;padding:4px 8px;border-radius:4px;">{} mensajes</span>',
            count
        )
    total_mensajes.short_description = "💬 Total"
    
    # ========================================
    # 📢 MEGÁFONO: ENVÍO MASIVO DE MENSAJES
    # ========================================
    
    @admin.action(description='📢 Megáfono: Enviar mensaje a seleccionados')
    def enviar_mensaje_masivo(self, request, queryset):
        """
        Permite enviar un mensaje personalizado a múltiples estudiantes seleccionados.
        Útil para avisos de mantenimiento, anuncios importantes, etc.
        
        IMPORTANTE: Valida la ventana de 24 horas de WhatsApp para evitar bloqueos.
        """
        from datetime import timedelta
        
        if 'aplicar' in request.POST:
            # El usuario confirmó el envío
            mensaje = request.POST.get('mensaje')
            if not mensaje:
                self.message_user(request, "⚠️ Debes escribir un mensaje", level=messages.ERROR)
                return
            
            # Verificar que no esté vacío
            if not mensaje.strip():
                self.message_user(request, "⚠️ El mensaje no puede estar vacío", level=messages.ERROR)
                return
            
            enviados = 0
            errores = 0
            errores_detalle = []
            
            for estudiante in queryset:
                try:
                    # Personalizar mensaje con nombre del estudiante
                    mensaje_personalizado = mensaje.replace('{nombre}', estudiante.nombre)
                    mensaje_personalizado = mensaje_personalizado.replace('{cedula}', estudiante.cedula)
                    
                    # Enviar mensaje
                    resultado = enviar_whatsapp(estudiante.telefono, mensaje_personalizado)
                    
                    if resultado:
                        enviados += 1
                        logger.info(f"📢 Megáfono: Mensaje enviado a {estudiante.nombre}")
                        
                        # Registrar en WhatsappLog
                        WhatsappLog.objects.create(
                            telefono=estudiante.telefono,
                            mensaje=mensaje_personalizado,
                            tipo='SENT',
                            estado='ENVIADO',
                            estudiante=estudiante,
                            mensaje_id=f'megafono_{timezone.now().timestamp()}'
                        )
                    else:
                        errores += 1
                        errores_detalle.append(f"{estudiante.nombre} ({estudiante.telefono})")
                        logger.error(f"❌ Megáfono: Error enviando a {estudiante.nombre}")
                        
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"{estudiante.nombre}: {str(e)}")
                    logger.error(f"❌ Megáfono: Error con {estudiante.nombre}: {e}")
            
            # Mostrar resultados
            if enviados > 0:
                self.message_user(
                    request,
                    f"✅ Mensaje enviado exitosamente a {enviados} estudiante(s)",
                    level=messages.SUCCESS
                )
            
            if errores > 0:
                mensaje_error = f"⚠️ Hubo {errores} error(es) al enviar"
                if errores_detalle:
                    mensaje_error += f": {', '.join(errores_detalle[:5])}"
                    if len(errores_detalle) > 5:
                        mensaje_error += f" y {len(errores_detalle) - 5} más..."
                self.message_user(request, mensaje_error, level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # Validar ventana de 24 horas ANTES de mostrar formulario
        hace_24h = timezone.now() - timedelta(hours=24)
        activos = []
        inactivos = []
        sin_interaccion = []
        
        for estudiante in queryset:
            # Obtener último mensaje RECIBIDO del estudiante (no enviado por nosotros)
            ultimo_msg = WhatsappLog.objects.filter(
                estudiante=estudiante,
                tipo_mensaje='recibido'  # Solo mensajes que EL estudiante nos envió
            ).order_by('-timestamp').first()
            
            if ultimo_msg:
                if ultimo_msg.timestamp >= hace_24h:
                    activos.append({
                        'estudiante': estudiante,
                        'ultima_interaccion': ultimo_msg.timestamp,
                        'hace': timezone.now() - ultimo_msg.timestamp
                    })
                else:
                    inactivos.append({
                        'estudiante': estudiante,
                        'ultima_interaccion': ultimo_msg.timestamp,
                        'hace': timezone.now() - ultimo_msg.timestamp
                    })
            else:
                # Nunca ha escrito
                sin_interaccion.append(estudiante)
        
        # Mostrar formulario de confirmación con advertencias
        return render(request, 'admin/enviar_mensaje_masivo.html', {
            'estudiantes': queryset,
            'total_estudiantes': queryset.count(),
            'activos': activos,
            'inactivos': inactivos,
            'sin_interaccion': sin_interaccion,
            'total_activos': len(activos),
            'total_inactivos': len(inactivos),
            'total_sin_interaccion': len(sin_interaccion),
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
            'title': 'Enviar Mensaje Masivo (Megáfono)',
        })
    
    # ========================================
    # 📢 ANUNCIO GRUPAL (Difusión Simulada)
    # ========================================
    
    @admin.action(description='📣 Enviar anuncio grupal (difusión)')
    def enviar_anuncio_grupal(self, request, queryset):
        """
        Envía el mismo mensaje a múltiples estudiantes usando template aprobado de Twilio.
        Ideal para: PQRS, anuncios de clases, notificaciones importantes.
        """
        from core.utils import enviar_whatsapp_twilio_content_template
        
        if 'aplicar' in request.POST:
            mensaje = request.POST.get('mensaje', '').strip()
            
            # Debug: Ver qué datos llegan
            logger.info(f"📋 POST data: {dict(request.POST)}")
            
            if not mensaje:
                self.message_user(request, "⚠️ Debes escribir un mensaje", level=messages.ERROR)
                # Regresar al formulario con los estudiantes seleccionados
                return render(request, 'admin/enviar_anuncio_grupal.html', {
                    'estudiantes': queryset,
                    'total_estudiantes': queryset.count(),
                    'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
                    'title': 'Enviar Anuncio Grupal (Difusión)',
                    'mensaje': mensaje,  # Preservar mensaje ingresado
                })
            
            # Reconstruir queryset desde los IDs en POST
            estudiante_ids = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
            if estudiante_ids:
                queryset = Estudiante.objects.filter(id__in=estudiante_ids)
            
            if not queryset.exists():
                self.message_user(request, "⚠️ No se encontraron estudiantes seleccionados", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')
            
            # Verificar que el template esté configurado
            content_sid = getattr(settings, 'TWILIO_TEMPLATE_ANUNCIO_GRUPAL', None)
            logger.info(f"🔧 Content SID: {content_sid}")
            
            if not content_sid:
                self.message_user(
                    request, 
                    "⚠️ Template de Twilio no configurado. Configura TWILIO_TEMPLATE_ANUNCIO_GRUPAL en .env",
                    level=messages.ERROR
                )
                # Regresar al formulario con el mensaje preservado
                return render(request, 'admin/enviar_anuncio_grupal.html', {
                    'estudiantes': queryset,
                    'total_estudiantes': queryset.count(),
                    'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
                    'title': 'Enviar Anuncio Grupal (Difusión)',
                    'mensaje': mensaje,
                })
            
            enviados = 0
            errores = 0
            errores_detalle = []
            
            for estudiante in queryset:
                try:
                    # Variables del template:
                    # {{1}} = Nombre del estudiante
                    # {{2}} = Contenido del mensaje
                    # {{3}} = Enlace del grupo
                    variables = {
                        '1': estudiante.nombre,
                        '2': mensaje
                    }
                    
                    # Enviar con template de Twilio
                    resultado = enviar_whatsapp_twilio_content_template(
                        telefono=estudiante.telefono,
                        content_sid=content_sid,
                        variables=variables
                    )
                    
                    if resultado.get('success'):
                        enviados += 1
                        logger.info(f"📣 Anuncio grupal enviado a {estudiante.nombre}")
                    else:
                        errores += 1
                        errores_detalle.append(estudiante.nombre)
                        logger.error(f"❌ Error enviando anuncio a {estudiante.nombre}: {resultado.get('response')}")
                        
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"{estudiante.nombre}: {str(e)}")
                    logger.error(f"❌ Error con {estudiante.nombre}: {e}")
            
            # Mostrar resultados
            if enviados > 0:
                self.message_user(
                    request,
                    f"✅ Anuncio enviado exitosamente a {enviados} estudiante(s)",
                    level=messages.SUCCESS
                )
            
            if errores > 0:
                mensaje_error = f"⚠️ Hubo {errores} error(es)"
                if errores_detalle:
                    mensaje_error += f": {', '.join(errores_detalle[:3])}"
                    if len(errores_detalle) > 3:
                        mensaje_error += f" y {len(errores_detalle) - 3} más"
                self.message_user(request, mensaje_error, level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # Mostrar formulario
        return render(request, 'admin/enviar_anuncio_grupal.html', {
            'estudiantes': queryset,
            'total_estudiantes': queryset.count(),
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
            'title': 'Enviar Anuncio Grupal (Difusión)',
        })
    
    # ========================================
    # 👥 INVITACIÓN A GRUPO DE WHATSAPP
    # ========================================
    
    @admin.action(description='🔗 Invitar a grupo de WhatsApp')
    def invitar_a_grupo_whatsapp(self, request, queryset):
        """
        Envía invitación con enlace al grupo de WhatsApp.
        Puede usar el enlace del curso o del cliente, o uno personalizado.
        """
        if 'aplicar' in request.POST:
            enlace_tipo = request.POST.get('enlace_tipo')  # 'curso', 'cliente', 'personalizado'
            curso_id = request.POST.get('curso_id')
            enlace_personalizado = request.POST.get('enlace_personalizado')
            mensaje_intro = request.POST.get('mensaje_intro', '')
            
            # Determinar el enlace a usar
            enlace_grupo = None
            
            if enlace_tipo == 'curso' and curso_id:
                try:
                    curso = Curso.objects.get(id=curso_id)
                    enlace_grupo = curso.enlace_grupo_whatsapp
                    if not enlace_grupo:
                        self.message_user(
                            request,
                            f"⚠️ El curso '{curso.nombre}' no tiene un enlace de grupo configurado",
                            level=messages.ERROR
                        )
                        return
                except Curso.DoesNotExist:
                    self.message_user(request, "⚠️ Curso no encontrado", level=messages.ERROR)
                    return
            
            elif enlace_tipo == 'cliente':
                # Usar el enlace del cliente del primer estudiante
                primer_estudiante = queryset.first()
                if primer_estudiante and primer_estudiante.cliente:
                    enlace_grupo = primer_estudiante.cliente.enlace_grupo_whatsapp
                    if not enlace_grupo:
                        self.message_user(
                            request,
                            f"⚠️ El cliente '{primer_estudiante.cliente.nombre}' no tiene un enlace de grupo configurado",
                            level=messages.ERROR
                        )
                        return
                else:
                    self.message_user(
                        request,
                        "⚠️ Los estudiantes no tienen cliente asignado o el cliente no tiene enlace de grupo",
                        level=messages.ERROR
                    )
                    return
            
            elif enlace_tipo == 'personalizado':
                enlace_grupo = enlace_personalizado
                if not enlace_grupo or not enlace_grupo.startswith('https://chat.whatsapp.com/'):
                    self.message_user(
                        request,
                        "⚠️ Enlace personalizado inválido. Debe ser: https://chat.whatsapp.com/xxxxx",
                        level=messages.ERROR
                    )
                    return
            
            if not enlace_grupo:
                self.message_user(request, "⚠️ No se pudo determinar el enlace del grupo", level=messages.ERROR)
                return
            
            # Construir mensaje
            mensaje_base = mensaje_intro if mensaje_intro else "¡Hola! Te invitamos a unirte a nuestro grupo de WhatsApp:"
            
            enviados = 0
            errores = 0
            errores_detalle = []
            
            for estudiante in queryset:
                try:
                    from core.utils import enviar_whatsapp_twilio_content_template
                    
                    # Verificar que el template esté configurado
                    content_sid = settings.TWILIO_TEMPLATE_INVITACION_GRUPO
                    if not content_sid:
                        self.message_user(
                            request,
                            "⚠️ Template de invitación no configurado. Configura TWILIO_TEMPLATE_INVITACION_GRUPO en .env",
                            level=messages.ERROR
                        )
                        return
                    
                    # Variables del template:
                    # {{1}} = Nombre del estudiante
                    # {{2}} = Mensaje de introducción (opcional)
                    # {{3}} = Enlace del grupo
                    mensaje_intro_final = mensaje_intro if mensaje_intro else "¡Hola! Te invitamos a unirte a nuestro grupo de WhatsApp."
                    
                    variables = {
                        '1': estudiante.nombre,
                        '2': mensaje_intro_final,
                        '3': enlace_grupo
                    }
                    
                    # Enviar con template de Twilio
                    resultado = enviar_whatsapp_twilio_content_template(
                        telefono=estudiante.telefono,
                        content_sid=content_sid,
                        variables=variables
                    )
                    
                    if resultado.get('success'):
                        enviados += 1
                        logger.info(f"🔗 Invitación a grupo enviada a {estudiante.nombre}")
                    else:
                        errores += 1
                        errores_detalle.append(estudiante.nombre)
                        logger.error(f"❌ Error enviando invitación a {estudiante.nombre}: {resultado.get('response')}")
                        
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"{estudiante.nombre}: {str(e)}")
                    logger.error(f"❌ Error con {estudiante.nombre}: {e}")
            
            # Mostrar resultados
            if enviados > 0:
                self.message_user(
                    request,
                    f"✅ Invitación enviada exitosamente a {enviados} estudiante(s)",
                    level=messages.SUCCESS
                )
            
            if errores > 0:
                mensaje_error = f"⚠️ {errores} error(es)"
                if errores_detalle:
                    mensaje_error += f": {', '.join(errores_detalle[:3])}"
                self.message_user(request, mensaje_error, level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # GET - Mostrar formulario
        # Obtener cursos y clientes con enlaces configurados
        cursos_con_enlace = Curso.objects.exclude(enlace_grupo_whatsapp='').values('id', 'nombre', 'emoji')
        
        # Verificar si los estudiantes tienen cliente
        clientes_ids = queryset.values_list('cliente_id', flat=True).distinct()
        clientes_con_enlace = Cliente.objects.filter(
            id__in=clientes_ids
        ).exclude(enlace_grupo_whatsapp='').values('id', 'nombre')
        
        return render(request, 'admin/invitar_grupo_whatsapp.html', {
            'estudiantes': queryset,
            'total_estudiantes': queryset.count(),
            'cursos_con_enlace': cursos_con_enlace,
            'clientes_con_enlace': clientes_con_enlace,
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
            'title': 'Invitar a Grupo de WhatsApp',
        })
    
    # ✅ Acción para enviar mensaje manual de prueba
    def enviar_mensaje_manual(self, request, queryset):
        """Permite enviar un mensaje de prueba a los estudiantes seleccionados"""
        if 'aplicar' in request.POST:
            # El usuario confirmó el envío
            mensaje = request.POST.get('mensaje')
            if not mensaje:
                self.message_user(request, "⚠️ Debes escribir un mensaje", level=messages.ERROR)
                return
            
            enviados = 0
            errores = 0
            for estudiante in queryset:
                try:
                    telefono = estudiante.telefono
                    if not telefono.startswith('whatsapp:'):
                        telefono = f'whatsapp:{telefono}'
                    
                    # Enviar con Twilio
                    resultado = enviar_whatsapp_twilio(
                        telefono=telefono,
                        texto=mensaje,
                        mensaje_id_referencia=None
                    )
                    
                    if resultado.get('success'):
                        # Registrar en log
                        WhatsappLog.objects.create(
                            telefono=estudiante.telefono,
                            mensaje=mensaje,
                            mensaje_id=resultado.get('mensaje_id'),
                            tipo='SENT',
                            estado='sent'
                        )
                        enviados += 1
                        logger.info(f"✅ Mensaje manual enviado a {estudiante.nombre} ({estudiante.telefono})")
                    else:
                        errores += 1
                        logger.error(f"❌ Error al enviar a {estudiante.telefono}: {resultado.get('error')}")
                
                except Exception as e:
                    errores += 1
                    logger.error(f"❌ Excepción al enviar a {estudiante.telefono}: {str(e)}")
            
            if enviados > 0:
                self.message_user(request, f"✅ Mensaje enviado exitosamente a {enviados} estudiante(s)", level=messages.SUCCESS)
            if errores > 0:
                self.message_user(request, f"⚠️ Hubo {errores} error(es) al enviar", level=messages.WARNING)
            
            # Redirigir a la lista de estudiantes
            return redirect('admin:core_estudiante_changelist')
        
        # Mostrar formulario de confirmación
        return render(request, 'admin/enviar_mensaje_manual.html', {
            'estudiantes': queryset,
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        })
    
    enviar_mensaje_manual.short_description = "📤 Enviar mensaje de prueba"
    
    @admin.action(description='📊 Exportar estudiantes con cursos a Excel')
    def exportar_estudiantes_por_curso(self, request, queryset):
        """Exporta estudiantes con información detallada de cursos"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        from core.export_estudiantes import limpiar_telefono
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Estudiantes por Curso"
        
        # Encabezados
        headers = [
            'Nombre', 'Apellido', 'Documento', 'Teléfono', 'Email',
            'Municipio', 'Departamento', 'Ciudad',
            'Organización', 'Grupo', 'Curso', 'Estado', 'Progreso (%)',
            'Fecha registro', 'Activo',
            'Cursos Inscritos', 'Cursos Completados', 'Progreso Promedio',
            'Total Mensajes',
        ]
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for estudiante in queryset.select_related('cliente').prefetch_related('grupos'):
            # Calcular estadísticas de cursos
            progresos = ProgresoEstudiante.objects.filter(estudiante=estudiante).select_related('curso')
            total_cursos = progresos.count()
            cursos_completados = progresos.filter(completado=True).count()
            
            # Calcular promedio de progreso
            if total_cursos > 0:
                progreso_promedio = sum([p.porcentaje_avance() for p in progresos]) / total_cursos
            else:
                progreso_promedio = 0

            primer_progreso = progresos.first()
            curso_nombre = primer_progreso.curso.nombre if primer_progreso and primer_progreso.curso_id else ''
            if total_cursos > 1:
                curso_nombre = f"{curso_nombre} (+{total_cursos - 1})" if curso_nombre else f"{total_cursos} cursos"
            estado_prog = 'Sin inscripción'
            pct_prog = 0
            if primer_progreso:
                pct_prog = primer_progreso.porcentaje_avance()
                estado_prog = 'Completado' if primer_progreso.completado else (
                    'En curso' if pct_prog > 0 else 'Sin avance'
                )
            
            # Total mensajes
            total_mensajes = WhatsappLog.objects.filter(telefono=estudiante.telefono).count()
            grupos_txt = ', '.join(g.nombre for g in estudiante.grupos.all()[:5])
            municipio = (estudiante.municipio or '').strip()
            departamento = (getattr(estudiante, 'departamento', '') or '').strip()
            ciudad = municipio
            
            ws.append([
                estudiante.nombre,
                '',
                estudiante.cedula,
                limpiar_telefono(estudiante.telefono),
                '',
                municipio,
                departamento,
                ciudad,
                estudiante.cliente.nombre if estudiante.cliente else 'Sin cliente',
                grupos_txt,
                curso_nombre,
                estado_prog,
                f"{pct_prog:.1f}%",
                estudiante.fecha_registro.strftime('%Y-%m-%d %H:%M'),
                'Sí' if estudiante.activo else 'No',
                total_cursos,
                cursos_completados,
                f"{progreso_promedio:.1f}%",
                total_mensajes,
            ])
        
        # Crear segunda hoja con detalle de cursos
        ws2 = wb.create_sheet("Detalle por Curso")
        headers2 = [
            'Documento', 'Estudiante', 'Teléfono', 'Municipio', 'Departamento',
            'Curso', 'Progreso', 'Estado', 'Fecha Inicio',
        ]
        ws2.append(headers2)
        
        # Estilo
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Detalle de cada curso
        for estudiante in queryset:
            progresos = ProgresoEstudiante.objects.filter(estudiante=estudiante).select_related('curso')
            for progreso in progresos:
                ws2.append([
                    estudiante.cedula,
                    estudiante.nombre,
                    limpiar_telefono(estudiante.telefono),
                    estudiante.municipio or '',
                    getattr(estudiante, 'departamento', '') or '',
                    progreso.curso.nombre,
                    f"{progreso.porcentaje_avance()}%",
                    "Completado" if progreso.completado else "En progreso",
                    progreso.fecha_inicio.strftime('%Y-%m-%d'),
                ])
        
        # Ajustar ancho de columnas
        for ws in [wb.active, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=estudiantes_cursos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        self.message_user(request, f"✅ Exportados {queryset.count()} estudiantes con detalle de cursos", level=messages.SUCCESS)
        return response
    
    def exportar_plantilla_importacion(self, request):
        """Genera plantilla Excel mejorada con curso y cliente"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.comments import Comment
        # from openpyxl.data_validation import DataValidation  # Eliminado, no se usa
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Estudiantes"
        
        # Encabezados LatAm: Tipo documento + Documento (+ columnas previas)
        headers = [
            'Tipo documento', 'Documento', 'Nombre Completo', 'Teléfono',
            'Municipio', 'Departamento', 'Género', 'Edad', 'Curso', 'Cliente',
        ]
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Comentarios con instrucciones
        ws['A1'].comment = Comment(
            "Tipo: CC, TI, CE, DUI (SV), CURP/INE (MX), DNI, RUT (CL), DPI (GT), CI, PP, OTRO",
            "eki",
        )
        ws['B1'].comment = Comment(
            "Número/ID sin puntos ni espacios. Puede tener letras (CURP, RUT).",
            "eki",
        )
        ws['C1'].comment = Comment("Nombre completo del estudiante", "eki")
        ws['D1'].comment = Comment(
            "WhatsApp CON código de país.\n"
            "IMPORTANTE: formatee la columna como TEXTO (no número).\n"
            "CO: 573001234567 | SV: 50371234567 | MX: 5215512345678",
            "eki",
        )
        ws['E1'].comment = Comment("Municipio / ciudad", "eki")
        ws['F1'].comment = Comment("Departamento / estado / provincia", "eki")
        ws['G1'].comment = Comment("Género: masculino, femenino, otro, no reporta", "eki")
        ws['H1'].comment = Comment("Edad en años", "eki")
        ws['I1'].comment = Comment(
            "Nombre EXACTO del curso (hoja Valores Disponibles).\n"
            "Ej. Cenipalma — Clases Aprende\n"
            "Si es Clases·Aprende, estudia en aula (sin *listo* por WA).",
            "eki",
        )
        ws['J1'].comment = Comment(
            "Nombre EXACTO del Cliente/org (ej. Cenipalma).",
            "eki",
        )
        
        # Obtener cursos y clientes para validación
        cursos = Curso.objects.filter(activo=True).order_by('nombre')
        clientes = Cliente.objects.filter(activo=True).order_by('nombre')
        
        # Agregar ejemplos con datos reales (teléfono como TEXTO para evitar notación científica)
        curso_ejemplo = cursos.first().nombre if cursos.exists() else 'Curso de Café'
        cliente_ejemplo = clientes.first().nombre if clientes.exists() else 'FNC'
        ws.append(['CC', '1234567890', 'Juan Pérez García', '573001234567', 'Manizales', 'Caldas', 'masculino', 35, curso_ejemplo, cliente_ejemplo])
        ws.append(['DUI', '012345678', 'Ana Martínez', '50371234567', 'San Salvador', 'San Salvador', 'femenino', 28, curso_ejemplo, cliente_ejemplo])
        ws.append(['CURP', 'PEGJ850101HDFRRN09', 'José Pérez', '5215512345678', 'CDMX', 'Ciudad de México', 'masculino', 40, '', ''])
        
        # Fila vacía para empezar
        ws.append(['', '', '', '', '', '', '', '', '', ''])

        # Forzar Texto en Documento (B) y Teléfono (D)
        from openpyxl.styles.numbers import FORMAT_TEXT
        for r in range(2, 202):
            ws[f'B{r}'].number_format = FORMAT_TEXT
            ws[f'D{r}'].number_format = FORMAT_TEXT
            for col in ('B', 'D'):
                v = ws[f'{col}{r}'].value
                if v is not None and v != '':
                    ws[f'{col}{r}'].value = str(v)
        
        # Estilo para ejemplos
        example_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        for row in [2, 3, 4]:
            for cell in ws[row]:
                cell.fill = example_fill
                cell.font = Font(italic=True, color="666666", size=10)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 28
        ws.column_dimensions['J'].width = 22
        
        # Crear hoja con listas de valores disponibles
        ws_ref = wb.create_sheet("Valores Disponibles")
        ws_ref.append(["CURSOS DISPONIBLES", "CLIENTES DISPONIBLES"])
        ws_ref['A1'].font = Font(bold=True, size=12, color="2196F3")
        ws_ref['B1'].font = Font(bold=True, size=12, color="2196F3")
        
        max_rows = max(cursos.count(), clientes.count())
        for i in range(max_rows):
            curso_nombre = cursos[i].nombre if i < cursos.count() else ""
            cliente_nombre = clientes[i].nombre if i < clientes.count() else ""
            ws_ref.append([curso_nombre, cliente_nombre])
        
        ws_ref.column_dimensions['A'].width = 35
        ws_ref.column_dimensions['B'].width = 35
        
        # Agregar hoja de instrucciones
        ws_inst = wb.create_sheet("Instrucciones")
        instrucciones = [
            ["GUÍA RÁPIDA - IMPORTAR ESTUDIANTES A eki (LatAm)"],
            [""],
            ["CAMPOS OBLIGATORIOS:"],
            ["   • Tipo documento: CC, DUI (SV), CURP/INE (MX), DNI, RUT, DPI, CI, PP, OTRO"],
            ["   • Documento: número/ID (puede tener letras, ej. CURP)"],
            ["   • Nombre: Nombre completo del estudiante"],
            ["   • Teléfono: columna D en TEXTO. Ej CO: 573001234567 (no notación científica)"],
            ["   • Municipio / Departamento / Género / Edad"],
            [""],
            ["CAMPOS OPCIONALES:"],
            ["   • Curso / Cliente: ver hoja 'Valores Disponibles'"],
            [""],
            ["COMPATIBILIDAD:"],
            ["   • Plantillas viejas (sin columna Tipo) siguen funcionando → se asume CC"],
            ["   • No uses números de 10 dígitos sin código país fuera de Colombia"],
            [""],
            ["EE.UU.: plantillas Marketing WhatsApp están bloqueadas por Meta (+1)."],
            [""],
            ["PROCESO:"],
            ["   1. Completa 'Plantilla Estudiantes'"],
            ["   2. Admin → Estudiantes → Importar desde Excel"],
            [""],
            ["ERRORES COMUNES: documento o teléfono duplicado; curso/cliente inexistente"],
        ]
        
        for row_data in instrucciones:
            ws_inst.append(row_data)
        
        # Estilo del título
        ws_inst['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws_inst['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar ancho
        ws_inst.column_dimensions['A'].width = 80
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=PLANTILLA_IMPORTACION_ESTUDIANTES_eki.xlsx'
        wb.save(response)
        
        # Solo mostrar mensaje si es una acción de admin
        self.message_user(request, "Plantilla de importación descargada. Comparte este archivo con tus clientes.", level=messages.SUCCESS)
        
        return response
    
    def delete_model(self, request, obj):
        """Eliminar un estudiante individual con manejo seguro de relaciones CASCADE"""
        try:
            nombre = obj.nombre or obj.cedula
            # Eliminar relaciones en orden para evitar timeouts
            from core.models import (
                WhatsappLog, EnvioLog, ProgresoEstudiante, ModuloCompletado,
                ResultadoExamen, RespuestaEjercicio, InteraccionLog,
                SolicitudSoporte, Certificado
            )
            from core.gamificacion import PerfilGamificacion, BadgeEstudiante, TransaccionPuntos
            from core.models_extras import EnvioProgramado, PQRS, InvitacionGrupo, GrupoEstudiantes
            from core.recompensas import CanjeRecompensa
            
            # Limpiar relaciones vinculadas al teléfono
            if obj.telefono:
                WhatsappLog.objects.filter(telefono=obj.telefono).delete()
            
            # Limpiar WhatsappLog por FK (SET_NULL)
            WhatsappLog.objects.filter(estudiante=obj).update(estudiante=None)
            
            # Limpiar relaciones por FK directas
            EnvioLog.objects.filter(estudiante=obj).delete()
            InteraccionLog.objects.filter(estudiante=obj).delete()
            SolicitudSoporte.objects.filter(estudiante=obj).delete()
            RespuestaEjercicio.objects.filter(estudiante=obj).delete()
            ResultadoExamen.objects.filter(estudiante=obj).delete()
            
            # Modelos extras
            EnvioProgramado.objects.filter(estudiante=obj).delete()
            PQRS.objects.filter(estudiante=obj).delete()
            InvitacionGrupo.objects.filter(estudiante=obj).delete()
            CanjeRecompensa.objects.filter(estudiante=obj).delete()
            
            # M2M relations
            for grupo in GrupoEstudiantes.objects.filter(estudiantes=obj):
                grupo.estudiantes.remove(obj)
            
            # Progresos y sus dependencias
            progresos = ProgresoEstudiante.objects.filter(estudiante=obj)
            for prog in progresos:
                ModuloCompletado.objects.filter(progreso=prog).delete()
            progresos.delete()
            
            Certificado.objects.filter(estudiante=obj).delete()
            TransaccionPuntos.objects.filter(perfil__estudiante=obj).delete()
            BadgeEstudiante.objects.filter(estudiante=obj).delete()
            PerfilGamificacion.objects.filter(estudiante=obj).delete()
            
            obj.delete()
            self.message_user(request, f"✅ Estudiante '{nombre}' eliminado correctamente", level=messages.SUCCESS)
        except Exception as e:
            self.message_user(request, f"❌ Error eliminando estudiante: {str(e)}", level=messages.ERROR)

    def delete_queryset(self, request, queryset):
        """Eliminar múltiples estudiantes con manejo seguro"""
        total = queryset.count()
        errores = 0
        for estudiante in queryset:
            try:
                self.delete_model(request, estudiante)
            except Exception:
                errores += 1
        if errores:
            self.message_user(request, f"⚠️ {total - errores} eliminados, {errores} con error", level=messages.WARNING)

    def eliminar_estudiantes_seguro(self, request, queryset):
        """Eliminar estudiantes seleccionados con limpieza de datos relacionados"""
        total = queryset.count()
        errores = 0
        for estudiante in queryset:
            try:
                self.delete_model(request, estudiante)
            except Exception as e:
                errores += 1
        if errores == 0:
            self.message_user(request, f"✅ {total} estudiante(s) eliminado(s) correctamente", level=messages.SUCCESS)
        else:
            self.message_user(request, f"⚠️ {total - errores} eliminados, {errores} con error", level=messages.WARNING)
    eliminar_estudiantes_seguro.short_description = "🗑️ Eliminar estudiantes (seguro)"

    @admin.action(description='🏢 Escoger cliente (asignación masiva)')
    def asignar_cliente_masivo(self, request, queryset):
        """Asigna el mismo Cliente (organización) a todos los estudiantes seleccionados."""
        if 'aplicar' in request.POST:
            cliente_id = (request.POST.get('cliente_id') or '').strip()
            if not cliente_id:
                self.message_user(request, "Seleccione un cliente.", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')

            cliente = Cliente.objects.filter(id=cliente_id, activo=True).first()
            if not cliente:
                self.message_user(request, "Cliente no válido o inactivo.", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')

            ids = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
            if not ids:
                self.message_user(request, "No se recibieron estudiantes seleccionados.", level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')

            qs = Estudiante.objects.filter(id__in=ids)
            n = qs.update(cliente=cliente)
            self.message_user(
                request,
                f"Se asignó el cliente «{cliente.nombre}» a {n} estudiante(s).",
                level=messages.SUCCESS,
            )
            return redirect('admin:core_estudiante_changelist')

        clientes = Cliente.objects.filter(activo=True).order_by('nombre')
        return render(
            request,
            'admin/asignar_cliente_estudiantes.html',
            {
                'estudiantes': queryset,
                'total_estudiantes': queryset.count(),
                'clientes': clientes,
                'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
                'title': 'Escoger cliente — asignación masiva',
            },
        )

    @admin.action(description='📚 Inscribir en curso (masivo)')
    def inscribir_curso_masivo(self, request, queryset):
        """Inscribe estudiantes seleccionados en un curso (sin Excel)."""
        from core.inscripcion_curso import inscribir_estudiante_en_curso

        if 'aplicar' in request.POST:
            curso_id = (request.POST.get('curso_id') or '').strip()
            if not curso_id:
                self.message_user(request, 'Seleccione un curso.', level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')
            curso = Curso.objects.filter(pk=curso_id, activo=True).first()
            if not curso:
                self.message_user(request, 'Curso no válido.', level=messages.ERROR)
                return redirect('admin:core_estudiante_changelist')
            ids = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
            qs = Estudiante.objects.filter(id__in=ids)
            creados = 0
            for est in qs:
                _prog, creado = inscribir_estudiante_en_curso(est, curso)
                if creado:
                    creados += 1
            self.message_user(
                request,
                f'Inscripción en «{curso.nombre}»: {creados} nuevo(s), '
                f'{qs.count() - creados} ya inscrito(s).',
                level=messages.SUCCESS,
            )
            return redirect('admin:core_estudiante_changelist')

        cursos = Curso.objects.filter(activo=True).order_by('nombre')
        return render(
            request,
            'admin/inscribir_curso_masivo.html',
            {
                'estudiantes': queryset,
                'total_estudiantes': queryset.count(),
                'cursos': cursos,
                'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
                'title': 'Inscribir en curso — masivo',
            },
        )

    @admin.action(description='📋 Exportar selección (CSV ligero)')
    def exportar_seleccion_ligera(self, request, queryset):
        """Exporta filas visibles/seleccionadas sin hoja de progreso pesada."""
        import csv
        from django.http import HttpResponse
        from datetime import datetime

        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            f'attachment; filename="estudiantes_{datetime.now():%Y%m%d_%H%M%S}.csv"'
        )
        w = csv.writer(resp)
        w.writerow(['nombre', 'cedula', 'telefono', 'cliente', 'activo'])
        for est in queryset.select_related('cliente'):
            w.writerow([
                est.nombre,
                est.cedula,
                est.telefono,
                est.cliente.nombre if est.cliente_id else '',
                'si' if est.activo else 'no',
            ])
        return resp

    def asignar_a_grupo_accion(self, request, queryset):
        """Asigna múltiples estudiantes a un grupo (existente o nuevo)"""
        if request.POST.get('aplicar') or 'apply' in request.POST:
            # Verificar si es grupo existente o nuevo
            grupo_tipo = request.POST.get('grupo_tipo', 'existente')
            
            if grupo_tipo == 'nuevo':
                # Crear nuevo grupo
                nuevo_nombre = request.POST.get('nuevo_nombre', '').strip()
                nuevo_emoji = request.POST.get('nuevo_emoji', '👥')
                nuevo_descripcion = request.POST.get('nuevo_descripcion', '').strip()
                
                if not nuevo_nombre:
                    self.message_user(request, "⚠️ Debes ingresar un nombre para el grupo", level=messages.WARNING)
                    return
                
                # Obtener el cliente del primer estudiante o None
                primer_estudiante = queryset.first()
                cliente = primer_estudiante.cliente if primer_estudiante else None
                
                # Crear el grupo
                grupo = GrupoEstudiantes.objects.create(
                    nombre=nuevo_nombre,
                    emoji=nuevo_emoji,
                    descripcion=nuevo_descripcion,
                    cliente=cliente,
                    creado_por=request.user,
                    activo=True
                )
                
                # Asignar estudiantes al nuevo grupo
                contador = 0
                for estudiante in queryset:
                    grupo.estudiantes.add(estudiante)
                    contador += 1
                
                self.message_user(
                    request,
                    f"✅ Grupo '{grupo.emoji} {grupo.nombre}' creado y {contador} estudiante(s) asignado(s)",
                    level=messages.SUCCESS
                )
                return redirect('admin:core_estudiante_changelist')
            
            else:
                # Usar grupo existente
                grupo_id = request.POST.get('grupo')
                if not grupo_id:
                    self.message_user(request, "⚠️ Debes seleccionar un grupo", level=messages.WARNING)
                    return
                
                try:
                    grupo = GrupoEstudiantes.objects.get(id=grupo_id)
                    contador = 0
                    for estudiante in queryset:
                        grupo.estudiantes.add(estudiante)
                        contador += 1
                    
                    self.message_user(
                        request,
                        f"✅ {contador} estudiante(s) asignado(s) al grupo '{grupo.nombre}'",
                        level=messages.SUCCESS
                    )
                    return redirect('admin:core_estudiante_changelist')
                except GrupoEstudiantes.DoesNotExist:
                    self.message_user(request, "❌ El grupo seleccionado no existe", level=messages.ERROR)
                    return redirect('admin:core_estudiante_changelist')
        
        # Mostrar formulario de selección de grupo
        grupos = GrupoEstudiantes.objects.all().order_by('nombre')
        context = {
            'title': f'Asignar {queryset.count()} estudiantes a un grupo',
            'queryset': queryset,
            'grupos': grupos,
            'opts': self.model._meta,
            'action_checkbox_name': helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, 'admin/asignar_grupo.html', context)
    asignar_a_grupo_accion.short_description = "👥 Asignar estudiantes a un grupo"


# @admin.register(Plantilla)  # Registro duplicado - movido arriba
class PlantillaAdminDuplicado(admin.ModelAdmin):
    """Gestión de plantillas de mensajes con Twilio Content Templates"""
    list_display = ('nombre_interno', 'categoria_emoji', 'temas_badge', 'twilio_status_badge', 'vista_previa', 'activa', 'veces_usada', 'fecha_modificacion')
    list_filter = ('categoria', 'temas', 'activa', 'aprobada_twilio', 'fecha_creacion')
    search_fields = ('nombre_interno', 'cuerpo_mensaje', 'twilio_template_sid', 'twilio_template_nombre')
    actions = ['enviar_plantilla_directa', 'duplicar_plantilla', 'activar_plantillas', 'desactivar_plantillas']
    readonly_fields = ('veces_usada', 'fecha_creacion', 'fecha_modificacion', 'preview_personalizado')
    filter_horizontal = ('temas',)
    
    fieldsets = (
        ('Información de la Plantilla', {
            'fields': ('nombre_interno', 'categoria', 'temas', 'activa'),
            'description': 'Dale un nombre descriptivo a tu plantilla, categoría y temas relacionados'
        }),
        ('Contenido del Mensaje', {
            'fields': ('cuerpo_mensaje',),
            'description': mark_safe('<strong>Variables disponibles:</strong> {nombre} {telefono} {curso}<br>'
                          '<strong>Ejemplo:</strong> "Hola {nombre}, te damos la bienvenida al curso {curso}"')
        }),
        ('Configuración Twilio Content Templates', {
            'fields': ('twilio_template_sid', 'twilio_template_nombre', 'aprobada_twilio'),
            'description': mark_safe('<div style="background:#e3f2fd;padding:15px;border-radius:4px;margin-bottom:15px;">'
                          '<strong>INSTRUCCIONES PARA USAR TWILIO:</strong><br><br>'
                          '1. Ve a <a href="https://console.twilio.com/us1/develop/sms/content-editor" target="_blank">Twilio Content Editor</a><br>'
                          '2. Crea una nueva plantilla con este mensaje<br>'
                          '3. Espera la aprobación de Twilio (1-2 días hábiles)<br>'
                          '4. Copia el <strong>Content SID</strong> (ej: HX1234...)<br>'
                          '5. Pégalo en el campo "Twilio Content SID" arriba<br>'
                          '6. Marca "Aprobada en Twilio" cuando esté lista<br><br>'
                          '<strong>Una vez aprobada, podrás usarla en campañas</strong></div>')
        }),
        ('Estadísticas y Vista Previa', {
            'fields': ('preview_personalizado', 'veces_usada', 'fecha_creacion', 'fecha_modificacion'),
            'classes': ('collapse',),
        }),
    )
    
    def temas_badge(self, obj):
        """Muestra temas asociados con badges"""
        temas = obj.temas.all()
        if not temas:
            return format_html('<span style="color:#999;">Sin temas</span>')
        
        badges = []
        for tema in temas:
            badges.append(f'<span style="background:#e3f2fd;color:#1976d2;padding:3px 10px;border-radius:12px;margin:2px;display:inline-block;">{tema}</span>')
        return format_html(''.join(badges))
    temas_badge.short_description = "Temas"
    
    def twilio_status_badge(self, obj):
        """Muestra estado de aprobación en Twilio"""
        if not obj.twilio_template_sid:
            return format_html('<span style="background:#f5f5f5;color:#666;padding:4px 10px;border-radius:12px;font-size:11px;">📝 Sin SID</span>')
        
        if obj.aprobada_twilio:
            return format_html(
                '<span style="background:#e8f5e9;color:#2e7d32;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:500;">✅ Aprobada</span>'
            )
        else:
            return format_html(
                '<span style="background:#fff3e0;color:#f57c00;padding:4px 10px;border-radius:12px;font-size:11px;font-weight:500;">⏳ Pendiente</span>'
            )
    twilio_status_badge.short_description = "🔵 Estado Twilio"
    
    def categoria_emoji(self, obj):
        """Muestra categoría con emoji"""
        return obj.get_categoria_display()
    categoria_emoji.short_description = "📂 Categoría"
    
    def vista_previa(self, obj):
        """Muestra preview del mensaje"""
        preview = obj.vista_previa()
        return format_html('<span style="color:#666;font-style:italic;">{}</span>', preview)
    vista_previa.short_description = "📄 Vista Previa"
    
    def preview_personalizado(self, obj):
        """Muestra cómo se vería el mensaje personalizado"""
        ejemplo = obj.cuerpo_mensaje.replace('{nombre}', 'Juan Pérez')
        ejemplo = ejemplo.replace('{telefono}', '+573001234567')
        ejemplo = ejemplo.replace('{curso}', 'Cultivo de Aguacate Hass')
        return format_html(
            '<div style="background:#f5f5f5;padding:15px;border-left:4px solid #4CAF50;border-radius:4px;">'
            '<strong>📱 Vista Previa Personalizada:</strong><br><br>{}</div>',
            ejemplo
        )
    preview_personalizado.short_description = "Vista Previa con Variables"
    
    @admin.action(description='📄 Duplicar plantilla(s) seleccionada(s)')
    def duplicar_plantilla(self, request, queryset):
        """Duplica plantillas seleccionadas"""
        duplicadas = 0
        for plantilla in queryset:
            plantilla.pk = None
            plantilla.nombre_interno = f"{plantilla.nombre_interno} (Copia)"
            plantilla.veces_usada = 0
            plantilla.save()
            duplicadas += 1
        self.message_user(request, f"✅ {duplicadas} plantilla(s) duplicada(s)", level=messages.SUCCESS)
    
    @admin.action(description='✅ Activar plantilla(s) seleccionada(s)')
    def activar_plantillas(self, request, queryset):
        """Activa plantillas seleccionadas"""
        actualizadas = queryset.update(activa=True)
        self.message_user(request, f"✅ {actualizadas} plantilla(s) activada(s)", level=messages.SUCCESS)
    
    @admin.action(description='❌ Desactivar plantilla(s) seleccionada(s)')
    def desactivar_plantillas(self, request, queryset):
        """Desactiva plantillas seleccionadas"""
        actualizadas = queryset.update(activa=False)
        self.message_user(request, f"⚠️ {actualizadas} plantilla(s) desactivada(s)", level=messages.WARNING)
    
    @admin.action(description='📱 Enviar a Meta para aprobación')
    def enviar_a_meta_accion(self, request, queryset):
        """Envía plantillas a Meta WhatsApp Business para aprobación"""
        from core.meta_templates import enviar_plantilla_a_meta
        
        enviadas = 0
        errores = 0
        
        for plantilla in queryset:
            try:
                # Verificar si ya fue enviada
                if plantilla.enviada_a_meta and plantilla.meta_template_status == 'APPROVED':
                    self.message_user(
                        request,
                        f"⚠️ '{plantilla.nombre_interno}' ya está aprobada en Meta",
                        level=messages.WARNING
                    )
                    continue
                
                # Enviar a Meta
                resultado = enviar_plantilla_a_meta(
                    nombre_plantilla=plantilla.nombre_interno,
                    contenido=plantilla.cuerpo_mensaje,
                    categoria=plantilla.categoria if plantilla.categoria in ['MARKETING', 'UTILITY'] else 'MARKETING',
                    idioma='es'
                )
                
                if resultado['success']:
                    # Actualizar plantilla con información de Meta
                    plantilla.enviada_a_meta = True
                    plantilla.meta_template_id = resultado['template_id']
                    plantilla.meta_template_status = resultado['status']
                    plantilla.meta_template_name = resultado['nombre_meta']
                    plantilla.save()
                    
                    enviadas += 1
                    self.message_user(
                        request,
                        f"✅ '{plantilla.nombre_interno}' enviada a Meta. ID: {resultado['template_id']}",
                        level=messages.SUCCESS
                    )
                else:
                    errores += 1
                    self.message_user(
                        request,
                        f"❌ Error con '{plantilla.nombre_interno}': {resultado['message']}",
                        level=messages.ERROR
                    )
            
            except Exception as e:
                errores += 1
                logger.error(f"Error enviando plantilla a Meta: {str(e)}")
                self.message_user(
                    request,
                    f"❌ Excepción con '{plantilla.nombre_interno}': {str(e)}",
                    level=messages.ERROR
                )
        
        # Resumen final
        if enviadas > 0:
            self.message_user(
                request,
                f"🎉 {enviadas} plantilla(s) enviada(s) a Meta para revisión",
                level=messages.SUCCESS
            )
        if errores > 0:
            self.message_user(
                request,
                f"⚠️ {errores} plantilla(s) con errores. Verifica las credenciales de Meta.",
                level=messages.WARNING
            )
    
    @admin.action(description='📤 Enviar plantilla a estudiantes')
    def enviar_plantilla_directa(self, request, queryset):
        """Permite enviar una plantilla directamente a estudiantes seleccionados"""
        
        if queryset.count() > 1:
            self.message_user(request, "⚠️ Solo puedes enviar una plantilla a la vez", level=messages.WARNING)
            return
        
        plantilla = queryset.first()
        
        # Si es POST con confirmación
        if 'aplicar' in request.POST:
            # Obtener estudiantes seleccionados
            estudiantes_ids = request.POST.getlist('estudiantes_seleccionados')
            if not estudiantes_ids:
                self.message_user(request, "⚠️ Debes seleccionar al menos un estudiante", level=messages.ERROR)
            else:
                enviados = 0
                errores = 0
                
                for est_id in estudiantes_ids:
                    try:
                        estudiante = Estudiante.objects.get(id=est_id)
                        
                        # Personalizar mensaje con nombre
                        mensaje = plantilla.cuerpo_mensaje.replace('{nombre}', estudiante.nombre)
                        mensaje = mensaje.replace('{estudiante}', estudiante.nombre)
                        
                        telefono = estudiante.telefono
                        if not telefono.startswith('whatsapp:'):
                            telefono = f'whatsapp:{telefono}'
                        
                        resultado = enviar_whatsapp_twilio(
                            telefono=telefono,
                            texto=mensaje,
                            mensaje_id_referencia=None
                        )
                        
                        if resultado.get('success'):
                            enviados += 1
                            # Registrar en WhatsappLog
                            WhatsappLog.objects.create(
                                telefono=estudiante.telefono,
                                mensaje=mensaje,
                                mensaje_id=resultado.get('mensaje_id'),
                                tipo='SENT',
                                estado='SENT'
                            )
                            logger.info(f"✅ Plantilla '{plantilla.nombre_interno}' enviada a {estudiante.nombre}")
                        else:
                            errores += 1
                            logger.error(f"❌ Error al enviar plantilla a {estudiante.telefono}")
                    
                    except Exception as e:
                        errores += 1
                        logger.error(f"❌ Excepción al enviar: {str(e)}")
                
                if enviados > 0:
                    self.message_user(request, f"✅ Plantilla enviada a {enviados} estudiante(s)", level=messages.SUCCESS)
                if errores > 0:
                    self.message_user(request, f"⚠️ Hubo {errores} error(es)", level=messages.WARNING)
                
                # Redirigir a la lista de plantillas después de enviar
                from django.urls import reverse
                return redirect(reverse('admin:core_plantilla_changelist'))
        
        # Mostrar formulario de selección de estudiantes
        estudiantes = Estudiante.objects.filter(activo=True).order_by('nombre')
        
        return render(request, 'admin/enviar_plantilla_directa.html', {
            'plantilla': plantilla,
            'estudiantes': estudiantes,
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        })


class EnvioProgramadoForm(forms.ModelForm):
    media_file_upload = forms.FileField(
        label='Subir archivo desde PC',
        required=False,
        help_text='Opcional. Guarda el archivo en storage/S3 y completa URL del Archivo.',
    )

    class Meta:
        model = EnvioProgramado
        exclude = (
            'fecha_envio_real',
            'total_destinatarios',
            'total_enviados',
            'total_fallidos',
            'error',
            'fecha_creacion',
            'creado_por',
        )

    def save(self, commit=True):
        instance = super().save(commit=False)
        uploaded_file = self.cleaned_data.get('media_file_upload')
        if uploaded_file:
            instance.media_url = guardar_upload_admin_media(
                uploaded_file,
                carpeta='envios_programados',
                prefix='media',
            )
            instance.incluir_media = True
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class EnvioProgramadoInline(admin.TabularInline):
    """Envíos programados dentro de una campaña"""
    model = EnvioProgramado
    form = EnvioProgramadoForm
    extra = 0
    fields = (
        'nombre', 'tipo', 'fecha_programada', 'estado', 'mensaje',
        'incluir_media', 'media_url', 'media_file_upload', 'fecha_envio_real',
    )
    readonly_fields = ('fecha_envio_real',)
    show_change_link = True


