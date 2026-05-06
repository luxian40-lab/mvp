"""
API REST para datos de progreso y tareas del estudiante.
Endpoints:
- GET /api/estudiante/{telefono}/ → datos del estudiante
- GET /api/estudiante/{telefono}/progreso/ → progreso detallado
- GET /api/estudiante/{telefono}/siguiente-tarea/ → siguiente tarea
"""
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.conf import settings
from django.db.models import Q
from django.db.models import Count
from django.db.models.functions import TruncDate
from datetime import datetime, timedelta
import json
import math
import hmac
import logging
from .models import (
    Estudiante,
    Campana,
    EnvioLog,
    AliadoEmpleabilidad,
    MisionEmpleabilidad,
    Curso,
    ProgresoEstudiante,
    ModuloCompletado,
    WhatsappLog,
)
from .services_estudiante import (
    get_estudiante_payload,
    get_estudiante_progreso_payload,
    get_estudiante_siguiente_tarea_payload,
)

logger = logging.getLogger(__name__)


def _haversine_metros(lat1, lon1, lat2, lon2):
    radio_tierra = 6371000.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radio_tierra * c


def _score_oportunidad(distancia_metros, aliado):
    """Prioriza oportunidades con distancia, prioridad del aliado y cupos."""
    distancia = float(distancia_metros if distancia_metros is not None else 999999)
    prioridad = int(getattr(aliado, 'prioridad', 3) or 3)
    cupos = int(getattr(aliado, 'cupos_disponibles', 0) or 0)

    score_dist = max(0.0, 1000.0 - distancia)
    score_prioridad = prioridad * 120.0
    score_cupos = min(cupos, 50) * 5.0
    return round(score_dist + score_prioridad + score_cupos, 2)


@csrf_exempt
@require_http_methods(["GET"])
def api_estudiante(request, telefono):
    """
    GET /api/estudiante/{telefono}/
    Devuelve información del estudiante.
    """
    try:
        logger.info(
            "api_estudiante_requested",
            extra={"request_id": getattr(request, "request_id", ""), "telefono": telefono},
        )
        return JsonResponse({
            'success': True,
            'estudiante': get_estudiante_payload(telefono),
        })
    except Estudiante.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Estudiante con teléfono {telefono} no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_estudiante_progreso(request, telefono):
    """
    GET /api/estudiante/{telefono}/progreso/
    Devuelve progreso del estudiante en cursos/módulos.
    """
    try:
        logger.info(
            "api_estudiante_progreso_requested",
            extra={"request_id": getattr(request, "request_id", ""), "telefono": telefono},
        )
        payload = get_estudiante_progreso_payload(telefono)
        return JsonResponse({
            'success': True,
            'estudiante': payload['estudiante'],
            'progreso': payload['progreso'],
        })
    except Estudiante.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Estudiante con teléfono {telefono} no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_estudiante_siguiente_tarea(request, telefono):
    """
    GET /api/estudiante/{telefono}/siguiente-tarea/
    Devuelve la siguiente tarea del estudiante.
    """
    try:
        logger.info(
            "api_estudiante_siguiente_tarea_requested",
            extra={"request_id": getattr(request, "request_id", ""), "telefono": telefono},
        )
        payload = get_estudiante_siguiente_tarea_payload(telefono)
        return JsonResponse({
            'success': True,
            'estudiante': payload['estudiante'],
            'siguiente_tarea': payload['siguiente_tarea'],
        })
    except Estudiante.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Estudiante con teléfono {telefono} no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_empleabilidad_oportunidades(request):
    """GET /api/empleabilidad/oportunidades/?telefono=...&latitud=...&longitud=..."""
    telefono = (request.GET.get('telefono', '') or '').strip()
    latitud_raw = request.GET.get('latitud')
    longitud_raw = request.GET.get('longitud')

    if not telefono:
        return JsonResponse({'success': False, 'error': 'telefono es obligatorio'}, status=400)
    if latitud_raw is None or longitud_raw is None:
        return JsonResponse({'success': False, 'error': 'latitud y longitud son obligatorios'}, status=400)

    try:
        estudiante = Estudiante.objects.select_related('cliente').get(telefono=telefono)
        latitud = float(latitud_raw)
        longitud = float(longitud_raw)
    except Estudiante.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Estudiante no encontrado'}, status=404)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Coordenadas inválidas'}, status=400)

    cliente = estudiante.cliente
    if cliente and not (getattr(cliente, 'habilitar_gamificacion_proximidad', False) or getattr(cliente, 'empleabilidad_exploracion_activa', False)):
        return JsonResponse({'success': True, 'activa': False, 'oportunidades': []})

    radio = int(getattr(cliente, 'empleabilidad_radio_metros', 800) if cliente else 800)
    hoy = timezone.localdate()
    aliados_qs = AliadoEmpleabilidad.objects.filter(vacantes_activas=True).filter(
        Q(vigencia_desde__isnull=True) | Q(vigencia_desde__lte=hoy)
    ).filter(
        Q(vigencia_hasta__isnull=True) | Q(vigencia_hasta__gte=hoy)
    )
    if cliente:
        aliados_qs = aliados_qs.filter(Q(cliente__isnull=True) | Q(cliente=cliente))

    oportunidades = []
    for aliado in aliados_qs:
        dist = _haversine_metros(latitud, longitud, aliado.latitud, aliado.longitud)
        if dist <= radio:
            score = _score_oportunidad(dist, aliado)
            oportunidades.append({
                'aliado_id': aliado.id,
                'nombre_empresa': aliado.nombre_empresa,
                'distancia_metros': round(dist, 1),
                'score_prioridad': score,
                'cupos_disponibles': int(getattr(aliado, 'cupos_disponibles', 0) or 0),
                'indicacion_sector': aliado.indicacion_sector or '',
            })

    oportunidades.sort(key=lambda x: (-x['score_prioridad'], x['distancia_metros']))
    return JsonResponse({
        'success': True,
        'activa': True,
        'radio_metros': radio,
        'oportunidades': oportunidades[:10],
    })


@csrf_exempt
@require_http_methods(["POST"])
def api_empleabilidad_claim(request):
    """POST /api/empleabilidad/claim/ con {telefono, aliado_id, latitud, longitud}."""
    try:
        body = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        body = request.POST

    telefono = (body.get('telefono', '') or '').strip()
    aliado_id = body.get('aliado_id')
    latitud = body.get('latitud')
    longitud = body.get('longitud')

    if not telefono or not aliado_id:
        return JsonResponse({'success': False, 'error': 'telefono y aliado_id son obligatorios'}, status=400)

    try:
        estudiante = Estudiante.objects.select_related('cliente').get(telefono=telefono)
        aliado = AliadoEmpleabilidad.objects.get(id=int(aliado_id), vacantes_activas=True)
        lat = float(latitud) if latitud is not None else None
        lon = float(longitud) if longitud is not None else None
    except Estudiante.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Estudiante no encontrado'}, status=404)
    except AliadoEmpleabilidad.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Aliado no encontrado o inactivo'}, status=404)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Datos inválidos'}, status=400)

    distancia = None
    if lat is not None and lon is not None:
        distancia = _haversine_metros(lat, lon, aliado.latitud, aliado.longitud)
    score_prioridad = _score_oportunidad(distancia, aliado)

    mision = MisionEmpleabilidad.objects.create(
        cliente=estudiante.cliente,
        estudiante=estudiante,
        aliado=aliado,
        estado='reclamada',
        estado_flujo='interesado',
        latitud=lat,
        longitud=lon,
        distancia_metros=round(distancia, 1) if distancia is not None else None,
        puntaje_prioridad=score_prioridad,
        fecha_reclamada=timezone.now(),
        fecha_interes=timezone.now(),
        metadata={'fuente': 'api_claim'},
    )

    return JsonResponse({
        'success': True,
        'mision_id': mision.id,
        'estado': mision.estado,
        'estado_flujo': mision.estado_flujo,
        'score_prioridad': mision.puntaje_prioridad,
        'codigo_hint': 'Solicita el código secreto al aliado para completar.',
    })


@csrf_exempt
@require_http_methods(["POST"])
def api_empleabilidad_completar(request):
    """POST /api/empleabilidad/completar/ con {telefono, mision_id, codigo}."""
    try:
        body = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        body = request.POST

    telefono = (body.get('telefono', '') or '').strip()
    mision_id = body.get('mision_id')
    codigo = (body.get('codigo', '') or '').strip().lower()

    if not telefono or not mision_id or not codigo:
        return JsonResponse({'success': False, 'error': 'telefono, mision_id y codigo son obligatorios'}, status=400)

    try:
        estudiante = Estudiante.objects.select_related('cliente').get(telefono=telefono)
        mision = MisionEmpleabilidad.objects.select_related('aliado').get(id=int(mision_id), estudiante=estudiante)
    except Estudiante.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Estudiante no encontrado'}, status=404)
    except MisionEmpleabilidad.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Misión no encontrada'}, status=404)

    if codigo != str(mision.aliado.codigo_secreto).strip().lower():
        return JsonResponse({'success': False, 'error': 'Código inválido'}, status=400)

    puntos = int(getattr(estudiante.cliente, 'empleabilidad_puntos_validacion', 30) if estudiante.cliente else 30)
    if mision.estado != 'completada':
        mision.estado = 'completada'
        mision.codigo_validado = True
        mision.puntos_otorgados = puntos
        mision.estado_flujo = 'postulado'
        mision.fecha_completada = timezone.now()
        mision.fecha_postulacion = timezone.now()
        mision.save(update_fields=['estado', 'codigo_validado', 'puntos_otorgados', 'estado_flujo', 'fecha_completada', 'fecha_postulacion'])

    return JsonResponse({
        'success': True,
        'mision_id': mision.id,
        'estado': mision.estado,
        'estado_flujo': mision.estado_flujo,
        'puntos': mision.puntos_otorgados,
    })


@csrf_exempt
@require_http_methods(["POST"])
def api_empleabilidad_flujo(request):
    """POST /api/empleabilidad/flujo/ con {mision_id, estado_flujo}."""
    try:
        body = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        body = request.POST

    mision_id = body.get('mision_id')
    estado_flujo = (body.get('estado_flujo', '') or '').strip().lower()
    estados_validos = {'descubierto', 'interesado', 'postulado', 'entrevista', 'vinculado', 'descartado'}

    if not mision_id or estado_flujo not in estados_validos:
        return JsonResponse({'success': False, 'error': 'mision_id y estado_flujo válido son obligatorios'}, status=400)

    try:
        mision = MisionEmpleabilidad.objects.get(id=int(mision_id))
    except MisionEmpleabilidad.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Misión no encontrada'}, status=404)

    mision.estado_flujo = estado_flujo
    ahora = timezone.now()
    update_fields = ['estado_flujo']
    if estado_flujo == 'interesado' and not mision.fecha_interes:
        mision.fecha_interes = ahora
        update_fields.append('fecha_interes')
    elif estado_flujo == 'postulado' and not mision.fecha_postulacion:
        mision.fecha_postulacion = ahora
        update_fields.append('fecha_postulacion')
    elif estado_flujo == 'entrevista' and not mision.fecha_entrevista:
        mision.fecha_entrevista = ahora
        update_fields.append('fecha_entrevista')
    elif estado_flujo == 'vinculado' and not mision.fecha_vinculacion:
        mision.fecha_vinculacion = ahora
        update_fields.append('fecha_vinculacion')

    mision.save(update_fields=update_fields)
    return JsonResponse({'success': True, 'mision_id': mision.id, 'estado_flujo': mision.estado_flujo})


@csrf_exempt
@require_http_methods(["GET"])
def api_empleabilidad_resumen(request):
    """GET /api/empleabilidad/resumen/?cliente_id=..."""
    cliente_id = request.GET.get('cliente_id')
    qs = MisionEmpleabilidad.objects.all()
    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)

    total = qs.count()
    por_estado = {k: v for k, v in qs.values_list('estado_flujo').annotate(c=Count('id'))}
    desc = por_estado.get('descubierto', 0)
    ints = por_estado.get('interesado', 0)
    post = por_estado.get('postulado', 0)
    entr = por_estado.get('entrevista', 0)
    vinc = por_estado.get('vinculado', 0)

    def tasa(a, b):
        return round((a / b * 100), 2) if b else 0

    return JsonResponse({
        'success': True,
        'total_misiones': total,
        'embudo': {
            'descubierto': desc,
            'interesado': ints,
            'postulado': post,
            'entrevista': entr,
            'vinculado': vinc,
        },
        'kpis': {
            'tasa_interes_vs_descubierto': tasa(ints, desc),
            'tasa_postulacion_vs_interes': tasa(post, ints),
            'tasa_entrevista_vs_postulado': tasa(entr, post),
            'tasa_vinculacion_vs_postulado': tasa(vinc, post),
            'tasa_vinculacion_vs_descubierto': tasa(vinc, desc),
        },
    })


def _integracion_apply_cors(request, resp):
    allowed_raw = str(getattr(settings, 'INTEGRACION_API_ALLOWED_ORIGINS', '*') or '*').strip()
    origin = (request.headers.get('Origin', '') or '').strip()

    allow_origin = '*'
    if allowed_raw != '*':
        allowed = [o.strip() for o in allowed_raw.split(',') if o.strip()]
        if origin and origin in allowed:
            allow_origin = origin
        elif allowed:
            allow_origin = allowed[0]

    resp['Access-Control-Allow-Origin'] = allow_origin
    resp['Vary'] = 'Origin'
    resp['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    resp['Access-Control-Allow-Headers'] = 'Authorization, Content-Type, X-API-Key'
    resp['Access-Control-Max-Age'] = '86400'
    return resp


def _integracion_auth_error(request):
    expected_key = str(getattr(settings, 'INTEGRACION_API_KEY', '') or '').strip()
    if not expected_key:
        return None

    auth_header = (request.headers.get('Authorization', '') or '').strip()
    provided = ''
    if auth_header.lower().startswith('bearer '):
        provided = auth_header.split(' ', 1)[1].strip()
    if not provided:
        provided = (request.headers.get('X-API-Key', '') or request.GET.get('api_key', '') or '').strip()

    if not provided or not hmac.compare_digest(provided, expected_key):
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=401)
    return None


def _integracion_parse_filtros(request, permitir_curso=False):
    cliente_id_raw = (request.GET.get('cliente_id') or '').strip()
    curso_id_raw = (request.GET.get('curso_id') or '').strip()
    fecha_raw = (request.GET.get('fecha') or '').strip()
    desde_raw = (request.GET.get('desde') or '').strip()
    hasta_raw = (request.GET.get('hasta') or '').strip()

    cliente_id = None
    if cliente_id_raw:
        try:
            cliente_id = int(cliente_id_raw)
        except ValueError:
            return None, JsonResponse({'success': False, 'error': 'cliente_id debe ser numérico'}, status=400)

    curso_id = None
    if permitir_curso and curso_id_raw:
        try:
            curso_id = int(curso_id_raw)
        except ValueError:
            return None, JsonResponse({'success': False, 'error': 'curso_id debe ser numérico'}, status=400)

    def _parse_iso(value: str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except Exception:
            return None

    if desde_raw or hasta_raw:
        fecha_desde = _parse_iso(desde_raw)
        fecha_hasta = _parse_iso(hasta_raw)
        if not fecha_desde or not fecha_hasta:
            return None, JsonResponse({'success': False, 'error': 'desde y hasta deben usar formato YYYY-MM-DD'}, status=400)
    else:
        fecha_unica = _parse_iso(fecha_raw) if fecha_raw else timezone.localdate()
        if not fecha_unica:
            return None, JsonResponse({'success': False, 'error': 'fecha debe usar formato YYYY-MM-DD'}, status=400)
        fecha_desde = fecha_unica
        fecha_hasta = fecha_unica

    if fecha_hasta < fecha_desde:
        return None, JsonResponse({'success': False, 'error': 'hasta no puede ser menor que desde'}, status=400)

    max_dias = int(getattr(settings, 'INTEGRACION_API_MAX_DIAS', 31) or 31)
    rango_dias = (fecha_hasta - fecha_desde).days + 1
    if rango_dias > max_dias:
        return None, JsonResponse({
            'success': False,
            'error': f'Rango máximo permitido: {max_dias} días',
        }, status=400)

    return {
        'cliente_id': cliente_id,
        'curso_id': curso_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'rango_dias': rango_dias,
    }, None


@csrf_exempt
@require_http_methods(["GET", "OPTIONS"])
def api_integracion_educativa_metricas(request):
    """GET /api/integracion/educativa/metricas/?cliente_id=...&curso_id=...&fecha=YYYY-MM-DD"""
    if request.method == 'OPTIONS':
        return _integracion_apply_cors(request, HttpResponse(status=204))

    auth_error = _integracion_auth_error(request)
    if auth_error:
        return _integracion_apply_cors(request, auth_error)

    filtros, parse_error = _integracion_parse_filtros(request, permitir_curso=True)
    if parse_error:
        return _integracion_apply_cors(request, parse_error)

    cliente_id = filtros['cliente_id']
    curso_id = filtros['curso_id']
    fecha_desde = filtros['fecha_desde']
    fecha_hasta = filtros['fecha_hasta']
    rango_dias = filtros['rango_dias']

    estudiantes_base = Estudiante.objects.filter(activo=True)
    if cliente_id is not None:
        estudiantes_base = estudiantes_base.filter(cliente_id=cliente_id)
    if curso_id is not None:
        estudiantes_base = estudiantes_base.filter(progresos__curso_id=curso_id).distinct()

    progreso_base = ProgresoEstudiante.objects.filter(estudiante__in=estudiantes_base)
    if curso_id is not None:
        progreso_base = progreso_base.filter(curso_id=curso_id)

    estudiantes_registrados_q = estudiantes_base.filter(
        fecha_registro__date__gte=fecha_desde,
        fecha_registro__date__lte=fecha_hasta,
    )
    inscripciones_q = progreso_base.filter(
        fecha_inicio__date__gte=fecha_desde,
        fecha_inicio__date__lte=fecha_hasta,
    )
    cursos_completados_q = progreso_base.filter(
        completado=True,
        fecha_completado__date__gte=fecha_desde,
        fecha_completado__date__lte=fecha_hasta,
    )
    modulos_completados_q = ModuloCompletado.objects.filter(
        progreso__in=progreso_base,
        fecha_completado__date__gte=fecha_desde,
        fecha_completado__date__lte=fecha_hasta,
    )

    telefonos_scope = estudiantes_base.exclude(telefono='').values_list('telefono', flat=True)
    mensajes_q = WhatsappLog.objects.filter(
        fecha__date__gte=fecha_desde,
        fecha__date__lte=fecha_hasta,
    )
    if cliente_id is not None or curso_id is not None:
        mensajes_q = mensajes_q.filter(
            Q(estudiante__in=estudiantes_base) | Q(telefono__in=telefonos_scope)
        ).distinct()

    estudiantes_activos_total = estudiantes_base.count()
    estudiantes_registrados_total = estudiantes_registrados_q.count()
    inscripciones_total = inscripciones_q.count()
    cursos_completados_total = cursos_completados_q.count()
    modulos_completados_total = modulos_completados_q.count()
    mensajes_total = mensajes_q.count()
    mensajes_enviados_total = mensajes_q.filter(tipo='SENT').count()
    mensajes_recibidos_total = mensajes_q.filter(tipo='INCOMING').count()

    tasa_completacion = round((cursos_completados_total / inscripciones_total * 100), 2) if inscripciones_total else 0

    cursos_scope = Curso.objects.filter(progresoestudiante__in=progreso_base).distinct()
    if curso_id is not None:
        cursos_scope = cursos_scope.filter(id=curso_id)

    filtro_progreso_base = Q(progresoestudiante__id__in=progreso_base.values('id'))
    filtro_inscritos = (
        filtro_progreso_base
        & Q(progresoestudiante__fecha_inicio__date__gte=fecha_desde)
        & Q(progresoestudiante__fecha_inicio__date__lte=fecha_hasta)
    )
    filtro_completados = (
        filtro_progreso_base
        & Q(progresoestudiante__completado=True)
        & Q(progresoestudiante__fecha_completado__date__gte=fecha_desde)
        & Q(progresoestudiante__fecha_completado__date__lte=fecha_hasta)
    )

    cursos_resumen_raw = list(
        cursos_scope.annotate(
            inscritos=Count('progresoestudiante', filter=filtro_inscritos, distinct=True),
            completados=Count('progresoestudiante', filter=filtro_completados, distinct=True),
        )
        .values('id', 'nombre', 'inscritos', 'completados')
        .order_by('-inscritos', 'nombre')[:20]
    )
    cursos_resumen = []
    for item in cursos_resumen_raw:
        inscritos = int(item.get('inscritos') or 0)
        completados = int(item.get('completados') or 0)
        cursos_resumen.append({
            'curso_id': int(item.get('id')),
            'curso_nombre': item.get('nombre') or '',
            'inscritos': inscritos,
            'completados': completados,
            'tasa_completacion': round((completados / inscritos * 100), 2) if inscritos else 0,
        })

    registros_por_dia = {
        str(row['metric_day']): int(row['c'])
        for row in estudiantes_registrados_q.annotate(metric_day=TruncDate('fecha_registro'))
        .values('metric_day')
        .annotate(c=Count('id'))
    }
    inscripciones_por_dia = {
        str(row['metric_day']): int(row['c'])
        for row in inscripciones_q.annotate(metric_day=TruncDate('fecha_inicio'))
        .values('metric_day')
        .annotate(c=Count('id'))
    }
    completados_por_dia = {
        str(row['metric_day']): int(row['c'])
        for row in cursos_completados_q.annotate(metric_day=TruncDate('fecha_completado'))
        .values('metric_day')
        .annotate(c=Count('id'))
    }
    modulos_por_dia = {
        str(row['metric_day']): int(row['c'])
        for row in modulos_completados_q.annotate(metric_day=TruncDate('fecha_completado'))
        .values('metric_day')
        .annotate(c=Count('id'))
    }
    mensajes_por_dia = {
        str(row['metric_day']): int(row['c'])
        for row in mensajes_q.annotate(metric_day=TruncDate('fecha'))
        .values('metric_day')
        .annotate(c=Count('id'))
    }
    enviados_por_dia = {
        str(row['metric_day']): int(row['c'])
        for row in mensajes_q.filter(tipo='SENT').annotate(metric_day=TruncDate('fecha'))
        .values('metric_day')
        .annotate(c=Count('id'))
    }
    recibidos_por_dia = {
        str(row['metric_day']): int(row['c'])
        for row in mensajes_q.filter(tipo='INCOMING').annotate(metric_day=TruncDate('fecha'))
        .values('metric_day')
        .annotate(c=Count('id'))
    }

    metrics = []
    day_cursor = fecha_desde
    while day_cursor <= fecha_hasta:
        metric_date = day_cursor.isoformat()
        metrics.append({
            'schema_version': 2,
            'metric_date': metric_date,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'metric_name': 'estudiantes_registrados',
            'metric_value': int(registros_por_dia.get(metric_date, 0)),
        })
        metrics.append({
            'schema_version': 2,
            'metric_date': metric_date,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'metric_name': 'inscripciones_curso',
            'metric_value': int(inscripciones_por_dia.get(metric_date, 0)),
        })
        metrics.append({
            'schema_version': 2,
            'metric_date': metric_date,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'metric_name': 'cursos_completados',
            'metric_value': int(completados_por_dia.get(metric_date, 0)),
        })
        metrics.append({
            'schema_version': 2,
            'metric_date': metric_date,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'metric_name': 'modulos_completados',
            'metric_value': int(modulos_por_dia.get(metric_date, 0)),
        })
        metrics.append({
            'schema_version': 2,
            'metric_date': metric_date,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'metric_name': 'mensajes_whatsapp_total',
            'metric_value': int(mensajes_por_dia.get(metric_date, 0)),
        })
        metrics.append({
            'schema_version': 2,
            'metric_date': metric_date,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'metric_name': 'mensajes_whatsapp_sent',
            'metric_value': int(enviados_por_dia.get(metric_date, 0)),
        })
        metrics.append({
            'schema_version': 2,
            'metric_date': metric_date,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'metric_name': 'mensajes_whatsapp_incoming',
            'metric_value': int(recibidos_por_dia.get(metric_date, 0)),
        })
        day_cursor += timedelta(days=1)

    formularios_gei = _integracion_resumen_gei(
        cliente_id=cliente_id,
        curso_id=curso_id,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )

    response = JsonResponse({
        'success': True,
        'meta': {
            'schema_version': 2,
            'tenant_id': cliente_id,
            'course_id': curso_id,
            'desde': fecha_desde.isoformat(),
            'hasta': fecha_hasta.isoformat(),
            'dias': rango_dias,
        },
        'resumen': {
            'estudiantes_activos_total': int(estudiantes_activos_total),
            'estudiantes_registrados_total': int(estudiantes_registrados_total),
            'inscripciones_total': int(inscripciones_total),
            'cursos_completados_total': int(cursos_completados_total),
            'modulos_completados_total': int(modulos_completados_total),
            'tasa_completacion_cursos': tasa_completacion,
            'mensajes_total': int(mensajes_total),
            'mensajes_enviados_total': int(mensajes_enviados_total),
            'mensajes_recibidos_total': int(mensajes_recibidos_total),
        },
        'cursos': cursos_resumen,
        'formularios_gei': formularios_gei,
        'metrics': metrics,
    })
    return _integracion_apply_cors(request, response)


def _integracion_resumen_gei(cliente_id=None, curso_id=None, fecha_desde=None, fecha_hasta=None):
    """Resumen agregado de FichaGEI para el endpoint LXP.

    Si la app `formulario` no está disponible o no hay tablas, devuelve un
    dict vacío con disponible=False (no rompe el endpoint principal).
    """
    try:
        from formulario.models import CAMPOS_GEI_7, FichaGEI, SesionFormulario
    except Exception:
        return {'disponible': False}

    try:
        f_q = FichaGEI.objects.all()
        s_q = SesionFormulario.objects.all()
        if cliente_id is not None:
            f_q = f_q.filter(cliente_id=cliente_id)
            s_q = s_q.filter(estudiante__cliente_id=cliente_id)
        if curso_id is not None:
            f_q = f_q.filter(curso_id=curso_id)
            s_q = s_q.filter(formulario__curso_id=curso_id)

        if fecha_desde is not None:
            f_q_periodo = f_q.filter(fecha_inicio__date__gte=fecha_desde)
        else:
            f_q_periodo = f_q
        if fecha_hasta is not None:
            f_q_periodo = f_q_periodo.filter(fecha_inicio__date__lte=fecha_hasta)

        fichas_totales = f_q.count()
        muestra = list(f_q.order_by('-fecha_update')[:500])

        fichas_completas = 0
        fichas_parciales = 0
        fichas_pendientes = 0
        suma_pct = 0
        nulos_por_campo = {c: 0 for c in CAMPOS_GEI_7}
        for ficha in muestra:
            pct = int(ficha.completitud_pct or 0)
            suma_pct += pct
            if pct == 100:
                fichas_completas += 1
            elif pct == 0:
                fichas_pendientes += 1
            else:
                fichas_parciales += 1
            for c in CAMPOS_GEI_7:
                v = getattr(ficha, c, None)
                if v is None or v == '':
                    nulos_por_campo[c] += 1

        muestra_n = len(muestra)
        completitud_promedio_pct = round(suma_pct / muestra_n, 1) if muestra_n else 0.0

        campo_menor = None
        if muestra_n:
            campo_menor_item = max(nulos_por_campo.items(), key=lambda kv: kv[1])
            if campo_menor_item[1] > 0:
                campo_menor = {
                    'campo': campo_menor_item[0],
                    'fichas_sin_dato': campo_menor_item[1],
                    'porcentaje_sin_dato': round(campo_menor_item[1] * 100 / muestra_n, 1),
                }

        sesiones_activas = s_q.filter(completado=False).count()
        sesiones_completadas_periodo = s_q.filter(completado=True)
        if fecha_desde is not None:
            sesiones_completadas_periodo = sesiones_completadas_periodo.filter(
                fecha_update__date__gte=fecha_desde
            )
        if fecha_hasta is not None:
            sesiones_completadas_periodo = sesiones_completadas_periodo.filter(
                fecha_update__date__lte=fecha_hasta
            )

        tiempo_promedio_min = None
        completadas_muestra = list(sesiones_completadas_periodo[:500])
        if completadas_muestra:
            total_seg = 0
            n = 0
            for s in completadas_muestra:
                if s.fecha_inicio and s.fecha_update:
                    total_seg += (s.fecha_update - s.fecha_inicio).total_seconds()
                    n += 1
            if n:
                tiempo_promedio_min = round((total_seg / n) / 60, 1)

        return {
            'disponible': True,
            'fichas_totales': int(fichas_totales),
            'fichas_periodo': int(f_q_periodo.count()),
            'fichas_completas': int(fichas_completas),
            'fichas_parciales': int(fichas_parciales),
            'fichas_pendientes': int(fichas_pendientes),
            'completitud_promedio_pct': float(completitud_promedio_pct),
            'completitud_muestra': int(muestra_n),
            'campo_con_menor_completitud': campo_menor,
            'sesiones_activas': int(sesiones_activas),
            'sesiones_completadas_periodo': int(sesiones_completadas_periodo.count()),
            'tiempo_promedio_completar_min': tiempo_promedio_min,
        }
    except Exception as exc:
        logger.warning("api_integracion_resumen_gei_error", extra={'error': str(exc)})
        return {'disponible': False, 'error': 'no se pudo calcular el resumen GEI'}


@csrf_exempt
@require_http_methods(["GET", "OPTIONS"])
def api_integracion_gei_detalle(request):
    """GET /api/integracion/gei/detalle/?cliente_id=...&curso_id=...&desde=...&hasta=...&page=1&page_size=50

    Devuelve fichas GEI individuales paginadas para consumo del LXP.
    """
    if request.method == 'OPTIONS':
        return _integracion_apply_cors(request, HttpResponse(status=204))

    auth_error = _integracion_auth_error(request)
    if auth_error:
        return _integracion_apply_cors(request, auth_error)

    filtros, parse_error = _integracion_parse_filtros(request, permitir_curso=True)
    if parse_error:
        return _integracion_apply_cors(request, parse_error)

    try:
        from formulario.models import CAMPOS_GEI_7, FichaGEI
    except Exception:
        return _integracion_apply_cors(
            request,
            JsonResponse({'success': False, 'error': 'app formulario no disponible'}, status=400),
        )

    try:
        page = max(1, int(request.GET.get('page', '1')))
    except ValueError:
        page = 1
    try:
        page_size = int(request.GET.get('page_size', '50'))
    except ValueError:
        page_size = 50
    page_size = max(1, min(page_size, 200))

    qs = FichaGEI.objects.select_related('estudiante', 'cliente', 'curso')
    if filtros['cliente_id'] is not None:
        qs = qs.filter(cliente_id=filtros['cliente_id'])
    if filtros['curso_id'] is not None:
        qs = qs.filter(curso_id=filtros['curso_id'])
    qs = qs.filter(
        fecha_inicio__date__gte=filtros['fecha_desde'],
        fecha_inicio__date__lte=filtros['fecha_hasta'],
    ).order_by('-fecha_update')

    total = qs.count()
    inicio = (page - 1) * page_size
    fin = inicio + page_size
    fichas = list(qs[inicio:fin])

    items = []
    for f in fichas:
        items.append({
            'id': f.id,
            'estudiante_id': f.estudiante_id,
            'estudiante_nombre': getattr(f.estudiante, 'nombre', '') or '',
            'cliente_id': f.cliente_id,
            'cliente_nombre': f.cliente.nombre if f.cliente_id else None,
            'curso_id': f.curso_id,
            'curso_nombre': f.curso.nombre if f.curso_id else None,
            'nombre_finca': f.nombre_finca or '',
            'area_ha': f.area_ha,
            'num_plantas': f.num_plantas,
            'fertilizante_kg': f.fertilizante_kg,
            'concentracion_n_pct': f.concentracion_n_pct,
            'tipo_combustible': f.tipo_combustible or '',
            'combustible_gal': f.combustible_gal,
            'energia_kwh': f.energia_kwh,
            'residuos_ton': f.residuos_ton,
            'manejo_residuos': f.manejo_residuos or '',
            'produccion_kg': f.produccion_kg,
            'tiene_bosque': f.tiene_bosque,
            'area_bosque_ha': f.area_bosque_ha,
            'completitud_pct': f.completitud_pct,
            'fecha_inicio': f.fecha_inicio.isoformat() if f.fecha_inicio else None,
            'fecha_update': f.fecha_update.isoformat() if f.fecha_update else None,
        })

    response = JsonResponse({
        'success': True,
        'meta': {
            'schema_version': 1,
            'tenant_id': filtros['cliente_id'],
            'course_id': filtros['curso_id'],
            'desde': filtros['fecha_desde'].isoformat(),
            'hasta': filtros['fecha_hasta'].isoformat(),
            'page': page,
            'page_size': page_size,
            'total': total,
            'campos_gei': list(CAMPOS_GEI_7),
        },
        'fichas': items,
    })
    return _integracion_apply_cors(request, response)


@csrf_exempt
@require_http_methods(["GET", "OPTIONS"])
def api_integracion_gei_exportar(request):
    """GET /api/integracion/gei/exportar/?cliente_id=...&curso_id=...&desde=...&hasta=...

    Exporta las fichas GEI del periodo a XLSX (mismas columnas que el admin).
    """
    if request.method == 'OPTIONS':
        return _integracion_apply_cors(request, HttpResponse(status=204))

    auth_error = _integracion_auth_error(request)
    if auth_error:
        return _integracion_apply_cors(request, auth_error)

    filtros, parse_error = _integracion_parse_filtros(request, permitir_curso=True)
    if parse_error:
        return _integracion_apply_cors(request, parse_error)

    try:
        from formulario.models import FichaGEI
    except Exception:
        return _integracion_apply_cors(
            request,
            JsonResponse({'success': False, 'error': 'app formulario no disponible'}, status=400),
        )

    try:
        import io
        import openpyxl
    except Exception:
        return _integracion_apply_cors(
            request,
            JsonResponse({'success': False, 'error': 'openpyxl no disponible en el servidor'}, status=500),
        )

    qs = FichaGEI.objects.select_related('estudiante', 'cliente', 'curso')
    if filtros['cliente_id'] is not None:
        qs = qs.filter(cliente_id=filtros['cliente_id'])
    if filtros['curso_id'] is not None:
        qs = qs.filter(curso_id=filtros['curso_id'])
    qs = qs.filter(
        fecha_inicio__date__gte=filtros['fecha_desde'],
        fecha_inicio__date__lte=filtros['fecha_hasta'],
    ).order_by('-fecha_update')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FichasGEI'
    encabezado = [
        'id', 'id_estudiante', 'nombre_estudiante', 'id_cliente', 'nombre_cliente',
        'id_curso', 'nombre_curso', 'nombre_finca', 'area_ha', 'num_plantas',
        'fertilizante_kg', 'concentracion_n_pct', 'tipo_combustible', 'combustible_gal',
        'energia_kwh', 'residuos_ton', 'manejo_residuos', 'produccion_kg',
        'tiene_bosque', 'area_bosque_ha', 'completitud_pct', 'fecha_inicio', 'fecha_update',
    ]
    for c, t in enumerate(encabezado, start=1):
        ws.cell(1, c, t)
    for r, f in enumerate(qs, start=2):
        fila = [
            f.id,
            f.estudiante_id,
            getattr(f.estudiante, 'nombre', '') or '',
            f.cliente_id or '',
            f.cliente.nombre if f.cliente_id else '',
            f.curso_id or '',
            f.curso.nombre if f.curso_id else '',
            f.nombre_finca or '',
            f.area_ha,
            f.num_plantas,
            f.fertilizante_kg,
            f.concentracion_n_pct,
            f.tipo_combustible or '',
            f.combustible_gal,
            f.energia_kwh,
            f.residuos_ton,
            f.manejo_residuos or '',
            f.produccion_kg,
            f.tiene_bosque,
            f.area_bosque_ha,
            f.completitud_pct,
            timezone.localtime(f.fecha_inicio).replace(tzinfo=None) if f.fecha_inicio else '',
            timezone.localtime(f.fecha_update).replace(tzinfo=None) if f.fecha_update else '',
        ]
        for c, v in enumerate(fila, start=1):
            ws.cell(r, c, v)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"fichas_gei_{filtros['fecha_desde'].isoformat()}_{filtros['fecha_hasta'].isoformat()}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return _integracion_apply_cors(request, response)


@csrf_exempt
@require_http_methods(["GET", "OPTIONS"])
def api_integracion_empleabilidad_metricas(request):
    """GET /api/integracion/empleabilidad/metricas/?cliente_id=...&fecha=YYYY-MM-DD"""
    if request.method == 'OPTIONS':
        return _integracion_apply_cors(request, HttpResponse(status=204))

    auth_error = _integracion_auth_error(request)
    if auth_error:
        return _integracion_apply_cors(request, auth_error)

    filtros, parse_error = _integracion_parse_filtros(request, permitir_curso=False)
    if parse_error:
        return _integracion_apply_cors(request, parse_error)

    cliente_id = filtros['cliente_id']
    fecha_desde = filtros['fecha_desde']
    fecha_hasta = filtros['fecha_hasta']
    rango_dias = filtros['rango_dias']

    qs = MisionEmpleabilidad.objects.filter(
        fecha_descubierta__date__gte=fecha_desde,
        fecha_descubierta__date__lte=fecha_hasta,
    )
    if cliente_id is not None:
        qs = qs.filter(cliente_id=cliente_id)

    agregados = list(
        qs.annotate(metric_day=TruncDate('fecha_descubierta'))
        .values('metric_day', 'estado_flujo')
        .annotate(c=Count('id'))
        .order_by('metric_day', 'estado_flujo')
    )

    metrics = []
    resumen_por_estado = {}
    resumen_por_dia = {}

    for row in agregados:
        day = row.get('metric_day')
        estado = row.get('estado_flujo') or 'sin_estado'
        count = int(row.get('c') or 0)
        day_txt = str(day or fecha_desde)

        resumen_por_estado[estado] = resumen_por_estado.get(estado, 0) + count
        resumen_por_dia[day_txt] = resumen_por_dia.get(day_txt, 0) + count

        metrics.append({
            'schema_version': 1,
            'metric_date': day_txt,
            'tenant_id': cliente_id,
            'metric_name': f'embudo_{estado}',
            'metric_value': count,
        })

    d = fecha_desde
    while d <= fecha_hasta:
        day_txt = d.isoformat()
        metrics.append({
            'schema_version': 1,
            'metric_date': day_txt,
            'tenant_id': cliente_id,
            'metric_name': 'misiones_total',
            'metric_value': int(resumen_por_dia.get(day_txt, 0)),
        })
        d += timedelta(days=1)

    response = JsonResponse({
        'success': True,
        'meta': {
            'schema_version': 1,
            'tenant_id': cliente_id,
            'desde': fecha_desde.isoformat(),
            'hasta': fecha_hasta.isoformat(),
            'dias': rango_dias,
        },
        'resumen': {
            'misiones_total': int(qs.count()),
            'por_estado': resumen_por_estado,
        },
        'metrics': metrics,
    })
    return _integracion_apply_cors(request, response)
