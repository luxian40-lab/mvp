"""
Vista de Dashboard con Métricas Visuales
Muestra estadísticas clave del sistema en tiempo real
"""
from django.shortcuts import render
from django.template.response import TemplateResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Q, Avg, Sum, F, Max
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.http import HttpResponse
from django.utils import timezone as dj_tz
from datetime import datetime, timedelta
import re
from collections import Counter, defaultdict
import json
import io
import openpyxl

from .models import (
    Estudiante, WhatsappLog, Plantilla, Campana, EnvioLog,
    ProgresoEstudiante, Curso, ModuloCompletado, Cliente,
    PerfilGamificacion, Badge, TransaccionPuntos, SolicitudSoporte,
    Certificado, Linea
)
from .gamificacion import BadgeEstudiante

try:
    from formulario.models import FichaGEI, SesionFormulario
except ImportError:
    FichaGEI = None
    SesionFormulario = None


def _excel_safe(value):
    """openpyxl no soporta datetime con timezone."""
    if hasattr(value, "tzinfo") and getattr(value, "tzinfo", None) is not None:
        try:
            return dj_tz.localtime(value).replace(tzinfo=None)
        except Exception:
            return value.replace(tzinfo=None)
    return value


@staff_member_required
def dashboard_metricas(request):
    """Dashboard excepcional con métricas completas de control"""

    # ========== PARÁMETROS DE TIEMPO ==========
    cliente_filtro_raw = (request.GET.get('cliente') or '').strip()
    cliente_filtro = int(cliente_filtro_raw) if cliente_filtro_raw.isdigit() else None
    ahora = datetime.now()
    ahora_tz = dj_tz.now()
    hace_1_dia = ahora - timedelta(days=1)
    hace_7_dias = ahora - timedelta(days=7)
    hace_30_dias = ahora - timedelta(days=30)
    hace_90_dias = ahora - timedelta(days=90)

    # ========== SCOPES (filtro por empresa) ==========
    clientes_q = Cliente.objects.all()
    if cliente_filtro:
        clientes_q = clientes_q.filter(id=cliente_filtro)

    estudiantes_q = Estudiante.objects.all()
    if cliente_filtro:
        estudiantes_q = estudiantes_q.filter(cliente_id=cliente_filtro)

    progresos_q = ProgresoEstudiante.objects.all()
    if cliente_filtro:
        progresos_q = progresos_q.filter(estudiante__cliente_id=cliente_filtro)

    telefonos_estudiantes = list(
        estudiantes_q.exclude(telefono__isnull=True).exclude(telefono='').values_list('telefono', flat=True)
    )
    whatsapp_q = WhatsappLog.objects.all()
    if cliente_filtro:
        whatsapp_q = whatsapp_q.filter(
            Q(telefono__in=telefonos_estudiantes) | Q(estudiante__cliente_id=cliente_filtro)
        ).distinct()

    # ========== 1. MÉTRICAS DE CLIENTES ==========
    total_clientes = clientes_q.count()
    clientes_activos = clientes_q.filter(activo=True).count() if hasattr(Cliente, 'activo') else total_clientes

    # Métricas detalladas por cliente
    clientes_detalle = []
    clientes_labels = []
    clientes_estudiantes_data = []
    clientes_mensajes_data = []
    
    for cliente in clientes_q:
        estudiantes_cliente = estudiantes_q.filter(cliente=cliente)
        estudiantes_activos_cliente = estudiantes_cliente.filter(activo=True).count()
        mensajes_cliente = whatsapp_q.filter(
            telefono__in=estudiantes_cliente.values_list('telefono', flat=True)
        ).count()
        
        clientes_detalle.append({
            'nombre': cliente.nombre,
            'total_estudiantes': estudiantes_cliente.count(),
            'estudiantes_activos': estudiantes_activos_cliente,
            'mensajes': mensajes_cliente,
        })
        
        # Datos para gráfica
        clientes_labels.append(cliente.nombre[:15])  # Limitar nombre
        clientes_estudiantes_data.append(estudiantes_cliente.count())
        clientes_mensajes_data.append(mensajes_cliente)

    # ========== 2. MÉTRICAS DE ESTUDIANTES ==========
    total_estudiantes = estudiantes_q.count()
    estudiantes_activos = estudiantes_q.filter(activo=True).count()
    estudiantes_nuevos_30d = estudiantes_q.filter(fecha_registro__gte=hace_30_dias).count()
    estudiantes_nuevos_7d = estudiantes_q.filter(fecha_registro__gte=hace_7_dias).count()

    # Distribución por cliente
    estudiantes_por_cliente = estudiantes_q.values('cliente__nombre').annotate(
        total=Count('id'),
        activos=Count('id', filter=Q(activo=True))
    ).order_by('-total')

    # ========== 3. MÉTRICAS DE MENSAJERÍA ==========
    total_mensajes = whatsapp_q.count()
    mensajes_24h = whatsapp_q.filter(fecha__gte=hace_1_dia).count()
    mensajes_7d = whatsapp_q.filter(fecha__gte=hace_7_dias).count()
    mensajes_30d = whatsapp_q.filter(fecha__gte=hace_30_dias).count()

    # Tipos de mensajes
    mensajes_enviados = whatsapp_q.filter(tipo='SENT').count()
    mensajes_recibidos = whatsapp_q.filter(tipo='INCOMING').count()
    mensajes_error = whatsapp_q.filter(
        tipo='SENT',
    ).filter(
        Q(estado__iexact='error') | Q(estado__iexact='failed') | Q(estado__iexact='undelivered')
    ).count()
    mensajes_recibidos_confirmados = whatsapp_q.filter(
        tipo='SENT',
        estado__in=['SENT', 'DELIVERED', 'READ', 'sent', 'delivered', 'read'],
    ).count()
    mensajes_abiertos = whatsapp_q.filter(
        tipo='SENT', estado__iexact='READ'
    ).count()
    mensajes_entregados = whatsapp_q.filter(
        tipo='SENT', estado__iexact='DELIVERED'
    ).count()
    mensajes_en_transito = whatsapp_q.filter(
        tipo='SENT',
        estado__in=['PENDING', 'QUEUED', 'SENDING', 'pending', 'queued', 'sending'],
    ).count()

    estudiantes_abrieron = Estudiante.objects.filter(
        telefono__in=whatsapp_q.filter(tipo='SENT', estado__iexact='READ').values_list('telefono', flat=True)
    ).values('id', 'nombre', 'telefono').order_by('nombre')[:50]

    # Aperturas (READ) con última fecha — incluye números sin estudiante (ej. bot comercial / prospectos)
    aperturas_rows = (
        whatsapp_q.filter(tipo='SENT', estado__iexact='READ')
        .values('telefono')
        .annotate(ultima=Max('fecha'))
        .order_by('-ultima')[:50]
    )
    tel_a_nombre = {}
    for est in estudiantes_q.filter(activo=True).only('nombre', 'telefono'):
        digits = re.sub(r'\D', '', est.telefono or '')
        if len(digits) >= 8:
            tel_a_nombre[digits] = est.nombre
            tel_a_nombre[digits[-10:]] = est.nombre
    aperturas_whatsapp = []
    for row in aperturas_rows:
        tel = row['telefono'] or ''
        d = re.sub(r'\D', '', tel)
        nombre = ''
        if len(d) >= 8:
            nombre = tel_a_nombre.get(d) or tel_a_nombre.get(d[-10:]) or ''
        aperturas_whatsapp.append(
            {'telefono': tel, 'nombre': nombre or '—', 'fecha': row['ultima']}
        )

    bot_comercial_sent = whatsapp_q.filter(
        tipo='SENT', agente_usado='BOT_COMERCIAL'
    ).count()
    bot_comercial_read = whatsapp_q.filter(
        tipo='SENT', agente_usado='BOT_COMERCIAL', estado__iexact='READ'
    ).count()
    bot_comercial_delivered = whatsapp_q.filter(
        tipo='SENT', agente_usado='BOT_COMERCIAL', estado__iexact='DELIVERED'
    ).count()

    # Tasa de éxito
    tasa_exito_mensajes = ((mensajes_enviados - mensajes_error) / mensajes_enviados * 100) if mensajes_enviados > 0 else 0

    # ========== 3B. MÉTRICAS FICHA GEI (recolección por WhatsApp) ==========
    gei_disponible = FichaGEI is not None and SesionFormulario is not None
    gei_fichas_total = gei_fichas_30d = 0
    gei_sesiones_activas = gei_sesiones_completadas_30d = 0
    gei_promedio_completitud = 0.0
    gei_fichas_completas = gei_fichas_parciales = gei_fichas_pendientes = 0
    gei_campo_menor = None
    gei_tiempo_promedio_min = None
    if gei_disponible:
        try:
            from formulario.models import CAMPOS_GEI_7
        except Exception:
            CAMPOS_GEI_7 = ()
        fichas_gei_q = FichaGEI.objects.all()
        ses_gei_q = SesionFormulario.objects.all()
        if cliente_filtro:
            fichas_gei_q = fichas_gei_q.filter(cliente_id=cliente_filtro)
            ses_gei_q = ses_gei_q.filter(estudiante__cliente_id=cliente_filtro)
        gei_fichas_total = fichas_gei_q.count()
        gei_fichas_30d = fichas_gei_q.filter(fecha_inicio__gte=ahora_tz - timedelta(days=30)).count()
        gei_sesiones_activas = ses_gei_q.filter(completado=False).count()
        gei_sesiones_completadas_30d = ses_gei_q.filter(
            completado=True, fecha_update__gte=ahora_tz - timedelta(days=30)
        ).count()
        _muestra = list(fichas_gei_q.order_by("-fecha_update")[:300])
        nulos_por_campo = {c: 0 for c in CAMPOS_GEI_7}
        if _muestra:
            suma_pct = 0
            for ficha in _muestra:
                pct = int(ficha.completitud_pct or 0)
                suma_pct += pct
                if pct == 100:
                    gei_fichas_completas += 1
                elif pct == 0:
                    gei_fichas_pendientes += 1
                else:
                    gei_fichas_parciales += 1
                for c in CAMPOS_GEI_7:
                    v = getattr(ficha, c, None)
                    if v is None or v == "":
                        nulos_por_campo[c] += 1
            gei_promedio_completitud = suma_pct / len(_muestra)
            if nulos_por_campo:
                cm = max(nulos_por_campo.items(), key=lambda kv: kv[1])
                if cm[1] > 0:
                    gei_campo_menor = {
                        "campo": cm[0],
                        "fichas_sin_dato": cm[1],
                        "porcentaje_sin_dato": round(cm[1] * 100 / len(_muestra), 1),
                    }

        completadas_30d_qs = list(
            ses_gei_q.filter(
                completado=True, fecha_update__gte=ahora_tz - timedelta(days=30)
            )[:300]
        )
        if completadas_30d_qs:
            total_seg = 0
            n_sesiones = 0
            for s in completadas_30d_qs:
                if s.fecha_inicio and s.fecha_update:
                    total_seg += (s.fecha_update - s.fecha_inicio).total_seconds()
                    n_sesiones += 1
            if n_sesiones:
                gei_tiempo_promedio_min = round((total_seg / n_sesiones) / 60, 1)

    # ========== 4. MÉTRICAS DE CAMPAÑAS ==========
    campanas_q = Campana.objects.filter(cliente_id=cliente_filtro) if cliente_filtro else Campana.objects.all()
    total_campanas = campanas_q.count()
    campanas_ejecutadas = campanas_q.filter(ejecutada=True).count()
    campanas_programadas = campanas_q.filter(fecha_programada__gte=ahora).count()

    # Envíos de campañas
    total_envios_campanas = EnvioLog.objects.count()
    envios_exitosos = EnvioLog.objects.filter(estado='exitoso').count()
    envios_fallidos = EnvioLog.objects.filter(estado='fallido').count()

    tasa_exito_campanas = (envios_exitosos / total_envios_campanas * 100) if total_envios_campanas > 0 else 0

    # ========== 5. MÉTRICAS EDUCATIVAS ==========
    cursos_q = Curso.objects.filter(cliente_id=cliente_filtro) if cliente_filtro else Curso.objects.all()
    total_cursos = cursos_q.count()
    cursos_activos = cursos_q.filter(activo=True).count()

    estudiantes_inscritos = progresos_q.count()
    cursos_completados = progresos_q.filter(completado=True).count()
    cursos_en_proceso = progresos_q.filter(completado=False).count()
    modulos_completados = ModuloCompletado.objects.filter(progreso__in=progresos_q).count()
    corte_stale = ahora_tz - timedelta(days=14)
    estudiantes_iniciaron_y_abandonaron = (
        progresos_q.filter(completado=False)
        .annotate(_mc=Count('modulos_completados'))
        .filter(_mc__gte=1)
        .filter(
            Q(fecha_ultimo_avance__lt=ahora_tz - timedelta(days=7))
            | Q(fecha_ultimo_avance__isnull=True, fecha_inicio__lt=corte_stale)
        )
        .select_related('estudiante', 'curso', 'modulo_actual')
        .order_by(F('fecha_ultimo_avance').asc(nulls_first=True))[:60]
    )

    tasa_completacion = (cursos_completados / estudiantes_inscritos * 100) if estudiantes_inscritos > 0 else 0

    # Progreso por curso
    progreso_cursos = []
    for curso in cursos_q[:10]:  # Top 10 cursos
        inscritos = progresos_q.filter(curso=curso).count()
        completados = progresos_q.filter(curso=curso, completado=True).count()
        porcentaje = (completados / inscritos * 100) if inscritos > 0 else 0

        progreso_cursos.append({
            'curso': curso,
            'inscritos': inscritos,
            'completados': completados,
            'porcentaje': round(porcentaje, 1)
        })

    # ========== 6. MÉTRICAS DE PLANTILLAS ==========
    total_plantillas = Plantilla.objects.count()
    plantillas_activas = Plantilla.objects.filter(activa=True).count()
    plantillas_twilio = Plantilla.objects.filter(twilio_template_sid__isnull=False, aprobada_twilio=True).count()

    # Plantillas más usadas
    top_plantillas = Plantilla.objects.filter(activa=True).order_by('-veces_usada')[:8]

    # ========== 7. MÉTRICAS DE GAMIFICACIÓN ==========
    total_puntos_otorgados = TransaccionPuntos.objects.filter(tipo='credito').aggregate(total=Sum('puntos'))['total'] or 0
    total_puntos_canjeados = TransaccionPuntos.objects.filter(tipo='debito').aggregate(total=Sum('puntos'))['total'] or 0

    total_badges = Badge.objects.count()
    badges_otorgados = BadgeEstudiante.objects.count()

    perfiles_gamificacion = PerfilGamificacion.objects.count()

    # ========== 8. MÉTRICAS DE SOPORTE ==========
    solicitudes_q = SolicitudSoporte.objects.filter(estudiante__cliente_id=cliente_filtro) if cliente_filtro else SolicitudSoporte.objects.all()
    total_solicitudes = solicitudes_q.count()
    solicitudes_pendientes = solicitudes_q.filter(estado='pendiente').count()
    solicitudes_atendiendo = solicitudes_q.filter(estado='en_atencion').count()
    solicitudes_resueltas = solicitudes_q.filter(estado='resuelta').count()

    # ========== 9. MÉTRICAS DE CERTIFICADOS ==========
    certs_q = Certificado.objects.filter(estudiante__cliente_id=cliente_filtro) if cliente_filtro else Certificado.objects.all()
    total_certificados = certs_q.count()
    certificados_generados_30d = certs_q.filter(fecha_emision__gte=hace_30_dias).count()

    # ========== 10. ANÁLISIS TEMPORAL ==========
    # Mensajes por día (últimos 14 días)
    mensajes_por_dia = whatsapp_q.filter(fecha__gte=hace_30_dias).annotate(
        dia=TruncDate('fecha')
    ).values('dia').annotate(
        total=Count('id'),
        enviados=Count('id', filter=Q(tipo='SENT')),
        recibidos=Count('id', filter=Q(tipo='INCOMING'))
    ).order_by('dia')

    # Preparar datos para gráficas
    fechas = []
    mensajes_diarios = []
    envios_diarios = []

    for item in mensajes_por_dia:
        fechas.append(item['dia'].strftime('%d/%m'))
        mensajes_diarios.append(item['total'])
        envios_diarios.append(item['enviados'])

    # ========== 11. ALERTAS Y NOTIFICACIONES ==========
    alertas = []

    # Alertas críticas
    if mensajes_error > 10:
        alertas.append({
            'tipo': 'error',
            'titulo': 'Altos errores de envío',
            'mensaje': f'{mensajes_error} mensajes con error en las últimas 24h',
            'icono': '⚠️'
        })

    if solicitudes_pendientes > 5:
        alertas.append({
            'tipo': 'warning',
            'titulo': 'Solicitudes de soporte pendientes',
            'mensaje': f'{solicitudes_pendientes} solicitudes esperando atención',
            'icono': '📞'
        })

    if estudiantes_nuevos_7d == 0:
        alertas.append({
            'tipo': 'info',
            'titulo': 'Sin nuevos estudiantes',
            'mensaje': 'No se registraron estudiantes en los últimos 7 días',
            'icono': '📈'
        })

    # ========== 12. KPIs PRINCIPALES ==========
    kpis = [
        {
            'titulo': 'Total Clientes',
            'valor': f"{total_clientes}",
            'cambio': f"{clientes_activos} activos",
            'tendencia': 'up',
            'icono': '🏢',
            'color': 'info'
        },
        {
            'titulo': 'Tasa de Éxito',
            'valor': f"{tasa_exito_campanas:.1f}%",
            'cambio': '+2.3%',
            'tendencia': 'up',
            'icono': '🎯',
            'color': 'primary'
        },
        {
            'titulo': 'Estudiantes Activos',
            'valor': f"{estudiantes_activos:,}",
            'cambio': f"+{estudiantes_nuevos_7d}",
            'tendencia': 'up',
            'icono': '👥',
            'color': 'info'
        },
        {
            'titulo': 'Mensajes Enviados',
            'valor': f"{mensajes_7d:,}",
            'cambio': f"+{mensajes_24h}",
            'tendencia': 'up',
            'icono': '💬',
            'color': 'warning'
        }
    ]

    # ========== CONTEXTO PARA TEMPLATE ==========
    context = {
        # KPIs principales
        'kpis': kpis,
        'alertas': alertas,

        # Clientes
        'total_clientes': total_clientes,
        'clientes_activos': clientes_activos,
        'clientes': Cliente.objects.all().order_by('nombre'),
        'cliente_filtro': cliente_filtro,
        'clientes_detalle': clientes_detalle,
        'clientes_labels_json': json.dumps(clientes_labels),
        'clientes_estudiantes_json': json.dumps(clientes_estudiantes_data),
        'clientes_mensajes_json': json.dumps(clientes_mensajes_data),

        # Estudiantes
        'total_estudiantes': total_estudiantes,
        'estudiantes_activos': estudiantes_activos,
        'estudiantes_nuevos_30d': estudiantes_nuevos_30d,
        'estudiantes_nuevos_7d': estudiantes_nuevos_7d,
        'estudiantes_por_cliente': estudiantes_por_cliente,

        # Mensajería
        'total_mensajes': total_mensajes,
        'mensajes_24h': mensajes_24h,
        'mensajes_7d': mensajes_7d,
        'mensajes_30d': mensajes_30d,
        'mensajes_enviados': mensajes_enviados,
        'mensajes_recibidos': mensajes_recibidos,
        'mensajes_error': mensajes_error,
        'mensajes_recibidos_confirmados': mensajes_recibidos_confirmados,
        'mensajes_entregados': mensajes_entregados,
        'mensajes_abiertos': mensajes_abiertos,
        'mensajes_en_transito': mensajes_en_transito,
        'estudiantes_abrieron': list(estudiantes_abrieron),
        'aperturas_whatsapp': aperturas_whatsapp,
        'bot_comercial_sent': bot_comercial_sent,
        'bot_comercial_read': bot_comercial_read,
        'bot_comercial_delivered': bot_comercial_delivered,
        'tasa_exito_mensajes': round(tasa_exito_mensajes, 1),

        # Ficha GEI (recolección de datos vía formulario en WhatsApp)
        'gei_disponible': gei_disponible,
        'gei_fichas_total': gei_fichas_total,
        'gei_fichas_30d': gei_fichas_30d,
        'gei_sesiones_activas': gei_sesiones_activas,
        'gei_sesiones_completadas_30d': gei_sesiones_completadas_30d,
        'gei_promedio_completitud': round(gei_promedio_completitud, 1),
        'gei_fichas_completas': gei_fichas_completas,
        'gei_fichas_parciales': gei_fichas_parciales,
        'gei_fichas_pendientes': gei_fichas_pendientes,
        'gei_campo_menor': gei_campo_menor,
        'gei_tiempo_promedio_min': gei_tiempo_promedio_min,

        # Campañas
        'total_campanas': total_campanas,
        'campanas_ejecutadas': campanas_ejecutadas,
        'campanas_programadas': campanas_programadas,
        'total_envios_campanas': total_envios_campanas,
        'envios_exitosos': envios_exitosos,
        'envios_fallidos': envios_fallidos,
        'tasa_exito_campanas': round(tasa_exito_campanas, 1),

        # Educación
        'total_cursos': total_cursos,
        'cursos_activos': cursos_activos,
        'estudiantes_inscritos': estudiantes_inscritos,
        'cursos_completados': cursos_completados,
        'cursos_en_proceso': cursos_en_proceso,
        'modulos_completados': modulos_completados,
        'estudiantes_iniciaron_y_abandonaron': estudiantes_iniciaron_y_abandonaron,
        'tasa_completacion': round(tasa_completacion, 1),
        'progreso_cursos': progreso_cursos,

        # Plantillas
        'total_plantillas': total_plantillas,
        'plantillas_activas': plantillas_activas,
        'plantillas_twilio': plantillas_twilio,
        'top_plantillas': top_plantillas,

        # Gamificación
        'total_puntos_otorgados': total_puntos_otorgados,
        'total_puntos_canjeados': total_puntos_canjeados,
        'total_badges': total_badges,
        'badges_otorgados': badges_otorgados,
        'perfiles_gamificacion': perfiles_gamificacion,

        # Soporte
        'total_solicitudes': total_solicitudes,
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_atendiendo': solicitudes_atendiendo,
        'solicitudes_resueltas': solicitudes_resueltas,

        # Certificados
        'total_certificados': total_certificados,
        'certificados_generados_30d': certificados_generados_30d,

        # Datos para gráficas
        'fechas_json': json.dumps(fechas),
        'mensajes_diarios_json': json.dumps(mensajes_diarios),
        'envios_diarios_json': json.dumps(envios_diarios),

        # Información temporal
        'fecha_actualizacion': ahora.strftime('%d/%m/%Y %H:%M'),
    }

    return TemplateResponse(request, 'admin/dashboard_metricas.html', context)


@staff_member_required
def dashboard_gerencial(request):
    """Dashboard gerencial con vista ejecutiva — mismas métricas, template diferente"""
    response = dashboard_metricas(request)
    response.template_name = 'admin/dashboard_gerencial.html'
    return response


@staff_member_required
def exportar_metricas_excel(request):
    cliente_raw = (request.GET.get('cliente') or '').strip()
    cliente_id = int(cliente_raw) if cliente_raw.isdigit() else None
    ahora = dj_tz.now()
    hace_30 = ahora - timedelta(days=30)

    estudiantes_q = Estudiante.objects.all().select_related('cliente')
    if cliente_id:
        estudiantes_q = estudiantes_q.filter(cliente_id=cliente_id)
    telefonos = list(estudiantes_q.exclude(telefono__isnull=True).exclude(telefono='').values_list('telefono', flat=True))

    whatsapp_q = WhatsappLog.objects.all()
    if cliente_id:
        whatsapp_q = whatsapp_q.filter(Q(telefono__in=telefonos) | Q(estudiante__cliente_id=cliente_id)).distinct()

    progresos_q = ProgresoEstudiante.objects.all().select_related('estudiante', 'curso', 'modulo_actual')
    if cliente_id:
        progresos_q = progresos_q.filter(estudiante__cliente_id=cliente_id)

    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["Métrica", "Valor", "Periodo"])
    ws_resumen.append(["Estudiantes activos", estudiantes_q.filter(activo=True).count(), "Actual"])
    ws_resumen.append(["Mensajes enviados", whatsapp_q.filter(tipo='SENT').count(), "Actual"])
    ws_resumen.append(["Mensajes recibidos", whatsapp_q.filter(tipo='INCOMING').count(), "Actual"])
    ws_resumen.append(["Cursos completados", progresos_q.filter(completado=True).count(), "Actual"])
    ws_resumen.append(["Módulos completados", ModuloCompletado.objects.filter(progreso__in=progresos_q).count(), "Actual"])
    ws_resumen.append(["Tránsito DELIVERED", whatsapp_q.filter(tipo='SENT', estado__iexact='DELIVERED').count(), "Actual"])
    ws_resumen.append(["Tránsito READ", whatsapp_q.filter(tipo='SENT', estado__iexact='READ').count(), "Actual"])
    ws_resumen.append(["Sesiones bot comercial", whatsapp_q.filter(agente_usado='BOT_COMERCIAL').count(), "Actual"])

    ws_est = wb.create_sheet("Estudiantes")
    ws_est.append(["Nombre", "Cédula", "Teléfono", "Cliente", "Estado", "Último mensaje", "Curso actual", "Módulo actual", "% Progreso", "Fecha inscripción"])
    ultimo_msg_por_tel = {
        row['telefono']: row['ultima']
        for row in whatsapp_q.values('telefono').annotate(ultima=Max('fecha'))
    }
    progreso_por_est = {p.estudiante_id: p for p in progresos_q.order_by('-fecha_ultimo_avance')}
    for est in estudiantes_q:
        prog = progreso_por_est.get(est.id)
        ws_est.append([
            est.nombre,
            est.cedula,
            est.telefono,
            est.cliente.nombre if est.cliente_id else '',
            est.estado_chat,
            _excel_safe(ultimo_msg_por_tel.get(est.telefono)),
            prog.curso.nombre if prog else '',
            f"M{prog.modulo_actual.numero}" if prog and prog.modulo_actual_id else '',
            prog.porcentaje_avance() if prog else 0,
            _excel_safe(est.fecha_registro),
        ])

    ws_prog = wb.create_sheet("Progreso por Curso")
    ws_prog.append(["Estudiante", "Curso", "Módulo actual", "Completado", "Fecha completado", "Puntos acumulados"])
    puntos_por_est = {
        p.estudiante_id: p.puntos_totales for p in PerfilGamificacion.objects.filter(estudiante_id__in=progresos_q.values_list('estudiante_id', flat=True))
    }
    for p in progresos_q:
        ws_prog.append([
            p.estudiante.nombre if p.estudiante_id else '',
            p.curso.nombre if p.curso_id else '',
            p.modulo_actual.numero if p.modulo_actual_id else '',
            "Sí" if p.completado else "No",
            _excel_safe(p.fecha_completado),
            puntos_por_est.get(p.estudiante_id, 0),
        ])

    ws_wa = wb.create_sheet("WhatsApp Logs (ultimos 30 dias)")
    ws_wa.append(["Fecha", "Teléfono", "Tipo", "Estado", "Mensaje (truncado 100 chars)", "Error detalle"])
    for log in whatsapp_q.filter(fecha__gte=hace_30).order_by('-fecha')[:5000]:
        ws_wa.append([_excel_safe(log.fecha), log.telefono, log.tipo, log.estado, (log.mensaje or '')[:100], (log.error_detalle or '')[:200]])

    ws_bot = wb.create_sheet("Bot Comercial")
    ws_bot.append(["Fecha", "Teléfono", "Pregunta", "Respuesta (truncado)", "Fuente RAG"])
    incomings = list(
        whatsapp_q.filter(agente_usado='BOT_COMERCIAL', tipo='INCOMING').order_by('-fecha')[:2000]
    )
    sents = list(
        whatsapp_q.filter(agente_usado='BOT_COMERCIAL', tipo='SENT').order_by('-fecha')[:2000]
    )
    # Emparejamiento simple por teléfono + cercanía temporal
    sents_por_tel = defaultdict(list)
    for s in sents:
        sents_por_tel[s.telefono].append(s)
    for inc in incomings:
        resp = next((s for s in sents_por_tel.get(inc.telefono, []) if s.fecha >= inc.fecha), None)
        fuente = 'RAG' if resp and ('información oficial' in (resp.mensaje or '').lower() or 'base técnica' in (resp.mensaje or '').lower()) else ''
        ws_bot.append([
            _excel_safe(inc.fecha),
            inc.telefono,
            (inc.mensaje or '')[:200],
            (resp.mensaje or '')[:240] if resp else '',
            fuente,
        ])

    cliente_nombre = "todos"
    if cliente_id:
        c = Cliente.objects.filter(id=cliente_id).first()
        if c:
            cliente_nombre = re.sub(r'[^a-zA-Z0-9_-]+', '_', c.nombre.strip())[:40] or "cliente"
    fecha_str = ahora.strftime("%Y%m%d_%H%M")
    filename = f"eki_metricas_{cliente_nombre}_{fecha_str}.xlsx"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
