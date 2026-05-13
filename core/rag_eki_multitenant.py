"""
RAG Multi-Tenant para EKI MVP
Cada combinación Cliente + Curso tiene su propia BD vectorial aislada.
Compañía A NO puede ver documentos de Compañía B.
Curso 1 NO puede ver documentos de Curso 2.
"""
import os
import logging
from typing import List, Dict, Tuple
from django.conf import settings

logger = logging.getLogger(__name__)

# =========================================================
# Import ChromaDB con fallback graceful
# =========================================================
try:
    import chromadb
    CHROMADB_DISPONIBLE = True
except ImportError:
    CHROMADB_DISPONIBLE = False
    logger.warning("[RAG] chromadb no disponible — RAG deshabilitado")


class RAGClienteCurso:
    """
    RAG independiente para CADA combinación Cliente + Curso.
    Totalmente aislado, sin mezcla de datos.

    Estructura:
        chroma_db/
        ├── cliente_1/
        │   ├── curso_1/   ← ChromaDB aislado
        │   └── curso_2/   ← ChromaDB aislado
        └── cliente_2/
            └── curso_1/   ← ChromaDB aislado
    """

    def __init__(self, cliente_id: int, curso_id: int):
        if not CHROMADB_DISPONIBLE:
            raise RuntimeError("chromadb no está instalado")

        self.cliente_id = cliente_id
        self.curso_id = curso_id

        # Ruta aislada por cliente + curso
        base = getattr(settings, 'CHROMA_DB_DIR', os.path.join(settings.BASE_DIR, 'chroma_db'))
        self.db_path = os.path.join(base, f"cliente_{cliente_id}", f"curso_{curso_id}")
        os.makedirs(self.db_path, exist_ok=True)

        # ChromaDB persistente y aislado
        self.client = chromadb.PersistentClient(path=self.db_path)
        collection_name = f"cli{cliente_id}_cur{curso_id}"
        self.collection = self.client.get_or_create_collection(
            name=collection_name,
            metadata={"hnsw:space": "cosine"}
        )

        logger.info(f"[RAG] Inicializado: Cliente {cliente_id}, Curso {curso_id} ({self.db_path})")

    # ==========================================
    # PROCESAMIENTO DE DOCUMENTOS
    # ==========================================

    def procesar_documento(
        self,
        ruta_archivo: str,
        nombre_documento: str,
        tipo: str = "contenido"
    ) -> int:
        """
        Procesa un documento y lo indexa SOLO para este cliente+curso.

        Args:
            ruta_archivo: Ruta al archivo (PDF, DOCX, TXT)
            nombre_documento: Identificador único del documento
            tipo: "contenido" | "manual" | "faq"

        Returns:
            Cantidad de chunks indexados
        """
        logger.info(f"[RAG] Procesando: {nombre_documento} (Cliente {self.cliente_id}, Curso {self.curso_id})")

        texto = self._extraer_texto(ruta_archivo)
        if not texto:
            logger.error(f"[RAG] No se pudo extraer texto de {ruta_archivo}")
            return 0

        chunks = self._dividir_chunks(texto)
        if not chunks:
            return 0

        # Eliminar chunks anteriores del mismo documento (re-upload)
        self._eliminar_por_fuente(nombre_documento)

        ids = []
        documents = []
        metadatas = []

        for i, chunk in enumerate(chunks):
            ids.append(f"{nombre_documento}__chunk_{i}")
            documents.append(chunk)
            metadatas.append({
                "source": nombre_documento,
                "tipo": tipo,
                "chunk_num": i,
                "cliente_id": self.cliente_id,
                "curso_id": self.curso_id,
            })

        # Agregar en batch
        self.collection.add(ids=ids, documents=documents, metadatas=metadatas)
        logger.info(f"[RAG] {nombre_documento}: {len(ids)} chunks indexados")
        return len(ids)

    def procesar_texto_directo(
        self,
        texto: str,
        nombre_documento: str,
        tipo: str = "contenido"
    ) -> int:
        """
        Indexa texto directo (ej: contenido de un módulo) sin archivo.
        """
        if not texto or not texto.strip():
            return 0

        chunks = self._dividir_chunks(texto)
        if not chunks:
            return 0

        self._eliminar_por_fuente(nombre_documento)

        ids = []
        documents = []
        metadatas = []

        for i, chunk in enumerate(chunks):
            ids.append(f"{nombre_documento}__chunk_{i}")
            documents.append(chunk)
            metadatas.append({
                "source": nombre_documento,
                "tipo": tipo,
                "chunk_num": i,
                "cliente_id": self.cliente_id,
                "curso_id": self.curso_id,
            })

        self.collection.add(ids=ids, documents=documents, metadatas=metadatas)
        logger.info(f"[RAG] Texto directo '{nombre_documento}': {len(ids)} chunks")
        return len(ids)

    # ==========================================
    # EXTRACCIÓN DE TEXTO
    # ==========================================

    def _extraer_texto(self, ruta: str) -> str:
        """Extrae texto de PDF, DOCX, TXT o Excel (XLSX/XLSM; p. ej. listas de precios)."""
        ext = os.path.splitext(ruta)[1].lower()

        if ext == '.pdf':
            return self._extraer_pdf(ruta)
        elif ext == '.docx':
            return self._extraer_docx(ruta)
        elif ext == '.txt':
            return self._extraer_txt(ruta)
        elif ext in ('.xlsx', '.xlsm'):
            return self._extraer_xlsx(ruta)
        else:
            logger.warning(f"[RAG] Tipo no soportado: {ext}")
            return ""

    @staticmethod
    def _extraer_pdf(ruta: str) -> str:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(ruta)
            return "\n".join(p.extract_text() or "" for p in reader.pages)
        except Exception as e:
            logger.error(f"[RAG] Error PDF: {e}")
            return ""

    @staticmethod
    def _extraer_docx(ruta: str) -> str:
        try:
            from docx import Document
            doc = Document(ruta)
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as e:
            logger.error(f"[RAG] Error DOCX: {e}")
            return ""

    @staticmethod
    def _extraer_xlsx(ruta: str) -> str:
        """Excel: filas como texto con encabezados (mejor para embeddings y precios por producto)."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(ruta, read_only=True, data_only=True)
            blocks: list[str] = []
            try:
                for sheet in wb.worksheets:
                    blocks.append(f"## Hoja: {sheet.title}")
                    header: list[str] | None = None
                    for row in sheet.iter_rows(values_only=True):
                        cells = [
                            str(c).strip() if c is not None and str(c).strip() != "" else ""
                            for c in (row or ())
                        ]
                        if not any(cells):
                            continue
                        if header is None:
                            header = cells
                            blocks.append("ENCABEZADOS: " + " | ".join(header))
                            continue
                        n = max(len(header), len(cells))
                        pairs: list[str] = []
                        for i in range(n):
                            h = header[i] if i < len(header) else ""
                            v = cells[i] if i < len(cells) else ""
                            if not h and not v:
                                continue
                            if h and v:
                                pairs.append(f"{h}: {v}")
                            elif v:
                                pairs.append(v)
                        if pairs:
                            blocks.append("FILA | " + " ; ".join(pairs))
            finally:
                wb.close()
            return "\n".join(blocks)
        except Exception as e:
            logger.error(f"[RAG] Error XLSX/XLSM: {e}")
            return ""

    @staticmethod
    def _extraer_txt(ruta: str) -> str:
        try:
            with open(ruta, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            logger.error(f"[RAG] Error TXT: {e}")
            return ""

    # ==========================================
    # CHUNKING
    # ==========================================

    @staticmethod
    def _dividir_chunks(texto: str, chunk_size: int = 800, overlap: int = 100) -> List[str]:
        """Divide texto en chunks con overlap, cortando en puntos naturales."""
        chunks = []
        start = 0
        while start < len(texto):
            end = start + chunk_size
            if end < len(texto):
                # Buscar punto de corte natural
                last_period = texto[start:end].rfind('.')
                last_newline = texto[start:end].rfind('\n')
                cut = max(last_period, last_newline)
                if cut > chunk_size // 2:
                    end = start + cut + 1
            chunk = texto[start:end].strip()
            if chunk:
                chunks.append(chunk)
            start = end - overlap
        return chunks

    # ==========================================
    # BÚSQUEDA (SOLO en este cliente+curso)
    # ==========================================

    def buscar(self, pregunta: str, top_k: int = 3) -> List[Dict]:
        """
        Busca documentos relevantes SOLO en este cliente+curso.

        Returns:
            Lista de dicts: {contenido, fuente, tipo, similitud}
        """
        try:
            count = self.collection.count()
            if count == 0:
                return []

            results = self.collection.query(
                query_texts=[pregunta],
                n_results=min(top_k, count)
            )

            docs = []
            for doc, meta, dist in zip(
                results['documents'][0],
                results['metadatas'][0],
                results['distances'][0]
            ):
                docs.append({
                    'contenido': doc,
                    'fuente': meta.get('source', ''),
                    'tipo': meta.get('tipo', ''),
                    'similitud': round(1 - dist, 3) if dist <= 2 else 0
                })
            return docs

        except Exception as e:
            logger.error(f"[RAG] Error buscando: {e}")
            return []

    def obtener_contexto_rag(self, pregunta: str, max_chars: int = 2000) -> str:
        """
        Obtiene contexto RAG formateado para inyectar en prompts de IA.
        Retorna string vacío si no hay documentos relevantes.
        """
        docs = self.buscar(pregunta, top_k=3)
        if not docs:
            return ""

        fragmentos = []
        chars_total = 0
        for doc in docs:
            fragmento = f"[Fuente: {doc['fuente']}]\n{doc['contenido']}"
            if chars_total + len(fragmento) > max_chars:
                break
            fragmentos.append(fragmento)
            chars_total += len(fragmento)

        if not fragmentos:
            return ""

        return (
            "\n\n📚 DOCUMENTOS DEL CURSO (usa esta información para responder):\n"
            + "\n---\n".join(fragmentos)
            + "\n\n⚠️ REGLA: Prioriza la información de estos documentos sobre tu conocimiento general.\n"
        )

    # ==========================================
    # RESPUESTA COMPLETA
    # ==========================================

    def responder(self, pregunta: str, personalidad: str = "geronimo") -> Tuple[str, List[Dict]]:
        """
        Genera respuesta usando SOLO documentos de este cliente+curso.

        Args:
            pregunta: Pregunta del estudiante
            personalidad: "geronimo" o "maria"

        Returns:
            (respuesta, documentos_usados)
        """
        docs = self.buscar(pregunta, top_k=3)

        if not docs:
            msg = "No encontré esa información en el material del curso. ¿Podrías reformular tu pregunta?" if personalidad == "geronimo" else "Déjame revisar el material del curso."
            return msg, []

        contexto = "\n\n---\n\n".join(
            f"[De: {d['fuente']}]\n{d['contenido']}" for d in docs
        )

        if personalidad == "geronimo":
            system = (
                "Eres el Profesor Gerónimo, educador agrícola de eki.\n"
                "REGLAS:\n"
                "1. SOLO usa información de los documentos proporcionados.\n"
                "2. Si no está en los documentos, di: 'No lo encontré en el material'.\n"
                "3. Máximo 60 palabras.\n"
                "4. Máximo 2 emojis.\n"
                "5. Método sandwich: concepto → ejemplo → pregunta.\n"
                "6. Siempre termina con una pregunta."
            )
        else:
            system = (
                "Eres María, mentora educativa de eki.\n"
                "REGLAS:\n"
                "1. SOLO usa documentos proporcionados.\n"
                "2. Si no está: 'Déjame revisar el material'.\n"
                "3. Resuelve dudas con ejemplos prácticos.\n"
                "4. Empática y motivadora."
            )

        user_msg = f"Documentos del curso:\n{contexto}\n\nPregunta: {pregunta}\n\nResponde SOLO con estos documentos."

        try:
            from openai import OpenAI
            client = OpenAI(api_key=getattr(settings, 'OPENAI_API_KEY', ''))
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user_msg}
                ],
                temperature=0.7,
                max_tokens=150,
                timeout=10
            )
            return response.choices[0].message.content.strip(), docs
        except Exception as e:
            logger.error(f"[RAG] Error OpenAI: {e}")
            return "Error procesando tu pregunta. Intenta de nuevo.", []

    # ==========================================
    # UTILIDADES
    # ==========================================

    def listar_documentos(self) -> List[Dict]:
        """Lista documentos indexados en este cliente+curso."""
        try:
            results = self.collection.get()
            if not results['metadatas']:
                return []
            docs_map = {}
            for meta in results['metadatas']:
                src = meta.get('source', 'desconocido')
                if src not in docs_map:
                    docs_map[src] = {'tipo': meta.get('tipo', ''), 'chunks': 0}
                docs_map[src]['chunks'] += 1
            return [{'nombre': k, 'tipo': v['tipo'], 'chunks': v['chunks']} for k, v in docs_map.items()]
        except Exception as e:
            logger.error(f"[RAG] Error listando: {e}")
            return []

    def contar_chunks(self) -> int:
        try:
            return self.collection.count()
        except Exception:
            return 0

    def eliminar_documento(self, nombre_documento: str) -> bool:
        return self._eliminar_por_fuente(nombre_documento)

    def _eliminar_por_fuente(self, nombre: str) -> bool:
        try:
            results = self.collection.get(where={"source": nombre})
            if results['ids']:
                self.collection.delete(ids=results['ids'])
                logger.info(f"[RAG] Eliminado: {nombre} ({len(results['ids'])} chunks)")
                return True
        except Exception as e:
            logger.error(f"[RAG] Error eliminando {nombre}: {e}")
        return False
