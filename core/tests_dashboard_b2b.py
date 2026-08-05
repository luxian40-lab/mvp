"""Tests dashboard unificado — Reportes B2B por grupo y exportación Excel."""
from __future__ import annotations

import io

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from django.test import Client
from django.test.utils import override_settings

from core.models import Cliente, Curso, Estudiante, ProgresoEstudiante
from core.models_extras import GrupoEstudiantes

pytestmark = pytest.mark.django_db


@override_settings(SECURE_SSL_REDIRECT=False)
def test_dashboard_b2b_filtra_por_grupo():
    User = get_user_model()
    User.objects.create_user("dashb2b", password="x", is_staff=True)

    cli = Cliente.objects.create(
        nombre="Org B2B",
        contacto_principal="C",
        email="b2b@example.com",
        telefono="573001112233",
    )
    curso = Curso.objects.create(nombre="Curso B2B", descripcion="d", cliente=cli)

    e1 = Estudiante.objects.create(
        cedula="B2B00001",
        nombre="En Grupo",
        telefono="57300000001",
        cliente=cli,
        activo=True,
    )
    e2 = Estudiante.objects.create(
        cedula="B2B00002",
        nombre="Sin Grupo",
        telefono="57300000002",
        cliente=cli,
        activo=True,
    )
    g = GrupoEstudiantes.objects.create(nombre="Grupo A", cliente=cli)
    g.estudiantes.add(e1)

    ProgresoEstudiante.objects.create(estudiante=e1, curso=curso, completado=False)
    ProgresoEstudiante.objects.create(estudiante=e2, curso=curso, completado=True)

    client = Client()
    assert client.login(username="dashb2b", password="x")

    r_all = client.get(
        "/admin/dashboard/",
        {"tab": "reportes", "cliente": str(cli.id), "detalle": "1"},
        HTTP_HOST="127.0.0.1",
    )
    assert r_all.status_code == 200
    assert b"En Grupo" in r_all.content
    assert b"Sin Grupo" in r_all.content

    r_g = client.get(
        "/admin/dashboard/",
        {"tab": "reportes", "cliente": str(cli.id), "grupo": str(g.id), "detalle": "1"},
        HTTP_HOST="127.0.0.1",
    )
    assert r_g.status_code == 200
    assert b"En Grupo" in r_g.content
    assert b"Sin Grupo" not in r_g.content


@override_settings(SECURE_SSL_REDIRECT=False)
def test_dashboard_b2b_excel_incluye_grupo_y_estado():
    User = get_user_model()
    User.objects.create_user("dashxls", password="y", is_staff=True)

    cli = Cliente.objects.create(
        nombre="Org XLS",
        contacto_principal="C",
        email="xls@example.com",
        telefono="573001112244",
    )
    curso = Curso.objects.create(nombre="Curso XLS", descripcion="d", cliente=cli)
    e = Estudiante.objects.create(
        cedula="XLS00001",
        nombre="Fila Excel Uno",
        telefono="57300000003",
        cliente=cli,
        activo=True,
    )
    g = GrupoEstudiantes.objects.create(nombre="GX", cliente=cli)
    g.estudiantes.add(e)
    ProgresoEstudiante.objects.create(estudiante=e, curso=curso, completado=False)

    client = Client()
    assert client.login(username="dashxls", password="y")

    resp = client.get(
        "/admin/dashboard/",
        {
            "exportar": "excel",
            "tab": "reportes",
            "cliente": str(cli.id),
            "grupo": str(g.id),
        },
        HTTP_HOST="127.0.0.1",
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Estudiantes B2B" in wb.sheetnames
    ws = wb["Estudiantes B2B"]
    assert ws.cell(1, 1).value == "Nombre"
    assert ws.cell(1, 3).value == "Teléfono"
    assert ws.cell(1, 6).value == "Grupo(s)"
    assert ws.cell(1, 11).value == "Estado avance"
    assert ws.cell(2, 1).value == "Fila Excel Uno"
    assert ws.cell(2, 3).value == "57300000003"
    assert "GX" in str(ws.cell(2, 6).value or "")


@override_settings(SECURE_SSL_REDIRECT=False)
def test_dashboard_reportes_default_sin_tablas_detalle():
    """Analítica abre rápido: sin ?detalle=1 no arma filas de estudiante."""
    User = get_user_model()
    User.objects.create_user("dashfast", password="z", is_staff=True)

    cli = Cliente.objects.create(
        nombre="Org Fast",
        contacto_principal="C",
        email="fast@example.com",
        telefono="573009998877",
    )
    curso = Curso.objects.create(nombre="Curso Fast", descripcion="d", cliente=cli)
    e = Estudiante.objects.create(
        cedula="FAST0001",
        nombre="NombreUnicoFastDetail",
        telefono="57300000999",
        cliente=cli,
        activo=True,
    )
    ProgresoEstudiante.objects.create(estudiante=e, curso=curso, completado=False)

    client = Client()
    assert client.login(username="dashfast", password="z")

    r_light = client.get(
        "/admin/dashboard/",
        {"tab": "learning", "section": "reportes", "cliente": str(cli.id)},
        HTTP_HOST="127.0.0.1",
    )
    assert r_light.status_code == 200
    assert r_light.context["detalle_cargado"] is False
    assert r_light.context["estudiantes_detalle"] == []
    assert b"Cargar tablas" in r_light.content

    r_heavy = client.get(
        "/admin/dashboard/",
        {"tab": "reportes", "cliente": str(cli.id), "detalle": "1"},
        HTTP_HOST="127.0.0.1",
    )
    assert r_heavy.status_code == 200
    assert r_heavy.context["detalle_cargado"] is True
    nombres = [row["nombre"] for row in r_heavy.context["estudiantes_detalle"]]
    assert any(n.lower() == "nombreunicofastdetail" for n in nombres)


@override_settings(SECURE_SSL_REDIRECT=False)
def test_dashboard_cursos_dropdown_solo_de_cliente():
    """Al filtrar por organización, el select de cursos no lista cursos ajenos."""
    User = get_user_model()
    User.objects.create_user("dashcurso", password="c", is_staff=True)

    cli_a = Cliente.objects.create(
        nombre="Org Alpha Filter",
        contacto_principal="A",
        email="alpha@example.com",
        telefono="573001110001",
    )
    cli_b = Cliente.objects.create(
        nombre="Org Beta Filter",
        contacto_principal="B",
        email="beta@example.com",
        telefono="573001110002",
    )
    Curso.objects.create(nombre="CursoSoloAlphaUnico", descripcion="d", cliente=cli_a)
    Curso.objects.create(nombre="CursoSoloBetaUnico", descripcion="d", cliente=cli_b)

    client = Client()
    assert client.login(username="dashcurso", password="c")

    r = client.get(
        "/admin/dashboard/",
        {"tab": "learning", "section": "reportes", "cliente": str(cli_a.id)},
        HTTP_HOST="127.0.0.1",
    )
    assert r.status_code == 200
    assert b"CursoSoloAlphaUnico" in r.content
    assert b"CursoSoloBetaUnico" not in r.content
