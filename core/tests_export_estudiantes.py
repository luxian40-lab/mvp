"""Tests exportación Excel de estudiantes."""
from __future__ import annotations

from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from core.export_estudiantes import limpiar_telefono
from core.models import Cliente, Estudiante, ProgresoEstudiante, Curso


class TestLimpiarTelefono(TestCase):
    def test_telefono_sin_prefijo(self):
        self.assertEqual(limpiar_telefono('+57 300 123 4567'), '3001234567')

    def test_telefono_whatsapp_format(self):
        self.assertEqual(limpiar_telefono('whatsapp:+573001234567'), '3001234567')


class TestColumnasExportExcel(TestCase):
    def test_columnas_presentes(self):
        cliente = Cliente.objects.create(
            nombre='Export Co',
            nit='900000002-0',
            contacto_principal='T',
            email='ex@example.com',
            telefono='573000088888',
            activo=True,
        )
        est = Estudiante.objects.create(
            cedula='7777777',
            nombre='Export Test',
            telefono='573001234567',
            municipio='Medellín',
            departamento='Antioquia',
            cliente=cliente,
            activo=True,
        )
        curso = Curso.objects.create(nombre='Curso X', cliente=cliente, activo=True)
        ProgresoEstudiante.objects.create(estudiante=est, curso=curso)

        from django.contrib import admin
        from core.admin import EstudianteAdmin

        ma = EstudianteAdmin(Estudiante, admin.site)
        from django.test import RequestFactory
        from django.contrib.auth import get_user_model
        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib.sessions.middleware import SessionMiddleware

        User = get_user_model()
        user = User.objects.create_superuser('admin_export', 'a@b.com', 'pass')
        request = RequestFactory().get('/')
        request.user = user
        middleware = SessionMiddleware(lambda r: None)
        middleware.process_request(request)
        request.session.save()
        request._messages = FallbackStorage(request)
        response = ma.exportar_estudiantes_por_curso(request, Estudiante.objects.filter(pk=est.pk))
        self.assertIsNotNone(response)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for col in ('Municipio', 'Departamento', 'Ciudad', 'Teléfono'):
            self.assertIn(col, headers)
        tel_idx = headers.index('Teléfono') + 1
        self.assertEqual(ws.cell(row=2, column=tel_idx).value, '3001234567')
