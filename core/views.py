from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt

def _construir_dashboard_unificado_contexto(request, incluir_detalle=True):
    """Construye contexto y payload JSON del dashboard unificado usando filtros consistentes."""
    from datetime import datetime, timedelta
    from django.db.models import Avg, Count, Q
    from django.db.models.functions import TruncDate
    from django.utils import timezone
    import json

    def _to_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _to_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            return None

    cliente_id = _to_int((request.GET.get('cliente') or '').strip())
    curso_id = _to_int((request.GET.get('curso') or '').strip())
    fecha_inicio_raw = (request.GET.get('fecha_inicio') or '').strip()
    fecha_fin_raw = (request.GET.get('fecha_fin') or '').strip()
    municipio_filtro = (request.GET.get('municipio') or '').strip()
    tab_raw = (request.GET.get('tab') or 'executive').strip().lower()
    from core.domains.dashboard import resolve_dashboard_tab, resolve_learning_section

    tab_actual = resolve_dashboard_tab(tab_raw)
    learning_section = resolve_learning_section(tab_raw, request.GET.get('section'))
    grupo_id = _to_int((request.GET.get('grupo_id') or request.GET.get('grupo') or '').strip())
    modulo_hasta_numero = _to_int((request.GET.get('modulo_hasta') or '').strip())

    fecha_inicio_dt = _to_date(fecha_inicio_raw)
    fecha_fin_dt = _to_date(fecha_fin_raw)
    if fecha_inicio_dt and fecha_fin_dt and fecha_inicio_dt > fecha_fin_dt:
        fecha_inicio_dt, fecha_fin_dt = fecha_fin_dt, fecha_inicio_dt

    fecha_inicio = fecha_inicio_dt.isoformat() if fecha_inicio_dt else ''
    fecha_fin = fecha_fin_dt.isoformat() if fecha_fin_dt else ''

    # Retención: contexto liviano (evita recalcular Executive/Learning completo).
    if tab_actual == 'retencion':
        from portal.retencion_service import analitica_retencion_portal

        clientes_all = Cliente.objects.all().order_by('nombre')
        cursos_all = Curso.objects.filter(activo=True).order_by('orden', 'nombre')
        if cliente_id:
            cursos_all = cursos_all.filter(cliente_id=cliente_id)
        grupos_qs = GrupoEstudiantes.objects.all().order_by('nombre')
        if cliente_id:
            grupos_qs = grupos_qs.filter(Q(cliente_id=cliente_id) | Q(cliente__isnull=True))

        retencion_data = None
        if cliente_id:
            org = Cliente.objects.filter(pk=cliente_id, activo=True).first()
            if org:
                desde_ret = (request.GET.get('desde') or fecha_inicio or '').strip() or None
                hasta_ret = (request.GET.get('hasta') or fecha_fin or '').strip() or None
                retencion_data = analitica_retencion_portal(
                    org,
                    curso_id=curso_id,
                    grupo_id=grupo_id,
                    desde=desde_ret,
                    hasta=hasta_ret,
                )

        context = {
            'tab_actual': tab_actual,
            'learning_section': learning_section,
            'clientes': Cliente.objects.all().order_by('nombre'),
            'cursos': cursos_all,
            'cursos_retencion': cursos_all,
            'cliente_filtro': cliente_id,
            'curso_filtro': curso_id,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'municipios': [],
            'municipio_filtro': municipio_filtro,
            'grupos': grupos_qs,
            'grupo_filtro': grupo_id,
            'modulo_hasta_filtro': modulo_hasta_numero,
            'retencion_data': retencion_data,
            'clientes_detalle': [],
            'estudiantes_detalle': [],
            'tickets_soporte': [],
            'eventos_ia_recientes': [],
            'resumen_payload_json': '{}',
            'chart_labels': '[]',
            'chart_values': '[]',
            'chart_ubicaciones_labels': '[]',
            'chart_ubicaciones_values': '[]',
            'chart_tipos_labels': '[]',
            'chart_tipos_values': '[]',
            'total_cursos': 0,
            'total_clientes': 0,
            'total_estudiantes': 0,
            'total_mensajes_whatsapp': 0,
            'mensajes_enviados': 0,
            'mensajes_recibidos': 0,
            'wa_entregados': 0,
            'wa_leidos': 0,
            'wa_en_transito': 0,
            'wa_bot_comercial_sent': 0,
            'wa_bot_comercial_read': 0,
            'total_audios': 0,
            'total_agentes_ia': 0,
            'total_progreso': 0,
            'total_modulos_completados': 0,
            'cursos_completados': 0,
            'total_certificados': 0,
            'total_perfiles_gam': 0,
            'puntos_promedio': 0,
            'top_estudiantes': [],
            'ranking_gamificacion_completo': [],
            'total_prospectos': 0,
            'tasa_completacion': 0,
            'ubicaciones_municipio': [],
            'progreso_por_curso': [],
        }
        return context, {}

    clientes_all = Cliente.objects.all().order_by('nombre')
    cursos_all = Curso.objects.all().order_by('nombre')

    estudiantes_q = Estudiante.objects.filter(activo=True)
    if cliente_id:
        estudiantes_q = estudiantes_q.filter(cliente_id=cliente_id)
    if curso_id:
        estudiantes_q = estudiantes_q.filter(progresos__curso_id=curso_id).distinct()
    if fecha_inicio_dt:
        estudiantes_q = estudiantes_q.filter(fecha_registro__date__gte=fecha_inicio_dt)
    if fecha_fin_dt:
        estudiantes_q = estudiantes_q.filter(fecha_registro__date__lte=fecha_fin_dt)
    if grupo_id:
        estudiantes_q = estudiantes_q.filter(grupos__id=grupo_id).distinct()

    progreso_q = ProgresoEstudiante.objects.select_related('estudiante', 'curso')
    if cliente_id:
        progreso_q = progreso_q.filter(estudiante__cliente_id=cliente_id)
    if curso_id:
        progreso_q = progreso_q.filter(curso_id=curso_id)
    if fecha_inicio_dt:
        progreso_q = progreso_q.filter(fecha_inicio__date__gte=fecha_inicio_dt)
    if fecha_fin_dt:
        progreso_q = progreso_q.filter(fecha_inicio__date__lte=fecha_fin_dt)
    if grupo_id:
        progreso_q = progreso_q.filter(estudiante__grupos__id=grupo_id).distinct()

    modulos_completados_q = ModuloCompletado.objects.select_related('progreso', 'modulo')
    if cliente_id:
        modulos_completados_q = modulos_completados_q.filter(progreso__estudiante__cliente_id=cliente_id)
    if curso_id:
        modulos_completados_q = modulos_completados_q.filter(progreso__curso_id=curso_id)
    if fecha_inicio_dt:
        modulos_completados_q = modulos_completados_q.filter(fecha_completado__date__gte=fecha_inicio_dt)
    if fecha_fin_dt:
        modulos_completados_q = modulos_completados_q.filter(fecha_completado__date__lte=fecha_fin_dt)
    if grupo_id:
        modulos_completados_q = modulos_completados_q.filter(
            progreso__estudiante__grupos__id=grupo_id
        ).distinct()

    whatsapp_q = WhatsappLog.objects.all()
    if fecha_inicio_dt:
        whatsapp_q = whatsapp_q.filter(fecha__date__gte=fecha_inicio_dt)
    if fecha_fin_dt:
        whatsapp_q = whatsapp_q.filter(fecha__date__lte=fecha_fin_dt)

    if cliente_id or curso_id:
        estudiantes_scope = Estudiante.objects.filter(activo=True)
        if cliente_id:
            estudiantes_scope = estudiantes_scope.filter(cliente_id=cliente_id)
        if curso_id:
            estudiantes_scope = estudiantes_scope.filter(progresos__curso_id=curso_id).distinct()
        if grupo_id:
            estudiantes_scope = estudiantes_scope.filter(grupos__id=grupo_id).distinct()

        telefonos_scope = estudiantes_scope.exclude(telefono='').values_list('telefono', flat=True)
        whatsapp_q = whatsapp_q.filter(
            Q(estudiante__in=estudiantes_scope) | Q(telefono__in=telefonos_scope)
        ).distinct()

    total_estudiantes = estudiantes_q.count()

    if cliente_id:
        total_clientes = Cliente.objects.filter(id=cliente_id).count()
    elif curso_id or fecha_inicio_dt or fecha_fin_dt:
        total_clientes = Cliente.objects.filter(estudiantes__in=estudiantes_q).distinct().count()
    else:
        total_clientes = Cliente.objects.count()

    progreso_filter_q = Q(progresoestudiante__id__in=progreso_q.values('id'))
    progreso_por_curso_qs = Curso.objects.all()
    if cliente_id or curso_id or fecha_inicio_dt or fecha_fin_dt:
        progreso_por_curso_qs = progreso_por_curso_qs.filter(progreso_filter_q)

    progreso_por_curso = progreso_por_curso_qs.annotate(
        total_estudiantes=Count('progresoestudiante', filter=progreso_filter_q, distinct=True),
        total_modulos_completados=Count('progresoestudiante__modulos_completados', filter=progreso_filter_q, distinct=True),
        completados=Count(
            'progresoestudiante',
            filter=progreso_filter_q & Q(progresoestudiante__completado=True),
            distinct=True,
        ),
    ).order_by('nombre')

    total_cursos = Curso.objects.count() if not (cliente_id or curso_id or fecha_inicio_dt or fecha_fin_dt) else progreso_por_curso.count()

    total_mensajes_whatsapp = whatsapp_q.count()
    mensajes_enviados = whatsapp_q.filter(tipo='SENT').count()
    mensajes_recibidos = whatsapp_q.filter(tipo='INCOMING').count()
    total_audios = whatsapp_q.filter(es_audio=True).count()
    total_agentes_ia = whatsapp_q.filter(agente_usado__isnull=False).exclude(agente_usado='').count()

    wa_entregados = whatsapp_q.filter(tipo='SENT', estado__iexact='DELIVERED').count()
    wa_leidos = whatsapp_q.filter(tipo='SENT', estado__iexact='READ').count()
    wa_en_transito = whatsapp_q.filter(
        tipo='SENT',
        estado__in=['PENDING', 'QUEUED', 'SENDING', 'pending', 'queued', 'sending'],
    ).count()
    wa_bot_comercial_sent = whatsapp_q.filter(tipo='SENT', agente_usado='BOT_COMERCIAL').count()
    wa_bot_comercial_read = whatsapp_q.filter(
        tipo='SENT', agente_usado='BOT_COMERCIAL', estado__iexact='READ'
    ).count()

    total_progreso = progreso_q.count()
    total_modulos_completados = modulos_completados_q.count()
    cursos_completados = progreso_q.filter(completado=True).count()

    try:
        from .models_certificados import Certificado
        certificados_q = Certificado.objects.all()
        if cliente_id:
            certificados_q = certificados_q.filter(estudiante__cliente_id=cliente_id)
        if curso_id:
            certificados_q = certificados_q.filter(curso_id=curso_id)
        if fecha_inicio_dt:
            certificados_q = certificados_q.filter(fecha_emision__date__gte=fecha_inicio_dt)
        if fecha_fin_dt:
            certificados_q = certificados_q.filter(fecha_emision__date__lte=fecha_fin_dt)
        total_certificados = certificados_q.count()
    except Exception:
        total_certificados = 0

    total_perfiles_gam = 0
    puntos_promedio = 0
    top_estudiantes = []
    ranking_gamificacion_completo = []
    try:
        from .gamificacion import PerfilGamificacion
        perfiles_q = PerfilGamificacion.objects.filter(estudiante_id__in=estudiantes_q.values('id'))
        total_perfiles_gam = perfiles_q.count()
        puntos_promedio = perfiles_q.aggregate(avg=Avg('puntos_totales'))['avg'] or 0
        if incluir_detalle:
            rank_qs = (
                perfiles_q.select_related('estudiante', 'estudiante__cliente')
                .order_by('-puntos_totales')
            )
            top_estudiantes = rank_qs[:10]
            ranking_gamificacion_completo = list(rank_qs[:2000])
    except Exception:
        pass

    ubicaciones_municipio = (
        estudiantes_q.exclude(municipio__isnull=True)
        .exclude(municipio='')
        .values('municipio')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    chart_ubicaciones_labels = [u['municipio'] for u in ubicaciones_municipio]
    chart_ubicaciones_values = [u['total'] for u in ubicaciones_municipio]

    hoy = fecha_fin_dt or timezone.localdate()
    hace_7_dias = hoy - timedelta(days=6)
    mensajes_por_dia = (
        whatsapp_q.filter(fecha__date__gte=hace_7_dias, fecha__date__lte=hoy)
        .annotate(dia=TruncDate('fecha'))
        .values('dia')
        .annotate(total=Count('id'))
        .order_by('dia')
    )
    mensajes_por_dia_map = {m['dia']: int(m['total']) for m in mensajes_por_dia}

    chart_labels = []
    chart_values = []
    for i in range(7):
        dia = hoy - timedelta(days=6 - i)
        chart_labels.append(dia.strftime('%d/%m'))
        chart_values.append(int(mensajes_por_dia_map.get(dia, 0)))

    tipos_msg = whatsapp_q.values('tipo').annotate(total=Count('id')).order_by('-total')
    chart_tipos_labels = [t['tipo'] or 'Otro' for t in tipos_msg]
    chart_tipos_values = [int(t['total']) for t in tipos_msg]

    try:
        from .models import ProspectoB2B
        prospectos_q = ProspectoB2B.objects.all()
        if fecha_inicio_dt:
            prospectos_q = prospectos_q.filter(fecha_captura__date__gte=fecha_inicio_dt)
        if fecha_fin_dt:
            prospectos_q = prospectos_q.filter(fecha_captura__date__lte=fecha_fin_dt)
        total_prospectos = prospectos_q.count()
    except Exception:
        total_prospectos = 0

    total_inscripciones = total_progreso
    tasa_completacion = round((cursos_completados / total_inscripciones * 100), 1) if total_inscripciones > 0 else 0

    municipios = list(
        estudiantes_q.exclude(municipio__isnull=True)
        .exclude(municipio='')
        .values_list('municipio', flat=True)
        .distinct()
        .order_by('municipio')
    )

    estudiantes_detalle = []
    clientes_detalle = []
    tickets_soporte = []
    if incluir_detalle:
        est_q = estudiantes_q.select_related('cliente').prefetch_related('grupos')
        if municipio_filtro:
            est_q = est_q.filter(municipio=municipio_filtro)

        est_ids = list(est_q[:200].values_list('id', flat=True))

        puntos_map = {}
        try:
            from .gamificacion import PerfilGamificacion as PG_detail
            puntos_map = dict(
                PG_detail.objects.filter(estudiante_id__in=est_ids).values_list('estudiante_id', 'puntos_totales')
            )
        except Exception:
            puntos_map = {}

        progresos_rows = (
            progreso_q.filter(estudiante_id__in=est_ids)
            .select_related('estudiante', 'estudiante__cliente', 'curso', 'modulo_actual')
            .prefetch_related('estudiante__grupos')
            .annotate(
                total_mods=Count('curso__modulos', distinct=True),
                mods_comp=Count('modulos_completados', distinct=True),
            )
            .order_by('estudiante__nombre', 'curso__nombre')
        )

        seen_estudiante_sin_progreso = set()
        from core.drip_schedule import avance_sobre_modulos, estudiante_llego_hasta_modulo, modulos_para_metricas

        for progreso in progresos_rows:
            est = progreso.estudiante
            if modulo_hasta_numero is not None and not estudiante_llego_hasta_modulo(progreso, modulo_hasta_numero):
                continue
            seen_estudiante_sin_progreso.add(est.id)
            total_mods = progreso.total_mods or 0
            mods_comp = progreso.mods_comp or 0
            avance = round(mods_comp / total_mods * 100) if total_mods > 0 else 0
            mods_drip = modulos_para_metricas(
                est,
                progreso.curso,
                modulo_hasta_numero=modulo_hasta_numero,
                usar_drip_calendario=modulo_hasta_numero is None,
            )
            comps_drip, total_drip, avance_drip = avance_sobre_modulos(progreso, mods_drip)
            if progreso.completado:
                estado_avance = 'Completado'
                modulo_txt = 'Curso completado'
            elif progreso.modulo_actual_id and progreso.modulo_actual:
                estado_avance = 'En curso'
                m = progreso.modulo_actual
                modulo_txt = f'M{m.numero} · {m.titulo}'
            elif avance > 0:
                estado_avance = 'En curso'
                modulo_txt = f'En curso ({mods_comp}/{total_mods} módulos)'
            else:
                estado_avance = 'Sin avance'
                modulo_txt = 'Sin iniciar'
            grupos_txt = ', '.join(sorted(g.nombre for g in est.grupos.all())) or '-'
            estudiantes_detalle.append({
                'nombre': est.nombre,
                'cedula': est.cedula,
                'telefono': (est.telefono or '').strip() or '-',
                'organizacion': est.cliente.nombre if est.cliente else '-',
                'municipio': est.municipio or '-',
                'curso': progreso.curso.nombre if progreso.curso_id else '-',
                'modulo_actual': modulo_txt,
                'modulos_completados': f'{mods_comp}/{total_mods}' if total_mods else '-',
                'modulos_drip': f'{comps_drip}/{total_drip}' if total_drip else '-',
                'avance': avance,
                'avance_drip': avance_drip,
                'puntos': puntos_map.get(est.id, 0),
                'grupos': grupos_txt,
                'estado_avance': estado_avance,
            })

        for est in est_q.filter(id__in=est_ids).exclude(id__in=seen_estudiante_sin_progreso):
            grupos_txt = ', '.join(sorted(g.nombre for g in est.grupos.all())) or '-'
            estudiantes_detalle.append({
                'nombre': est.nombre,
                'cedula': est.cedula,
                'telefono': (est.telefono or '').strip() or '-',
                'organizacion': est.cliente.nombre if est.cliente else '-',
                'municipio': est.municipio or '-',
                'curso': '-',
                'modulo_actual': '-',
                'modulos_completados': '-',
                'modulos_drip': '-',
                'avance': 0,
                'avance_drip': 0,
                'puntos': puntos_map.get(est.id, 0),
                'grupos': grupos_txt,
                'estado_avance': 'Sin inscripción',
            })

        clientes_iter = clientes_all if not cliente_id else clientes_all.filter(id=cliente_id)
        for c in clientes_iter:
            est_cliente = estudiantes_q.filter(cliente=c)
            n_est = est_cliente.count()
            tels = est_cliente.exclude(telefono='').values_list('telefono', flat=True)
            progreso_cliente = progreso_q.filter(estudiante__cliente=c)

            whatsapp_cliente = whatsapp_q.filter(
                Q(estudiante__cliente=c) | Q(telefono__in=tels)
            ).distinct()

            clientes_detalle.append({
                'nombre': c.nombre,
                'cursos': progreso_cliente.values('curso_id').distinct().count(),
                'estudiantes': n_est,
                'uso_audio': whatsapp_cliente.filter(es_audio=True).count(),
                'uso_ia': whatsapp_cliente.filter(agente_usado__isnull=False).exclude(agente_usado='').count(),
                'cursos_completados': progreso_cliente.filter(completado=True).count(),
            })

        try:
            from .models import SolicitudSoporte
            tickets_soporte_q = SolicitudSoporte.objects.select_related('estudiante')
            if cliente_id:
                tickets_soporte_q = tickets_soporte_q.filter(estudiante__cliente_id=cliente_id)
            if curso_id:
                tickets_soporte_q = tickets_soporte_q.filter(estudiante__progresos__curso_id=curso_id).distinct()
            if fecha_inicio_dt:
                tickets_soporte_q = tickets_soporte_q.filter(fecha_solicitud__date__gte=fecha_inicio_dt)
            if fecha_fin_dt:
                tickets_soporte_q = tickets_soporte_q.filter(fecha_solicitud__date__lte=fecha_fin_dt)
            tickets_soporte = tickets_soporte_q.order_by('-fecha_solicitud')[:50]
        except Exception:
            tickets_soporte = []

    grupos_qs = GrupoEstudiantes.objects.all().order_by('nombre')
    if cliente_id:
        grupos_qs = grupos_qs.filter(Q(cliente_id=cliente_id) | Q(cliente__isnull=True))

    resumen_payload = {
        'success': True,
        'generated_at': timezone.now().isoformat(),
        'kpis': {
            'total_cursos': int(total_cursos),
            'total_clientes': int(total_clientes),
            'total_estudiantes': int(total_estudiantes),
            'total_mensajes_whatsapp': int(total_mensajes_whatsapp),
            'mensajes_enviados': int(mensajes_enviados),
            'mensajes_recibidos': int(mensajes_recibidos),
            'wa_entregados': int(wa_entregados),
            'wa_leidos': int(wa_leidos),
            'wa_en_transito': int(wa_en_transito),
            'wa_bot_comercial_sent': int(wa_bot_comercial_sent),
            'wa_bot_comercial_read': int(wa_bot_comercial_read),
            'total_audios': int(total_audios),
            'total_agentes_ia': int(total_agentes_ia),
            'total_progreso': int(total_progreso),
            'total_modulos_completados': int(total_modulos_completados),
            'cursos_completados': int(cursos_completados),
            'total_certificados': int(total_certificados),
            'total_perfiles_gam': int(total_perfiles_gam),
            'puntos_promedio': round(float(puntos_promedio), 1),
            'total_prospectos': int(total_prospectos),
            'tasa_completacion': float(tasa_completacion),
        },
        'chart_mensajes': {
            'labels': chart_labels,
            'values': chart_values,
        },
        'chart_ubicaciones': {
            'labels': chart_ubicaciones_labels,
            'values': chart_ubicaciones_values,
        },
        'chart_tipos': {
            'labels': chart_tipos_labels,
            'values': chart_tipos_values,
        },
    }

    eventos_ia_recientes = []
    try:
        from core.models import EventoIA

        eventos_ia_recientes = list(
            EventoIA.objects.select_related('estudiante', 'curso', 'modulo')
            .order_by('-created_at')[:20]
        )
    except Exception:
        eventos_ia_recientes = []

    context = {
        'total_cursos': total_cursos,
        'total_clientes': total_clientes,
        'total_estudiantes': total_estudiantes,
        'total_mensajes_whatsapp': total_mensajes_whatsapp,
        'mensajes_enviados': mensajes_enviados,
        'mensajes_recibidos': mensajes_recibidos,
        'wa_entregados': wa_entregados,
        'wa_leidos': wa_leidos,
        'wa_en_transito': wa_en_transito,
        'wa_bot_comercial_sent': wa_bot_comercial_sent,
        'wa_bot_comercial_read': wa_bot_comercial_read,
        'total_audios': total_audios,
        'total_agentes_ia': total_agentes_ia,
        'total_progreso': total_progreso,
        'total_modulos_completados': total_modulos_completados,
        'cursos_completados': cursos_completados,
        'total_certificados': total_certificados,
        'total_perfiles_gam': total_perfiles_gam,
        'puntos_promedio': round(puntos_promedio, 1),
        'top_estudiantes': top_estudiantes,
        'ranking_gamificacion_completo': ranking_gamificacion_completo,
        'total_prospectos': total_prospectos,
        'tasa_completacion': tasa_completacion,
        'ubicaciones_municipio': ubicaciones_municipio,
        'progreso_por_curso': progreso_por_curso,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
        'chart_ubicaciones_labels': json.dumps(chart_ubicaciones_labels),
        'chart_ubicaciones_values': json.dumps(chart_ubicaciones_values),
        'chart_tipos_labels': json.dumps(chart_tipos_labels),
        'chart_tipos_values': json.dumps(chart_tipos_values),
        'resumen_payload_json': json.dumps(resumen_payload),
        'clientes': clientes_all,
        'cursos': cursos_all,
        'cliente_filtro': cliente_id,
        'curso_filtro': curso_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'municipios': municipios,
        'municipio_filtro': municipio_filtro,
        'tab_actual': tab_actual,
        'learning_section': learning_section,
        'grupos': grupos_qs,
        'grupo_filtro': grupo_id,
        'modulo_hasta_filtro': modulo_hasta_numero,
        'clientes_detalle': clientes_detalle,
        'estudiantes_detalle': estudiantes_detalle,
        'tickets_soporte': tickets_soporte,
        'eventos_ia_recientes': eventos_ia_recientes,
        'retencion_data': None,
        'cursos_retencion': Curso.objects.none(),
    }

    return context, resumen_payload


# Vista unificada del dashboard admin
@staff_member_required
def dashboard_unificado(request):
    """
    Dashboard profesional unificado de eki.
    Métricas reales: cursos, clientes, estudiantes, certificados, gamificación,
    WhatsApp, IA, y progreso educativo.
    """
    context, resumen_payload = _construir_dashboard_unificado_contexto(request, incluir_detalle=True)

    # --- Excel export (todas las pestañas + datos de gráficos) ---
    if request.GET.get('exportar') == 'excel':
        from analytics.exports import export_dashboard_excel

        return export_dashboard_excel(
            context=context,
            resumen_payload=resumen_payload,
            tab=context.get('tab_actual', 'executive'),
            learning_section=context.get('learning_section', 'reportes'),
        )

    return render(request, 'admin/dashboard.html', context)


@staff_member_required
def dashboard_unificado_resumen_data(request):
    """Endpoint JSON para refresco en tiempo real del Resumen Ejecutivo."""
    _, payload = _construir_dashboard_unificado_contexto(request, incluir_detalle=False)
    return JsonResponse(payload)


@staff_member_required
def bot_comercial_admin_view(request):
    """Vista administrativa para operación del Bot Comercial IA."""
    from datetime import timedelta
    from django.urls import reverse
    from django.utils import timezone as dj_tz

    from .models import DocumentoRAGComercial, WhatsappLog
    from .rag_comercial_manager import rag_comercial_manager

    endpoint_path = '/webhook/ia-bot-comercial/'
    endpoint_url = request.build_absolute_uri(endpoint_path)
    cliente_id = int(
        getattr(settings, 'BOT_COMERCIAL_CLIENTE_ID', 0) or 0
    )
    canal_rag = str(getattr(settings, 'BOT_COMERCIAL_RAG_CANAL', 'bot_comercial') or 'bot_comercial')

    docs_qs = DocumentoRAGComercial.objects.filter(cliente_id=cliente_id if cliente_id > 0 else None, canal=canal_rag)
    total_docs = docs_qs.count()
    total_docs_indexados = docs_qs.filter(estado='indexado').count()

    chunks_total = 0
    if rag_comercial_manager.disponible:
        chunks_total = rag_comercial_manager.contar_chunks(cliente_id=cliente_id, canal=canal_rag)

    hace_7 = dj_tz.now() - timedelta(days=7)
    bc_in = WhatsappLog.objects.filter(
        tipo='INCOMING', agente_usado='BOT_COMERCIAL', fecha__gte=hace_7
    ).count()
    bc_out = WhatsappLog.objects.filter(
        tipo='SENT', agente_usado='BOT_COMERCIAL', fecha__gte=hace_7
    ).count()
    bc_read = WhatsappLog.objects.filter(
        tipo='SENT', agente_usado='BOT_COMERCIAL', estado__iexact='READ'
    ).count()
    bc_delivered = WhatsappLog.objects.filter(
        tipo='SENT', agente_usado='BOT_COMERCIAL', estado__iexact='DELIVERED'
    ).count()
    callback_general = request.build_absolute_uri('/webhook/whatsapp/')

    context = {
        'endpoint_path': endpoint_path,
        'endpoint_url': endpoint_url,
        'bot_comercial_whatsapp_number': getattr(settings, 'BOT_COMERCIAL_WHATSAPP_NUMBER', ''),
        'bot_comercial_cliente_id': getattr(settings, 'BOT_COMERCIAL_CLIENTE_ID', ''),
        'bot_comercial_rag_canal': canal_rag,
        'bot_comercial_force_routing': bool(getattr(settings, 'BOT_COMERCIAL_FORCE_ROUTING', False)),
        'bot_comercial_openai_model': getattr(settings, 'BOT_COMERCIAL_OPENAI_MODEL', ''),
        'bot_comercial_vision_model': getattr(settings, 'BOT_COMERCIAL_VISION_MODEL', ''),
        'rag_comercial_disponible': rag_comercial_manager.disponible,
        'rag_comercial_total_docs': total_docs,
        'rag_comercial_docs_indexados': total_docs_indexados,
        'rag_comercial_chunks_total': chunks_total,
        'bot_comercial_incoming_7d': bc_in,
        'bot_comercial_sent_7d': bc_out,
        'bot_comercial_read_total': bc_read,
        'bot_comercial_delivered_total': bc_delivered,
        'memory_turnos': int(getattr(settings, 'BOT_COMERCIAL_MEMORY_TURNOS', 12) or 12),
        'memory_chars': int(getattr(settings, 'BOT_COMERCIAL_MEMORY_MAX_CHARS', 3600) or 3600),
        'twilio_status_callback_configured': bool(
            str(getattr(settings, 'TWILIO_STATUS_CALLBACK_URL', '') or '').strip()
        ),
        'twilio_status_callback_example': callback_general,
        'whatsapp_log_admin_url': reverse('admin:core_whatsapplog_changelist'),
    }
    return render(request, 'admin/bot_comercial.html', context)

from django.http import HttpResponse, JsonResponse, FileResponse, HttpResponseBadRequest
from django.core.files.storage import default_storage
import mimetypes
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q, Max, Count
from django.shortcuts import get_object_or_404
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
import math
import requests
import tempfile
import os
import logging

# Logger para debugging
logger = logging.getLogger(__name__)

from .models import Campana, Estudiante, WhatsappLog, EnvioLog, Cliente, Curso, ProgresoEstudiante, ModuloCompletado
from .models_extras import ArchivoModulo, GrupoEstudiantes
from .utils import enviar_whatsapp, enviar_whatsapp_twilio
from .intent_detector import detect_intent, mensaje_indica_listo as _mensaje_indica_listo
from .response_templates import (
    MENSAJE_CAPTION_SOLO_MEDIA as TWILIO_CAPTION_ADJUNTO,
    get_response_for_intent,
    parte_mensaje_con_media,
)

from django.views.decorators.csrf import csrf_exempt

# Endpoint proxy para servir archivos de S3 desde el dominio propio
@csrf_exempt
def serve_media_proxy(request, filename):
    s3_url = f"https://eki-produccion.s3.us-east-2.amazonaws.com/{filename}"
    r = requests.get(s3_url, stream=True)
    if r.status_code == 200:
        content_type = r.headers.get('Content-Type', 'application/octet-stream')
        content_length = r.headers.get('Content-Length')
        # Forzar Content-Disposition inline para WhatsApp
        content_disposition = f'inline; filename="{filename}"'
        response = FileResponse(r.raw, content_type=content_type)
        if content_length:
            response['Content-Length'] = content_length
        response['Content-Disposition'] = content_disposition
        # WhatsApp/Twilio requieren CORS headers a veces
        response['Access-Control-Allow-Origin'] = '*'
        return response
    else:
        return HttpResponseBadRequest("Archivo no encontrado o error en S3")


def _transcribir_audio_twilio(media_url, media_type='audio/ogg'):
    """
    Transcribe un audio de Twilio.
    
    Prioridad: OpenAI Whisper (confiable en producción), luego VOSK (offline).
    
    Args:
        media_url: URL del audio en Twilio
        media_type: MIME type del audio
    
    Returns:
        str: Texto transcrito o None si falla
    """
    try:
        # Obtener credenciales de Twilio para descargar el audio
        account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
        auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
        
        # Descargar el audio
        response = requests.get(media_url, auth=(account_sid, auth_token))
        response.raise_for_status()
        
        audio_size = len(response.content)
        print(f"🎤 Transcribiendo audio ({audio_size} bytes)...")
        
        # Guardar temporalmente con extensión acorde al MIME enviado por Twilio
        # Normalizar: quitar parámetros MIME (e.g. "audio/ogg; codecs=opus" → "audio/ogg")
        mime_base = (media_type or '').split(';')[0].strip().lower()
        suffix_map = {
            'audio/ogg': '.ogg',
            'audio/opus': '.ogg',
            'audio/mpeg': '.mp3',
            'audio/mp3': '.mp3',
            'audio/mp4': '.m4a',
            'audio/aac': '.ogg',   # Whisper no soporta .aac, guardar como .ogg
            'audio/amr': '.ogg',   # Whisper no soporta .amr, guardar como .ogg
            'audio/webm': '.webm',
            'audio/wav': '.wav',
        }
        suffix = suffix_map.get(mime_base, '.ogg')

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            tmp_file.write(response.content)
            audio_path = tmp_file.name
        
        try:
            # OPCIÓN 1: WHISPER (OpenAI — confiable en producción)
            openai_api_key = getattr(settings, 'OPENAI_API_KEY', '')
            if openai_api_key:
                max_whisper_size = 20 * 1024 * 1024  # 20MB
                if audio_size <= max_whisper_size:
                    print(f"🎤 Transcribiendo con Whisper (archivo: {suffix}, {audio_size} bytes)...")
                    try:
                        from openai import OpenAI
                        client = OpenAI(api_key=openai_api_key)

                        with open(audio_path, 'rb') as audio_file:
                            transcription = client.audio.transcriptions.create(
                                model="whisper-1",
                                file=audio_file,
                                language="es",
                            )

                        texto = (transcription.text or '').strip()
                        print(f"✅ Whisper transcribió: '{texto}'")
                        return texto if texto else None
                    except Exception as whisper_error:
                        print(f"❌ Error Whisper: {whisper_error}")
                        import traceback; traceback.print_exc()
                else:
                    print(f"⚠️ Audio demasiado grande para Whisper ({audio_size} bytes)")
            else:
                print(f"⚠️ OPENAI_API_KEY no configurada — Whisper no disponible")

            # OPCIÓN 2: VOSK (gratuito, offline — si modelo disponible)
            try:
                texto = _transcribir_con_vosk(audio_path)
                if texto:
                    print(f"✅ Vosk transcribió: '{texto}'")
                    return texto
                print(f"⚠️ Vosk retornó vacío")
            except Exception as vosk_error:
                print(f"⚠️ Vosk no disponible: {vosk_error}")
            
            # OPCIÓN 3: FALLBACK — no se pudo transcribir
            print("⚠️ Sin transcripción disponible - retornando None")
            return None
            
        finally:
            # Eliminar archivo temporal
            if os.path.exists(audio_path):
                os.remove(audio_path)
    
    except Exception as e:
        print(f"❌ Error en transcripción: {e}")
        return None


def _transcribir_con_vosk(audio_path):
    """
    Transcribe audio usando VOSK (gratuito, offline).
    
    Instalación requerida:
    - pip install vosk
    - Descargar modelo: https://alphacephei.com/vosk/models
    - Colocar en: models/vosk-model-small-es-0.42/
    """
    try:
        import json
        from vosk import Model, KaldiRecognizer
        from pydub import AudioSegment
        import wave
        
        # Ruta al modelo de Vosk (configurar en settings)
        model_path = getattr(settings, 'VOSK_MODEL_PATH', 'models/vosk-model-small-es-0.42')
        
        if not os.path.exists(model_path):
            raise FileNotFoundError(f"Modelo Vosk no encontrado en {model_path}")
        
        # Cargar modelo (se cachea automáticamente)
        model = Model(model_path)
        
        # Convertir audio a formato WAV 16kHz mono (requerido por Vosk)
        audio = AudioSegment.from_file(audio_path)
        audio = audio.set_frame_rate(16000).set_channels(1)

        # Guardar como WAV temporal — SIEMPRE en ruta distinta al original
        base, _ = os.path.splitext(audio_path)
        wav_path = base + '_vosk.wav'
        audio.export(wav_path, format='wav')

        try:
            # Transcribir usando wave module (más confiable)
            recognizer = KaldiRecognizer(model, 16000)

            wf = wave.open(wav_path, "rb")

            # Procesar por chunks
            while True:
                data = wf.readframes(4000)
                if len(data) == 0:
                    break
                recognizer.AcceptWaveform(data)

            wf.close()

            # Obtener resultado final
            result = json.loads(recognizer.FinalResult())
            texto = result.get('text', '').strip()

            print(f"✅ Vosk transcribió: '{texto}'")
            return texto if texto else None
        finally:
            # Limpiar archivo WAV temporal (siempre, incluso en error)
            if os.path.exists(wav_path):
                os.remove(wav_path)

    except Exception as e:
        print(f"❌ Error Vosk: {e}")
        raise  # Re-lanzar para que el fallback funcione


@staff_member_required
def dashboard_view(request):
    """Redirige al dashboard de métricas completo (contexto histórico incompleto aquí rompía la plantilla)."""
    from django.shortcuts import redirect
    from django.urls import reverse

    target = reverse('dashboard_metrics')
    qs = request.META.get('QUERY_STRING', '')
    if qs:
        return redirect(f'{target}?{qs}')
    return redirect(target)


# ---------- Vista de instrucciones ----------
@staff_member_required
def instrucciones_view(request):
    """Vista para mostrar el instructivo completo de eki."""
    return render(request, 'admin/instrucciones.html')


# ---------- Vista de importación de prospectos B2B ----------
@staff_member_required
def importar_prospectos(request):
    """Importar prospectos B2B desde archivo Excel.
    Formato: Teléfono | Nombre Contacto | Email | Empresa
    """
    import re
    context = {}

    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        if not archivo:
            context['error'] = "Por favor selecciona un archivo Excel"
            return render(request, 'admin/importar_prospectos.html', context)

        try:
            if not archivo.name.endswith(('.xlsx', '.xls')):
                context['error'] = 'El archivo debe ser .xlsx o .xls'
                return render(request, 'admin/importar_prospectos.html', context)

            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active

            creados = 0
            actualizados = 0
            errores = []

            def _normalizar_celda(val):
                if val is None:
                    return ''
                if isinstance(val, (int, float)):
                    return str(int(val)) if isinstance(val, float) and val == int(val) else str(val) if isinstance(val, float) else str(val)
                return str(val).strip()

            def _normalizar_telefono(raw):
                tel = re.sub(r'\D', '', raw)
                if tel.startswith('57') and len(tel) == 12:
                    return tel
                if len(tel) == 10 and tel.startswith('3'):
                    return '57' + tel
                if len(tel) == 7 or len(tel) == 10:
                    return '57' + tel
                return tel

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row[:2]):
                    continue

                try:
                    telefono_raw = _normalizar_celda(row[0]) if len(row) > 0 else ''
                    nombre = _normalizar_celda(row[1]) if len(row) > 1 else ''
                    email = _normalizar_celda(row[2]) if len(row) > 2 else ''
                    empresa = _normalizar_celda(row[3]) if len(row) > 3 else ''

                    if not telefono_raw:
                        errores.append(f"Fila {row_idx}: Teléfono vacío")
                        continue

                    telefono = _normalizar_telefono(telefono_raw)
                    if not telefono or len(telefono) < 10:
                        errores.append(f"Fila {row_idx}: Teléfono inválido '{telefono_raw}'")
                        continue

                    from .models import ProspectoB2B
                    prospecto, created = ProspectoB2B.objects.update_or_create(
                        telefono=telefono,
                        defaults={
                            'nombre_contacto': nombre or '',
                            'email': email or '',
                            'empresa': empresa or '',
                            'origen': 'excel',
                        }
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1

                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")

            context.update({
                'exito': True,
                'creados': creados,
                'actualizados': actualizados,
                'total': creados + actualizados,
                'advertencias': errores[:20] if errores else [],
            })
        except Exception as e:
            context['error'] = f"Error procesando archivo: {str(e)}"

    return render(request, 'admin/importar_prospectos.html', context)


# ---------- Vista de importación de estudiantes ----------
@staff_member_required
def importar_estudiantes(request):
    """Vista para importar estudiantes desde un archivo Excel.
    Campos obligatorios: Cédula | Nombre | Teléfono.
    Campos opcionales: Municipio | Departamento | Género | Edad | Curso | Cliente.
    """
    context = {}
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        
        if not archivo:
            context['error'] = "Por favor selecciona un archivo Excel"
            return render(request, 'admin/importar_estudiantes.html', context)
        
        try:
            if not archivo.name.endswith(('.xlsx', '.xls')):
                context['error'] = 'El archivo debe ser .xlsx o .xls'
                return render(request, 'admin/importar_estudiantes.html', context)
            
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            estudiantes_creados = 0
            estudiantes_actualizados = 0
            inscritos = 0
            errores = []
            
            import re
            from django.db import IntegrityError
            
            def _normalizar_celda(val):
                """Convierte celdas Excel a string limpio (int/float → str sin decimales)."""
                if val is None:
                    return ''
                if isinstance(val, float):
                    if val == int(val):
                        return str(int(val))
                    return str(val)
                if isinstance(val, int):
                    return str(val)
                return str(val).strip()
            
            def _limpiar_texto(val):
                """Limpia texto: strip, lower, elimina espacios dobles."""
                if not val:
                    return ''
                return re.sub(r'\s+', ' ', val.strip().lower())
            
            def _normalizar_telefono(raw):
                """Normaliza teléfono colombiano: solo dígitos, prefijo 57."""
                tel = re.sub(r'\D', '', raw)
                if tel.startswith('57') and len(tel) == 12:
                    return tel
                if len(tel) == 10 and tel.startswith('3'):
                    return '57' + tel
                if len(tel) == 7 or len(tel) == 10:
                    return '57' + tel
                return tel
            
            GENEROS_VALIDOS = {'m': 'M', 'f': 'F', 'o': 'O', 'masculino': 'M', 'femenino': 'F', 
                               'otro': 'O', 'hombre': 'M', 'mujer': 'F', 'nr': 'NR', 'no reporta': 'NR'}
            
            # Columnas esperadas: A=Cédula | B=Nombre | C=Teléfono | D=Municipio | E=Departamento | F=Género | G=Edad | H=Curso | I=Cliente
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row[:3]):
                    continue
                
                try:
                    cedula = _normalizar_celda(row[0]) if len(row) > 0 else ''
                    nombre = _normalizar_celda(row[1]) if len(row) > 1 else ''
                    telefono_raw = _normalizar_celda(row[2]) if len(row) > 2 else ''
                    municipio = _limpiar_texto(_normalizar_celda(row[3])) if len(row) > 3 else ''
                    departamento = _limpiar_texto(_normalizar_celda(row[4])) if len(row) > 4 else ''
                    genero_raw = _limpiar_texto(_normalizar_celda(row[5])) if len(row) > 5 else ''
                    edad_raw = _normalizar_celda(row[6]) if len(row) > 6 else ''
                    curso_nombre = _normalizar_celda(row[7]) if len(row) > 7 else ''
                    cliente_nombre = _normalizar_celda(row[8]) if len(row) > 8 else ''
                    
                    # Validar campos obligatorios mínimos (Cédula, Nombre, Teléfono)
                    campos_faltantes = []
                    if not cedula: campos_faltantes.append('Cédula')
                    if not nombre: campos_faltantes.append('Nombre')
                    if not telefono_raw: campos_faltantes.append('Teléfono')
                    
                    if campos_faltantes:
                        errores.append(f"Fila {row_idx}: Faltan campos obligatorios: {', '.join(campos_faltantes)}")
                        continue
                    
                    # Normalizar teléfono
                    telefono = _normalizar_telefono(telefono_raw)
                    if not telefono or len(telefono) < 10:
                        errores.append(f"Fila {row_idx}: Teléfono inválido '{telefono_raw}'")
                        continue
                    
                    # Normalizar género (opcional, default NR)
                    genero = GENEROS_VALIDOS.get(genero_raw, '') if genero_raw else ''
                    if not genero:
                        genero = 'NR'
                    
                    # Validar edad (opcional)
                    edad = None
                    if edad_raw:
                        try:
                            edad_str = re.sub(r'\D', '', str(edad_raw))
                            if edad_str:
                                edad = int(edad_str)
                                if edad < 1 or edad > 120:
                                    errores.append(f"Fila {row_idx}: Edad '{edad_raw}' fuera de rango (1-120)")
                                    continue
                        except (ValueError, TypeError):
                            errores.append(f"Fila {row_idx}: Edad '{edad_raw}' no es un número válido")
                            continue
                    
                    # Buscar cliente
                    cliente = None
                    if cliente_nombre:
                        try:
                            cliente = Cliente.objects.get(nombre__iexact=cliente_nombre.strip())
                        except Cliente.DoesNotExist:
                            errores.append(f"Fila {row_idx}: Cliente '{cliente_nombre}' no encontrado")
                    
                    # Crear o actualizar por CÉDULA (clave única)
                    defaults = {
                        'nombre': nombre.strip().title(),
                        'telefono': telefono,
                        'municipio': municipio,
                        'departamento': departamento,
                        'genero': genero,
                        'edad': edad,
                        'tipo_documento': 'CC',
                        'estado_onboarding': 'completado',
                        'estado_chat': 'ACTIVO',
                        'acepto_terminos': True,
                        'activo': True,
                    }
                    if cliente:
                        defaults['cliente'] = cliente
                    
                    try:
                        estudiante, creado = Estudiante.objects.update_or_create(
                            cedula=cedula,
                            defaults=defaults
                        )
                        if creado:
                            estudiantes_creados += 1
                        else:
                            estudiantes_actualizados += 1
                    except IntegrityError as e:
                        if 'telefono' in str(e).lower():
                            errores.append(f"Fila {row_idx}: Teléfono '{telefono}' ya registrado para otro estudiante")
                        else:
                            errores.append(f"Fila {row_idx}: Error de integridad - {str(e)}")
                        continue
                    
                    # Inscribir en curso si se especificó
                    if curso_nombre:
                        try:
                            curso = Curso.objects.get(nombre__iexact=curso_nombre.strip())
                            progreso, prog_creado = ProgresoEstudiante.objects.get_or_create(
                                estudiante=estudiante,
                                curso=curso,
                                defaults={'progreso': 0, 'completado': False}
                            )
                            if prog_creado:
                                inscritos += 1
                        except Curso.DoesNotExist:
                            errores.append(f"Fila {row_idx}: Curso '{curso_nombre}' no encontrado")
                
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
            
            context['exito'] = True
            context['creados'] = estudiantes_creados
            context['actualizados'] = estudiantes_actualizados
            context['inscritos'] = inscritos
            context['total'] = estudiantes_creados + estudiantes_actualizados
            
            if errores:
                context['advertencias'] = errores[:20]
        
        except Exception as e:
            context['error'] = f'Error al procesar el archivo: {str(e)}'
    
    return render(request, 'admin/importar_estudiantes.html', context)


# ---------- Vista de descarga de reportes ----------
@staff_member_required
def descargar_reportes(request):
    """Vista para descargar reportes en Excel filtrando por fechas."""
    context = {}
    
    if request.method == 'POST':
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        tipo_reporte = request.POST.get('tipo_reporte', 'todos')  # todos, envios, whatsapp
        
        try:
            # Parsear fechas
            inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d') if fecha_inicio else None
            fin = datetime.strptime(fecha_fin, '%Y-%m-%d') if fecha_fin else None
            
            # Ajustar fin de día
            if fin:
                fin = fin.replace(hour=23, minute=59, second=59)
            
            # Crear workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Eliminar hoja por defecto
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # ========== ENVÍOS ==========
            if tipo_reporte in ['todos', 'envios']:
                ws_envios = wb.create_sheet('Envíos')
                
                # Filtrar por fecha
                queryset = EnvioLog.objects.all()
                if inicio:
                    queryset = queryset.filter(fecha_envio__gte=inicio)
                if fin:
                    queryset = queryset.filter(fecha_envio__lte=fin)
                queryset = queryset.order_by('-fecha_envio')
                
                # Encabezados
                headers = ['ID', 'Estudiante', 'Teléfono', 'Campaña', 'Plantilla', 'Estado', 'Fecha', 'Respuesta API']
                ws_envios.append(headers)
                
                # Aplicar estilos a encabezados
                for cell in ws_envios[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Datos
                for log in queryset:
                    fecha_str = log.fecha_envio.strftime('%Y-%m-%d %H:%M:%S') if log.fecha_envio else ''
                    row = [
                        log.id,
                        log.estudiante.nombre,
                        log.estudiante.telefono,
                        log.campana.nombre,
                        log.campana.plantilla.nombre_interno,
                        log.estado,
                        fecha_str,
                        log.respuesta_api or ''
                    ]
                    ws_envios.append(row)
                
                # Ajustar ancho de columnas
                ws_envios.column_dimensions['A'].width = 8
                ws_envios.column_dimensions['B'].width = 20
                ws_envios.column_dimensions['C'].width = 15
                ws_envios.column_dimensions['D'].width = 20
                ws_envios.column_dimensions['E'].width = 20
                ws_envios.column_dimensions['F'].width = 12
                ws_envios.column_dimensions['G'].width = 20
                ws_envios.column_dimensions['H'].width = 30
            
            # ========== WHATSAPP ==========
            if tipo_reporte in ['todos', 'whatsapp']:
                ws_whatsapp = wb.create_sheet('WhatsApp')
                
                # Filtrar por fecha
                queryset = WhatsappLog.objects.all()
                if inicio:
                    queryset = queryset.filter(fecha__gte=inicio)
                if fin:
                    queryset = queryset.filter(fecha__lte=fin)
                queryset = queryset.order_by('-fecha')
                
                # Encabezados
                headers = ['ID', 'Teléfono', 'Tipo', 'Estado', 'Mensaje', 'Fecha', 'ID Mensaje']
                ws_whatsapp.append(headers)
                
                # Aplicar estilos a encabezados
                for cell in ws_whatsapp[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Datos
                for log in queryset:
                    fecha_str = log.fecha.strftime('%Y-%m-%d %H:%M:%S') if log.fecha else ''
                    tipo = '📥 Entrante' if log.tipo == 'INCOMING' else '📤 Saliente'
                    row = [
                        log.id,
                        log.telefono,
                        tipo,
                        log.estado,
                        log.mensaje or '',
                        fecha_str,
                        log.mensaje_id or ''
                    ]
                    ws_whatsapp.append(row)
                
                # Ajustar ancho de columnas
                ws_whatsapp.column_dimensions['A'].width = 8
                ws_whatsapp.column_dimensions['B'].width = 15
                ws_whatsapp.column_dimensions['C'].width = 15
                ws_whatsapp.column_dimensions['D'].width = 12
                ws_whatsapp.column_dimensions['E'].width = 50
                ws_whatsapp.column_dimensions['F'].width = 20
                ws_whatsapp.column_dimensions['G'].width = 25
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            response['Content-Disposition'] = f'attachment; filename="Reporte_eki_{fecha_str}.xlsx"'
            wb.save(response)
            return response
        
        except Exception as e:
            context['error'] = f"Error al generar reporte: {str(e)}"
    
    # GET: mostrar formulario
    # Calcular primer día del mes actual y último día
    hoy = datetime.now()
    primer_dia_mes = hoy.replace(day=1)
    if hoy.month == 12:
        ultimo_dia_mes = primer_dia_mes.replace(year=hoy.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        ultimo_dia_mes = primer_dia_mes.replace(month=hoy.month + 1, day=1) - timedelta(days=1)
    
    context['fecha_inicio_default'] = primer_dia_mes.strftime('%Y-%m-%d')
    context['fecha_fin_default'] = ultimo_dia_mes.strftime('%Y-%m-%d')
    
    return render(request, 'admin/descargar_reportes.html', context)


_TWILIO_STATUS_CALLBACKS = frozenset({
    'queued', 'sending', 'sent', 'delivered', 'undelivered', 'failed', 'read',
})


def _twilio_post_plano(post_data) -> dict:
    if hasattr(post_data, 'keys'):
        return {k: post_data.get(k) for k in post_data.keys()}
    return dict(post_data)


def _es_status_callback_twilio(post_data) -> bool:
    message_status = (post_data.get('MessageStatus') or post_data.get('SmsStatus') or '').lower()
    return bool(message_status and message_status in _TWILIO_STATUS_CALLBACKS)


def _encolar_twilio_edu_si_async(post_data) -> bool:
    """Encola webhook educativo en Celery si WEBHOOK_CELERY_ASYNC=true."""
    if not getattr(settings, 'WEBHOOK_CELERY_ASYNC', False):
        return False
    if _es_status_callback_twilio(post_data):
        return False
    from core.tasks import procesar_twilio_webhook_async

    procesar_twilio_webhook_async.delay(_twilio_post_plano(post_data))
    logger.info("📤 Webhook educativo encolado en Celery | sid=%s", post_data.get('MessageSid', ''))
    return True


# ---------- Webhook para WhatsApp Cloud API ----------
@csrf_exempt
def whatsapp_webhook(request):
    """
    Webhook universal para WhatsApp (Meta + Twilio)
    GET: Verificación del token
    POST: Procesa mensajes entrantes de ambos proveedores
    """
    if request.method == 'GET':
        # Verificación para Meta WhatsApp
        verify_token = request.GET.get('hub.verify_token')
        challenge = request.GET.get('hub.challenge')
        expected = getattr(settings, 'WHATSAPP_VERIFY_TOKEN', 'eki_whatsapp_verify_token_2025')
        if verify_token and expected and verify_token == expected:
            return HttpResponse(challenge)
        return HttpResponse('Forbidden', status=403)

    if request.method == 'POST':
        import sys
        print("🔵 WEBHOOK RECIBIÓ POST", flush=True)
        logger.info("🔵 WEBHOOK RECIBIÓ POST")

        def _numero_limpio(valor):
            import re
            return re.sub(r'\D', '', str(valor or ''))

        def _es_destino_bot_comercial(data):
            if bool(getattr(settings, 'BOT_COMERCIAL_FORCE_ROUTING', False)):
                return True
            to_limpio = _numero_limpio(data.get('To', ''))
            if not to_limpio:
                return False
            sandbox_number = _numero_limpio(getattr(settings, 'BOT_COMERCIAL_SANDBOX_NUMBER', '14155238886'))
            candidatos = {
                _numero_limpio(getattr(settings, 'BOT_COMERCIAL_WHATSAPP_NUMBER', '')),
            }
            candidatos.discard('')
            # En Sandbox de Twilio, forzar siempre canal comercial para no mezclar con bot educativo.
            if sandbox_number and to_limpio == sandbox_number:
                return True
            return bool(candidatos and to_limpio in candidatos)
        
        try:
            # Intentar parsear como JSON (Meta)
            payload = json.loads(request.body.decode('utf-8'))
            print(f"🔵 Payload (JSON): {payload}", flush=True)
            logger.info(f"🔵 Payload (JSON): {payload}")
            
            # Detectar si es Meta o Twilio
            if 'entry' in payload:
                # ===== META WHATSAPP =====
                logger.info("📍 Detectado: META WhatsApp")
                _procesar_meta_webhook(payload)
            else:
                # Podría ser Twilio con JSON — intentar procesarlo como Twilio también
                print("⚠️ JSON recibido pero no es Meta — intentando como Twilio", flush=True)
                logger.info("⚠️ JSON recibido pero no es Meta, verificando si tiene datos Twilio")
                # Algunos webhooks de Twilio pueden llegar como JSON
                if 'Body' in payload or 'From' in payload or 'MessageStatus' in payload:
                    print("🔵 JSON con datos Twilio detectado — procesando", flush=True)
                    logger.info(
                        "🧪 Twilio JSON inbound | path=%s | from=%s | to=%s | sid=%s",
                        request.path,
                        payload.get('From', ''),
                        payload.get('To', ''),
                        payload.get('MessageSid', ''),
                    )
                    if _es_destino_bot_comercial(payload):
                        logger.info("🧭 Router webhook: Twilio destino comercial/agro detectado (JSON)")
                        _procesar_bot_comercial_twilio_webhook(payload)
                    else:
                        logger.info("🧭 Router webhook: Twilio destino educativo detectado (JSON)")
                        if not _encolar_twilio_edu_si_async(payload):
                            _procesar_twilio_webhook(payload)
                else:
                    print("⚠️ JSON desconocido — ignorando", flush=True)
                return HttpResponse('OK')
                
        except json.JSONDecodeError:
            # Podría ser Twilio (form-data)
            print("🔵 Payload (Form-Data) - Probablemente Twilio", flush=True)
            print(f"POST keys: {list(request.POST.keys())}", flush=True)
            logger.info("🔵 Payload (Form-Data) - Probablemente Twilio")
            logger.info(
                "🧪 Twilio Form inbound | path=%s | from=%s | to=%s | sid=%s",
                request.path,
                request.POST.get('From', ''),
                request.POST.get('To', ''),
                request.POST.get('MessageSid', ''),
            )
            if _es_destino_bot_comercial(request.POST):
                logger.info("🧭 Router webhook: Twilio destino comercial/agro detectado (Form-Data)")
                _procesar_bot_comercial_twilio_webhook(request.POST)
                twilio_result = None
            else:
                logger.info("🧭 Router webhook: Twilio destino educativo detectado (Form-Data)")
                if _encolar_twilio_edu_si_async(request.POST):
                    twilio_result = None
                else:
                    twilio_result = _procesar_twilio_webhook(request.POST)
            # Si _procesar_twilio_webhook devuelve TwiML HttpResponse, retornarlo
            if isinstance(twilio_result, HttpResponse):
                return twilio_result
        
        except Exception as e:
            print(f"❌ Error en webhook: {str(e)}", flush=True)
            logger.error(f"❌ Error en webhook: {str(e)}")
            import traceback
            traceback.print_exc()
            return HttpResponse('Error', status=500)

        return HttpResponse('OK')


def _cliente_en_ventana(cliente, campo_habilitar, campo_inicio, campo_fin):
    """Evalúa si una funcionalidad está habilitada para un cliente según ventana de fechas."""
    if not cliente:
        return True
    if not getattr(cliente, campo_habilitar, False):
        return False

    hoy = timezone.localdate()
    inicio = getattr(cliente, campo_inicio, None)
    fin = getattr(cliente, campo_fin, None)

    if inicio and hoy < inicio:
        return False
    if fin and hoy > fin:
        return False
    return True


def _cliente_habilita_pregunta_abierta_final(cliente):
    return _cliente_en_ventana(
        cliente,
        'habilitar_pregunta_abierta_final',
        'fecha_inicio_pregunta_abierta_final',
        'fecha_fin_pregunta_abierta_final',
    )


def _cliente_habilita_proximidad(cliente):
    habilitado_legacy = _cliente_en_ventana(
        cliente,
        'habilitar_gamificacion_proximidad',
        'fecha_inicio_gamificacion_proximidad',
        'fecha_fin_gamificacion_proximidad',
    )
    return bool(habilitado_legacy or (cliente and getattr(cliente, 'empleabilidad_exploracion_activa', False)))


def _bot_comercial_sin_contexto_natural(pregunta: str, cliente=None) -> str:
    """Respuestas cuando aún no hay contexto RAG suficiente (tono formal agrónomo)."""
    from core.nati import obtener_nombre_bot

    nombre = obtener_nombre_bot(cliente)
    q = (pregunta or '').strip().lower()

    saludos = {'hola', 'buenas', 'buenos dias', 'buen día', 'buenas tardes', 'buenas noches', 'hey'}
    if q in saludos or any(s in q for s in ['hola', 'buenas', 'buen dia', 'buen día']):
        return (
            f"Buenos días. Soy {nombre}, agrónoma virtual de eki.\n\n"
            "Le acompaño en manejo técnico de su cultivo y, si usted lo solicita, "
            "en orientación de catálogo con base en información oficial.\n\n"
            "Indíqueme, por favor, su cultivo y qué necesita resolver."
        )

    if 'gulupa' in q or 'gulupa' in q:
        return (
            "Perfecto, trabajemos *gulupa*.\n\n"
            "Puedo orientarle en suelo, pH, altitud, nutrición y manejo sanitario "
            "con la información técnica disponible.\n\n"
            "Para una respuesta precisa, indíqueme el punto puntual que desea resolver primero."
        )

    return (
        "Entendido. Vamos a resolverlo de forma técnica y práctica.\n\n"
        "Indíqueme cultivo y objetivo puntual (suelo, nutrición, plaga o enfermedad) "
        "y le responderé con base en la información oficial disponible."
    )


def _extraer_texto_archivo_simple(ruta_archivo: str, *, xlsx_max_rows=None) -> str:
    """Extracción liviana de texto para fallback cuando RAG vectorial no retorna contexto.

    xlsx_max_rows: en Excel limita filas extraídas (evita bloquear Gunicorn con hojas enormes).
    """
    import os

    ext = os.path.splitext(ruta_archivo)[1].lower()
    try:
        if ext == '.txt':
            with open(ruta_archivo, 'r', encoding='utf-8', errors='ignore') as f:
                t = f.read()
                return t[:120000]
        if ext == '.pdf':
            from PyPDF2 import PdfReader

            reader = PdfReader(ruta_archivo)
            parts = []
            for i, p in enumerate(reader.pages):
                if i >= 45:
                    break
                parts.append(p.extract_text() or "")
            return "\n".join(parts)[:120000]
        if ext == '.docx':
            from docx import Document

            doc = Document(ruta_archivo)
            return "\n".join(p.text for p in doc.paragraphs)[:80000]
        if ext in ('.xlsx', '.xlsm'):
            from core.rag_eki_multitenant import RAGClienteCurso

            mr = xlsx_max_rows if xlsx_max_rows is not None else None
            return (RAGClienteCurso._extraer_xlsx(ruta_archivo, max_rows=mr) or '').strip()[:120000]
    except Exception:
        return ""
    return ""


def _contexto_fallback_desde_documentos(
    cliente_ids: list,
    pregunta: str,
    max_chars: int = 1800,
    *,
    max_docs: int = 8,
    xlsx_max_rows: int | None = 3200,
) -> str:
    """Fallback semántico simple sobre documentos indexados para evitar respuestas vacías."""
    from django.db.models import Q
    from .models import DocumentoRAGComercial
    import re

    tokens = [t for t in re.findall(r"[a-zA-ZáéíóúñÁÉÍÓÚÑ0-9]{3,}", (pregunta or '').lower())][:12]

    p_q = (pregunta or '').lower()
    consulta_catalogo = bool(
        re.search(
            r'precio|precios|cotiz|lista|tarifa|valor|cu[aá]nto|cuesta|insumo|producto|'
            r'cat[aá]logo|bulto|arroba|\bkg\b|kilo|dosis|paquete|mezcla|fertil|herbic|fungic',
            p_q,
        )
    )

    base_qs = DocumentoRAGComercial.objects.filter(estado='indexado').filter(
        Q(cliente_id__in=cliente_ids) | Q(cliente__isnull=True)
    )
    if consulta_catalogo:
        from django.db.models import Case, IntegerField, When

        qs = base_qs.annotate(
            _catprio=Case(
                When(cliente__isnull=True, then=0),
                default=1,
                output_field=IntegerField(),
            )
        ).order_by('_catprio', '-fecha_indexado', '-fecha_subida')[: max(4, max_docs + 4)]
    else:
        qs = base_qs.order_by('-fecha_indexado', '-fecha_subida')[: max(4, max_docs + 2)]

    fragmentos = []
    chars_total = 0
    docs_procesados = 0
    ya_extracto_general = False
    for doc in qs:
        if docs_procesados >= max_docs:
            break
        try:
            ruta = doc.archivo.path
        except Exception:
            ruta = doc._descargar_temp()
        if not ruta:
            continue

        texto = _extraer_texto_archivo_simple(ruta, xlsx_max_rows=xlsx_max_rows)
        docs_procesados += 1
        if not texto:
            continue

        texto_norm = texto.lower()
        score = sum(1 for tk in tokens if tk in texto_norm)
        usar_extracto_inicial = (
            consulta_catalogo
            and doc.cliente_id is None
            and score <= 0
            and not ya_extracto_general
        )
        if score <= 0 and tokens and not usar_extracto_inicial:
            continue

        pos = 0
        if score > 0:
            for tk in tokens:
                p = texto_norm.find(tk)
                if p >= 0:
                    pos = p
                    break
        if usar_extracto_inicial:
            ini, fin = 0, min(len(texto), 1600)
            ya_extracto_general = True
        else:
            ini = max(0, pos - 260)
            fin = min(len(texto), pos + 640)
        snippet = texto[ini:fin].strip().replace('\x00', ' ')
        if not snippet:
            continue

        bloque = f"[Fuente: {doc.nombre}]\n{snippet}"
        if chars_total + len(bloque) > max_chars:
            break
        fragmentos.append(bloque)
        chars_total += len(bloque)

    if not fragmentos:
        return ""

    return (
        "\n\n📚 CONTEXTO DOCUMENTAL (fallback por archivo indexado):\n"
        + "\n---\n".join(fragmentos)
        + "\n\n⚠️ Usa esta información como base técnica prioritaria."
    )


def _normalizar_consulta_web(pregunta: str) -> str:
    """Normaliza consulta para búsquedas web/académicas de respaldo."""
    import re

    tokens = re.findall(r"[a-zA-ZáéíóúñÁÉÍÓÚÑ0-9]{3,}", (pregunta or '').lower())
    stop = {
        'para', 'como', 'donde', 'cuando', 'cuanto', 'cuantos', 'cual', 'cuales',
        'tengo', 'necesito', 'quiero', 'sobre', 'desde', 'hasta', 'entre', 'esta',
        'este', 'estos', 'estas', 'porque', 'favor', 'hola', 'buenas', 'dias',
        'tarde', 'noche', 'gracias', 'cultivo', 'agricola', 'agricultura',
    }
    claves = [t for t in tokens if t not in stop]
    if not claves:
        return (pregunta or '').strip()[:120]
    return ' '.join(claves[:10]).strip()


def _contexto_fallback_web_agro(pregunta: str, max_chars: int = 1800) -> str:
    """Busca referencias externas cuando aún no hay contexto RAG suficiente."""
    if not bool(getattr(settings, 'BOT_COMERCIAL_WEB_FALLBACK_ENABLED', True)):
        return ""

    try:
        timeout = float(getattr(settings, 'BOT_COMERCIAL_WEB_FALLBACK_TIMEOUT', 6.0) or 6.0)
    except Exception:
        timeout = 6.0
    try:
        max_fuentes = int(getattr(settings, 'BOT_COMERCIAL_WEB_FALLBACK_MAX_FUENTES', 4) or 4)
    except Exception:
        max_fuentes = 4

    query_base = _normalizar_consulta_web(pregunta)
    if not query_base:
        return ""

    import re
    from html import unescape
    from urllib.parse import quote_plus

    fuentes = []
    vistos = set()

    # 1) Base académica abierta (Crossref) como aproximación a literatura técnica.
    try:
        r = requests.get(
            'https://api.crossref.org/works',
            params={
                'query': query_base,
                'rows': max_fuentes,
                'sort': 'relevance',
                'order': 'desc',
                'select': 'title,container-title,published-print,published-online,created,DOI,URL,abstract',
            },
            timeout=timeout,
            headers={'User-Agent': 'eki-bot-comercial/1.0'},
        )
        r.raise_for_status()
        payload = r.json() or {}
        for item in (payload.get('message') or {}).get('items', [])[:max_fuentes]:
            titulo = ((item.get('title') or [''])[0] or '').strip()
            if not titulo:
                continue

            anio = ''
            for k in ['published-print', 'published-online', 'created']:
                dp = ((item.get(k) or {}).get('date-parts') or [])
                if dp and dp[0]:
                    anio = str(dp[0][0])
                    break

            revista = ((item.get('container-title') or [''])[0] or '').strip()
            doi = (item.get('DOI') or '').strip()
            url = (item.get('URL') or '').strip()
            resumen = re.sub(r'<[^>]+>', ' ', unescape((item.get('abstract') or '').strip()))
            resumen = re.sub(r'\s+', ' ', resumen).strip()[:280]

            clave = (titulo.lower(), url.lower())
            if clave in vistos:
                continue
            vistos.add(clave)

            fuentes.append({
                'origen': 'Academico/Crossref',
                'titulo': titulo,
                'resumen': resumen or f"Referencia técnica {anio or ''} {revista}".strip(),
                'url': url or (f"https://doi.org/{doi}" if doi else ''),
            })
    except Exception as e:
        logger.info("🌐 Fallback académico no disponible: %s", e)

    # 2) Internet general: DuckDuckGo Instant Answer.
    try:
        r = requests.get(
            'https://api.duckduckgo.com/',
            params={
                'q': f"{query_base} manejo agronomico",
                'format': 'json',
                'no_html': 1,
                'skip_disambig': 1,
            },
            timeout=timeout,
            headers={'User-Agent': 'eki-bot-comercial/1.0'},
        )
        r.raise_for_status()
        data = r.json() or {}

        if data.get('AbstractText') and data.get('AbstractURL'):
            titulo_abs = (data.get('Heading') or 'Resumen web técnico').strip()
            clave = (titulo_abs.lower(), (data.get('AbstractURL') or '').lower())
            if clave not in vistos:
                vistos.add(clave)
                fuentes.append({
                    'origen': 'Internet/DuckDuckGo',
                    'titulo': titulo_abs,
                    'resumen': str(data.get('AbstractText') or '').strip()[:280],
                    'url': str(data.get('AbstractURL') or '').strip(),
                })

        def _iter_topics(items):
            for it in items or []:
                if isinstance(it, dict) and 'Text' in it and 'FirstURL' in it:
                    yield it
                for sub in (it.get('Topics') if isinstance(it, dict) else []) or []:
                    if isinstance(sub, dict) and 'Text' in sub and 'FirstURL' in sub:
                        yield sub

        for it in _iter_topics(data.get('RelatedTopics')):
            titulo = str(it.get('Text') or '').strip()
            url = str(it.get('FirstURL') or '').strip()
            if not titulo or not url:
                continue
            clave = (titulo.lower(), url.lower())
            if clave in vistos:
                continue
            vistos.add(clave)
            fuentes.append({
                'origen': 'Internet/DuckDuckGo',
                'titulo': titulo[:160],
                'resumen': titulo[:280],
                'url': url,
            })
            if len(fuentes) >= max_fuentes * 2:
                break
    except Exception as e:
        logger.info("🌐 Fallback internet no disponible: %s", e)

    # 3) Respaldo Wikipedia ES cuando no hay suficiente info externa.
    if len(fuentes) < max_fuentes:
        try:
            r = requests.get(
                'https://es.wikipedia.org/w/api.php',
                params={
                    'action': 'query',
                    'list': 'search',
                    'srsearch': f"{query_base} agricultura",
                    'srlimit': max_fuentes,
                    'format': 'json',
                    'utf8': 1,
                },
                timeout=timeout,
                headers={'User-Agent': 'eki-bot-comercial/1.0'},
            )
            r.raise_for_status()
            data = r.json() or {}
            for item in ((data.get('query') or {}).get('search') or []):
                titulo = str(item.get('title') or '').strip()
                snippet = re.sub(r'<[^>]+>', ' ', unescape(str(item.get('snippet') or '')))
                snippet = re.sub(r'\s+', ' ', snippet).strip()[:280]
                if not titulo:
                    continue
                url = f"https://es.wikipedia.org/wiki/{quote_plus(titulo.replace(' ', '_'))}"
                clave = (titulo.lower(), url.lower())
                if clave in vistos:
                    continue
                vistos.add(clave)
                fuentes.append({
                    'origen': 'Internet/Wikipedia',
                    'titulo': titulo,
                    'resumen': snippet or 'Referencia general de agricultura.',
                    'url': url,
                })
                if len(fuentes) >= max_fuentes * 2:
                    break
        except Exception as e:
            logger.info("🌐 Fallback Wikipedia no disponible: %s", e)

    if not fuentes:
        return ""

    bloques = []
    chars_total = 0
    for idx, f in enumerate(fuentes[: max_fuentes * 2], start=1):
        bloque = (
            f"[Fuente externa {idx} | {f['origen']}]\n"
            f"{f['titulo']}\n"
            f"{f['resumen']}\n"
            f"URL: {f['url'] or 'N/D'}"
        )
        if chars_total + len(bloque) > max_chars:
            break
        bloques.append(bloque)
        chars_total += len(bloque)

    if not bloques:
        return ""

    return (
        "\n\n🌐 INFORMACION COMPLEMENTARIA DE WEB (solo si la oficial de eki no alcanza):\n"
        + "\n---\n".join(bloques)
        + "\n\n⚠️ Si hay información oficial de eki, esa SIEMPRE tiene prioridad sobre estas fuentes externas."
    )


def _bot_comercial_respuesta_catalogo(
    pregunta: str,
    contexto_rag: str,
    diagnostico_vision: str = '',
    contexto_web: str = '',
    historial_chat: str = '',
    cliente=None,
    sesion_comercial=None,
    bloque_contexto_agro: str = '',
    routing=None,
    rag_chunks=None,
    ctx_agro=None,
) -> str:
    """Genera respuesta técnica/comercial estricta basada en contexto RAG (sin alucinaciones).

    Si se pasa `cliente` (instancia de `core.Cliente`), se usa su `system_prompt_extra`
    y `nombre_bot` para personalizar la identidad de Nat.
    """
    api_key = getattr(settings, 'OPENAI_API_KEY', '')
    contexto_base = contexto_rag or contexto_web
    tiene_rag = bool(contexto_rag)
    if not api_key:
        if not contexto_base:
            return _bot_comercial_sin_contexto_natural(pregunta, cliente=cliente)

        if tiene_rag:
            encabezado = "Con base en la información oficial de eki:"
            cierre = "Si aplica cotización, compárteme cantidad, municipio y cultivo."
        else:
            encabezado = (
                "Aún no tengo información oficial de eki suficiente para resolver eso. "
                "Le comparto un respaldo técnico general:"
            )
            cierre = (
                "Si ya tiene ficha técnica o precios del producto, le pido que los hagamos "
                "llegar al equipo de eki para darle una recomendación exacta por cultivo y necesidad."
            )

        return (
            f"{encabezado}\n\n"
            f"{contexto_base[:1200]}\n\n"
            f"{cierre}"
        )

    try:
        from openai import OpenAI
        from core.nati import armar_instruccion_modo, armar_messages_para_openai, armar_system_prompt
        from core.nat_router import decidir_routing_nat
        import time
        inicio = time.time()
        client = OpenAI(api_key=api_key)

        if routing is None:
            routing = decidir_routing_nat(
                pregunta,
                rag_chunks=rag_chunks or [],
                tiene_rag_texto=bool(contexto_rag),
                contexto_rag_chars=len(contexto_rag or ''),
                ctx_agro=ctx_agro,
                diagnostico_vision=diagnostico_vision,
            )

        system_prompt = armar_system_prompt(cliente=cliente)
        bloque_agro = (bloque_contexto_agro or '').strip()
        bloque_ctx_prompt = (
            f"CONTEXTO AGRONÓMICO DEL PRODUCTOR (estructurado):\n{bloque_agro}\n\n"
            if bloque_agro else ''
        )
        bloque_modo = armar_instruccion_modo(routing.modo, routing.escala_premium)
        bloque_consulta = (
            f"CONSULTA DEL PRODUCTOR:\n{pregunta}\n\n"
            f"{bloque_ctx_prompt}"
            f"{bloque_modo}"
            f"DIAGNÓSTICO VISIÓN (si aplica):\n{diagnostico_vision or 'N/A'}\n\n"
            f"INFORMACIÓN OFICIAL DE EKI (fuente principal — use esto con precisión):\n{contexto_rag or '[VACIO]'}\n\n"
            f"INFORMACIÓN COMPLEMENTARIA WEB (solo si la oficial no alcanza):\n{contexto_web or '[VACIO]'}\n\n"
            "Recuerde: nunca mencione al productor términos como RAG, base de "
            "conocimiento, fragmento o documento indexado. Hable como agrónoma formal de eki (siempre de usted).\n"
            "Si la consulta parece error de tipeo, ofrezca 1–2 interpretaciones plausibles "
            "con '¿Quiso decir...?' antes de conclusiones fuertes.\n"
            "Si INFORMACIÓN OFICIAL incluye listas Excel (producto, precio, dosis), "
            "use solo esas cifras; si no aparecen, no las invente."
        )
        if sesion_comercial is not None:
            # Memoria en SesionComercial; no duplicar WhatsappLog en el prompt.
            messages = armar_messages_para_openai(
                sesion=sesion_comercial,
                nuevo_mensaje=bloque_consulta,
                cliente=cliente,
            )
        else:
            user_prompt = bloque_consulta
            if historial_chat:
                user_prompt = (
                    f"{bloque_consulta}\n\n"
                    f"HISTORIAL RECIENTE (referencia):\n{historial_chat}"
                )
            messages = [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': user_prompt},
            ]
        modelo = routing.modelo
        temperatura = 0.1 if routing.escala_premium else 0.15
        try:
            max_out = int(getattr(settings, 'BOT_COMERCIAL_OPENAI_MAX_TOKENS', 650) or 650)
        except (TypeError, ValueError):
            max_out = 650
        max_out = max(400, min(max_out, 1200))
        from core.openai_compat import chat_completion_token_kwargs
        completion = client.chat.completions.create(
            model=modelo,
            messages=messages,
            **chat_completion_token_kwargs(modelo, max_out, temperatura),
        )
        texto = (completion.choices[0].message.content or '').strip()
        if not texto:
            # Segundo intento: effort mínimo si el modelo gastó el cupo en reasoning
            completion = client.chat.completions.create(
                model=modelo,
                messages=messages,
                **chat_completion_token_kwargs(
                    modelo, max(max_out, 900), temperatura, reasoning_effort='minimal',
                ),
            )
            texto = (completion.choices[0].message.content or '').strip()
        latencia_ms = int((time.time() - inicio) * 1000)
        usage = getattr(completion, 'usage', None)
        try:
            from core.eventos_ia import emit_ia_agent_triggered

            emit_ia_agent_triggered(
                cliente=cliente,
                agente='nati',
                mensaje=pregunta,
                respuesta=texto,
                modelo=modelo,
                latencia_ms=latencia_ms,
                tokens_in=getattr(usage, 'prompt_tokens', None) if usage else None,
                tokens_out=getattr(usage, 'completion_tokens', None) if usage else None,
                canal='whatsapp_comercial',
                metadata={
                    'tiene_rag': bool(contexto_rag),
                    'tiene_web': bool(contexto_web),
                    'routing_modo': routing.modo,
                    'routing_razon': routing.razon,
                    'escala_premium': routing.escala_premium,
                    'rag_max_similitud': routing.rag_max_similitud,
                },
            )
        except Exception:
            pass
        return texto or "No logré construir una respuesta válida. Intenta con otra consulta."
    except Exception as e:
        logger.warning(f"⚠️ Bot Comercial LLM fallback: {e}")
        if contexto_rag:
            return (
                "Le comparto lo que encontré en la información oficial de eki:\n\n"
                f"{contexto_rag[:1000]}\n\n"
                "Para cotizar, indíqueme cantidad y municipio."
            )
        if contexto_web:
            return (
                "Aún no tengo suficiente contenido indexado para responder con exactitud. "
                "Encontré estas referencias externas de apoyo:\n\n"
                f"{contexto_web[:1000]}\n\n"
                "Si sube fichas técnicas o precios al equipo eki, podré darle una recomendación directa."
            )
        return _bot_comercial_sin_contexto_natural(pregunta, cliente=cliente)


def _bot_comercial_historial_reciente(telefono: str, max_turnos: int = 6, max_chars: int = 1200) -> str:
    """Construye memoria corta de conversación comercial desde WhatsappLog."""
    if not telefono:
        return ''

    try:
        max_turnos = int(max_turnos or 6)
    except (TypeError, ValueError):
        max_turnos = 6
    max_turnos = max(2, min(max_turnos, 12))

    try:
        max_chars = int(max_chars or 1200)
    except (TypeError, ValueError):
        max_chars = 1200
    max_chars = max(300, min(max_chars, 2500))

    logs = list(
        WhatsappLog.objects.filter(
            telefono=telefono,
            agente_usado='BOT_COMERCIAL',
        )
        .exclude(mensaje__isnull=True)
        .exclude(mensaje='')
        .order_by('-fecha')[: (max_turnos * 3)]
    )

    if not logs:
        return ''

    import re

    lineas = []
    for log in reversed(logs):
        texto = re.sub(r'\s+', ' ', str(log.mensaje or '').strip())
        if not texto:
            continue
        if texto.startswith('[MEDIA:'):
            continue

        if len(texto) > 260:
            texto = f"{texto[:257]}..."

        tipo = str(log.tipo or '').upper()
        rol = 'Cliente' if tipo == 'INCOMING' else 'Bot'
        lineas.append(f"{rol}: {texto}")

    if not lineas:
        return ''

    # Recorta por caracteres conservando las líneas más recientes.
    salida = []
    total = 0
    for linea in reversed(lineas):
        costo = len(linea) + 1
        if total + costo > max_chars:
            break
        salida.append(linea)
        total += costo

    salida.reverse()
    return '\n'.join(salida)


def _obtener_o_crear_sesion_comercial(telefono: str, cliente=None, horas_expira: int = 4):
    from django.utils import timezone as _tz
    from .models import SesionComercial

    telefono = (telefono or '').strip()
    if not telefono:
        return None
    ahora = _tz.now()
    delta = timedelta(hours=max(1, int(horas_expira or 4)))
    sesion = (
        SesionComercial.objects.filter(telefono=telefono)
        .order_by('-fecha_ultimo_mensaje')
        .first()
    )
    if not sesion:
        return SesionComercial.objects.create(telefono=telefono, cliente=cliente)

    if sesion.fecha_ultimo_mensaje and sesion.fecha_ultimo_mensaje < (ahora - delta):
        sesion.historial_mensajes = []
        sesion.cliente = cliente
        sesion.save(update_fields=['historial_mensajes', 'cliente', 'fecha_ultimo_mensaje'])
    elif cliente and sesion.cliente_id != getattr(cliente, 'id', None):
        sesion.cliente = cliente
        sesion.save(update_fields=['cliente', 'fecha_ultimo_mensaje'])
    return sesion


def _actualizar_sesion_comercial(sesion, mensaje_usuario: str, respuesta_bot: str):
    if sesion is None:
        return
    historial = list(sesion.historial_mensajes or [])
    historial.append({'role': 'user', 'content': (mensaje_usuario or '')[:3000]})
    historial.append({'role': 'assistant', 'content': (respuesta_bot or '')[:3000]})
    if len(historial) > 20:
        historial = historial[-20:]
    sesion.historial_mensajes = historial
    sesion.save(update_fields=['historial_mensajes', 'fecha_ultimo_mensaje'])


def _bot_comercial_diagnosticar_imagen(media_url: str, media_type: str, cliente=None) -> str:
    """Diagnóstico preliminar de imagen de cultivo usando visión (si está disponible)."""
    if not media_url or not media_type.startswith('image'):
        return ''

    from core.ai_capabilities import resolver_ai_capability

    if not resolver_ai_capability('diagnostico_agro', cliente=cliente):
        return (
            "Recibí su imagen. El diagnóstico visual automático no está habilitado "
            "para su organización; describa los síntomas y el cultivo, por favor."
        )

    api_key = getattr(settings, 'OPENAI_API_KEY', '')
    if not api_key:
        return "Imagen recibida. Diagnóstico visual no disponible en este entorno."

    try:
        from openai import OpenAI
        from core.openai_compat import chat_completion_token_kwargs
        client = OpenAI(api_key=api_key)
        vision_model = getattr(settings, 'BOT_COMERCIAL_VISION_MODEL', 'gpt-4o-mini')
        resp = client.chat.completions.create(
            model=vision_model,
            messages=[
                {
                    'role': 'system',
                    'content': (
                        "Usted es agrónoma de campo en Colombia. Entregue un diagnóstico preliminar "
                        "breve y prudente, tratando al productor de usted. No afirme certeza absoluta. "
                        "Indique posible plaga, enfermedad o estrés y síntomas observados."
                    ),
                },
                {
                    'role': 'user',
                    'content': [
                        {'type': 'text', 'text': 'Analiza esta imagen de cultivo y da un diagnóstico preliminar.'},
                        {'type': 'image_url', 'image_url': {'url': media_url}},
                    ],
                },
            ],
            **chat_completion_token_kwargs(vision_model, 280, 0.2),
        )
        return (resp.choices[0].message.content or '').strip()
    except Exception as e:
        logger.warning(f"⚠️ Bot Comercial visión no disponible: {e}")
        return (
            "Recibí su imagen, pero no pude completar el diagnóstico visual en este momento. "
            "Describa los síntomas y el cultivo, por favor."
        )


def _procesar_bot_comercial_twilio_webhook(post_data, forzar_canal=False):
    """Webhook Twilio dedicado para bot comercial (texto libre, voz e imagen)."""
    from .rag_comercial_manager import rag_comercial_manager

    status = post_data.get('MessageStatus', post_data.get('SmsStatus', ''))
    if status and status.lower() in ['queued', 'sending', 'sent', 'delivered', 'undelivered', 'failed', 'read']:
        _registrar_estado_twilio_callback(post_data)
        return

    msg_body = (post_data.get('Body', '') or '').strip()
    msg_from = post_data.get('From', '')
    msg_to = post_data.get('To', '')
    from_number_respuesta = msg_to
    msg_sid = post_data.get('MessageSid', f'botcom_{timezone.now().timestamp()}')
    num_media = int(post_data.get('NumMedia', 0) or 0)
    media_type = post_data.get('MediaContentType0', '') or ''
    media_url = post_data.get('MediaUrl0', '') or ''

    # Evita doble respuesta si Twilio reintenta el mismo webhook.
    if msg_sid and WhatsappLog.objects.filter(
        mensaje_id=msg_sid,
        tipo='INCOMING',
        agente_usado='BOT_COMERCIAL',
    ).exists():
        logger.info("♻️ Bot comercial: webhook duplicado ignorado | sid=%s", msg_sid)
        return

    from core.eventos_ia import set_trace_id
    set_trace_id()

    if msg_from.startswith('whatsapp:'):
        msg_from = msg_from.replace('whatsapp:', '')
    if msg_to.startswith('whatsapp:'):
        msg_to = msg_to.replace('whatsapp:', '')
    import re
    telefono_limpio = re.sub(r'\D', '', msg_from)
    to_limpio = re.sub(r'\D', '', msg_to)
    if len(telefono_limpio) == 10:
        telefono_limpio = f"57{telefono_limpio}"

    numero_bot_comercial = re.sub(
        r'\D',
        '',
        str(getattr(settings, 'BOT_COMERCIAL_WHATSAPP_NUMBER', '') or ''),
    )
    if (not forzar_canal) and numero_bot_comercial and to_limpio and to_limpio != numero_bot_comercial:
        logger.info(
            "Webhook bot comercial ignorado por número destino distinto: to=%s esperado=%s",
            to_limpio,
            numero_bot_comercial,
        )
        return

    if num_media > 0 and ('audio' in media_type or 'ogg' in media_type):
        transcripcion = _transcribir_audio_twilio(media_url, media_type=media_type)
        if transcripcion:
            msg_body = transcripcion
        elif not msg_body:
            msg_body = '[AUDIO_NO_TRANSCRITO]'

    if not msg_body and num_media == 0:
        logger.info("ℹ️ Bot comercial: webhook sin contenido ignorado | sid=%s", msg_sid)
        return

    WhatsappLog.objects.create(
        telefono=telefono_limpio,
        mensaje=msg_body or f'[MEDIA:{media_type}]',
        mensaje_id=msg_sid,
        tipo='INCOMING',
        es_audio=('audio' in media_type),
        agente_usado='BOT_COMERCIAL',
    )

    try:
        from core.eventos_ia import emit_webhook_recibido
        from core.ai_capabilities import resolver_ai_capability

        if resolver_ai_capability('eventos_ia'):
            emit_webhook_recibido(
                mensaje=msg_body or f'[MEDIA:{media_type}]',
                telefono=telefono_limpio,
                canal='whatsapp_comercial',
            )
    except Exception:
        pass

    msg_normalizado = re.sub(r'\s+', ' ', (msg_body or '').strip().lower())
    memoria_turnos = int(getattr(settings, 'BOT_COMERCIAL_MEMORY_TURNOS', 12) or 12)
    memoria_chars = int(getattr(settings, 'BOT_COMERCIAL_MEMORY_MAX_CHARS', 3600) or 3600)
    historial_chat = _bot_comercial_historial_reciente(
        telefono=telefono_limpio,
        max_turnos=memoria_turnos,
        max_chars=memoria_chars,
    )
    if historial_chat:
        logger.info(
            "🧠 Memoria chat comercial aplicada | telefono=%s | chars=%s",
            telefono_limpio,
            len(historial_chat),
        )

    es_saludo = bool(
        re.match(
            r'^(hola|buenas|buenos dias|buen día|buenas tardes|buenas noches|hey|que tal|qué tal)\b',
            msg_normalizado,
        )
    )

    cliente_id_cfg = int(
        getattr(settings, 'BOT_COMERCIAL_CLIENTE_ID', 0) or 0
    )
    canal_rag = str(getattr(settings, 'BOT_COMERCIAL_RAG_CANAL', 'bot_comercial') or 'bot_comercial')

    from core.nati import armar_saludo_inicial, armar_saludo_menu, armar_cliente_ids_rag, resolver_cliente_desde_numero_whatsapp

    cliente_nati = resolver_cliente_desde_numero_whatsapp(msg_to)
    if not cliente_nati and cliente_id_cfg:
        try:
            cliente_nati = Cliente.objects.filter(id=cliente_id_cfg, activo=True).first()
        except Exception as e:
            logger.warning(
                "Bot comercial: no se pudo cargar Cliente id=%s para Nat: %s",
                cliente_id_cfg, e,
            )

    sesion_comercial = _obtener_o_crear_sesion_comercial(
        telefono=telefono_limpio,
        cliente=cliente_nati,
        horas_expira=int(getattr(settings, 'BOT_COMERCIAL_SESSION_HOURS', 4) or 4),
    )

    consulta = msg_body or 'Necesito asesoría agrícola'
    rag_chunks = []

    ctx_agro = None
    bloque_contexto_agro = ''
    try:
        from core.ai_capabilities import resolver_ai_capability
        from core.contexto_agro import actualizar_contexto_desde_mensaje, formatear_bloque_contexto_para_prompt

        if resolver_ai_capability('nati_structured_context', cliente=cliente_nati):
            ctx_agro = actualizar_contexto_desde_mensaje(sesion_comercial, msg_body)
            bloque_contexto_agro = formatear_bloque_contexto_para_prompt(ctx_agro)
    except Exception:
        pass

    routing = None
    if es_saludo:
        texto_respuesta = armar_saludo_inicial(cliente_nati)
    elif msg_normalizado in ['listo', 'continuar', 'menu', 'menú']:
        texto_respuesta = armar_saludo_menu(cliente_nati)
    else:
        diagnostico_vision = ''
        if num_media > 0 and media_type.startswith('image') and media_url:
            diagnostico_vision = _bot_comercial_diagnosticar_imagen(
                media_url, media_type, cliente=cliente_nati,
            )

        consulta = msg_body or 'Necesito asesoría agrícola'
        if diagnostico_vision:
            consulta = f"{consulta}\n\nDiagnóstico preliminar imagen: {diagnostico_vision}"

        texto_respuesta = None
        try:
            from core.nat_diagnostico import siguiente_pregunta_diagnostico

            pregunta_diag = siguiente_pregunta_diagnostico(
                ctx_agro,
                msg_body,
                tiene_imagen=bool(diagnostico_vision),
            )
            if pregunta_diag:
                texto_respuesta = pregunta_diag
        except Exception:
            pass

        rag_chunks: list = []
        routing = None
        if texto_respuesta is None:
            cliente_ids_consulta = armar_cliente_ids_rag(cliente_nati)

            contexto_rag = ''
            contexto_web = ''

            from core.catalogo_precios import (
                buscar_precios,
                es_consulta_catalogo,
                formatear_contexto_precios,
            )
            contexto_precios_db = ''
            if es_consulta_catalogo(consulta) and cliente_ids_consulta:
                productos_precio = buscar_precios(
                    [i for i in cliente_ids_consulta if i != 0] or cliente_ids_consulta,
                    consulta,
                )
                if productos_precio:
                    contexto_precios_db = formatear_contexto_precios(productos_precio)
                    logger.info(
                        "💰 Precios Postgres | hits=%s | clientes=%s",
                        len(productos_precio),
                        cliente_ids_consulta,
                    )

            if rag_comercial_manager.disponible:
                canales_consulta = []
                for c in [canal_rag, 'bot_comercial']:
                    if c and c not in canales_consulta:
                        canales_consulta.append(c)

                rag_max = int(getattr(settings, 'BOT_COMERCIAL_RAG_MAX_CHARS', 2500) or 2500)
                rag_max = max(400, min(rag_max, 4000))
                try:
                    top_k = int(getattr(settings, 'BOT_COMERCIAL_RAG_TOP_K', 9) or 9)
                except (TypeError, ValueError):
                    top_k = 9
                top_k = max(3, min(top_k, 20))

                for canal in canales_consulta:
                    rag_result = rag_comercial_manager.obtener_contexto_varios_clientes(
                        cliente_ids_consulta,
                        canal,
                        consulta,
                        max_chars=rag_max,
                        top_k_por_scope=top_k,
                        retornar_chunks=True,
                    )
                    contexto_rag, rag_chunks = rag_result
                    if contexto_rag:
                        logger.info(
                            "🧠 RAG comercial unificado | canal=%s | contexto_chars=%s | clientes=%s",
                            canal,
                            len(contexto_rag),
                            cliente_ids_consulta,
                        )
                        try:
                            from core.eventos_ia import emit_rag_query_executed

                            emit_rag_query_executed(
                                pregunta=consulta,
                                cliente=cliente_nati,
                                canal='whatsapp_comercial',
                                chunks_count=len(rag_chunks),
                                contexto_chars=len(contexto_rag),
                                chunks=rag_chunks,
                                metadata={'origen': 'rag_comercial', 'canal_rag': canal},
                            )
                        except Exception:
                            pass
                        break

            if not contexto_rag and getattr(settings, 'BOT_COMERCIAL_RAG_FILE_FALLBACK', True):
                rag_fb = int(getattr(settings, 'BOT_COMERCIAL_RAG_MAX_CHARS', 1600) or 1600)
                rag_fb = max(400, min(rag_fb + 200, 4000))
                fb_docs = int(getattr(settings, 'BOT_COMERCIAL_RAG_FALLBACK_MAX_DOCS', 2) or 2)
                fb_rows = int(getattr(settings, 'BOT_COMERCIAL_RAG_FALLBACK_XLSX_ROWS', 800) or 800)
                contexto_rag = _contexto_fallback_desde_documentos(
                    cliente_ids=cliente_ids_consulta,
                    pregunta=consulta,
                    max_chars=rag_fb,
                    max_docs=fb_docs,
                    xlsx_max_rows=fb_rows,
                )
                if contexto_rag:
                    logger.info("🧠 RAG fallback documental usado | contexto_chars=%s", len(contexto_rag))

            try:
                from core.agrosavia_connector import buscar_agrosavia, formatear_contexto_agrosavia

                if len(contexto_rag or '') < 800:
                    ctx_agrosavia = formatear_contexto_agrosavia(buscar_agrosavia(consulta, size=3))
                    if ctx_agrosavia:
                        contexto_rag = (
                            f"{contexto_rag}\n\n{ctx_agrosavia}".strip()
                            if contexto_rag
                            else ctx_agrosavia
                        )
                        logger.info("🌾 AGROSAVIA live | chars=%s", len(ctx_agrosavia))
            except Exception:
                pass

            if contexto_precios_db:
                contexto_rag = (
                    f"{contexto_precios_db}\n\n{contexto_rag}".strip()
                    if contexto_rag
                    else contexto_precios_db
                )

            from core.nat_router import decidir_routing_nat

            routing = decidir_routing_nat(
                consulta,
                rag_chunks=rag_chunks,
                tiene_rag_texto=bool(contexto_rag),
                contexto_rag_chars=len(contexto_rag or ''),
                ctx_agro=ctx_agro,
                diagnostico_vision=diagnostico_vision,
            )
            logger.info(
                "🧭 Nat routing | modelo=%s modo=%s razon=%s web=%s sim=%s",
                routing.modelo,
                routing.modo,
                routing.razon,
                routing.usar_web,
                routing.rag_max_similitud,
            )

            if routing.usar_web and not contexto_web:
                from core.nati import buscar_en_web_colombia

                contexto_web = buscar_en_web_colombia(consulta) or _contexto_fallback_web_agro(
                    pregunta=consulta,
                    max_chars=1800,
                )
                if contexto_web:
                    logger.info("🌐 Web complementaria (RAG débil/ausente) | chars=%s", len(contexto_web))

            texto_respuesta = _bot_comercial_respuesta_catalogo(
                pregunta=consulta,
                contexto_rag=contexto_rag,
                diagnostico_vision=diagnostico_vision,
                contexto_web=contexto_web,
                historial_chat=historial_chat,
                cliente=cliente_nati,
                sesion_comercial=sesion_comercial,
                bloque_contexto_agro=bloque_contexto_agro,
                routing=routing,
                rag_chunks=rag_chunks,
                ctx_agro=ctx_agro,
            )

    if (
        not es_saludo
        and msg_normalizado not in ['listo', 'continuar', 'menu', 'menú']
        and routing is not None
        and routing.modo in ('tecnico', 'catalogo', 'ambiguo')
    ):
        try:
            import uuid

            from core.eventos_ia import get_or_create_trace_id
            from core.knowledge_studio import crear_candidata_hitl

            tid = get_or_create_trace_id()
            trace_uuid = uuid.UUID(tid) if tid else None
            crear_candidata_hitl(
                cliente=cliente_nati,
                sesion=sesion_comercial,
                telefono=telefono_limpio,
                pregunta=consulta,
                respuesta_nati=texto_respuesta,
                contexto_agro=ctx_agro.to_dict() if ctx_agro else {},
                chunks_rag=rag_chunks,
                trace_id=trace_uuid,
            )
        except Exception:
            pass

    try:
        resultado_envio = enviar_whatsapp_twilio(
            telefono_limpio,
            texto_respuesta,
            from_number=from_number_respuesta,
        )
        if not resultado_envio.get('success'):
            raise RuntimeError(str(resultado_envio.get('response') or 'Error enviando por Twilio'))

        try:
            from core.eventos_ia import emit_mensaje_enviado

            emit_mensaje_enviado(
                telefono=telefono_limpio,
                texto=texto_respuesta,
                mensaje_id=resultado_envio.get('mensaje_id'),
                cliente=cliente_nati,
                canal='whatsapp_comercial',
                agente='nati',
            )
        except Exception:
            pass

        WhatsappLog.objects.create(
            telefono=telefono_limpio,
            mensaje=texto_respuesta[:1500],
            mensaje_id=resultado_envio.get('mensaje_id'),
            tipo='SENT',
            agente_usado='BOT_COMERCIAL',
        )
        _actualizar_sesion_comercial(
            sesion=sesion_comercial,
            mensaje_usuario=msg_body or consulta,
            respuesta_bot=texto_respuesta,
        )
    except Exception as e:
        logger.error(f"❌ Error respondiendo bot comercial: {e}")


@csrf_exempt
def bot_comercial_webhook(request):
    """Webhook dedicado para el número de WhatsApp del bot comercial."""
    if request.method != 'POST':
        return HttpResponse('Method Not Allowed', status=405)

    try:
        try:
            payload = json.loads(request.body.decode('utf-8'))
            # Endpoint dedicado: siempre procesa como bot comercial/agro
            _procesar_bot_comercial_twilio_webhook(payload, forzar_canal=True)
        except json.JSONDecodeError:
            # Endpoint dedicado: siempre procesa como bot comercial/agro
            _procesar_bot_comercial_twilio_webhook(request.POST, forzar_canal=True)
        return HttpResponse('OK')
    except Exception as e:
        logger.error(f"❌ Error webhook bot comercial: {e}")
        return HttpResponse('Error', status=500)


def _escape_twiml(text):
    """Escapa caracteres especiales para TwiML XML."""
    if not text:
        return text
    text = text.replace('&', '&amp;')
    text = text.replace('<', '&lt;')
    text = text.replace('>', '&gt;')
    return text


def _registrar_estado_twilio_callback(post_data):
    """
    Actualiza estado de WhatsappLog (SENT) con callbacks de Twilio.
    Permite métricas de entregado/abierto/fallido por mensaje.
    """
    try:
        sid = (
            post_data.get('MessageSid')
            or post_data.get('SmsSid')
            or post_data.get('message_sid')
            or ''
        )
        status_raw = (
            post_data.get('MessageStatus')
            or post_data.get('SmsStatus')
            or post_data.get('status')
            or ''
        )
        status = str(status_raw).strip().upper()

        if not sid or not status:
            return

        # Tomar el último envío con este SID para no tocar INCOMING.
        log = (
            WhatsappLog.objects.filter(mensaje_id=sid, tipo='SENT')
            .order_by('-fecha')
            .first()
        )
        if log:
            if log.estado != status:
                log.estado = status
                log.save(update_fields=['estado'])
            return

        # Fallback para no perder trazabilidad de callbacks huérfanos.
        WhatsappLog.objects.create(
            telefono='desconocido',
            mensaje=f'[STATUS_CALLBACK:{status}]',
            mensaje_id=sid,
            tipo='SENT',
            estado=status,
            error_detalle='Callback Twilio sin log SENT previo.',
        )
    except Exception as e:
        logger.warning(f"⚠️ No se pudo registrar callback Twilio: {e}")


def _twilio_max_body_chars() -> int:
    """Retorna límite seguro para cuerpo de mensajes Twilio WhatsApp."""
    try:
        valor = int(getattr(settings, 'TWILIO_MAX_BODY_CHARS', 1500) or 1500)
    except (TypeError, ValueError):
        valor = 1500

    if valor < 200 or valor > 1590:
        return 1500
    return valor


def _segmentar_texto_twilio(texto: str, max_chars: int = None) -> list:
    """Divide texto en segmentos compatibles con límite de Twilio."""
    texto = str(texto or '').strip()
    if not texto:
        return ['']

    limite = max_chars or _twilio_max_body_chars()

    try:
        from .response_templates import dividir_contenido_seguro
        chunks = dividir_contenido_seguro(texto, max_chars=limite)
        if chunks:
            return chunks
    except Exception:
        pass

    # Fallback simple por longitud si la utilidad no está disponible.
    return [texto[i:i + limite] for i in range(0, len(texto), limite)]


def youtube_hace_solo_enlace_en_texto(url: str) -> bool:
    """True si la URL es página/embed de YouTube, no archivo .mp4 directo (Twilio no puede adjuntarla como media)."""
    u = (url or '').strip().lower()
    if not u:
        return False
    if u.endswith('.mp4') or '/videoplayback' in u:
        return False
    return 'youtube.com/watch' in u or 'youtu.be/' in u or 'youtube.com/shorts/' in u


def _enviar_mensaje_twilio_segmentado(client, from_number: str, to_number: str, body: str, media_url: str = None) -> list:
    """Envía mensaje Twilio en segmentos seguros y devuelve [(sid, texto_enviado), ...]."""
    body_limpio = str(body or '').strip()
    media_limpia = str(media_url or '').strip() or None
    if not body_limpio and media_limpia:
        body_limpio = TWILIO_CAPTION_ADJUNTO

    segmentos = _segmentar_texto_twilio(body_limpio)
    enviados = []

    from_limpio = str(from_number or '').strip()
    to_limpio = str(to_number or '').strip()

    for idx, segmento in enumerate(segmentos):
        seg_txt = (segmento or '').strip()
        if not seg_txt and media_limpia and idx == 0:
            seg_txt = TWILIO_CAPTION_ADJUNTO
        params = {
            'body': seg_txt if seg_txt else (' ' if media_limpia else ''),
            'from_': from_limpio,
            'to': to_limpio,
        }

        # Adjuntar media solo en el primer segmento.
        if media_limpia and idx == 0:
            params['media_url'] = [media_limpia]

        try:
            mensaje = client.messages.create(**params)
        except Exception as media_err:
            err_str = str(media_err)
            if '63019' in err_str and media_limpia and idx == 0:
                params.pop('media_url', None)
                body_base = (params.get('body') or '').strip()
                params['body'] = f"{body_base}\n\n📎 Archivo: {media_limpia}".strip()
                mensaje = client.messages.create(**params)
            else:
                raise

        enviados.append((mensaje, params.get('body', '') or seg_txt))

    return enviados


def _haversine_metros(lat1, lon1, lat2, lon2):
    """Distancia Haversine en metros."""
    radio_tierra = 6371000.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radio_tierra * c


def _mensaje_bloqueo_drip_view(fecha_desbloqueo):
    from .response_templates import _mensaje_bloqueo_drip
    return _mensaje_bloqueo_drip(fecha_desbloqueo)


def _activar_radar_empleabilidad_si_aplica(estudiante):
    from django.db.models import Q
    from .models import AliadoEmpleabilidad

    if not _cliente_habilita_proximidad(estudiante.cliente):
        return False

    if estudiante.cliente_id:
        hay_aliados = AliadoEmpleabilidad.objects.filter(vacantes_activas=True).filter(
            Q(cliente__isnull=True) | Q(cliente=estudiante.cliente)
        ).exists()
    else:
        hay_aliados = AliadoEmpleabilidad.objects.filter(vacantes_activas=True).exists()

    if not hay_aliados:
        return False

    ctx = estudiante.contexto_temporal or {}
    ctx['radar_empleabilidad_activo'] = True
    ctx['empleabilidad_habilitado_en'] = timezone.now().isoformat()
    estudiante.contexto_temporal = ctx
    estudiante.save(update_fields=['contexto_temporal'])
    return True


def _pregunta_abierta_final_pendiente(estudiante, progreso):
    from .models import PreguntaAbiertaFinalCurso, RespuestaAbiertaFinal

    preguntas_qs = PreguntaAbiertaFinalCurso.objects.filter(
        curso=progreso.curso,
        activa=True
    ).order_by('orden', 'id')

    if not preguntas_qs.exists():
        return None

    cliente_habilita = _cliente_habilita_pregunta_abierta_final(estudiante.cliente)
    curso_habilita = bool(getattr(progreso.curso, 'habilitar_pregunta_abierta_final', False))
    if not (cliente_habilita and curso_habilita):
        logger.info(
            "⚠️ Fallback pregunta abierta final por configuración | estudiante_id=%s | curso_id=%s | cliente_habilita=%s | curso_habilita=%s",
            estudiante.id,
            progreso.curso.id,
            cliente_habilita,
            curso_habilita,
        )

    preguntas = list(preguntas_qs[:3])

    for pregunta in preguntas:
        existe = RespuestaAbiertaFinal.objects.filter(
            pregunta=pregunta,
            estudiante=estudiante
        ).exists()
        if not existe:
            return pregunta

    return None


def _procesar_ubicacion_empleabilidad(estudiante, latitud, longitud):
    """
    Evalúa proximidad del estudiante a aliados activos y construye respuesta.
    Guarda aliado objetivo en contexto para validación de código secreto.
    """
    from django.db.models import Q
    from .models import AliadoEmpleabilidad, MisionEmpleabilidad

    if not _cliente_habilita_proximidad(estudiante.cliente):
        return (
            "📍 El radar de empleabilidad por ubicación no está activo para tu organización en esta fecha. "
            "Si lo esperabas, escribe *ayuda* para que el equipo valide tu acceso."
        )

    if estudiante.cliente_id:
        aliados = AliadoEmpleabilidad.objects.filter(vacantes_activas=True).filter(
            Q(cliente__isnull=True) | Q(cliente=estudiante.cliente)
        )
    else:
        aliados = AliadoEmpleabilidad.objects.filter(vacantes_activas=True)

    aliados = list(aliados)
    if not aliados:
        return "📍 En este momento no hay vacantes activas de aliados en tu zona. Te avisaremos cuando se habiliten."

    cliente_cfg = estudiante.cliente
    radio_metros = int(getattr(cliente_cfg, 'empleabilidad_radio_metros', 800) or 800)
    max_misiones_dia = int(getattr(cliente_cfg, 'empleabilidad_max_misiones_dia', 3) or 3)
    hoy = timezone.localdate()
    misiones_hoy = MisionEmpleabilidad.objects.filter(
        estudiante=estudiante,
        fecha_descubierta__date=hoy,
    ).exclude(estado='cancelada').count()
    if misiones_hoy >= max_misiones_dia:
        return (
            f"📌 Ya completaste tu límite diario de exploración ({max_misiones_dia} misiones).\n"
            "Vuelve mañana para descubrir nuevas oportunidades."
        )

    mejor = None
    mejor_dist = None
    for aliado in aliados:
        dist = _haversine_metros(latitud, longitud, aliado.latitud, aliado.longitud)
        if mejor_dist is None or dist < mejor_dist:
            mejor_dist = dist
            mejor = aliado

    if not mejor:
        return "No pude procesar tu ubicación en este momento. Inténtalo nuevamente."

    if mejor_dist > radio_metros:
        return (
            "📍 Aún no hay oportunidades dentro de tu radio de exploración actual.\n\n"
            f"Distancia más cercana a {mejor.nombre_empresa}: *{int(round(mejor_dist))} m*.\n"
            f"Radio activo de tu organización: *{int(radio_metros)} m*."
        )

    mision = MisionEmpleabilidad.objects.create(
        cliente=estudiante.cliente,
        estudiante=estudiante,
        aliado=mejor,
        estado='descubierta',
        latitud=latitud,
        longitud=longitud,
        distancia_metros=round(mejor_dist, 1),
        metadata={'fuente': 'whatsapp_location'},
    )

    ctx = estudiante.contexto_temporal or {}
    ctx['radar_empleabilidad_activo'] = True
    ctx['aliado_empleabilidad_objetivo_id'] = mejor.id
    ctx['mision_empleabilidad_id'] = mision.id
    ctx['distancia_aliado_m'] = round(mejor_dist, 1)
    estudiante.contexto_temporal = ctx
    estudiante.estado_onboarding = 'esperando_codigo_empleabilidad'
    estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])

    if mejor_dist <= 100:
        return (
            f"🎯 *¡Estás a {int(round(mejor_dist))} metros de {mejor.nombre_empresa}!*\n\n"
            "Acércate a la entrada y envía el *código secreto* que verás en la puerta."
        )

    sector = mejor.indicacion_sector or "del parque principal"
    return (
        "📍 Aún estás lejos de nuestras empresas aliadas.\n\n"
        f"Vas por buen camino. Acércate al sector *{sector}* y vuelve a enviarme tu ubicación.\n"
        f"Distancia aproximada actual a {mejor.nombre_empresa}: *{int(round(mejor_dist))} m*."
    )


def _es_respuesta_liberar_certificado(msg: str) -> bool:
    """
    Respuesta del estudiante que abre ventana WhatsApp tras plantilla de certificado.
    Acepta OK, gracias, cualquier texto o número. NO acepta 'listo'/'continuar'
    (comandos de avance de curso).
    """
    import re

    t = (msg or '').strip().lower()
    if not t:
        return False
    limpio = re.sub(r'[^0-9a-záéíóúñ ]', '', t).strip()
    if not limpio:
        return False
    if limpio in ('listo', 'continuar', 'continuar curso', 'siguiente', 'menu', 'menú'):
        return False
    return True


def _es_ack_certificado(msg: str) -> bool:
    """Alias retrocompatible para tests y plantilla inicial (pide OK)."""
    return _es_respuesta_liberar_certificado(msg)


def _intentar_responder_envio_certificado(estudiante, msg_body, telefono_limpio, msg_from):
    """
    Si el estudiante tiene un certificado pendiente (tras plantilla de aviso) y responde
    cualquier mensaje, envía el diploma (la ventana de 24 h quedó abierta por su respuesta).
    """
    ctx = estudiante.contexto_temporal or {}
    pend = ctx.get('cert_envio_pendiente')
    if not pend:
        return False
    if not _es_respuesta_liberar_certificado(msg_body):
        return False

    cert_id = pend.get('certificado_id')
    from .models_certificados import Certificado
    from .certificado_service import enviar_certificado_whatsapp
    from .certificado_presencial_service import limpiar_cert_envio_pendiente

    cert = (
        Certificado.objects.filter(id=cert_id, emitido=True)
        .select_related('estudiante', 'curso')
        .first()
    )
    if not cert:
        limpiar_cert_envio_pendiente(estudiante)
        return False

    ok = False
    try:
        ok = enviar_certificado_whatsapp(cert)
    except Exception as e:
        logger.error('🎓 Envío certificado tras OK falló est=%s: %s', estudiante.id, e, exc_info=True)

    if ok:
        limpiar_cert_envio_pendiente(estudiante)
        logger.info(
            '🎓 Certificado %s entregado tras respuesta de est=%s',
            cert.codigo_verificacion,
            estudiante.id,
        )
    else:
        logger.error(
            '🎓 Certificado %s NO pudo enviarse tras respuesta de est=%s (sigue pendiente)',
            cert.codigo_verificacion,
            estudiante.id,
        )
    # La respuesta era para el certificado: no seguir onboarding/curso.
    return True


def _procesar_twilio_webhook(post_data):
    """Procesa webhooks de Twilio WhatsApp"""
    # ============================================================
    # FILTRO 1: Ignorar status callbacks de Twilio (queued/sent/delivered)
    # Twilio envía callbacks de estado al mismo webhook, NO son mensajes
    # ============================================================
    message_status = post_data.get('MessageStatus', post_data.get('SmsStatus', ''))
    if message_status and message_status.lower() in ['queued', 'sending', 'sent', 'delivered', 'undelivered', 'failed', 'read']:
        _registrar_estado_twilio_callback(post_data)
        logger.debug(f"Status callback procesado: {message_status}")
        return
    
    # FILTRO 2: Ignorar si no hay Body ni Media (status callback sin MessageStatus)
    raw_body = post_data.get('Body', '')
    raw_media = int(post_data.get('NumMedia', 0))
    if not raw_body and raw_media == 0 and not post_data.get('From', ''):
        logger.debug("Webhook vacio ignorado (sin Body ni Media)")
        return
    
    try:
        logger.info("🔵 TWILIO: Procesando...")
        
        # Twilio envía datos en formato form-data
        msg_body = post_data.get('Body', '')
        msg_from = post_data.get('From', '')  # whatsapp:+573001234567
        msg_to = post_data.get('To', '')      # whatsapp:+14155238886
        msg_sid = post_data.get('MessageSid', f'twilio_{timezone.now().timestamp()}')

        # Evita procesar dos veces el mismo mensaje entrante si Twilio reintenta.
        if msg_sid and WhatsappLog.objects.filter(mensaje_id=msg_sid, tipo='INCOMING').exists():
            logger.info("♻️ Twilio inbound duplicado ignorado | sid=%s", msg_sid)
            return

        from core.eventos_ia import set_trace_id
        set_trace_id()
        
        logger.info(f"📱 Body: {msg_body} | From: {msg_from} | To: {msg_to}")

        # 🎤 DETECTAR AUDIO: Twilio envía audios como MediaUrl
        num_media = int(post_data.get('NumMedia', 0))
        media_type = post_data.get('MediaContentType0', '')
        media_url = post_data.get('MediaUrl0', '')
        latitud_raw = post_data.get('Latitude')
        longitud_raw = post_data.get('Longitude')
        print(f"🎤 DEBUG AUDIO: NumMedia={num_media}, MediaType='{media_type}', MediaUrl={bool(media_url)}, Body='{msg_body[:30] if msg_body else ''}'", flush=True)

        es_audio = num_media > 0 and ('audio' in media_type or 'ogg' in media_type)
        if num_media > 0:
            if es_audio:
                print(f"🎤 Audio detectado: {media_url} (type={media_type})")
                try:
                    transcripcion = _transcribir_audio_twilio(media_url, media_type=media_type)
                    print(f"✅ Audio transcrito: '{transcripcion}'")
                    if transcripcion:
                        msg_body = transcripcion
                    elif not msg_body:
                        msg_body = "[AUDIO_NO_TRANSCRITO]"
                except Exception as e:
                    print(f"❌ Error transcribiendo audio: {e}")
                    import traceback; traceback.print_exc()
                    if not msg_body:
                        msg_body = "[AUDIO_NO_TRANSCRITO]"
            elif not msg_body:
                # Imagen u otro media sin texto — ignorar media, no es audio
                print(f"📎 Media no-audio recibido: type={media_type}")
                msg_body = "listo"
        
        # Limpiar número (quitar whatsapp: y normalizar igual que el modelo)
        if msg_from.startswith('whatsapp:'):
            msg_from = msg_from.replace('whatsapp:', '')
        
        # Normalizar teléfono igual que el modelo (sin +, sin espacios, sin guiones)
        import re
        telefono_limpio = re.sub(r'\D', '', msg_from)  # Solo dígitos
        if len(telefono_limpio) == 10:
            telefono_limpio = f"57{telefono_limpio}"
        
        logger.info(f"📱 De: {msg_from} → Limpio: {telefono_limpio} | Mensaje: {msg_body}")
        logger.info(f"TWILIO MSG: From={telefono_limpio} Body='{msg_body[:50]}'")
        
        # 1. Guardar mensaje entrante con teléfono limpio
        WhatsappLog.objects.create(
            telefono=telefono_limpio,
            mensaje=msg_body,
            mensaje_id=msg_sid,
            tipo='INCOMING',
            es_audio=es_audio,
        )
        logger.info(f"✅ Guardado INCOMING")

        try:
            from core.eventos_ia import emit_webhook_recibido
            from core.ai_capabilities import resolver_ai_capability

            if resolver_ai_capability('eventos_ia'):
                emit_webhook_recibido(
                    mensaje=msg_body,
                    telefono=telefono_limpio,
                    canal='whatsapp_edu',
                )
        except Exception:
            pass
        
        # ============================================================
        # FASE 0: INTERCEPCIÓN DE NO REGISTRADOS (Lead Generation)
        # Si el número no existe en Estudiante, activar "Modo Ventas"
        # ============================================================
        try:
            estudiante = Estudiante.objects.select_related('cliente').get(telefono=telefono_limpio)
            logger.info(f"Estudiante encontrado: {estudiante.nombre} (ID: {estudiante.id})")

            # Respuesta campaña única (Sí/No, Asistiré/No asistiré) — antes del flujo del curso
            try:
                from .campana_respuestas import intentar_registrar_respuesta_campana_unica

                _ack_campana = intentar_registrar_respuesta_campana_unica(
                    telefono_limpio=telefono_limpio,
                    post_data=post_data,
                    msg_body=msg_body,
                    estudiante=estudiante,
                    mensaje_sid=msg_sid,
                )
                if _ack_campana:
                    from twilio.rest import Client as TwilioClient

                    account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                    auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                    twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                    client_tw = TwilioClient(account_sid, auth_token)
                    destino = (
                        f'whatsapp:{msg_from}'
                        if not str(msg_from).startswith('whatsapp:')
                        else msg_from
                    )
                    client_tw.messages.create(
                        body=_ack_campana,
                        from_=str(twilio_number).strip(),
                        to=str(destino).strip(),
                    )
                    WhatsappLog.objects.create(
                        telefono=telefono_limpio,
                        mensaje=_ack_campana,
                        tipo='SENT',
                    )
                    logger.info(
                        '📣 [campana] ack enviado | est=%s tel=%s',
                        estudiante.id,
                        telefono_limpio[-6:],
                    )
                    return
            except Exception as _e_camp:
                logger.warning('📣 [campana] error registrando respuesta: %s', _e_camp, exc_info=True)

            # Ubicación de WhatsApp (Twilio): Latitude/Longitude
            if latitud_raw is not None and longitud_raw is not None:
                try:
                    latitud = float(latitud_raw)
                    longitud = float(longitud_raw)
                    texto_geo = _procesar_ubicacion_empleabilidad(estudiante, latitud, longitud)
                    try:
                        from twilio.rest import Client as TwilioClient
                        account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                        auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                        twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                        client_tw = TwilioClient(account_sid, auth_token)
                        destino = f'whatsapp:{msg_from}' if not str(msg_from).startswith('whatsapp:') else msg_from
                        client_tw.messages.create(body=texto_geo, from_=str(twilio_number).strip(), to=str(destino).strip())
                        WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_geo, tipo='SENT')
                    except Exception as e:
                        logger.error(f"❌ Error enviando respuesta de ubicación: {e}")
                    return
                except ValueError:
                    logger.warning(f"⚠️ Coordenadas inválidas recibidas: lat={latitud_raw}, lon={longitud_raw}")

            # Certificado presencial pendiente: ANTES de Habeas/onboarding.
            # Si el estudiante respondió OK a la plantilla inicial, "ok" no debe
            # interpretarse como aceptación de política de datos.
            try:
                if _intentar_responder_envio_certificado(
                    estudiante, msg_body, telefono_limpio, msg_from,
                ):
                    return
            except Exception as e:
                logger.exception('Intercept certificado pendiente omitido: %s', e)

        except Estudiante.DoesNotExist:
            # Verificar si ya es un prospecto B2B existente
            from .models import ProspectoB2B
            prospecto = None
            try:
                prospecto = ProspectoB2B.objects.get(telefono=telefono_limpio)
            except ProspectoB2B.DoesNotExist:
                pass
            
            msg_lower = msg_body.strip().lower()
            
            if prospecto:
                # Prospecto existente - procesar su respuesta
                if prospecto.esperando_email:
                    # Validar si parece un email
                    import re as re_email
                    email_match = re_email.search(r'[\w.+-]+@[\w-]+\.[\w.]+', msg_body)
                    if email_match:
                        prospecto.email = email_match.group(0)
                        prospecto.esperando_email = False
                        prospecto.fecha_ultimo_contacto = timezone.now()
                        prospecto.save()
                        
                        # Notificar al equipo de ventas por email
                        try:
                            from django.core.mail import send_mail
                            send_mail(
                                subject=f"🏢 Nuevo Lead B2B - {prospecto.empresa or prospecto.telefono}",
                                message=f"Nuevo prospecto capturado por el bot:\n\nTeléfono: {prospecto.telefono}\nEmpresa: {prospecto.empresa}\nEmail: {prospecto.email}\nMensaje: {prospecto.mensaje_original}",
                                from_email=settings.DEFAULT_FROM_EMAIL,
                                recipient_list=[getattr(settings, 'EMAIL_SOPORTE', 'comunidad.educativa@eki.com.co')],
                                fail_silently=True
                            )
                        except Exception:
                            pass
                        
                        texto_respuesta = (
                            "✅ *¡Perfecto!*\n\n"
                            f"Hemos registrado tu correo: *{prospecto.email}*\n\n"
                            "Nuestro equipo de ventas te contactará muy pronto "
                            "para contarte todo sobre las capacitaciones de eki. 🚜\n\n"
                            "¡Gracias por tu interés! 🌱"
                        )
                    else:
                        texto_respuesta = "📧 Por favor envía un correo electrónico válido (ej: nombre@empresa.com)"
                elif msg_lower in ['1', 'empresa', 'eki para mi empresa']:
                    prospecto.esperando_email = True
                    prospecto.fecha_ultimo_contacto = timezone.now()
                    prospecto.save()
                    texto_respuesta = (
                        "🏢 *¡Excelente!*\n\n"
                        "Nos encantaría ayudar a capacitar a tu equipo.\n\n"
                        "📧 Por favor envíanos tu *correo electrónico* "
                        "y un asesor de ventas te contactará:\n\n"
                        "👉 Ejemplo: juan@miempresa.com"
                    )
                elif msg_lower in ['3', 'ayuda', 'soy estudiante', 'estudiante']:
                    texto_respuesta = (
                        "🙋‍♂️ *¡Entendido!*\n\n"
                        "Si eres estudiante y cambiaste de número, "
                        "por favor contacta a tu coordinador o escribe a:\n\n"
                        "📧 comunidad.educativa@eki.com.co\n\n"
                        "Incluye tu nombre completo y número de cédula para que podamos ayudarte."
                    )
                else:
                    from .whatsapp_service import enviar_mensaje_ventas
                    enviar_mensaje_ventas(msg_from)
                    return
                
                # Enviar respuesta al prospecto
                try:
                    from twilio.rest import Client as TwilioClient
                    account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                    auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                    twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                    client_tw = TwilioClient(account_sid, auth_token)
                    destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                    client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                except Exception as e:
                    print(f"❌ Error enviando a prospecto: {e}")
                return
            
            else:
                # Nuevo prospecto - crear y enviar mensaje de ventas
                ProspectoB2B.objects.create(
                    telefono=telefono_limpio,
                    mensaje_original=msg_body,
                    origen='whatsapp_bot'
                )
                from .whatsapp_service import enviar_mensaje_ventas
                enviar_mensaje_ventas(msg_from)
                logger.info(f"🏢 Nuevo prospecto B2B capturado: {telefono_limpio}")
                return
        
        # ============================================================
        # MÁQUINA DE ESTADOS B2B (Onboarding con Botones Twilio)
        # ============================================================
        estado_chat = getattr(estudiante, 'estado_chat', None)
        logger.info(f"📍 Estado estudiante {estudiante.nombre}: estado_chat={estado_chat}, onboarding={estudiante.estado_onboarding}, acepto={estudiante.acepto_terminos}")
        
        # Migrar estudiantes legacy al nuevo sistema
        if not estado_chat or estado_chat in ('', None):
            if estudiante.acepto_terminos and estudiante.estado_onboarding == 'completado':
                estudiante.estado_chat = 'ACTIVO'
            elif estudiante.acepto_terminos:
                estudiante.estado_chat = 'ESPERANDO_CEDULA'
            else:
                estudiante.estado_chat = 'ESPERANDO_HABEAS_DATA'
            estudiante.save()
            estado_chat = estudiante.estado_chat
            logger.info(f"📍 Legacy migration: estado_chat → {estado_chat}")
        
        # Auto-corregir: admin creó estudiante con acepto_terminos=True pero estado_chat quedó en ESPERANDO_HABEAS_DATA
        if estado_chat == 'ESPERANDO_HABEAS_DATA' and estudiante.acepto_terminos:
            if estudiante.estado_onboarding == 'completado':
                estudiante.estado_chat = 'ACTIVO'
            else:
                estudiante.estado_chat = 'ESPERANDO_CEDULA'
            estudiante.save()
            estado_chat = estudiante.estado_chat
            logger.info(f"📍 Auto-corrección admin: estado_chat → {estado_chat}")

        # ============================================================
        # PRIORIDAD GLOBAL: menu + corregir datos (post-habeas)
        # Debe funcionar en cualquier estado del bot una vez el usuario
        # aceptó términos.
        # ============================================================
        from .correccion_datos import (
            construir_menu_principal_texto,
            es_keyword_correccion,
            es_keyword_menu,
            estudiante_en_flujo_correccion,
            iniciar_flujo_correccion,
            normalizar_texto,
            procesar_flujo_correccion,
        )
        texto_norm = normalizar_texto(msg_body)

        # Si el usuario quedó en estado legacy por PQRS/corrección, normalizar
        # para que "listo" retome el curso sin falso "sin curso".
        if texto_norm in {"listo", "continuar"}:
            from .models import ProgresoEstudiante
            if ProgresoEstudiante.objects.filter(
                estudiante=estudiante, completado=False, curso__activo=True
            ).exists():
                cambios = []
                if estudiante.estado_chat != "ACTIVO":
                    estudiante.estado_chat = "ACTIVO"
                    cambios.append("estado_chat")
                if estudiante.estado_onboarding not in ("completado",):
                    estudiante.estado_onboarding = "completado"
                    cambios.append("estado_onboarding")
                if cambios:
                    estudiante.save(update_fields=cambios)

        if estudiante.acepto_terminos:
            from .utils import enviar_whatsapp_twilio
            if estudiante_en_flujo_correccion(estudiante):
                texto_respuesta = procesar_flujo_correccion(estudiante, msg_body)
                enviar_whatsapp_twilio(msg_from, texto_respuesta)
                return

            if es_keyword_correccion(texto_norm):
                texto_respuesta = iniciar_flujo_correccion(estudiante)
                enviar_whatsapp_twilio(msg_from, texto_respuesta)
                return

            from .flujo_whatsapp_b2b import (
                es_estudiante_b2b,
                es_keyword_retomar,
                salir_seleccion_curso_legacy,
            )
            from .response_templates import get_response_for_intent

            if es_estudiante_b2b(estudiante) and es_keyword_retomar(texto_norm):
                salir_seleccion_curso_legacy(estudiante)
                texto_respuesta = get_response_for_intent(
                    'continuar_leccion',
                    estudiante.nombre,
                    estudiante_id=estudiante.id,
                    mensaje_original=msg_body,
                )
                enviar_whatsapp_twilio(msg_from, texto_respuesta)
                return

            if es_keyword_menu(texto_norm):
                texto_respuesta = construir_menu_principal_texto(estudiante)
                enviar_whatsapp_twilio(msg_from, texto_respuesta)
                return
        
        # --- BARRERA 1: HABEAS DATA ---
        if estado_chat == 'ESPERANDO_HABEAS_DATA':
            if not (estudiante.contexto_temporal or {}).get('cert_envio_pendiente'):
                msg_lower = msg_body.strip().lower()
                keywords_acepto = ['acepto', 'sí', 'si', 'aceptar', 'ok', 'yes', 'acepto', 'de acuerdo']
                keywords_no = ['no acepto', 'no', 'rechazo', 'rechazar']

                if any(k in msg_lower for k in keywords_acepto):
                    estudiante.acepto_terminos = True
                    estudiante.fecha_aceptacion_terminos = timezone.now()
                    estudiante.estado_chat = 'ESPERANDO_CEDULA'
                    estudiante.save()

                    texto_respuesta = (
                        "✅ *¡Gracias por aceptar!*\n\n"
                        "Para verificar tu identidad, por favor escribe "
                        "tu *número de cédula* (solo los números, sin puntos ni espacios).\n\n"
                        "👉 Ejemplo: 1234567890"
                    )
                elif any(k in msg_lower for k in keywords_no):
                    texto_respuesta = (
                        "😔 Entendemos tu decisión.\n\n"
                        "Sin la aceptación de la política de datos no podemos "
                        "activar tu cuenta en la plataforma.\n\n"
                        "Si cambias de opinión, escríbenos en cualquier momento. 🌱"
                    )
                else:
                    # Enviar primero template Twilio (cliente > global > fallback eki).
                    # Si falla, degradar a texto plano con URL para no bloquear onboarding.
                    from .whatsapp_service import enviar_habeas_data
                    resultado_tpl = enviar_habeas_data(msg_from, cliente=estudiante.cliente)
                    if resultado_tpl.get('success'):
                        return

                    # Fallback texto: mostrar URL efectiva del cliente.
                    from .security_handler import _url_politica_datos_cliente
                    url_politica = _url_politica_datos_cliente(estudiante=estudiante)
                    texto_respuesta = (
                        "👋 *¡Bienvenido a eki!*\n\n"
                        "🚜 Tu plataforma de soluciones educativas por WhatsApp\n\n"
                        "📜 *Protección de Datos Personales*\n"
                        "Antes de comenzar, necesitamos tu autorización para usar "
                        "tus datos de acuerdo con la Ley 1581 de 2012.\n\n"
                        f"🔗 Lee nuestra política completa aquí:\n{url_politica}\n\n"
                        "*¿Aceptas el tratamiento de tus datos?*\n\n"
                        "👉 Escribe *Acepto* o *No acepto*"
                    )

                # Enviar y cortar
                try:
                    from twilio.rest import Client as TwilioClient
                    account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                    auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                    twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                    client_tw = TwilioClient(account_sid, auth_token)
                    destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                    client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                    WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
                except Exception as e:
                    logger.error(f"❌ Error enviando habeas data: {e}")
                    import traceback; traceback.print_exc()
                return  # CORTAR EJECUCIÓN
        
        # --- BARRERA 2: VALIDACIÓN 2FA (Cédula) ---
        if estado_chat == 'ESPERANDO_CEDULA':
            # Limpiar input del usuario
            cedula_input = re.sub(r'[\s\.\-]', '', msg_body.strip())
            msg_lower_cedula = msg_body.strip().lower()
            
            # Detectar "ayuda" → crear ticket de soporte
            if msg_lower_cedula in ['ayuda', 'help', 'soporte']:
                from .models import SolicitudSoporte
                solicitud = SolicitudSoporte.objects.create(
                    estudiante=estudiante,
                    mensaje_original=f"Ayuda en verificación de cédula - no coincide con registros",
                    keyword_usada='ayuda_cedula',
                    asunto='Problema con verificación de cédula',
                    prioridad='media'
                )
                texto_respuesta = (
                    f"🆘 *Ticket de Soporte #{solicitud.id}*\n\n"
                    f"Hola {estudiante.nombre}, hemos registrado tu solicitud.\n\n"
                    "📝 Un asesor revisará tu caso y te contactará pronto.\n"
                    "🕐 *Tiempo de respuesta:* menos de 24 horas.\n\n"
                    "Si recuerdas tu cédula, puedes intentar de nuevo escribiéndola aquí."
                )
            # Comparar con la cédula sanitizada en BD
            elif cedula_input == estudiante.cedula:
                estudiante.estado_chat = 'CONFIRMANDO_DATOS'
                estudiante.save()
                
                # Enviar confirmación con datos + botones (5 variables)
                org_nombre = estudiante.cliente.nombre if estudiante.cliente else 'eki'
                from .whatsapp_service import enviar_confirmacion_datos
                resultado = enviar_confirmacion_datos(
                    msg_from,
                    estudiante.nombre,
                    f"{estudiante.tipo_documento} {estudiante.cedula}",
                    org_nombre,
                    edad=estudiante.edad,
                    municipio=estudiante.municipio,
                )
                if resultado.get('success'):
                    return  # Template enviado
                
                # Fallback texto plano
                texto_respuesta = (
                    "✅ *¡Cédula verificada!*\n\n"
                    "Tus datos registrados:\n\n"
                    f"👤 *Nombre:* {estudiante.nombre}\n"
                    f"🆔 *Documento:* {estudiante.tipo_documento} {estudiante.cedula}\n"
                    f"📍 *Municipio:* {estudiante.municipio or 'No registrado'}\n"
                    f"🏢 *Organización:* {org_nombre}\n"
                    f"🎂 *Edad:* {estudiante.edad or 'No registrada'}\n"
                    f"👫 *Género:* {estudiante.get_genero_display() if estudiante.genero else 'No registrado'}\n\n"
                    "*¿Tus datos están correctos?*\n\n"
                    "👉 Escribe *Sí* si todo está bien\n"
                    "👉 Escribe *No* si hay un error"
                )
            else:
                texto_respuesta = (
                    "❌ *Cédula no coincide*\n\n"
                    "El número que ingresaste no coincide con "
                    "nuestros registros.\n\n"
                    "Por favor verifica y escribe tu cédula nuevamente "
                    "(solo números, sin puntos ni espacios).\n\n"
                    "👉 Ejemplo: 1234567890\n\n"
                    "Si crees que hay un error, escribe *ayuda*"
                )
            
            try:
                from twilio.rest import Client as TwilioClient
                account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                client_tw = TwilioClient(account_sid, auth_token)
                destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
            except Exception as e:
                logger.error(f"❌ Error enviando validación 2FA: {e}")
                import traceback; traceback.print_exc()
            return  # CORTAR EJECUCIÓN
        
        # --- BARRERA 3: CONFIRMACIÓN DE DATOS ---
        if estado_chat == 'CONFIRMANDO_DATOS':
            msg_lower = msg_body.strip().lower()
            keywords_si = ['sí', 'si', 'todo bien', 'correcto', 'bien', 'ok', 'yes', 'confirmo', 'confirmar']
            keywords_modificar = ['modificar', 'no', 'error', 'mal', 'incorrecto', 'hay un error', 'cambiar']
            
            if any(k in msg_lower for k in keywords_si):
                estudiante.estado_chat = 'ACTIVO'
                estudiante.estado_onboarding = 'completado'  # Legacy compat
                estudiante.save()
                
                # Enviar curso directamente (sin menú)
                org_nombre = estudiante.cliente.nombre if estudiante.cliente else 'eki'
                try:
                    from .models import ProgresoEstudiante
                    from .response_templates import obtener_video_url
                    from .selector_curso import resolver_curso_post_confirmacion

                    curso = resolver_curso_post_confirmacion(estudiante)
                    if curso:
                        progreso, creado = ProgresoEstudiante.objects.get_or_create(
                            estudiante=estudiante,
                            curso=curso,
                            defaults={'completado': False}
                        )
                        modulo = progreso.modulo_actual
                        if not modulo:
                            modulo = curso.modulos.order_by('numero').first()
                            if modulo:
                                progreso.modulo_actual = modulo
                                from .module_steps import reset_progreso_pasos_modulo
                                reset_progreso_pasos_modulo(progreso, save=False)
                                progreso.save(
                                    update_fields=[
                                        'modulo_actual',
                                        'paso_actual_modulo',
                                        'esperando_respuesta_evaluacion_paso',
                                        'paso_evaluacion_paso_id',
                                    ]
                                )
                        if modulo:
                            # Get agent names: Cliente > Curso > defaults
                            cliente_obj = estudiante.cliente
                            nombre_tutor = (
                                (cliente_obj.nombre_agente_tutor if cliente_obj and hasattr(cliente_obj, 'nombre_agente_tutor') and cliente_obj.nombre_agente_tutor else '') or
                                curso.nombre_agente_tutor or 'Claudia'
                            )
                            nombre_asistente = (
                                (cliente_obj.nombre_agente_asistente if cliente_obj and hasattr(cliente_obj, 'nombre_agente_asistente') and cliente_obj.nombre_agente_asistente else '') or
                                curso.nombre_agente_asistente or 'Darío'
                            )
                            
                            # Presentación de agentes
                            from .tutor_ia_modulo import generar_presentacion_agentes
                            msg_tutor, msg_asistente = generar_presentacion_agentes(
                                curso_nombre=curso.nombre,
                                estudiante_nombre=estudiante.nombre or 'Estudiante',
                                nombre_tutor=nombre_tutor,
                                nombre_asistente=nombre_asistente
                            )
                            
                            # Gamification explanation message
                            msg_gamificacion = ""
                            usar_gamificacion = (cliente_obj.usar_gamificacion if cliente_obj else True)
                            if usar_gamificacion:
                                msg_gamificacion = (
                                    "🎮 *Nuestra experiencia de formación funciona a través de puntos*\n\n"
                                    "A medida que avances en el curso, tendrás retos que evaluar.\n\n"
                                    "¡Vamos a aprender y avanzar juntos! 💪"
                                )

                            # --- Mensaje 1: Bienvenida + Gamificación + Agentes (TODO EN UNO) ---
                            partes_intro = [
                                f"✅ *¡Datos confirmados, {estudiante.nombre}!*\n\nBienvenido al programa de *{org_nombre}*"
                            ]
                            partes_intro.append(msg_tutor)
                            partes_intro.append(msg_asistente)
                            if msg_gamificacion:
                                partes_intro.append(msg_gamificacion)
                            partes_intro.append("📚 *Comenzamos con el primer módulo de tu curso...* 👇")
                            msg_intro = "\n\n".join(partes_intro)

                            from .module_steps import (
                                modulo_usa_pasos,
                                pasos_activos_qs,
                                reset_progreso_pasos_modulo,
                                entregar_bloque_secciones_desde_paso,
                                log_y_mensaje_modo_pasos_sin_pasos,
                            )

                            if modulo_usa_pasos(modulo):
                                if not pasos_activos_qs(modulo).exists():
                                    _fb_conf = log_y_mensaje_modo_pasos_sin_pasos(
                                        modulo, 'confirmando_datos_primer_modulo'
                                    )
                                    texto_respuesta = '[MULTI_MSG]' + msg_intro + '[SEP]' + _fb_conf
                                else:
                                    reset_progreso_pasos_modulo(progreso, save=True)
                                    partes_mod0_conf: list[str] = []
                                    if modulo.numero == 0 and (modulo.contenido or '').strip():
                                        from .response_templates import dividir_contenido_seguro as _div0
                                        _mh = f"📖 *{modulo.numero}. {modulo.titulo}*\n\n"
                                        _ch0 = _div0(modulo.contenido, max_chars=1500)
                                        if _ch0:
                                            _mm0 = _mh + _ch0[0]
                                            for _ck in _ch0[1:]:
                                                if len(_mm0) + len(_ck) + 4 < 1500:
                                                    _mm0 += "\n\n" + _ck
                                                else:
                                                    break
                                            partes_mod0_conf.append(_mm0)
                                    msg_pasos_conf = entregar_bloque_secciones_desde_paso(
                                        progreso, modulo, 1
                                    )
                                    _inner_conf = msg_pasos_conf[len('[MULTI_MSG]') :]
                                    _pieces = partes_mod0_conf + [
                                        p for p in _inner_conf.split('[SEP]') if p
                                    ]
                                    texto_respuesta = '[MULTI_MSG]' + msg_intro + '[SEP]' + '[SEP]'.join(_pieces)
                            else:
                                video_url = obtener_video_url(modulo)
                                archivos_multimedia = modulo.archivos_multimedia.filter(activo=True)
                                archivos_msg = ""
                                primera_media_url = None
                                extra_media_urls = []
                                if archivos_multimedia.exists():
                                    archivos_msg = ""
                                    for idx, archivo in enumerate(archivos_multimedia):
                                        icono = {'video': '🎥', 'imagen': '🖼️', 'infografia': '📊', 'pdf': '📄', 'audio': '🎵'}.get(archivo.tipo, '📁')
                                        url = archivo.get_url_para_envio()
                                        if url:
                                            if not primera_media_url:
                                                primera_media_url = url
                                                archivos_msg += f"\n{icono} {archivo.titulo} (adjunto)"
                                            else:
                                                extra_media_urls.append((url, archivo.titulo, icono))
                                                archivos_msg += f"\n{icono} {archivo.titulo} (adjunto)"
                                        else:
                                            archivos_msg += f"\n{icono} {archivo.titulo}"
                                if not archivos_multimedia.exists() and video_url:
                                    primera_media_url = video_url

                                # --- Mensaje 2: Contenido del módulo (con multimedia) ---
                                from .response_templates import dividir_contenido_seguro
                                contenido_modulo = modulo.contenido or ''
                                chunks = dividir_contenido_seguro(contenido_modulo, max_chars=1300)
                                modulo_header = f"📖 *Módulo {modulo.numero}: {modulo.titulo}*\n\n"
                                if chunks:
                                    msg_modulo = modulo_header + chunks[0]
                                    for chunk in chunks[1:]:
                                        if len(msg_modulo) + len(chunk) + 4 < 1400:
                                            msg_modulo += "\n\n" + chunk
                                        else:
                                            break
                                else:
                                    msg_modulo = modulo_header + (modulo.descripcion or '')
                                # v1.9.8: No mostrar labels de archivos en texto (se envían como mensajes separados)

                                # Orden: intro (con agentes) → módulo TEXTO → video(s) → [DELAY] → "escribe listo"
                                texto_respuesta = "[MULTI_MSG]" + msg_intro + "[SEP]" + msg_modulo
                                # Video principal como mensaje separado después del texto
                                hay_media_conf = False
                                if primera_media_url:
                                    texto_respuesta += f"[SEP]{parte_mensaje_con_media(primera_media_url)}"
                                    hay_media_conf = True
                                for extra_url, extra_titulo, extra_icono in extra_media_urls:
                                    cap_extra = (
                                        f'{extra_icono} {extra_titulo}'.strip()
                                        if extra_titulo
                                        else None
                                    )
                                    texto_respuesta += f"[SEP]{parte_mensaje_con_media(extra_url, cap_extra)}"
                                    hay_media_conf = True
                                # "Escribe listo" AL FINAL — solo si hay más módulos
                                hay_mas_modulos = curso.modulos.filter(numero__gt=modulo.numero).exists()
                                if hay_mas_modulos:
                                    if hay_media_conf:
                                        texto_respuesta += "[SEP][DELAY:5]"
                                    from .avance_whatsapp import CTX_FIN_ENTREGA_MODULO, resolver_cta_listo
                                    texto_respuesta += "[SEP]" + resolver_cta_listo(
                                        estudiante, curso, CTX_FIN_ENTREGA_MODULO
                                    )
                        else:
                            texto_respuesta = f"✅ *¡Datos confirmados!* Bienvenido al programa de *{org_nombre}*.\n\nEl curso aún no tiene módulos configurados. Te notificaremos cuando estén listos."
                    else:
                        texto_respuesta = f"✅ *¡Datos confirmados!* Bienvenido al programa de *{org_nombre}*.\n\nAún no hay cursos disponibles. Te notificaremos cuando estén listos."
                except Exception as e:
                    logger.error(f"❌ Error enviando curso directo: {e}")
                    import traceback; traceback.print_exc()
                    texto_respuesta = f"✅ *¡Datos confirmados!* Bienvenido al programa de *{org_nombre}*.\n\nTu organización te notificará cuando estén listos los cursos. Escribe *ayuda* si necesitas asistencia."
                
                # MENÚ OCULTO (no eliminado del código):
                # from .whatsapp_service import enviar_menu_principal
                # resultado = enviar_menu_principal(msg_from, estudiante.nombre)
                # if resultado.get('success'):
                #     return
            elif any(k in msg_lower for k in keywords_modificar):
                # Botón "Modificar" presionado → crear ticket de soporte directamente
                from .models import SolicitudSoporte
                SolicitudSoporte.objects.create(
                    estudiante=estudiante,
                    mensaje_original=f"Solicitud de corrección de datos desde verificación. Datos actuales: Nombre={estudiante.nombre}, Cédula={estudiante.cedula}, Municipio={estudiante.municipio}",
                    keyword_usada='correccion_datos',
                    asunto='Corrección de datos desde verificación',
                    estado='pendiente'
                )
                texto_respuesta = (
                    "📝 *Solicitud de Corrección Recibida*\n\n"
                    f"Hola {estudiante.nombre}, hemos creado un ticket de soporte "
                    "para la corrección de tus datos.\n\n"
                    "📧 *Nuestro equipo te contactará pronto.*"
                )
                estudiante.estado_chat = 'ACTIVO'
                estudiante.estado_onboarding = 'completado'
                estudiante.save()
            else:
                # Re-enviar la plantilla de confirmación (tiene botones Confirmar/Modificar)
                from .whatsapp_service import enviar_confirmacion_datos
                org_nombre = estudiante.cliente.nombre if estudiante.cliente else 'eki'
                resultado_reenvio = enviar_confirmacion_datos(
                    msg_from,
                    estudiante.nombre,
                    f"{estudiante.tipo_documento} {estudiante.cedula}",
                    org_nombre,
                    edad=estudiante.edad,
                    municipio=estudiante.municipio,
                )
                if resultado_reenvio.get('success'):
                    return  # Template reenviado, no necesita texto
                texto_respuesta = (
                    "Por favor revisa tus datos y toca *Confirmar* o *Modificar* en la plantilla."
                )
            
            try:
                from twilio.rest import Client as TwilioClient
                account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                client_tw = TwilioClient(account_sid, auth_token)
                destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                
                # Handle multi-message (process [MULTI_MSG]/[SEP]/[MEDIA:] markers)
                if texto_respuesta.startswith('[MULTI_MSG]'):
                    import re as re_conf
                    partes_conf = texto_respuesta.replace('[MULTI_MSG]', '', 1).split('[SEP]')
                    for parte_c in partes_conf:
                        if not parte_c.strip():
                            continue
                        parte_texto_c = parte_c.strip()
                        # [DELAY:N] — pausa intencional para que WhatsApp entregue videos
                        delay_m_c = re_conf.match(r'^\[DELAY:(\d+)\]$', parte_texto_c)
                        if delay_m_c:
                            import time; time.sleep(int(delay_m_c.group(1)))
                            continue
                        # Detectar Content Template → enviar como template, NO como texto
                        if parte_texto_c.startswith('[SEND_TEMPLATE:'):
                            tmpl_m_c = re_conf.match(r'\[SEND_TEMPLATE:(HX[a-f0-9]+)\]', parte_texto_c)
                            if tmpl_m_c:
                                from .whatsapp_service import enviar_template_twilio
                                tel_limpio_t = msg_from.replace('whatsapp:', '').replace('+', '')
                                enviar_template_twilio(tel_limpio_t, tmpl_m_c.group(1))
                            import time; time.sleep(0.5)
                            continue
                        parte_media_c = None
                        media_m_c = re_conf.search(r'\[MEDIA:(.*?)\]', parte_texto_c)
                        if media_m_c:
                            parte_media_c = (media_m_c.group(1) or '').strip()
                            parte_texto_c = parte_texto_c.replace(media_m_c.group(0), '').strip()
                        if parte_media_c and not (parte_texto_c or '').strip():
                            parte_texto_c = TWILIO_CAPTION_ADJUNTO
                        mp_c = {'body': parte_texto_c, 'from_': str(twilio_number).strip(), 'to': str(destino).strip()}
                        if parte_media_c:
                            mp_c['media_url'] = [parte_media_c]
                        try:
                            msg_sent = client_tw.messages.create(**mp_c)
                        except Exception as media_err_c:
                            if '63019' in str(media_err_c) and parte_media_c:
                                mp_c.pop('media_url', None)
                                mp_c['body'] += f"\n\n📎 Archivo: {parte_media_c}"
                                msg_sent = client_tw.messages.create(**mp_c)
                            else:
                                raise
                        import time; time.sleep(0.5)
                        WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=parte_texto_c[:500], tipo='SENT')
                else:
                    client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                    WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
            except Exception as e:
                logger.error(f"❌ Error enviando confirmación: {e}")
                import traceback; traceback.print_exc()
            return  # CORTAR EJECUCIÓN
        
        # --- BARRERA 3B: LEGACY AUTO-CORRECCIÓN → Redirigir a ACTIVO ---
        if estado_chat in ('ESPERANDO_AYUDA_MODIFICAR', 'ESPERANDO_CORRECCION_DATOS'):
            # Ya no hay menú de corrección — redirigir al flujo normal
            estudiante.estado_chat = 'ACTIVO'
            estudiante.estado_onboarding = 'completado'
            estudiante.save()
            
            msg_lower = msg_body.strip().lower()
            
            # Si dice "continuar", "sí" o similar → enviar al curso
            keywords_continuar = ['continuar', 'sí', 'si', 'ok', 'listo', 'seguir', 'menu', 'menú']
            keywords_ayuda = ['ayuda', 'soporte', 'ticket', 'problema']
            
            if any(k in msg_lower for k in keywords_ayuda):
                from .models import SolicitudSoporte
                SolicitudSoporte.objects.create(
                    estudiante=estudiante,
                    mensaje_original=f"Solicitud de soporte: {msg_body}",
                    keyword_usada='soporte_legacy',
                    asunto='Soporte desde flujo legacy',
                    estado='pendiente'
                )
                texto_respuesta = (
                    "🆘 *Solicitud Recibida*\n\n"
                    f"Hola {estudiante.nombre}, hemos registrado tu solicitud.\n"
                    "Nuestro equipo te contactará pronto.\n\n"
                    "👉 Escribe *continuar* para seguir con tu curso"
                )
            else:
                texto_respuesta = (
                    f"✅ *¡Listo, {estudiante.nombre}!*\n\n"
                    "Continuemos con tu curso."
                )
            
            try:
                from twilio.rest import Client as TwilioClient
                account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                client_tw = TwilioClient(account_sid, auth_token)
                destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
            except Exception as e:
                logger.error(f"❌ Error enviando ayuda modificar: {e}")
                import traceback; traceback.print_exc()
            return
        
        # ============================================================
        # ESTUDIANTE ACTIVO - Procesar acciones del menú y flujo normal
        # ============================================================
        # Detectar acciones del menú principal (tanto ACTIVO como completado)
        if estado_chat == 'ACTIVO' or estudiante.estado_onboarding == 'completado':
            msg_lower = msg_body.strip().lower()
            logger.info(f"📍 ACTIVO handler: msg='{msg_lower}', onboarding={estudiante.estado_onboarding}")

            if estado_chat == 'ACTIVO':
                try:
                    from formulario.routing import debe_usar_agente_formulario
                    if debe_usar_agente_formulario(estudiante):
                        from formulario.agent import manejar_mensaje_formulario
                        texto_respuesta = manejar_mensaje_formulario(estudiante, msg_body)
                        try:
                            from twilio.rest import Client as TwilioClient
                            account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                            auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                            twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                            client_tw = TwilioClient(account_sid, auth_token)
                            destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                            client_tw.messages.create(
                                body=texto_respuesta,
                                from_=str(twilio_number).strip(),
                                to=str(destino).strip()
                            )
                            WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
                        except Exception as e:
                            logger.error(f"❌ Error enviando respuesta formulario GEI: {e}")
                        return
                except Exception as e:
                    logger.error(f"❌ Agente formulario: {e}", exc_info=True)
            
            # PRIORIDAD: Si está seleccionando curso, NO interceptar números
            if estudiante.estado_onboarding == 'esperando_seleccion_curso':
                from .flujo_whatsapp_b2b import (
                    es_estudiante_b2b,
                    mensaje_digitos_sin_menu,
                    salir_seleccion_curso_legacy,
                )
                if es_estudiante_b2b(estudiante):
                    salir_seleccion_curso_legacy(estudiante)
                    from .response_templates import get_response_for_intent
                    if msg_body.strip().isdigit():
                        texto_respuesta = mensaje_digitos_sin_menu(estudiante)
                    else:
                        texto_respuesta = get_response_for_intent(
                            'continuar_leccion',
                            estudiante.nombre,
                            estudiante_id=estudiante.id,
                            mensaje_original=msg_body,
                        )
                elif msg_lower in ['menu', 'menú']:
                    estudiante.estado_onboarding = 'completado'
                    estudiante.contexto_temporal = None
                    estudiante.save()
                    from .response_templates import get_response_for_intent
                    texto_respuesta = get_response_for_intent('saludo', estudiante.nombre, estudiante_id=estudiante.id)
                else:
                    # Extraer número del curso: soporta "tomar 1", "1", "tomar1"
                    import re as re_curso
                    indice = None
                    match_tomar = re_curso.match(r'^tomar\s*(\d+)$', msg_lower)
                    if match_tomar:
                        indice = int(match_tomar.group(1))
                    elif msg_body.strip().isdigit():
                        indice = int(msg_body.strip())
                    
                    if indice is not None:
                        from .selector_curso import continuar_curso_seleccionado
                        estudiante.estado_onboarding = 'completado'
                        estudiante.save(update_fields=['estado_onboarding'])
                        texto_respuesta = continuar_curso_seleccionado(estudiante.id, indice, msg_body)
                        logger.info(f"✅ Curso seleccionado: {indice}")
                    else:
                        # No es número ni menú → resetear y procesar normalmente
                        estudiante.estado_onboarding = 'completado'
                        estudiante.contexto_temporal = None
                        estudiante.save()
                        from core.eventos_ia import detectar_intent_con_evento
                        from .response_templates import get_response_for_intent
                        _p0 = estudiante.progresos.order_by('-fecha_inicio').first()
                        intent = detectar_intent_con_evento(
                            msg_body,
                            estudiante=estudiante,
                            curso=_p0.curso if _p0 else None,
                            modulo=_p0.modulo_actual if _p0 else None,
                        )
                        if intent != 'desconocido':
                            texto_respuesta = get_response_for_intent(intent, estudiante.nombre, estudiante_id=estudiante.id, mensaje_original=msg_body)
                        else:
                            texto_respuesta = "No entendí tu selección. Escribe *tomar 1* para escoger un curso o *menú* para volver."
                # — Enviar respuesta de selección de curso y CORTAR —
                try:
                    from twilio.rest import Client as TwilioClient
                    account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                    auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                    twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                    client_tw = TwilioClient(account_sid, auth_token)
                    destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                    # Check for multi-message or media markers
                    if texto_respuesta.startswith('[MULTI_MSG]'):
                        partes = texto_respuesta.replace('[MULTI_MSG]', '', 1).split('[SEP]')
                        for parte in partes:
                            if not parte.strip():
                                continue
                            import re as re_multi
                            parte_texto = parte.strip()
                            # [DELAY:N] — pausa intencional para entrega de videos
                            delay_m = re_multi.match(r'^\[DELAY:(\d+)\]$', parte_texto)
                            if delay_m:
                                import time; time.sleep(int(delay_m.group(1)))
                                continue
                            # Detectar Content Template
                            if parte_texto.startswith('[SEND_TEMPLATE:'):
                                tmpl_m = re_multi.match(r'\[SEND_TEMPLATE:(HX[a-f0-9]+)\]', parte_texto)
                                if tmpl_m:
                                    from .whatsapp_service import enviar_template_twilio
                                    tel_limpio_t = msg_from.replace('whatsapp:', '').replace('+', '')
                                    enviar_template_twilio(tel_limpio_t, tmpl_m.group(1))
                                import time; time.sleep(0.5)
                                continue
                            parte_media = None
                            media_m = re_multi.search(r'\[MEDIA:(.*?)\]', parte_texto)
                            if media_m:
                                parte_media = media_m.group(1).strip()
                                parte_texto = parte_texto.replace(media_m.group(0), '').strip()
                            if parte_media and not parte_texto:
                                parte_texto = TWILIO_CAPTION_ADJUNTO
                            mp = {'body': parte_texto, 'from_': str(twilio_number).strip(), 'to': str(destino).strip()}
                            if parte_media:
                                mp['media_url'] = [parte_media]
                            try:
                                client_tw.messages.create(**mp)
                            except Exception:
                                mp.pop('media_url', None)
                                client_tw.messages.create(**mp)
                            import time; time.sleep(0.5)
                    else:
                        media_url_sel = None
                        import re as re_sel
                        media_m = re_sel.search(r'\[MEDIA:(.*?)\]', texto_respuesta)
                        if media_m:
                            media_url_sel = media_m.group(1).strip()
                            texto_respuesta = texto_respuesta.replace(media_m.group(0), '').strip()
                        if media_url_sel and not texto_respuesta:
                            texto_respuesta = TWILIO_CAPTION_ADJUNTO
                        mp = {'body': texto_respuesta, 'from_': str(twilio_number).strip(), 'to': str(destino).strip()}
                        if media_url_sel:
                            mp['media_url'] = [media_url_sel]
                        try:
                            client_tw.messages.create(**mp)
                        except Exception:
                            mp.pop('media_url', None)
                            client_tw.messages.create(**mp)
                    WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta[:500], tipo='SENT')
                except Exception as e:
                    logger.error(f"❌ Error enviando selección curso: {e}")
                return  # CORTAR EJECUCIÓN

            # ============================================================
            # PRIORIDAD: Si el estudiante está interactuando con un AGENTE
            # (Darío, Facilitadora, Tutor IA, etc.), NO interceptar con
            # el gate de "solo listo" — dejar que el flujo caiga al handler
            # del agente más abajo en el código.
            # ============================================================
            estados_agente = [
                'esperando_respuesta_asistente',     # Darío
                'esperando_respuesta_reto',          # Facilitadora
                'esperando_respuesta_tutor_ia',      # Tutor legacy
                'esperando_respuesta_progreso',      # María
                'esperando_respuesta_modulo',         # Evaluación módulo
                'esperando_respuesta_pregunta_abierta_final',  # Respuesta final calificada por facilitadora
            ]
            if estudiante.estado_onboarding in estados_agente:
                logger.info(f"🤖 Agente activo ({estudiante.estado_onboarding}) — bypassing gate listo")
                # No interceptar — caerá al flujo EXISTENTE más abajo (handlers de agente)

            # Detectar "Mis cursos" → rama legacy deshabilitada (curso sin menú)
            elif False and msg_lower in ['1', 'mis cursos', 'cursos', '📚 mis cursos']:
                pass

            else:
                # 🆘 PQRS / ayuda con contexto — antes del gate «solo listo» (evita «No entendí»)
                try:
                    from .pqrs_agent import (
                        intentar_procesar_seguimiento_pqrs_whatsapp,
                        mensaje_activa_soporte,
                        respuesta_ayuda_con_ticket_abierto,
                    )
                    from .security_handler import procesar_solicitud_soporte

                    _resp_pqrs_gate = None
                    if mensaje_activa_soporte(msg_body):
                        _resp_pqrs_gate = respuesta_ayuda_con_ticket_abierto(estudiante, msg_body)
                        if not _resp_pqrs_gate:
                            _resp_pqrs_gate = procesar_solicitud_soporte(
                                estudiante, msg_body, 'curso_ayuda',
                            )
                    else:
                        _resp_pqrs_gate = intentar_procesar_seguimiento_pqrs_whatsapp(
                            estudiante, msg_body,
                        )

                    if _resp_pqrs_gate:
                        try:
                            from twilio.rest import Client as TwilioClient

                            account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                            auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                            twilio_number = getattr(
                                settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806',
                            )
                            client_tw = TwilioClient(account_sid, auth_token)
                            destino = (
                                f'whatsapp:{msg_from}'
                                if not msg_from.startswith('whatsapp:')
                                else msg_from
                            )
                            client_tw.messages.create(
                                body=_resp_pqrs_gate,
                                from_=str(twilio_number).strip(),
                                to=str(destino).strip(),
                            )
                            WhatsappLog.objects.create(
                                telefono=telefono_limpio,
                                mensaje=_resp_pqrs_gate[:500],
                                tipo='SENT',
                            )
                        except Exception as e:
                            logger.error('❌ Error enviando PQRS/ayuda en gate curso: %s', e)
                        return
                except Exception:
                    logger.exception('PQRS gate curso omitido')

                # 📚 Paso módulo: evaluación / reto — antes del gate "solo listo" (A, texto libre, etc.)
                if estado_chat == 'ACTIVO':
                    from .module_steps import procesar_respuesta_evaluacion_paso
                    from .models import ProgresoEstudiante

                    _prog_eval = (
                        ProgresoEstudiante.objects.filter(
                            estudiante=estudiante,
                            completado=False,
                        )
                        .order_by('-fecha_inicio')
                        .first()
                    )
                    _ctx_eval = estudiante.contexto_temporal or {}
                    _fid_eval = _ctx_eval.get('curso_activo_id')
                    if (
                        _fid_eval
                        and _prog_eval
                        and _prog_eval.curso_id != int(_fid_eval)
                    ):
                        _prog_eval = (
                            ProgresoEstudiante.objects.filter(
                                estudiante=estudiante,
                                completado=False,
                                curso_id=int(_fid_eval),
                            ).first()
                            or _prog_eval
                        )
                    if _prog_eval:
                        _resp_paso = procesar_respuesta_evaluacion_paso(
                            estudiante, _prog_eval, msg_body
                        )
                        if _resp_paso is not None:
                            logger.info(
                                "📚 [pasos] respuesta evaluación procesada en webhook | est=%s",
                                estudiante.id,
                            )
                            try:
                                from twilio.rest import Client as TwilioClient
                                import re as _re_paso

                                account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                                auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                                twilio_number = getattr(
                                    settings,
                                    'TWILIO_PHONE_NUMBER',
                                    'whatsapp:+573202948806',
                                )
                                client_tw = TwilioClient(account_sid, auth_token)
                                destino = (
                                    f'whatsapp:{msg_from}'
                                    if not msg_from.startswith('whatsapp:')
                                    else msg_from
                                )
                                twilio_from = str(twilio_number).strip()
                                destino_st = str(destino).strip()

                                def _log_twilio_segments_pasos(mensajes_enviados, parte_texto, parte_media):
                                    for _seg_idx, (mensaje, texto_enviado) in enumerate(
                                        mensajes_enviados, start=1
                                    ):
                                        texto_log = texto_enviado or parte_texto or (
                                            f'[MEDIA:{parte_media}]' if parte_media else ''
                                        )
                                        WhatsappLog.objects.create(
                                            telefono=telefono_limpio,
                                            mensaje=texto_log[:1500],
                                            mensaje_id=mensaje.sid,
                                            tipo='SENT',
                                        )

                                if _resp_paso.startswith('[MULTI_MSG]'):
                                    partes = _resp_paso.replace('[MULTI_MSG]', '', 1).split('[SEP]')
                                    for parte in partes:
                                        if not parte.strip():
                                            continue
                                        parte_texto = parte.strip()
                                        delay_m = _re_paso.match(
                                            r'^\[DELAY:(\d+)\]$', parte_texto
                                        )
                                        if delay_m:
                                            import time as _time_p

                                            _time_p.sleep(int(delay_m.group(1)))
                                            continue
                                        parte_media = None
                                        if '[MEDIA:' in parte_texto:
                                            mm = _re_paso.search(r'\[MEDIA:(.*?)\]', parte_texto)
                                            if mm:
                                                parte_media = (mm.group(1) or '').strip()
                                                parte_texto = parte_texto.replace(
                                                    mm.group(0), ''
                                                ).strip()
                                                if parte_media and youtube_hace_solo_enlace_en_texto(
                                                    parte_media
                                                ):
                                                    logger.warning(
                                                        '📎 [pasos] URL parece YouTube (no MP4 directo) | est=%s',
                                                        estudiante.id,
                                                    )
                                                if not parte_texto.strip() and parte_media:
                                                    parte_texto = TWILIO_CAPTION_ADJUNTO
                                        enviados_p = _enviar_mensaje_twilio_segmentado(
                                            client=client_tw,
                                            from_number=twilio_from,
                                            to_number=destino_st,
                                            body=parte_texto,
                                            media_url=parte_media,
                                        )
                                        _log_twilio_segments_pasos(enviados_p, parte_texto, parte_media)
                                        import time as _time_p

                                        _time_p.sleep(0.5)
                                else:
                                    parte_media = None
                                    parte_texto = (_resp_paso or '').strip()
                                    if '[MEDIA:' in parte_texto:
                                        mm = _re_paso.search(r'\[MEDIA:(.*?)\]', parte_texto)
                                        if mm:
                                            parte_media = (mm.group(1) or '').strip()
                                            parte_texto = parte_texto.replace(
                                                mm.group(0), ''
                                            ).strip()
                                            if not parte_texto.strip() and parte_media:
                                                parte_texto = TWILIO_CAPTION_ADJUNTO
                                    enviados_p = _enviar_mensaje_twilio_segmentado(
                                        client=client_tw,
                                        from_number=twilio_from,
                                        to_number=destino_st,
                                        body=parte_texto,
                                        media_url=parte_media,
                                    )
                                    _log_twilio_segments_pasos(enviados_p, parte_texto, parte_media)
                            except Exception as e:
                                logger.error(f"❌ Error enviando respuesta paso módulo: {e}", exc_info=True)
                            return

                keywords_corregir_curso = [
                    '4', 'corregir datos', 'corregir mis datos', 'cambiar datos', 'cambiar mis datos',
                    'me equivoqué', 'me equivoque', 'editar datos', 'modificar datos', 'modificar',
                    'datos incorrectos', 'mis datos', 'actualizar datos'
                ]

                if msg_lower in keywords_corregir_curso:
                    from .correccion_datos import iniciar_flujo_correccion
                    texto_respuesta = iniciar_flujo_correccion(estudiante)
                    try:
                        from twilio.rest import Client as TwilioClient
                        account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                        auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                        twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                        client_tw = TwilioClient(account_sid, auth_token)
                        destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                        client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                        WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
                    except Exception as e:
                        logger.error(f"❌ Error enviando corrección datos: {e}")
                    return

                if not _mensaje_indica_listo(msg_body):
                    if msg_body.strip() == '[AUDIO_NO_TRANSCRITO]':
                        texto_respuesta = "⚠️ No pude escuchar tu audio. Por favor intenta de nuevo o escríbeme. Para avanzar escribe *listo* o *continuar*."
                    else:
                        texto_respuesta = "No entendí. Si quieres avanzar de módulo escribe *listo* o *continuar*. Si necesitas ayuda, escribe *ayuda*."
                    try:
                        from twilio.rest import Client as TwilioClient
                        account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                        auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                        twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                        client_tw = TwilioClient(account_sid, auth_token)
                        destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                        client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                        WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
                    except Exception as e:
                        logger.error(f"❌ Error enviando respuesta no entendida: {e}")
                    return

            # Detectar "Mis puntos" (botón o texto) - rama legacy
            if msg_lower in ['2', 'mis puntos', 'puntos', '🏆 mis puntos']:
                from .whatsapp_service import enviar_gamificacion_visual
                enviar_gamificacion_visual(msg_from, estudiante)
                return
            
            # Detectar "Necesito ayuda" / PQRS / Soporte (todo unificado)
            elif msg_lower in ['3', 'necesito ayuda', 'ayuda', '🙋‍♂️ necesito ayuda', 'pqrs', 'soporte', 'queja', 'reclamo', 'solicitud']:
                from .security_handler import procesar_solicitud_soporte
                respuesta = procesar_solicitud_soporte(estudiante, msg_body, 'menu_ayuda')
                try:
                    from twilio.rest import Client as TwilioClient
                    account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                    auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                    twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                    client_tw = TwilioClient(account_sid, auth_token)
                    destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                    client_tw.messages.create(body=respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                    WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=respuesta, tipo='SENT')
                except Exception:
                    pass
                return
            
            # Detectar "corregir datos" / "me equivoqué" -> iniciar autocorreccion guiada
            elif msg_lower in ['4', 'corregir datos', 'corregir mis datos', 'cambiar datos', 'cambiar mis datos',
                               'me equivoqué', 'me equivoque', 'editar datos', 'modificar datos',
                               'datos incorrectos', 'mis datos', 'actualizar datos']:
                from .correccion_datos import iniciar_flujo_correccion
                texto_respuesta = iniciar_flujo_correccion(estudiante)
                try:
                    from twilio.rest import Client as TwilioClient
                    account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                    auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                    twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                    client_tw = TwilioClient(account_sid, auth_token)
                    destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                    client_tw.messages.create(body=texto_respuesta, from_=str(twilio_number).strip(), to=str(destino).strip())
                    WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta, tipo='SENT')
                except Exception as e:
                    logger.error(f"❌ Error enviando corrección datos: {e}")
                return
            
            # Detectar "menú" → enviar directamente al curso asignado (sin lista)
            elif msg_lower in ['menu', 'menú', 'inicio', 'hola']:
                if estudiante.estado_onboarding in estados_agente:
                    logger.info(
                        "menu/inicio/hola durante agente pedagógico (%s) — omitir reenvío de módulo",
                        estudiante.estado_onboarding,
                    )
                else:
                    from .models import Curso, ProgresoEstudiante
                    from .response_templates import obtener_video_url
                    org = estudiante.cliente
                    cursos = Curso.objects.filter(cliente=org, activo=True).order_by('orden', 'nombre') if org else Curso.objects.filter(activo=True).order_by('orden', 'nombre')
                    progreso_existente = ProgresoEstudiante.objects.filter(
                        estudiante=estudiante, completado=False, curso__activo=True
                    ).first()
                    curso = progreso_existente.curso if progreso_existente else cursos.first()
                    if curso:
                        progreso, _ = ProgresoEstudiante.objects.get_or_create(
                            estudiante=estudiante, curso=curso, defaults={'completado': False}
                        )
                        modulo = progreso.modulo_actual
                        if not modulo:
                            modulo = curso.modulos.order_by('numero').first()
                            if modulo:
                                progreso.modulo_actual = modulo
                                from .module_steps import reset_progreso_pasos_modulo
                                reset_progreso_pasos_modulo(progreso, save=False)
                                progreso.save(
                                    update_fields=[
                                        'modulo_actual',
                                        'paso_actual_modulo',
                                        'esperando_respuesta_evaluacion_paso',
                                        'paso_evaluacion_paso_id',
                                    ]
                                )
                        if modulo:
                            from .module_steps import (
                                modulo_usa_pasos,
                                mensaje_recordatorio_paso_actual,
                                pasos_activos_qs,
                                log_y_mensaje_modo_pasos_sin_pasos,
                            )
                            if modulo_usa_pasos(modulo):
                                if not pasos_activos_qs(modulo).exists():
                                    texto_respuesta = log_y_mensaje_modo_pasos_sin_pasos(
                                        modulo, 'views_menu_inicio'
                                    )
                                else:
                                    _np_m = pasos_activos_qs(modulo).count()
                                    if (
                                        progreso.paso_actual_modulo > _np_m
                                        and not progreso.esperando_respuesta_evaluacion_paso
                                    ):
                                        texto_respuesta = (
                                            "✅ Ya recibiste todo el material de esta unidad.\n\n"
                                            "Escribe *listo* para registrar tu avance y seguir 👇"
                                        )
                                    else:
                                        _rem_m = mensaje_recordatorio_paso_actual(progreso, modulo)
                                        texto_respuesta = _rem_m or (
                                            f"📖 *Módulo {modulo.numero}: {modulo.titulo}*\n\n"
                                            "Escribe *continuar* o *listo* cuando quieras seguir."
                                        )
                            else:
                                video_url = obtener_video_url(modulo)
                                archivos_multimedia = modulo.archivos_multimedia.filter(activo=True)
                                archivos_msg = ""
                                primera_media_url = None
                                extra_media_urls = []
                                if archivos_multimedia.exists():
                                    archivos_msg = ""
                                    for idx, archivo in enumerate(archivos_multimedia):
                                        icono = {'video': '🎥', 'imagen': '🖼️', 'infografia': '📊', 'pdf': '📄', 'audio': '🎵'}.get(archivo.tipo, '📁')
                                        url = archivo.get_url_para_envio()
                                        if url:
                                            if not primera_media_url:
                                                primera_media_url = url
                                                archivos_msg += f"\n{icono} {archivo.titulo} (adjunto)"
                                            else:
                                                extra_media_urls.append((url, archivo.titulo, icono))
                                                archivos_msg += f"\n{icono} {archivo.titulo} (adjunto)"
                                        else:
                                            archivos_msg += f"\n{icono} {archivo.titulo}"
                                if not archivos_multimedia.exists() and video_url:
                                    primera_media_url = video_url
                                msg_texto_menu = (
                                    f"📖 *Módulo {modulo.numero}: {modulo.titulo}*\n\n"
                                    f"{modulo.contenido}"
                                )
                                partes_menu = [msg_texto_menu]
                                if primera_media_url:
                                    partes_menu.append(parte_mensaje_con_media(primera_media_url))
                                for extra_url, extra_titulo, extra_icono in extra_media_urls:
                                    cap_m = f'{extra_icono} {extra_titulo}'.strip() if extra_titulo else None
                                    partes_menu.append(parte_mensaje_con_media(extra_url, cap_m))
                                if len(partes_menu) > 1:
                                    texto_respuesta = "[MULTI_MSG]" + "[SEP]".join(partes_menu)
                                else:
                                    texto_respuesta = msg_texto_menu
                        else:
                            texto_respuesta = f"📚 Tu curso *{curso.nombre}* aún no tiene módulos configurados."
                    else:
                        texto_respuesta = (
                            "📚 Aún no tienes un curso asignado. "
                            "Tu coordinador te lo asignará pronto.\n\n"
                            "Si acabas de reportar un problema, cuando quieras retomar escribe *listo* o *menú*."
                        )
                    try:
                        from twilio.rest import Client as TwilioClient
                        account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                        auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                        twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                        client_tw = TwilioClient(account_sid, auth_token)
                        destino = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
                        twilio_from = str(twilio_number).strip()

                        if texto_respuesta.startswith('[MULTI_MSG]'):
                            import re as re_menu_m
                            partes_mm = texto_respuesta.replace('[MULTI_MSG]', '', 1).split('[SEP]')
                            for parte_mm in partes_mm:
                                if not parte_mm.strip():
                                    continue
                                parte_txt = parte_mm.strip()
                                delay_mm = re_menu_m.match(r'^\[DELAY:(\d+)\]$', parte_txt)
                                if delay_mm:
                                    import time
                                    time.sleep(int(delay_mm.group(1)))
                                    continue
                                media_mm = re_menu_m.search(r'\[MEDIA:(.*?)\]', parte_txt)
                                parte_media_mm = None
                                if media_mm:
                                    parte_media_mm = media_mm.group(1).strip()
                                    parte_txt = parte_txt.replace(media_mm.group(0), '').strip()
                                cuerpo_mm = parte_txt.strip() if parte_txt else ''
                                if parte_media_mm and not cuerpo_mm:
                                    cuerpo_mm = TWILIO_CAPTION_ADJUNTO
                                mp_mm = {
                                    'body': cuerpo_mm if cuerpo_mm else ' ',
                                    'from_': twilio_from,
                                    'to': str(destino).strip(),
                                }
                                if parte_media_mm:
                                    mp_mm['media_url'] = [parte_media_mm]
                                try:
                                    client_tw.messages.create(**mp_mm)
                                except Exception as ex_mm:
                                    if '63019' in str(ex_mm) and parte_media_mm:
                                        mp_mm.pop('media_url', None)
                                        mp_mm['body'] = (mp_mm.get('body') or '').strip() + f"\n\n📎 Archivo: {parte_media_mm}"
                                        client_tw.messages.create(**mp_mm)
                                    else:
                                        raise
                                import time
                                time.sleep(0.5)
                                WhatsappLog.objects.create(
                                    telefono=telefono_limpio, mensaje=(cuerpo_mm or parte_txt)[:500], tipo='SENT'
                                )
                        else:
                            media_url_menu = None
                            import re as re_menu
                            media_m = re_menu.search(r'\[MEDIA:(.*?)\]', texto_respuesta)
                            if media_m:
                                media_url_menu = media_m.group(1).strip()
                                texto_respuesta = texto_respuesta.replace(media_m.group(0), '').strip()
                            mp = {'body': texto_respuesta, 'from_': twilio_from, 'to': str(destino).strip()}
                            if media_url_menu:
                                mp['media_url'] = [media_url_menu]
                            try:
                                client_tw.messages.create(**mp)
                            except Exception:
                                mp.pop('media_url', None)
                                client_tw.messages.create(**mp)
                            WhatsappLog.objects.create(telefono=telefono_limpio, mensaje=texto_respuesta[:500], tipo='SENT')
                    except Exception as e:
                        logger.error(f"❌ Error enviando curso: {e}")
                    return
        
        # ============================================================
        # PQRS: seguimiento de ticket abierto (máx. 2 preguntas de clarificación)
        # ============================================================
        pqrs_atendido = False
        respuesta_pqrs = None
        try:
            from .pqrs_agent import intentar_procesar_seguimiento_pqrs_whatsapp

            respuesta_pqrs = intentar_procesar_seguimiento_pqrs_whatsapp(estudiante, msg_body)
            if respuesta_pqrs:
                pqrs_atendido = True
                logger.info(
                    '🆘 PQRS seguimiento — respuesta automática estudiante_id=%s',
                    estudiante.id,
                )
        except Exception as e:
            logger.exception('PQRS seguimiento omitido: %s', e)

        # ============================================================
        # FLUJO EXISTENTE: Procesamiento normal (IA tutors, módulos, etc.)
        # ============================================================
        
        # 3. 🛡️ PRIORIDAD 1: Verificar seguridad (Habeas Data) - Legacy
        from .security_handler import verificar_seguridad_completa
        if pqrs_atendido:
            bloqueado = True
            respuesta_seguridad = respuesta_pqrs
        else:
            bloqueado, respuesta_seguridad, estudiante = verificar_seguridad_completa(
                estudiante,
                msg_body,
                telefono_limpio,
                numero_destino=msg_to,
            )
        print(f"🛡️ Seguridad: bloqueado={bloqueado} | estudiante={estudiante} | estado={getattr(estudiante, 'estado_onboarding', 'N/A')}", flush=True)
        
        # Default safety - will be overwritten by any branch below
        texto_respuesta = "No entendí. Si quieres avanzar de módulo escribe *listo*. Si necesitas ayuda, escribe *ayuda*."
        
        if bloqueado:
            print(f"🛡️ Bloqueado por seguridad/habeas data", flush=True)
            texto_respuesta = respuesta_seguridad
        else:
            # Si el contexto sigue en Darío pero el estado quedó desincronizado, continuar_leccion
            # podía ejecutarse antes que la rama del asistente y saltar la facilitadora.
            try:
                _ctx_dario_sync = estudiante.contexto_temporal or {}
                if _ctx_dario_sync.get('tipo') == 'asistente_dario':
                    if estudiante.estado_onboarding != 'esperando_respuesta_asistente':
                        _ob_prev = estudiante.estado_onboarding
                        estudiante.estado_onboarding = 'esperando_respuesta_asistente'
                        estudiante.save(update_fields=['estado_onboarding'])
                        logger.warning(
                            "🔄 Darío: estado resincronizado %s → esperando_respuesta_asistente "
                            "(ctx=asistente_dario) estudiante_id=%s",
                            _ob_prev,
                            estudiante.id,
                        )
            except Exception as e:
                logger.warning("⚠️ Darío sync omitido: %s", e)

            if estudiante.estado_onboarding == 'esperando_codigo_empleabilidad':
                from .models import AliadoEmpleabilidad, MisionEmpleabilidad
                from .gamificacion import PerfilGamificacion, Badge, BadgeEstudiante
                ctx_emp = estudiante.contexto_temporal or {}
                aliado_id = ctx_emp.get('aliado_empleabilidad_objetivo_id')
                mision_id = ctx_emp.get('mision_empleabilidad_id')
                aliado = None
                if aliado_id:
                    aliado = AliadoEmpleabilidad.objects.filter(id=aliado_id, vacantes_activas=True).first()
                mision = None
                if mision_id:
                    mision = MisionEmpleabilidad.objects.filter(id=mision_id, estudiante=estudiante).first()

                if aliado and msg_body.strip().lower() == str(aliado.codigo_secreto).strip().lower():
                    perfil, _ = PerfilGamificacion.objects.get_or_create(estudiante=estudiante)
                    puntos_cfg = int(getattr(estudiante.cliente, 'empleabilidad_puntos_validacion', 30) or 30)
                    perfil.agregar_puntos(puntos_cfg, f"Radar Empleabilidad: {aliado.nombre_empresa}")
                    badge = Badge.objects.filter(tipo='ESPECIAL', activo=True, nombre__icontains='emple').first()
                    if badge:
                        BadgeEstudiante.objects.get_or_create(estudiante=estudiante, badge=badge)

                    if mision and mision.estado != 'completada':
                        mision.estado = 'completada'
                        mision.codigo_validado = True
                        mision.puntos_otorgados = puntos_cfg
                        mision.fecha_completada = timezone.now()
                        mision.save(update_fields=['estado', 'codigo_validado', 'puntos_otorgados', 'fecha_completada'])

                    from .tasks import enviar_email_org_admin_async
                    try:
                        asunto = f"Match empleabilidad: {estudiante.nombre} - {aliado.nombre_empresa}"
                        mensaje_html = (
                            f"<p>El estudiante <strong>{estudiante.nombre}</strong> "
                            f"(tel: {estudiante.telefono}) validó el código secreto de "
                            f"<strong>{aliado.nombre_empresa}</strong>.</p>"
                            f"<p>Iniciar contacto para proceso de contratación.</p>"
                        )
                        enviar_email_org_admin_async.delay(estudiante.id, asunto, mensaje_html)
                    except Exception as e:
                        logger.warning(f"⚠️ No se pudo encolar notificación de empleabilidad: {e}")

                    ctx_emp['ultimo_match_empleabilidad_aliado_id'] = aliado.id
                    ctx_emp['ultimo_match_empleabilidad_fecha'] = timezone.now().isoformat()
                    estudiante.contexto_temporal = ctx_emp
                    estudiante.estado_onboarding = 'curso_finalizado'
                    estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])
                    texto_respuesta = (
                        f"🏆 *¡Logro desbloqueado!*\n\n"
                        f"Validaste el código de *{aliado.nombre_empresa}*.\n"
                        "✅ Ya notificamos al equipo para iniciar el proceso de empleabilidad."
                    )
                    estudiante.estado_onboarding = 'curso_finalizado'
                else:
                    if mision and mision.estado == 'descubierta':
                        mision.estado = 'reclamada'
                        mision.fecha_reclamada = timezone.now()
                        mision.save(update_fields=['estado', 'fecha_reclamada'])
                    texto_respuesta = (
                        "🔐 El código no coincide.\n\n"
                        "Verifica el código secreto en la entrada de la empresa y vuelve a enviarlo."
                    )
                    estudiante.estado_onboarding = 'esperando_codigo_empleabilidad'
                estudiante.save(update_fields=['estado_onboarding'])

            elif estudiante.estado_onboarding == 'esperando_respuesta_pregunta_abierta_final':
                from .models import PreguntaAbiertaFinalCurso, RespuestaAbiertaFinal, ProgresoEstudiante
                if msg_body.strip() == '[AUDIO_NO_TRANSCRITO]':
                    texto_respuesta = (
                        "⚠️ No pude escuchar tu audio. Por favor intenta de nuevo "
                        "o escríbeme tu respuesta abierta final."
                    )
                else:
                    ctx_open = estudiante.contexto_temporal or {}
                    pregunta_id = ctx_open.get('pregunta_abierta_final_id')
                    progreso_id = ctx_open.get('progreso_id')
                    pregunta = PreguntaAbiertaFinalCurso.objects.filter(id=pregunta_id, activa=True).first()
                    progreso = ProgresoEstudiante.objects.filter(id=progreso_id).first() if progreso_id else None

                    if pregunta:
                        logger.info(
                            "🧾 Pregunta abierta final en contexto | estudiante_id=%s | curso_id=%s | pregunta_id=%s | orden=%s | texto=%s",
                            estudiante.id,
                            getattr(pregunta.curso, 'id', None),
                            pregunta.id,
                            getattr(pregunta, 'orden', None),
                            (pregunta.pregunta or '')[:180],
                        )
                        respuesta_final, _ = RespuestaAbiertaFinal.objects.update_or_create(
                            pregunta=pregunta,
                            estudiante=estudiante,
                            defaults={
                                'curso': pregunta.curso,
                                'progreso': progreso,
                                'respuesta_texto': msg_body,
                                'fecha_respuesta': timezone.now(),
                                'estado': 'pendiente',
                            }
                        )
                        logger.info(
                            "📝 Respuesta abierta final registrada | estudiante_id=%s | curso_id=%s | progreso_id=%s",
                            estudiante.id,
                            getattr(pregunta.curso, 'id', None),
                            getattr(progreso, 'id', None),
                        )

                        # Evaluar con la misma rúbrica de facilitadora para dar feedback
                        # inmediato y mantener coherencia de gamificación.
                        curso_obj = pregunta.curso if pregunta and pregunta.curso_id else (progreso.curso if progreso else None)
                        modulos_eval = list(curso_obj.modulos.all().order_by('numero')) if curso_obj else []
                        puntaje_final_10 = 7
                        feedback_final = (
                            "1. Gracias por su respuesta; usted sí propone una línea de acción.\n\n"
                            "2. Para subir nivel, faltó precisión objetiva en dos partes: "
                            "(a) diagnóstico: qué señales mediría, cuánto y en qué tiempo; "
                            "(b) control: qué acción exacta aplicaría, con qué frecuencia y criterio de verificación.\n\n"
                            "3. Puntaje total: 7/10\n"
                            "4. Desglose: Enfoque 2/3 | Fundamentación 3/4 | Claridad 2/3\n\n"
                            "Diagnóstico: parcial | Acción/Control: parcial."
                        )
                        try:
                            from .tutor_ia_modulo import evaluar_reto_facilitador
                            from core.gamificacion_modo import (
                                get_modo_gamificacion,
                                modo_usa_calificacion,
                                gamificacion_otorga_puntos,
                                formatear_nota,
                                registrar_nota_gamificacion,
                                resumen_calificaciones_estudiante,
                            )

                            modo_gami = get_modo_gamificacion(
                                getattr(estudiante, 'cliente', None),
                            )
                            puntaje_final, feedback_final = evaluar_reto_facilitador(
                                modulos_eval,
                                msg_body,
                                pregunta.pregunta,
                                estudiante_nombre=estudiante.nombre or "Estudiante",
                                curso_nombre=(getattr(curso_obj, 'nombre', None) if curso_obj else None),
                                modo_gamificacion=modo_gami,
                            )
                        except Exception as e:
                            logger.warning(f"⚠️ No se pudo evaluar pregunta abierta final con IA: {e}")

                        respuesta_final.estado = 'calificada'
                        if modo_usa_calificacion(getattr(estudiante, 'cliente', None)):
                            nota = float(puntaje_final)
                            respuesta_final.calificacion = int(round(nota * 10))
                        else:
                            respuesta_final.calificacion = int(
                                max(1, min(10, int(puntaje_final))) * 10,
                            )
                        respuesta_final.retroalimentacion = feedback_final
                        respuesta_final.fecha_calificacion = timezone.now()
                        respuesta_final.save(update_fields=['estado', 'calificacion', 'retroalimentacion', 'fecha_calificacion'])

                        puntos_msg = ""
                        try:
                            if modo_usa_calificacion(getattr(estudiante, 'cliente', None)):
                                nota_f = float(puntaje_final)
                                registrar_nota_gamificacion(
                                    estudiante,
                                    nota_f,
                                    'pregunta_abierta',
                                    curso=curso_obj,
                                    detalle='Pregunta abierta final',
                                )
                                res_n = resumen_calificaciones_estudiante(
                                    estudiante,
                                    curso_obj.id if curso_obj else None,
                                )
                                prom = res_n.get('promedio')
                                extra_prom = (
                                    f"\n📊 *Promedio acumulado:* {formatear_nota(prom)}/5"
                                    if prom is not None else ''
                                )
                                puntos_msg = (
                                    f"\n\n📋 *Nota:* {formatear_nota(nota_f)}/5{extra_prom}"
                                )
                            elif gamificacion_otorga_puntos(
                                getattr(estudiante, 'cliente', None), curso_obj,
                            ):
                                from .gamificacion import PerfilGamificacion
                                from .response_templates import _barra_progreso

                                perfil, _ = PerfilGamificacion.objects.get_or_create(estudiante=estudiante)
                                puntaje_10 = int(puntaje_final)
                                puntos_abierta = int(max(1, min(10, puntaje_10)) * 5)
                                perfil.agregar_puntos(
                                    puntos_abierta,
                                    f"Pregunta abierta final: {puntaje_10}/10",
                                )
                                perfil.refresh_from_db()

                                porcentaje = progreso.porcentaje_avance() if progreso else 100
                                barra = _barra_progreso(porcentaje)
                                puntos_msg = (
                                    f"\n\n💰 *+{puntos_abierta} puntos* → Total: *{perfil.puntos_totales} pts*\n"
                                    f"{barra} {porcentaje}%"
                                )
                        except Exception as e:
                            logger.warning(f"⚠️ No se pudieron aplicar puntos por pregunta abierta final: {e}")

                        # Si hay más preguntas abiertas pendientes (máximo 3),
                        # continuar en secuencia antes de emitir certificado.
                        siguiente_pregunta = _pregunta_abierta_final_pendiente(estudiante, progreso) if progreso else None
                        if siguiente_pregunta:
                            logger.info(
                                "🔁 Siguiente pregunta abierta final | estudiante_id=%s | curso_id=%s | pregunta_id=%s | orden=%s | texto=%s",
                                estudiante.id,
                                progreso.curso.id,
                                siguiente_pregunta.id,
                                getattr(siguiente_pregunta, 'orden', None),
                                (siguiente_pregunta.pregunta or '')[:180],
                            )
                            estudiante.contexto_temporal = {
                                'tipo': 'pregunta_abierta_final',
                                'curso_id': progreso.curso.id,
                                'progreso_id': progreso.id,
                                'pregunta_abierta_final_id': siguiente_pregunta.id,
                            }
                            estudiante.estado_onboarding = 'esperando_respuesta_pregunta_abierta_final'
                            estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])

                            texto_respuesta = (
                                "[MULTI_MSG]"
                                f"📋 *Facilitadora*\n\n{feedback_final}{puntos_msg}"
                                "[SEP]"
                                "📝 *Siguiente pregunta abierta final*\n\n"
                                f"{siguiente_pregunta.pregunta}\n\n"
                                "✍️ Responde con tus propias palabras (texto o audio)."
                            )
                        else:
                            # Mantener el flujo general: después de la última pregunta final,
                            # se entrega certificado y se cierra el curso.
                            msg_cert_img = ""
                            if curso_obj:
                                try:
                                    from .certificado_service import crear_certificado_automatico, obtener_url_certificado_twilio
                                    cert = crear_certificado_automatico(estudiante, curso_obj)
                                    if cert and cert.archivo_imagen:
                                        cert_url = obtener_url_certificado_twilio(cert)
                                        if cert_url:
                                            msg_cert_img = f"🎓 *¡Tu certificado!*\n\n[MEDIA:{cert_url}]"
                                        else:
                                            s3_key = str(cert.archivo_imagen.name)
                                            cert_url = f"https://eki-produccion.s3.us-east-2.amazonaws.com/{s3_key}"
                                            msg_cert_img = f"🎓 *¡Tu certificado!*\n\n[MEDIA:{cert_url}]"
                                    elif cert and cert.archivo_pdf:
                                        msg_cert_img = f"🎓 *¡Tu certificado!*\n📄 Descárgalo aquí: {cert.archivo_pdf.url}"
                                    else:
                                        msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."
                                except Exception as e:
                                    logger.error(f"❌ Error certificado tras pregunta abierta final: {e}", exc_info=True)
                                    msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."

                            radar_msg = ""
                            if _activar_radar_empleabilidad_si_aplica(estudiante):
                                radar_msg = (
                                    "📍 *¡Radar de Empleos desbloqueado!*\n\n"
                                    "Ve al parque principal de Subachoque y envíame tu *Ubicación* "
                                    "usando el clip de WhatsApp (📎)."
                                )

                            estudiante.estado_onboarding = 'curso_finalizado'
                            estudiante.contexto_temporal = None
                            estudiante.save(update_fields=['estado_onboarding', 'contexto_temporal'])

                            partes_finales = [f"📋 *Facilitadora*\n\n{feedback_final}{puntos_msg}"]
                            if radar_msg:
                                partes_finales.append(radar_msg)
                            if msg_cert_img:
                                partes_finales.append(msg_cert_img)
                            texto_respuesta = "[MULTI_MSG]" + "[SEP]".join(partes_finales)
                    else:
                        texto_respuesta = (
                            "No encuentro una pregunta abierta final activa para tu curso en este momento. "
                            "Escribe *ayuda* para soporte."
                        )

            # v1.9.8g: Post-certificate cutoff — no more interaction
            elif estudiante.estado_onboarding == 'curso_finalizado':
                print(f"🚫 Curso finalizado — sin interacción post-certificado")
                # Check if student has a new active course
                from .models import ProgresoEstudiante
                nuevo_progreso = ProgresoEstudiante.objects.filter(
                    estudiante=estudiante, completado=False
                ).first()
                if nuevo_progreso:
                    # New course assigned — reset state
                    estudiante.estado_onboarding = 'completado'
                    estudiante.save()
                    texto_respuesta = (
                        f"🎉 *¡Tienes un nuevo curso asignado!*\n\n"
                        f"📚 *{nuevo_progreso.curso.nombre}*\n\n"
                        f"Escribe *listo* para comenzar."
                    )
                else:
                    texto_respuesta = (
                        "✅ *Tu proceso del curso ya finalizó y tu certificado fue enviado.*\n\n"
                        "Cuando tu organización te inscriba en un nuevo curso, te notificaremos. "
                        "¡Gracias por tu participación! 🎓"
                    )
            
            # v1.9.8g: Asistente (compañero) — hasta 2 preguntas antes del reto
            elif estudiante.estado_onboarding == 'esperando_respuesta_asistente':
                ctx = estudiante.contexto_temporal or {}
                preguntas_hechas = ctx.get('preguntas_hechas', 0)
                modulos_reto_ids = ctx.get('modulos_reto_ids', [])
                progreso_id = ctx.get('progreso_id')
                _cli = estudiante.cliente if getattr(estudiante, 'cliente_id', None) else None
                _na_cli = (
                    (_cli.nombre_agente_asistente if _cli and getattr(_cli, 'nombre_agente_asistente', None) else '')
                    or ''
                )
                _na_curso = ''
                if progreso_id:
                    try:
                        from .models import ProgresoEstudiante
                        _pp = ProgresoEstudiante.objects.select_related('curso').get(id=progreso_id)
                        _na_curso = (_pp.curso.nombre_agente_asistente if _pp.curso else '') or ''
                    except ProgresoEstudiante.DoesNotExist:
                        pass
                nombre_asistente = (_na_cli or _na_curso or 'Darío').strip() or 'Darío'
                print(f"💬 Asistente ({nombre_asistente}): Esperando respuesta del compañero IA")
                
                msg_lower = msg_body.strip().lower()
                if msg_lower in ['ayuda', 'soporte', 'ticket']:
                    from .security_handler import procesar_solicitud_soporte
                    texto_respuesta = procesar_solicitud_soporte(estudiante, msg_body, 'asistente_ayuda')
                elif msg_body.strip() == '[AUDIO_NO_TRANSCRITO]':
                    # Audio no pudo ser transcrito — NO contar como pregunta
                    print(f"🎤 Audio no transcrito en asistente — pidiendo reintento")
                    preguntas_restantes = 2 - preguntas_hechas
                    texto_respuesta = (
                        f"💬 *{nombre_asistente}*\n\n"
                        f"⚠️ No pude escuchar tu audio. Por favor intenta de nuevo "
                        f"o escríbeme tu pregunta.\n\n"
                        f"Te quedan {preguntas_restantes} pregunta(s). "
                        f"Si no tienes preguntas, escribe *listo*."
                    )
                elif _mensaje_indica_listo(msg_body) or preguntas_hechas >= 2:
                    # Flujo exigido: Darío -> Facilitadora (reto) al escribir listo
                    print("🎯 Asistente terminó → Activando Facilitadora con reto")
                    from .models import ProgresoEstudiante
                    from .tutor_ia_modulo import (
                        cargar_modulos_reto,
                        generar_reto_facilitador,
                        listar_modulos_cobertura_reto,
                    )
                    from .models import Modulo as ModuloRetoCtx
                    try:
                        progreso = ProgresoEstudiante.objects.get(id=progreso_id)
                    except ProgresoEstudiante.DoesNotExist:
                        progreso = None
                    modulos_reto = cargar_modulos_reto(
                        modulos_reto_ids, progreso.curso_id if progreso else None
                    )
                    if not modulos_reto and progreso and ctx.get('modulo_id'):
                        _mchk = ModuloRetoCtx.objects.filter(
                            id=ctx['modulo_id'], curso_id=progreso.curso_id
                        ).first()
                        if _mchk:
                            modulos_reto = listar_modulos_cobertura_reto(_mchk, progreso.curso)
                            modulos_reto_ids = [m.id for m in modulos_reto]

                    if modulos_reto and progreso:
                        _cliente = estudiante.cliente if hasattr(estudiante, 'cliente') and estudiante.cliente else None
                        nombre_tutor = (
                            (_cliente.nombre_agente_tutor if _cliente and hasattr(_cliente, 'nombre_agente_tutor') and _cliente.nombre_agente_tutor else '') or
                            progreso.curso.nombre_agente_tutor or 'Claudia'
                        )
                        reto = generar_reto_facilitador(
                            modulos_reto,
                            progreso.curso.nombre,
                            estudiante_nombre=estudiante.nombre or "Estudiante",
                            preguntas_ejemplo=progreso.curso.preguntas_ejemplo_ia or ""
                        )
                        _prev_ts = (estudiante.contexto_temporal or {}).get('_ts_leccion', 0)
                        estudiante.contexto_temporal = {
                            'tipo': 'reto_facilitador',
                            'modulos_reto_ids': modulos_reto_ids,
                            'reto_texto': reto,
                            'progreso_id': progreso_id,
                            'es_final': ctx.get('es_reto_final', False),
                            '_ts_leccion': _prev_ts,
                        }
                        estudiante.estado_onboarding = 'esperando_respuesta_reto'
                        estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])
                        texto_respuesta = (
                            f"📋 *{nombre_tutor}*\n\n"
                            f"{reto}\n\n"
                            "✍️ _Escriba o envíe un audio con su respuesta._"
                        )
                    else:
                        logger.warning(
                            "reto asistente vacío | progreso_id=%s modulo_ctx=%s ids_ctx=%s",
                            progreso_id,
                            (ctx or {}).get('modulo_id'),
                            modulos_reto_ids,
                        )
                        estudiante.estado_onboarding = 'completado'
                        estudiante.contexto_temporal = None
                        estudiante.save(update_fields=['estado_onboarding', 'contexto_temporal'])
                        texto_respuesta = (
                            "Seguimos con tu curso. Escribí *listo* para continuar "
                            "o *ayuda* si necesitás soporte."
                        )
                else:
                    # Student asked a question to Darío — answer from RAG (max 2)
                    preguntas_hechas += 1
                    from .models import ProgresoEstudiante
                    from .tutor_ia_modulo import cargar_modulos_reto, generar_respuesta_asistente
                    try:
                        _pr_dario = ProgresoEstudiante.objects.get(id=progreso_id) if progreso_id else None
                    except ProgresoEstudiante.DoesNotExist:
                        _pr_dario = None
                    modulos_reto = cargar_modulos_reto(
                        modulos_reto_ids, _pr_dario.curso_id if _pr_dario else None
                    )
                    
                    respuesta_dario = generar_respuesta_asistente(
                        modulos_reto,
                        msg_body,
                        estudiante_nombre=estudiante.nombre or "Estudiante",
                        nombre_asistente=nombre_asistente,
                    )
                    
                    ctx['preguntas_hechas'] = preguntas_hechas
                    estudiante.contexto_temporal = ctx
                    estudiante.save()
                    
                    if preguntas_hechas >= 2:
                        texto_respuesta = (
                            f"💬 *{nombre_asistente}*\n\n{respuesta_dario}\n\n"
                            f"Ya respondí tus 2 preguntas. Ahora la facilitadora te tiene un reto. "
                            f"Escribe *listo* cuando estés preparado."
                        )
                    else:
                        texto_respuesta = (
                            f"💬 *{nombre_asistente}*\n\n{respuesta_dario}\n\n"
                            f"¿Tienes otra pregunta? Te queda {2 - preguntas_hechas} pregunta más.\n"
                            f"Ejemplos: \"¿qué reviso primero en campo?\", \"¿cómo confirmo el diagnóstico?\". "
                            f"Si no, escribe *listo*."
                        )
            
            # v1.9.8g: Facilitadora — evaluando respuesta al reto
            elif estudiante.estado_onboarding == 'esperando_respuesta_reto':
                print(f"📋 Facilitadora: Evaluando respuesta al reto")
                ctx = estudiante.contexto_temporal or {}
                modulos_reto_ids = ctx.get('modulos_reto_ids', [])
                reto_texto = ctx.get('reto_texto', '')
                progreso_id = ctx.get('progreso_id')
                
                msg_lower = msg_body.strip().lower()
                if msg_lower in ['ayuda', 'soporte', 'ticket']:
                    from .security_handler import procesar_solicitud_soporte
                    texto_respuesta = procesar_solicitud_soporte(estudiante, msg_body, 'reto_ayuda')
                elif msg_body.strip() == '[AUDIO_NO_TRANSCRITO]':
                    # Audio no pudo ser transcrito — pedir reintento sin evaluar
                    print(f"🎤 Audio no transcrito en reto — pidiendo reintento")
                    texto_respuesta = (
                        "⚠️ No pude escuchar tu audio. Por favor intenta de nuevo "
                        "o escríbeme tu respuesta al reto.\n\n"
                        "✍️ _Escriba o envíe un audio con su respuesta._"
                    )
                else:
                    from .models import ProgresoEstudiante
                    from .tutor_ia_modulo import cargar_modulos_reto, evaluar_reto_facilitador
                    try:
                        progreso = ProgresoEstudiante.objects.get(id=progreso_id)
                    except ProgresoEstudiante.DoesNotExist:
                        progreso = None
                    modulos_reto = cargar_modulos_reto(
                        modulos_reto_ids, progreso.curso_id if progreso else None
                    )
                    
                    from core.gamificacion_modo import get_modo_gamificacion, construir_mensaje_evaluacion_reto

                    _cliente = estudiante.cliente if hasattr(estudiante, 'cliente') and estudiante.cliente else None
                    modo_gami = get_modo_gamificacion(_cliente)
                    puntaje, feedback = evaluar_reto_facilitador(
                        modulos_reto, msg_body, reto_texto,
                        estudiante_nombre=estudiante.nombre or "Estudiante",
                        curso_nombre=(progreso.curso.nombre if progreso else None),
                        modo_gamificacion=modo_gami,
                    )

                    nombre_tutor = (
                        (_cliente.nombre_agente_tutor if _cliente and hasattr(_cliente, 'nombre_agente_tutor') and _cliente.nombre_agente_tutor else '') or
                        'Claudia'
                    )
                    try:
                        progreso = ProgresoEstudiante.objects.get(id=progreso_id)
                        nombre_tutor = progreso.curso.nombre_agente_tutor or nombre_tutor
                    except ProgresoEstudiante.DoesNotExist:
                        progreso = None

                    msg_eval = construir_mensaje_evaluacion_reto(
                        estudiante, progreso, puntaje, feedback, nombre_tutor,
                    )
                    
                    es_final = ctx.get('es_final', False)
                    
                    if es_final and progreso:
                        pregunta_abierta = None
                        pregunta_abierta_ctx_id = ctx.get('pregunta_abierta_final_id')
                        if pregunta_abierta_ctx_id:
                            from .models import PreguntaAbiertaFinalCurso, RespuestaAbiertaFinal
                            pregunta_ctx = PreguntaAbiertaFinalCurso.objects.filter(
                                id=pregunta_abierta_ctx_id,
                                curso=progreso.curso,
                                activa=True,
                            ).first()
                            if pregunta_ctx:
                                ya_respondio_ctx = RespuestaAbiertaFinal.objects.filter(
                                    pregunta=pregunta_ctx,
                                    estudiante=estudiante,
                                ).exists()
                                if not ya_respondio_ctx:
                                    pregunta_abierta = pregunta_ctx

                        if not pregunta_abierta:
                            pregunta_abierta = _pregunta_abierta_final_pendiente(estudiante, progreso)

                        if pregunta_abierta:
                            logger.info(
                                "🧭 Pregunta abierta final seleccionada post-reto | estudiante_id=%s | curso_id=%s | pregunta_id=%s | orden=%s | texto=%s",
                                estudiante.id,
                                progreso.curso.id,
                                pregunta_abierta.id,
                                getattr(pregunta_abierta, 'orden', None),
                                (pregunta_abierta.pregunta or '')[:180],
                            )
                            estudiante.contexto_temporal = {
                                'tipo': 'pregunta_abierta_final',
                                'curso_id': progreso.curso.id,
                                'progreso_id': progreso.id,
                                'pregunta_abierta_final_id': pregunta_abierta.id,
                            }
                            estudiante.estado_onboarding = 'esperando_respuesta_pregunta_abierta_final'
                            estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])
                            texto_respuesta = (
                                f"{msg_eval}\n\n"
                                "📝 *Antes del cierre final*, responde esta pregunta abierta:\n\n"
                                f"{pregunta_abierta.pregunta}\n\n"
                                "✍️ Tu facilitadora revisará y calificará tu respuesta."
                            )
                        else:
                            logger.info(
                                "⚠️ Post-reto final sin pregunta abierta final | estudiante_id=%s | curso_id=%s | ctx_pregunta_id=%s",
                                estudiante.id,
                                progreso.curso.id,
                                pregunta_abierta_ctx_id,
                            )
                            # v1.9.8h: Final reto — issue certificate
                            estudiante.estado_onboarding = 'curso_finalizado'
                            estudiante.contexto_temporal = None
                            estudiante.save(update_fields=['estado_onboarding', 'contexto_temporal'])
                            
                            msg_cert_img = ""
                            try:
                                from .certificado_service import crear_certificado_automatico, obtener_url_certificado_twilio
                                cert = crear_certificado_automatico(estudiante, progreso.curso)
                                if cert and cert.archivo_imagen:
                                    cert_url = obtener_url_certificado_twilio(cert)
                                    if cert_url:
                                        msg_cert_img = f"🎓 *¡Tu certificado!*\n\n[MEDIA:{cert_url}]"
                                    else:
                                        s3_key = str(cert.archivo_imagen.name)
                                        cert_url = f"https://eki-produccion.s3.us-east-2.amazonaws.com/{s3_key}"
                                        msg_cert_img = f"🎓 *¡Tu certificado!*\n\n[MEDIA:{cert_url}]"
                                elif cert and cert.archivo_pdf:
                                    msg_cert_img = f"🎓 *¡Tu certificado!*\n📄 Descárgalo aquí: {cert.archivo_pdf.url}"
                                else:
                                    msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."
                            except Exception as e:
                                logger.error(f"❌ Error certificado post-reto final: {e}", exc_info=True)
                                msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."

                            radar_msg = ""
                            if _activar_radar_empleabilidad_si_aplica(estudiante):
                                radar_msg = (
                                    "📍 *¡Radar de Empleos desbloqueado!*\n\n"
                                    "Ve al parque principal de Subachoque y envíame tu *Ubicación* "
                                    "usando el clip de WhatsApp (📎)."
                                )
                            
                            msg_final = (
                                f"{msg_eval}\n\n"
                                f"🎓 *¡FELICITACIONES!*\n\n"
                                f"Ha completado el curso: *{progreso.curso.nombre}*"
                            )
                            partes_finales = [msg_final]
                            if radar_msg:
                                partes_finales.append(radar_msg)
                            partes_finales.append(msg_cert_img)
                            texto_respuesta = "[MULTI_MSG]" + "[SEP]".join(partes_finales)
                    else:
                        _prev_ctx = estudiante.contexto_temporal or {}
                        from .helpers_examenes import contexto_temporal_tras_cerrar_agente
                        from .drip_schedule import mensaje_bloqueo_avance_siguiente_modulo

                        _base_ctx = contexto_temporal_tras_cerrar_agente(progreso, _prev_ctx) or {}
                        estudiante.estado_onboarding = 'completado'
                        # Tras Darío + facilitadora: el drip/calendario aplica igual que sin agentes.
                        # No adelantar modulo_actual hasta que desbloquee el siguiente módulo.
                        if progreso:
                            modulo_cerrado = progreso.modulo_actual
                            _blk_tras_reto = mensaje_bloqueo_avance_siguiente_modulo(
                                estudiante, progreso, modulo_cerrado
                            )
                            if _blk_tras_reto:
                                estudiante.contexto_temporal = _base_ctx
                                estudiante.save(
                                    update_fields=['contexto_temporal', 'estado_onboarding']
                                )
                                texto_respuesta = f"{msg_eval}\n\n{_blk_tras_reto}"
                            else:
                                siguiente = progreso.curso.modulos.filter(
                                    numero__gt=progreso.modulo_actual.numero
                                ).order_by('numero').first()
                                if siguiente:
                                    progreso.modulo_actual = siguiente
                                    from .module_steps import reset_progreso_pasos_modulo
                                    reset_progreso_pasos_modulo(progreso, save=False)
                                    progreso.save(
                                        update_fields=[
                                            'modulo_actual',
                                            'paso_actual_modulo',
                                            'esperando_respuesta_evaluacion_paso',
                                            'paso_evaluacion_paso_id',
                                        ]
                                    )
                                    _base_ctx['post_reto_entregar_modulo_id'] = siguiente.id
                                estudiante.contexto_temporal = _base_ctx
                                estudiante.save(
                                    update_fields=['contexto_temporal', 'estado_onboarding']
                                )
                                texto_respuesta = (
                                    f"{msg_eval}\n\n"
                                    "✅ Escribe *continuar* para recibir el contenido del siguiente módulo.\n"
                                    "Cuando lo hayas revisado, responde *listo* para seguir."
                                )
                        else:
                            estudiante.contexto_temporal = _base_ctx
                            estudiante.save(
                                update_fields=['contexto_temporal', 'estado_onboarding']
                            )
                            texto_respuesta = (
                                f"{msg_eval}\n\n"
                                "✅ Escribe *continuar* para recibir el contenido del siguiente módulo.\n"
                                "Cuando lo hayas revisado, responde *listo* para seguir."
                            )
                    print(f"✅ Facilitadora reto evaluado | modo={modo_gami} | valor={puntaje}", flush=True)
            
            # 3.5a PRIORIDAD: Si está respondiendo al TUTOR IA (legacy)
            elif estudiante.estado_onboarding == 'esperando_respuesta_tutor_ia':
                from .gamificacion import PerfilGamificacion
                print(f"🎓 Evaluando respuesta del Facilitador (legacy tutor IA)")
                ctx = estudiante.contexto_temporal or {}
                modulo_id = ctx.get('modulo_id')
                pregunta_tutor = ctx.get('pregunta_tutor', '')
                intentos = ctx.get('intentos_tutor', 0)
                
                # Detectar si el usuario quiere omitir el tutor
                msg_lower = msg_body.strip().lower()
                palabras_skip = ['listo', 'continuar', 'saltar', 'omitir', 'siguiente', 'pasar', 'menu', 'menú']
                
                if msg_body.strip() == '[AUDIO_NO_TRANSCRITO]':
                    # Audio no pudo ser transcrito — pedir reintento
                    texto_respuesta = (
                        "⚠️ No pude escuchar tu audio. Por favor intenta de nuevo "
                        "o escríbeme tu respuesta.\n\n"
                        "Si prefieres continuar sin responder, escribe *listo*."
                    )
                elif any(p in msg_lower for p in palabras_skip):
                    # Usuario quiere seguir sin responder al tutor
                    estudiante.contexto_temporal = None
                    estudiante.estado_onboarding = 'completado'
                    estudiante.save()
                    print(f"⏭️ Profesor Gerónimo omitido por usuario")
                    
                    # Si dijo "menu", retomar curso (B2B) o menú sandbox
                    if msg_lower in ['menu', 'menú']:
                        from .flujo_whatsapp_b2b import respuesta_tras_keyword_menu
                        texto_respuesta = respuesta_tras_keyword_menu(
                            estudiante, estudiante.nombre, msg_body
                        )
                    else:
                        # v1.9.6: NO llamar continuar_leccion — el skip solo cierra el tutor.
                        # El estudiante ya tiene el contenido del módulo. Cuando lo estudie
                        # y escriba "listo" de nuevo, avanzará por el flujo normal (estado=completado).
                        texto_respuesta = "👍 Sin problema.\n\nContinúa revisando el contenido del módulo 👆\n\nCuando termines, escribe *listo* para avanzar al siguiente."
                        print(f"✅ v1.9.6: Skip Gerónimo sin avance automático", flush=True)
                else:
                    from .tutor_ia_modulo import evaluar_respuesta_modulo
                    from .models import Modulo
                    
                    try:
                        modulo = Modulo.objects.get(id=modulo_id) if modulo_id else None
                    except Modulo.DoesNotExist:
                        modulo = None
                    
                    if modulo:
                        aprobado, feedback = evaluar_respuesta_modulo(
                            modulo, msg_body, pregunta_tutor,
                            estudiante_nombre=estudiante.nombre or "Estudiante"
                        )
                        
                        # v1.9.8: Siempre 1 sola interacción — feedback y continúa
                        estudiante.contexto_temporal = None
                        estudiante.estado_onboarding = 'completado'
                        estudiante.save()
                        if aprobado:
                            perfil, _ = PerfilGamificacion.objects.get_or_create(estudiante=estudiante)
                            perfil.agregar_puntos(5, "Respuesta correcta - Profesor Gerónimo")
                            texto_respuesta = f"{feedback}\n\n💰 *+5 puntos bonus* por tu respuesta 💪\n\nContinúa revisando el módulo 👆\nCuando termines, escribe *listo* para avanzar."
                            print(f"✅ v1.9.8: Gerónimo aprobado — 1 interacción", flush=True)
                        else:
                            texto_respuesta = f"{feedback}\n\n✅ *¡Buen esfuerzo!* Sigue estudiando el módulo 👆\n\nCuando termines, escribe *listo* para avanzar."
                            print(f"✅ v1.9.8: Gerónimo incorrecto — feedback y continúa", flush=True)
                    else:
                        estudiante.contexto_temporal = None
                        estudiante.estado_onboarding = 'completado'
                        estudiante.save()
                        # v1.9.6: Solo enviar feedback, NO auto-avanzar
                        texto_respuesta = "✅ Gracias por tu respuesta!\n\nContinúa revisando el módulo 👆\nCuando termines, escribe *listo* para avanzar."
                        print(f"✅ v1.9.6: Gerónimo sin módulo — feedback sin auto-avance", flush=True)
            
            # 3.5a2 PRIORIDAD: Si está respondiendo a la REVISIÓN DE PROGRESO
            elif estudiante.estado_onboarding == 'esperando_respuesta_progreso':
                from .gamificacion import PerfilGamificacion
                print(f"�‍🏫 Evaluando respuesta de María (Revisión de Progreso)")
                ctx = estudiante.contexto_temporal or {}
                pregunta_tutor = ctx.get('pregunta_tutor', '')
                modulos_info = ctx.get('modulos_info', '')
                intentos = ctx.get('intentos_tutor', 0)
                modulos_reto_ids = ctx.get('modulos_reto_ids', [])
                progreso_id = ctx.get('progreso_id')
                es_reto_final = ctx.get('es_reto_final', False)

                def _activar_reto_despues_de_maria(prefijo=""):
                    from .models import ProgresoEstudiante, Modulo
                    from .tutor_ia_modulo import (
                        cargar_modulos_reto,
                        generar_reto_facilitador,
                        listar_modulos_cobertura_reto,
                    )
                    progreso_r = (
                        ProgresoEstudiante.objects.filter(id=progreso_id)
                        .select_related('curso', 'modulo_actual')
                        .first()
                        if progreso_id
                        else None
                    )
                    _mids = list(modulos_reto_ids) if modulos_reto_ids else []
                    modulos_reto = (
                        cargar_modulos_reto(_mids, progreso_r.curso_id if progreso_r else None)
                        if _mids
                        else []
                    )
                    if not modulos_reto and progreso_r:
                        _mchk = None
                        _outer = estudiante.contexto_temporal or {}
                        mid = _outer.get('modulo_id')
                        if mid:
                            _mchk = Modulo.objects.filter(
                                id=mid, curso_id=progreso_r.curso_id
                            ).first()
                        if not _mchk and getattr(progreso_r, 'modulo_actual_id', None):
                            _mchk = progreso_r.modulo_actual
                        if _mchk:
                            modulos_reto = listar_modulos_cobertura_reto(_mchk, progreso_r.curso)
                            _mids = [m.id for m in modulos_reto]
                    if not (progreso_r and modulos_reto):
                        logger.warning(
                            "reto post-María vacío | progreso_id=%s modulo_ctx=%s ids_ctx=%s",
                            progreso_id,
                            (estudiante.contexto_temporal or {}).get('modulo_id'),
                            modulos_reto_ids,
                        )
                        estudiante.contexto_temporal = None
                        estudiante.estado_onboarding = 'completado'
                        estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])
                        return (
                            f"{prefijo}\n\n"
                            "Seguimos con tu curso. Escribí *listo* para continuar "
                            "o *ayuda* si necesitás soporte."
                        ).strip()
                    _cliente = estudiante.cliente if hasattr(estudiante, 'cliente') and estudiante.cliente else None
                    nombre_tutor = (
                        (_cliente.nombre_agente_tutor if _cliente and hasattr(_cliente, 'nombre_agente_tutor') and _cliente.nombre_agente_tutor else '') or
                        progreso_r.curso.nombre_agente_tutor or 'Claudia'
                    )
                    reto = generar_reto_facilitador(
                        modulos_reto,
                        progreso_r.curso.nombre,
                        estudiante_nombre=estudiante.nombre or "Estudiante",
                        preguntas_ejemplo=progreso_r.curso.preguntas_ejemplo_ia or ""
                    )
                    _prev_ts = (estudiante.contexto_temporal or {}).get('_ts_leccion', 0)
                    estudiante.contexto_temporal = {
                        'tipo': 'reto_facilitador',
                        'modulos_reto_ids': _mids,
                        'reto_texto': reto,
                        'progreso_id': progreso_id,
                        'es_final': es_reto_final,
                        '_ts_leccion': _prev_ts,
                    }
                    estudiante.estado_onboarding = 'esperando_respuesta_reto'
                    estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])
                    bloque_reto = (
                        f"📋 *{nombre_tutor}*\n\n"
                        f"{reto}\n\n"
                        "✍️ _Escriba o envíe un audio con su respuesta._"
                    )
                    return f"{prefijo}\n\n{bloque_reto}".strip()
                
                # Detectar si el usuario quiere omitir
                msg_lower = msg_body.strip().lower()
                palabras_skip = ['listo', 'continuar', 'saltar', 'omitir', 'siguiente', 'pasar', 'menu', 'menú']
                
                if msg_body.strip() == '[AUDIO_NO_TRANSCRITO]':
                    texto_respuesta = (
                        "⚠️ No pude escuchar tu audio. Por favor intenta de nuevo "
                        "o escríbeme tu respuesta.\n\n"
                        "Si prefieres continuar sin responder, escribe *listo*."
                    )
                elif any(p in msg_lower for p in palabras_skip):
                    print(f"⏭️ María omitida por usuario → activando reto")
                    
                    if msg_lower in ['menu', 'menú']:
                        from .flujo_whatsapp_b2b import respuesta_tras_keyword_menu
                        texto_respuesta = respuesta_tras_keyword_menu(
                            estudiante, estudiante.nombre, msg_body
                        )
                    else:
                        texto_respuesta = _activar_reto_despues_de_maria("👍 Perfecto, pasemos al reto de la facilitadora.")
                else:
                    from .tutor_ia_modulo import evaluar_respuesta_progreso
                    
                    resuelta, feedback = evaluar_respuesta_progreso(
                        modulos_info, msg_body, pregunta_tutor,
                        estudiante_nombre=estudiante.nombre or "Estudiante"
                    )
                    
                    if resuelta:
                        perfil, _ = PerfilGamificacion.objects.get_or_create(estudiante=estudiante)
                        perfil.agregar_puntos(3, "Revisión de progreso - María")
                        texto_respuesta = _activar_reto_despues_de_maria(f"{feedback}\n\n💰 *+3 puntos* por tu reflexión 💪")
                        print(f"✅ María resuelta — activando reto", flush=True)
                    else:
                        texto_respuesta = _activar_reto_despues_de_maria(f"{feedback}\n\n✅ *Buena reflexión!*")
                        print(f"✅ María evaluada — activando reto", flush=True)

            # 3.5a2 PRIORIDAD: Si está respondiendo pregunta de RECUPERACIÓN (<70 pts)
            elif estudiante.estado_onboarding == 'esperando_respuesta_recuperacion':
                ctx = estudiante.contexto_temporal or {}
                pregunta_data = ctx.get('pregunta_data', {})
                correcta = pregunta_data.get('correcta', 'A')
                explicacion = pregunta_data.get('explicacion', '')
                
                from .tutor_ia_modulo import evaluar_respuesta_recuperacion
                es_correcta, msg_evaluacion = evaluar_respuesta_recuperacion(msg_body, correcta, explicacion)
                
                # Si acertó, dar puntos bonus
                if es_correcta:
                    try:
                        from .gamificacion import PerfilGamificacion
                        perfil_rec = PerfilGamificacion.objects.get(estudiante=estudiante)
                        perfil_rec.agregar_puntos(15, "🏆 Pregunta de recuperación correcta")
                    except Exception:
                        pass
                
                # Limpiar estado y continuar al certificado
                curso_id = ctx.get('curso_id')
                estudiante.estado_onboarding = 'completado'
                estudiante.contexto_temporal = None
                estudiante.save()
                
                # Generar certificado y resumen
                from .response_templates import _generar_completado_final
                msg_final = _generar_completado_final(estudiante, curso_id)
                
                texto_respuesta = f"[MULTI_MSG]{msg_evaluacion}[SEP]{msg_final}"

            # 3.5b PRIORIDAD: Si está respondiendo pregunta de módulo (examen clásico)
            elif estudiante.estado_onboarding == 'esperando_respuesta_modulo':
                # Si el usuario dice "menu", salir del examen y mostrar menú
                msg_lower_exam = msg_body.strip().lower()
                if msg_body.strip() == '[AUDIO_NO_TRANSCRITO]':
                    texto_respuesta = (
                        "⚠️ No pude escuchar tu audio. Por favor intenta de nuevo "
                        "o escríbeme tu respuesta."
                    )
                elif msg_lower_exam in ['menu', 'menú']:
                    estudiante.estado_onboarding = 'completado'
                    estudiante.save()
                    from .flujo_whatsapp_b2b import respuesta_tras_keyword_menu
                    texto_respuesta = respuesta_tras_keyword_menu(
                        estudiante, estudiante.nombre, msg_body
                    )
                else:
                    # Validar respuesta a pregunta de módulo
                    from .pregunta_handler import validar_respuesta, procesar_respuesta_abierta_ia
                    print(f"📝 Validando respuesta a pregunta de módulo")
                    
                    # Verificar si la pregunta es abierta (IA) o de opciones
                    ctx = estudiante.contexto_temporal or {}
                    es_pregunta_ia = ctx.get('tipo') == 'pregunta_tutor_ia'
                    
                    # Fallback: verificar ultima_pregunta_data si contexto no tiene tipo IA
                    if not es_pregunta_ia:
                        pregunta_data = None
                        if hasattr(estudiante, 'ultima_pregunta_data') and estudiante.ultima_pregunta_data:
                            import ast
                            try:
                                pregunta_data = ast.literal_eval(estudiante.ultima_pregunta_data)
                            except Exception:
                                pregunta_data = None
                        if pregunta_data and (not pregunta_data.get('opciones')):
                            es_pregunta_ia = True
                    
                    if es_pregunta_ia:
                        # Pregunta IA abierta — evaluar con IA
                        es_correcta, mensaje_respuesta = procesar_respuesta_abierta_ia(estudiante, msg_body)
                        modulo_completado = None
                    else:
                        es_correcta, mensaje_respuesta, modulo_completado = validar_respuesta(estudiante, msg_body)

                    # Obtener progreso para avanzar al siguiente módulo
                    if modulo_completado or es_pregunta_ia:
                        from .helpers_examenes import puede_avanzar_modulo, es_modulo_checkpoint_reto_ia
                        
                        if modulo_completado:
                            progreso = modulo_completado.progreso
                            modulo_actual = modulo_completado.modulo
                        else:
                            # Para preguntas IA abierta, obtener progreso desde contexto
                            from .models import ProgresoEstudiante, Modulo
                            modulo_id = ctx.get('modulo_id')
                            progreso_id = ctx.get('progreso_id')
                            try:
                                modulo_actual = Modulo.objects.get(id=modulo_id) if modulo_id else None
                                progreso = ProgresoEstudiante.objects.get(id=progreso_id) if progreso_id else None
                            except Exception:
                                modulo_actual = None
                                progreso = None
                            
                            if progreso and modulo_actual:
                                # Crear ModuloCompletado para la pregunta IA abierta
                                modulo_abierto, created_abierto = ModuloCompletado.objects.get_or_create(
                                    progreso=progreso,
                                    modulo=modulo_actual
                                )
                                if created_abierto:
                                    progreso.fecha_ultimo_avance = timezone.now()
                                    progreso.save(update_fields=['fecha_ultimo_avance'])
                        
                        _skip_avance = False
                        if not (progreso and modulo_actual):
                            texto_respuesta = mensaje_respuesta
                            _skip_avance = True
                        
                        if not _skip_avance:
                            # VERIFICAR EXAMEN OBLIGATORIO ANTES DE AVANZAR
                            puede_avanzar, mensaje_examen, detalles = puede_avanzar_modulo(estudiante, modulo_actual)
                        
                            if not puede_avanzar:
                                # NO puede avanzar - examen obligatorio no aprobado
                                mensaje_respuesta += f"""


🔒 *Examen Obligatorio*

{mensaje_examen}

Para continuar al siguiente módulo debes aprobar el examen de este módulo.

Escribe *"examen"* cuando estés listo para intentarlo."""
                            
                                texto_respuesta = mensaje_respuesta
                                # Fall through to Twilio API send
                        
                            else:
                                # Buscar siguiente módulo
                                siguiente_modulo = progreso.curso.modulos.filter(
                                    numero__gt=modulo_actual.numero
                                ).order_by('numero').first()
                                drip_bloqueado = False

                                if siguiente_modulo:
                                    from .drip_schedule import mensaje_bloqueo_avance_siguiente_modulo
                                    from .helpers_examenes import evaluar_checkpoint_reto_ia

                                    total_modulos = progreso.curso.modulos.count()
                                    usar_agentes_ia_curso = bool(
                                        getattr(progreso.curso, 'usar_agentes_ia', True)
                                    )
                                    decision_cp = evaluar_checkpoint_reto_ia(
                                        modulo_actual,
                                        total_modulos,
                                        usar_agentes_ia_curso,
                                    )
                                    es_modulo_reto = decision_cp.es_reto
                                    try:
                                        from core.eventos_ia import emit_checkpoint_evaluado

                                        emit_checkpoint_evaluado(
                                            decision_cp,
                                            estudiante=estudiante,
                                            curso=progreso.curso,
                                            modulo=modulo_actual,
                                            origen='pregunta_modulo',
                                        )
                                    except Exception:
                                        pass

                                    _blk_v = mensaje_bloqueo_avance_siguiente_modulo(
                                        estudiante, progreso, modulo_actual
                                    )
                                    # Igual que en continuar_leccion: el drip no debe tapar el checkpoint IA.
                                    if _blk_v and not es_modulo_reto:
                                        texto_respuesta = f"{mensaje_respuesta}\n\n{_blk_v}"
                                        drip_bloqueado = True
                                        siguiente_modulo = None

                                if siguiente_modulo:
                                    _fcp = getattr(modulo_actual, 'facilitador_checkpoint', None)
                                    logger.info(
                                        '🎯 [checkpoint-pregunta-modulo] curso=%s mod_num=%s total_mod=%s '
                                        'usar_ia=%s facilitador_checkpoint=%s regla=%s -> es_reto=%s | est=%s',
                                        progreso.curso_id,
                                        getattr(modulo_actual, 'numero', None),
                                        total_modulos,
                                        usar_agentes_ia_curso,
                                        _fcp,
                                        decision_cp.regla_aplicada,
                                        es_modulo_reto,
                                        estudiante.id,
                                    )
                                    
                                    if not es_modulo_reto:
                                        # Normal: advance pointer
                                        progreso.modulo_actual = siguiente_modulo
                                        from .module_steps import reset_progreso_pasos_modulo
                                        reset_progreso_pasos_modulo(progreso, save=False)
                                        progreso.save(
                                            update_fields=[
                                                'modulo_actual',
                                                'paso_actual_modulo',
                                                'esperando_respuesta_evaluacion_paso',
                                                'paso_evaluacion_paso_id',
                                            ]
                                        )
                                    # else: pointer stays — will advance after reto
                                
                                    estudiante.preguntas_ia_restantes = 3
                                    estudiante.save()
                                
                                    porcentaje = progreso.porcentaje_avance()
                                    from .response_templates import obtener_video_url
                                    video_url = obtener_video_url(siguiente_modulo)
                                
                                    archivos_multimedia = siguiente_modulo.archivos_multimedia.filter(activo=True)
                                    archivos_msg = ""
                                    primera_media_url = None
                                    extra_media_urls = []
                                
                                    if archivos_multimedia.exists():
                                        archivos_msg = ""
                                        for idx, archivo in enumerate(archivos_multimedia):
                                            icono = {'video': '🎥', 'imagen': '🖼️', 'infografia': '📊', 'pdf': '📄', 'audio': '🎵'}.get(archivo.tipo, '📁')
                                            url = archivo.get_url_para_envio()
                                            if url:
                                                if not primera_media_url:
                                                    primera_media_url = url
                                                    archivos_msg += f"\n{icono} {archivo.titulo} (adjunto)"
                                                else:
                                                    extra_media_urls.append((url, archivo.titulo, icono))
                                                    archivos_msg += f"\n{icono} {archivo.titulo} (adjunto)"
                                            else:
                                                archivos_msg += f"\n{icono} {archivo.titulo}"

                                    if not archivos_multimedia.exists() and video_url:
                                        primera_media_url = video_url
                                
                                    msg_completado = mensaje_respuesta
                                
                                    # v1.9.8h: Agentes — check reto
                                    nombre_tutor = progreso.curso.nombre_agente_tutor or 'Claudia'
                                    nombre_asistente = progreso.curso.nombre_agente_asistente or 'Darío'
                                    
                                    if es_modulo_reto:
                                        from .tutor_ia_modulo import (
                                            descripcion_rango_modulos_reto_esp,
                                            listar_modulos_cobertura_reto,
                                        )

                                        modulos_reto = listar_modulos_cobertura_reto(
                                            modulo_actual, progreso.curso
                                        )
                                        modulos_reto_range = descripcion_rango_modulos_reto_esp(
                                            modulos_reto
                                        )

                                        dario_msg = (
                                            f"💬 *{nombre_asistente}*\n\n"
                                            f"¡Hola! Es hora de una pausa para repasar conceptos. "
                                            f"{nombre_tutor} te va a recibir con un reto sobre {modulos_reto_range}.\n\n"
                                            f"Te puedo ayudar a resolver un par de preguntas antes. "
                                            f"¿Tienes alguna pregunta sobre lo que hemos visto?\n\n"
                                            f"Ejemplos:\n"
                                            f"• ¿Cómo identifico a tiempo el daño en campo?\n"
                                            f"• ¿Qué error debo evitar al aplicar control?\n"
                                            f"• ¿Qué reviso primero antes de decidir tratamiento?\n\n"
                                            f"Envíame un audio o escríbeme; si no tienes preguntas, escribe *listo*."
                                        )

                                        _prev_ts = (estudiante.contexto_temporal or {}).get('_ts_leccion', 0)
                                        estudiante.contexto_temporal = {
                                            'tipo': 'asistente_dario',
                                            'curso_activo_id': progreso.curso_id,
                                            'modulo_id': modulo_actual.id,
                                            'progreso_id': progreso.id,
                                            'modulos_reto_ids': [m.id for m in modulos_reto],
                                            'preguntas_hechas': 0,
                                            '_ts_leccion': _prev_ts,
                                        }
                                        estudiante.estado_onboarding = 'esperando_respuesta_asistente'
                                        estudiante.save()
                                        
                                        texto_respuesta = "[MULTI_MSG]" + "[SEP]".join([msg_completado, dario_msg])
                                    else:
                                        # v1.9.8i: Normal module — exam result + next module (no completado msg)
                                        estudiante.estado_onboarding = 'completado'
                                        estudiante.save()
                                        
                                        msg_modulo = f"📖 *Módulo {siguiente_modulo.numero}: {siguiente_modulo.titulo}*\n\n{siguiente_modulo.descripcion}\n\n{siguiente_modulo.contenido}"
                                    
                                        if siguiente_modulo.examen_obligatorio:
                                            msg_modulo += f"\n\n⚠️ *Este módulo tiene examen obligatorio ({siguiente_modulo.puntaje_minimo_aprobacion}% para aprobar)*"
                                    
                                        partes = [msg_completado, msg_modulo]
                                        hay_media_exam = False
                                        if primera_media_url:
                                            partes.append(parte_mensaje_con_media(primera_media_url))
                                            hay_media_exam = True
                                        for extra_url, extra_titulo, extra_icono in extra_media_urls:
                                            cap_e = f'{extra_icono} {extra_titulo}'.strip() if extra_titulo else None
                                            partes.append(parte_mensaje_con_media(extra_url, cap_e))
                                            hay_media_exam = True
                                        if hay_media_exam:
                                            partes.append("[DELAY:5]")
                                        from .avance_whatsapp import CTX_FIN_ENTREGA_MODULO, resolver_cta_listo
                                        partes.append(
                                            resolver_cta_listo(
                                                estudiante, progreso.curso, CTX_FIN_ENTREGA_MODULO
                                            )
                                        )
                                        texto_respuesta = "[MULTI_MSG]" + "[SEP]".join(partes)

                                elif not drip_bloqueado:
                                    # Completó todos los módulos
                                    progreso.completado = True
                                    progreso.fecha_completado = timezone.now()
                                    progreso.save()
                                
                                    # === v1.9.8h: RETO FINAL en lugar de pregunta de recuperación ===
                                    _skip_cert = False
                                    pregunta_abierta = _pregunta_abierta_final_pendiente(estudiante, progreso)
                                    usar_gamificacion_final = bool(
                                        progreso.curso.usar_gamificacion or
                                        (estudiante.cliente.usar_gamificacion if getattr(estudiante, 'cliente', None) else False)
                                    )
                                    usar_agentes_ia_final = bool(getattr(progreso.curso, 'usar_agentes_ia', True))
                                    # Reto final solo si el curso usa gamificación Y agentes IA.
                                    activar_reto_final = usar_gamificacion_final and usar_agentes_ia_final

                                    if activar_reto_final:
                                        try:
                                            nombre_tutor_final = progreso.curso.nombre_agente_tutor or 'Claudia'
                                            nombre_asist_final = progreso.curso.nombre_agente_asistente or 'Darío'
                                            modulos_all = list(progreso.curso.modulos.filter(numero__gte=4).order_by('numero'))
                                            if not modulos_all:
                                                modulos_all = list(progreso.curso.modulos.all().order_by('numero'))
                                            modulos_final_range = "los módulos finales del curso"
                                            if len(modulos_all) >= 2:
                                                modulos_final_range = f"los módulos {modulos_all[0].numero} a {modulos_all[-1].numero}"

                                            _prev_ts = (estudiante.contexto_temporal or {}).get('_ts_leccion', 0)
                                            estudiante.contexto_temporal = {
                                                'tipo': 'asistente_dario',
                                                'curso_activo_id': progreso.curso_id,
                                                'curso_id': progreso.curso.id,
                                                'modulo_id': modulo_actual.id,
                                                'progreso_id': progreso.id,
                                                'modulos_reto_ids': [m.id for m in modulos_all],
                                                'preguntas_hechas': 0,
                                                'es_reto_final': True,
                                                '_ts_leccion': _prev_ts,
                                            }
                                            if pregunta_abierta:
                                                estudiante.contexto_temporal['pregunta_abierta_final_id'] = pregunta_abierta.id
                                            estudiante.estado_onboarding = 'esperando_respuesta_asistente'
                                            estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])

                                            texto_respuesta = (
                                                f"{mensaje_respuesta}\n\n"
                                                f"🎉 *¡Completaste todos los módulos del curso!*\n\n"
                                                f"💬 *{nombre_asist_final}*\n\n"
                                                f"Antes de tu certificado, {nombre_tutor_final} te planteará un reto final sobre {modulos_final_range}.\n\n"
                                                "¿Tienes dudas antes del reto?\n"
                                                "Ejemplos:\n"
                                                "• ¿Cómo diferencio un daño leve de uno económico?\n"
                                                "• ¿Qué paso práctico recomienda para validar en campo?\n"
                                                "• ¿Qué indicador debo monitorear cada semana?\n\n"
                                                "Envíame tu pregunta (texto o audio).\n"
                                                "Si no tienes dudas, escribe *listo* para pasar con la facilitadora."
                                            )
                                            logger.info(
                                                f"🎯 Reto final activado | estudiante_id={estudiante.id} | curso_id={progreso.curso.id} | "
                                                f"curso_usar_gamificacion={bool(progreso.curso.usar_gamificacion)} | "
                                                f"cliente_usar_gamificacion={bool(estudiante.cliente.usar_gamificacion) if getattr(estudiante, 'cliente', None) else False} | "
                                                f"pregunta_abierta_id={getattr(pregunta_abierta, 'id', None)}"
                                            )
                                            _skip_cert = True
                                        except Exception as e:
                                            logger.warning(f"⚠️ Reto final exam: {e}")
                                    
                                    if not _skip_cert:
                                        if pregunta_abierta:
                                            estudiante.contexto_temporal = {
                                                'tipo': 'pregunta_abierta_final',
                                                'curso_id': progreso.curso.id,
                                                'progreso_id': progreso.id,
                                                'pregunta_abierta_final_id': pregunta_abierta.id,
                                            }
                                            estudiante.estado_onboarding = 'esperando_respuesta_pregunta_abierta_final'
                                            estudiante.save(update_fields=['contexto_temporal', 'estado_onboarding'])
                                            texto_respuesta = (
                                                f"{mensaje_respuesta}\n\n"
                                                "📝 *Antes del cierre final*, responde esta pregunta abierta:\n\n"
                                                f"{pregunta_abierta.pregunta}\n\n"
                                                "✍️ Tu facilitadora revisará y calificará tu respuesta."
                                            )
                                        else:
                                            estudiante.estado_onboarding = 'curso_finalizado'
                                            estudiante.save(update_fields=['estado_onboarding'])

                                            msg_final = (
                                                f"{mensaje_respuesta}\n\n"
                                                f"🎓 *¡FELICITACIONES!*\n\n"
                                                f"Ha completado el curso: *{progreso.curso.nombre}*\n\n"
                                                f"🏆 Su certificado se está generando..."
                                            )

                                            msg_cert_img = ""
                                            try:
                                                from .certificado_service import crear_certificado_automatico, obtener_url_certificado_twilio
                                                logger.info(f"🎓 Iniciando generación de certificado para {estudiante.nombre} - {progreso.curso.nombre}")
                                                cert = crear_certificado_automatico(estudiante, progreso.curso)
                                                logger.info(f"🎓 Certificado resultado: cert={cert}, imagen={cert.archivo_imagen if cert else 'N/A'}, pdf={cert.archivo_pdf if cert else 'N/A'}")
                                            
                                                if cert and cert.archivo_imagen:
                                                    cert_url = obtener_url_certificado_twilio(cert)
                                                    if cert_url:
                                                        msg_cert_img = f"🎓 *¡Tu certificado!*\n\n[MEDIA:{cert_url}]"
                                                        logger.info(f"✅ Certificado URL para Twilio: {cert_url}")
                                                    else:
                                                        s3_key = str(cert.archivo_imagen.name)
                                                        cert_url = f"https://eki-produccion.s3.us-east-2.amazonaws.com/{s3_key}"
                                                        msg_cert_img = f"🎓 *¡Tu certificado!*\n\n[MEDIA:{cert_url}]"
                                                        logger.info(f"✅ Certificado URL fallback: {cert_url}")
                                                elif cert and cert.archivo_pdf:
                                                    cert_url = cert.archivo_pdf.url
                                                    msg_cert_img = f"🎓 *¡Tu certificado!*\n📄 Descárgalo aquí: {cert_url}"
                                                    logger.info(f"📄 Certificado PDF URL: {cert_url}")
                                                elif cert:
                                                    logger.warning(f"⚠️ Cert creado sin archivo, forzando regeneración...")
                                                    from .certificado_service import generar_y_guardar_certificado
                                                    generar_y_guardar_certificado(cert, force=True)
                                                    cert.refresh_from_db()
                                                    if cert.archivo_imagen:
                                                        cert_url = obtener_url_certificado_twilio(cert)
                                                        if cert_url:
                                                            msg_cert_img = f"🎓 *¡Tu certificado!*\n\n[MEDIA:{cert_url}]"
                                                            logger.info(f"✅ Certificado PRESIGNED URL (retry): {cert_url[:100]}...")
                                                        else:
                                                            msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."
                                                    else:
                                                        msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."
                                                else:
                                                    msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."
                                                    logger.warning(f"❌ crear_certificado_automatico retornó None para {estudiante.nombre}")
                                            except Exception as e:
                                                logger.error(f"❌ Error generando certificado: {e}", exc_info=True)
                                                import traceback; traceback.print_exc()
                                                msg_cert_img = "🎓 Tu certificado se está generando. Te lo enviaremos pronto."

                                            radar_msg = ""
                                            if _activar_radar_empleabilidad_si_aplica(estudiante):
                                                radar_msg = (
                                                    "📍 *¡Radar de Empleos desbloqueado!*\n\n"
                                                    "Ve al parque principal de Subachoque y envíame tu *Ubicación* "
                                                    "usando el clip de WhatsApp (📎)."
                                                )

                                            partes = [msg_final]
                                            if radar_msg:
                                                partes.append(radar_msg)
                                            partes.append(msg_cert_img)
                                            texto_respuesta = "[MULTI_MSG]" + "[SEP]".join(partes)
                    
                    if not texto_respuesta:
                        texto_respuesta = mensaje_respuesta
                    print(f"✅ Respuesta validada: {'Correcta' if es_correcta else 'Incorrecta'}")
            
            # 3.5c PRIORIDAD: Si está seleccionando un curso de la lista
            elif estudiante.estado_onboarding == 'esperando_seleccion_curso':
                from .flujo_whatsapp_b2b import (
                    es_estudiante_b2b,
                    mensaje_digitos_sin_menu,
                    salir_seleccion_curso_legacy,
                )
                msg_sel = msg_body.strip().lower()
                if es_estudiante_b2b(estudiante):
                    salir_seleccion_curso_legacy(estudiante)
                    from .response_templates import get_response_for_intent
                    if msg_body.strip().isdigit():
                        texto_respuesta = mensaje_digitos_sin_menu(estudiante)
                    else:
                        texto_respuesta = get_response_for_intent(
                            'continuar_leccion',
                            estudiante.nombre,
                            estudiante_id=estudiante.id,
                            mensaje_original=msg_body,
                        )
                elif msg_sel in ['menu', 'menú']:
                    estudiante.estado_onboarding = 'completado'
                    estudiante.contexto_temporal = None
                    estudiante.save()
                    from .response_templates import get_response_for_intent
                    texto_respuesta = get_response_for_intent('saludo', estudiante.nombre, estudiante_id=estudiante.id)
                else:
                    import re as re_curso_sel
                    indice_sel = None
                    match_tomar_sel = re_curso_sel.match(r'^tomar\s*(\d+)$', msg_sel)
                    if match_tomar_sel:
                        indice_sel = int(match_tomar_sel.group(1))
                    elif msg_body.strip().isdigit():
                        indice_sel = int(msg_body.strip())
                    if indice_sel is not None:
                        from .selector_curso import continuar_curso_seleccionado
                        estudiante.estado_onboarding = 'completado'
                        estudiante.save(update_fields=['estado_onboarding'])
                        texto_respuesta = continuar_curso_seleccionado(estudiante.id, indice_sel, msg_body)
                        print(f"✅ Curso seleccionado: {indice_sel}")
                    else:
                        # Si no es número/tomar ni menú, resetear estado y procesar normalmente
                        estudiante.estado_onboarding = 'completado'
                        estudiante.contexto_temporal = None
                        estudiante.save()
                        from core.eventos_ia import detectar_intent_con_evento
                        from .response_templates import get_response_for_intent
                        _p0 = estudiante.progresos.order_by('-fecha_inicio').first()
                        intent = detectar_intent_con_evento(
                            msg_body,
                            estudiante=estudiante,
                            curso=_p0.curso if _p0 else None,
                            modulo=_p0.modulo_actual if _p0 else None,
                        )
                        if intent != 'desconocido':
                            texto_respuesta = get_response_for_intent(intent, estudiante.nombre, estudiante_id=estudiante.id, mensaje_original=msg_body)
                        else:
                            texto_respuesta = "No entendí tu selección. Escribe *menú* para ver las opciones."
            
            # 4. Detectar intent y usar templates primero
            else:
                from core.eventos_ia import detectar_intent_con_evento
                from .response_templates import get_response_for_intent

                _prog = None
                try:
                    _prog = estudiante.progresos.order_by('-fecha_inicio').first()
                except Exception:
                    pass
                intent = detectar_intent_con_evento(
                    msg_body,
                    estudiante=estudiante,
                    curso=_prog.curso if _prog else None,
                    modulo=_prog.modulo_actual if _prog else None,
                )
                print(f"🎯 Intent detectado: {intent}")
            
                # Intent especial: corregir datos → redirigir al flujo de corrección
                if intent == 'corregir_datos':
                    estudiante.estado_chat = 'ESPERANDO_CORRECCION_DATOS'
                    estudiante.save()
                    texto_respuesta = (
                        "📝 *Corrección de Datos*\n\n"
                        "Puedes corregir cualquiera de tus datos.\n\n"
                        "Escribe el campo que deseas cambiar seguido del nuevo valor:\n\n"
                        "1️⃣ *nombre:* Tu nombre completo\n"
                        "2️⃣ *municipio:* Tu municipio\n"
                        "3️⃣ *departamento:* Tu departamento\n"
                        "4️⃣ *documento:* Tipo y número (CC, TI, CE, PP)\n"
                        "5️⃣ *edad:* Tu edad\n"
                        "6️⃣ *genero:* M, F, Otro, NR\n\n"
                        "📝 _Ejemplos:_\n"
                        "_nombre: María García López_\n"
                        "_municipio: Bogotá_\n"
                        "_edad: 35_\n\n"
                        "👉 Escribe *menú* cuando termines"
                    )
                # Si hay un intent conocido, usar template
                elif intent != 'desconocido':
                    texto_respuesta = get_response_for_intent(
                        intent, 
                        estudiante.nombre,
                        estudiante_id=estudiante.id,
                        mensaje_original=msg_body
                    )
                    print(f"✅ Respuesta desde template: {texto_respuesta[:50]}...")
                else:
                    # Solo si no hay intent, usar IA para preguntas sobre agricultura
                    # 🛑 ANTI-ABUSO IA: Verificar preguntas restantes
                    print(f"🤖 Usando IA para pregunta sobre agricultura")
                    if estudiante.preguntas_ia_restantes <= 0:
                        # Freno de mano: IA pausada
                        from .avance_whatsapp import CTX_FIN_ENTREGA_MODULO, resolver_cta_listo
                        _prog_ia = estudiante.progresos.order_by('-fecha_inicio').first()
                        _curso_ia = _prog_ia.curso if _prog_ia else None
                        texto_respuesta = (
                            "[MULTI_MSG]⚠️ *Has agotado tus preguntas libres a la IA para este modulo.*\n\n"
                            "Para desbloquear mas preguntas, necesitas responder "
                            "la pregunta de evaluacion del modulo actual."
                            "[SEP]"
                            + resolver_cta_listo(estudiante, _curso_ia, CTX_FIN_ENTREGA_MODULO)
                        )
                    else:
                        try:
                            from .ai_assistant import responder_con_ia
                            texto_respuesta = responder_con_ia(msg_body, telefono_limpio)
                            # Restar pregunta usada (anti-abuso silencioso)
                            estudiante.preguntas_ia_restantes = max(0, estudiante.preguntas_ia_restantes - 1)
                            estudiante.save()
                            print(f"✅ IA generó respuesta: {texto_respuesta[:50]}...")
                        except Exception as e:
                            print(f"❌ Error IA: {e}, usando respuesta genérica")
                            texto_respuesta = "Disculpa, tengo problemas técnicos. Vuelve a escribir tu mensaje para continuar."
        
        # 3. Enviar respuesta via Twilio
        print(f"📤 ENVIANDO RESPUESTA: '{texto_respuesta[:80]}...' (len={len(texto_respuesta)})", flush=True)
        if (
            'Pregunta abierta final de la Facilitadora' in texto_respuesta
            or 'Siguiente pregunta abierta final' in texto_respuesta
        ):
            logger.info(
                "📨 Respuesta de pregunta abierta final enviada | estudiante_id=%s | preview=%s",
                estudiante.id,
                texto_respuesta[:500],
            )
        # Detectar si hay media_url en la respuesta (marcado con [MEDIA:url])
        # NOTA: Solo extraer [MEDIA:] para mensajes simples (no MULTI_MSG)
        # MULTI_MSG maneja su propio [MEDIA:] por cada parte del split
        media_url_to_send = None
        if not texto_respuesta.startswith('[MULTI_MSG]') and '[MEDIA:' in texto_respuesta:
            import re
            media_match = re.search(r'\[MEDIA:(.*?)\]', texto_respuesta)
            if media_match:
                media_url_to_send = media_match.group(1)
                texto_respuesta = texto_respuesta.replace(media_match.group(0), '').strip()
                print(f"🖼️ Media URL detectada: {media_url_to_send}")
        
        try:
            from twilio.rest import Client
            account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
            auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
            twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
            twilio_number = str(twilio_number).strip()
            print(f"DEBUG TWILIO FROM (views.py): '{twilio_number}'")
            
            if not account_sid or not auth_token:
                print("❌ Credenciales Twilio faltantes")
                return
            
            client = Client(account_sid, auth_token)

            # Usar el teléfono original (con +) para enviar por Twilio
            destino_formateado = f'whatsapp:{msg_from}' if not msg_from.startswith('whatsapp:') else msg_from
            destino_formateado = str(destino_formateado).strip()
            
            # Check if response is a standalone template send
            if texto_respuesta.strip().startswith('[SEND_TEMPLATE:'):
                import re
                tmpl_match = re.match(r'\[SEND_TEMPLATE:(HX[a-f0-9]+)\]', texto_respuesta.strip())
                if tmpl_match:
                    template_sid = tmpl_match.group(1)
                    print(f"📋 Enviando Content Template standalone: {template_sid}")
                    try:
                        from .whatsapp_service import enviar_template_twilio
                        tel_limpio = msg_from.replace('whatsapp:', '').replace('+', '')
                        enviar_template_twilio(tel_limpio, template_sid)
                        print(f"✅ Template {template_sid} enviado OK")
                    except Exception as tmpl_err:
                        print(f"⚠️ Error enviando template standalone: {tmpl_err}")
                    return
            
            # Check if response is a multi-message (marked with [MULTI_MSG])
            if texto_respuesta.startswith('[MULTI_MSG]'):
                # Extract and send multiple messages
                partes = texto_respuesta.replace('[MULTI_MSG]', '', 1).split('[SEP]')
                
                for idx, parte in enumerate(partes):
                    if not parte.strip():
                        continue
                    
                    parte_texto = parte.strip()
                    parte_media = None
                    
                    # [DELAY:N] — pausa intencional para que WhatsApp entregue videos antes del texto siguiente
                    import re as re_delay
                    delay_match = re_delay.match(r'^\[DELAY:(\d+)\]$', parte_texto)
                    if delay_match:
                        delay_secs = int(delay_match.group(1))
                        print(f"⏳ Pausa de {delay_secs}s para entrega de videos...")
                        import time
                        time.sleep(delay_secs)
                        continue
                    
                    # Detectar si esta parte es un Content Template de Twilio
                    if parte_texto.startswith('[SEND_TEMPLATE:'):
                        import re
                        tmpl_match = re.match(r'\[SEND_TEMPLATE:(HX[a-f0-9]+)\]', parte_texto)
                        if tmpl_match:
                            template_sid = tmpl_match.group(1)
                            print(f"📋 Enviando Content Template {template_sid} como parte {idx+1}")
                            try:
                                from .whatsapp_service import enviar_template_twilio
                                tel_limpio = msg_from.replace('whatsapp:', '').replace('+', '')
                                enviar_template_twilio(tel_limpio, template_sid)
                                print(f"✅ Template {template_sid} enviado OK")
                            except Exception as tmpl_err:
                                print(f"⚠️ Error enviando template: {tmpl_err}")
                            import time
                            time.sleep(0.5)
                            continue
                    
                    # Extraer [MEDIA:url] de esta parte
                    if '[MEDIA:' in parte_texto:
                        import re
                        media_match_p = re.search(r'\[MEDIA:(.*?)\]', parte_texto)
                        if media_match_p:
                            parte_media = (media_match_p.group(1) or '').strip()
                            parte_texto = parte_texto.replace(media_match_p.group(0), '').strip()
                            logger.info(
                                '📎 [MEDIA] parte %s url=%s',
                                idx + 1,
                                parte_media[:500] + ('…' if len(parte_media) > 500 else ''),
                            )
                            if parte_media and youtube_hace_solo_enlace_en_texto(parte_media):
                                logger.warning(
                                    '📎 [MEDIA] URL parece página de YouTube (no MP4 directo); '
                                    'Twilio suele necesitar enlace al archivo. parte=%s',
                                    idx + 1,
                                )
                            if not parte_texto.strip() and parte_media:
                                parte_texto = TWILIO_CAPTION_ADJUNTO
                    
                    mensajes_enviados = _enviar_mensaje_twilio_segmentado(
                        client=client,
                        from_number=twilio_number,
                        to_number=destino_formateado,
                        body=parte_texto,
                        media_url=parte_media,
                    )

                    for seg_idx, (mensaje, texto_enviado) in enumerate(mensajes_enviados, start=1):
                        print(f"✅ Mensaje {idx+1}.{seg_idx} enviado via Twilio: {mensaje.sid}")
                        texto_log = texto_enviado or parte_texto or (f"[MEDIA:{parte_media}]" if parte_media else '')
                        WhatsappLog.objects.create(
                            telefono=telefono_limpio,
                            mensaje=texto_log[:1500],
                            mensaje_id=mensaje.sid,
                            tipo='SENT'
                        )
                        print(f"✅ Guardado SENT")
                    
                    # Small delay between messages to avoid rate limiting
                    import time
                    time.sleep(0.5)
            else:
                # Single message (original behavior)
                mensajes_enviados = _enviar_mensaje_twilio_segmentado(
                    client=client,
                    from_number=twilio_number,
                    to_number=destino_formateado,
                    body=texto_respuesta,
                    media_url=media_url_to_send,
                )

                for seg_idx, (mensaje, texto_enviado) in enumerate(mensajes_enviados, start=1):
                    if len(mensajes_enviados) == 1:
                        print(f"✅ Mensaje enviado via Twilio: {mensaje.sid}")
                    else:
                        print(f"✅ Mensaje segmento {seg_idx}/{len(mensajes_enviados)} enviado via Twilio: {mensaje.sid}")

                    texto_log = texto_enviado or texto_respuesta or (f"[MEDIA:{media_url_to_send}]" if media_url_to_send else '')
                    WhatsappLog.objects.create(
                        telefono=telefono_limpio,
                        mensaje=texto_log[:1500],
                        mensaje_id=mensaje.sid,
                        tipo='SENT'
                    )
                    print(f"✅ Guardado SENT")
            
        except Exception as e:
            print(f"❌ Error enviando respuesta Twilio: {str(e)}")
            import traceback
            traceback.print_exc()
    
    except Exception as e:
        print(f"❌ Error en _procesar_twilio_webhook: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # ============================================================
        # SAFETY NET: Siempre enviar ALGO al usuario, nunca quedar mudo
        # ============================================================
        try:
            msg_from_fallback = post_data.get('From', '')
            if msg_from_fallback:
                from twilio.rest import Client
                account_sid = getattr(settings, 'TWILIO_ACCOUNT_SID', '')
                auth_token = getattr(settings, 'TWILIO_AUTH_TOKEN', '')
                twilio_number = getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806')
                twilio_number = str(twilio_number).strip()
                
                if account_sid and auth_token:
                    client = Client(account_sid, auth_token)
                    destino = f'whatsapp:{msg_from_fallback}' if not msg_from_fallback.startswith('whatsapp:') else msg_from_fallback
                    destino = str(destino).strip()
                    client.messages.create(
                        body="⚠️ Tuvimos un problema técnico momentáneo. Por favor vuelve a escribir tu mensaje para continuar.",
                        from_=twilio_number,
                        to=destino
                    )
                    print(f"✅ Safety net: mensaje de error enviado a {destino}")
        except Exception as fallback_err:
            print(f"❌ Safety net también falló: {fallback_err}")


def _procesar_meta_webhook(payload):
    """Procesa webhooks de Meta WhatsApp (mantiene compatibilidad)"""
    try:
        print("🔵 META: Procesando...")
        entries = payload.get('entry', [])
        
        for entry in entries:
            changes = entry.get('changes', [])
            for change in changes:
                value = change.get('value', {})
                meta_to = (
                    (value.get('metadata') or {}).get('display_phone_number')
                    or (value.get('metadata') or {}).get('phone_number_id')
                    or ''
                )
                
                # Mensajes entrantes
                messages = value.get('messages', [])
                for m in messages:
                    phone = m.get('from')
                    msg_id = m.get('id')
                    text = ''
                    if 'text' in m and isinstance(m['text'], dict):
                        text = m['text'].get('body', '')
                    
                    # Guardar mensaje
                    WhatsappLog.objects.create(
                        telefono=phone,
                        mensaje=text,
                        mensaje_id=msg_id,
                        tipo='INCOMING'
                    )
                    
                    # Obtener o crear estudiante
                    estudiante, _ = Estudiante.objects.get_or_create(
                        telefono=phone,
                        defaults={'nombre': 'Usuario', 'activo': True, 'cedula': f'META_{phone[-10:]}'}
                    )
                    
                    # Verificar seguridad primero
                    from .security_handler import verificar_seguridad_completa
                    bloqueado, respuesta_seguridad, estudiante = verificar_seguridad_completa(
                        estudiante,
                        text,
                        telefono=phone,
                        numero_destino=meta_to,
                    )
                    
                    if bloqueado:
                        texto_respuesta = respuesta_seguridad
                    else:
                        # Detectar intent
                        intent = detect_intent(text)
                        
                        if intent != 'desconocido':
                            # Usar template
                            texto_respuesta = get_response_for_intent(
                                intent, 
                                estudiante.nombre,
                                estudiante_id=estudiante.id,
                                mensaje_original=text
                            )
                        else:
                            # Usar IA solo para preguntas
                            try:
                                from .ai_assistant import responder_con_ia
                                texto_respuesta = responder_con_ia(text, phone)
                            except Exception as e:
                                print(f"Error IA: {e}")
                                texto_respuesta = "Disculpa, tengo problemas técnicos. Vuelve a escribir tu mensaje para continuar."
                    
                    # Enviar respuesta
                    resultado_envio = enviar_whatsapp(phone, texto_respuesta)
                    
                    if resultado_envio.get('success'):
                        WhatsappLog.objects.create(
                            telefono=phone,
                            mensaje=texto_respuesta,
                            mensaje_id=resultado_envio.get('mensaje_id'),
                            tipo='SENT'
                        )
    
    except Exception as e:
        print(f"❌ Error en _procesar_meta_webhook: {str(e)}")
        import traceback
        traceback.print_exc()


@staff_member_required
def probar_twilio_view(request):
    """Vista para probar integración con Twilio WhatsApp"""
    context = {
        'mensaje': None,
        'error': False,
        'resultado': None
    }
    
    if request.method == 'POST':
        try:
            from twilio.rest import Client
            import os
            
            # Obtener datos del formulario
            tipo_mensaje = request.POST.get('tipo_mensaje')
            usar_template = request.POST.get('usar_template') == 'on'
            telefono = request.POST.get('telefono', '').strip()
            mensaje_texto = request.POST.get('mensaje', '').strip()
            url_imagen = request.POST.get('url_imagen', '').strip()
            
            # Validar credenciales
            account_sid = os.environ.get('TWILIO_ACCOUNT_SID')
            auth_token = os.environ.get('TWILIO_AUTH_TOKEN')
            template_sid = os.environ.get('TWILIO_TEMPLATE_SID')
            
            if not account_sid or not auth_token:
                context['mensaje'] = '<strong>❌ Error:</strong> Las credenciales de Twilio no están configuradas en el archivo .env'
                context['error'] = True
                return render(request, 'admin/probar_twilio.html', context)
            
            # Validar teléfono
            if not telefono:
                context['mensaje'] = '<strong>❌ Error:</strong> Debes proporcionar un número de teléfono'
                context['error'] = True
                return render(request, 'admin/probar_twilio.html', context)
            
            # Asegurar formato whatsapp:
            if not telefono.startswith('+'):
                telefono = f'+{telefono}'
            if not telefono.startswith('whatsapp:'):
                telefono_whatsapp = f'whatsapp:{telefono}'
            else:
                telefono_whatsapp = telefono
            
            # Crear cliente Twilio
            client = Client(account_sid, auth_token)
            
            # Si se usa template aprobado
            if usar_template and template_sid:
                message = client.messages.create(
                    content_sid=template_sid,
                    from_=getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806'),
                    to=telefono_whatsapp
                )
            else:
                # Preparar parámetros del mensaje libre
                params = {
                    "to": telefono_whatsapp,
                    "from_": getattr(settings, 'TWILIO_PHONE_NUMBER', 'whatsapp:+573202948806'),
                    "body": mensaje_texto
                }
                
                # Si es mensaje con imagen/video, generar URL firmada y agregar media_url
                if tipo_mensaje == 'imagen' and url_imagen:
                    # Revisar si la URL es de S3 y necesita firma
                    from core.utils import generar_url_firmada_s3_v4
                    import re
                    s3_pattern = r'https://([\w\-]+)\.s3[\w\-\.]*\.amazonaws\.com/(.+)'
                    match = re.match(s3_pattern, url_imagen)
                    if match:
                        bucket_name = match.group(1)
                        object_name = match.group(2)
                        url_firmada = generar_url_firmada_s3_v4(bucket_name, object_name)
                        params["media_url"] = [url_firmada]
                    else:
                        params["media_url"] = [url_imagen]
                
                # Enviar mensaje
                message = client.messages.create(**params)
            
            # Crear resultado formateado
            resultado_texto = f"""
✅ MENSAJE ENVIADO EXITOSAMENTE

📝 SID: {message.sid}
📊 Estado: {message.status}
📅 Fecha: {message.date_created}
📱 Destino: {telefono}
"""
            
            if usar_template and template_sid:
                resultado_texto += f"📋 Template SID: {template_sid}\n"
            else:
                resultado_texto += f"💬 Mensaje: {mensaje_texto[:100]}{'...' if len(mensaje_texto) > 100 else ''}\n"
                if tipo_mensaje == 'imagen' and url_imagen:
                    resultado_texto += f"🖼️  Imagen: {url_imagen}\n"
            
            context['mensaje'] = f'<strong>✅ ¡Éxito!</strong> El mensaje fue enviado correctamente. SID: {message.sid}'
            context['error'] = False
            context['resultado'] = resultado_texto
            
            # Guardar log
            WhatsappLog.objects.create(
                telefono=telefono.replace('whatsapp:', '').replace('+', ''),
                mensaje=mensaje_texto,
                mensaje_id=message.sid,
                estado='SENT'
            )
            
        except Exception as e:
            context['mensaje'] = f'<strong>❌ Error al enviar:</strong> {str(e)}'
            context['error'] = True
            context['resultado'] = f"ERROR:\n{str(e)}"
    
    return render(request, 'admin/probar_twilio.html', context)


@staff_member_required
def calendario_campanas_view(request):
    """Vista de calendario de campañas programadas"""
    from django.utils import timezone
    
    ahora = timezone.now()
    
    # Campañas pendientes (programadas pero no ejecutadas)
    campanas_pendientes = Campana.objects.filter(
        fecha_programada__isnull=False,
        ejecutada=False
    ).order_by('fecha_programada')
    
    # Campañas ejecutadas que tenían programación
    campanas_ejecutadas = Campana.objects.filter(
        fecha_programada__isnull=False,
        ejecutada=True
    ).order_by('-fecha_programada')[:10]
    
    context = {
        'campanas_pendientes': campanas_pendientes,
        'campanas_ejecutadas': campanas_ejecutadas,
    }
    
    return render(request, 'admin/calendario_campanas.html', context)


@staff_member_required
def conversaciones_view(request):
    """Vista de conversaciones estilo WhatsApp Web (pantalla completa)."""
    from core.conversaciones_service import construir_contexto_inbox

    cliente_filtro_raw = (request.GET.get("cliente") or "").strip()
    cliente_filtro_id = int(cliente_filtro_raw) if cliente_filtro_raw.isdigit() else None
    estudiante_raw = (request.GET.get("estudiante") or "").strip()
    estudiante_id = int(estudiante_raw) if estudiante_raw.isdigit() else None
    page_raw = (request.GET.get("page") or "1").strip()
    page = int(page_raw) if page_raw.isdigit() else 1

    context = construir_contexto_inbox(
        cliente_filtro_id=cliente_filtro_id,
        estudiante_id=estudiante_id,
        telefono=(request.GET.get("telefono") or "").strip() or None,
        page=page,
    )
    context.update({
        "inbox_base_url": "/admin/conversaciones/",
        "inbox_volver_url": "/admin/",
        "inbox_volver_label": "Panel admin",
        "inbox_modo": "admin",
    })
    return render(request, "admin/conversaciones.html", context)


@staff_member_required
def chat_prueba_view(request):
    """Vista para probar la IA sin necesidad de WhatsApp/ngrok"""
    return render(request, 'admin/chat_prueba.html')


@staff_member_required
def chat_prueba_api(request):
    """API para el chat de prueba"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            mensaje = data.get('mensaje', '')
            telefono = data.get('telefono', 'test_chat')
            
            print(f"🔵 Chat de prueba - Mensaje: {mensaje}")
            
            # Guardar mensaje entrante
            WhatsappLog.objects.create(
                telefono=telefono,
                mensaje=mensaje,
                mensaje_id=f"test_{timezone.now().timestamp()}",
                tipo='INCOMING'
            )
            
            # Obtener respuesta de la IA
            try:
                from .ai_assistant import responder_con_ia
                respuesta = responder_con_ia(mensaje, telefono)
                print(f"✅ IA respondió: {respuesta}")
            except Exception as e:
                print(f"❌ Error en IA: {e}")
                # Fallback
                from .intent_detector import detect_intent
                from .response_templates import get_response_for_intent
                intent = detect_intent(mensaje)
                respuesta = get_response_for_intent(intent, 'Usuario')
            
            # Guardar respuesta
            WhatsappLog.objects.create(
                telefono=telefono,
                mensaje=respuesta,
                mensaje_id=f"test_response_{timezone.now().timestamp()}",
                tipo='SENT'
            )
            
            return JsonResponse({
                'success': True,
                'respuesta': respuesta
            })
            
        except Exception as e:
            print(f"❌ Error en chat de prueba: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def obtener_archivos_modulo_view(request, modulo_id):
    """
    API para obtener archivos multimedia de un módulo específico
    Usado por estudiantes para ver contenido disponible
    """
    from .models import Modulo
    
    try:
        modulo = get_object_or_404(Modulo, id=modulo_id)
        archivos = modulo.archivos_multimedia.filter(activo=True).order_by('orden', 'id')
        
        archivos_data = []
        for archivo in archivos:
            archivos_data.append({
                'id': archivo.id,
                'tipo': archivo.get_tipo_display(),
                'titulo': archivo.titulo,
                'descripcion': archivo.descripcion,
                'url_descarga': f'/media/descargar-archivo/{archivo.id}/' if archivo.archivo else None,
                'url_externa': archivo.url_externa,
                'url_proxy': archivo.get_url_para_envio(),
                'disponible_offline': archivo.disponible_offline,
                'tamano_mb': archivo.tamano_mb(),
                'duracion_segundos': archivo.duracion_segundos,
            })
        
        return JsonResponse({
            'success': True,
            'modulo': {
                'id': modulo.id,
                'titulo': modulo.titulo,
                'numero': modulo.numero,
            },
            'archivos': archivos_data,
            'total': len(archivos_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


def descargar_archivo_multimedia(request, archivo_id):
    """
    Descarga un archivo multimedia específico
    Permite descarga offline si está habilitada
    """
    try:
        archivo = get_object_or_404(ArchivoModulo, id=archivo_id)
        
        if not archivo.archivo:
            return JsonResponse({
                'error': 'Este archivo no tiene descarga disponible. Usa la URL externa.'
            }, status=400)
        
        # Verificar si la descarga offline está permitida
        if not archivo.disponible_offline:
            return JsonResponse({
                'error': 'La descarga offline no está habilitada para este archivo.'
            }, status=403)
        
        # Retornar el archivo para descarga
        response = FileResponse(archivo.archivo.open('rb'))
        response['Content-Disposition'] = f'attachment; filename="{archivo.titulo}.{archivo.archivo.name.split(".")[-1]}"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error al descargar archivo: {str(e)}'
        }, status=500)


def stream_media(request):
    """Proxy simple para servir archivos multimedia almacenados (oculta la URL S3).

    Parámetros: ?path=<ruta_relativa_en_storage>
    Ej: /media/stream/?path=modulos/2026/02/video.mp4
    """
    path = request.GET.get('path')
    if not path:
        return HttpResponseBadRequest('Falta parámetro path')

    # Normalizar y evitar traversal
    path = path.lstrip('/')
    if '..' in path:
        return HttpResponseBadRequest('Ruta inválida')

    try:
        f = default_storage.open(path, 'rb')
        content_type = mimetypes.guess_type(path)[0] or 'application/octet-stream'
        return FileResponse(f, content_type=content_type)
    except Exception:
        return HttpResponse(status=404)


@staff_member_required
def test_email_gmail_view(request):
    """Vista para probar la configuración de Gmail"""
    from .email_test import test_gmail_connection, format_email_status_html
    
    context = {
        'title': 'Probar Conexión Gmail',
        'status_html': format_email_status_html(),
        'resultado': None
    }
    
    if request.method == 'POST':
        success, message = test_gmail_connection()
        context['resultado'] = {
            'success': success,
            'message': message
        }
    
    return render(request, 'admin/test_email.html', context)


# ========================================
# VISTAS PARA GENERACIÓN DE CURSOS CON IA
# ========================================

def _contexto_crear_curso_ia(request):
    from .models import Cliente
    from .utils_ia import MODELO_IA_DEFAULT, modelos_ia_habilitados

    historial = request.session.get('chat_curso_ia') or []
    modelos = modelos_ia_habilitados()
    return {
        'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
        'modelos_ia': modelos,
        'modelos_ia_disponibles': bool(modelos),
        'historial_chat': historial,
        'modelo_actual': request.session.get('modelo_usado', MODELO_IA_DEFAULT),
    }


def _guardar_sesion_curso_ia(request, estructura, cliente_id, fuente_nombre, modelo_ia, texto_fuente=''):
    request.session['estructura_curso'] = estructura
    request.session['cliente_id'] = str(cliente_id)
    request.session['archivo_nombre'] = fuente_nombre or 'prompt-ia'
    request.session['modelo_usado'] = modelo_ia
    request.session['texto_fuente_curso'] = (texto_fuente or '')[:12000]
    request.session.modified = True


@staff_member_required
def subir_documento_curso(request):
    """Paso 1: chat + prompt largo o archivo → estructura JSON del curso."""
    from django.shortcuts import redirect
    from .models import Cliente
    from .utils_ia import (
        extraer_texto_documento,
        generar_estructura_curso_con_ia,
        validar_estructura_curso,
    )

    context = _contexto_crear_curso_ia(request)

    if request.method == 'POST':
        try:
            accion = request.POST.get('accion', 'generar')
            cliente_id = request.POST.get('cliente_id')
            modelo_ia = request.POST.get('modelo_ia', 'gpt-4o-mini')
            prompt_usuario = (request.POST.get('prompt') or '').strip()
            archivo = request.FILES.get('documento')

            if accion == 'limpiar_chat':
                request.session.pop('chat_curso_ia', None)
                return redirect('subir_documento_curso')

            if not cliente_id:
                context['error'] = 'Selecciona una organización'
                return render(request, 'admin/subir_documento_curso.html', context)

            try:
                cliente = Cliente.objects.get(id=cliente_id)
            except Cliente.DoesNotExist:
                context['error'] = 'Organización no encontrada'
                return render(request, 'admin/subir_documento_curso.html', context)

            texto = ''
            fuente = 'prompt-ia'
            if archivo:
                nombre = archivo.name.lower()
                if not (nombre.endswith('.pdf') or nombre.endswith('.docx') or nombre.endswith('.txt')):
                    context['error'] = 'Solo PDF, Word (.docx) o TXT'
                    return render(request, 'admin/subir_documento_curso.html', context)
                texto = extraer_texto_documento(archivo)
                fuente = archivo.name
            elif prompt_usuario:
                texto = prompt_usuario
                fuente = 'prompt-chat'
            else:
                context['error'] = 'Escribe un prompt o sube un documento'
                return render(request, 'admin/subir_documento_curso.html', context)

            if len(texto) < 200:
                context['error'] = 'El contenido es muy corto (mínimo 200 caracteres)'
                return render(request, 'admin/subir_documento_curso.html', context)

            from .utils_ia import validar_modelo_ia_disponible
            try:
                validar_modelo_ia_disponible(modelo_ia)
            except ValueError as e:
                context['error'] = str(e)
                return render(request, 'admin/subir_documento_curso.html', context)

            historial = list(request.session.get('chat_curso_ia') or [])
            historial.append({'rol': 'user', 'texto': prompt_usuario or f'[Archivo: {fuente}]'})
            request.session['chat_curso_ia'] = historial[-20:]

            use_async = not getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
            if use_async:
                import uuid
                from django.core.cache import cache
                from core.tasks import generar_curso_ia_async, _curso_ia_cache_key

                job_id = str(uuid.uuid4())
                cache.set(_curso_ia_cache_key(job_id), {'status': 'pending'}, 3600)
                request.session['curso_ia_job_id'] = job_id
                request.session['curso_ia_pending'] = {
                    'cliente_id': str(cliente_id),
                    'fuente': fuente,
                    'modelo_ia': modelo_ia,
                    'texto': texto[:12000],
                    'prompt_usuario': prompt_usuario,
                }
                request.session.modified = True
                try:
                    generar_curso_ia_async.delay(job_id, texto, modelo_ia)
                    return redirect('generando_curso_ia')
                except Exception as celery_err:
                    logger.warning('Celery no disponible para curso IA, modo sync: %s', celery_err)
                    request.session.pop('curso_ia_job_id', None)
                    request.session.pop('curso_ia_pending', None)

            estructura = generar_estructura_curso_con_ia(texto, modelo=modelo_ia)
            es_valida, errores = validar_estructura_curso(estructura)
            if not es_valida:
                context['error'] = f'Estructura inválida: {", ".join(errores)}'
                return render(request, 'admin/subir_documento_curso.html', context)

            historial.append({
                'rol': 'assistant',
                'texto': f'Generé «{estructura.get("titulo", "Curso")}» con {len(estructura.get("modulos", []))} módulos.',
            })
            request.session['chat_curso_ia'] = historial[-20:]
            _guardar_sesion_curso_ia(request, estructura, cliente_id, fuente, modelo_ia, texto)
            return redirect('vista_previa_curso_ia')

        except ValueError as e:
            context['error'] = str(e)
        except Exception as e:
            context['error'] = f'Error: {e}'
            logger.error(f'Error en subir_documento_curso: {e}', exc_info=True)

    return render(request, 'admin/subir_documento_curso.html', context)


@staff_member_required
def generando_curso_ia(request):
    """Pantalla de espera mientras Celery genera la estructura (evita 504)."""
    job_id = request.session.get('curso_ia_job_id')
    if not job_id:
        return redirect('subir_documento_curso')
    return render(request, 'admin/generando_curso_ia.html', {'job_id': job_id})


@staff_member_required
def api_estado_curso_ia(request):
    """Polling JSON del job de generación IA."""
    from django.core.cache import cache
    from django.http import JsonResponse
    from core.tasks import _curso_ia_cache_key

    job_id = request.GET.get('job_id') or request.session.get('curso_ia_job_id')
    if not job_id:
        return JsonResponse({'status': 'missing'}, status=404)
    data = cache.get(_curso_ia_cache_key(job_id)) or {'status': 'pending'}
    if data.get('status') == 'ok':
        pending = request.session.get('curso_ia_pending') or {}
        estructura = data.get('estructura')
        if estructura and pending.get('cliente_id'):
            historial = list(request.session.get('chat_curso_ia') or [])
            historial.append({
                'rol': 'assistant',
                'texto': f'Generé «{estructura.get("titulo", "Curso")}» con {len(estructura.get("modulos", []))} módulos.',
            })
            request.session['chat_curso_ia'] = historial[-20:]
            _guardar_sesion_curso_ia(
                request,
                estructura,
                pending['cliente_id'],
                pending.get('fuente', 'prompt-ia'),
                pending.get('modelo_ia', 'gpt-4o-mini'),
                pending.get('texto', ''),
            )
            request.session.pop('curso_ia_job_id', None)
            request.session.pop('curso_ia_pending', None)
            return JsonResponse({
                'status': 'ok',
                'redirect': '/admin/vista-previa-curso-ia/',
                'titulo': estructura.get('titulo'),
                'modulos': len(estructura.get('modulos', [])),
            })
    if data.get('status') == 'error':
        request.session.pop('curso_ia_job_id', None)
        request.session.pop('curso_ia_pending', None)
    return JsonResponse({
        'status': data.get('status', 'pending'),
        'error': data.get('error'),
    })


@staff_member_required
def vista_previa_curso_ia(request):
    """Paso 2: revisión humana, edición y regeneración por módulo."""
    from django.contrib import messages
    from django.shortcuts import redirect
    from .models import Cliente
    from .utils_ia import guardar_curso_desde_estructura, regenerar_modulo_en_estructura

    estructura = request.session.get('estructura_curso')
    cliente_id = request.session.get('cliente_id')
    archivo_nombre = request.session.get('archivo_nombre', 'prompt-ia')
    modelo_usado = request.session.get('modelo_usado', 'gpt-4o-mini')
    texto_fuente = request.session.get('texto_fuente_curso', '')

    if not estructura or not cliente_id:
        messages.error(request, 'No hay borrador de curso. Genera uno primero.')
        return redirect('subir_documento_curso')

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        messages.error(request, 'Organización no encontrada')
        return redirect('subir_documento_curso')

    context = {
        'estructura': estructura,
        'estructura_json': json.dumps(estructura, ensure_ascii=False, indent=2),
        'cliente': cliente,
        'archivo_nombre': archivo_nombre,
        'modelo_usado': modelo_usado,
        'total_modulos': len(estructura.get('modulos', [])),
        'total_lecciones': sum(len(m.get('lecciones', [])) for m in estructura.get('modulos', [])),
        'historial_chat': request.session.get('chat_curso_ia') or [],
    }

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'guardar':
            try:
                estructura['titulo'] = request.POST.get('titulo', estructura['titulo'])
                estructura['descripcion'] = request.POST.get('descripcion', estructura['descripcion'])
                estructura['duracion_estimada'] = request.POST.get(
                    'duracion_estimada', estructura.get('duracion_estimada', '4 semanas')
                )
                estructura['nivel'] = request.POST.get('nivel', estructura.get('nivel', 'Intermedio'))
                estructura['puntos_por_leccion'] = int(
                    request.POST.get('puntos_por_leccion', estructura.get('puntos_por_leccion', 50))
                )
                curso = guardar_curso_desde_estructura(estructura, cliente, archivo_nombre)
                for key in (
                    'estructura_curso', 'cliente_id', 'archivo_nombre',
                    'modelo_usado', 'texto_fuente_curso', 'chat_curso_ia',
                ):
                    request.session.pop(key, None)
                messages.success(
                    request,
                    f'Curso «{curso.nombre}» creado (inactivo). Revísalo en el admin antes de activar.',
                )
                return redirect(f'/admin/core/curso/{curso.id}/change/')
            except Exception as e:
                context['error'] = f'Error al guardar: {e}'
                logger.error(f'Error guardando curso IA: {e}', exc_info=True)

        elif accion == 'regenerar_modulo':
            try:
                idx = int(request.POST.get('modulo_indice', 0))
                instrucciones = (request.POST.get('instrucciones_regenerar') or '').strip()
                estructura = regenerar_modulo_en_estructura(
                    estructura,
                    idx,
                    texto_fuente=texto_fuente,
                    instrucciones=instrucciones,
                    modelo=modelo_usado,
                )
                request.session['estructura_curso'] = estructura
                historial = list(request.session.get('chat_curso_ia') or [])
                historial.append({
                    'rol': 'assistant',
                    'texto': f'Regeneré el módulo {idx + 1}: {estructura["modulos"][idx].get("nombre", "")}',
                })
                request.session['chat_curso_ia'] = historial[-20:]
                messages.success(request, f'Módulo {idx + 1} regenerado.')
                return redirect('vista_previa_curso_ia')
            except Exception as e:
                context['error'] = f'No se pudo regenerar: {e}'

        elif accion == 'cancelar':
            for key in (
                'estructura_curso', 'cliente_id', 'archivo_nombre',
                'modelo_usado', 'texto_fuente_curso', 'chat_curso_ia',
            ):
                request.session.pop(key, None)
            messages.info(request, 'Creación cancelada')
            return redirect('subir_documento_curso')

        context['estructura'] = estructura
        context['estructura_json'] = json.dumps(estructura, ensure_ascii=False, indent=2)
        context['total_modulos'] = len(estructura.get('modulos', []))

    return render(request, 'admin/vista_previa_curso_ia.html', context)
