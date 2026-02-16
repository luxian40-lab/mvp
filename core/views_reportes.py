"""
Views para reportes avanzados y análisis de datos
"""
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Import seguro de modelos
try:
    from .models import Estudiante, WhatsappLog, InteraccionLog
except ImportError:
    Estudiante = WhatsappLog = InteraccionLog = None

import logging
logger = logging.getLogger(__name__)



@staff_member_required
def dashboard_reportes_avanzados(request):
    """
    Dashboard avanzado con análisis de municipios, canales de comunicación y más
    """
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        municipio_filtro = request.GET.get('municipio')

        # Configurar rango de fechas (últimos 30 días por defecto)
        if fecha_inicio and fecha_fin:
            from datetime import datetime
            try:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            except (ValueError, TypeError):
                fecha_fin = timezone.now()
                fecha_inicio = fecha_fin - timedelta(days=30)
        else:
            fecha_fin = timezone.now()
            fecha_inicio = fecha_fin - timedelta(days=30)

        estudiantes_query = Estudiante.objects.filter(fecha_registro__range=[fecha_inicio, fecha_fin]) if Estudiante else []
        if municipio_filtro and hasattr(estudiantes_query, 'filter'):
            estudiantes_query = estudiantes_query.filter(municipio__icontains=municipio_filtro)

        municipios = estudiantes_query.values('municipio').annotate(
            total=Count('id'),
            activos=Count('id', filter=Q(activo=True)),
            completados=Count('id', filter=Q(progresos__completado=True))
        ).order_by('-total')[:20] if hasattr(estudiantes_query, 'values') else []

        logs_query = WhatsappLog.objects.filter(
            fecha__range=[fecha_inicio, fecha_fin],
            tipo='INCOMING'
        ) if WhatsappLog else []
        if municipio_filtro and hasattr(logs_query, 'filter'):
            logs_query = logs_query.filter(estudiante__municipio__icontains=municipio_filtro)

        total_mensajes = logs_query.count() if hasattr(logs_query, 'count') else 0
        mensajes_audio = logs_query.filter(
            Q(mensaje__icontains='audio') |
            Q(mensaje__icontains='🎤') |
            Q(mensaje__icontains='voice')
        ).count() if hasattr(logs_query, 'filter') else 0
        mensajes_texto = total_mensajes - mensajes_audio

        analisis_estudiantes = []
        if hasattr(estudiantes_query, 'select_related'):
            for estudiante in estudiantes_query.select_related().prefetch_related('progresos')[:100]:
                mensajes = WhatsappLog.objects.filter(
                    estudiante=estudiante,
                    fecha__range=[fecha_inicio, fecha_fin],
                    tipo='INCOMING'
                ) if WhatsappLog else []
                total_msgs = mensajes.count() if hasattr(mensajes, 'count') else 0
                audio_msgs = mensajes.filter(
                    Q(mensaje__icontains='audio') |
                    Q(mensaje__icontains='🎤')
                ).count() if hasattr(mensajes, 'filter') else 0
                texto_msgs = total_msgs - audio_msgs
                if total_msgs > 0:
                    preferencia_audio = (audio_msgs / total_msgs) * 100
                    canal_preferido = '🎤 Audio' if audio_msgs > texto_msgs else '📝 Texto'
                else:
                    preferencia_audio = 0
                    canal_preferido = 'Sin interacción'
                # porcentaje_avance is a method, not a DB field — calculate manually
                _progresos = list(estudiante.progresos.all())
                progreso_promedio = (sum(p.porcentaje_avance() for p in _progresos) / len(_progresos)) if _progresos else 0
                analisis_estudiantes.append({
                    'estudiante': estudiante,
                    'total_mensajes': total_msgs,
                    'mensajes_audio': audio_msgs,
                    'mensajes_texto': texto_msgs,
                    'preferencia_audio': round(preferencia_audio, 1),
                    'canal_preferido': canal_preferido,
                    'progreso_promedio': round(progreso_promedio or 0, 1)
                })
        analisis_estudiantes.sort(key=lambda x: x['total_mensajes'], reverse=True)

        total_estudiantes = estudiantes_query.count() if hasattr(estudiantes_query, 'count') else 0
        estudiantes_activos = estudiantes_query.filter(activo=True).count() if hasattr(estudiantes_query, 'filter') else 0
        estudiantes_audio = sum(1 for e in analisis_estudiantes if e['canal_preferido'] == '🎤 Audio')
        estudiantes_texto = sum(1 for e in analisis_estudiantes if e['canal_preferido'] == '📝 Texto')

        todos_municipios = Estudiante.objects.values_list('municipio', flat=True).distinct().order_by('municipio') if Estudiante else []

        context = {
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
            'municipio_filtro': municipio_filtro or '',
            'todos_municipios': [m for m in todos_municipios if m],
            'total_estudiantes': total_estudiantes,
            'estudiantes_activos': estudiantes_activos,
            'total_mensajes': total_mensajes,
            'mensajes_audio': mensajes_audio,
            'mensajes_texto': mensajes_texto,
            'estudiantes_audio': estudiantes_audio,
            'estudiantes_texto': estudiantes_texto,
            'municipios': municipios,
            'analisis_estudiantes': analisis_estudiantes[:50],
            'porcentaje_audio': round((mensajes_audio / total_mensajes * 100) if total_mensajes > 0 else 0, 1),
            'porcentaje_texto': round((mensajes_texto / total_mensajes * 100) if total_mensajes > 0 else 0, 1),
        }
        return render(request, 'admin/dashboard_reportes_avanzados.html', context)
    except Exception as e:
        # logger.error(f"Error en dashboard_reportes_avanzados: {e}", exc_info=True)
        return render(request, 'admin/error_page.html', {
            'error_message': 'Ocurrió un error al cargar el dashboard de reportes avanzados.',
            'error_detail': str(e) if hasattr(request, 'user') and getattr(request.user, 'is_superuser', False) else None
        }, status=500)


@staff_member_required
def descargar_reporte_xlsx(request):
    """
    Descarga reporte completo en formato XLSX
    """
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    municipio_filtro = request.GET.get('municipio')
    
    # Configurar rango de fechas
    if fecha_inicio and fecha_fin:
        from datetime import datetime
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
    else:
        fecha_fin = timezone.now()
        fecha_inicio = fecha_fin - timedelta(days=30)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # === HOJA 1: ANÁLISIS POR MUNICIPIO ===
    ws1 = wb.active
    ws1.title = "Por Municipio"
    
    # Encabezados
    headers1 = ['Municipio', 'Total Estudiantes', 'Activos', 'Completados', '% Activos', '% Completados']
    ws1.append(headers1)
    
    # Estilos de encabezado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    estudiantes_query = Estudiante.objects.filter(fecha_registro__range=[fecha_inicio, fecha_fin])
    if municipio_filtro:
        estudiantes_query = estudiantes_query.filter(municipio__icontains=municipio_filtro)
    
    municipios = estudiantes_query.values('municipio').annotate(
        total=Count('id'),
        activos=Count('id', filter=Q(activo=True)),
        completados=Count('id', filter=Q(progresos__completado=True))
    ).order_by('-total')
    
    for m in municipios:
        total = m['total']
        activos = m['activos']
        completados = m['completados']
        ws1.append([
            m['municipio'] or 'Sin municipio',
            total,
            activos,
            completados,
            f"{(activos/total*100) if total > 0 else 0:.1f}%",
            f"{(completados/total*100) if total > 0 else 0:.1f}%"
        ])
    
    # Ajustar anchos
    for i, column in enumerate(ws1.columns, 1):
        ws1.column_dimensions[get_column_letter(i)].width = 20
    
    # === HOJA 2: ANÁLISIS POR CANAL ===
    ws2 = wb.create_sheet("Por Canal")
    
    headers2 = ['Nombre', 'Teléfono', 'Municipio', 'Total Mensajes', 'Audio', 'Texto', '% Audio', 'Canal Preferido', 'Progreso']
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for estudiante in estudiantes_query:
        mensajes = WhatsappLog.objects.filter(
            estudiante=estudiante,
            fecha__range=[fecha_inicio, fecha_fin],
            tipo='INCOMING'
        )
        
        total_msgs = mensajes.count()
        audio_msgs = mensajes.filter(
            Q(mensaje__icontains='audio') | 
            Q(mensaje__icontains='🎤')
        ).count()
        texto_msgs = total_msgs - audio_msgs
        
        if total_msgs > 0:
            pref_audio = (audio_msgs / total_msgs) * 100
            canal_pref = 'Audio' if audio_msgs > texto_msgs else 'Texto'
        else:
            pref_audio = 0
            canal_pref = 'Sin interacción'
        
        _progs = list(estudiante.progresos.all())
        progreso = (sum(p.porcentaje_avance() for p in _progs) / len(_progs)) if _progs else 0
        
        ws2.append([
            estudiante.nombre,
            estudiante.telefono,
            estudiante.municipio or 'Sin municipio',
            total_msgs,
            audio_msgs,
            texto_msgs,
            f"{pref_audio:.1f}%",
            canal_pref,
            f"{progreso:.1f}%"
        ])
    
    # Ajustar anchos
    for i, column in enumerate(ws2.columns, 1):
        ws2.column_dimensions[get_column_letter(i)].width = 18
    
    # === HOJA 3: RESUMEN EJECUTIVO ===
    ws3 = wb.create_sheet("Resumen")
    
    total_estudiantes = estudiantes_query.count()
    total_activos = estudiantes_query.filter(activo=True).count()
    total_mensajes = WhatsappLog.objects.filter(
        fecha__range=[fecha_inicio, fecha_fin],
        tipo='INCOMING'
    ).count()
    
    ws3.append(['REPORTE DE ANÁLISIS EKI'])
    ws3.append([''])
    ws3.append(['Período', f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"])
    ws3.append(['Fecha de generación', timezone.now().strftime('%d/%m/%Y %H:%M')])
    ws3.append([''])
    ws3.append(['ESTADÍSTICAS GENERALES'])
    ws3.append(['Total de estudiantes', total_estudiantes])
    ws3.append(['Estudiantes activos', total_activos, f"{(total_activos/total_estudiantes*100) if total_estudiantes > 0 else 0:.1f}%"])
    ws3.append(['Total mensajes recibidos', total_mensajes])
    ws3.append([''])
    ws3.append(['CANALES DE COMUNICACIÓN'])
    
    mensajes_audio_total = WhatsappLog.objects.filter(
        fecha__range=[fecha_inicio, fecha_fin],
        tipo='INCOMING'
    ).filter(Q(mensaje__icontains='audio') | Q(mensaje__icontains='🎤')).count()
    mensajes_texto_total = total_mensajes - mensajes_audio_total
    
    ws3.append(['Mensajes por audio', mensajes_audio_total, f"{(mensajes_audio_total/total_mensajes*100) if total_mensajes > 0 else 0:.1f}%"])
    ws3.append(['Mensajes por texto', mensajes_texto_total, f"{(mensajes_texto_total/total_mensajes*100) if total_mensajes > 0 else 0:.1f}%"])
    
    # Estilos
    ws3['A1'].font = Font(size=16, bold=True)
    ws3['A6'].font = Font(bold=True)
    ws3['A11'].font = Font(bold=True)
    
    for i in range(1, 15):
        ws3.row_dimensions[i].height = 20
    
    # Generar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_eki_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
