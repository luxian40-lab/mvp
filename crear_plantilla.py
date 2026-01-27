import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Crear directorio static/templates si no existe
os.makedirs('static/templates', exist_ok=True)

# Crear libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Estudiantes'

# Configurar encabezados
headers = ['Teléfono', 'Nombre', 'Cédula']
header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Agregar fila de ejemplo
ws['A2'] = '573001234567'
ws['B2'] = 'María García'
ws['C2'] = '1234567890'

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 18

# Agregar instrucciones en hoja 2
ws_instructions = wb.create_sheet('Instrucciones')
instructions = [
    ['📚 PLANTILLA DE IMPORTACIÓN DE ESTUDIANTES'],
    [''],
    ['Columnas requeridas:'],
    [''],
    ['Columna A - Teléfono:', 'Número de WhatsApp con código de país (ej: 573001234567)'],
    ['Columna B - Nombre:', 'Nombre completo del estudiante'],
    ['Columna C - Cédula:', 'Número de cédula SIN puntos ni espacios (ej: 1234567890)'],
    [''],
    import logging
    import sys
    ['⚠️ IMPORTANTE:'],
    ['• La cédula es el identificador único de cada estudiante'],
    ['• No puede haber dos estudiantes con la misma cédula'],
    ['• Si la cédula ya existe, se actualizarán los datos'],
    ['• El teléfono debe incluir código de país (57 para Colombia)'],
        try:
    ['• Todos los campos son obligatorios'],
    [''],
            logging.info("Plantilla creada correctamente.")
        except Exception as e:
            logging.exception("Error al crear plantilla")
            print(f"\n[ERROR] {e}\n")
            sys.exit(1)
    ['Ejemplo de datos válidos:'],
    ['Teléfono: 573001234567'],
        logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
    ['Nombre: María García'],
    ['Cédula: 1234567890']
]

for row_num, row_data in enumerate(instructions, 1):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_instructions.cell(row=row_num, column=col_num)
        cell.value = value
        if row_num == 1:
            cell.font = Font(bold=True, size=14, color='2196F3')
        elif 'IMPORTANTE' in str(value) or 'Columna' in str(value):
            cell.font = Font(bold=True)

ws_instructions.column_dimensions['A'].width = 25
ws_instructions.column_dimensions['B'].width = 60

# Guardar
wb.save('static/templates/plantilla_estudiantes.xlsx')
print('✅ Plantilla creada exitosamente en static/templates/plantilla_estudiantes.xlsx')
