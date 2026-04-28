import io
from datetime import datetime

import openpyxl
from django.contrib import admin, messages
from django.http import HttpResponse
from django.utils import timezone

from .models import FichaGEI, FlujoPregunta, ResultadoGEI, SesionFormulario, TipoFormulario


class ResultadoGEIInline(admin.StackedInline):
    model = ResultadoGEI
    extra = 0
    max_num = 1
    can_delete = False
    readonly_fields = (
        "em_fertilizante_kg",
        "em_combustible_kg",
        "em_energia_kg",
        "em_residuos_kg",
        "em_total_kg",
        "rem_bosque_kg",
        "balance_neto_tco2e",
        "intensidad_kg_co2e_por_kg",
        "evaluacion",
        "completitud_calculo_pct",
        "campos_faltantes",
        "fecha_calculo",
    )


class FlujoPreguntaInline(admin.TabularInline):
    model = FlujoPregunta
    extra = 0
    ordering = ("orden",)
    fields = (
        "orden",
        "campo_destino",
        "pregunta_texto",
        "tipo_dato",
        "rango_min",
        "rango_max",
        "unidad_parseo",
        "opciones_choice",
        "usar_llm_parseo",
        "es_opcional",
        "texto_reintento",
    )


@admin.register(TipoFormulario)
class TipoFormularioAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "curso", "modulo", "cliente_label", "activo", "fecha_creacion")
    list_filter = ("activo", "curso", "cliente")
    search_fields = ("nombre", "descripcion", "cliente__nombre")
    autocomplete_fields = ("cliente",)
    inlines = (FlujoPreguntaInline,)

    @admin.display(description="Cliente", ordering="cliente__nombre")
    def cliente_label(self, obj: "TipoFormulario") -> str:
        return obj.cliente.nombre if obj.cliente_id else "(global)"


@admin.register(FlujoPregunta)
class FlujoPreguntaAdmin(admin.ModelAdmin):
    list_display = ("id", "formulario", "orden", "campo_destino", "tipo_dato", "es_opcional")
    list_filter = ("tipo_dato", "es_opcional")
    search_fields = ("pregunta_texto", "campo_destino")


@admin.register(ResultadoGEI)
class ResultadoGEIAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "ficha",
        "em_total_kg",
        "balance_neto_tco2e",
        "intensidad_kg_co2e_por_kg",
        "evaluacion",
        "completitud_calculo_pct",
        "fecha_calculo",
    )
    list_filter = ("evaluacion",)
    search_fields = ("ficha__estudiante__nombre", "ficha__estudiante__telefono", "ficha__nombre_finca")
    readonly_fields = (
        "ficha",
        "em_fertilizante_kg",
        "em_combustible_kg",
        "em_energia_kg",
        "em_residuos_kg",
        "em_total_kg",
        "rem_bosque_kg",
        "balance_neto_tco2e",
        "intensidad_kg_co2e_por_kg",
        "evaluacion",
        "completitud_calculo_pct",
        "campos_faltantes",
        "fecha_calculo",
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return True

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


@admin.register(SesionFormulario)
class SesionFormularioAdmin(admin.ModelAdmin):
    list_display = ("id", "estudiante", "formulario", "paso_actual", "completado", "fecha_update")
    list_filter = ("completado", "formulario")
    search_fields = ("estudiante__nombre", "estudiante__telefono")
    readonly_fields = (
        "estudiante",
        "formulario",
        "ficha",
        "ficha_destino_id",
        "completado",
        "paso_actual",
        "reintentos_paso",
        "fecha_inicio",
        "fecha_update",
        "progreso",
        "modulo_siguiente",
    )

    def has_add_permission(self, request):
        return False


def _fila_ficha(f: FichaGEI) -> list:
    return [
        f.id,
        f.estudiante_id,
        f.estudiante.nombre,
        f.cliente_id or "",
        f.cliente.nombre if f.cliente_id else "",
        f.curso_id or "",
        f.nombre_finca,
        f.area_ha,
        f.num_plantas,
        f.fertilizante_kg,
        f.concentracion_n_pct,
        f.get_tipo_combustible_display() if f.tipo_combustible else "",
        f.combustible_gal,
        f.energia_kwh,
        f.residuos_ton,
        f.get_manejo_residuos_display() if f.manejo_residuos else "",
        f.produccion_kg,
        f.tiene_bosque,
        f.area_bosque_ha,
        f.completitud_pct,
        f.fecha_inicio,
        f.fecha_update,
    ]


@admin.register(FichaGEI)
class FichaGEIAdmin(admin.ModelAdmin):
    inlines = (ResultadoGEIInline,)
    list_display = (
        "id",
        "estudiante",
        "cliente",
        "curso",
        "nombre_finca",
        "area_ha",
        "num_plantas",
        "fertilizante_kg",
        "concentracion_n_pct",
        "tipo_combustible",
        "combustible_gal",
        "energia_kwh",
        "residuos_ton",
        "manejo_residuos",
        "produccion_kg",
        "tiene_bosque",
        "area_bosque_ha",
        "completitud_col",
    )
    list_filter = ("cliente", "curso", "manejo_residuos", "tipo_combustible")
    search_fields = ("estudiante__nombre", "nombre_finca", "estudiante__cedula", "estudiante__telefono")
    readonly_fields = ("fecha_inicio", "fecha_update", "completitud_col", "ficha_id_display")

    @admin.display(description="Completitud (%)", ordering="id")
    def completitud_col(self, obj: FichaGEI) -> int:
        return obj.completitud_pct

    @admin.display(description="ID")
    def ficha_id_display(self, obj: FichaGEI) -> int:
        return obj.id

    actions = ("exportar_a_excel",)

    @admin.action(description="Exportar selección a Excel (.xlsx)")
    def exportar_a_excel(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, "No hay filas que exportar.", level=messages.WARNING)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FichasGEI"
        encabezado = [
            "id",
            "id_estudiante",
            "nombre_estudiante",
            "id_cliente",
            "nombre_cliente",
            "id_curso",
            "nombre_finca",
            "area_ha",
            "num_plantas",
            "fertilizante_kg",
            "concentracion_n_pct",
            "tipo_combustible",
            "combustible_gal",
            "energia_kwh",
            "residuos_ton",
            "manejo_residuos",
            "produccion_kg",
            "tiene_bosque",
            "area_bosque_ha",
            "completitud_pct",
            "fecha_inicio",
            "fecha_update",
        ]
        for c, t in enumerate(encabezado, 1):
            ws.cell(1, c, t)
        for r, f in enumerate(queryset.select_related("estudiante", "cliente", "curso"), start=2):
            for c, v in enumerate(_fila_ficha(f), start=1):
                if isinstance(v, (datetime,)):
                    v = timezone.localtime(v) if v.tzinfo else v
                ws.cell(r, c, v)
        b = io.BytesIO()
        wb.save(b)
        b.seek(0)
        r = HttpResponse(
            b.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r["Content-Disposition"] = f'attachment; filename="fichas_gei_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        return r
