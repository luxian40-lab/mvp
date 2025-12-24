"""
Script para generar la plantilla Excel de importación de estudiantes.
Ejecutar: python manage.py shell < generate_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Crear archivo
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Estudiantes'

# Estilos
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
header_alignment = Alignment(horizontal='center', vertical='center')

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados
headers = ['Nombre', 'Teléfono']
ws.append(headers)

# Aplicar estilos a encabezados
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# Datos de ejemplo
examples = [
    ['Juan Pérez', '573001234567'],
    ['María García', '573009876543'],
    ['Carlos López', '573005555555'],
    ['Ana Martínez', '573004444444'],
    ['Pedro Rodríguez', '573003333333'],
]

# Agregar ejemplos
for row in examples:
    ws.append(row)

# Aplicar estilos a los datos
for row in ws.iter_rows(min_row=2, max_row=len(examples)+1, min_col=1, max_col=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 18

# Bloquear formato
from openpyxl.worksheet.datavalidation import DataValidation

# Crear validación para teléfono (debe ser numérico)
dv = DataValidation(type="custom", formula1='=NOT(ISERROR(VALUE(B2:B1000)))', showErrorMessage=True)
dv.error = 'El teléfono debe ser un número'
dv.errorTitle = 'Error de Validación'
ws.add_data_validation(dv)
dv.add('B2:B1000')

# Guardar
static_dir = 'static/templates'
if not os.path.exists(static_dir):
    os.makedirs(static_dir)

output_path = os.path.join(static_dir, 'plantilla_estudiantes.xlsx')
wb.save(output_path)

print(f"✅ Plantilla creada exitosamente en: {output_path}")
