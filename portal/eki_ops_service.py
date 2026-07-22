"""Métricas multi-organización para semi-admin eki (portal ops)."""
from __future__ import annotations

from io import BytesIO

from django.utils import timezone

from core.models import Cliente, Curso, Estudiante
from core.models_certificados import Certificado
from portal.dashboard_ops import resumen_dashboard_rapido


def listar_orgs_activas():
    return Cliente.objects.filter(activo=True).order_by('nombre')


def metricas_por_organizacion(*, org_id: int | None = None) -> list[dict]:
    """
    Una fila por organización con KPIs ligeros (solo aggregates).
    No llama calcular_metricas_empresa.
    """
    qs = listar_orgs_activas()
    if org_id:
        qs = qs.filter(pk=org_id)

    filas = []
    for org in qs:
        kpis = resumen_dashboard_rapido(org)
        cursos = Curso.objects.filter(cliente=org, activo=True).count()
        estudiantes = Estudiante.objects.filter(cliente=org).count()
        certificados = Certificado.objects.filter(
            estudiante__cliente=org, emitido=True,
        ).count()
        total = int(kpis.get('total_inscritos') or 0)
        fin = int(kpis.get('finalizados') or 0)
        avance_pct = round(fin * 100 / total) if total else 0
        filas.append({
            'org_id': org.pk,
            'org_nombre': org.nombre,
            'portal_productos': (org.portal_productos or org.tipo_proyecto or '')[:40],
            'cursos': cursos,
            'estudiantes': estudiantes,
            'inscritos': total,
            'finalizados': fin,
            'en_curso': int(kpis.get('en_curso') or 0),
            'no_iniciados': int(kpis.get('no_iniciados') or 0),
            'avance_pct': avance_pct,
            'certificados': certificados,
        })
    return filas


def totales_globales(filas: list[dict]) -> dict:
    keys = ('cursos', 'estudiantes', 'inscritos', 'finalizados', 'en_curso', 'no_iniciados', 'certificados')
    out = {k: sum(int(f.get(k) or 0) for f in filas) for k in keys}
    insc = out['inscritos']
    out['avance_pct'] = round(out['finalizados'] * 100 / insc) if insc else 0
    out['orgs'] = len(filas)
    return out


def exportar_metricas_excel(filas: list[dict]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Metricas orgs'
    headers = [
        'org_id', 'organizacion', 'productos', 'cursos', 'estudiantes',
        'inscritos', 'finalizados', 'en_curso', 'no_iniciados', 'avance_pct', 'certificados',
    ]
    ws.append(headers)
    fill = PatternFill(start_color='5F3A6E', end_color='5F3A6E', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill

    for f in filas:
        ws.append([
            f['org_id'], f['org_nombre'], f['portal_productos'], f['cursos'], f['estudiantes'],
            f['inscritos'], f['finalizados'], f['en_curso'], f['no_iniciados'],
            f['avance_pct'], f['certificados'],
        ])

    ws2 = wb.create_sheet('Totales')
    t = totales_globales(filas)
    ws2.append(['generado', timezone.now().isoformat()])
    for k, v in t.items():
        ws2.append([k, v])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
