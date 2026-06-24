"""Exportaciones Excel desde el portal B2B."""
from __future__ import annotations

import io
from typing import Iterable

from django.http import HttpResponse

from core.drip_schedule import estudiante_llego_hasta_modulo, max_modulo_alcanzado
from core.export_estudiantes import limpiar_telefono
from core.models import Curso, Estudiante, ProgresoEstudiante


def _genero_etiqueta(est) -> str:
    g = (getattr(est, 'genero', None) or '').strip()
    return {'M': 'masculino', 'F': 'femenino', 'O': 'otro', 'NR': 'no reporta'}.get(g, g or '')


def filas_reenganche_sin_modulo(
    org,
    *,
    curso_id: int | None = None,
    grupo_id: int | None = None,
    modulo_objetivo: int | None = None,
) -> list[dict]:
    """
    Una fila por inscripción (estudiante + curso) que NO alcanzó el módulo objetivo.
    Si modulo_objetivo es 5, incluye quienes están en M4 o menos / sin iniciar.
    """
    progreso_q = ProgresoEstudiante.objects.filter(
        estudiante__cliente_id=org.pk,
    ).select_related('estudiante', 'curso', 'modulo_actual')

    if curso_id:
        progreso_q = progreso_q.filter(curso_id=curso_id)
    if grupo_id:
        progreso_q = progreso_q.filter(estudiante__grupos__id=grupo_id).distinct()

    filas = []
    for p in progreso_q.order_by('estudiante__nombre', 'curso__nombre'):
        if modulo_objetivo and estudiante_llego_hasta_modulo(p, modulo_objetivo):
            continue
        est = p.estudiante
        max_m = max_modulo_alcanzado(p)
        grupos = ', '.join(
            g.nombre for g in est.grupos.filter(cliente_id=org.pk).order_by('nombre')[:5]
        )
        filas.append({
            'cedula': est.cedula or '',
            'nombre': est.nombre or '',
            'telefono': limpiar_telefono(est.telefono),
            'municipio': est.municipio or '',
            'departamento': getattr(est, 'departamento', '') or '',
            'genero': _genero_etiqueta(est),
            'edad': est.edad if getattr(est, 'edad', None) else '',
            'curso': p.curso.nombre if p.curso_id else '',
            'cliente': org.nombre,
            'grupo': grupos,
            'modulo_alcanzado': max_m,
            'estado': 'Completado' if p.completado else (
                'En curso' if max_m > 0 else 'Sin avance'
            ),
            'avance_pct': p.porcentaje_avance(),
        })
    return filas


def respuesta_excel_plantilla(
    filas: Iterable[dict],
    *,
    nombre_archivo: str = 'estudiantes_reenganche.xlsx',
) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla reenganche'

    headers = [
        'Cédula', 'Nombre Completo', 'Teléfono', 'Municipio', 'Departamento',
        'Género', 'Edad', 'Curso', 'Cliente',
        'Grupo', 'Módulo alcanzado', 'Estado avance', 'Progreso %',
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row in filas:
        ws.append([
            row['cedula'],
            row['nombre'],
            row['telefono'],
            row['municipio'],
            row['departamento'],
            row['genero'],
            row['edad'],
            row['curso'],
            row['cliente'],
            row['grupo'],
            row['modulo_alcanzado'],
            row['estado'],
            row['avance_pct'],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


def validar_filtros_export(org, curso_id, grupo_id) -> tuple[int | None, int | None]:
    curso_id_int = int(curso_id) if curso_id and str(curso_id).isdigit() else None
    grupo_id_int = int(grupo_id) if grupo_id and str(grupo_id).isdigit() else None
    if curso_id_int and not Curso.objects.filter(pk=curso_id_int, cliente_id=org.pk).exists():
        curso_id_int = None
    if grupo_id_int and not org.grupos_estudiantes.filter(pk=grupo_id_int, activo=True).exists():
        grupo_id_int = None
    return curso_id_int, grupo_id_int


def respuesta_excel_avance_estudiantes(
    filas: Iterable[dict],
    *,
    nombre_archivo: str = 'avance_estudiantes.xlsx',
) -> HttpResponse:
    """Excel con avance detallado por estudiante (portal clientes)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Avance estudiantes'

    headers = [
        'Nombre', 'Cédula', 'Teléfono', 'Grupo(s)', 'Curso',
        'Módulo actual', 'Módulos completados', 'Estado', 'Avance %', 'Puntos',
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row in filas:
        ws.append([
            row.get('nombre', ''),
            row.get('cedula', ''),
            row.get('telefono', ''),
            row.get('grupos', ''),
            row.get('curso', ''),
            row.get('modulo_actual', ''),
            row.get('modulos_completados', ''),
            row.get('estado_avance', ''),
            row.get('avance', 0),
            row.get('puntos', 0),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response
