"""Exportación Excel de fichas GEI desde el portal."""
from __future__ import annotations

import io

from django.http import HttpResponse
from django.utils import timezone

from .gei_service import queryset_fichas_org


def _excel_datetime(dt):
    """openpyxl no acepta datetimes con tzinfo."""
    if not dt:
        return ''
    if timezone.is_aware(dt):
        return timezone.localtime(dt).replace(tzinfo=None)
    return dt


def respuesta_excel_fichas_gei(org, filtros: dict, nombre_archivo: str = 'fichas_gei.xlsx') -> HttpResponse:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        return HttpResponse(
            f'Exportación Excel no disponible (falta openpyxl): {exc}',
            status=500,
            content_type='text/plain; charset=utf-8',
        )

    qs = queryset_fichas_org(org, filtros)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FichasGEI'
    encabezado = [
        'id', 'productor', 'telefono', 'curso', 'nombre_finca', 'area_ha', 'num_plantas',
        'fertilizante_kg', 'concentracion_n_pct', 'tipo_combustible', 'combustible_gal',
        'energia_kwh', 'residuos_ton', 'manejo_residuos', 'produccion_kg',
        'tiene_bosque', 'area_bosque_ha', 'completitud_pct',
        'balance_tco2e', 'evaluacion', 'fecha_inicio', 'fecha_update', 'es_sandbox',
    ]
    ws.append(encabezado)
    header_fill = PatternFill(start_color='9A6CAC', end_color='9A6CAC', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for f in qs.iterator(chunk_size=200):
        res = getattr(f, 'resultado', None)
        ws.append([
            f.id,
            getattr(f.estudiante, 'nombre', '') if f.estudiante_id else '',
            getattr(f.estudiante, 'telefono', '') if f.estudiante_id else '',
            f.curso.nombre if f.curso_id else '',
            f.nombre_finca or '',
            f.area_ha,
            f.num_plantas,
            f.fertilizante_kg,
            f.concentracion_n_pct,
            f.get_tipo_combustible_display() if f.tipo_combustible else '',
            f.combustible_gal,
            f.energia_kwh,
            f.residuos_ton,
            f.get_manejo_residuos_display() if f.manejo_residuos else '',
            f.produccion_kg,
            f.tiene_bosque,
            f.area_bosque_ha,
            f.completitud_pct,
            res.balance_neto_tco2e if res else '',
            res.get_evaluacion_display() if res and res.evaluacion else '',
            _excel_datetime(f.fecha_inicio),
            _excel_datetime(f.fecha_update),
            bool(getattr(f, 'es_sandbox', False)),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response
