"""CRUD de catálogo de recomendaciones y lista de precios para Nat (portal)."""
from __future__ import annotations

from decimal import Decimal, InvalidOperation

from django.core.files.uploadedfile import UploadedFile
from django.db import IntegrityError
from django.db.models import Q

from core.models import ProductoCatalogo, ProductoComercial

_MAX_IMAGEN_BYTES = 5 * 1024 * 1024
_IMAGEN_CONTENT_TYPES = {
    'image/jpeg',
    'image/jpg',
    'image/png',
    'image/webp',
}


def _txt(data, key: str, default: str = '') -> str:
    return (data.get(key) or default).strip()


def _bool_activo(data) -> bool:
    return str(data.get('activo') or '').lower() in ('1', 'true', 'on', 'yes', 'si', 'sí')


def _decimal_or_none(raw, *, required: bool = False, label: str = 'Precio'):
    s = (raw or '').strip().replace(',', '.')
    if not s:
        if required:
            raise ValueError(f'{label} es obligatorio.')
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f'{label} no es un número válido.') from exc


def _int_or_none(raw, *, label: str = 'Stock'):
    s = (raw or '').strip()
    if not s:
        return None
    try:
        val = int(float(s.replace(',', '.')))
    except (TypeError, ValueError) as exc:
        raise ValueError(f'{label} debe ser un número entero.') from exc
    if val < 0:
        raise ValueError(f'{label} no puede ser negativo.')
    return val


def _date_or_none(raw: str):
    from datetime import datetime

    s = (raw or '').strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError as exc:
        raise ValueError('Fecha inválida (use AAAA-MM-DD).') from exc


def _validar_imagen(archivo: UploadedFile | None) -> UploadedFile | None:
    if not archivo:
        return None
    name = (getattr(archivo, 'name', '') or '').lower()
    ctype = (getattr(archivo, 'content_type', '') or '').lower()
    if ctype not in _IMAGEN_CONTENT_TYPES and not name.endswith(('.jpg', '.jpeg', '.png', '.webp')):
        raise ValueError('La foto debe ser JPG, PNG o WebP.')
    size = getattr(archivo, 'size', None)
    if size is not None and size > _MAX_IMAGEN_BYTES:
        raise ValueError('La foto no puede superar 5 MB.')
    return archivo


def listar_catalogo(org, *, q: str = '', solo_activos: bool = False):
    qs = ProductoCatalogo.objects.filter(cliente_id=org.pk)
    if solo_activos:
        qs = qs.filter(activo=True)
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) | Q(categoria__icontains=q) | Q(sku__icontains=q)
        )
    return qs.order_by('categoria', 'nombre')


def listar_precios(org, *, q: str = '', solo_activos: bool = False):
    """Solo SKUs de esta org (nunca el catálogo general cliente=null)."""
    qs = ProductoComercial.objects.filter(cliente_id=org.pk)
    if solo_activos:
        qs = qs.filter(activo=True)
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) | Q(sku__icontains=q) | Q(categoria__icontains=q)
        )
    return qs.order_by('nombre', 'sku')


def crear_catalogo(org, data, files=None) -> ProductoCatalogo:
    nombre = _txt(data, 'nombre')
    if not nombre:
        raise ValueError('El nombre es obligatorio.')
    descripcion = _txt(data, 'descripcion')
    problema = _txt(data, 'problema_que_resuelve')
    if not descripcion:
        raise ValueError('La descripción es obligatoria.')
    if not problema:
        raise ValueError('Indique qué problemas resuelve (Nat lo usa para recomendar).')

    imagen = _validar_imagen((files or {}).get('imagen') if files else None)
    obj = ProductoCatalogo(
        cliente=org,
        nombre=nombre,
        sku=_txt(data, 'sku'),
        descripcion=descripcion,
        problema_que_resuelve=problema,
        ingrediente_activo=_txt(data, 'ingrediente_activo'),
        categoria=_txt(data, 'categoria'),
        cultivos_objetivo=_txt(data, 'cultivos_objetivo'),
        dosis=_txt(data, 'dosis'),
        precio_cop=_decimal_or_none(data.get('precio_cop'), label='Precio de referencia'),
        unidad=_txt(data, 'unidad'),
        url_producto=_txt(data, 'url_producto'),
        activo=_bool_activo(data),
    )
    if imagen:
        obj.imagen = imagen
    try:
        obj.save()
    except IntegrityError as exc:
        raise ValueError(f'Ya existe un producto llamado «{nombre}» en su catálogo.') from exc
    return obj


def actualizar_catalogo(obj: ProductoCatalogo, data, files=None) -> ProductoCatalogo:
    nombre = _txt(data, 'nombre')
    if not nombre:
        raise ValueError('El nombre es obligatorio.')
    descripcion = _txt(data, 'descripcion')
    problema = _txt(data, 'problema_que_resuelve')
    if not descripcion:
        raise ValueError('La descripción es obligatoria.')
    if not problema:
        raise ValueError('Indique qué problemas resuelve.')

    obj.nombre = nombre
    obj.sku = _txt(data, 'sku')
    obj.descripcion = descripcion
    obj.problema_que_resuelve = problema
    obj.ingrediente_activo = _txt(data, 'ingrediente_activo')
    obj.categoria = _txt(data, 'categoria')
    obj.cultivos_objetivo = _txt(data, 'cultivos_objetivo')
    obj.dosis = _txt(data, 'dosis')
    obj.precio_cop = _decimal_or_none(data.get('precio_cop'), label='Precio de referencia')
    obj.unidad = _txt(data, 'unidad')
    obj.url_producto = _txt(data, 'url_producto')
    obj.activo = _bool_activo(data)

    if str(data.get('quitar_imagen') or '').lower() in ('1', 'true', 'on', 'yes'):
        if obj.imagen:
            obj.imagen.delete(save=False)
        obj.imagen = None
    else:
        imagen = _validar_imagen((files or {}).get('imagen') if files else None)
        if imagen:
            obj.imagen = imagen

    try:
        obj.save()
    except IntegrityError as exc:
        raise ValueError(f'Ya existe un producto llamado «{nombre}» en su catálogo.') from exc
    return obj


def crear_precio(org, data) -> ProductoComercial:
    sku = _txt(data, 'sku')
    nombre = _txt(data, 'nombre')
    if not sku:
        raise ValueError('El SKU / código es obligatorio.')
    if not nombre:
        raise ValueError('El nombre es obligatorio.')
    precio = _decimal_or_none(data.get('precio'), required=True, label='Precio')

    obj = ProductoComercial(
        cliente=org,
        sku=sku,
        nombre=nombre,
        presentacion=_txt(data, 'presentacion'),
        unidad=_txt(data, 'unidad'),
        stock=_int_or_none(data.get('stock')),
        precio=precio,
        moneda=_txt(data, 'moneda', 'COP') or 'COP',
        categoria=_txt(data, 'categoria'),
        notas=_txt(data, 'notas'),
        vigencia_desde=_date_or_none(_txt(data, 'vigencia_desde')),
        vigencia_hasta=_date_or_none(_txt(data, 'vigencia_hasta')),
        activo=_bool_activo(data),
    )
    try:
        obj.save()
    except IntegrityError as exc:
        raise ValueError(f'Ya existe el SKU «{sku}» en su lista de precios.') from exc
    return obj


def actualizar_precio(obj: ProductoComercial, data) -> ProductoComercial:
    sku = _txt(data, 'sku')
    nombre = _txt(data, 'nombre')
    if not sku:
        raise ValueError('El SKU / código es obligatorio.')
    if not nombre:
        raise ValueError('El nombre es obligatorio.')
    precio = _decimal_or_none(data.get('precio'), required=True, label='Precio')

    obj.sku = sku
    obj.nombre = nombre
    obj.presentacion = _txt(data, 'presentacion')
    obj.unidad = _txt(data, 'unidad')
    obj.stock = _int_or_none(data.get('stock'))
    obj.precio = precio
    obj.moneda = _txt(data, 'moneda', 'COP') or 'COP'
    obj.categoria = _txt(data, 'categoria')
    obj.notas = _txt(data, 'notas')
    obj.vigencia_desde = _date_or_none(_txt(data, 'vigencia_desde'))
    obj.vigencia_hasta = _date_or_none(_txt(data, 'vigencia_hasta'))
    obj.activo = _bool_activo(data)
    try:
        obj.save()
    except IntegrityError as exc:
        raise ValueError(f'Ya existe el SKU «{sku}» en su lista de precios.') from exc
    return obj


# ── Import / plantilla Excel ───────────────────────────────────────────────

import re
import unicodedata
from io import BytesIO


def _norm_header(raw) -> str:
    s = str(raw or '').strip().lower()
    s = ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )
    s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')
    return s


_ALIASES_PRODUCTO = {
    'nombre': 'nombre',
    'producto': 'nombre',
    'nombre_producto': 'nombre',
    'sku': 'sku',
    'codigo': 'sku',
    'codigo_sku': 'sku',
    'descripcion': 'descripcion',
    'descripcion_para_que_sirve': 'descripcion',
    'problema_que_resuelve': 'problema_que_resuelve',
    'problemas': 'problema_que_resuelve',
    'problemas_que_resuelve': 'problema_que_resuelve',
    'ingrediente_activo': 'ingrediente_activo',
    'ingrediente': 'ingrediente_activo',
    'categoria': 'categoria',
    'cultivos_objetivo': 'cultivos_objetivo',
    'cultivos': 'cultivos_objetivo',
    'cultivo': 'cultivos_objetivo',
    'dosis': 'dosis',
    'dosis_recomendada': 'dosis',
    'precio_cop': 'precio_cop',
    'precio_referencia': 'precio_cop',
    'precio_ref': 'precio_cop',
    'unidad': 'unidad',
    'url_producto': 'url_producto',
    'link': 'url_producto',
    'url': 'url_producto',
    'activo': 'activo',
}

_ALIASES_PRECIO = {
    'sku': 'sku',
    'codigo': 'sku',
    'codigo_sku': 'sku',
    'nombre': 'nombre',
    'producto': 'nombre',
    'presentacion': 'presentacion',
    'unidad': 'unidad',
    'stock': 'stock',
    'existencia': 'stock',
    'inventario': 'stock',
    'cantidad': 'stock',
    'precio': 'precio',
    'precio_cop': 'precio',
    'moneda': 'moneda',
    'categoria': 'categoria',
    'notas': 'notas',
    'vigencia_desde': 'vigencia_desde',
    'desde': 'vigencia_desde',
    'vigencia_hasta': 'vigencia_hasta',
    'hasta': 'vigencia_hasta',
    'activo': 'activo',
}


def _cell_str(val) -> str:
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _map_headers(row_values, aliases: dict) -> dict[int, str]:
    mapping = {}
    for idx, raw in enumerate(row_values):
        key = aliases.get(_norm_header(raw))
        if key:
            mapping[idx] = key
    return mapping


def _row_to_dict(cells, mapping: dict[int, str]) -> dict:
    data = {}
    for idx, key in mapping.items():
        if idx < len(cells):
            data[key] = _cell_str(cells[idx])
    if 'activo' not in data or data.get('activo') == '':
        data['activo'] = '1'
    return data


def generar_plantilla_excel_nat() -> BytesIO:
    """Plantilla con hojas Productos y Precios + fila de ejemplo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws_p = wb.active
    ws_p.title = 'Productos'
    headers_p = [
        'nombre', 'sku', 'descripcion', 'problema_que_resuelve', 'ingrediente_activo',
        'categoria', 'cultivos_objetivo', 'dosis', 'precio_cop', 'unidad',
        'url_producto', 'activo',
    ]
    ws_p.append(headers_p)
    ws_p.append([
        'Fungicida Café Plus',
        'FUNG-CAFE-500G',
        'Fungicida sistémico para control de roya en café.',
        'Roya del cafeto, manchas foliares, hongos en hojas.',
        'Triazol ejemplo',
        'fungicida',
        'café',
        '300-500 g por 200 L de agua',
        '125000',
        '500 g',
        '',
        '1',
    ])
    fill = PatternFill(start_color='0F6E6A', end_color='0F6E6A', fill_type='solid')
    for cell in ws_p[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill

    ws_r = wb.create_sheet('Precios')
    headers_r = [
        'sku', 'nombre', 'presentacion', 'unidad', 'stock', 'precio', 'moneda',
        'categoria', 'notas', 'vigencia_desde', 'vigencia_hasta', 'activo',
    ]
    ws_r.append(headers_r)
    ws_r.append([
        'FUNG-CAFE-500G',
        'Fungicida Café Plus 500g',
        'frasco 500 g',
        'unidad',
        '24',
        '125000',
        'COP',
        'fungicida',
        'Precio referencia tienda',
        '',
        '',
        '1',
    ])
    for cell in ws_r[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill

    ws_help = wb.create_sheet('Instrucciones')
    ws_help.append(['Cómo usar esta plantilla'])
    ws_help.append(['1. Complete la hoja Productos (recomendaciones de Nat).'])
    ws_help.append(['2. Complete la hoja Precios (SKU oficiales).'])
    ws_help.append(['3. Suba el archivo en Portal → Productos o Precios → Importar Excel.'])
    ws_help.append(['4. Si el nombre (productos) o SKU (precios) ya existe, se actualiza.'])
    ws_help.append(['5. activo: 1 = sí, 0 = no.'])
    ws_help.append(['6. sku en Productos: opcional; si coincide con Precios, une ficha + stock.'])
    ws_help.append(['7. stock en Precios: unidades en bodega (opcional).'])
    ws_help.append(['8. La foto del producto se sube en el portal (no por Excel).'])
    ws_help.append(['Columnas obligatorias Productos: nombre, descripcion, problema_que_resuelve'])
    ws_help.append(['Columnas obligatorias Precios: sku, nombre, precio'])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def importar_excel_nat(org, archivo) -> dict:
    """
    Importa hojas Productos y/o Precios.
    Upsert: ProductoCatalogo por nombre; ProductoComercial por sku.
    """
    from openpyxl import load_workbook

    wb = load_workbook(archivo, data_only=True)
    creados_p = actualizados_p = errores_p = 0
    creados_r = actualizados_r = errores_r = 0
    detalle: list[str] = []

    ws = None
    for name in wb.sheetnames:
        if _norm_header(name) in ('productos', 'catalogo', 'producto', 'recomendaciones'):
            ws = wb[name]
            break
    if ws is None and wb.sheetnames:
        first = wb[wb.sheetnames[0]]
        headers = [_cell_str(c.value) for c in next(first.iter_rows(min_row=1, max_row=1))]
        if any(_norm_header(h) in _ALIASES_PRODUCTO for h in headers):
            if 'sku' not in {_norm_header(h) for h in headers} or 'nombre' in {
                _ALIASES_PRODUCTO.get(_norm_header(h)) for h in headers
            }:
                # Prefer productos sheet only if it looks like catalog (has problema or descripcion)
                norms = {_norm_header(h) for h in headers}
                if 'problema_que_resuelve' in norms or 'descripcion' in norms or (
                    'nombre' in norms and 'sku' not in norms
                ):
                    ws = first

    if ws is not None:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            mapping = _map_headers(rows[0], _ALIASES_PRODUCTO)
            if 'nombre' not in mapping.values():
                detalle.append('Hoja Productos: falta columna nombre.')
                errores_p += 1
            else:
                for i, row in enumerate(rows[1:], start=2):
                    if not row or all(v is None or str(v).strip() == '' for v in row):
                        continue
                    data = _row_to_dict(list(row), mapping)
                    try:
                        nombre = _txt(data, 'nombre')
                        if not nombre:
                            raise ValueError('nombre vacío')
                        existing = ProductoCatalogo.objects.filter(
                            cliente=org, nombre__iexact=nombre
                        ).first()
                        if existing:
                            actualizar_catalogo(existing, data)
                            actualizados_p += 1
                        else:
                            crear_catalogo(org, data)
                            creados_p += 1
                    except ValueError as exc:
                        errores_p += 1
                        detalle.append(f'Productos fila {i}: {exc}')

    ws_r = None
    for name in wb.sheetnames:
        if _norm_header(name) in ('precios', 'precio', 'sku', 'lista_de_precios'):
            ws_r = wb[name]
            break

    if ws_r is not None:
        rows = list(ws_r.iter_rows(values_only=True))
        if rows:
            mapping = _map_headers(rows[0], _ALIASES_PRECIO)
            if 'sku' not in mapping.values() or 'precio' not in mapping.values():
                detalle.append('Hoja Precios: faltan columnas sku y/o precio.')
                errores_r += 1
            else:
                for i, row in enumerate(rows[1:], start=2):
                    if not row or all(v is None or str(v).strip() == '' for v in row):
                        continue
                    data = _row_to_dict(list(row), mapping)
                    try:
                        sku = _txt(data, 'sku')
                        if not sku:
                            raise ValueError('sku vacío')
                        if not _txt(data, 'nombre'):
                            data['nombre'] = sku
                        existing = ProductoComercial.objects.filter(
                            cliente=org, sku__iexact=sku
                        ).first()
                        if existing:
                            actualizar_precio(existing, data)
                            actualizados_r += 1
                        else:
                            crear_precio(org, data)
                            creados_r += 1
                    except ValueError as exc:
                        errores_r += 1
                        detalle.append(f'Precios fila {i}: {exc}')

    if ws is None and ws_r is None:
        raise ValueError(
            'No se encontró hoja «Productos» ni «Precios». '
            'Descargue la plantilla e intente de nuevo.'
        )

    return {
        'productos_creados': creados_p,
        'productos_actualizados': actualizados_p,
        'productos_errores': errores_p,
        'precios_creados': creados_r,
        'precios_actualizados': actualizados_r,
        'precios_errores': errores_r,
        'detalle': detalle[:30],
    }
