"""Métricas Nat / comercial para el portal B2B."""
from __future__ import annotations

from collections import Counter
from datetime import timedelta

from django.db.models import Count
from django.utils import timezone

# Umbrales ops (Fase C): sin catálogo/precio = bloqueante, no solo aviso suave.
MIN_CHARS_PROBLEMA = 20
MIN_CHARS_DESCRIPCION = 15


def checklist_preparacion_nat(org) -> list[dict]:
    """
    Semáforo operativo endurecido.
    Sin línea / catálogo / precios / conocimiento indexado → Nat no está lista
    (nivel bad). Calidad débil del catálogo → warn.
    """
    from core.models import BibliotecaConocimiento, DocumentoRAGComercial, ProductoCatalogo, ProductoComercial

    linea = (getattr(org, 'numero_whatsapp_nat', '') or '').strip()
    catalogo_qs = ProductoCatalogo.objects.filter(cliente_id=org.pk, activo=True)
    catalogo_n = catalogo_qs.count()
    catalogo_ok_calidad = sum(
        1
        for p in catalogo_qs.only('problema_que_resuelve', 'descripcion')
        if len((p.problema_que_resuelve or '').strip()) >= MIN_CHARS_PROBLEMA
        and len((p.descripcion or '').strip()) >= MIN_CHARS_DESCRIPCION
    )
    precios_n = ProductoComercial.objects.filter(cliente_id=org.pk, activo=True).count()
    bib_n = BibliotecaConocimiento.objects.filter(cliente_id=org.pk).count()
    bib_idx = BibliotecaConocimiento.objects.filter(
        cliente_id=org.pk, estado_rag='indexado'
    ).count()
    bib_err = BibliotecaConocimiento.objects.filter(
        cliente_id=org.pk, estado_rag='error'
    ).count()
    rag_legacy = DocumentoRAGComercial.objects.filter(
        cliente_id=org.pk, canal='bot_comercial'
    ).count()
    tiene_conocimiento = bib_idx > 0 or rag_legacy > 0

    items = [
        {
            'clave': 'linea',
            'ok': bool(linea),
            'nivel': 'ok' if linea else 'bad',
            'bloqueante': True,
            'titulo': 'Línea WhatsApp Nat',
            'detalle': (
                f'Configurada: {linea}. Debe coincidir con el To de Twilio.'
                if linea
                else 'Sin línea: los mensajes pueden ir al bot educativo. Pida a eki configurarla.'
            ),
            'url': '/portal/perfil/' if not linea else None,
        },
        {
            'clave': 'catalogo',
            'ok': catalogo_n > 0,
            'nivel': 'ok' if catalogo_n > 0 else 'bad',
            'bloqueante': True,
            'titulo': 'Catálogo de recomendaciones',
            'detalle': (
                f'{catalogo_n} producto(s) activos.'
                if catalogo_n
                else 'Sin catálogo: Nat no puede recomendar su oferta (avisará que falta portafolio).'
            ),
            'url': '/portal/catalogo/nuevo/' if catalogo_n == 0 else '/portal/catalogo/',
        },
        {
            'clave': 'calidad_catalogo',
            'ok': catalogo_n == 0 or catalogo_ok_calidad >= max(1, catalogo_n // 2),
            'nivel': (
                'ok'
                if catalogo_n == 0 or catalogo_ok_calidad >= max(1, catalogo_n // 2)
                else 'warn'
            ),
            'bloqueante': False,
            'titulo': 'Calidad del catálogo',
            'detalle': (
                f'{catalogo_ok_calidad}/{catalogo_n} con descripción y problemas bien escritos '
                f'(≥{MIN_CHARS_PROBLEMA} caracteres). Mejora el match con la consulta del productor.'
                if catalogo_n
                else 'Cargue productos primero; luego complete problemas que resuelve y dosis.'
            ),
            'url': '/portal/catalogo/',
        },
        {
            'clave': 'precios',
            'ok': precios_n > 0,
            'nivel': 'ok' if precios_n > 0 else 'bad',
            'bloqueante': True,
            'titulo': 'Lista de precios (SKU)',
            'detalle': (
                f'{precios_n} ítem(s) de precio oficial.'
                if precios_n
                else 'Sin precios: Nat avisará que no hay lista oficial cuando pregunten costo.'
            ),
            'url': '/portal/precios/nuevo/' if precios_n == 0 else '/portal/precios/',
        },
        {
            'clave': 'biblioteca',
            'ok': tiene_conocimiento,
            'nivel': 'ok' if tiene_conocimiento else 'bad',
            'bloqueante': True,
            'titulo': 'Conocimiento (Biblioteca)',
            'detalle': (
                f'Biblioteca: {bib_idx}/{bib_n} indexados'
                + (f', {bib_err} con error' if bib_err else '')
                + f'. RAG legacy: {rag_legacy}.'
            ),
            'url': '/portal/biblioteca/nuevo/' if bib_n == 0 else '/portal/biblioteca/',
        },
        {
            'clave': 'indexacion',
            'ok': bib_err == 0,
            'nivel': 'ok' if bib_err == 0 else 'warn',
            'bloqueante': False,
            'titulo': 'Indexación sin errores',
            'detalle': (
                'Todos los documentos de biblioteca están indexados o pendientes.'
                if bib_err == 0
                else f'{bib_err} documento(s) con error de indexación: Nat no los usará hasta corregirlos.'
            ),
            'url': '/portal/biblioteca/' if bib_err else None,
        },
    ]
    return items


def _mensaje_texto(msg) -> str:
    if not isinstance(msg, dict):
        return ''
    return (msg.get('content') or msg.get('text') or '').strip()


def _rol(msg) -> str:
    if not isinstance(msg, dict):
        return ''
    return (msg.get('role') or '').strip().lower()


def _ultimas_preguntas_sesiones(sesiones, *, limite: int = 40) -> list[dict]:
    out: list[dict] = []
    for s in sesiones:
        hist = list(s.historial_mensajes or [])
        for msg in reversed(hist):
            if _rol(msg) in ('user', 'usuario', 'human') and _mensaje_texto(msg):
                out.append({
                    'telefono': s.telefono,
                    'fecha': s.fecha_ultimo_mensaje,
                    'pregunta': _mensaje_texto(msg)[:280],
                    'turnos': len(hist),
                })
                break
        if len(out) >= limite:
            break
    return out


def _recomendaciones_detectadas(sesiones, nombres_productos: list[str], *, limite: int = 25) -> list[dict]:
    """Heurística: productos del catálogo mencionados en respuestas del asistente."""
    nombres = [n.strip() for n in nombres_productos if (n or '').strip()]
    if not nombres:
        return []
    contador: Counter[str] = Counter()
    ejemplos: dict[str, str] = {}
    for s in sesiones:
        asistente = ' '.join(
            _mensaje_texto(m)
            for m in (s.historial_mensajes or [])
            if _rol(m) in ('assistant', 'asistente', 'bot', 'model')
        )
        if not asistente:
            continue
        low = asistente.lower()
        for nombre in nombres:
            if nombre.lower() in low:
                contador[nombre] += 1
                ejemplos.setdefault(nombre, s.telefono)
    ranked = contador.most_common(limite)
    return [
        {'producto': nombre, 'menciones': n, 'ejemplo_tel': ejemplos.get(nombre, '')}
        for nombre, n in ranked
    ]


def reporte_ops_nat(org, *, dias: int = 30) -> dict:
    """Qué preguntan / qué se recomienda — operación del negocio con Nat."""
    from core.models import (
        BibliotecaConocimiento,
        ContextoAgroSession,
        ConversacionRAGCandidata,
        ProductoCatalogo,
        ProductoComercial,
        SesionComercial,
        SolicitudSoporte,
    )

    from .capabilities import categorias_pqrs_portal

    dias = max(1, min(int(dias or 30), 90))
    desde = timezone.now() - timedelta(days=dias)

    sesiones_qs = SesionComercial.objects.filter(
        cliente_id=org.pk,
        fecha_ultimo_mensaje__gte=desde,
    ).order_by('-fecha_ultimo_mensaje')
    sesiones = list(sesiones_qs[:80])
    sesiones_n = sesiones_qs.count()

    nombres = list(
        ProductoCatalogo.objects.filter(cliente_id=org.pk, activo=True)
        .values_list('nombre', flat=True)[:80]
    )
    preguntas = _ultimas_preguntas_sesiones(sesiones, limite=40)
    recomendaciones = _recomendaciones_detectadas(sesiones, nombres)

    contextos = (
        ContextoAgroSession.objects.filter(sesion__cliente_id=org.pk, updated_at__gte=desde)
        .exclude(cultivo='')
        .values_list('cultivo', flat=True)[:200]
    )
    top_cultivos = Counter(c.strip().lower() for c in contextos if c and c.strip()).most_common(8)

    problemas = (
        ContextoAgroSession.objects.filter(sesion__cliente_id=org.pk, updated_at__gte=desde)
        .exclude(problema='')
        .values_list('problema', flat=True)[:200]
    )
    top_problemas = Counter(p.strip().lower() for p in problemas if p and p.strip()).most_common(8)

    hitl_qs = ConversacionRAGCandidata.objects.filter(
        cliente_id=org.pk, fecha_creacion__gte=desde
    )
    hitl_pend = hitl_qs.filter(estado=ConversacionRAGCandidata.ESTADO_PENDIENTE).count()
    hitl_preguntas = list(
        hitl_qs.order_by('-fecha_creacion')[:20].values('pregunta', 'estado', 'fecha_creacion')
    )

    pqrs_q = SolicitudSoporte.objects.filter(
        estudiante__cliente=org,
        fecha_solicitud__gte=desde,
    )
    cats = categorias_pqrs_portal(org)
    if cats is not None:
        pqrs_q = pqrs_q.filter(categoria__in=cats)
    pqrs_por_cat = list(
        pqrs_q.values('categoria').annotate(n=Count('id')).order_by('-n')[:10]
    )

    checklist = checklist_preparacion_nat(org)
    bloqueantes = [i for i in checklist if i.get('bloqueante') and not i.get('ok')]
    listo = len(bloqueantes) == 0

    return {
        'dias': dias,
        'desde': desde,
        'listo_operar': listo,
        'bloqueantes': bloqueantes,
        'avisos_nat': [
            i['titulo'] + ': ' + i['detalle']
            for i in bloqueantes
        ],
        'sesiones_periodo': sesiones_n,
        'preguntas_recientes': preguntas,
        'recomendaciones_detectadas': recomendaciones,
        'top_cultivos': [{'nombre': k, 'n': v} for k, v in top_cultivos],
        'top_problemas': [{'nombre': k, 'n': v} for k, v in top_problemas],
        'hitl_pendientes': hitl_pend,
        'hitl_preguntas': hitl_preguntas,
        'pqrs_periodo': pqrs_q.count(),
        'pqrs_por_categoria': pqrs_por_cat,
        'catalogo_total': ProductoCatalogo.objects.filter(cliente_id=org.pk, activo=True).count(),
        'precios_total': ProductoComercial.objects.filter(cliente_id=org.pk, activo=True).count(),
        'bib_indexados': BibliotecaConocimiento.objects.filter(
            cliente_id=org.pk, estado_rag='indexado'
        ).count(),
    }


def analitica_nat(org) -> dict:
    from core.models import (
        BibliotecaConocimiento,
        ConversacionRAGCandidata,
        ProductoCatalogo,
        ProductoComercial,
        SesionComercial,
        SolicitudSoporte,
    )

    from .capabilities import categorias_pqrs_portal

    hace_7 = timezone.now() - timedelta(days=7)
    hace_30 = timezone.now() - timedelta(days=30)

    sesiones = SesionComercial.objects.filter(cliente_id=org.pk)
    sesiones_total = sesiones.count()
    sesiones_7d = sesiones.filter(fecha_ultimo_mensaje__gte=hace_7).count()
    sesiones_30d = sesiones.filter(fecha_ultimo_mensaje__gte=hace_30).count()

    catalogo_total = ProductoCatalogo.objects.filter(cliente_id=org.pk, activo=True).count()
    precios_total = ProductoComercial.objects.filter(cliente_id=org.pk, activo=True).count()
    bib_total = BibliotecaConocimiento.objects.filter(cliente_id=org.pk).count()
    bib_indexados = BibliotecaConocimiento.objects.filter(
        cliente_id=org.pk, estado_rag='indexado'
    ).count()
    catalogo_top = list(
        ProductoCatalogo.objects.filter(cliente_id=org.pk, activo=True)
        .order_by('categoria', 'nombre')[:12]
        .values('nombre', 'categoria', 'precio_cop')
    )

    pqrs_q = SolicitudSoporte.objects.filter(estudiante__cliente=org)
    cats = categorias_pqrs_portal(org)
    if cats is not None:
        pqrs_q = pqrs_q.filter(categoria__in=cats)

    hitl_q = ConversacionRAGCandidata.objects.filter(cliente_id=org.pk)
    hitl_pendientes = hitl_q.filter(estado=ConversacionRAGCandidata.ESTADO_PENDIENTE).count()
    hitl_recientes = list(
        hitl_q.select_related('sesion').order_by('-fecha_creacion')[:15]
    )

    sesiones_recientes = list(
        sesiones.order_by('-fecha_ultimo_mensaje')[:15]
    )

    checklist = checklist_preparacion_nat(org)
    checklist_ok = sum(1 for i in checklist if i.get('ok'))
    checklist_pendientes = [i for i in checklist if not i.get('ok')]
    bloqueantes = [i for i in checklist if i.get('bloqueante') and not i.get('ok')]
    listo_operar = len(bloqueantes) == 0

    return {
        'sesiones_total': sesiones_total,
        'sesiones_7d': sesiones_7d,
        'sesiones_30d': sesiones_30d,
        'catalogo_total': catalogo_total,
        'precios_total': precios_total,
        'bib_total': bib_total,
        'bib_indexados': bib_indexados,
        'catalogo_top': catalogo_top,
        'pqrs_total': pqrs_q.count(),
        'pqrs_pendientes': pqrs_q.filter(estado='pendiente').count(),
        'hitl_pendientes': hitl_pendientes,
        'hitl_recientes': hitl_recientes,
        'sesiones_recientes': sesiones_recientes,
        'linea_nat': (org.numero_whatsapp_nat or '').strip() or None,
        'checklist_nat': checklist,
        'checklist_ok': checklist_ok,
        'checklist_total': len(checklist),
        'checklist_pendientes': checklist_pendientes,
        'listo_operar': listo_operar,
        'bloqueantes_nat': bloqueantes,
    }


def exportar_ops_nat_excel(org, *, dias: int = 30):
    """Excel con preguntas recientes y recomendaciones detectadas."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = reporte_ops_nat(org, dias=dias)
    wb = Workbook()

    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['Organización', org.nombre])
    ws.append(['Periodo (días)', data['dias']])
    ws.append(['Listo para operar', 'Sí' if data['listo_operar'] else 'No'])
    ws.append(['Sesiones en periodo', data['sesiones_periodo']])
    ws.append(['HITL pendientes', data['hitl_pendientes']])
    ws.append(['PQRS en periodo', data['pqrs_periodo']])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet('Preguntas')
    ws2.append(['Fecha', 'Teléfono', 'Turnos', 'Última pregunta'])
    for p in data['preguntas_recientes']:
        ws2.append([
            p['fecha'].strftime('%Y-%m-%d %H:%M') if p.get('fecha') else '',
            p.get('telefono', ''),
            p.get('turnos', 0),
            p.get('pregunta', ''),
        ])

    ws3 = wb.create_sheet('Recomendaciones')
    ws3.append(['Producto', 'Menciones (heurística)', 'Ejemplo teléfono'])
    for r in data['recomendaciones_detectadas']:
        ws3.append([r['producto'], r['menciones'], r.get('ejemplo_tel', '')])

    ws4 = wb.create_sheet('Cultivos')
    ws4.append(['Cultivo', 'Contexto N'])
    for c in data['top_cultivos']:
        ws4.append([c['nombre'], c['n']])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
