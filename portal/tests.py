import io

from django.contrib.auth.models import User
from django.test import Client, TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from core.models import Cliente, Curso, Estudiante, Modulo, ProgresoEstudiante
from core.models_extras import GrupoEstudiantes
from portal.capabilities import modulos_portal
from portal.exports import filas_reenganche_sin_modulo
from portal.middleware import PORTAL_SESSION_KEY
from portal.models import PortalUsuario


class PortalPerfilTests(TestCase):
    def setUp(self):
        self.cliente = Cliente.objects.create(
            nombre='Cooperativa Test',
            contacto_principal='Ana',
            email='ana@test.com',
            telefono='573001234567',
            activo=True,
            fecha_fin_suscripcion='2099-12-31',
        )
        self.user_admin = User.objects.create_user('portal_admin', password='pass1234')
        self.user_viewer = User.objects.create_user('portal_viewer', password='pass1234')
        PortalUsuario.objects.create(
            user=self.user_admin, organizacion=self.cliente, rol='admin',
        )
        PortalUsuario.objects.create(
            user=self.user_viewer, organizacion=self.cliente, rol='viewer',
        )
        self.http = Client()

    def _login(self, username):
        user = User.objects.get(username=username)
        pu = PortalUsuario.objects.get(user=user)
        session = self.http.session
        session[PORTAL_SESSION_KEY] = pu.pk
        session.save()

    def test_perfil_requiere_sesion(self):
        r = self.http.get('/portal/perfil/')
        self.assertEqual(r.status_code, 302)

    def test_viewer_no_puede_post_perfil(self):
        self._login('portal_viewer')
        r = self.http.post('/portal/perfil/', {'portal_subtitulo': 'Nuevo'})
        self.assertEqual(r.status_code, 200)
        self.cliente.refresh_from_db()
        self.assertEqual(self.cliente.portal_subtitulo, '')

    def test_admin_guarda_subtitulo(self):
        self._login('portal_admin')
        r = self.http.post('/portal/perfil/', {'portal_subtitulo': 'Programa 2026'})
        self.assertEqual(r.status_code, 200)
        self.cliente.refresh_from_db()
        self.assertEqual(self.cliente.portal_subtitulo, 'Programa 2026')


class PortalExportDetalleTests(TestCase):
    def setUp(self):
        self.cliente = Cliente.objects.create(
            nombre='Org Export',
            contacto_principal='Ana',
            email='export@test.com',
            telefono='573009990001',
            activo=True,
            fecha_fin_suscripcion='2099-12-31',
        )
        self.otro_cliente = Cliente.objects.create(
            nombre='Otra Org',
            contacto_principal='Bob',
            email='otro@test.com',
            telefono='573009990002',
            activo=True,
            fecha_fin_suscripcion='2099-12-31',
        )
        user = User.objects.create_user('portal_exp', password='pass1234')
        PortalUsuario.objects.create(user=user, organizacion=self.cliente, rol='admin')
        self.http = Client()
        session = self.http.session
        session[PORTAL_SESSION_KEY] = PortalUsuario.objects.get(user=user).pk
        session.save()

        self.curso = Curso.objects.create(
            nombre='Curso Test',
            descripcion='Desc',
            cliente=self.cliente,
            activo=True,
        )
        self.m1 = Modulo.objects.create(
            curso=self.curso, numero=1, titulo='M1', descripcion='d', contenido='c',
        )
        self.m2 = Modulo.objects.create(
            curso=self.curso, numero=2, titulo='M2', descripcion='d', contenido='c',
        )
        self.est_bajo = Estudiante.objects.create(
            cedula='9001', nombre='Bajo Modulo', telefono='573111111101', cliente=self.cliente,
        )
        self.est_alto = Estudiante.objects.create(
            cedula='9002', nombre='Alto Modulo', telefono='573111111102', cliente=self.cliente,
        )
        self.est_otro = Estudiante.objects.create(
            cedula='9003', nombre='Externo', telefono='573111111103', cliente=self.otro_cliente,
        )
        ProgresoEstudiante.objects.create(
            estudiante=self.est_bajo, curso=self.curso, modulo_actual=self.m1,
        )
        ProgresoEstudiante.objects.create(
            estudiante=self.est_alto, curso=self.curso, modulo_actual=self.m2,
        )
        self.grupo = GrupoEstudiantes.objects.create(
            nombre='Grupo Norte', cliente=self.cliente, activo=True,
        )
        self.grupo.estudiantes.add(self.est_bajo)

    def test_export_sin_modulo_redirige(self):
        r = self.http.get('/portal/estudiantes/exportar/')
        self.assertEqual(r.status_code, 302)
        self.assertIn('export_error=modulo', r.url)

    def test_export_excel_excluye_quien_llego_al_modulo(self):
        r = self.http.get('/portal/estudiantes/exportar/?sin_modulo=2')
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        nombres = {row[1] for row in rows}
        self.assertIn('Bajo Modulo', nombres)
        self.assertNotIn('Alto Modulo', nombres)

    def test_export_filtra_por_grupo(self):
        filas = filas_reenganche_sin_modulo(
            self.cliente, curso_id=self.curso.pk, grupo_id=self.grupo.pk, modulo_objetivo=2,
        )
        self.assertEqual(len(filas), 1)
        self.assertEqual(filas[0]['nombre'], 'Bajo Modulo')

    def test_estudiante_detalle_y_aislamiento(self):
        r = self.http.get(f'/portal/estudiantes/{self.est_bajo.pk}/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Bajo Modulo')
        r2 = self.http.get(f'/portal/estudiantes/{self.est_otro.pk}/')
        self.assertEqual(r2.status_code, 404)

    def test_metricas_acepta_filtro_grupo(self):
        r = self.http.get(f'/portal/metricas/?grupo={self.grupo.pk}&curso={self.curso.pk}')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Grupo Norte')
        self.assertContains(r, 'Descargar reportes')

    def test_portal_reportes_page(self):
        r = self.http.get('/portal/reportes/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Reportes y descargas')
        self.assertContains(r, 'Avance de estudiantes')

    def test_export_avance_excel(self):
        r = self.http.get(f'/portal/estudiantes/exportar/avance/?curso={self.curso.pk}')
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        nombres = {row[0] for row in rows}
        self.assertIn('Bajo Modulo', nombres)
        self.assertIn('Alto Modulo', nombres)


class PortalModulosTests(TestCase):
    def test_modulos_desde_portal_productos(self):
        c = Cliente.objects.create(
            nombre='Combo',
            contacto_principal='X',
            email='combo@test.com',
            telefono='573009990099',
            activo=True,
            tipo_proyecto='cursos',
            portal_productos='cursos,gei,nat',
        )
        m = modulos_portal(c)
        self.assertTrue(m['cursos'] and m['gei'] and m['nat'])
        self.assertFalse(m['empleabilidad'])

    def test_modulos_solo_tipo_principal(self):
        c = Cliente.objects.create(
            nombre='Solo GEI',
            contacto_principal='X',
            email='gei@test.com',
            telefono='573009990098',
            activo=True,
            tipo_proyecto='gei',
        )
        m = modulos_portal(c)
        self.assertFalse(m['cursos'])
        self.assertTrue(m['gei'])

    def test_modulos_gei_por_ficha_en_curso(self):
        from core.models import Curso

        c = Cliente.objects.create(
            nombre='Cursos+GEI ficha',
            contacto_principal='X',
            email='ficha@test.com',
            telefono='573009990096',
            activo=True,
            tipo_proyecto='cursos',
            portal_productos='cursos',
        )
        Curso.objects.create(
            cliente=c,
            nombre='Curso con GEI',
            activo=True,
            tiene_formulario_gei=True,
        )
        m = modulos_portal(c)
        self.assertTrue(m['cursos'])
        self.assertTrue(m['gei'])


class PortalViewerGeiTests(TestCase):
    def setUp(self):
        self.cliente = Cliente.objects.create(
            nombre='Viewer Org',
            contacto_principal='A',
            email='v@test.com',
            telefono='573009990097',
            activo=True,
            portal_productos='cursos',
        )
        user = User.objects.create_user('viewer_only', password='pass1234')
        PortalUsuario.objects.create(user=user, organizacion=self.cliente, rol='viewer')
        self.http = Client()
        session = self.http.session
        session[PORTAL_SESSION_KEY] = PortalUsuario.objects.get(user=user).pk
        session.save()

    def test_viewer_no_guarda_metas(self):
        r = self.http.post('/portal/metricas/', {'meta_finalizacion_porcentaje': '99'})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'solo lectura')

    def test_gei_lista_requiere_modulo(self):
        r = self.http.get('/portal/gei/')
        self.assertEqual(r.status_code, 302)


class PortalMetricasReorganizacionTests(TestCase):
    def setUp(self):
        self.cliente = Cliente.objects.create(
            nombre='Org Métricas',
            contacto_principal='Ana',
            email='metricas@test.com',
            telefono='573009990099',
            activo=True,
            fecha_fin_suscripcion='2099-12-31',
            portal_productos='cursos',
        )
        user = User.objects.create_user('portal_metricas', password='pass1234')
        PortalUsuario.objects.create(user=user, organizacion=self.cliente, rol='admin')
        self.http = Client()
        session = self.http.session
        session[PORTAL_SESSION_KEY] = PortalUsuario.objects.get(user=user).pk
        session.save()

    def test_gamificacion_redirige_a_dashboard(self):
        r = self.http.get('/portal/gamificacion/')
        self.assertEqual(r.status_code, 302)
        self.assertIn('/portal/dashboard/', r['Location'])
        self.assertIn('#ranking', r['Location'])

    def test_metricas_detalladas_titulo(self):
        r = self.http.get('/portal/metricas/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Métricas detalladas')
        self.assertContains(r, 'Resumen ejecutivo')
        self.assertContains(r, 'Analítica de aprendizaje')

    def test_dashboard_sin_reporte_b2b(self):
        r = self.http.get('/portal/dashboard/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Inicio')
        self.assertNotContains(r, 'Reporte B2B')
        self.assertContains(r, 'Métricas detalladas')


class PortalCertificadosFlujoTests(TestCase):
    def setUp(self):
        self.cliente = Cliente.objects.create(
            nombre='Org Portal CF',
            contacto_principal='Ana',
            email='cf@test.com',
            telefono='573007777771',
            activo=True,
            fecha_fin_suscripcion='2099-12-31',
            tipo_proyecto='cursos',
        )
        self.otro = Cliente.objects.create(
            nombre='Otra Org CF',
            contacto_principal='Bob',
            email='otrocf@test.com',
            telefono='573007777772',
            activo=True,
            fecha_fin_suscripcion='2099-12-31',
        )
        user = User.objects.create_user('portal_cf', password='pass1234')
        PortalUsuario.objects.create(user=user, organizacion=self.cliente, rol='viewer')
        self.http = Client()
        session = self.http.session
        session[PORTAL_SESSION_KEY] = PortalUsuario.objects.get(user=user).pk
        session.save()

        self.curso = Curso.objects.create(
            nombre='Curso Embudo', cliente=self.cliente, activo=True,
        )
        self.m1 = Modulo.objects.create(
            curso=self.curso, numero=1, titulo='Intro', descripcion='d', contenido='c',
        )
        self.m2 = Modulo.objects.create(
            curso=self.curso, numero=2, titulo='Avanzado', descripcion='d', contenido='c',
        )
        self.est1 = Estudiante.objects.create(
            cedula='cf1', nombre='Uno', telefono='573007777773', cliente=self.cliente,
        )
        self.est2 = Estudiante.objects.create(
            cedula='cf2', nombre='Dos', telefono='573007777774', cliente=self.cliente,
        )
        self.est_otro = Estudiante.objects.create(
            cedula='cf3', nombre='Externo', telefono='573007777775', cliente=self.otro,
        )
        ProgresoEstudiante.objects.create(
            estudiante=self.est1, curso=self.curso, modulo_actual=self.m2,
        )
        ProgresoEstudiante.objects.create(
            estudiante=self.est2, curso=self.curso, modulo_actual=self.m1,
        )

    def test_flujo_curso_solo_lectura(self):
        r = self.http.get(f'/portal/cursos/{self.curso.id}/flujo/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Flujo del curso')
        self.assertContains(r, 'solo lectura')
        self.assertContains(r, 'M1')
        self.assertContains(r, 'M2')

    def test_flujo_no_expone_curso_otra_org(self):
        curso_otro = Curso.objects.create(nombre='Ajeno', cliente=self.otro, activo=True)
        r = self.http.get(f'/portal/cursos/{curso_otro.id}/flujo/')
        self.assertEqual(r.status_code, 302)
        self.assertIn('/portal/cursos/', r['Location'])

    def test_certificados_solo_org(self):
        from core.models_certificados import Certificado
        from django.utils import timezone

        Certificado.objects.create(
            estudiante=self.est1,
            curso=self.curso,
            calificacion_final=100,
            fecha_inicio='2026-01-01',
            emitido=True,
            fecha_emision=timezone.now(),
        )
        curso_otro = Curso.objects.create(nombre='Ajeno', cliente=self.otro, activo=True)
        Certificado.objects.create(
            estudiante=self.est_otro,
            curso=curso_otro,
            calificacion_final=100,
            fecha_inicio='2026-01-01',
            emitido=True,
            fecha_emision=timezone.now(),
        )
        r = self.http.get('/portal/certificados/')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Uno')
        self.assertNotContains(r, 'Externo')
